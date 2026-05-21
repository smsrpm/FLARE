"""
flare_ua_worker.py — background worker for FLARE uncertainty analysis.

Launched by flare_ua.py with:
    python -u flare_ua_worker.py <path-to-ua_worker_config.json>

The worker runs independently of the Streamlit browser session, writes
ua_status.json atomically, and writes per-sample console logs in the UA run
folder so the UI can poll progress and recover after a browser disconnect.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

STATUS_FILE = "ua_status.json"
ABORT_FILE  = "ua_abort_requested.json"

def _now():
    return datetime.now().isoformat(timespec="seconds")

def _write_json_atomic(path: Path, data: dict):
    """Write JSON status robustly on Windows/OneDrive.

    OneDrive, antivirus, or the Streamlit UI can briefly lock ua_status.json
    while the worker is updating it.  A status-write failure must not kill a
    long UA run, so this function retries the atomic replace and then falls
    back to direct write.  Final failure is logged and ignored.
    """
    payload = json.dumps(data, indent=2)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Unique temp name avoids collisions/stale locks on a fixed .tmp file.
    tmp = path.with_name(f"{path.name}.{os.getpid()}.{int(time.time() * 1000)}.tmp")

    last_err = None
    for attempt in range(8):
        try:
            tmp.write_text(payload, encoding="utf-8")
            os.replace(str(tmp), str(path))
            return
        except PermissionError as e:
            last_err = e
            time.sleep(0.05 * (attempt + 1))
        except OSError as e:
            last_err = e
            time.sleep(0.05 * (attempt + 1))

    # Fallback: direct write.  This is not as atomic, but is better than
    # terminating the worker because a status heartbeat could not be written.
    try:
        path.write_text(payload, encoding="utf-8")
        return
    except Exception as e:
        last_err = e

    try:
        warn = path.parent / "ua_status_write_warnings.log"
        with warn.open("a", encoding="utf-8", errors="replace") as f:
            f.write(f"[{_now()}] Could not update {path.name}: {last_err}\n")
    except Exception:
        pass

    try:
        tmp.unlink(missing_ok=True)
    except Exception:
        pass

def _read_json(path: Path, default=None):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default

def _kill_tree(pid: int):
    if not pid:
        return
    try:
        if os.name == "nt":
            subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"],
                           capture_output=True, text=True, timeout=10)
        else:
            os.kill(pid, 15)
    except Exception:
        pass

def _aborted(run_dir: Path) -> bool:
    return (run_dir / ABORT_FILE).exists()

def sample_values(dist, p1, p2, base, n, rng):
    if dist == "uniform":
        if p1 >= p2:
            raise ValueError(f"Uniform requires Lower < Upper, got {p1} >= {p2}")
        return rng.uniform(p1, p2, n)
    if dist == "normal":
        if p2 <= 0:
            raise ValueError(f"Normal requires Std deviation > 0, got {p2}")
        raw = rng.normal(p1, p2, n)
        return np.clip(raw, p1 - 4*p2, p1 + 4*p2)
    if dist == "lognormal":
        if p2 <= 0:
            raise ValueError(f"Lognormal requires ln(std) > 0, got {p2}")
        return np.exp(rng.normal(p1, p2, n))
    if dist == "triangular":
        if p1 >= p2:
            raise ValueError(f"Triangular requires Lower < Upper, got {p1} >= {p2}")
        if not (p1 <= base <= p2):
            raise ValueError(f"Triangular mode {base} must be between {p1} and {p2}")
        return rng.triangular(p1, base, p2, n)
    return np.full(n, base)

def _copy_input_snapshot(src: Path, run_dir: Path) -> Path:
    """Copy an input workbook to a private snapshot before openpyxl reads it.

    This mirrors the PWR Simulator pattern.  The user may have the original
    workbook open in Excel, and OneDrive may briefly lock it during save/sync.
    We retry the snapshot copy and then operate only on the private copy.
    """
    run_dir.mkdir(parents=True, exist_ok=True)
    snap = run_dir / f".~ua_input_snapshot_{src.stem}_{os.getpid()}_{uuid.uuid4().hex[:8]}.xlsx"
    last_err = None
    for attempt in range(12):
        try:
            shutil.copy2(src, snap)
            return snap
        except PermissionError as e:
            last_err = e
            time.sleep(0.15 * (attempt + 1))
        except OSError as e:
            last_err = e
            time.sleep(0.15 * (attempt + 1))
    raise PermissionError(
        f"Could not copy input workbook after retries: {src}. "
        f"Close/save the workbook or pause OneDrive sync if this persists. Last error: {last_err}"
    )


# Directories that FLARE creates at runtime and should not be searched for base input decks.
_EXCLUDE_INPUT_DIR_PREFIXES = ("sim_", "risk_", "ua_", ".sim_all_", "__pycache__")

def _is_generated_or_hidden_dir(path: Path, work_dir: Path) -> bool:
    """Return True if *path* is inside a generated FLARE output/control folder."""
    try:
        rel_parts = path.relative_to(work_dir).parts
    except Exception:
        rel_parts = path.parts
    for part in rel_parts:
        if part in (".", ""):
            continue
        if part.startswith(_EXCLUDE_INPUT_DIR_PREFIXES):
            return True
    return False

def resolve_base_input_path(work_dir: Path, base_case: str, configured_path=None) -> Path:
    """Resolve the base input workbook for UA.

    The recursive-input workflow intentionally stores Case*_in.xlsx files in
    subfolders and ignores root-level input decks.  Older UI/config files may
    omit base_input_path or may provide the historical root fallback.  If the
    configured path is absent or invalid, search all non-generated subfolders
    for <base_case>_in.xlsx.
    """
    candidates = []

    if configured_path:
        cp = Path(configured_path)
        if not cp.is_absolute():
            cp = work_dir / cp
        # Accept a valid explicit path even if it is in the root; this preserves
        # backwards compatibility for old run configurations.
        if cp.exists():
            return cp

    target = f"{base_case}_in.xlsx"
    for f in work_dir.rglob(target):
        if not f.is_file():
            continue
        # Do not search the FLARE root in the new organization.
        if f.parent == work_dir:
            continue
        if _is_generated_or_hidden_dir(f.parent, work_dir):
            continue
        if f.name.startswith(".~"):
            continue
        candidates.append(f)

    candidates = sorted(candidates, key=lambda x: str(x).lower())
    if not candidates:
        searched = str(work_dir)
        raise FileNotFoundError(
            f"Base input workbook not found for {base_case}: expected {target} "
            f"in a non-generated subfolder under {searched}."
        )
    if len(candidates) > 1:
        # Deterministic choice; log enough detail to diagnose duplicate decks.
        print(
            "[ua-worker] WARNING: multiple base input workbooks found for "
            f"{base_case}; using {candidates[0]}. Other matches: "
            + "; ".join(str(c) for c in candidates[1:]),
            flush=True,
        )
    return candidates[0]

def build_ua_input(work_dir: Path, run_dir: Path, base_case: str, overrides: dict, sample_id: int, base_input_path: Path = None) -> str:
    src      = Path(base_input_path) if base_input_path is not None else work_dir / f"{base_case}_in.xlsx"
    tmp_name = f"ua_{base_case}_{sample_id}"
    dst      = run_dir / f"{tmp_name}_in.xlsx"

    snap = _copy_input_snapshot(src, run_dir)
    wb = None
    try:
        wb = load_workbook(snap)
        ws = wb[f"{base_case}_in"]
        if f"{base_case}_out" in wb.sheetnames:
            wb[f"{base_case}_out"].title = f"{tmp_name}_out"
        ws.title = f"{tmp_name}_in"

        time_row = None
        for row in ws.iter_rows(max_col=1):
            v = row[0].value
            if isinstance(v, str) and v.strip().startswith("Time"):
                time_row = row[0].row
                break
        if time_row is None:
            time_row = ws.max_row

        n_ins = len(overrides) + 1
        ws.insert_rows(time_row, amount=n_ins)
        ws.cell(row=time_row, column=1).value = "# UA overrides (last-assignment-wins)"
        for i, (var, val) in enumerate(overrides.items(), start=1):
            ws.cell(row=time_row + i, column=1).value = f"{var} = {val:.8g}"

        # Save the perturbed input directly in the UA run folder, where flare_sim.py will run.
        wb.save(dst)
        return tmp_name
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
        try:
            snap.unlink(missing_ok=True)
        except Exception:
            pass

def move_outputs(work_dir: Path, run_dir: Path, tmp_case: str):
    # Move known outputs plus diagnostic and figure artifacts for this temporary case.
    suffixes = (
        "_in.xlsx", "_in.csv", "_out.xlsx", "_out.csv", "_diag.csv",
        "_fail.csv", "_figures.pdf", "_console.log"
    )
    for sfx in suffixes:
        src = work_dir / f"{tmp_case}{sfx}"
        if src.exists():
            try:
                shutil.move(str(src), str(run_dir / src.name))
            except Exception:
                pass

    # Move any additional input echo/snapshot CSVs generated by flare_sim.py.
    # Some simulator versions emit ua_<case>_<sample>_in.csv separately from
    # the workbook; keep it with the rest of the UA sample artifacts.
    for f in work_dir.glob(f"{tmp_case}*_in.csv"):
        try:
            shutil.move(str(f), str(run_dir / f.name))
        except Exception:
            pass

    # Remove/openpyxl temp input copy if the simulator created it.
    tmp = work_dir / f".~{tmp_case}_in.xlsx"
    try:
        if tmp.exists():
            tmp.unlink()
    except Exception:
        pass

    # Move generic figures, if any; in fast mode these usually do not exist.
    for pat in (f"{tmp_case}_*.png", f"{tmp_case}_*.pdf", "figure_*.png"):
        for f in work_dir.glob(pat):
            try:
                shutil.move(str(f), str(run_dir / f.name))
            except Exception:
                pass

def extract_scalars(base_case: str, sample_id: int, run_dir: Path) -> dict:
    csv_path = run_dir / f"ua_{base_case}_{sample_id}_out.csv"
    if not csv_path.exists():
        return {}
    try:
        df = pd.read_csv(csv_path)
        r = {}
        if "RCS Pressure (kPa)" in df.columns:
            r["P_min_kPa"] = df["RCS Pressure (kPa)"].min()
            r["P_max_kPa"] = df["RCS Pressure (kPa)"].max()
        if "RCS Temperature (K)" in df.columns:
            r["T_max_K"] = df["RCS Temperature (K)"].max()
        if "RK Total Power (MW)" in df.columns:
            r["P_peak_MW"] = df["RK Total Power (MW)"].max()
        if "DNBR" in df.columns:
            dn = df["DNBR"].replace(0, np.nan)
            r["DNBR_min"] = dn.min()
        if "Clad Surface Temp (K)" in df.columns:
            tw = df["Clad Surface Temp (K)"]
            r["avg_clad_peak_K"]  = tw.max()
            r["avg_clad_final_K"] = tw.dropna().iloc[-1] if len(tw.dropna()) else np.nan
        if "Hot Pin Clad Temp (K)" in df.columns:
            thc = df["Hot Pin Clad Temp (K)"]
            r["hot_pin_clad_peak_K"]  = thc.max()
            r["hot_pin_clad_final_K"] = thc.dropna().iloc[-1] if len(thc.dropna()) else np.nan
        if "Hot Pin Fuel Temp (K)" in df.columns:
            thf = df["Hot Pin Fuel Temp (K)"]
            r["hot_pin_fuel_peak_K"] = thf.max()
        if "Rod Failures DNB (est.)" in df.columns:
            r["N_fail_DNB"] = int(df["Rod Failures DNB (est.)"].max())
        if "Rod Failures Gap (est.)" in df.columns:
            r["N_fail_gap"] = int(df["Rod Failures Gap (est.)"].max())
        if "Rod Failures EarlyIV (est.)" in df.columns:
            r["N_fail_eiv"] = int(df["Rod Failures EarlyIV (est.)"].max())
        return r
    except Exception as e:
        return {"error": str(e)}


# UA plot export defaults.  Keep this list aligned with the PWR Simulator
# figures and the UA Results selector in flare_ua.py.
UA_PLOTTED_TS_COLUMNS = [
    "RCS Pressure (kPa)", "Equilibrium Quality (-)", "Void Fraction (-)",
    "Total Mass Scaled", "Vessel Level (m)", "Break Flow (kg/s)",
    "PORV Mass Flow (kg/s)", "Accumulator Pressure (kPa)",
    "Accumulator Temperature (K)", "Accumulator Level (m)",
    "Accumulator Flow (kg/s)", "CVCS Makeup (kg/s)", "CVCS Letdown (kg/s)",
    "HPSI Flow (kg/s)", "LPSI Flow (kg/s)", "SI Pumped Total (kg/s)",
    "RK Total Power (MW)", "Core Power (MW)", "Pump Speed (rpm)",
    "Pump Velocity (m/s)", "SG Heat Removal (MW)",
    "Clad Surface Temp (K)", "RCS Temperature (K)", "Fuel Avg Temp (K)",
    "Hot Pin Clad Temp (K)", "Hot Pin Fuel Temp (K)", "Clad HTC (W/m2-K)",
    "DNBR", "Reactivity scram (pcm)", "Reactivity ext (pcm)",
    "Reactivity Boron (pcm)", "Reactivity Doppler (pcm)",
    "Reactivity Moderator (pcm)", "Reactivity net (pcm)",
    "Pressurizer Level (m)", "Pressurizer Level (norm)",
    "Zr Oxidation Hot Pin ECR (%)", "Zr Oxidation Mean Oxidizing Rod ECR (%)",
    "H2 Generated (kg)", "H2 Full Core Cladding Reaction (kg)",
    "Zr Oxidizing Rods (est.)",
]

UNIT_CONV_WORKER = {
    "(K)":    ("°C", lambda v: v - 273.15),
    "(kPa)":  ("kPa", lambda v: v),
    "(kg/s)": ("kg/s", lambda v: v),
    "(MW)":   ("MW", lambda v: v),
    "[-]":    ("", lambda v: v),
}

SCALAR_UNIT_WORKER = {
    "hot_pin_clad_peak_K":  "(K)",
    "hot_pin_clad_final_K": "(K)",
    "avg_clad_peak_K":      "(K)",
    "avg_clad_final_K":     "(K)",
    "hot_pin_fuel_peak_K":  "(K)",
    "T_max_K":              "(K)",
    "P_min_kPa":            "(kPa)",
    "P_max_kPa":            "(kPa)",
    "P_peak_MW":            "(MW)",
}

def _safe_plot_token(text, max_len=48):
    return (re.sub(r"[^A-Za-z0-9_.-]+", "_", str(text)).strip("_")[:max_len] or "plot")

def _convert_ts_metric(values, col_name):
    arr = pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(dtype=float)
    for suffix, (unit, fn) in UNIT_CONV_WORKER.items():
        if str(col_name).endswith(suffix):
            return fn(arr), unit
    return arr, ""

def _convert_scalar_metric(values, col_name):
    arr = pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(dtype=float)
    suffix = SCALAR_UNIT_WORKER.get(str(col_name))
    if suffix and suffix in UNIT_CONV_WORKER:
        unit, fn = UNIT_CONV_WORKER[suffix]
        return fn(arr), unit
    return arr, ""

def create_default_ua_plots(run_dir: Path, base_case: str, results_df: pd.DataFrame):
    """Automatically create the default UA PNG/PDF plot set.

    This is called only when Fast mode is not selected.  It mirrors the default
    UI selections: Clad Surface Temp time-series and hot-pin clad peak scalar
    when available, with fallback choices if a case lacks those outputs.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    df = results_df.copy()
    if df is None or df.empty or "status" not in df.columns:
        return None
    df_ok = df[df["status"] == "OK"].copy()
    if df_ok.empty:
        return None

    ts_entries = []
    available_cols = set()
    for f in sorted(run_dir.glob(f"ua_{base_case}_*_out.csv")):
        m = re.search(r"_(\d+)_out\.csv$", f.name)
        if not m:
            continue
        try:
            ts_df = pd.read_csv(f)
            ts_entries.append({"sample": int(m.group(1)), "df": ts_df})
            available_cols.update(c for c in ts_df.columns if c != "Time (s)")
        except Exception:
            pass
    ts_choices = [c for c in UA_PLOTTED_TS_COLUMNS if c in available_cols]
    selected_ts = next((c for c in ["Clad Surface Temp (K)", "Hot Pin Clad Temp (K)", "RCS Pressure (kPa)"] if c in ts_choices), ts_choices[0] if ts_choices else None)

    out_cols = [c for c in df_ok.columns if not str(c).startswith("in_") and c not in ("sample", "status", "error")]
    selected_out = "hot_pin_clad_peak_K" if "hot_pin_clad_peak_K" in out_cols else (out_cols[0] if out_cols else None)
    if not selected_out:
        return None

    out_dir = run_dir / "ua_plots"
    out_dir.mkdir(parents=True, exist_ok=True)
    tag = f"default_{_safe_plot_token(selected_ts)}_{_safe_plot_token(selected_out)}"
    pdf_path = out_dir / f"ua_{_safe_plot_token(base_case)}_{tag}_plots.pdf"
    pngs = []

    y_vals, y_unit = _convert_scalar_metric(df_ok[selected_out], selected_out)
    y_label = selected_out.replace("_", " ") + (f" [{y_unit}]" if y_unit else "")
    in_cols = [c for c in df_ok.columns if str(c).startswith("in_")]

    def save(fig, name, pdf):
        png = out_dir / f"ua_{_safe_plot_token(base_case)}_{tag}_{name}.png"
        fig.tight_layout()
        fig.savefig(png, dpi=180, bbox_inches="tight")
        pngs.append(str(png))
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

    with PdfPages(pdf_path) as pdf:
        if selected_ts and ts_entries:
            fig, ax = plt.subplots(figsize=(9.5, 5.2))
            nplot = 0
            unit = ""
            for entry in ts_entries:
                ts_df = entry["df"]
                if selected_ts in ts_df.columns and "Time (s)" in ts_df.columns:
                    x = pd.to_numeric(ts_df["Time (s)"], errors="coerce")
                    y, unit = _convert_ts_metric(ts_df[selected_ts], selected_ts)
                    ax.plot(x, y, linewidth=1.0, alpha=0.55, label=f"S{entry['sample']}")
                    nplot += 1
            if nplot:
                base_label = selected_ts
                for sfx in UNIT_CONV_WORKER:
                    if selected_ts.endswith(sfx):
                        base_label = selected_ts[:-len(sfx)].rstrip()
                        break
                ax.set_title(f"UA Time-Series Overlay — {selected_ts}")
                ax.set_xlabel("Time [s]")
                ax.set_ylabel(f"{base_label} [{unit}]" if unit else base_label)
                ax.grid(True, alpha=0.35)
                if nplot <= 20:
                    ax.legend(fontsize=7, ncol=2)
                save(fig, "fig01_timeseries", pdf)
            else:
                plt.close(fig)

        valid = np.isfinite(y_vals)
        if valid.any():
            ys = np.sort(y_vals[valid])
            cdf = np.arange(1, len(ys) + 1) / len(ys)
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.plot(ys, cdf, linewidth=2.0)
            ax.set_title(f"UA CDF — {selected_out}")
            ax.set_xlabel(y_label)
            ax.set_ylabel("Cumulative probability")
            ax.set_ylim(0, 1)
            ax.grid(True, alpha=0.35)
            save(fig, "fig02_cdf", pdf)

        corr_pairs = []
        fig_no = 3
        for in_col in in_cols:
            x = pd.to_numeric(df_ok[in_col], errors="coerce").to_numpy(dtype=float)
            mask = np.isfinite(x) & np.isfinite(y_vals)
            if mask.sum() < 2:
                continue
            r = np.corrcoef(x[mask], y_vals[mask])[0, 1] if mask.sum() > 2 else np.nan
            if np.isfinite(r):
                corr_pairs.append((in_col.replace("in_", ""), r))
            fig, ax = plt.subplots(figsize=(7.5, 5))
            ax.scatter(x[mask], y_vals[mask], s=24, alpha=0.7)
            title = f"{in_col.replace('in_', '')} vs {selected_out}"
            if np.isfinite(r):
                title += f"  (r={r:.3f})"
            ax.set_title(title)
            ax.set_xlabel(in_col.replace("in_", ""))
            ax.set_ylabel(y_label)
            ax.grid(True, alpha=0.35)
            save(fig, f"fig{fig_no:02d}_scatter_{_safe_plot_token(in_col.replace('in_', ''))}", pdf)
            fig_no += 1

        if corr_pairs:
            corr_pairs = sorted(corr_pairs, key=lambda p: abs(p[1]), reverse=True)
            labels = [p[0] for p in corr_pairs]
            vals = [p[1] for p in corr_pairs]
            fig_h = max(4.5, 0.35 * len(labels) + 1.5)
            fig, ax = plt.subplots(figsize=(8, fig_h))
            ax.barh(labels[::-1], vals[::-1])
            ax.axvline(0.0, linewidth=1.0)
            ax.set_xlim(-1, 1)
            ax.set_xlabel("Pearson correlation coefficient, r")
            ax.set_title(f"Pearson Ranking — {selected_out}")
            ax.grid(True, axis="x", alpha=0.35)
            save(fig, f"fig{fig_no:02d}_pearson_ranking", pdf)

    return {"pdf": str(pdf_path), "pngs": pngs, "out_dir": str(out_dir)} if pngs else None

def main(config_path: Path) -> int:
    cfg = json.loads(config_path.read_text(encoding="utf-8"))
    work_dir = Path(cfg["work_dir"])
    run_dir  = Path(cfg["run_dir"])
    base_case = cfg["base_case"]
    # Path to the selected base input workbook.  Prefer the explicit path
    # written by the recursive-inputs UI, but tolerate older configs by
    # searching non-generated subfolders for <base_case>_in.xlsx.
    base_input_path = resolve_base_input_path(
        work_dir, base_case, cfg.get("base_input_path")
    )
    active_vars = cfg["active_vars"]
    n_samples = int(cfg["n_samples"])
    fast_mode = bool(cfg.get("fast_mode", True))
    timeout_s = int(cfg.get("timeout_s", 600))

    run_dir.mkdir(parents=True, exist_ok=True)
    status_path = run_dir / STATUS_FILE

    if not base_input_path.exists():
        raise FileNotFoundError(f"Base input workbook not found: {base_input_path}")

    status = {
        "status": "starting",
        "message": "UA worker starting.",
        "base_case": base_case,
        "base_input_path": str(base_input_path),
        "run_dir": str(run_dir),
        "worker_pid": os.getpid(),
        "total_samples": n_samples,
        "completed_samples": 0,
        "failed_samples": 0,
        "current_sample": None,
        "current_pid": None,
        "current_log": None,
        "started": _now(),
        "last_update": _now(),
    }
    _write_json_atomic(status_path, status)

    print(f"[ua-worker] Started: {run_dir}", flush=True)
    print(f"[ua-worker] Base case: {base_case}; samples={n_samples}; vars={len(active_vars)}", flush=True)

    try:
        rng = np.random.default_rng()
        sample_matrix = {}
        for var, vcfg in active_vars.items():
            sample_matrix[var] = sample_values(
                vcfg["dist"], float(vcfg["p1"]), float(vcfg["p2"]),
                float(vcfg["base"]), n_samples, rng
            )

        df_samp = pd.DataFrame({"sample": range(1, n_samples + 1)})
        for var in active_vars:
            df_samp[var] = sample_matrix[var]
        df_samp.to_csv(run_dir / f"ua_{base_case}_samples.csv", index=False)

        all_results = []
        n_ok = 0
        n_fail = 0

        for i in range(1, n_samples + 1):
            if _aborted(run_dir):
                status.update({
                    "status": "aborted",
                    "message": "UA run aborted by user before starting next sample.",
                    "last_update": _now(),
                })
                _write_json_atomic(status_path, status)
                print("[ua-worker] Aborted before sample start.", flush=True)
                break

            overrides = {var: float(sample_matrix[var][i - 1]) for var in active_vars}
            tmp_case = build_ua_input(work_dir, run_dir, base_case, overrides, i, base_input_path)
            log_path = run_dir / f"ua_{base_case}_{i}_console.log"

            status.update({
                "status": "running",
                "message": f"Running sample {i} of {n_samples}.",
                "current_sample": i,
                "current_pid": None,
                "current_log": str(log_path),
                "last_update": _now(),
            })
            _write_json_atomic(status_path, status)
            print(f"[ua-worker] Running sample {i}/{n_samples}: {tmp_case}", flush=True)

            env = os.environ.copy()
            env["PYTHONUTF8"] = "1"
            env["MPLBACKEND"] = "Agg"

            cmd = [sys.executable, "-u", str(work_dir / "flare_sim.py"), tmp_case]
            if fast_mode:
                cmd.append("--no-figures")

            start_t = time.time()
            with open(log_path, "w", encoding="utf-8", errors="replace") as log:
                log.write(f"[ua-worker] Command: {' '.join(cmd)}\n")
                log.flush()
                flags = 0
                try:
                    flags = subprocess.CREATE_NEW_PROCESS_GROUP
                except AttributeError:
                    flags = 0
                proc = subprocess.Popen(
                    cmd,
                    cwd=str(run_dir),
                    stdout=log,
                    stderr=subprocess.STDOUT,
                    env=env,
                    creationflags=flags,
                )

                status.update({"current_pid": proc.pid, "last_update": _now()})
                _write_json_atomic(status_path, status)

                timed_out = False
                aborted = False
                while True:
                    rc = proc.poll()
                    elapsed = time.time() - start_t
                    if rc is not None:
                        break
                    if _aborted(run_dir):
                        aborted = True
                        _kill_tree(proc.pid)
                        break
                    if elapsed > timeout_s:
                        timed_out = True
                        _kill_tree(proc.pid)
                        break

                    # Heartbeat for UI polling.
                    status.update({
                        "status": "running",
                        "message": f"Running sample {i} of {n_samples} ({elapsed:.0f} s elapsed).",
                        "elapsed_current_s": round(elapsed, 1),
                        "last_update": _now(),
                    })
                    _write_json_atomic(status_path, status)
                    time.sleep(2.0)

                rc = proc.poll()
                log.write(f"\n[ua-worker] Return code: {rc}\n")
                if timed_out:
                    log.write(f"[ua-worker] TIMEOUT after {timeout_s} s.\n")
                if aborted:
                    log.write("[ua-worker] ABORT requested by user.\n")
                log.flush()

            move_outputs(work_dir, run_dir, tmp_case)

            log_txt = ""
            try:
                log_txt = log_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                pass

            ok = (
                not timed_out
                and not aborted
                and (rc == 0)
                and "Traceback" not in log_txt
                and "Total execution time" in log_txt
                and (run_dir / f"{tmp_case}_out.csv").exists()
            )

            scalars = extract_scalars(base_case, i, run_dir)
            scalars["sample"] = i
            scalars["status"] = "OK" if ok else ("ABORTED" if aborted else "FAIL")
            for var in active_vars:
                scalars[f"in_{var}"] = float(sample_matrix[var][i - 1])
            if not ok:
                tail_lines = [l for l in log_txt.splitlines() if l.strip()][-8:]
                scalars["error"] = " | ".join(tail_lines)

            all_results.append(scalars)

            if aborted:
                status.update({
                    "status": "aborted",
                    "message": f"UA run aborted during sample {i}.",
                    "current_pid": None,
                    "last_update": _now(),
                })
                _write_json_atomic(status_path, status)
                print("[ua-worker] Aborted during sample.", flush=True)
                break

            if ok:
                n_ok += 1
            else:
                n_fail += 1

            # Write partial results after every sample.
            df_results = pd.DataFrame(all_results)
            rename_map = {c: c[3:] for c in df_results.columns if c.startswith("in_")}
            df_results.rename(columns=rename_map).to_csv(
                run_dir / f"ua_{base_case}_results.csv", index=False
            )

            status.update({
                "status": "running",
                "message": f"Sample {i} complete: {'OK' if ok else 'FAIL'}.",
                "completed_samples": n_ok,
                "failed_samples": n_fail,
                "current_pid": None,
                "last_update": _now(),
            })
            _write_json_atomic(status_path, status)

        # Ensure final CSV exists.
        if all_results:
            df_results = pd.DataFrame(all_results)
            rename_map = {c: c[3:] for c in df_results.columns if c.startswith("in_")}
            df_results.rename(columns=rename_map).to_csv(
                run_dir / f"ua_{base_case}_results.csv", index=False
            )


        # Automatically create the default UA plot set unless this was a fast
        # no-figures run.  The UI can later create additional plot sets for the
        # currently selected variables.
        if all_results and not fast_mode and status.get("status") != "aborted":
            try:
                status.update({
                    "status": "running",
                    "message": "Creating UA default plot set.",
                    "last_update": _now(),
                })
                _write_json_atomic(status_path, status)
                create_default_ua_plots(run_dir, base_case, pd.DataFrame(all_results))
            except Exception as _plot_err:
                try:
                    with (run_dir / "ua_plot_warnings.log").open("a", encoding="utf-8", errors="replace") as _pf:
                        _pf.write(f"[{_now()}] Could not create default UA plots: {_plot_err}\n")
                except Exception:
                    pass

        if status.get("status") == "aborted":
            final_status = "aborted"
            msg = f"UA aborted: {n_ok}/{n_samples} samples completed; {n_fail} failed."
        elif n_ok > 0:
            final_status = "complete"
            msg = f"UA complete: {n_ok}/{n_samples} samples completed; {n_fail} failed."
        else:
            final_status = "failed"
            msg = f"UA failed: 0/{n_samples} samples completed; {n_fail} failed."

        status.update({
            "status": final_status,
            "message": msg,
            "completed_samples": n_ok,
            "failed_samples": n_fail,
            "current_sample": None,
            "current_pid": None,
            "current_log": None,
            "last_update": _now(),
        })
        _write_json_atomic(status_path, status)
        print(f"[ua-worker] {msg}", flush=True)
        return 0 if final_status in ("complete", "aborted") else 1

    except Exception as e:
        status.update({
            "status": "failed",
            "message": f"UA worker failed: {e}",
            "current_pid": None,
            "last_update": _now(),
        })
        _write_json_atomic(status_path, status)
        print(f"[ua-worker] FAILED: {e}", flush=True)
        return 1

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python flare_ua_worker.py <ua_worker_config.json>", file=sys.stderr)
        sys.exit(2)
    sys.exit(main(Path(sys.argv[1])))
