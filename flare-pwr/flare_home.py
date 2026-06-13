"""flare_home.py — FLARE launcher

Run with:
    streamlit run flare_home.py

This revision replaces the Streamlit st.columns-based home-page icon layout
with HTML/CSS grids so iPhone/iPad browsers do not collapse the tool icons
into one giant icon per row.  The tool routing behavior is preserved.
"""

import asyncio
import base64
import logging
import os
import sys
from pathlib import Path

import requests
import streamlit as st


# ── Quiet benign Streamlit/Tornado websocket disconnect noise ────────────────
class _FLAREWebSocketDisconnectFilter(logging.Filter):
    _BENIGN_EXC_NAMES = {"WebSocketClosedError", "StreamClosedError"}
    _BENIGN_TEXT = (
        "WebSocketClosedError",
        "StreamClosedError",
        "Stream is closed",
        "keepalive ping failed",
    )

    def filter(self, record):
        try:
            logger_name = getattr(record, "name", "")
            msg = record.getMessage()
            if record.exc_info and record.exc_info[1] is not None:
                exc_name = record.exc_info[1].__class__.__name__
                if exc_name in self._BENIGN_EXC_NAMES:
                    return False
                if logger_name.startswith("websockets") and exc_name == "AssertionError":
                    return False
            if any(token in msg for token in self._BENIGN_TEXT):
                return False
        except Exception:
            pass
        return True


def _install_flare_log_filter():
    flt = _FLAREWebSocketDisconnectFilter()
    for name in (
        "asyncio",
        "tornado",
        "tornado.application",
        "tornado.general",
        "tornado.access",
        "streamlit",
        "streamlit.runtime",
        "websockets",
        "websockets.server",
        "websockets.client",
        "websockets.legacy",
        "websockets.legacy.protocol",
    ):
        logging.getLogger(name).addFilter(flt)

    root = logging.getLogger()
    root.addFilter(flt)
    for handler in root.handlers:
        handler.addFilter(flt)

    try:
        loop = asyncio.get_event_loop()
        default_handler = loop.default_exception_handler

        def _quiet_loop_exception_handler(loop, context):
            exc = context.get("exception")
            msg = str(context.get("message", ""))
            if exc is not None and exc.__class__.__name__ in flt._BENIGN_EXC_NAMES:
                return
            if exc is not None and exc.__class__.__name__ == "AssertionError":
                if "keepalive" in msg.lower() or "websocket" in msg.lower():
                    return
            if any(token in msg for token in flt._BENIGN_TEXT):
                return
            default_handler(context)

        loop.set_exception_handler(_quiet_loop_exception_handler)
    except Exception:
        pass


_install_flare_log_filter()

WORK_DIR = Path(__file__).parent
if str(WORK_DIR) not in sys.path:
    sys.path.insert(0, str(WORK_DIR))


def _runtime_dir(base_dir: Path | None = None) -> Path:
    """Return the FLARE runtime folder, preferring runtime/ then Runtime/."""
    base = Path(base_dir) if base_dir is not None else WORK_DIR
    lower = base / "runtime"
    upper = base / "Runtime"
    if lower.exists():
        rt = lower
    elif upper.exists():
        rt = upper
    else:
        rt = lower
    rt.mkdir(parents=True, exist_ok=True)
    return rt


def _runtime_file(name: str, base_dir: Path | None = None) -> Path:
    return _runtime_dir(base_dir) / name


# ── API key / config helpers ──────────────────────────────────────────────────
def _read_config(key):
    """Read a key=value entry from flare_config.txt."""
    cfg = _runtime_file("flare_config.txt")
    if not cfg.exists():
        return None
    try:
        for line in cfg.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k.strip().lower() == key.lower():
                return v.strip().strip('"').strip("'")
    except Exception:
        return None
    return None


def _write_config_value(key, value):
    """Create/update a key=value entry in flare_config.txt, preserving other lines."""
    cfg = _runtime_file("flare_config.txt")
    lines = []
    if cfg.exists():
        try:
            lines = cfg.read_text(encoding="utf-8").splitlines()
        except Exception:
            lines = []

    out = []
    replaced = False
    for line in lines:
        if "=" in line and not line.lstrip().startswith("#"):
            k, _ = line.split("=", 1)
            if k.strip().lower() == key.lower():
                out.append(f'{key} = "{value}"')
                replaced = True
                continue
        out.append(line)

    if not replaced:
        if out and out[-1].strip():
            out.append("")
        out.append(f'{key} = "{value}"')

    cfg.write_text("\n".join(out) + "\n", encoding="utf-8")


def _delete_config_key(key):
    """Remove a key=value entry from flare_config.txt if present."""
    cfg = _runtime_file("flare_config.txt")
    if not cfg.exists():
        return
    try:
        lines = cfg.read_text(encoding="utf-8").splitlines()
        out = [
            l
            for l in lines
            if not (
                "=" in l
                and not l.lstrip().startswith("#")
                and l.split("=", 1)[0].strip().lower() == key.lower()
            )
        ]
        cfg.write_text("\n".join(out) + "\n", encoding="utf-8")
    except Exception:
        pass


def load_api_key():
    """Read ANTHROPIC_API_KEY from runtime/flare_config.txt or the environment."""
    return _read_config("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY", "")


def load_model():
    """Read ANTHROPIC_MODEL from flare_config.txt; fall back to claude-sonnet-4-5."""
    return _read_config("ANTHROPIC_MODEL") or "claude-sonnet-4-5"


_ANTHROPIC_MODEL = load_model()

st.set_page_config(
    page_title="FLARE",
    page_icon="🔥",
    layout="wide",
    initial_sidebar_state="expanded",
)

if "page" not in st.session_state:
    st.session_state.page = "home"

# Sync page from URL.
if "page" in st.query_params:
    st.session_state.page = st.query_params["page"]


# ── Routing ──────────────────────────────────────────────────────────────────
if st.session_state.page != "home":
    import streamlit as _st

    _st.set_page_config = lambda *a, **kw: None
    page = st.session_state.page
    target = {
        "simulator": "flare_ui.py",
        "ua": "flare_ua.py",
        "editor": "flare_model_editor.py",
        "analyzer": "flare_analyzer.py",
        "risk": "flare_risk.py",
    }.get(page)

    if target:
        _globals = {
            "__file__": str(WORK_DIR / target),
            "FLARE_WORK_DIR": str(WORK_DIR),
        }
        import os as _os

        _prev_cwd = _os.getcwd()
        _target_path = WORK_DIR / target
        if not _target_path.exists():
            st.error(
                f"**{target}** was not found in the FLARE folder.\n\n"
                f"Expected location: `{_target_path}`\n\n"
                "This tool may not yet be installed. Please check your FLARE installation."
            )
            if st.button("← Back to Home", key="back_home_missing"):
                st.session_state.page = "home"
                st.rerun()
        else:
            try:
                _os.chdir(str(WORK_DIR))
                exec(open(_target_path, encoding="utf-8").read(), _globals)
            except SystemExit:
                pass
            except Exception as _exc:
                st.error(
                    f"**{target}** encountered an error:\n\n"
                    f"```\n{type(_exc).__name__}: {_exc}\n```"
                )
                if st.button("← Back to Home", key="back_home_error"):
                    st.session_state.page = "home"
                    st.rerun()
            finally:
                _os.chdir(_prev_cwd)
    st.stop()


# ── Image data URIs ───────────────────────────────────────────────────────────
ICON_DIRS = [WORK_DIR / "icons", WORK_DIR / "Icons"]


def _icon_path(filename):
    """Return the first matching icon path under icons/ or Icons/."""
    for _dir in ICON_DIRS:
        _p = _dir / filename
        if _p.exists():
            return _p
    return None


def load_image(path, fmt=None):
    """Return a data URI for an image path, or None if it is unavailable."""
    if path is None:
        return None
    try:
        _fmt = (fmt or path.suffix.lstrip(".") or "png").lower()
        if _fmt == "jpg":
            _fmt = "jpeg"
        with open(path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        return f"data:image/{_fmt};base64,{encoded}"
    except FileNotFoundError:
        return None


def load_icon(filename):
    """Load a FLARE home-screen icon from icons/ or Icons/."""
    return load_image(_icon_path(filename))


IMG_SIM = load_icon("icon_sim.jpg")
IMG_UA = load_icon("icon_ua.jpg")
IMG_EDIT = load_icon("icon_editor.png")
IMG_PA = load_icon("icon_analyzer.png")
IMG_TM = load_icon("icon_theory_manual.png")
IMG_UM = load_icon("icon_user_manual.png")
IMG_RISK = load_icon("icon_risk.png")
IMG_BUDDY = load_icon("FLAREBUDDY.png")


def _manual_path(filename):
    """Return the path to a manual stored in manuals/ or Manuals/."""
    for folder in ("manuals", "Manuals"):
        p = WORK_DIR / folder / filename
        if p.exists():
            return p
    return None


def load_docx(path):
    """Return manual docx bytes from manuals/ or Manuals/, or None."""
    p = _manual_path(path)
    if p is not None:
        return p.read_bytes()
    return None


# ── Home screen styling ──────────────────────────────────────────────────────
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Share+Tech+Mono&display=swap');

.stApp {
    background: #070b13;
    color: #e8eef7;
}

.block-container {
    padding-top: 1.0rem !important;
    max-width: 1500px !important;
}

.page-bg {
    position: fixed;
    inset: 0;
    z-index: -2;
    background:
        linear-gradient(rgba(255,255,255,0.028) 1px, transparent 1px),
        linear-gradient(90deg, rgba(255,255,255,0.028) 1px, transparent 1px),
        radial-gradient(circle at 50% 20%, rgba(232,83,10,0.18), transparent 34%),
        #070b13;
    background-size: 80px 80px, 80px 80px, auto, auto;
}

.page-glow {
    position: fixed;
    inset: 0;
    z-index: -1;
    pointer-events: none;
    background: radial-gradient(circle at 50% 12%, rgba(232,83,10,0.16), transparent 32%);
}

.hero {
    text-align: center;
    margin-top: 0.1rem;
    margin-bottom: 1.0rem;
}

.flare-title {
    font-family: 'Share Tech Mono', monospace;
    font-size: clamp(4.2rem, 10vw, 9.0rem);
    letter-spacing: 0.18em;
    color: #e8530a;
    text-shadow: 0 0 30px rgba(232,83,10,0.50);
    line-height: 0.95;
}

.flare-sub {
    font-family: 'Share Tech Mono', monospace;
    text-transform: uppercase;
    letter-spacing: 0.34em;
    font-size: clamp(0.88rem, 2.0vw, 1.65rem);
    line-height: 1.8;
    margin-top: 0.7rem;
}

.orange { color: #f07623; }
.white { color: #e8eef7; }

.divider {
    width: 160px;
    height: 1px;
    margin: 1.2rem auto 0 auto;
    background: linear-gradient(90deg, transparent, #e8530a, transparent);
}

.version-tag {
    position: fixed;
    right: 1.0rem;
    bottom: 0.4rem;
    opacity: 0.20;
    font-family: 'Share Tech Mono', monospace;
    letter-spacing: 0.25em;
    color: #b6c7df;
    font-size: 0.78rem;
}

/* Forced desktop-like home icon layout.
   The wrapper scrolls sideways on narrow screens instead of letting Streamlit
   stack the columns vertically. */
.flare-grid-scroll {
    width: 100%;
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    padding: 0.2rem 0 0.8rem 0;
}

.flare-tool-grid {
    display: grid;
    grid-template-columns: repeat(5, minmax(118px, 1fr));
    gap: clamp(12px, 1.7vw, 24px);
    min-width: 680px;
    max-width: 1450px;
    margin: 1.6rem auto 1.25rem auto;
    align-items: center;
}

.flare-tool-card {
    display: block;
    text-decoration: none !important;
}

.flare-tool-card img,
.flare-manual-card img {
    width: 100%;
    height: auto;
    display: block;
    border-radius: 12px;
    border: 2px solid #1c2d44;
    transition: transform 0.15s ease, border-color 0.15s ease, box-shadow 0.15s ease;
}

.flare-tool-card:hover img,
.flare-manual-card:hover img {
    transform: scale(1.02);
    border-color: #e8530a;
    box-shadow: 0 0 28px rgba(232,83,10,0.45);
}

.flare-risk-fallback {
    border: 2px solid #1c2d44;
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
    background: #0d1a28;
    color: #e8530a;
    font-family: 'Share Tech Mono', monospace;
    font-size: 0.9rem;
    cursor: pointer;
    min-height: 120px;
    display: flex;
    align-items: center;
    justify-content: center;
}

.flare-manual-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(100px, 160px));
    gap: 22px;
    justify-content: center;
    align-items: center;
    margin: 0.6rem auto 2.0rem auto;
}

@media (min-width: 900px) {
    .flare-tool-grid {
        min-width: 0;
    }
}

@media (max-width: 700px) {
    .block-container {
        padding-left: 0.8rem !important;
        padding-right: 0.8rem !important;
    }

    .hero {
        margin-bottom: 0.4rem;
    }

    .flare-title {
        font-size: 4.2rem;
        letter-spacing: 0.12em;
    }

    .flare-sub {
        font-size: 0.82rem;
        letter-spacing: 0.22em;
    }

    .flare-tool-grid {
        grid-template-columns: repeat(5, 126px);
        min-width: calc(5 * 126px + 4 * 12px);
        gap: 12px;
        justify-content: start;
        margin-top: 1.0rem;
    }

    .flare-manual-grid {
        grid-template-columns: repeat(2, 118px);
        gap: 18px;
    }
}

.flare-assistant-title {
    font-family:'Share Tech Mono', monospace;
    color:#e8530a;
    text-transform: uppercase;
    letter-spacing:0.25em;
    margin-top:1.0rem;
    margin-bottom:0.35rem;
}

.flare-user-msg {
    background:#112033;
    border-left:3px solid #e8530a;
    padding:0.65rem 0.85rem;
    border-radius:8px;
    margin:0.4rem 0;
}

.flare-assistant-msg {
    background:#0c1624;
    border-left:3px solid #5a7a99;
    padding:0.65rem 0.85rem;
    border-radius:8px;
    margin:0.4rem 0;
}
</style>
<div class="page-bg"></div>
<div class="page-glow"></div>
""",
    unsafe_allow_html=True,
)

_buddy_hero_img = (
    f"<img src='{IMG_BUDDY}' style='"
    "height:clamp(6.0rem,14vw,10rem);width:auto;object-fit:contain;"
    "filter:drop-shadow(0 0 18px rgba(232,83,10,0.45));flex-shrink:0;'/>"
    if IMG_BUDDY
    else ""
)

st.markdown(
    f"""
<div class="hero">
  <div style="display:flex;align-items:center;justify-content:center;gap:1.2rem;flex-wrap:wrap;">
    <div class="flare-title">FLARE</div>
    {_buddy_hero_img}
  </div>
  <div class="flare-sub">
    <span class="orange">Fast Licensing</span>
    <span class="white"> Accident Response Engine</span><br>
    <span class="white">by Robert P. Martin</span>
  </div>
  <div class="divider"></div>
</div>
<div class="version-tag">FLARE &middot; PWR/SMR Safety Analysis</div>
""",
    unsafe_allow_html=True,
)


# ── Clickable Image Buttons ──────────────────────────────────────────────────
_TOOL_TIPS = {
    "simulator": "PWR Simulator — run design-basis accident simulations",
    "ua": "Uncertainty Analysis — Monte Carlo parameter sensitivity",
    "editor": "Model Editor — create and edit FLARE input decks",
    "analyzer": "Plant Analyzer — replay and compare simulation results",
    "risk": "Risk Assessment — probabilistic safety screening",
}


def image_link(img_base64, page, label):
    """Return a clickable icon card for the home-page CSS grid."""
    if img_base64 is None:
        if page == "risk":
            return (
                f'<a href="?page={page}" target="_self" class="flare-tool-card" '
                f'title="{_TOOL_TIPS.get(page, label)}">'
                f'<div class="flare-risk-fallback">⚛️<br><br>RISK<br>ASSESSMENT</div></a>'
            )
        return ""
    tip = _TOOL_TIPS.get(page, label)
    return f"""
    <a href="?page={page}" target="_self" class="flare-tool-card" title="{tip}">
        <img src="{img_base64}" alt="{label}"/>
    </a>
    """


st.markdown(
    f"""
<div class="flare-grid-scroll" aria-label="FLARE tool launcher">
  <div class="flare-tool-grid">
    {image_link(IMG_SIM, "simulator", "PWR Simulator")}
    {image_link(IMG_UA, "ua", "Uncertainty Analysis")}
    {image_link(IMG_EDIT, "editor", "Model Editor")}
    {image_link(IMG_PA, "analyzer", "Plant Analyzer")}
    {image_link(IMG_RISK, "risk", "Risk Assessment")}
  </div>
</div>
""",
    unsafe_allow_html=True,
)


# ── FLARE Assistant support ──────────────────────────────────────────────────
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []


def _manuals_dir():
    """Return the Manuals/manuals folder if present, else None."""
    for _name in ("manuals", "Manuals"):
        _p = WORK_DIR / _name
        if _p.exists() and _p.is_dir():
            return _p
    return None


def _load_corpus_file(fname):
    """Load FLARE Assistant corpus text from Manuals/manuals only."""
    _mdir = _manuals_dir()
    if _mdir is None:
        return f"[{fname} not found: Manuals/manuals folder not found]"
    p = _mdir / fname
    try:
        return p.read_text(encoding="utf-8") if p.exists() else f"[{fname} not found in {_mdir.name}/]"
    except Exception as e:
        return f"[{fname} load error from {_mdir.name}/: {e}]"


def _build_system_prompt():
    """Build system prompt, loading reference content from files at runtime."""
    base = """You are the FLARE Assistant — an expert on the FLARE (Fast Licensing Accident Response Engine) nuclear reactor safety analysis suite. FLARE is a Python/Streamlit application for PWR and SMR design-basis accident analysis. It includes PWR Simulator, Uncertainty Analysis, Model Editor, Plant Analyzer, and Risk Assessment tools.

Answer only questions related to FLARE, its models, its input/output, its manuals, or its use. If a question is not related to FLARE, respond exactly: "That question is not relevant to FLARE. I can only help with questions about the FLARE safety analysis suite."

Be specific and technical. Reference parameter names, UI control names, and manual sections where helpful. Answer in 1-2 paragraphs unless the user asks for more detail.
"""
    return (
        base
        + "\n\n=== FLARE CODE THEORY MANUAL ===\n\n"
        + _load_corpus_file("flare_theory_manual.txt")
        + "\n\n=== FLARE USER'S MANUAL ===\n\n"
        + _load_corpus_file("flare_users_manual.txt")
        + "\n\n=== DBA BOOK CONTENT ===\n\n"
        + _load_corpus_file("flare_dba_book.txt")
    )


_SYSTEM_PROMPT = _build_system_prompt()


def _summarize_flare_xlsx_context(file_bytes, file_name):
    """Return a compact technical summary of a FLARE input workbook."""
    try:
        from io import BytesIO
        import re as _re
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        ws_title = ws.title
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        cmd_lines = []
        time_row_idx = None
        for i, row in enumerate(rows, start=1):
            v = row[0] if row else None
            if isinstance(v, str) and v.strip().lower().startswith("time"):
                time_row_idx = i
                break
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                time_row_idx = i
                break
            if isinstance(v, str) and v.strip():
                cmd_lines.append(v.strip())

        params = []
        for line in cmd_lines:
            if line.lstrip().startswith("#"):
                continue
            m = _re.match(r"^\s*([A-Za-z_]\w*)\s*=\s*(.+?)\s*(?:#.*)?$", line)
            if m:
                params.append((m.group(1), m.group(2).strip()))

        lines = [
            f"Uploaded FLARE input workbook: {file_name}",
            f"Worksheet: {ws_title}",
            f"Command-block lines: {len(cmd_lines)}",
            f"Parsed parameter assignments: {len(params)}",
        ]
        if params:
            lines.append("Parsed parameters:")
            for k, v in params[:140]:
                lines.append(f" - {k} = {v}")

        if time_row_idx is not None:
            header_idx = time_row_idx
            if time_row_idx > 1:
                prev = rows[time_row_idx - 2]
                if prev and isinstance(prev[0], str) and "time" in prev[0].lower():
                    header_idx = time_row_idx - 1
            data_rows = rows[header_idx:] if header_idx < len(rows) else []
            numeric_rows = [
                r for r in data_rows if r and isinstance(r[0], (int, float)) and not isinstance(r[0], bool)
            ]
            if numeric_rows:
                times = [float(r[0]) for r in numeric_rows]
                lines.append(f"Time table rows: {len(numeric_rows)}")
                lines.append(f"Time range: {min(times):.6g} to {max(times):.6g} s")

        return "\n".join(lines)
    except Exception as e:
        return f"Uploaded workbook {file_name} could not be parsed: {e}"


def _summarize_flare_csv_context(file_bytes, file_name):
    """Return a compact summary of a FLARE output or diagnostic CSV file."""
    try:
        from io import BytesIO
        import numpy as _np
        import pandas as _pd

        df = _pd.read_csv(BytesIO(file_bytes))
        lines = [
            f"Uploaded FLARE CSV file: {file_name}",
            f"Rows: {len(df)}; columns: {len(df.columns)}",
            "Columns: " + ", ".join(str(c) for c in df.columns[:90]),
        ]
        if "Time (s)" in df.columns:
            t = _pd.to_numeric(df["Time (s)"], errors="coerce").dropna()
            if not t.empty:
                lines.append(f"Time range: {t.min():.6g} to {t.max():.6g} s")

        summary_specs = [
            ("RCS Pressure (kPa)", "minmax"),
            ("RCS Temperature (K)", "minmax"),
            ("RK Total Power (MW)", "max"),
            ("Core Power (MW)", "max"),
            ("SG Heat Removal (MW)", "max"),
            ("Pump Mass Flow (kg/s)", "minmax"),
            ("Break Flow (kg/s)", "max"),
            ("PORV Mass Flow (kg/s)", "max"),
            ("Accumulator Flow (kg/s)", "max"),
            ("HPSI Flow (kg/s)", "max"),
            ("LPSI Flow (kg/s)", "max"),
            ("SI Pumped Total (kg/s)", "max"),
            ("Vessel Level (m)", "minmax"),
            ("Accumulator Level (m)", "minmax"),
            ("DNBR", "min"),
            ("Clad Surface Temp (K)", "max"),
            ("Hot Pin Clad Temp (K)", "max"),
            ("Hot Pin Fuel Temp (K)", "max"),
            ("H2 Generated (kg)", "max"),
            ("UA_sg_rated Used (MW/K)", "minmax"),
        ]
        for col, mode in summary_specs:
            if col not in df.columns:
                continue
            s = _pd.to_numeric(df[col], errors="coerce").replace([_np.inf, -_np.inf], _np.nan).dropna()
            if s.empty:
                continue
            if mode == "max":
                idx = s.idxmax()
                ttxt = f" at t={float(df.loc[idx, 'Time (s)']):.6g} s" if "Time (s)" in df.columns else ""
                lines.append(f"{col}: max={s.max():.6g}{ttxt}; final={s.iloc[-1]:.6g}")
            elif mode == "min":
                idx = s.idxmin()
                ttxt = f" at t={float(df.loc[idx, 'Time (s)']):.6g} s" if "Time (s)" in df.columns else ""
                lines.append(f"{col}: min={s.min():.6g}{ttxt}; final={s.iloc[-1]:.6g}")
            else:
                lines.append(
                    f"{col}: initial={s.iloc[0]:.6g}; min={s.min():.6g}; "
                    f"max={s.max():.6g}; final={s.iloc[-1]:.6g}"
                )
        return "\n".join(lines)
    except Exception as e:
        return f"Uploaded CSV {file_name} could not be parsed: {e}"


def _summarize_flare_text_context(text, source_name="pasted text"):
    """Return a compact summary of pasted FLARE command-block or CSV text."""
    import re as _re

    raw = str(text or "")
    lines = [f"User-supplied FLARE text context: {source_name}", f"Characters: {len(raw)}"]
    params = []
    for line in raw.splitlines():
        m = _re.match(r"^\s*([A-Za-z_]\w*)\s*=\s*(.+?)\s*(?:#.*)?$", line)
        if m:
            params.append((m.group(1), m.group(2).strip()))
    if params:
        lines.append(f"Parsed parameter assignments: {len(params)}")
        for k, v in params[:140]:
            lines.append(f" - {k} = {v}")
    excerpt = raw.strip()[:6000]
    if excerpt:
        lines.append("Text excerpt:")
        lines.append(excerpt)
    return "\n".join(lines)


def _build_uploaded_file_context(uploaded_file):
    """Classify and summarize uploaded FLARE input/output context."""
    if uploaded_file is None:
        return ""
    name = getattr(uploaded_file, "name", "uploaded file")
    suffix = Path(name).suffix.lower()
    data = uploaded_file.getvalue()
    if suffix == ".xlsx":
        return _summarize_flare_xlsx_context(data, name)
    if suffix == ".csv":
        return _summarize_flare_csv_context(data, name)
    try:
        text = data.decode("utf-8", errors="replace")
    except Exception:
        text = str(data[:6000])
    return _summarize_flare_text_context(text, name)


def _augment_history_with_file_context(history, file_context):
    """Attach current FLARE file context to the most recent user message."""
    if not file_context:
        return history
    msgs = [dict(m) for m in history]
    for i in range(len(msgs) - 1, -1, -1):
        if msgs[i].get("role") == "user":
            msgs[i]["content"] = (
                "USER-SUPPLIED FLARE FILE CONTEXT\n"
                "Use this context only to answer the user's current question. "
                "Do not treat it as part of the permanent manuals.\n\n"
                f"{file_context}\n\n"
                "USER QUESTION\n"
                f"{msgs[i].get('content', '')}"
            )
            return msgs
    return msgs


def call_claude(history, api_key="", file_context=""):
    """Call the Anthropic API with the full conversation history."""
    if not api_key:
        return "⚠️ No API key. Enter an Anthropic API key in the password field on the FLARE home page."
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": _ANTHROPIC_MODEL,
                "max_tokens": 1024,
                "system": _SYSTEM_PROMPT,
                "messages": _augment_history_with_file_context(history, file_context),
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        return data["content"][0]["text"]
    except requests.exceptions.Timeout:
        return "⚠️ Request timed out. Please try again."
    except Exception as e:
        return f"⚠️ API error: {e}"


# ── Manual icons — second row, centred ───────────────────────────────────────
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def manual_icon_html(img_b64, docx_name, label):
    """Return HTML for a clickable icon that downloads a docx file."""
    docx_bytes = load_docx(docx_name)
    img_src = img_b64 or ""
    if docx_bytes and img_b64:
        b64_doc = base64.b64encode(docx_bytes).decode()
        href = f"data:{DOCX_MIME};base64,{b64_doc}"
        return (
            f'<a href="{href}" download="{docx_name}" target="_self" '
            f'class="flare-manual-card" title="{label}">'
            f'<img src="{img_src}" alt="{label}"/></a>'
        )
    if img_b64:
        return (
            f'<div class="flare-manual-card">'
            f'<img src="{img_src}" alt="{label}" style="opacity:0.4;"/>'
            f'<p style="color:#5a7a99;font-size:0.75rem;margin-top:0.3rem;">'
            f'{docx_name} not found in manuals/ or Manuals/.</p></div>'
        )
    return ""


st.markdown(
    f"""
<div class="flare-manual-grid" aria-label="FLARE manuals">
    {manual_icon_html(IMG_TM, "FLARE_Code_Theory_Manual.docx", "Theory Manual")}
    {manual_icon_html(IMG_UM, "FLARE_Users_Manual.docx", "User's Manual")}
</div>
""",
    unsafe_allow_html=True,
)


# ── FLARE Assistant UI ───────────────────────────────────────────────────────
st.markdown(
    (
        '<div class="flare-assistant-title">'
        + (f'<img src="{IMG_BUDDY}" style="height:1.05rem;vertical-align:-0.2rem;margin-right:0.35rem;">' if IMG_BUDDY else "⚛️ ")
        + "FLARE Assistant — ask a question about FLARE</div>"
    ),
    unsafe_allow_html=True,
)

for msg in st.session_state.chat_history:
    if msg.get("role") == "user":
        st.markdown(f'<div class="flare-user-msg">{msg.get("content", "")}</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="flare-assistant-msg">{msg.get("content", "")}</div>', unsafe_allow_html=True)

if "user_api_key" not in st.session_state:
    st.session_state.user_api_key = load_api_key() or ""
if "api_key_input_counter" not in st.session_state:
    st.session_state.api_key_input_counter = 0
if "chat_pending" not in st.session_state:
    st.session_state.chat_pending = False
if "flare_file_context" not in st.session_state:
    st.session_state.flare_file_context = ""
if "flare_context_upload_counter" not in st.session_state:
    st.session_state.flare_context_upload_counter = 0

_active_key = st.session_state.user_api_key.strip()

_FAQS = [
    "— Select a frequently asked question —",
    "What is FLARE and what types of analyses does it support?",
    "How do I set up a new input case in FLARE?",
    "What is the structure of the FLARE input Excel file?",
    "How do I run a Large-Break LOCA case?",
    "How do I model a boron dilution event in FLARE?",
    "What does the core_flag parameter control?",
    "What is the difference between Core Power and RK Total Power in the output?",
    "How does FLARE calculate DNBR?",
    "How do I perform an uncertainty analysis in FLARE?",
    "What output files does FLARE produce?",
    "What are the four Condition categories for design-basis events?",
    "What acceptance criteria apply to the PWR large-break LOCA?",
    "What is the Evaluation Model approach to safety analysis?",
    "What is the difference between deterministic and BEPU methods?",
    "How is two-phase critical flow modelled in PWR LOCA analysis?",
    "What reactivity feedback mechanisms are important during a PWR ATWS?",
    "How is peak cladding temperature calculated during a LOCA?",
    "What is the role of ECCS in mitigating a large-break LOCA?",
    "What is the single-failure criterion and how does it apply to safety analysis?",
    "How is decay heat modelled in design-basis accident analysis?",
]

if "faq_prev" not in st.session_state:
    st.session_state.faq_prev = _FAQS[0]

_faq_choice = st.selectbox(
    "Frequently asked questions",
    _FAQS,
    index=0,
    key="faq_dropdown",
    label_visibility="collapsed",
    help="Select a question — it will be submitted automatically",
)

if (
    _faq_choice != _FAQS[0]
    and _faq_choice != st.session_state.faq_prev
    and _active_key
    and not st.session_state.chat_pending
):
    st.session_state.faq_prev = _faq_choice
    st.session_state.chat_history.append({"role": "user", "content": _faq_choice})
    with st.spinner("⚛️ FLARE Assistant is thinking…"):
        _faq_answer = call_claude(
            st.session_state.chat_history,
            api_key=_active_key,
            file_context=st.session_state.get("flare_file_context", ""),
        )
    st.session_state.chat_history.append({"role": "assistant", "content": _faq_answer})
    st.rerun()

st.markdown(
    "Optional: attach a FLARE input workbook, output CSV, diagnostic CSV, or pasted command block so the Assistant can answer questions about that specific case.",
    unsafe_allow_html=True,
)

with st.expander("Optional FLARE file context", expanded=False):
    _ctx_upload = st.file_uploader(
        "Upload FLARE input/output file",
        type=["xlsx", "csv", "txt"],
        key=f"flare_context_upload_{st.session_state.flare_context_upload_counter}",
        help="Upload a FLARE *_in.xlsx input file, *_out.csv output, *_diag.csv diagnostic file, or text export.",
    )
    _pasted_ctx = st.text_area(
        "Or paste FLARE command-block / CSV text",
        height=150,
        key="flare_context_paste",
    )

    _ctx_parts = []
    if _ctx_upload is not None:
        _ctx_parts.append(_build_uploaded_file_context(_ctx_upload))
    if _pasted_ctx.strip():
        _ctx_parts.append(_summarize_flare_text_context(_pasted_ctx, "pasted FLARE text"))

    if _ctx_parts:
        st.session_state.flare_file_context = "\n\n---\n\n".join(_ctx_parts)
        st.success("FLARE file context attached for the Assistant.")
        with st.expander("Preview parsed context", expanded=False):
            st.code(st.session_state.flare_file_context[:9000], language="text")
    elif st.session_state.flare_file_context:
        st.info("Using previously attached FLARE file context.")
        with st.expander("Preview parsed context", expanded=False):
            st.code(st.session_state.flare_file_context[:9000], language="text")
    else:
        st.caption("No file context attached.")

    if st.button("Clear attached file context", key="clear_flare_file_context"):
        st.session_state.flare_file_context = ""
        st.session_state.flare_context_upload_counter += 1
        st.rerun()

if "chat_input_counter" not in st.session_state:
    st.session_state.chat_input_counter = 0
if "chat_enter_pressed" not in st.session_state:
    st.session_state.chat_enter_pressed = False


def _on_chat_enter():
    _key = f"chat_input_{st.session_state.chat_input_counter}"
    if st.session_state.get(_key, "").strip():
        st.session_state.chat_enter_pressed = True


_qcol, _bcol = st.columns([10, 1])
with _qcol:
    user_input = st.text_input(
        label="flare_question",
        label_visibility="collapsed",
        placeholder="e.g., What is FLARE?",
        key=f"chat_input_{st.session_state.chat_input_counter}",
        on_change=_on_chat_enter,
    )
with _bcol:
    send_btn = st.button("Ask", type="primary", key="chat_send", disabled=not bool(_active_key))

if st.session_state.chat_history:
    if st.button("Clear conversation", key="chat_clear"):
        st.session_state.chat_history = []
        st.session_state.chat_input_counter += 1
        st.rerun()

if (send_btn or st.session_state.chat_enter_pressed) and user_input.strip() and not st.session_state.chat_pending:
    st.session_state.chat_enter_pressed = False
    st.session_state.chat_pending = True
    q = user_input.strip()
    st.session_state.chat_history.append({"role": "user", "content": q})
    st.session_state.chat_input_counter += 1
    with st.spinner("⚛️ FLARE Assistant is thinking…"):
        answer = call_claude(
            st.session_state.chat_history,
            api_key=_active_key,
            file_context=st.session_state.get("flare_file_context", ""),
        )
    st.session_state.chat_history.append({"role": "assistant", "content": answer})
    st.session_state.chat_pending = False
    st.rerun()


# ── API key field — final page object ────────────────────────────────────────
st.markdown('<div class="flare-assistant-title">Anthropic API Key</div>', unsafe_allow_html=True)


def _sync_api_key_from_input():
    current_key = f"api_key_input_{st.session_state.api_key_input_counter}"
    key = str(st.session_state.get(current_key, "") or "").strip()
    st.session_state.user_api_key = key
    if key:
        try:
            _write_config_value("ANTHROPIC_API_KEY", key)
        except Exception:
            pass
    else:
        _delete_config_key("ANTHROPIC_API_KEY")


_current_api_key_widget = f"api_key_input_{st.session_state.api_key_input_counter}"
_key_col, _clear_col, _hint_col = st.columns([3, 1.2, 2])
with _key_col:
    st.text_input(
        "Anthropic API key",
        label_visibility="collapsed",
        type="password",
        placeholder="sk-ant-...",
        key=_current_api_key_widget,
        on_change=_sync_api_key_from_input,
        help="Enter your Anthropic API key for this active browser session.",
    )
with _clear_col:
    if st.button("Forget key", key="clear_api_key_btn", width="stretch"):
        _delete_config_key("ANTHROPIC_API_KEY")
        st.session_state.user_api_key = ""
        st.session_state.api_key_input_counter += 1
        st.rerun()
with _hint_col:
    if st.session_state.user_api_key.strip():
        st.markdown("✅ &nbsp;API key active", unsafe_allow_html=True)
    else:
        st.markdown("⚠️ &nbsp;No API key — Ask button disabled", unsafe_allow_html=True)
