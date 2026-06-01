# flare_model_editor.py
# FLARE Model Editor with Validator and subfolder input-file discovery
#
# Supports FLARE input files stored in subfolders below the FLARE root.
# Root-level *_in.xlsx files are intentionally ignored to match the current
# FLARE project organization.

import streamlit as st
from openpyxl import load_workbook
from pathlib import Path
import shutil
import re

st.set_page_config(layout="wide", page_title="FLARE Model Editor")

WORK_DIR = Path(__file__).parent

# ── Input discovery helpers ──────────────────────────────────────────────────

_EXCLUDE_DIR_PREFIXES = (
    "sim_",
    "risk_",
    "ua_",
    ".sim_all_",
)
_EXCLUDE_DIR_NAMES = {
    "__pycache__",
    "runtime",
    "Runtime",
    "icons",
    "Icons",
    "manuals",
    "Manuals",
    "install",
    ".git",
    ".streamlit",
}


def _is_generated_or_support_path(path: Path) -> bool:
    """Return True if path is inside a generated/support folder to exclude."""
    try:
        rel_parts = path.relative_to(WORK_DIR).parts
    except Exception:
        rel_parts = path.parts

    for part in rel_parts[:-1]:
        if part in _EXCLUDE_DIR_NAMES:
            return True
        if any(part.startswith(prefix) for prefix in _EXCLUDE_DIR_PREFIXES):
            return True
    return False


def find_case_files():
    """
    Find FLARE input files in subfolders below WORK_DIR.

    Root-level *_in.xlsx files are intentionally ignored.  Generated folders
    such as sim_*, risk_*, ua_*, and .sim_all_* are also ignored.
    """
    found = []
    for p in WORK_DIR.rglob("*_in.xlsx"):
        if p.parent == WORK_DIR:
            continue
        if p.name.startswith(".~") or p.stem.startswith("ua_"):
            continue
        if _is_generated_or_support_path(p):
            continue
        found.append(p)

    # Deduplicate and sort by case name, then folder.
    found = sorted(set(found), key=lambda x: (x.stem.lower(), str(x.parent).lower()))
    return found


def case_label(path: Path) -> str:
    case_name = path.stem.replace("_in", "")
    rel_dir = path.parent.relative_to(WORK_DIR)
    return f"{case_name}  —  {rel_dir}"


# ── Styling ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
section[data-testid="stSidebar"] { background: #1a1f2e !important; }
section[data-testid="stSidebar"] button {
    background-color: #2a3145 !important;
    color: #e6edf3 !important;
    border: 1px solid #3b435c !important;
    border-radius: 8px !important;
}
section[data-testid="stSidebar"] button:hover {
    background-color: #343c55 !important;
    box-shadow: 0 0 6px rgba(100,150,255,0.6);
}
section[data-testid="stSidebar"] * { color: #e6edf3 !important; }
/* Exclude selectbox internals from the wildcard so Streamlit renders its own default colors */
section[data-testid="stSidebar"] [data-baseweb="select"] *,
section[data-testid="stSidebar"] [data-baseweb="select"] input {
    color: unset !important;
}

/* Force dark text inside selectbox controls */
section[data-testid="stSidebar"] [data-baseweb="select"] span,
section[data-testid="stSidebar"] [data-baseweb="select"] div,
section[data-testid="stSidebar"] [data-baseweb="select"] input,
section[data-testid="stSidebar"] [data-baseweb="select"] * {
    color: #111111 !important;
    -webkit-text-fill-color: #111111 !important;
}

/* Selected value background */
section[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: #ffffff !important;
}

/* Dropdown menu items */
div[role="listbox"] *,
ul[role="listbox"] * {
    color: #111111 !important;
    -webkit-text-fill-color: #111111 !important;
}
/* Caption code blocks: transparent bg so path doesn't appear as a dark block */
section[data-testid="stSidebar"] code {
    background: transparent !important;
}

/* ── Editable input fields: visible border + tinted background ── */
div[data-testid="stTextInput"] input {
    background-color: #f0f4ff !important;
    border: 1.5px solid #7090d0 !important;
    border-radius: 5px !important;
    color: #0d1117 !important;
    font-family: monospace !important;
}
div[data-testid="stTextInput"] input:focus {
    border-color: #3366cc !important;
    background-color: #e8eeff !important;
    box-shadow: 0 0 0 2px rgba(51,102,204,0.25) !important;
}
</style>
""", unsafe_allow_html=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def load_rows(path, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    # Use named sheet if given; otherwise always use the first sheet.
    # wb.active reflects whichever tab was last selected when saved — that
    # may be FLARECON rather than the RCS input sheet.
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_a = row[0] if row else None
        # Collect the last non-empty text value to the right of col A as a tooltip
        # fallback (typically a Description column further right than Value/Units).
        right_candidates = [
            str(v).strip() for v in row[1:]
            if v is not None and str(v).strip()
            and not str(v).strip().lstrip("-").replace(".", "", 1).isdigit()
        ]
        right_text = right_candidates[-1] if right_candidates else ""
        rows.append((i, col_a, right_text))
    wb.close()
    return rows


def has_flarecon_sheet(path):
    """Return the FLARECON sheet name (case-insensitive) or None."""
    wb = load_workbook(path, read_only=True, data_only=True)
    match = next((s for s in wb.sheetnames if s.upper() == "FLARECON"), None)
    wb.close()
    return match


def _is_section_label(stripped):
    """
    Return True if a line looks like a section header label.

    Accepts both:
      - '#'-prefixed headers used in FLARE RCS input  (e.g. '# Job Control')
      - Plain ALL-CAPS or Title-Case labels used in FLARECON input
        (e.g. 'DESIGN PARAMETERS', 'Initial Conditions')
    A line is NOT a section label if it contains '=' (it is a key=value pair)
    or if it looks like a data value (numeric, boolean, or quoted string).
    """
    if "=" in stripped:
        return False
    if stripped.startswith("#"):
        return True
    # Reject bare numbers, booleans, quoted strings, and bracketed units
    first_tok = stripped.split()[0] if stripped.split() else ""
    try:
        float(first_tok)
        return False
    except ValueError:
        pass
    if first_tok.lower() in ("true", "false"):
        return False
    if stripped.startswith(("'", '"')):
        return False
    # Reject table-header lines like 'time_spray [s]' — they contain '[' but no '='
    # We still want to allow them to be skipped gracefully; they are not section labels.
    if stripped.startswith(("time_", "q_", "mdot_")) and ("[" in stripped or not stripped.replace("_", "").replace(" ", "").isalpha()):
        return False
    return True


def build_sections(rows, skip_banner=False):
    sections, current = [], None
    banner_lines = []   # list of (row_number, text)
    prev_blank, in_banner = True, not skip_banner

    for i, row_tuple in enumerate(rows):
        r, text = row_tuple[0], row_tuple[1]
        right_text = row_tuple[2] if len(row_tuple) > 2 else ""
        raw = "" if text is None else str(text)
        stripped = raw.strip()

        if stripped == "":
            prev_blank = True
            continue

        # lookahead: find next non-blank line and whether there's a section label between
        next_nonblank, intervening_label = None, False
        for j in range(i + 1, len(rows)):
            nxt = rows[j][1]
            if nxt is None:
                continue
            nxt = str(nxt).strip()
            if not nxt:
                continue
            if _is_section_label(nxt):
                intervening_label = True
                break
            next_nonblank = nxt
            break

        if in_banner:
            if stripped.startswith("#"):
                if next_nonblank and "=" in next_nonblank and not intervening_label:
                    in_banner = False
                else:
                    banner_lines.append((r, stripped[1:].strip()))
                    prev_blank = False
                    continue
            else:
                in_banner = False

        if _is_section_label(stripped) and prev_blank:
            label = stripped[1:].strip() if stripped.startswith("#") else stripped
            current = {"name": label, "vars": []}
            sections.append(current)
            prev_blank = False
            continue

        if "=" in stripped:
            if current is None:
                current = {"name": "Unlabeled (Top Block)", "vars": []}
                sections.append(current)

            key, val = [x.strip() for x in stripped.split("=", 1)]
            # Split value from inline comment (e.g. "1500    # end time [s]")
            val_parts = val.split("#", 1)
            val_clean = val_parts[0].strip()
            # Prefer inline comment; fall back to any text in columns to the right
            tooltip = val_parts[1].strip() if len(val_parts) > 1 else right_text
            try:
                parsed = float(val_clean)
            except Exception:
                parsed = val_clean

            current["vars"].append({"row": r, "key": key, "value": parsed, "tooltip": tooltip})

        prev_blank = False

    banner_text = "\n".join(t for _, t in banner_lines)
    return banner_text, banner_lines, sections


def validate(rows):
    issues, prev_blank, in_banner = [], True, True
    for row_tuple in rows:
        r, text = row_tuple[0], row_tuple[1]
        raw = "" if text is None else str(text)
        stripped = raw.strip()

        if stripped == "":
            prev_blank = True
            continue

        if in_banner:
            if stripped.startswith("#"):
                prev_blank = False
                continue
            else:
                in_banner = False

        if stripped.startswith("#") and not prev_blank:
            issues.append(f"Line {r}: '#' not preceded by blank line")

        prev_blank = False
    return issues


def apply_updates(path, updates, banner_rows, sheet=None):
    wb = load_workbook(path)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]

    # Write back banner lines to their original row numbers
    for row_num, line in banner_rows:
        ws.cell(row=row_num, column=1).value = "# " + line

    for row, key, new_val, tooltip in updates:
        cell = ws.cell(row=row, column=1)
        if isinstance(cell.value, str) and "=" in cell.value:
            # Rebuild as "key = value    # tooltip" preserving the comment.
            new_cell = f"{key} = {new_val}"
            if tooltip:
                new_cell += f"    # {tooltip}"
            cell.value = new_cell

    wb.save(path)
    wb.close()
    return path


# ── FLARE Home button ────────────────────────────────────────────────────────
st.sidebar.markdown("""
    <style>
    [data-testid="stSidebar"] button[kind="secondary"] {
        background: transparent !important;
        border: 1px solid #e8530a !important;
        border-radius: 4px !important;
        color: #f97316 !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.08em !important;
        font-weight: 700 !important;
    }
    [data-testid="stSidebar"] button[kind="secondary"]:hover {
        background: rgba(232,83,10,0.18) !important;
        box-shadow: 0 0 14px rgba(232,83,10,0.45) !important;
        color: #ffffff !important;
    }
    </style>""", unsafe_allow_html=True)
if st.sidebar.button("\U0001f525  FLARE Home", key="home_btn", width="stretch"):
    st.session_state.page = "home"
    st.query_params.clear()
    st.rerun()
st.sidebar.divider()

# ── UI Modes ─────────────────────────────────────────────────────────────────
mode = st.sidebar.radio("Mode", ["Model Editor", "Validator", "New Model"])

case_files = find_case_files()

if not case_files:
    st.error(
        "No *_in.xlsx files found in FLARE subfolders.\n\n"
        "Input files are now expected in subfolders below the FLARE root. "
        "Root-level input files and generated folders such as sim_*, risk_*, "
        "ua_*, and .sim_all_* are ignored."
    )
    st.stop()

label_to_path = {case_label(p): p for p in case_files}
selected_label = st.sidebar.selectbox("Select Case", list(label_to_path.keys()))
file_path = label_to_path[selected_label]
rows = load_rows(file_path)

st.sidebar.markdown("---")
st.sidebar.caption(f"Selected file:\n`{file_path.relative_to(WORK_DIR)}`")

with open(file_path, "rb") as _fh:
    st.sidebar.download_button(
        label="⬇ Download Model",
        data=_fh,
        file_name=file_path.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── NEW MODEL MODE ───────────────────────────────────────────────────────────
if mode == "New Model":
    st.title("Create New Model from Template")
    st.markdown(f"Template: `{file_path.relative_to(WORK_DIR)}`")

    new_name = st.text_input("Enter new model name (without _in.xlsx)")
    same_folder = st.checkbox("Create in same folder as template", value=True)
    if same_folder:
        target_dir = file_path.parent
    else:
        target_dir_label = st.selectbox(
            "Target folder",
            sorted({str(p.parent.relative_to(WORK_DIR)) for p in case_files}),
        )
        target_dir = WORK_DIR / target_dir_label

    if st.button("Create Model"):
        if not new_name:
            st.error("Please provide a name.")
        else:
            safe_name = new_name.strip()
            if safe_name.endswith("_in.xlsx"):
                new_file = target_dir / safe_name
            elif safe_name.endswith(".xlsx"):
                new_file = target_dir / safe_name
            else:
                new_file = target_dir / f"{safe_name}_in.xlsx"

            if new_file.exists():
                st.error(f"File already exists: `{new_file.relative_to(WORK_DIR)}`")
            else:
                shutil.copy(file_path, new_file)
                st.success(f"Created: `{new_file.relative_to(WORK_DIR)}`")

# ── VALIDATOR MODE ───────────────────────────────────────────────────────────
elif mode == "Validator":
    st.title("FLARE Input Validator")
    st.subheader(file_path.stem.replace("_in", ""))
    st.caption(str(file_path.relative_to(WORK_DIR)))

    issues = validate(rows)

    if not issues:
        st.success("File complies with formatting rules.")
    else:
        st.error(f"{len(issues)} issue(s) found")
        for issue in issues:
            st.write(issue)

# ── EDITOR MODE ──────────────────────────────────────────────────────────────
else:
    st.title("FLARE Input Model Editor")
    case_name = file_path.stem.replace("_in", "")
    st.subheader(case_name)
    st.caption(str(file_path.relative_to(WORK_DIR)))

    banner_text, banner_rows, sections = build_sections(rows)
    flarecon_sheet = has_flarecon_sheet(file_path)

    # ── Sidebar: RCS sections ─────────────────────────────────────────────────
    if "section" not in st.session_state:
        st.session_state.section = "Banner"

    st.sidebar.markdown("---")
    st.sidebar.header("RCS Model Sections")

    if st.sidebar.button("Banner Header", use_container_width=True, key="rcs_sec_Banner"):
        st.session_state.section = "Banner"
        st.session_state.active_tab = "rcs"

    for sec in sections:
        if st.sidebar.button(sec["name"], use_container_width=True,
                              key=f"rcs_sec_{sec['name']}"):
            st.session_state.section = sec["name"]
            st.session_state.active_tab = "rcs"

    # ── Sidebar: FLARECON sections (only when sheet present) ──────────────────
    if flarecon_sheet:
        con_rows = load_rows(file_path, sheet=flarecon_sheet)
        con_banner, con_banner_rows, con_sections = build_sections(con_rows, skip_banner=True)

        # Default to first section (not Banner) since FLARECON has no banner
        if "con_section" not in st.session_state:
            st.session_state.con_section = (
                con_sections[0]["name"] if con_sections else "Banner"
            )

        st.sidebar.markdown("---")
        st.sidebar.header("FLARECON Sections")

        if st.sidebar.button("CON Banner Header", use_container_width=True,
                              key="con_sec_Banner"):
            st.session_state.con_section = "Banner"
            st.session_state.active_tab = "con"

        for sec in con_sections:
            if st.sidebar.button(sec["name"], use_container_width=True,
                                  key=f"con_sec_{sec['name']}"):
                st.session_state.con_section = sec["name"]
                st.session_state.active_tab = "con"

    # ── Main panel: tabbed when FLARECON present, plain otherwise ─────────────
    if flarecon_sheet:
        tab_rcs, tab_con = st.tabs(["FLARE RCS Model", "FLARECON Model"])
    else:
        tab_rcs = st.container()
        tab_con = None

    # ── RCS editor panel ──────────────────────────────────────────────────────
    with tab_rcs:
        if st.session_state.section == "Banner":
            st.header("Banner Header")
            banner_text = st.text_area("Edit Header", value=banner_text, height=200)
        else:
            section = next((s for s in sections if s["name"] == st.session_state.section), None)
            if section is None:
                st.session_state.section = "Banner"
                st.rerun()

            st.header(section["name"])

            for var in section["vars"]:
                col1, col2 = st.columns([2, 2])
                with col1:
                    tip = var.get("tooltip", "")
                    if tip:
                        st.markdown(
                            f"<div style='padding-top:8px;font-family:monospace;font-size:0.9rem;"
                            f"cursor:help;text-decoration:underline dotted #7090d0' "
                            f"title='{tip}'>{var['key']}</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            f"<div style='padding-top:8px;font-family:monospace;font-size:0.9rem'>"
                            f"{var['key']}</div>",
                            unsafe_allow_html=True,
                        )
                with col2:
                    wk = f"{file_path}_{section['name']}_{var['key']}"
                    if wk not in st.session_state:
                        st.session_state[wk] = str(var["value"])
                    st.text_input(
                        var["key"],
                        key=wk,
                        label_visibility="collapsed",
                    )

        # Build updates from ALL sections using session_state widget values where
        # available (visited sections) and original parsed values as fallback.
        updates = []
        for sec in sections:
            for var in sec["vars"]:
                widget_key = f"{file_path}_{sec['name']}_{var['key']}"
                val = st.session_state.get(widget_key, str(var["value"]))
                updates.append((var["row"], var["key"], val, var.get("tooltip", "")))

        if st.button("Save RCS Model", key="save_rcs"):
            # Re-pair edited banner text lines with original row numbers
            edited_lines = banner_text.split("\n")
            paired = [(banner_rows[i][0], edited_lines[i]) if i < len(edited_lines) else banner_rows[i]
                      for i in range(len(banner_rows))]
            out = apply_updates(file_path, updates, paired)
            st.success(f"Saved to `{out.relative_to(WORK_DIR)}`")

    # ── FLARECON editor panel ─────────────────────────────────────────────────
    if tab_con is not None:
        with tab_con:
            if st.session_state.con_section == "Banner":
                st.header("FLARECON Banner Header")
                con_banner = st.text_area(
                    "Edit FLARECON Header", value=con_banner,
                    height=200, key="con_banner_area"
                )
            else:
                con_sec = next(
                    (s for s in con_sections if s["name"] == st.session_state.con_section),
                    None
                )
                if con_sec is None:
                    st.session_state.con_section = "Banner"
                    st.rerun()

                st.header(con_sec["name"])

                for var in con_sec["vars"]:
                    col1, col2 = st.columns([2, 2])
                    with col1:
                        tip = var.get("tooltip", "")
                        if tip:
                            st.markdown(
                                f"<div style='padding-top:8px;font-family:monospace;font-size:0.9rem;"
                                f"cursor:help;text-decoration:underline dotted #7090d0' "
                                f"title='{tip}'>{var['key']}</div>",
                                unsafe_allow_html=True,
                            )
                        else:
                            st.markdown(
                                f"<div style='padding-top:8px;font-family:monospace;font-size:0.9rem'>"
                                f"{var['key']}</div>",
                                unsafe_allow_html=True,
                            )
                    with col2:
                        wk = f"con_{file_path}_{con_sec['name']}_{var['key']}"
                        if wk not in st.session_state:
                            st.session_state[wk] = str(var["value"])
                        st.text_input(
                            var["key"],
                            key=wk,
                            label_visibility="collapsed",
                        )

            # Build con_updates from ALL FLARECON sections via session_state.
            con_updates = []
            for sec in con_sections:
                for var in sec["vars"]:
                    widget_key = f"con_{file_path}_{sec['name']}_{var['key']}"
                    val = st.session_state.get(widget_key, str(var["value"]))
                    con_updates.append((var["row"], var["key"], val, var.get("tooltip", "")))

            if st.button("Save FLARECON Model", key="save_con"):
                edited_con_lines = con_banner.split("\n")
                con_paired = [(con_banner_rows[i][0], edited_con_lines[i]) if i < len(edited_con_lines) else con_banner_rows[i]
                              for i in range(len(con_banner_rows))]
                out = apply_updates(
                    file_path, con_updates, con_paired, sheet=flarecon_sheet
                )
                st.success(f"Saved to `{out.relative_to(WORK_DIR)}`")

st.sidebar.markdown("---")
st.sidebar.caption(f"Working Dir:\n`{WORK_DIR}`")
