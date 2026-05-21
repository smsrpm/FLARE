"""
flare_sim_batch_worker.py — durable worker for FLARE PWR Simulator "Run All Cases".

Launched by flare_ui.py. Runs each Case*_in.xlsx simulation in an independent
subprocess, writes each case output to a normal sim_<case>_<timestamp> folder,
and updates a small regular control/status folder so the Streamlit UI can reconnect after browser disconnects.
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path


def _runtime_dir(work_dir: Path) -> Path:
    """Return the FLARE runtime folder, preferring runtime/ then Runtime/."""
    base = Path(work_dir)
    lower = base / "runtime"
    upper = base / "Runtime"
    if lower.exists():
        rt = lower
    elif upper.exists():
        rt = upper
    else:
        rt = lower
        rt.mkdir(parents=True, exist_ok=True)
    return rt

def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def write_json_resilient(path: Path, payload: dict, *, warnings_name: str = "sim_all_status_write_warnings.log") -> None:
    """Best-effort JSON write tolerant of transient OneDrive/Windows locks."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=2)
    warning_path = path.parent / warnings_name

    last_err = None
    for attempt in range(10):
        tmp = path.with_name(f"{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
        try:
            tmp.write_text(text, encoding="utf-8")
            try:
                os.replace(str(tmp), str(path))
                return
            except PermissionError as e:
                last_err = e
                try:
                    tmp.unlink(missing_ok=True)
                except Exception:
                    pass
                time.sleep(0.05 * (attempt + 1))
        except Exception as e:
            last_err = e
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
            time.sleep(0.05 * (attempt + 1))

    # Fallback direct write. If it fails, log and continue; status failure must
    # not terminate the batch run.
    try:
        path.write_text(text, encoding="utf-8")
        return
    except Exception as e:
        last_err = e

    try:
        with open(warning_path, "a", encoding="utf-8") as f:
            f.write(f"[{now_iso()}] Could not write {path.name}: {last_err}\n")
    except Exception:
        pass


def read_json(path: Path, default=None):
    try:
        p = Path(path)
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {} if default is None else default


def abort_requested(run_dir: Path) -> bool:
    return (run_dir / "sim_all_abort_requested.json").exists()


def terminate_process_tree(proc: subprocess.Popen) -> None:
    try:
        if proc.poll() is not None:
            return
        if sys.platform.startswith("win"):
            subprocess.run(["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                           capture_output=True, text=True, check=False)
        else:
            proc.terminate()
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


def copy_input_with_retries(src: Path, dst: Path, *, attempts: int = 10) -> None:
    """Copy an input workbook, tolerating short Excel/OneDrive locks."""
    last_err = None
    for i in range(attempts):
        try:
            shutil.copy2(str(src), str(dst))
            return
        except PermissionError as e:
            last_err = e
            time.sleep(0.15 * (i + 1))
        except Exception as e:
            last_err = e
            time.sleep(0.05 * (i + 1))
    raise last_err if last_err is not None else RuntimeError(f"Could not copy {src} to {dst}")


# ── Final Report support ─────────────────────────────────────────────────────
def _read_config(work_dir: Path, key: str):
    cfg = _runtime_dir(Path(work_dir)) / "flare_config.txt"
    try:
        if cfg.exists():
            for line in cfg.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                if k.strip().lower() == key.lower():
                    return v.strip().strip('"').strip("'")
    except Exception:
        pass
    return None


def _load_api_key(work_dir: Path):
    return _read_config(work_dir, "ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY")


def _load_model(work_dir: Path):
    return _read_config(work_dir, "ANTHROPIC_MODEL") or "claude-sonnet-4-5"



NARRATIVE_QUALITY_GUIDANCE = """
AI narrative quality requirements:
- Use a tiered scope. For verification, null, or benign cases with no meaningful excursion, use only three concise sections: Steady-State/Initial Condition Performance, Code Behavior, and Regulatory Significance. Do not force full accident-analysis boilerplate onto benign cases.
- Reserve detailed Fuel Integrity and Core Damage and Source Term and Radiological Consequences sections for cases with nonzero ECR, hydrogen generation, rod failures, radionuclide releases, or nonzero dose.
- If all radionuclide group releases are zero, state that in one sentence and do not enumerate radionuclide groups individually.
- If ECR and hydrogen generation are zero, state that in one sentence and do not expand into a full oxidation discussion.
- The 10 CFR 50.46 PCT limit is 2,200°F (1,204°C / 1,477 K). State clearly whether the result passes or fails this criterion. If peak hot-pin clad temperature exceeds 1,477 K, call it a limit exceedance; do not hedge with phrases such as 'slightly exceeded' while also implying compliance.
- Describe deterministic code behavior declaratively. Avoid hedging phrases such as 'likely', 'suggests', 'may indicate', or 'appears to be' when explaining model responses already contained in the output.
- DNBR equal to NaN indicates that the DNB correlation is not applicable under the final conditions, not a numerical failure. Acknowledge it in one sentence only if needed.
- If initial SG heat removal is less than 50% of initial reactor power, state that the case was initialized at partial-load, hot-zero-power, or otherwise non-full-power conditions; do not call it a normal full-power case.
- If any dose receptor exceeds its regulatory limit, state this explicitly in conclusions, not only in tabulated data.
- Do not reproduce source-term or dose tables in prose. The numerical tables are appended separately; discuss only the important implications.
- Use PDF-safe scientific notation. Write large numbers as 1.41e21 or 1.41 x 10^21, not with Unicode superscript digits such as 10²¹.
"""

def _fmt(x, nd=3):
    try:
        if x is None:
            return "—"
        x = float(x)
        if abs(x) >= 1000:
            return f"{x:,.0f}"
        return f"{x:.{nd}f}"
    except Exception:
        return str(x)


def _read_case_df(case_run_dir: Path, case: str):
    try:
        import pandas as pd
        p = case_run_dir / f"{case}_out.csv"
        if p.exists():
            return pd.read_csv(p)
    except Exception:
        pass
    return None


def _event_time(df, col, threshold=0.0, mode="gt"):
    try:
        if col not in df.columns:
            return None
        s = df[col]
        mask = s > threshold if mode == "gt" else s < threshold
        if bool(mask.any()):
            return float(df.loc[mask, "Time (s)"].iloc[0])
    except Exception:
        pass
    return None


def _build_ic_final_rows(df):
    rows = []
    if df is None or df.empty:
        return rows
    specs = [
        ("RCS Pressure", "RCS Pressure (kPa)", "kPa", 0),
        ("RCS Temperature", "RCS Temperature (K)", "K", 1),
        ("RK Total Power", "RK Total Power (MW)", "MW", 3),
        ("Core Power", "Core Power (MW)", "MW", 3),
        ("Vessel Level", "Vessel Level (m)", "m", 3),
        ("Accumulator Level", "Accumulator Level (m)", "m", 3),
        ("Pump Mass Flow", "Pump Mass Flow (kg/s)", "kg/s", 2),
        ("SG Heat Removal", "SG Heat Removal (MW)", "MW", 3),
        ("Hot Pin Clad Temp", "Hot Pin Clad Temp (K)", "K", 1),
        ("DNBR", "DNBR", "—", 3),
        ("H2 Generated", "H2 Generated (kg)", "kg", 3),
    ]
    for label, col, unit, nd in specs:
        if col in df.columns:
            try:
                rows.append([label, _fmt(df[col].iloc[0], nd), _fmt(df[col].iloc[-1], nd), unit])
            except Exception:
                pass
    return rows


def _build_soe_rows(df):
    rows = []
    if df is None or df.empty or "Time (s)" not in df.columns:
        return rows
    t = df["Time (s)"]
    rows.append((0.0, "Simulation start"))
    events = [
        ("Break Flow (kg/s)", 0.1, "Break flow begins"),
        ("PORV Mass Flow (kg/s)", 0.01, "PORV opens / relief flow begins"),
        ("Accumulator Flow (kg/s)", 0.1, "Accumulator injection begins"),
        ("HPSI Flow (kg/s)", 0.1, "HPSI injection begins"),
        ("LPSI Flow (kg/s)", 0.1, "LPSI injection begins"),
        ("CVCS Makeup (kg/s)", 0.01, "CVCS makeup begins"),
    ]
    for col, threshold, label in events:
        et = _event_time(df, col, threshold, "gt")
        if et is not None:
            rows.append((et, label))
    if "Reactivity scram (pcm)" in df.columns:
        try:
            mask = df["Reactivity scram (pcm)"].abs() > 1.0
            if bool(mask.any()):
                rows.append((float(t[mask].iloc[0]), "Reactor scram reactivity inserted"))
        except Exception:
            pass
    extrema = [
        ("RCS Pressure (kPa)", "min", "Minimum RCS pressure"),
        ("RCS Pressure (kPa)", "max", "Maximum RCS pressure"),
        ("RCS Temperature (K)", "max", "Peak RCS temperature"),
        ("RK Total Power (MW)", "max", "Peak reactor power"),
        ("Hot Pin Clad Temp (K)", "max", "Peak hot-pin cladding temperature"),
        ("DNBR", "min", "Minimum DNBR"),
        ("Zr Oxidation Mean Oxidizing Rod ECR (%)", "max", "Peak mean oxidizing-rod ECR"),
    ]
    for col, kind, label in extrema:
        try:
            if col in df.columns:
                s = df[col].replace([float("inf"), -float("inf")], float("nan")).dropna()
                if not s.empty:
                    idx = s.idxmin() if kind == "min" else s.idxmax()
                    rows.append((float(df.loc[idx, "Time (s)"]), f"{label}: {_fmt(df.loc[idx, col], 3)}"))
        except Exception:
            pass
    rows.append((float(t.iloc[-1]), "Simulation end"))
    # de-duplicate by rounded time + label, then sort
    seen = set()
    out = []
    for tm, label in sorted(rows, key=lambda x: (x[0], x[1])):
        key = (round(float(tm), 3), label)
        if key not in seen:
            seen.add(key)
            out.append([float(tm), label])
    return out


def _build_case_summary_text(df, case: str):
    if df is None or df.empty:
        return f"Case {case}: no output CSV was available."
    lines = [f"Case: {case}", f"Duration: {_fmt(df['Time (s)'].iloc[0],1)} to {_fmt(df['Time (s)'].iloc[-1],1)} s"]
    def add_ext(col, label, units="", kind="max"):
        if col in df.columns:
            try:
                s = df[col].replace([float("inf"), -float("inf")], float("nan")).dropna()
                if s.empty: return
                idx = s.idxmax() if kind == "max" else s.idxmin()
                lines.append(f"{label}: initial {_fmt(df[col].iloc[0])} {units}, {kind} {_fmt(df.loc[idx, col])} {units} at t={_fmt(df.loc[idx, 'Time (s)'],1)} s, final {_fmt(df[col].iloc[-1])} {units}")
            except Exception:
                pass
    add_ext("RCS Pressure (kPa)", "RCS pressure", "kPa", "min")
    add_ext("RCS Temperature (K)", "RCS temperature", "K", "max")
    add_ext("RK Total Power (MW)", "Reactor power", "MW", "max")
    add_ext("Hot Pin Clad Temp (K)", "Hot-pin clad temperature", "K", "max")
    add_ext("DNBR", "DNBR", "", "min")
    add_ext("SG Heat Removal (MW)", "SG heat removal", "MW", "max")
    add_ext("Zr Oxidation Mean Oxidizing Rod ECR (%)", "Mean oxidizing-rod ECR", "%", "max")
    add_ext("H2 Generated (kg)", "Hydrogen generated", "kg", "max")
    try:
        _pwr0 = float(df["RK Total Power (MW)"].iloc[0]) if "RK Total Power (MW)" in df.columns else 0.0
        _sg0 = float(df["SG Heat Removal (MW)"].iloc[0]) if "SG Heat Removal (MW)" in df.columns else 0.0
        if _pwr0 > 0 and _sg0 < 0.5 * _pwr0:
            lines.append("Initial-condition check: initial SG heat removal is less than 50% of initial reactor power; do not describe this as normal full-power steady-state operation unless other data support that conclusion.")
    except Exception:
        pass
    try:
        _pct_limit_K = 1477.0
        _pct = float(pd.to_numeric(df["Hot Pin Clad Temp (K)"], errors="coerce").max()) if "Hot Pin Clad Temp (K)" in df.columns else None
        if _pct is not None:
            if _pct >= _pct_limit_K:
                lines.append(f"10 CFR 50.46 PCT check: peak hot-pin clad temperature {_pct:.1f} K exceeds the 1477 K limit; state this as a limit exceedance.")
            else:
                lines.append(f"10 CFR 50.46 PCT check: peak hot-pin clad temperature {_pct:.1f} K is below the 1477 K limit.")
    except Exception:
        pass
    try:
        _ecr = float(pd.to_numeric(df.get("Zr Oxidation Mean Oxidizing Rod ECR (%)", pd.Series([0.0])), errors="coerce").fillna(0.0).max())
        _h2 = float(pd.to_numeric(df.get("H2 Generated (kg)", pd.Series([0.0])), errors="coerce").fillna(0.0).max())
        _rods = 0.0
        for _c in ["Rod Failures DNB (est.)", "Rod Failures Gap (est.)", "Rod Failures EarlyIV (est.)"]:
            if _c in df.columns:
                _rods = max(_rods, float(pd.to_numeric(df[_c], errors="coerce").fillna(0.0).max()))
        if _ecr <= 0.0 and _h2 <= 0.0 and _rods <= 0.0:
            lines.append("Fuel-integrity scope: ECR, hydrogen generation, and rod failures are zero; collapse fuel-integrity discussion to one concise sentence.")
        if case.lower().endswith("null") or (_ecr <= 0.0 and _h2 <= 0.0 and _rods <= 0.0 and _pct is not None and _pct < 900.0):
            lines.append("Narrative scope: treat this as a verification/null or benign case; use a short three-section structure rather than a full accident narrative.")
    except Exception:
        pass
    try:
        if "DNBR" in df.columns:
            _dn_final = pd.to_numeric(df["DNBR"], errors="coerce").iloc[-1]
            if pd.isna(_dn_final):
                lines.append("DNBR interpretation: final DNBR is NaN because the DNB correlation is not applicable under the final conditions; this is expected behavior, not a numerical convergence issue.")
    except Exception:
        pass
    try:
        case_dir = globals().get("_CURRENT_CASE_RUN_DIR_FOR_SUMMARY")
        if case_dir:
            st_lines = _source_term_lines_for_report(Path(case_dir), case)
            if st_lines:
                lines.append("")
                lines.append("SOURCE TERM AND RADIOLOGICAL DATA FOR NARRATIVE:")
                lines.extend(st_lines)
    except Exception:
        pass
    return "\n".join(lines)




def _read_workbook_sheet(case_run_dir: Path, case: str, sheet_name: str):
    try:
        import pandas as pd
        p = Path(case_run_dir) / f"{case}_out.xlsx"
        if p.exists():
            return pd.read_excel(p, sheet_name=sheet_name, engine="openpyxl")
    except Exception:
        pass
    return None


def _read_workbook_rows(case_run_dir: Path, case: str, sheet_name: str):
    try:
        from openpyxl import load_workbook
        p = Path(case_run_dir) / f"{case}_out.xlsx"
        if not p.exists():
            return None
        wb = load_workbook(p, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close(); return None
        ws = wb[sheet_name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        wb.close()
        return rows
    except Exception:
        return None


def _format_source_term_option_worker(value):
    if value is None:
        return None
    raw = str(value).strip().strip('"').strip("'")
    if not raw or raw.lower() in {"nan", "none"}:
        return None
    norm = raw.lower().replace("_", "-")
    aliases = {
        "licensing-auto": "licensing-auto", "licensing auto": "licensing-auto",
        "thermal-failure": "thermal-failure", "rg1183-loca": "RG1183-LOCA",
        "rg-1183-loca": "RG1183-LOCA", "rg1183-nonloca": "RG1183-nonLOCA",
        "rg1183-non-loca": "RG1183-nonLOCA", "rg-1183-nonloca": "RG1183-nonLOCA",
        "rg1183-nonloca-dnb": "RG1183-nonLOCA-DNB", "rg1183-non-loca-dnb": "RG1183-nonLOCA-DNB",
    }
    return aliases.get(norm, raw.replace("_", "-"))


def _source_term_lines_for_report(case_run_dir: Path, case: str, max_groups: int = 12):
    """Compact source-term / dose facts for final-report AI narratives."""
    lines = []
    try:
        import pandas as pd
        st_df = _read_workbook_sheet(case_run_dir, case, "Source Term")
        if st_df is not None and not st_df.empty:
            option = None; applied = None
            if "Group" in st_df.columns:
                value_cols = [c for c in ["Total release frac", "Value", "Model", "Released inventory (Ci)"] if c in st_df.columns]
                for _, row in st_df.iterrows():
                    label = str(row.get("Group", "")).strip().lower()
                    val = None
                    for c in value_cols:
                        vv = row.get(c)
                        if vv is not None and not pd.isna(vv) and str(vv).strip() != "":
                            val = vv; break
                    if label in {"fuel source-term option", "source term option", "source_term_model"}:
                        option = val
                    elif label in {"model selected", "applied source-term model", "applied model"}:
                        applied = val
            option = _format_source_term_option_worker(option)
            applied = _format_source_term_option_worker(applied)
            if option:
                if option == "licensing-auto":
                    lines.append(f"Source-term option: licensing-auto; applied model: {applied or 'not reported'}")
                else:
                    lines.append(f"Source-term option: {option}")

            inv_col = "Core inventory (Ci)"; rel_col = "Released inventory (Ci)"; frac_col = "Total release frac"
            if inv_col in st_df.columns and "Group" in st_df.columns:
                df = st_df.copy()
                for c in [inv_col, rel_col, frac_col, "Gap release %", "Early IV %", "Total release %", "BDBE F factor", "BDBE ECR adjustment %"]:
                    if c in df.columns:
                        df[c] = pd.to_numeric(df[c], errors="coerce")
                main = df[df[inv_col].notna()].copy()
                if not main.empty:
                    total_rows = main[main["Group"].astype(str).str.upper().eq("TOTAL")]
                    non_total = main[~main["Group"].astype(str).str.upper().eq("TOTAL")].copy()
                    _release_cols = [c for c in [rel_col, frac_col, "Total release %"] if c in non_total.columns]
                    _all_zero_release = False
                    if _release_cols:
                        _vals = []
                        for _c in _release_cols:
                            _vals.extend(pd.to_numeric(non_total[_c], errors="coerce").fillna(0.0).abs().tolist())
                        _all_zero_release = bool(_vals) and max(_vals) <= 0.0
                    if _all_zero_release:
                        lines.append("All radionuclide group releases are zero; source-term prose should state this once and should not enumerate individual groups.")
                    else:
                        if rel_col in non_total.columns:
                            non_total = non_total.sort_values(rel_col, ascending=False)
                        rows = list(non_total.head(max_groups).to_dict("records"))
                        if not total_rows.empty:
                            rows.append(total_rows.iloc[0].to_dict())
                        lines.append("Source-term group releases shown in the Results panel:")
                        for r in rows:
                            g = str(r.get("Group", "")).strip()
                            if not g or g.lower() == "nan": continue
                            parts = [g]
                            if "Total release %" in r and pd.notna(r.get("Total release %")):
                                parts.append(f"total release {float(r['Total release %']):.4g}%")
                            elif frac_col in r and pd.notna(r.get(frac_col)):
                                parts.append(f"total release fraction {float(r[frac_col]):.4e}")
                            if rel_col in r and pd.notna(r.get(rel_col)):
                                parts.append(f"released inventory {float(r[rel_col]):.4e} Ci")
                            if "Gap release %" in r and pd.notna(r.get("Gap release %")):
                                parts.append(f"gap {float(r['Gap release %']):.4g}%")
                            if "Early IV %" in r and pd.notna(r.get("Early IV %")):
                                parts.append(f"early-IV {float(r['Early IV %']):.4g}%")
                            if "BDBE F factor" in r and pd.notna(r.get("BDBE F factor")):
                                parts.append(f"BDBE F {float(r['BDBE F factor']):.4g}")
                            if "BDBE ECR adjustment %" in r and pd.notna(r.get("BDBE ECR adjustment %")):
                                parts.append(f"ECR adjustment {float(r['BDBE ECR adjustment %']):.4g}%")
                            lines.append(" - " + "; ".join(parts))

            if "Group" in st_df.columns:
                meta_labels = {"Model selected", "Fuel source-term option", "Inventory model", "Estimated fissions", "Severe-event ECR release flag", "Severe-event ECR threshold (%)", "Predicted mean oxidizing-rod ECR (%)", "Severe-event ECR active", "NOTBADTRAD gap release duration (hr)", "NOTBADTRAD early-IV release duration (hr)", "NOTBADTRAD EAB integration time (hr)", "NOTBADTRAD LPZ integration time (hr)"}
                value_cols = [c for c in ["Total release frac", "Value", "Model", "Released inventory (Ci)"] if c in st_df.columns]
                for _, row in st_df[st_df["Group"].isin(meta_labels)].iterrows():
                    label = str(row.get("Group", "")).strip()
                    val = None
                    for c in value_cols:
                        vv = row.get(c)
                        if vv is not None and not pd.isna(vv) and str(vv).strip() != "":
                            val = vv; break
                    if val is not None:
                        lines.append(f"Source-term basis: {label} = {val}")
        for sheet, prefix in [("Dose", "NOTBADTRAD dose screening"), ("Iodine Spike", "Iodine spike pre-existing coolant activity dose")]:
            rows = _read_workbook_rows(case_run_dir, case, sheet)
            if rows:
                bits=[]
                for r in rows:
                    if r and r[0] in ("EAB", "LPZ", "Control Room") and len(r) >= 5:
                        bits.append(f"{r[0]} TEDE {r[1]} rem vs limit {r[2]}: {r[4]}")
                if bits:
                    lines.append(prefix + ": " + "; ".join(bits))
    except Exception as e:
        lines.append(f"Source-term narrative data unavailable: {e}")
    return lines



def _is_number_like(v):
    try:
        if v is None:
            return False
        import math
        f = float(v)
        return not math.isnan(f)
    except Exception:
        return False


def _fmt_report_value(v, fmt=None):
    try:
        import pandas as pd
        if v is None or pd.isna(v):
            return ""
    except Exception:
        if v is None:
            return ""
    if fmt:
        try:
            return fmt.format(float(v))
        except Exception:
            return str(v)
    return _normalize_pdf_text(v)


def _stringify_df_for_report(df, fmt_map=None, text_cols=None):
    fmt_map = fmt_map or {}
    text_cols = set(text_cols or [])
    out = df.copy()
    for c in out.columns:
        fmt = fmt_map.get(c)
        out[c] = out[c].map(lambda v, _fmt=None if c in text_cols else fmt: _fmt_report_value(v, _fmt))
    return out


def _source_term_dose_tables_for_report(case_run_dir: Path, case: str):
    """Build the same source-term/dose tables shown in the PWR Simulator Results panel."""
    tables = []
    try:
        import pandas as pd
        import numpy as np
        st_df = _read_workbook_sheet(case_run_dir, case, "Source Term")
        if st_df is not None and not st_df.empty:
            option = None; applied = None
            if "Group" in st_df.columns:
                value_cols = [c for c in ["Total release frac", "Value", "Model", "Released inventory (Ci)"] if c in st_df.columns]
                for _, row in st_df.iterrows():
                    label = str(row.get("Group", "")).strip().lower()
                    val = None
                    for c in value_cols:
                        vv = row.get(c)
                        if vv is not None and not pd.isna(vv) and str(vv).strip() != "":
                            val = vv; break
                    if label in {"fuel source-term option", "source term option", "source_term_model"}:
                        option = val
                    elif label in {"model selected", "applied source-term model", "applied model"}:
                        applied = val
            option = _format_source_term_option_worker(option)
            applied = _format_source_term_option_worker(applied)
            basis_rows = []
            if option:
                if option == "licensing-auto":
                    basis_rows.append(["Source-term option", option])
                    basis_rows.append(["Applied model", applied or "not reported in this output file"])
                else:
                    basis_rows.append(["Source-term option", option])
            if basis_rows:
                tables.append({"section":"Source Term  -  RG 1.183 / NBT", "title":"Source-term option", "df": pd.DataFrame(basis_rows, columns=["Field", "Value"])})

            inv_col = "Core inventory (Ci)"; rel_col = "Released inventory (Ci)"; frac_col = "Total release frac"
            if inv_col in st_df.columns and "Group" in st_df.columns:
                main = st_df[st_df[inv_col].map(_is_number_like)].copy()
                display_cols = [c for c in ["Group", "NBT key", inv_col, "Gap release %", "Early IV %", "Total release %", frac_col, rel_col] if c in main.columns]
                if display_cols and not main.empty:
                    fmt = {inv_col:"{:.4e}", rel_col:"{:.4e}", frac_col:"{:.4e}", "Gap release %":"{:.3f}", "Early IV %":"{:.3f}", "Total release %":"{:.3f}"}
                    tables.append({"section":"Source Term  -  RG 1.183 / NBT", "title":"Group inventories and releases", "df": _stringify_df_for_report(main[display_cols].reset_index(drop=True), fmt, text_cols=["Group", "NBT key"])})
            elif "Group" in st_df.columns:
                display_cols = [c for c in ["Group", "Gap release %", "Early IV %", "Total release %"] if c in st_df.columns]
                if display_cols:
                    tables.append({"section":"Source Term  -  RG 1.183 / NBT", "title":"Source-term releases", "df": _stringify_df_for_report(st_df[display_cols].reset_index(drop=True))})

            meta_labels = {"Model selected", "Fuel source-term option", "Inventory model", "Estimated fissions", "NOTBADTRAD gap release duration (hr)", "NOTBADTRAD early-IV release duration (hr)", "NOTBADTRAD EAB integration time (hr)", "NOTBADTRAD LPZ integration time (hr)", "Severe-event ECR release flag", "Severe-event ECR threshold (%)", "Predicted mean oxidizing-rod ECR (%)", "Severe-event ECR active"}
            if "Group" in st_df.columns:
                meta_df = st_df[st_df["Group"].isin(meta_labels)].copy()
                meta_cols = [c for c in ["Group", frac_col] if c in meta_df.columns]
                if meta_cols and not meta_df.empty:
                    tables.append({"section":"Source Term  -  RG 1.183 / NBT", "title":"Source-term basis", "df": _stringify_df_for_report(meta_df[meta_cols].reset_index(drop=True))})

        def add_dose_tables(sheet_name, section_title):
            rows = _read_workbook_rows(case_run_dir, case, sheet_name)
            if not rows:
                return
            summary = [r[:5] for r in rows if r and r[0] in ("EAB", "LPZ", "Control Room")]
            if summary:
                df = pd.DataFrame(summary, columns=["Location", "TEDE (rem)", "Limit (rem)", "Margin (rem)", "Result"])
                tables.append({"section": section_title, "title":"Summary", "df": _stringify_df_for_report(df, {"TEDE (rem)":"{:.4e}", "Margin (rem)":"{:.4e}"}, text_cols=["Location", "Limit (rem)", "Result"])})
            grp_hdr = next((i for i,r in enumerate(rows) if r and r[0] == "Group"), None)
            if grp_hdr is not None:
                grp_rows=[]
                for r in rows[grp_hdr+1:]:
                    if not any(r): break
                    grp_rows.append(r)
                if grp_rows:
                    df = pd.DataFrame([r[:6] for r in grp_rows], columns=["Group", "Released (Ci)", "EAB (rem)", "LPZ (rem)", "CR (rem)", "Release frac"])
                    tables.append({"section": section_title, "title":"Group contributions", "df": _stringify_df_for_report(df, {"Released (Ci)":"{:.4e}", "EAB (rem)":"{:.4e}", "LPZ (rem)":"{:.4e}", "CR (rem)":"{:.4e}", "Release frac":"{:.4e}"}, text_cols=["Group"])})
            if sheet_name == "Dose":
                fit_row = next((r for r in rows if r and r[0] == "Power-law fit"), None)
                if fit_row and len(fit_row) > 1 and fit_row[1]:
                    fit_err = next((r[1] for r in rows if r and r[0] == "Fit error (%)"), None)
                    data = [["Power-law fit", fit_row[1]]]
                    if fit_err is not None:
                        data.append(["Fit error (%)", fit_err])
                    tables.append({"section": section_title, "title":"TEDE vs Distance fit", "df": pd.DataFrame(data, columns=["Field", "Value"])})
                dist_hdr = next((i for i,r in enumerate(rows) if r and r[0] == "Distance (m)"), None)
                if dist_hdr is not None:
                    dt_rows=[]
                    for r in rows[dist_hdr+1:]:
                        if not any(r): break
                        dt_rows.append(r)
                    if dt_rows:
                        hdr=rows[dist_hdr]
                        if len(hdr) >= 4 and "chi" in str(hdr[1]).lower() and "EAB" in str(hdr[2]) and "LPZ" in str(hdr[3]):
                            df=pd.DataFrame([r[:4] for r in dt_rows], columns=["Distance (m)", "χ/Q (s/m³)", "TEDE EAB interval (rem)", "TEDE LPZ interval (rem)"])
                            tables.append({"section": section_title, "title":"TEDE vs Distance", "df": _stringify_df_for_report(df, {"χ/Q (s/m³)":"{:.4e}", "TEDE EAB interval (rem)":"{:.4e}", "TEDE LPZ interval (rem)":"{:.4e}"})})
                        elif len(hdr) >= 3 and "EAB" in str(hdr[1]) and "LPZ" in str(hdr[2]):
                            df=pd.DataFrame([r[:3] for r in dt_rows], columns=["Distance (m)", "TEDE EAB interval (rem)", "TEDE LPZ interval (rem)"])
                            tables.append({"section": section_title, "title":"TEDE vs Distance", "df": _stringify_df_for_report(df, {"TEDE EAB interval (rem)":"{:.4e}", "TEDE LPZ interval (rem)":"{:.4e}"})})

        add_dose_tables("Dose", "NOTBADTRAD Dose Screening")

        isp_rows = _read_workbook_rows(case_run_dir, case, "Iodine Spike")
        if isp_rows:
            metric_rows=[]
            keys = ["Model", "Coolant activity (uCi/g)", "Primary mass (kg)", "Spike inventory (Ci)", "Spike multiplier", "Scram occurred", "PORV opened", "DNB occurred", "N rods DNB/dryout", "Equilibrium spike frac", "Accident spike frac (DNB)"]
            for k in keys:
                row = next((r for r in isp_rows if r and r[0] == k), None)
                if row and len(row) > 1 and row[1] is not None:
                    metric_rows.append([k, row[1]])
            if metric_rows:
                tables.append({"section":"Iodine Spike — Pre-existing Coolant Activity", "title":"Inputs and trigger flags", "df": pd.DataFrame(metric_rows, columns=["Field", "Value"])})
        add_dose_tables("Iodine Spike", "Iodine Spike — Pre-existing Coolant Activity")
    except Exception as e:
        try:
            import pandas as pd
            tables.append({"section":"Source Term and Dose Tables", "title":"Parsing warning", "df": pd.DataFrame([[str(e)]], columns=["Warning"])})
        except Exception:
            pass
    return tables
def _call_anthropic(prompt: str, work_dir: Path, max_tokens: int = 1100):
    key = _load_api_key(work_dir)
    if not key:
        return None, "No Anthropic API key configured."
    try:
        import requests
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": _load_model(work_dir),
                "max_tokens": int(max_tokens),
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=90,
        )
        resp.raise_for_status()
        return resp.json()["content"][0]["text"], None
    except Exception as e:
        return None, str(e)


def _make_fallback_narrative(case: str, summary: str):
    lines = summary.splitlines()
    source_started = False
    core_lines, source_lines, other_lines = [], [], []
    for line in lines:
        if line.strip().startswith("SOURCE TERM AND RADIOLOGICAL DATA"):
            source_started = True
            continue
        if source_started:
            source_lines.append(line)
        elif any(k in line for k in ["ECR", "Hydrogen", "Hot-pin clad", "DNBR", "Rod failures"]):
            core_lines.append(line)
        else:
            other_lines.append(line)
    text = f"### Event Narrative — {case}\n\n"
    text += "An automated AI narrative was not generated for this case. The following deterministic summary was prepared from the FLARE output data.\n\n"
    text += "#### Event Overview and Thermal-Hydraulic Response\n" + "\n".join(f"- {line}" for line in other_lines if line.strip()) + "\n\n"
    if core_lines:
        text += "#### Fuel Integrity and Core Damage\n" + "\n".join(f"- {line}" for line in core_lines if line.strip()) + "\n\n"
    if source_lines:
        text += "#### Source Term and Radiological Consequences\n" + "\n".join(f"- {line}" for line in source_lines if line.strip())
    return text

def _generate_case_narrative(case: str, case_run_dir: Path, work_dir: Path, detail_level: float = 0.5):
    df = _read_case_df(case_run_dir, case)
    globals()["_CURRENT_CASE_RUN_DIR_FOR_SUMMARY"] = str(case_run_dir)
    summary = _build_case_summary_text(df, case)
    soe = _build_soe_rows(df)
    ic = _build_ic_final_rows(df)
    try:
        detail_level = max(0.0, min(1.0, float(detail_level)))
    except Exception:
        detail_level = 0.5
    if detail_level < 0.25:
        detail_instruction = "Write a brief engineering summary using separate headings for event overview, fuel integrity/core damage, source term/radiological consequences when applicable, and final safety significance."
    elif detail_level < 0.75:
        detail_instruction = "Write a standard engineering event narrative using explicit headings. Discuss initiating event, thermal-hydraulic response, safety-system response, fuel integrity/core damage, source term/radiological consequences when applicable, and final state."
    else:
        detail_instruction = "Write a detailed technical narrative using explicit headings. Include important timing, limiting parameters, safety-system performance, fuel integrity/core damage, source-term/radiological consequences when applicable, and final-state/regulatory significance."
    max_tokens = int(700 + detail_level * 1900)
    prompt = (
        "You are preparing an engineering event narrative for a FLARE PWR simulator final report. "
        "Use only the supplied data. Do not invent facts. "
        + detail_instruction + "\n\n"
        + NARRATIVE_QUALITY_GUIDANCE + "\n"
        "Mandatory organization: keep fuel integrity/core-damage topics separate from source-term and radiological topics. "
        "Discuss ECR, hydrogen generation, DNB, cladding temperature, and rod failures only under a heading such as "
        "'Fuel Integrity and Core Damage'. Discuss radionuclide release fractions, source-term model, NOTBADTRAD dose, "
        "iodine-spike dose, and radiological consequences only under 'Source Term and Radiological Consequences'. "
        "Do not merge those topics. If source-term data are supplied, include the important numerical group-release and dose results "
        "from that data in the source-term section.\n\n"
        f"{summary}\n\n"
        "Sequence of events and initial/final tables will be appended separately; do not reproduce them in markdown table form."
    )
    narrative, err = _call_anthropic(prompt, work_dir, max_tokens=max_tokens)
    if not narrative:
        narrative = _make_fallback_narrative(case, summary)
        if err:
            narrative += f"\n\n_AI narrative note: {err}_"
    narrative = _normalize_pdf_text(narrative)
    # save artifacts
    md_path = case_run_dir / f"{case}_narrative.md"
    soe_path = case_run_dir / f"{case}_sequence_of_events.csv"
    ic_path = case_run_dir / f"{case}_initial_final.csv"
    try:
        md_path.write_text(narrative, encoding="utf-8")
    except Exception:
        pass
    try:
        import pandas as pd
        pd.DataFrame(soe, columns=["Time (s)", "Event"]).to_csv(soe_path, index=False)
        pd.DataFrame(ic, columns=["Parameter", "Initial", "Final", "Units"]).to_csv(ic_path, index=False)
    except Exception:
        pass
    return {"case": case, "summary": summary, "narrative": narrative, "soe": soe, "ic": ic, "run_dir": str(case_run_dir)}


def _find_case_pngs(case_run_dir: Path):
    pngs = sorted(case_run_dir.glob("figure_*.png"))
    if not pngs:
        pngs = sorted(case_run_dir.glob("*.png"))
    def key(p):
        import re
        m = re.search(r"(\d+)", p.stem)
        return int(m.group(1)) if m else 9999
    return sorted(pngs, key=key)



_SUPERSCRIPT_TRANSLATION = str.maketrans({
    "⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4",
    "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9",
    "⁺": "+", "⁻": "-", "⁽": "(", "⁾": ")",
})
_SUBSCRIPT_TRANSLATION = str.maketrans({
    "₀": "0", "₁": "1", "₂": "2", "₃": "3", "₄": "4",
    "₅": "5", "₆": "6", "₇": "7", "₈": "8", "₉": "9",
    "₊": "+", "₋": "-", "₍": "(", "₎": ")",
})


def _normalize_pdf_text(text) -> str:
    """Normalize AI/report text to glyphs supported by ReportLab base fonts.

    The AI sometimes emits scientific notation such as 1.41 × 10²¹.  ReportLab's
    built-in PDF fonts can render the first superscript digit inconsistently,
    producing a black square in the final report.  Convert such notation to
    ASCII before creating Paragraphs/Tables.
    """
    if text is None:
        return ""
    s = str(text)
    # Convert superscript runs immediately following 10, e.g. 10²¹ -> 10^21.
    sup_chars = "⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁽⁾"
    s = re.sub(
        r"10([" + sup_chars + r"]+)",
        lambda m: "10^" + m.group(1).translate(_SUPERSCRIPT_TRANSLATION),
        s,
    )
    # Convert any remaining superscript/subscript glyphs to inline ASCII.
    s = s.translate(_SUPERSCRIPT_TRANSLATION)
    s = s.translate(_SUBSCRIPT_TRANSLATION)
    # Prefer ASCII-safe symbols in generated PDF prose.
    s = s.replace("×", "x")
    s = s.replace("−", "-")
    s = s.replace("–", "-")
    s = s.replace("—", "-")
    s = s.replace("≈", "~")
    s = s.replace("≤", "<=")
    s = s.replace("≥", ">=")
    return s


def _paragraphs_from_markdown(text: str):
    paras = []
    for block in str(text).split("\n\n"):
        b = block.strip()
        if not b:
            continue
        if b.startswith("###"):
            b = b.lstrip("#").strip()
        paras.append(_normalize_pdf_text(b.replace("**", "")))
    return paras


def _build_final_report(work_dir: Path, control_dir: Path, case_reports: list, results: list, detail_level: float = 0.5):
    pdf_path = Path(control_dir) / "FLARE_Run_All_Final_Report.pdf"
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, Image
    except Exception as e:
        raise RuntimeError(f"ReportLab is required for final PDF report: {e}")

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=TA_CENTER, fontSize=20, leading=24))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8.5, leading=10.5))
    styles.add(ParagraphStyle(name="Body11", parent=styles["BodyText"], fontSize=10.5, leading=13))
    styles.add(ParagraphStyle(name="H1Blue", parent=styles["Heading1"], textColor=colors.HexColor("#1f4e79")))
    styles.add(ParagraphStyle(name="H2Blue", parent=styles["Heading2"], textColor=colors.HexColor("#1f4e79")))

    doc = SimpleDocTemplate(str(pdf_path), pagesize=letter, rightMargin=0.65*inch, leftMargin=0.65*inch, topMargin=0.65*inch, bottomMargin=0.65*inch)
    story = []
    story.append(Paragraph(_normalize_pdf_text("FLARE Run All Cases — Final Report"), styles["TitleCenter"]))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Small"]))
    story.append(Spacer(1, 0.2*inch))

    ok = sum(1 for r in results if r.get("status") == "OK")
    failed = sum(1 for r in results if r.get("status") != "OK")
    intro = (
        f"This report compiles the output from a FLARE PWR Simulator Run All Cases batch. "
        f"The batch processed {len(results)} cases: {ok} completed successfully and {failed} were reported as failed, aborted, or incomplete. "
        "For each completed case, the report includes an event narrative, a sequence-of-events table, an initial/final state table, and the generated FLARE plot images."
    )
    story.append(Paragraph("Introduction", styles["H1Blue"]))
    story.append(Paragraph(_normalize_pdf_text(intro), styles["Body11"]))
    story.append(Spacer(1, 0.15*inch))

    # AI executive summary, with deterministic fallback.
    compact = "\n".join([cr.get("summary", "")[:1500] for cr in case_reports])
    try:
        detail_level = max(0.0, min(1.0, float(detail_level)))
    except Exception:
        detail_level = 0.5
    exec_paras = "2 to 3 concise paragraphs" if detail_level < 0.25 else ("3 to 5 concise paragraphs" if detail_level < 0.75 else "5 to 7 detailed paragraphs")
    exec_prompt = (
        "Prepare an executive summary for a FLARE final report covering a batch of PWR transient simulations. "
        f"Write {exec_paras}. Highlight the range of event types, main safety-system responses, limiting cases, and any failed/incomplete cases. "
        "Use cross-case context: identify the limiting case, compare it explicitly to its nearest parametric variant when one exists (for example hot-leg vs cold-leg LOCA variants), and identify any receptor dose exceedances in the executive summary and conclusions. "
        "Apply the same anti-boilerplate rules: benign/null cases should be summarized compactly, and zero source term or zero ECR/hydrogen should not be enumerated. "
        "Use only the supplied summaries.\n\n" + compact
    )
    exec_text, err = _call_anthropic(exec_prompt, work_dir, max_tokens=int(800 + detail_level * 1600))
    if not exec_text:
        exec_text = (
            f"The batch completed {ok} of {len(results)} cases successfully. "
            "The case sections that follow provide the detailed event narratives, sequence-of-events tables, initial/final states, and plots. "
            "No AI-generated executive summary was available."
        )
        if err:
            exec_text += f" AI summary note: {err}"
    exec_text = _normalize_pdf_text(exec_text)
    story.append(Paragraph("Executive Summary", styles["H1Blue"]))
    for para in _paragraphs_from_markdown(exec_text):
        story.append(Paragraph(para, styles["Body11"]))
        story.append(Spacer(1, 0.07*inch))

    # Summary table
    rows = [["Case", "Status", "Elapsed (s)", "Output Folder"]]
    for r in results:
        rows.append([_normalize_pdf_text(r.get("case", "")), _normalize_pdf_text(r.get("status", "")), _normalize_pdf_text(_fmt(r.get("elapsed_s"), 1)), _normalize_pdf_text(Path(r.get("run_dir", "")).name)])
    tbl = Table(rows, repeatRows=1, colWidths=[1.45*inch, 0.8*inch, 0.8*inch, 3.6*inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(Spacer(1, 0.15*inch))
    story.append(tbl)

    def _append_report_dataframe(section_title, table_title, df_tbl):
        if df_tbl is None or getattr(df_tbl, "empty", True):
            return
        story.append(Paragraph(table_title, styles["Small"]))
        # Wrap all cells as Paragraphs so long labels/fit equations do not overflow.
        cols = [str(c) for c in df_tbl.columns]
        data = [[Paragraph(c, styles["Small"]) for c in cols]]
        for _, row in df_tbl.iterrows():
            data.append([Paragraph(str(row.get(c, "")), styles["Small"]) for c in cols])
        avail_w = 7.2 * inch
        ncol = max(len(cols), 1)
        if ncol >= 8:
            col_widths = [avail_w / ncol] * ncol
        elif ncol == 6:
            col_widths = [1.25*inch, 1.1*inch, 0.95*inch, 0.95*inch, 0.95*inch, 1.1*inch]
        elif ncol == 5:
            col_widths = [1.25*inch, 1.15*inch, 1.4*inch, 1.15*inch, 1.0*inch]
        elif ncol == 4:
            col_widths = [1.25*inch, 1.65*inch, 2.0*inch, 2.0*inch]
        elif ncol == 2:
            col_widths = [2.4*inch, 4.6*inch]
        else:
            col_widths = [avail_w / ncol] * ncol
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#d9eaf7")),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 6.8),
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#dddddd")),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.08*inch))

    story.append(PageBreak())

    for cr in case_reports:
        case = cr.get("case", "Case")
        case_dir = Path(cr.get("run_dir", ""))
        story.append(Paragraph(_normalize_pdf_text(case), styles["H1Blue"]))
        story.append(Paragraph("Event Narrative", styles["H2Blue"]))
        for para in _paragraphs_from_markdown(cr.get("narrative", "")):
            story.append(Paragraph(para, styles["Body11"]))
            story.append(Spacer(1, 0.06*inch))
        # Append the same source-term and dose tables shown in the PWR Simulator Results panel.
        report_tables = _source_term_dose_tables_for_report(case_dir, case)
        if report_tables:
            current_section = None
            for item in report_tables:
                section = item.get("section", "Source Term and Dose Tables")
                if section != current_section:
                    story.append(Paragraph(section, styles["H2Blue"]))
                    current_section = section
                _append_report_dataframe(section, item.get("title", "Table"), item.get("df"))

        if cr.get("soe"):
            story.append(Paragraph("Sequence of Events", styles["H2Blue"]))
            soe_rows = [["Time (s)", "Event"]] + [[_fmt(r[0], 1), str(r[1])] for r in cr["soe"][:40]]
            t = Table(soe_rows, repeatRows=1, colWidths=[0.9*inch, 5.6*inch])
            t.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#d9eaf7")), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 8), ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#dddddd")), ("VALIGN", (0,0), (-1,-1), "TOP")]))
            story.append(t)
            story.append(Spacer(1, 0.1*inch))
        if cr.get("ic"):
            story.append(Paragraph("Initial Conditions and Final State", styles["H2Blue"]))
            ic_rows = [["Parameter", "Initial", "Final", "Units"]] + cr["ic"]
            t = Table(ic_rows, repeatRows=1, colWidths=[2.2*inch, 1.3*inch, 1.3*inch, 0.8*inch])
            t.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#d9eaf7")), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 8), ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#dddddd")), ("VALIGN", (0,0), (-1,-1), "TOP")]))
            story.append(t)
            story.append(Spacer(1, 0.1*inch))
        pngs = _find_case_pngs(case_dir)
        if pngs:
            story.append(PageBreak())
            story.append(Paragraph(_normalize_pdf_text(f"{case} — Generated Plots"), styles["H1Blue"]))
            # Put plots two per page.  Each image is scaled to fit one half-page
            # including its caption, preserving government-readable plot labels.
            for j in range(0, len(pngs), 2):
                pair = pngs[j:j+2]
                if j > 0:
                    story.append(PageBreak())
                    story.append(Paragraph(_normalize_pdf_text(f"{case} — Generated Plots"), styles["H1Blue"]))
                for p in pair:
                    try:
                        story.append(Paragraph(_normalize_pdf_text(p.name), styles["Small"]))
                        img = Image(str(p))
                        max_w, max_h = 6.7*inch, 3.05*inch
                        scale = min(max_w / float(img.imageWidth), max_h / float(img.imageHeight), 1.0)
                        img.drawWidth = float(img.imageWidth) * scale
                        img.drawHeight = float(img.imageHeight) * scale
                        story.append(img)
                        story.append(Spacer(1, 0.12*inch))
                    except Exception as e:
                        story.append(Paragraph(_normalize_pdf_text(f"Could not include {p.name}: {e}"), styles["Small"]))
        story.append(PageBreak())

    doc.build(story)
    return pdf_path


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python flare_sim_batch_worker.py <sim_all_worker_config.json>", flush=True)
        return 2

    cfg_path = Path(sys.argv[1])
    cfg = read_json(cfg_path, {})
    work_dir = Path(cfg.get("work_dir", ".")).resolve()
    # run_dir is now a small regular control/status directory.  Simulation outputs
    # are written to ordinary per-case sim_<case>_<timestamp> folders.
    run_dir = Path(cfg.get("run_dir", work_dir / f"sim_all_{time.strftime('%Y%m%d_%H%M%S')}"))
    case_entries = list(cfg.get("case_entries") or [])
    if case_entries:
        cases = [str(e.get("case")) for e in case_entries]
    else:
        cases = list(cfg.get("cases", []))
        case_entries = [{"case": c, "input_path": str(Path(cfg.get("work_dir", ".")) / f"{c}_in.xlsx")} for c in cases]
    batch_tag = str(cfg.get("batch_tag") or time.strftime("%Y%m%d_%H%M%S"))
    fast_mode = bool(cfg.get("fast_mode", False))
    final_report = bool(cfg.get("final_report", False))
    final_report_detail = float(cfg.get("final_report_detail", 0.5) or 0.5)
    final_report_detail = max(0.0, min(1.0, final_report_detail))
    if final_report:
        fast_mode = False
    timeout_s = int(cfg.get("timeout_s", 600))

    run_dir.mkdir(parents=True, exist_ok=True)
    status_path = run_dir / "sim_all_status.json"
    results_path = run_dir / "sim_all_results.json"
    results_csv = run_dir / "FLARE_sim_all_results.csv"

    print(f"[sim-batch-worker] Started: {run_dir}", flush=True)

    status = {
        "status": "running",
        "message": "Batch simulation worker started.",
        "total_runs": len(cases),
        "completed_runs": 0,
        "failed_runs": 0,
        "current_case": None,
        "current_case_pid": None,
        "current_case_elapsed_s": None,
        "current_case_console_log": None,
        "run_dir": str(run_dir),
        "control_dir": str(run_dir),
        "output_mode": "per_case_sim_dirs",
        "batch_tag": batch_tag,
        "final_report": final_report,
        "final_report_detail": final_report_detail,
        "final_report_pdf": None,
        "current_case_run_dir": None,
        "worker_pid": os.getpid(),
        "abort_requested": False,
        "started": now_iso(),
        "last_update": now_iso(),
    }
    write_json_resilient(status_path, status)

    results = []
    case_reports = []
    completed = 0
    failed = 0

    for idx, case_entry in enumerate(case_entries, start=1):
        case = str(case_entry.get("case"))
        input_path = Path(case_entry.get("input_path") or (work_dir / f"{case}_in.xlsx"))
        if abort_requested(run_dir):
            status.update({
                "status": "aborted",
                "message": "Batch run aborted before next case.",
                "abort_requested": True,
                "last_update": now_iso(),
            })
            write_json_resilient(status_path, status)
            return 130

        case_run_dir = work_dir / f"sim_{case}_{batch_tag}"
        case_run_dir.mkdir(parents=True, exist_ok=True)
        console_log = case_run_dir / f"{case}_console.log"
        src = input_path
        dst = case_run_dir / f"{case}_in.xlsx"

        status.update({
            "status": "running",
            "message": f"Preparing {case} ({idx}/{len(cases)}).",
            "current_case": case,
            "current_case_pid": None,
            "current_case_elapsed_s": 0.0,
            "current_case_console_log": str(console_log),
            "current_case_run_dir": str(case_run_dir),
            "last_update": now_iso(),
        })
        write_json_resilient(status_path, status)

        case_result = {
            "case": case,
            "index": idx,
            "status": "unknown",
            "returncode": None,
            "elapsed_s": None,
            "run_dir": str(case_run_dir),
            "out_csv": str(case_run_dir / f"{case}_out.csv"),
            "out_xlsx": str(case_run_dir / f"{case}_out.xlsx"),
            "console_log": str(console_log),
            "error": None,
        }

        try:
            if not src.exists():
                raise FileNotFoundError(f"Input workbook not found: {src}")
            copy_input_with_retries(src, dst)
        except Exception as e:
            failed += 1
            case_result.update({"status": "error", "error": f"Input copy failed: {e}"})
            results.append(case_result)
            completed += 1
            write_json_resilient(results_path, {r["case"]: r for r in results})
            status.update({
                "completed_runs": completed,
                "failed_runs": failed,
                "message": f"Input copy failed for {case} ({idx}/{len(cases)}).",
                "current_case_pid": None,
                "current_case_elapsed_s": None,
                "last_update": now_iso(),
            })
            write_json_resilient(status_path, status)
            continue

        cmd = [sys.executable, "-u", str(work_dir / "flare_sim.py"), case]
        if fast_mode:
            cmd.append("--no-figures")

        print(f"[sim-batch-worker] Running {case} ({idx}/{len(cases)})", flush=True)
        t0 = time.time()
        proc = None
        try:
            with open(console_log, "wb", buffering=0) as logf:
                env = {**os.environ, "PYTHONUTF8": "1", "MPLBACKEND": "Agg"}
                creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0) if sys.platform.startswith("win") else 0
                proc = subprocess.Popen(
                    cmd,
                    cwd=str(case_run_dir),
                    stdout=logf,
                    stderr=subprocess.STDOUT,
                    stdin=subprocess.DEVNULL,
                    env=env,
                    creationflags=creationflags,
                    close_fds=(not sys.platform.startswith("win")),
                )

                while True:
                    rc = proc.poll()
                    elapsed = time.time() - t0
                    status.update({
                        "status": "running",
                        "message": f"Running {case} ({idx}/{len(cases)}).",
                        "current_case": case,
                        "current_case_pid": int(proc.pid),
                        "current_case_elapsed_s": round(elapsed, 1),
                        "current_case_console_log": str(console_log),
                        "current_case_run_dir": str(case_run_dir),
                        "last_update": now_iso(),
                    })
                    write_json_resilient(status_path, status)

                    if abort_requested(run_dir):
                        terminate_process_tree(proc)
                        case_result.update({
                            "status": "aborted",
                            "returncode": proc.poll(),
                            "elapsed_s": round(time.time() - t0, 2),
                            "error": "User requested abort.",
                        })
                        results.append(case_result)
                        write_json_resilient(results_path, {r["case"]: r for r in results})
                        status.update({
                            "status": "aborted",
                            "message": "Batch run aborted by user.",
                            "abort_requested": True,
                            "current_case_pid": None,
                            "last_update": now_iso(),
                        })
                        write_json_resilient(status_path, status)
                        return 130

                    if rc is not None:
                        break
                    if elapsed > timeout_s:
                        terminate_process_tree(proc)
                        raise TimeoutError(f"{case} exceeded timeout of {timeout_s} s")
                    time.sleep(1.0)

                rc = proc.returncode
                elapsed = time.time() - t0
                out_csv = case_run_dir / f"{case}_out.csv"
                out_xlsx = case_run_dir / f"{case}_out.xlsx"
                log_tail = ""
                try:
                    data = console_log.read_bytes()[-12000:]
                    log_tail = data.decode("utf-8", errors="replace")
                except Exception:
                    pass

                ok = (rc == 0 and (out_csv.exists() or out_xlsx.exists()) and "Traceback" not in log_tail)
                if ok:
                    case_result.update({"status": "OK", "returncode": rc, "elapsed_s": round(elapsed, 2)})
                    if final_report:
                        status.update({
                            "message": f"Generating final-report narrative for {case} ({idx}/{len(cases)}).",
                            "current_case_pid": None,
                            "last_update": now_iso(),
                        })
                        write_json_resilient(status_path, status)
                        try:
                            case_report = _generate_case_narrative(case, case_run_dir, work_dir, detail_level=final_report_detail)
                            case_reports.append(case_report)
                            case_result["narrative_md"] = str(case_run_dir / f"{case}_narrative.md")
                            case_result["sequence_of_events_csv"] = str(case_run_dir / f"{case}_sequence_of_events.csv")
                            case_result["initial_final_csv"] = str(case_run_dir / f"{case}_initial_final.csv")
                        except Exception as _nr_err:
                            case_result["narrative_error"] = str(_nr_err)
                else:
                    failed += 1
                    case_result.update({
                        "status": "FAIL",
                        "returncode": rc,
                        "elapsed_s": round(elapsed, 2),
                        "error": "Simulation failed or expected output missing.",
                    })
                completed += 1
                results.append(case_result)
                write_json_resilient(results_path, {r["case"]: r for r in results})

                status.update({
                    "completed_runs": completed,
                    "failed_runs": failed,
                    "current_case": case,
                    "current_case_pid": None,
                    "current_case_elapsed_s": None,
                    "current_case_console_log": str(console_log),
                    "current_case_run_dir": str(case_run_dir),
                    "message": f"Completed {case} ({idx}/{len(cases)}).",
                    "last_update": now_iso(),
                })
                write_json_resilient(status_path, status)

        except Exception as e:
            if proc is not None and proc.poll() is None:
                terminate_process_tree(proc)
            elapsed = time.time() - t0
            failed += 1
            completed += 1
            case_result.update({
                "status": "FAIL",
                "returncode": getattr(proc, "returncode", None),
                "elapsed_s": round(elapsed, 2),
                "error": str(e),
            })
            results.append(case_result)
            write_json_resilient(results_path, {r["case"]: r for r in results})
            status.update({
                "completed_runs": completed,
                "failed_runs": failed,
                "current_case_pid": None,
                "message": f"Failed {case}: {e}",
                "last_update": now_iso(),
            })
            write_json_resilient(status_path, status)

        # Write/update a CSV summary after every case.
        try:
            with open(results_csv, "w", newline="", encoding="utf-8") as f:
                fieldnames = ["case", "index", "status", "returncode", "elapsed_s", "error", "run_dir", "out_csv", "out_xlsx", "console_log"]
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for r in results:
                    writer.writerow({k: r.get(k, "") for k in fieldnames})
        except Exception as e:
            try:
                with open(run_dir / "sim_all_csv_write_warnings.log", "a", encoding="utf-8") as f:
                    f.write(f"[{now_iso()}] Could not write summary CSV: {e}\n")
            except Exception:
                pass

    final_report_pdf = None
    if final_report:
        status.update({
            "status": "running",
            "message": "Compiling Run All Final Report PDF.",
            "current_case": None,
            "current_case_pid": None,
            "current_case_elapsed_s": None,
            "current_case_console_log": None,
            "last_update": now_iso(),
        })
        write_json_resilient(status_path, status)
        try:
            final_report_pdf = _build_final_report(work_dir, run_dir, case_reports, results, detail_level=final_report_detail)
        except Exception as _fr_err:
            try:
                with open(run_dir / "final_report_error.log", "a", encoding="utf-8") as f:
                    f.write(f"[{now_iso()}] Final report failed: {_fr_err}\n")
            except Exception:
                pass

    status.update({
        "status": "complete",
        "message": f"Batch run complete: {completed - failed}/{len(cases)} OK, {failed} failed." + (" Final report compiled." if final_report_pdf else (" Final report requested but could not be compiled." if final_report else "")),
        "final_report": final_report,
        "final_report_detail": final_report_detail,
        "final_report_pdf": str(final_report_pdf) if final_report_pdf else None,
        "completed_runs": completed,
        "failed_runs": failed,
        "current_case": None,
        "current_case_pid": None,
        "current_case_elapsed_s": None,
        "current_case_console_log": None,
        "current_case_run_dir": None,
        "last_update": now_iso(),
    })
    write_json_resilient(status_path, status)
    print(f"[sim-batch-worker] Complete: {completed - failed}/{len(cases)} OK, {failed} failed", flush=True)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
