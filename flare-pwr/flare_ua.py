"""
flare_ua.py   -   Streamlit uncertainty analysis UI for flare_sim.py
===============================================================================
Run with:
    streamlit run flare_ua.py
"""

import io
import os
import re
import sys
import csv
import json
import base64
import subprocess
import requests
import time as _time
import shutil as _shutil
from datetime import datetime
from datetime import datetime
import datetime as _datetime
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from pathlib import Path
from openpyxl import load_workbook

# ── Optional heavy deps (PDF / stats) ────────────────────────────────────────
try:
    from scipy.special import gammaln as _gammaln
    from scipy.stats import pearsonr as _pearsonr
    _SCIPY_OK = True
except ImportError:
    _SCIPY_OK = False

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as _mticker
from matplotlib.backends.backend_pdf import PdfPages as _PdfPages

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FLARE · Uncertainty Analysis",
    page_icon="⚛",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Styling (matches flare_ui.py) ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;600&display=swap');

:root {
    --bg:       #f5f7fa;
    --surface:  #ffffff;
    --sidebar:  #1a1f2e;
    --border:   #d0d7de;
    --accent:   #0969da;
    --accent2:  #1a7f37;
    --warn:     #9a6700;
    --danger:   #cf222e;
    --text:     #1f2328;
    --muted:    #57606a;
    --sid-text: #e6edf3;
    --sid-muted:#8b949e;
    --dark-surface: #0d1117;
    --dark-border:  #30363d;
}

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
    color: var(--text);
}
.stApp { background: var(--bg); }
header { background: transparent !important; }

section[data-testid="stSidebar"] {
    background: var(--sidebar) !important;
    border-right: 1px solid var(--dark-border);
    color: var(--sid-text);
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label {
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] .stCaption {
    color: var(--sid-muted) !important;
}
/* Keep expander header dark when expanded so label stays visible */
section[data-testid="stSidebar"] details,
section[data-testid="stSidebar"] details[open],
section[data-testid="stSidebar"] [data-testid="stExpander"],
section[data-testid="stSidebar"] [data-testid="stExpander"] > details {
    background: var(--dark-surface) !important;
    border: 1px solid var(--dark-border) !important;
    border-radius: 4px;
}
section[data-testid="stSidebar"] details summary,
section[data-testid="stSidebar"] details[open] summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    background: var(--dark-surface) !important;
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] details summary p,
section[data-testid="stSidebar"] details summary span,
section[data-testid="stSidebar"] details[open] summary p,
section[data-testid="stSidebar"] details[open] summary span {
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] {
    background: var(--dark-surface);
    border: 1px solid var(--dark-border);
    border-radius: 6px;
}
section[data-testid="stSidebar"] .stSelectbox span { color: #ffffff !important; }
div[role="listbox"] { background: var(--dark-surface) !important; }
div[role="option"] { color: #ffffff !important; }
div[role="option"]:hover { background: #21262d !important; }
section[data-testid="stSidebar"] input {
    background: var(--dark-surface) !important;
    color: #ffffff !important;
    border: 1px solid var(--dark-border) !important;
}
section[data-testid="stSidebar"] button[kind="primary"] {
    background: var(--danger) !important;
    color: white !important;
}
section[data-testid="stSidebar"] button[kind="primary"]:hover {
    background: #a40e26 !important;
}

/* Sidebar secondary/default buttons: keep Load buttons readable on dark theme. */
section[data-testid="stSidebar"] div.stButton > button,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button,
section[data-testid="stSidebar"] button[kind="secondary"] {
    background: #263142 !important;
    color: #ffffff !important;
    border: 1px solid #f97316 !important;
    border-radius: 6px !important;
    opacity: 1 !important;
}
section[data-testid="stSidebar"] div.stButton > button:hover,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button:hover,
section[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background: #f97316 !important;
    color: #ffffff !important;
    border-color: #fb923c !important;
}
section[data-testid="stSidebar"] div.stButton > button:disabled,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button:disabled,
section[data-testid="stSidebar"] button[kind="secondary"]:disabled,
section[data-testid="stSidebar"] button[disabled] {
    background: #374151 !important;
    color: #d1d5db !important;
    border: 1px solid #6b7280 !important;
    opacity: 1 !important;
}

section[data-testid="stSidebar"] code {
    color: #ffffff !important;
    background: var(--dark-border) !important;
    padding: 2px 6px; border-radius: 4px;
}
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace; }
button[role="tab"] { color: var(--muted); }
button[role="tab"][aria-selected="true"] {
    color: var(--accent);
    border-bottom: 2px solid var(--accent);
}
.console {
    background: #0d1117;
    border: 1px solid #30363d;
    border-radius: 6px;
    padding: 1rem;
    color: #7ee787;
    font-family: 'IBM Plex Mono', monospace;
    white-space: pre-wrap;
    font-size: 0.85rem;
}
.hdiv { border-top: 1px solid var(--border); margin: 1.5rem 0; }
.metric-grid-labeled { margin-bottom: 0.25rem; }
.metric-grid-title {
    font-family: 'IBM Plex Sans', sans-serif;
    font-size: 0.85rem;
    font-weight: 600;
    color: var(--text-muted, #888);
    margin-bottom: 0.4rem;
    letter-spacing: 0.02em;
    text-transform: uppercase;
}
.metric-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:0.75rem; }
.metric-tile {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1rem;
    text-align: center;
}
.metric-tile .val {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.4rem;
    color: var(--accent);
}
.metric-tile.ok  .val { color: var(--accent2); }
.metric-tile.warn .val { color: var(--warn); }
.metric-tile.danger .val { color: var(--danger); }
/* ── Tooltip (?) icon — targets the actual SVG stroke ── */
section[data-testid="stSidebar"] .stTooltipHoverTarget svg.icon {
    stroke: #f97316 !important;
}
section[data-testid="stSidebar"] .stTooltipHoverTarget:hover svg.icon {
    stroke: #ffffff !important;
}
.stTooltipHoverTarget svg.icon {
    stroke: #0969da !important;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────

WORK_DIR = Path(__file__).parent

def _load_buddy_b64():
    """Return a base64 data URI for the FLARE Buddy icon, or None."""
    for _dir in (WORK_DIR / "icons", WORK_DIR / "Icons"):
        _p = _dir / "FLAREBUDDY.png"
        if _p.exists():
            try:
                return "data:image/png;base64," + base64.b64encode(_p.read_bytes()).decode()
            except Exception:
                return None
    return None

_BUDDY_B64 = _load_buddy_b64()

def _buddy_expander_label(title: str) -> str:
    """Return an expander label with an inline FLARE Buddy thumbnail, or fall back to 🤖."""
    if _BUDDY_B64:
        return (
            f"<img src='{_BUDDY_B64}' style='height:1.3rem;width:1.3rem;"
            f"border-radius:50%;object-fit:cover;vertical-align:middle;"
            f"margin-right:0.4rem;'/>{title}"
        )
    return f"🤖  {title}"



def _runtime_dir(base_dir: Path | None = None) -> Path:
    """Return the FLARE runtime folder, preferring runtime/ then Runtime/."""
    base = Path(base_dir) if base_dir is not None else WORK_DIR
    lower = base / "runtime"
    upper = base / "Runtime"
    if lower.exists():
        rt = lower
    elif upper.exists():
        rt = upper
    else:
        rt = lower
        rt.mkdir(parents=True, exist_ok=True)
    return rt


def _runtime_file(name: str, base_dir: Path | None = None) -> Path:
    return _runtime_dir(base_dir) / name



# ── AI narrative helpers ─────────────────────────────────────────────────────
def _read_config(key):
    """Read KEY=value from runtime/flare_config.txt (or Runtime/flare_config.txt)."""
    cfg = _runtime_file("flare_config.txt")
    if cfg.exists():
        for line in cfg.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k.strip().upper() == key.upper():
                return v.strip().strip('"').strip("'")
    return None

def _load_api_key():
    return (st.session_state.get("user_api_key")
            or _read_config("ANTHROPIC_API_KEY")
            or os.environ.get("ANTHROPIC_API_KEY"))

def _load_model():
    return (_read_config("ANTHROPIC_MODEL")
            or os.environ.get("ANTHROPIC_MODEL")
            or "claude-sonnet-4-5")

def _anthropic_text(system_prompt, user_prompt, max_tokens=4000, timeout=120):
    """Call Claude and return (text, stop_reason).

    Raises RuntimeError on API / HTTP errors.  The stop_reason lets callers
    detect truncation ('max_tokens') and warn the user.
    """
    api_key = _load_api_key()
    if not api_key:
        raise RuntimeError("No API key found. Add ANTHROPIC_API_KEY to runtime/flare_config.txt (or Runtime/flare_config.txt) or save it on the FLARE Home page.")
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": _load_model(),
            "max_tokens": max_tokens,
            "temperature": 0.2,
            "system": system_prompt,
            "messages": [{"role": "user", "content": user_prompt}],
        },
        timeout=(20, timeout),
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"Anthropic API error {resp.status_code}: {resp.text[:600]}")
    data = resp.json()
    text = "\n".join(
        block.get("text", "") for block in data.get("content", [])
        if block.get("type") == "text"
    ).strip()
    stop_reason = data.get("stop_reason", "")
    return text, stop_reason

# Hardwired variable catalogue: {var_name: (label, default_dist, base, p1, p2, help)}
DEFAULT_UA_VARIABLES = {
    "F_r":           ("Radial peaking F_r",          "uniform",   1.30,    1.20,    1.45,
                      "Radial hot-channel factor: ratio of peak-to-average pin power. "
                      "Drives hot-pin clad temperature and DNBR. Typical range 1.2–1.45."),
    "F_z":           ("Axial peaking F_z",            "uniform",   1.55,    1.45,    1.65,
                      "Axial peaking factor (chopped cosine). Combined with F_r for "
                      "hot-pin heat flux. Typical fresh-core range 1.45–1.65."),
    "total_power":   ("Rated power [MW]",             "normal",  575.0,  575.0,    10.0,
                      "Rated thermal power. Uncertainty reflects calorimetric measurement "
                      "error (~1–2% of rated). Modelled as normal."),
    "temp_core_exit":("Core-exit temp [°C]",          "normal",  312.8,  312.8,     3.0,
                      "Initial bulk coolant temperature at core exit. Affects initial "
                      "enthalpy and moderator reactivity feedback. Uncertainty reflects "
                      "thermocouple calibration and flow mixing (~±3°C)."),
    "diameter_break":("Break diameter [m]",           "uniform",   0.20,    0.15,    0.25,
                      "Equivalent break diameter. Drives blowdown rate and peak clad "
                      "temperature. Use a uniform distribution to represent break spectrum "
                      "uncertainty."),
    "Cd_sub":        ("Break Cd (subcooled)",         "uniform",   0.93,    0.85,    1.00,
                      "Discharge coefficient for subcooled critical flow at the break. "
                      "Accounts for vena contracta and non-equilibrium effects. "
                      "Range 0.85–1.0 per NUREG validation studies."),
    "k_fuel":        ("Fuel conductivity [W/m·K]",   "uniform",   3.0,     2.5,     3.5,
                      "UO₂ thermal conductivity. Decreases with burnup and temperature. "
                      "Controls fuel centreline temperature and stored energy. "
                      "Typical BOL ~3 W/m·K; EOL ~2.5 W/m·K."),
    "h_gap":         ("Gap conductance [W/m²·K]",    "uniform", 6000.0, 3000.0,  9000.0,
                      "Fuel-clad gap conductance — the largest single contributor to fuel "
                      "temperature uncertainty. Depends on gap size, fill gas, and contact "
                      "pressure. Wide uniform range reflects as-fabricated variation."),
    "r_fuel_m":      ("Fuel radius [m]",              "uniform", 0.00418, 0.00390, 0.00440,
                      "Fuel pellet outer radius. Manufacturing tolerance ~±0.025 mm. "
                      "Affects fuel heat capacity, stored energy, and gap conductance area."),
    "rho_fuel":      ("Fuel density [kg/m³]",         "uniform", 10400.0, 10000.0, 10800.0,
                      "UO₂ pellet density (theoretical ~10 960 kg/m³). "
                      "Affects fuel heat capacity and total stored energy in the core."),
    "cp_fuel":       ("Fuel Cp [J/kg·K]",             "uniform",   330.0,   300.0,   360.0,
                      "UO₂ specific heat capacity. Increases with temperature. "
                      "Controls the rate of fuel temperature rise during a transient."),
    "delta_clad":    ("Clad thickness [m]",           "uniform", 0.00057, 0.00050, 0.00065,
                      "Zircaloy cladding wall thickness. Manufacturing tolerance ~±0.05 mm. "
                      "Affects conduction resistance between fuel and coolant."),
    "lambda_clad":   ("Clad conductivity [W/m·K]",   "uniform",   13.0,    11.0,    15.0,
                      "Zircaloy-4 thermal conductivity. Relatively insensitive to "
                      "temperature in the operating range. Uncertainty ~±15%."),
    "htc_post_chf":  ("Bromley HTC cap [W/m²·K]",   "uniform",   450.0,   300.0,   600.0,
                      "Upper limit on the Bromley film-boiling HTC after CHF. "
                      "Represents uncertainty in the inverted-annular flow regime at "
                      "high pressure. Lower values give more conservative PCT."),
    "htc_core_mult": ("Core HTC multiplier",          "uniform",   1.0,     0.8,     1.2,
                      "Multiplicative factor on the forced-convection (Dittus-Boelter / "
                      "Churchill-Chu) core HTC during normal pre-CHF operation. "
                      "Values < 1 reduce cladding-to-coolant heat transfer, raising clad "
                      "temperature and advancing DNB onset. Conservative direction: < 1. "
                      "Typical UA range ±20% (0.8–1.2)."),
    "htc_fb_mult":   ("Film boiling HTC multiplier",  "uniform",   1.0,     0.5,     2.0,
                      "Multiplicative factor on the post-CHF film boiling HTC — applied "
                      "to both the Bromley (inverted-annular) result and the steam "
                      "Dittus-Boelter (reflood / uncovery) result before the htc_post_chf "
                      "cap is enforced. Values < 1 worsen post-CHF cooling and raise PCT. "
                      "Conservative direction: < 1. "
                      "Wide range reflects factor-of-2 scatter in film boiling data."),
    "rewet_dT_K":    ("Leidenfrost ΔT [K]",          "uniform",   200.0,   150.0,   300.0,
                      "Maximum cladding superheat above T_sat at which rewetting "
                      "(return to nucleate boiling) is permitted. "
                      "Higher values delay rewetting and produce higher PCT."),
}

DIST_OPTIONS = ["uniform", "normal", "lognormal", "triangular"]

DIST_PARAM_LABELS = {
    "uniform":    ("Lower bound",  "Upper bound"),
    "normal":     ("Mean",         "Std deviation"),
    "lognormal":  ("ln(mean)",     "ln(std)"),
    "triangular": ("Lower bound",  "Upper bound"),
}

UA_VARIABLES_FILE = _runtime_file("flare_ua_variables.json")

def _normalize_ua_variable_catalog(raw):
    """Return UA variable catalogue in the internal tuple format.

    JSON format expected:
      {
        "parameter_name": {
          "label": "Display label",
          "distribution": "uniform|normal|lognormal|triangular",
          "base": 1.0,
          "p1": 0.8,
          "p2": 1.2,
          "help": "Sidebar help text"
        }
      }

    Keys beginning with '_' are ignored so the JSON file may contain comments
    such as "_note" without becoming a model parameter.
    """
    if isinstance(raw, list):
        items = []
        for entry in raw:
            if not isinstance(entry, dict):
                continue
            name = entry.get("name") or entry.get("variable") or entry.get("var")
            if name:
                items.append((str(name), entry))
    elif isinstance(raw, dict):
        items = [(str(k), v) for k, v in raw.items()
                 if not str(k).startswith("_")]
    else:
        raise ValueError("UA variable catalogue must be a JSON object or list.")

    catalog = {}
    for var, cfg in items:
        if not isinstance(cfg, dict):
            raise ValueError(f"UA variable '{var}' must be a JSON object.")
        label = str(cfg.get("label", var))
        dist  = str(cfg.get("distribution", cfg.get("dist", "uniform"))).lower()
        if dist not in DIST_OPTIONS:
            raise ValueError(
                f"UA variable '{var}' has invalid distribution '{dist}'. "
                f"Allowed: {', '.join(DIST_OPTIONS)}"
            )
        try:
            base = float(cfg.get("base"))
            p1   = float(cfg.get("p1", cfg.get("param1")))
            p2   = float(cfg.get("p2", cfg.get("param2")))
        except Exception as exc:
            raise ValueError(
                f"UA variable '{var}' must define numeric base, p1, and p2 values."
            ) from exc
        help_text = str(cfg.get("help", cfg.get("description", "")))
        sheet = str(cfg.get("sheet", "")).strip()
        catalog[var] = (label, dist, base, p1, p2, help_text, sheet)

    if not catalog:
        raise ValueError("UA variable catalogue contains no usable variables.")
    return catalog

def _write_default_ua_variables_file(path: Path):
    """Create an editable JSON catalogue from the built-in defaults."""
    raw = {}
    for var, cfg in DEFAULT_UA_VARIABLES.items():
        label, dist, base, p1, p2, help_text = cfg[:6]
        sheet = cfg[6] if len(cfg) > 6 else ""
        raw[var] = {
            "label": label,
            "distribution": dist,
            "base": base,
            "p1": p1,
            "p2": p2,
            "help": help_text,
        }
        if sheet:
            raw[var]["sheet"] = sheet
    path.write_text(json.dumps(raw, indent=2, ensure_ascii=False), encoding="utf-8")

def load_ua_variables():
    """Load sidebar UA parameter catalogue from runtime/flare_ua_variables.json.

    If the file is absent, it is created from the built-in defaults. If the
    file is malformed, the UI continues with the built-in defaults and shows a
    warning rather than preventing FLARE from loading.
    """
    try:
        if not UA_VARIABLES_FILE.exists():
            _write_default_ua_variables_file(UA_VARIABLES_FILE)
        raw = json.loads(UA_VARIABLES_FILE.read_text(encoding="utf-8"))
        return _normalize_ua_variable_catalog(raw)
    except Exception as exc:
        st.warning(
            f"Could not load `{UA_VARIABLES_FILE.name}`; using built-in UA variable defaults. "
            f"Reason: {exc}"
        )
        return DEFAULT_UA_VARIABLES.copy()

UA_VARIABLES = load_ua_variables()

PLOT_BG    = "#ffffff"
PLOT_PAPER = "#f5f7fa"
PLOT_GRID  = "#d0d7de"
PLOT_TEXT  = "#1f2328"
C = ["#0969da","#cf222e","#1a7f37","#9a6700","#6e40c9","#bc4c00"]


# Time-series variables offered in the UA overlay selector.
# This list intentionally mirrors the variables plotted by the PWR Simulator
# Results tab (flare_ui.py / flare_sim.py), so the UA plot selector stays
# focused on quantities that are already meaningful in normal FLARE review.
UA_PLOTTED_TS_COLUMNS = [
    # Fig 1-5: primary response, inventory, and discharge
    "RCS Pressure (kPa)",
    "Equilibrium Quality (-)",
    "Void Fraction (-)",
    "Total Mass Scaled",
    "Vessel Level (m)",
    "Break Flow (kg/s)",
    "PORV Mass Flow (kg/s)",

    # Fig 6-9: accumulator and injection / letdown flows
    "Accumulator Pressure (kPa)",
    "Accumulator Temperature (K)",
    "Accumulator Level (m)",
    "Accumulator Flow (kg/s)",
    "CVCS Makeup (kg/s)",
    "CVCS Letdown (kg/s)",
    "HPSI Flow (kg/s)",
    "LPSI Flow (kg/s)",
    "SI Pumped Total (kg/s)",

    # Fig 10-13: power, pump, and SG response
    "RK Total Power (MW)",
    "Core Power (MW)",
    "Pump Speed (rpm)",
    "Pump Velocity (m/s)",
    "SG Heat Removal (MW)",

    # Fig 14-16: core temperatures, HTC, and DNBR
    "Clad Surface Temp (K)",
    "RCS Temperature (K)",
    "Fuel Avg Temp (K)",
    "Hot Pin Clad Temp (K)",
    "Hot Pin Fuel Temp (K)",
    "Clad HTC (W/m2-K)",
    "DNBR",

    # Fig 17: reactivity components
    "Reactivity scram (pcm)",
    "Reactivity ext (pcm)",
    "Reactivity Boron (pcm)",
    "Reactivity Doppler (pcm)",
    "Reactivity Moderator (pcm)",
    "Reactivity net (pcm)",

    # Supplemental PWR Simulator plots
    "Pressurizer Level (m)",
    "Pressurizer Level (norm)",
    "Zr Oxidation Hot Pin ECR (%)",
    "Zr Oxidation Mean Oxidizing Rod ECR (%)",
    "H2 Generated (kg)",
    "H2 Full Core Cladding Reaction (kg)",
    "Zr Oxidizing Rods (est.)",
]

# Maps scalar output variable names to (TS CSV column, offset, unit label)
TS_MAP = {
    "hot_pin_clad_peak_K":  ("Hot Pin Clad Temp",           -273.15, "°C"),
    "hot_pin_clad_final_K": ("Hot Pin Clad Temp",           -273.15, "°C"),
    "avg_clad_peak_K":      ("Clad Surface Temp",           -273.15, "°C"),
    "avg_clad_final_K":     ("Clad Surface Temp",           -273.15, "°C"),
    "hot_pin_fuel_peak_K":  ("Hot Pin Fuel Temp",           -273.15, "°C"),
    "P_min_kPa":            ("RCS Pressure (kPa)",           0,      "kPa"),
    "P_max_kPa":            ("RCS Pressure (kPa)",           0,      "kPa"),
    "T_max_K":              ("RCS Temperature",             -273.15, "°C"),
    "DNBR_min":             ("DNBR",                         0,      ""),
    "N_fail_DNB":           ("Rod Failures DNB (est.)",      0,      "rods"),
    "N_fail_gap":           ("Rod Failures Gap (est.)",      0,      "rods"),
    "N_fail_eiv":           ("Rod Failures EarlyIV (est.)",  0,      "rods"),
    "P_peak_MW":            ("RK Total Power (MW)",          0,      "MW"),
}

# Unit conversion: each entry is (si_label, eng_label, si_to_eng_fn)
# Keyed by column name suffix patterns
# UNIT_CONV: suffix → (si_label, eng_label, si_fn, eng_fn, eng_scale)
# eng_scale is the multiplicative factor only (no offset)  -  used for std dev.
UNIT_CONV = {
    # Temperature columns (stored in K)
    "(K)":    ("°C",   "°F",   lambda v: v - 273.15, lambda v: (v - 273.15)*9/5 + 32, 9/5),
    # Pressure
    "(kPa)":  ("kPa",  "psia", lambda v: v,          lambda v: v * 0.145038,           0.145038),
    # Mass flow
    "(kg/s)": ("kg/s", "lb/s", lambda v: v,          lambda v: v * 2.20462,            2.20462),
    # Power  -  keep MW in both
    "(MW)":   ("MW",   "MW",   lambda v: v,          lambda v: v,                      1.0),
    # Dimensionless / other  -  no conversion
    "[-]":    ("",     "",     lambda v: v,          lambda v: v,                      1.0),
}

# Scalar result column name → unit suffix for lookup in UNIT_CONV
SCALAR_UNIT = {
    "hot_pin_clad_peak_K":  "(K)",
    "hot_pin_clad_final_K": "(K)",
    "avg_clad_peak_K":      "(K)",
    "avg_clad_final_K":     "(K)",
    "hot_pin_fuel_peak_K":  "(K)",
    "T_max_K":              "(K)",
    "P_min_kPa":            "(kPa)",
    "P_max_kPa":            "(kPa)",
    "P_peak_MW":            "(MW)",
}


def _safe_plot_token(text, max_len=48):
    """Return a filesystem-safe token for UA plot filenames."""
    token = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(text)).strip("_")
    return (token[:max_len] or "plot")


def _ua_element_key(prefix, *parts):
    """Build a stable, unique Streamlit key for dynamically rendered UA elements.

    Streamlit now raises StreamlitDuplicateElementId when multiple Plotly
    charts are rendered with the same inferred identity.  UA result tabs can
    render the same chart type repeatedly for different selected cases,
    response variables, time-series variables, or FLARECON scalar loops.
    Including the current rendering context in the key prevents those
    collisions while keeping widget identity deterministic across reruns.
    """
    safe_parts = [_safe_plot_token(prefix)]
    for part in parts:
        if part is None:
            continue
        safe_parts.append(_safe_plot_token(part))
    return "ua__" + "__".join(safe_parts)




# Narrative-detail slider has 21 detents (0.00 through 1.00 in 0.05
# increments).  Keep every detent mapped to a distinct expected length so the
# user sees a meaningful change at each slider step.
UA_NARRATIVE_DETAIL_WORD_TARGETS = {
    0.00: 300,
    0.05: 400,
    0.10: 500,
    0.15: 650,
    0.20: 800,
    0.25: 1000,
    0.30: 1200,
    0.35: 1450,
    0.40: 1700,
    0.45: 2000,
    0.50: 2300,
    0.55: 2600,
    0.60: 2900,
    0.65: 3200,
    0.70: 3500,
    0.75: 3800,
    0.80: 4100,
    0.85: 4400,
    0.90: 4700,
    0.95: 5000,
    1.00: 5300,
}

def _ua_detail_level(detail: float) -> float:
    """Snap narrative detail to the nearest 0.05 slider detent."""
    try:
        level = round(float(detail) / 0.05) * 0.05
    except Exception:
        level = 0.55
    return round(max(0.0, min(1.0, level)), 2)

def _ua_detail_word_target(detail: float) -> int:
    """Return the target narrative word count for a 0.05 slider detent."""
    return int(UA_NARRATIVE_DETAIL_WORD_TARGETS.get(_ua_detail_level(detail), 2600))

def _ua_detail_word_range(detail: float) -> tuple[int, int]:
    """Return an allowed word-count range around the detent target.

    The minimum setting is intentionally abstract-length, while all higher
    detents use a +/-15% tolerance around the target.
    """
    target = _ua_detail_word_target(detail)
    lo = max(250, int(round(target * 0.85)))
    hi = int(round(target * 1.15))
    return lo, hi



def _ua_trim_to_word_limit(text: str, max_words: int) -> str:
    """Hard-limit AI-generated markdown section length by word count.

    Anthropic length instructions are generally good but can drift high when a
    report is generated in multiple independent section calls.  This clamp keeps
    the sum of the section calls aligned with the user-selected narrative detail.
    """
    try:
        max_words = int(max_words)
    except Exception:
        return text or ""
    if max_words <= 0 or not text:
        return text or ""
    words = re.findall(r"\S+", text)
    if len(words) <= max_words:
        return text
    clipped = " ".join(words[:max_words]).strip()
    # Prefer ending at a sentence boundary near the requested limit.
    floor = max(0, int(len(clipped) * 0.85))
    sentence_end = max(clipped.rfind(". ", floor), clipped.rfind("; ", floor), clipped.rfind("! ", floor), clipped.rfind("? ", floor))
    if sentence_end > floor:
        clipped = clipped[:sentence_end + 1].strip()
    return clipped



def _ua_normalize_ai_section_markdown(text: str, section_title: str | None = None) -> str:
    """Keep section headings, but force body text to render as normal paragraphs.

    The section generator may return the requested heading and the first body
    sentence on the same physical Markdown line, e.g.
    ``### Summary This analysis ...``.  Streamlit renders the entire line as a
    heading.  This normalizer splits that pattern into a heading line followed
    by ordinary paragraph text.  It also strips paragraph-length full-line bold
    markup and demotes any extra intra-section headings.
    """
    if not text:
        return ""

    def _strip_inline_bold(s: str) -> str:
        s = s.strip()
        m = re.match(r"^\*\*(.+?)\*\*$", s)
        return m.group(1).strip() if m else s

    def _split_expected_title(heading_text: str, expected_title: str | None):
        title_text = _strip_inline_bold(heading_text.strip().strip("#").strip())
        if expected_title:
            exp = str(expected_title).strip()
            # Match the expected section title at the start, tolerating case
            # differences and optional punctuation before the body sentence.
            if title_text.lower().startswith(exp.lower()):
                rest = title_text[len(exp):].strip()
                rest = re.sub(r"^[-–—:.;\s]+", "", rest).strip()
                return exp, rest
        return title_text, ""

    lines = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    out: list[str] = []
    kept_primary_heading = False
    preserve_all_headings = section_title is None

    for raw in lines:
        line = raw.rstrip()
        stripped = line.strip()
        if not stripped:
            out.append("")
            continue

        m = re.match(r"^#{1,6}\s+(.+?)\s*$", stripped)
        if m:
            title, rest = _split_expected_title(m.group(1), section_title)
            use_heading = preserve_all_headings or not kept_primary_heading
            if use_heading:
                out.append(f"### {title}")
                kept_primary_heading = True
            else:
                words = title.split()
                if len(words) <= 8 and not re.search(r"[.;:,]$", title):
                    out.append(f"**{title}**")
                else:
                    out.append(title)
            if rest:
                out.append("")
                out.append(rest)
            continue

        # If the model made an entire paragraph bold, strip the bold markers for
        # paragraph-length text while preserving genuinely short emphasis lines.
        m = re.match(r"^\*\*(.+?)\*\*\s*$", stripped)
        if m and len(m.group(1).split()) > 10:
            out.append(m.group(1).strip())
            continue

        out.append(line)

    # Collapse excessive blank lines while preserving paragraph separation.
    cleaned = "\n".join(out)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    if section_title and not re.match(r"^#{1,6}\s+", cleaned):
        cleaned = f"### {section_title}\n\n{cleaned}".strip()
    return cleaned

def _ua_detail_descriptor(detail: float) -> str:
    """Return a concise descriptive label for the narrative detent."""
    level = _ua_detail_level(detail)
    if level <= 0.05:
        return "technical abstract / extended abstract"
    if level <= 0.15:
        return "short executive technical summary"
    if level <= 0.30:
        return "brief technical note"
    if level <= 0.50:
        return "moderate technical discussion"
    if level <= 0.70:
        return "full technical narrative"
    if level <= 0.90:
        return "long-form technical discussion"
    return "comprehensive technical report narrative"


def _ua_convert_series_for_plot(values, col_name, use_english=False):
    arr = pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(dtype=float)
    for suffix, conv in UNIT_CONV.items():
        if str(col_name).endswith(suffix):
            si_lbl, eng_lbl, si_fn, eng_fn = conv[:4]
            fn = eng_fn if use_english else si_fn
            return fn(arr), (eng_lbl if use_english else si_lbl)
    return arr, ""


def _ua_convert_scalar_for_plot(values, col_name, use_english=False):
    arr = pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(dtype=float)
    suffix = SCALAR_UNIT.get(str(col_name))
    if suffix and suffix in UNIT_CONV:
        si_lbl, eng_lbl, si_fn, eng_fn, _scale = UNIT_CONV[suffix]
        fn = eng_fn if use_english else si_fn
        return fn(arr), (eng_lbl if use_english else si_lbl)
    return arr, ""


# FLARECON scalar outputs are stored in UA results columns without normal unit
# suffixes, so they need a separate display-unit map when the UI is set to
# English units.
CON_SCALAR_UNIT = {
    "CON_P_peak_kPa":        ("kPa",  "psia", lambda v: v, lambda v: v * 0.145038),
    "CON_T_peak_C":          ("°C",   "°F",   lambda v: v, lambda v: v * 9.0 / 5.0 + 32.0),
    "CON_H2_peak_volpct":    ("vol%", "vol%", lambda v: v, lambda v: v),
    "CON_sump_level_peak_m": ("m",    "ft",   lambda v: v, lambda v: v * 3.28084),
}

CON_SCALAR_BASE_LABEL = {
    "CON_P_peak_kPa":        "Containment Peak Pressure",
    "CON_T_peak_C":          "Containment Peak Gas Temp",
    "CON_H2_peak_volpct":    "Peak H₂ Concentration",
    "CON_sump_level_peak_m": "Peak Sump Level",
}


def _ua_convert_con_scalar_for_plot(values, col_name, use_english=False):
    arr = pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(dtype=float)
    unit_cfg = CON_SCALAR_UNIT.get(str(col_name))
    if unit_cfg:
        si_lbl, eng_lbl, si_fn, eng_fn = unit_cfg
        return (eng_fn(arr), eng_lbl) if use_english else (si_fn(arr), si_lbl)
    return arr, ""


def _ua_con_scalar_display_label(col_name, use_english=False):
    base = CON_SCALAR_BASE_LABEL.get(str(col_name), str(col_name).replace("_", " "))
    _dummy, unit = _ua_convert_con_scalar_for_plot([0.0], col_name, use_english)
    return f"{base} [{unit}]" if unit else base


# ══════════════════════════════════════════════════════════════════════════════
# ── QD-RELAP / BEPU synthesis methods ────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

# ── Wilks order-statistic confidence ─────────────────────────────────────────

def _wilks_confidence_beta(n: int, m: int, gamma: float) -> float:
    """Return the one-sided confidence β for the m-th largest of n samples
    at coverage γ, using the exact binomial formula (Martin 2011, Eq. 1).
    """
    if not _SCIPY_OK:
        # Fallback: normal approximation
        return float(np.clip(1 - gamma ** n, 0, 1))
    total = 0.0
    for j in range(1, m + 1):
        i = n - m + j
        log_binom = (_gammaln(n + 1)
                     - _gammaln(i + 1)
                     - _gammaln(n - i + 1))
        total += np.exp(log_binom) * gamma ** i * (1 - gamma) ** (n - i)
    return float(1.0 - total)


# ── Stepwise importance analysis (Martin, Nucl. Technol. 175, 2011) ──────────

def _build_importance_table(df_ok: pd.DataFrame, out_col: str,
                             in_cols: list, r_min: float = 0.0):
    """Stepwise multiple-regression importance analysis.

    Returns (table_df, sqrt_sum, ratio_to_actual, actual_std, residual_std).
    """
    y_raw = pd.to_numeric(df_ok[out_col], errors="coerce").values.copy()
    valid = np.isfinite(y_raw)
    if valid.sum() < 4:
        return pd.DataFrame(), 0.0, 0.0, 0.0, 0.0

    y = y_raw[valid]
    # Build input matrix (only finite rows)
    x_labels = []
    x_mat_cols = []
    for ic in in_cols:
        xv = pd.to_numeric(df_ok[ic], errors="coerce").values[valid]
        if np.isfinite(xv).all():
            x_labels.append(ic)
            x_mat_cols.append(xv)
    if not x_labels:
        return pd.DataFrame(), 0.0, 0.0, 0.0, 0.0

    sv_mat = np.column_stack(x_mat_cols)
    active_labels = x_labels[:]
    remaining = y.copy()
    rem_sv = sv_mat.copy()
    table_rows = []
    actual_std = float(np.std(y, ddof=1))

    def _adj_r2(yy, yhat, n_obs, n_params):
        ss_res = np.sum((yy - yhat) ** 2)
        ss_tot = np.sum((yy - yy.mean()) ** 2)
        r2 = 1 - ss_res / (ss_tot + 1e-12)
        return 1 - (1 - r2) * (n_obs - 1) / max(n_obs - n_params - 1, 1)

    for _step in range(len(active_labels)):
        if not active_labels:
            break
        if _SCIPY_OK:
            cors = [float(_pearsonr(remaining, rem_sv[:, i])[0])
                    for i in range(len(active_labels))]
        else:
            cors = []
            for i in range(len(active_labels)):
                xv = rem_sv[:, i]
                if np.std(xv) > 0 and np.std(remaining) > 0:
                    cors.append(float(np.corrcoef(remaining, xv)[0, 1]))
                else:
                    cors.append(0.0)
        max_abs_r = max(abs(c) for c in cors)
        if max_abs_r < r_min:
            break

        best_idx = int(np.argmax([abs(c) for c in cors]))
        best_lbl = active_labels[best_idx]
        best_cor = cors[best_idx]
        param_x = rem_sv[:, best_idx]
        n_obs = len(remaining)

        # Best polynomial degree by adjusted R²
        c1 = np.polyfit(param_x, remaining, 1)
        f1 = np.polyval(c1, param_x)
        ar2_1 = _adj_r2(remaining, f1, n_obs, 1)
        c2 = np.polyfit(param_x, remaining, 2)
        f2 = np.polyval(c2, param_x)
        ar2_2 = _adj_r2(remaining, f2, n_obs, 2)

        if ar2_2 > ar2_1:
            deg, coeffs, fitted = 2, c2, f2
        else:
            deg, coeffs, fitted = 1, c1, f1

        residuals = remaining - fitted
        var_contributor = float(np.var(fitted, ddof=1))
        delta_std = float(np.sqrt(max(0.0, var_contributor)))

        # Human-readable label: strip "in_" prefix
        display_lbl = best_lbl.replace("in_", "")
        # Look up the UA_VARIABLES catalogue for a nicer label
        _var_cfg = UA_VARIABLES.get(display_lbl)
        if _var_cfg:
            display_lbl = _var_cfg[0]  # tuple index 0 is the label string

        table_rows.append({
            "Response":    (out_col if _step == 0 else table_rows[-1]["Parameter"]),
            "Variance":    round(float(np.var(remaining, ddof=1)), 2),
            "ΔStd Dev":    round(delta_std, 2),
            "Parameter":   display_lbl,
            "Correlation": round(best_cor, 4),
            "Fit Degree":  deg,
            "Adj R²":      round(max(ar2_1, ar2_2), 4),
            "_key":        best_lbl,
            "_coeffs":     list(coeffs),
            "_xv":         list(param_x),
            "_yv":         list(remaining),
        })

        remaining = residuals
        rem_sv = np.delete(rem_sv, best_idx, axis=1)
        active_labels = [l for i, l in enumerate(active_labels) if i != best_idx]

    sqrt_sum = float(np.sqrt(sum(r["ΔStd Dev"] ** 2 for r in table_rows)))
    ratio_to_actual = sqrt_sum / actual_std if actual_std > 0 else 0.0
    residual_std = float(np.std(remaining, ddof=1))
    return (pd.DataFrame(table_rows), sqrt_sum, ratio_to_actual,
            actual_std, residual_std)



def _display_input_label(input_col: str) -> str:
    """Return a readable label for an input-column name such as in_F_r."""
    raw = str(input_col).replace("in_", "", 1)
    cfg = UA_VARIABLES.get(raw)
    return cfg[0] if cfg else raw


def _terminal_stepwise_row(df_ok: pd.DataFrame, out_col: str, in_cols: list,
                           imp_df: pd.DataFrame, r_min: float) -> dict | None:
    """Build the final diagnostic row for a stepwise-regression table.

    The normal rows show each parameter accepted by the Student-t correlation
    screen.  This terminal row shows the residual data set produced after the
    last accepted parameter and identifies the next highest-correlated
    remaining parameter.  That makes the stop condition visible in the table:
    if the next |r| is below r_min, the stepwise loop correctly terminates.
    """
    try:
        y_raw = pd.to_numeric(df_ok[out_col], errors="coerce").values.copy()
        valid = np.isfinite(y_raw)
        if valid.sum() < 4:
            return None

        # Reconstruct the current residual vector after the last accepted fit.
        if imp_df is not None and len(imp_df) > 0:
            accepted = [str(v) for v in imp_df.get("_key", []) if str(v)]
            last = imp_df.iloc[-1]
            yv = np.asarray(last.get("_yv", []), dtype=float)
            xv = np.asarray(last.get("_xv", []), dtype=float)
            coeffs = np.asarray(last.get("_coeffs", []), dtype=float)
            if yv.size and xv.size and coeffs.size:
                remaining = yv - np.polyval(coeffs, xv)
                response_lbl = f"Residual after {last.get('Parameter', 'last accepted parameter')}"
            else:
                remaining = y_raw[valid]
                response_lbl = str(out_col)
        else:
            accepted = []
            remaining = y_raw[valid]
            response_lbl = str(out_col)

        # Evaluate the next candidate among unaccepted input variables.
        candidates = []
        for ic in in_cols:
            if ic in accepted:
                continue
            xv = pd.to_numeric(df_ok[ic], errors="coerce").values[valid]
            if len(xv) != len(remaining) or not np.isfinite(xv).all():
                continue
            if np.std(xv) <= 0 or np.std(remaining) <= 0:
                corr = 0.0
            elif _SCIPY_OK:
                corr = float(_pearsonr(remaining, xv)[0])
            else:
                corr = float(np.corrcoef(remaining, xv)[0, 1])
            if not np.isfinite(corr):
                corr = 0.0
            candidates.append((ic, corr, xv))

        variance = float(np.var(remaining, ddof=1)) if len(remaining) > 1 else 0.0
        residual_std = float(np.sqrt(max(0.0, variance)))

        if not candidates:
            return {
                "Response": response_lbl,
                "Variance": variance,
                "ΔStd Dev": residual_std,
                "Parameter": "No remaining candidate",
                "Correlation": np.nan,
                "Fit Degree": "—",
                "Adj R²": np.nan,
                "Status": "terminal diagnostic",
                "_terminal": True,
                "_r_min": float(r_min),
            }

        best_ic, best_corr, best_x = max(candidates, key=lambda t: abs(t[1]))

        # Estimate the standard deviation explained by the next candidate even
        # if it does not pass the threshold, so the final line quantifies the
        # residual effect size of the stopping candidate.
        try:
            c1 = np.polyfit(best_x, remaining, 1)
            f1 = np.polyval(c1, best_x)
            delta_std = float(np.sqrt(max(0.0, np.var(f1, ddof=1))))
            fit_deg = 1
            adj_r2 = float(best_corr ** 2)
        except Exception:
            delta_std = 0.0
            fit_deg = "—"
            adj_r2 = np.nan

        return {
            "Response": response_lbl,
            "Variance": variance,
            "ΔStd Dev": delta_std,
            "Parameter": f"Next candidate: {_display_input_label(best_ic)}",
            "Correlation": best_corr,
            "Fit Degree": fit_deg,
            "Adj R²": adj_r2,
            "Status": "terminal diagnostic",
            "_terminal": True,
            "_r_min": float(r_min),
        }
    except Exception:
        return None


def _stepwise_display_table(df_ok: pd.DataFrame, out_col: str, in_cols: list,
                            imp_df: pd.DataFrame, r_min: float) -> pd.DataFrame:
    """Return the Stepwise Regression display table plus the terminal row."""
    cols = ["Response", "Variance", "ΔStd Dev", "Parameter",
            "Correlation", "Fit Degree", "Adj R²"]
    if imp_df is not None and len(imp_df) > 0:
        disp = imp_df[[c for c in cols if c in imp_df.columns]].copy()
    else:
        disp = pd.DataFrame(columns=cols)
    term = _terminal_stepwise_row(df_ok, out_col, in_cols, imp_df, r_min)
    if term:
        disp = pd.concat([disp, pd.DataFrame([{c: term.get(c, np.nan) for c in cols}])],
                         ignore_index=True)
        disp["Row Type"] = ["accepted"] * (len(disp) - 1) + ["terminal"]
    else:
        disp["Row Type"] = "accepted"
    return disp


def _corr_threshold_style(val, r_min: float):
    """Color correlations relative to the Student-t acceptance threshold.

    Red is reserved for values that fail the Student-t screen.  Amber is used
    only for values that pass, but are within 10% above the threshold.  Black
    identifies values comfortably above the threshold.
    """
    try:
        ar = abs(float(val))
        r0 = abs(float(r_min))
    except Exception:
        return "color: #57606a;"
    if not np.isfinite(ar) or not np.isfinite(r0):
        return "color: #57606a;"
    if ar < r0:
        return "color: #cf222e; font-weight: 700;"
    if ar <= 1.10 * r0:
        return "color: #9a6700; font-weight: 700;"
    return "color: #000000; font-weight: 700;"


# ── matplotlib print rcParams (US-Government preflight) ──────────────────────

_MPL_RC = {
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":          12,
    "axes.titlesize":     14,
    "axes.labelsize":     13,
    "xtick.labelsize":    11,
    "ytick.labelsize":    11,
    "legend.fontsize":    11,
    "lines.linewidth":    1.8,
    "lines.markersize":   7,
    "axes.linewidth":     1.0,
    "grid.linewidth":     0.6,
    "grid.color":         "#cccccc",
    "axes.grid":          True,
    "figure.facecolor":   "white",
    "axes.facecolor":     "white",
    "axes.edgecolor":     "black",
    "text.color":         "black",
    "xtick.color":        "black",
    "ytick.color":        "black",
    "axes.labelcolor":    "black",
    "savefig.dpi":        180,
    "savefig.bbox":       None,
    "savefig.facecolor":  "white",
}

_BAR_GREYS  = ["#444444", "#666666", "#888888", "#aaaaaa",
               "#333333", "#555555", "#777777", "#999999"]
_LINE_GREYS = ["black", "#444444", "#777777", "#aaaaaa"]


def _clean_html(s):
    """Strip HTML tags from Plotly labels."""
    return re.sub(r"<[^>]+>", "", str(s or "")).strip()


def _to_list(v):
    if v is None:
        return []
    if hasattr(v, "tolist"):
        v = v.tolist()
    return list(v)


def _render_into_ax(ax, plotly_fig, poly_eq: str = ""):
    """Draw a Plotly figure's data into an existing matplotlib Axes."""
    data   = plotly_fig.data
    layout = plotly_fig.layout
    line_idx = 0

    for trace in data:
        ttype = getattr(trace, "type", "scatter")
        if ttype in ("scatter", "scattergl"):
            xs = _to_list(getattr(trace, "x", None))
            ys = _to_list(getattr(trace, "y", None))
            if not xs:
                continue
            mode   = getattr(trace, "mode", None) or "markers"
            colour = _LINE_GREYS[line_idx % len(_LINE_GREYS)]
            line_idx += 1
            lw = _MPL_RC["lines.linewidth"]
            ms = _MPL_RC["lines.markersize"]
            if "lines" in mode and "markers" in mode:
                ax.plot(xs, ys, "-o", color=colour, linewidth=lw, markersize=ms)
            elif "lines" in mode:
                ax.plot(xs, ys, "-", color=colour, linewidth=lw)
            else:
                mc = getattr(trace.marker, "color", None)
                if isinstance(mc, (list, tuple, np.ndarray)):
                    for xp, yp, c in zip(xs, ys, mc):
                        clr = ("black"
                               if str(c).lower() in ("#ff4444", "black", "red",
                                                     "#cf222e", "#danger")
                               else "#888888")
                        ax.plot(xp, yp, "o", color=clr, markersize=ms,
                                markeredgewidth=0)
                else:
                    ax.plot(xs, ys, "o", color=colour, markersize=ms,
                            markeredgewidth=0)
        elif ttype == "bar":
            orientation = getattr(trace, "orientation", None) or "v"
            xs = _to_list(getattr(trace, "x", None))
            ys = _to_list(getattr(trace, "y", None))
            if not xs:
                continue
            n = len(xs)
            colours = [_BAR_GREYS[i % len(_BAR_GREYS)] for i in range(n)]
            if orientation == "h":
                ax.barh(range(n), xs, color=colours,
                        edgecolor="black", linewidth=0.6)
                ax.set_yticks(range(n))
                ax.set_yticklabels([str(y) for y in ys],
                                   fontsize=_MPL_RC["ytick.labelsize"])
            else:
                try:
                    xs_num = np.asarray(xs, dtype=float)
                    ys_num = np.asarray(ys, dtype=float)
                    numeric_x = len(xs_num) == len(ys_num) and len(xs_num) > 0
                except Exception:
                    numeric_x = False
                if numeric_x:
                    diffs = np.diff(np.sort(xs_num))
                    diffs = diffs[np.isfinite(diffs) & (diffs > 0)]
                    width = float(np.median(diffs) * 0.90) if len(diffs) else 0.8
                    ax.bar(xs_num, ys_num, width=width, color=colours,
                           edgecolor="black", linewidth=0.6)
                else:
                    ax.bar(range(n), ys, color=colours,
                           edgecolor="black", linewidth=0.6)
                    ax.set_xticks(range(n))
                    ax.set_xticklabels([str(x) for x in xs],
                                       rotation=30, ha="right",
                                       fontsize=_MPL_RC["xtick.labelsize"])
        elif ttype == "histogram":
            xs = _to_list(getattr(trace, "x", None))
            if not xs:
                continue
            nbins = int(getattr(trace, "nbinsx", None) or 20)
            ax.hist(xs, bins=nbins, color="#666666",
                    edgecolor="black", linewidth=0.6, alpha=0.85)

    # Reference lines
    for shape in (getattr(layout, "shapes", None) or []):
        if shape is None:
            continue
        if getattr(shape, "type", "") != "line":
            continue
        x0 = getattr(shape, "x0", None)
        x1 = getattr(shape, "x1", None)
        y0 = getattr(shape, "y0", None)
        y1 = getattr(shape, "y1", None)
        sl = getattr(shape, "line", None)
        lw = max(1.0, float(getattr(sl, "width", 1.5) or 1.5))
        dash = str(getattr(sl, "dash", "") or "")
        ls = "--" if "dash" in dash else "-."
        try:
            if (y0 is not None and y1 is not None
                    and abs(float(y0) - float(y1)) < 1e-9):
                ax.axhline(float(y0), color="black", linewidth=lw,
                           linestyle=ls, zorder=3)
            elif (x0 is not None and x1 is not None
                    and abs(float(x0) - float(x1)) < 1e-9):
                ax.axvline(float(x0), color="black", linewidth=lw,
                           linestyle=ls, zorder=3)
        except Exception:
            pass

    # Axis labels
    def _get_axis_title(axis_obj):
        if axis_obj is None:
            return ""
        t = getattr(axis_obj, "title", None)
        if t is None:
            return ""
        return _clean_html(getattr(t, "text", "") or "")

    x_lbl = _get_axis_title(getattr(layout, "xaxis", None))
    y_lbl = _get_axis_title(getattr(layout, "yaxis", None))
    if x_lbl:
        ax.set_xlabel(x_lbl, fontsize=_MPL_RC["axes.labelsize"])
    if y_lbl:
        ax.set_ylabel(y_lbl, fontsize=_MPL_RC["axes.labelsize"])

    ax.tick_params(labelsize=_MPL_RC["xtick.labelsize"])
    ax.set_facecolor("white")
    ax.grid(True, color=_MPL_RC["grid.color"],
            linewidth=_MPL_RC["grid.linewidth"], zorder=0)
    for sp in ax.spines.values():
        sp.set_edgecolor("black")
        sp.set_linewidth(1.0)

    if poly_eq:
        ax.text(0.02, 0.04, poly_eq,
                transform=ax.transAxes,
                fontsize=9, color="black",
                ha="left", va="bottom",
                style="italic",
                bbox=dict(facecolor="white", edgecolor="none",
                          alpha=0.75, pad=2))


def _build_ua_pdf(figures_with_titles, run_info: dict,
                  importance_df, out_path: str):
    """Write all FLARE UA result figures to a portrait-letter PDF.

    Layout: cover page (run parameters + stepwise regression table),
    then one-column / two-rows of plots per page — identical to the
    QD-RELAP BEPU dashboard export.
    """
    PW, PH   = 8.5, 11.0
    MARGIN   = 0.55
    LEFT_M   = 1.65
    RIGHT_M  = 0.35
    USABLE_W = PW - LEFT_M - RIGHT_M
    USABLE_H = PH - 2 * MARGIN
    TITLE_H  = 0.28
    V_GAP    = 0.42
    FIG_W    = USABLE_W
    FIG_H    = (USABLE_H - V_GAP - 2 * TITLE_H) / 2.0

    ts = _datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

    with plt.rc_context(_MPL_RC):
        with _PdfPages(out_path) as pdf:

            # ── Cover page ────────────────────────────────────────────────
            cfig = plt.figure(figsize=(PW, PH), facecolor="white")
            cfig.text(0.5, 0.955, "FLARE Uncertainty Analysis",
                      ha="center", va="top", fontsize=18, fontweight="bold",
                      color="black")
            cfig.text(0.5, 0.918, "Best-Estimate Plus Uncertainty — Plot Export",
                      ha="center", va="top", fontsize=13, color="black")
            cfig.text(0.5, 0.890,
                      "Export layout: portrait page, two plots vertically stacked",
                      ha="center", va="top", fontsize=10, color="#444444")
            cfig.text(0.5, 0.865, f"Generated: {ts}",
                      ha="center", va="top", fontsize=10, color="#444444")

            # Run-parameters table
            rp_rows = [[k, str(v)] for k, v in run_info.items()]
            cfig.text(0.07, 0.815, "Run Parameters",
                      fontsize=11, fontweight="bold", color="black")
            rp_h = min(0.22, 0.028 * (len(rp_rows) + 1))
            rp_ax = cfig.add_axes([0.07, 0.815 - rp_h - 0.01, 0.86, rp_h])
            rp_ax.axis("off")
            rp_tbl = rp_ax.table(
                cellText=rp_rows,
                colLabels=["Parameter", "Value"],
                loc="upper left", cellLoc="left",
            )
            rp_tbl.auto_set_font_size(False)
            rp_tbl.set_fontsize(8.5)
            rp_tbl.auto_set_column_width([0, 1])
            for (r, c), cell in rp_tbl.get_celld().items():
                cell.set_edgecolor("#aaaaaa")
                cell.set_linewidth(0.5)
                if r == 0:
                    cell.set_facecolor("#dddddd")
                    cell.set_text_props(fontweight="bold", color="black")
                else:
                    cell.set_facecolor("white" if r % 2 == 1 else "#f4f4f4")
                    cell.set_text_props(color="black")

            # Stepwise regression table
            if importance_df is not None and len(importance_df) > 0:
                disp_cols = ["Response", "Variance", "ΔStd Dev", "Parameter",
                             "Correlation", "Fit Degree", "Adj R²"]
                disp_cols = [c for c in disp_cols if c in importance_df.columns]
                sr_rows = [
                    [str(round(v, 3)) if isinstance(v, float) else str(v)
                     for v in row]
                    for row in importance_df[disp_cols].itertuples(index=False)
                ]
                cfig.text(0.07, 0.500, "Stepwise Regression Table — Importance Analysis",
                          fontsize=11, fontweight="bold", color="black")
                sr_h = min(0.42, 0.030 * (len(sr_rows) + 1))
                sr_ax = cfig.add_axes([0.07, 0.500 - sr_h - 0.01, 0.86, sr_h])
                sr_ax.axis("off")
                sr_tbl = sr_ax.table(
                    cellText=sr_rows, colLabels=disp_cols,
                    loc="upper left", cellLoc="center",
                )
                sr_tbl.auto_set_font_size(False)
                sr_tbl.set_fontsize(7.8)
                sr_tbl.auto_set_column_width(list(range(len(disp_cols))))
                for (r, c), cell in sr_tbl.get_celld().items():
                    cell.set_edgecolor("#aaaaaa")
                    cell.set_linewidth(0.5)
                    if r == 0:
                        cell.set_facecolor("#dddddd")
                        cell.set_text_props(fontweight="bold", color="black")
                    else:
                        cell.set_facecolor("white" if r % 2 == 1 else "#f4f4f4")
                        cell.set_text_props(color="black")

            pdf.savefig(cfig, bbox_inches=None)
            plt.close(cfig)

            # ── Plot pages: 2-up vertically stacked ──────────────────────
            pairs = list(figures_with_titles)
            if len(pairs) % 2 != 0:
                pairs.append(("", None))

            for i in range(0, len(pairs), 2):
                pfig = plt.figure(figsize=(PW, PH), facecolor="white")
                for row_idx, entry in enumerate(pairs[i:i + 2]):
                    if len(entry) == 3:
                        ftitle, plfig, meta = entry
                    else:
                        ftitle, plfig = entry
                        meta = {}
                    if plfig is None:
                        continue
                    bottom_in = MARGIN + (1 - row_idx) * (FIG_H + TITLE_H + V_GAP)
                    ax_rect = [LEFT_M / PW,
                               bottom_in / PH,
                               FIG_W / PW,
                               FIG_H / PH]
                    new_ax = pfig.add_axes(ax_rect)
                    poly_eq = meta.get("poly_eq", "")
                    try:
                        _render_into_ax(new_ax, plfig, poly_eq=poly_eq)
                    except Exception as _e:
                        new_ax.text(0.5, 0.5, f"Render error:\n{_e}",
                                    ha="center", va="center",
                                    fontsize=10, color="black",
                                    transform=new_ax.transAxes)
                        new_ax.set_facecolor("white")
                    title_y = (bottom_in + FIG_H + 0.05) / PH
                    pfig.text(0.5, title_y, ftitle,
                              ha="center", va="bottom",
                              fontsize=12, fontweight="bold", color="black")
                pdf.savefig(pfig, bbox_inches=None)
                plt.close(pfig)

            d = pdf.infodict()
            d["Title"]   = "FLARE Uncertainty Analysis"
            d["Author"]  = "FLARE UA"
            d["Subject"] = "Best-Estimate Plus Uncertainty"
            d["CreationDate"] = _datetime.datetime.now()


def _make_ua_plot_set(df_results, ts_list, run_dir: Path, base_case: str,
                      selected_ts: str, selected_out: str, use_english=False,
                      tag="current"):
    """Create PNG files and a multipage PDF for the selected UA plot set.

    Now uses the QD-RELAP PDF builder (portrait, 2-up, cover page with
    run-parameter and stepwise-regression tables).
    """
    run_dir = Path(run_dir or WORK_DIR)
    out_dir = run_dir / "ua_plots"
    out_dir.mkdir(parents=True, exist_ok=True)

    df = _restore_input_columns_for_stats(df_results, st.session_state.get("ua_samples"))
    if df is None or len(df) == 0 or "status" not in df.columns:
        raise ValueError("No UA results are available for plotting.")
    df_ok = df[df["status"] == "OK"].copy()
    if len(df_ok) == 0:
        raise ValueError("No successful UA samples are available for plotting.")
    if selected_out not in df_ok.columns:
        raise ValueError(f"Selected scalar output is not available: {selected_out}")

    in_cols = [c for c in df_ok.columns if str(c).startswith("in_")]
    y_raw = pd.to_numeric(df_ok[selected_out], errors="coerce")
    y_vals, y_unit = _ua_convert_scalar_for_plot(y_raw, selected_out, use_english)
    y_label = selected_out.replace("_", " ") + (f" [{y_unit}]" if y_unit else "")

    export_figs = []   # (title, plotly_fig) or (title, plotly_fig, meta)
    pdf_path = out_dir / f"ua_{_safe_plot_token(base_case)}_{_safe_plot_token(tag)}_plots.pdf"

    # 1. Time-series overlay
    if ts_list and selected_ts and selected_ts not in ("(no data yet)", "(no plotted variables available)"):
        fig_ts = go.Figure()
        plotted = 0
        for j, entry in enumerate(ts_list):
            try:
                ts_df = entry["df"]
                if selected_ts not in ts_df.columns or "Time (s)" not in ts_df.columns:
                    continue
                x = pd.to_numeric(ts_df["Time (s)"], errors="coerce")
                yv, unit = _ua_convert_series_for_plot(ts_df[selected_ts], selected_ts, use_english)
                fig_ts.add_trace(go.Scatter(x=x, y=yv, mode="lines",
                                            line=dict(width=1.0), opacity=0.55,
                                            name=f"S{entry.get('sample', j+1)}",
                                            showlegend=(len(ts_list) <= 20)))
                plotted += 1
            except Exception:
                continue
        if plotted:
            _base_lbl = selected_ts
            for sfx in UNIT_CONV:
                if selected_ts.endswith(sfx):
                    _base_lbl = selected_ts[:-len(sfx)].rstrip()
                    break
            _, _unit_ts = _ua_convert_series_for_plot(np.array([0.0]), selected_ts, use_english)
            fig_ts.update_layout(
                title=dict(text=f"Time-Series Overlay — {selected_ts}  (n={plotted})"),
                xaxis_title="Time [s]",
                yaxis_title=f"{_base_lbl} [{_unit_ts}]" if _unit_ts else _base_lbl,
            )
            export_figs.append(("UA Time-Series Overlay", fig_ts))

    # 2. PCT Distribution histogram
    valid_mask = np.isfinite(y_vals)
    if valid_mask.any():
        ys_sorted = np.sort(y_vals[valid_mask])
        counts, edges = np.histogram(ys_sorted, bins=25)
        mids = 0.5 * (edges[:-1] + edges[1:])
        fig_hist = go.Figure()
        fig_hist.add_trace(go.Bar(x=mids, y=counts, name="Count"))
        fig_hist.update_layout(
            title=dict(text=f"Output Distribution — {selected_out}  (n={valid_mask.sum()})"),
            xaxis_title=y_label,
            yaxis_title="Count",
        )
        export_figs.append(("Output Histogram", fig_hist))

    # 3. CDF
    if valid_mask.any():
        cdf = np.arange(1, len(ys_sorted) + 1) / len(ys_sorted)
        fig_cdf = go.Figure()
        fig_cdf.add_trace(go.Scatter(x=ys_sorted, y=cdf, mode="lines",
                                     line=dict(width=2)))
        fig_cdf.update_layout(
            title=dict(text=f"Empirical CDF — {selected_out}"),
            xaxis_title=y_label,
            yaxis_title="Cumulative probability",
        )
        export_figs.append(("Empirical CDF", fig_cdf))

    # 4. Ordered-statistic plot
    if valid_mask.any():
        ord_n = min(20, len(ys_sorted))
        fig_ord = go.Figure()
        fig_ord.add_trace(go.Scatter(
            x=list(range(1, ord_n + 1)),
            y=list(ys_sorted[-ord_n:][::-1]),
            mode="lines+markers",
            line=dict(width=2), marker=dict(size=7),
        ))
        fig_ord.update_layout(
            title=dict(text=f"Top-{ord_n} Ordered Values — {selected_out}"),
            xaxis_title="Order statistic (m)",
            yaxis_title=y_label,
        )
        export_figs.append(("Ordered Statistics", fig_ord))

    # 5. Wilks β vs m curve
    n_ok = int(valid_mask.sum())
    if n_ok >= 2:
        ms_list = list(range(1, min(11, n_ok + 1)))
        betas_95  = [_wilks_confidence_beta(n_ok, m, 0.95) * 100 for m in ms_list]
        fig_beta = go.Figure()
        fig_beta.add_trace(go.Scatter(
            x=ms_list, y=betas_95, mode="lines+markers",
            line=dict(width=2), marker=dict(size=6), name="γ = 95%",
        ))
        fig_beta.add_hline(y=95, line_dash="dot")
        fig_beta.update_layout(
            title=dict(text=f"Wilks Confidence β vs. Order m  (n={n_ok})"),
            xaxis_title="Order statistic (m)",
            yaxis_title="Confidence β (%)",
        )
        export_figs.append(("Wilks Confidence β vs. m", fig_beta))

    # 6. Importance analysis scatter plots + tornado
    imp_df, sqrt_sum, ratio, actual_std, resid_std = _build_importance_table(
        df_ok, selected_out, in_cols, r_min=0.0
    )

    if len(imp_df) > 0:
        # Tornado (ΔStd Dev)
        bar_df = imp_df[["Parameter", "ΔStd Dev"]].sort_values("ΔStd Dev")
        fig_bar = go.Figure(go.Bar(
            y=bar_df["Parameter"], x=bar_df["ΔStd Dev"],
            orientation="h",
        ))
        fig_bar.update_layout(
            title=dict(text=f"Importance — ΔStd Dev per parameter"),
            xaxis_title="ΔStd Dev",
            yaxis_title="",
        )
        export_figs.append(("Importance Ranking — ΔStd Dev", fig_bar))

        # Scatter per parameter
        for _, row in imp_df.iterrows():
            xv  = np.array(row["_xv"])
            yv  = np.array(row["_yv"])
            cx  = np.array(row["_coeffs"])
            xf  = np.linspace(xv.min(), xv.max(), 80)
            yf  = np.polyval(cx, xf)
            deg_ = int(row["Fit Degree"])
            r_p  = float(row["Correlation"])

            # Build polynomial equation label for the PDF
            powers_ = list(range(deg_, -1, -1))
            parts_  = []
            for pw, c in zip(powers_, list(cx)):
                if pw == 0:
                    parts_.append(f"{c:+.4g}")
                elif pw == 1:
                    parts_.append(f"{c:+.4g}·x")
                else:
                    parts_.append(f"{c:+.4g}·x^{pw}")
            poly_eq = (f"y ≈ {'  '.join(parts_)}"
                       f"   [deg {deg_}  adj R²={row['Adj R²']:.3f}]")

            fig_sc = go.Figure()
            fig_sc.add_trace(go.Scatter(x=list(xv), y=list(yv), mode="markers",
                                        marker=dict(size=5, opacity=0.6),
                                        showlegend=False))
            fig_sc.add_trace(go.Scatter(x=list(xf), y=list(yf), mode="lines",
                                        line=dict(width=2),
                                        showlegend=False))
            fig_sc.update_layout(
                title=dict(text=f"{row['Parameter']}  r={r_p:+.3f}"),
                xaxis_title=row["Parameter"],
                yaxis_title=selected_out.replace("_", " "),
            )
            export_figs.append((f"Scatter: {row['Parameter']}", fig_sc,
                                 {"poly_eq": poly_eq}))

    # Build run info for cover page
    run_info = {
        "Case":              base_case,
        "Scalar output":     selected_out,
        "Successful samples": int(valid_mask.sum()),
        "Actual std dev":    f"{actual_std:.4g} {y_unit}".strip() if actual_std else "—",
        "Convolution √Σ Var": f"{sqrt_sum:.4g} {y_unit}".strip() if sqrt_sum else "—",
        "Ratio est/actual":  f"{ratio:.4f}" if ratio else "—",
        "Generated":         _datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    _build_ua_pdf(export_figs, run_info, imp_df if len(imp_df) > 0 else None,
                  str(pdf_path))

    return {"pdf": pdf_path, "pngs": [], "out_dir": out_dir}


# ── Helpers ───────────────────────────────────────────────────────────────────

_EXCLUDED_INPUT_DIR_PREFIXES = ("sim_", "risk_", "ua_", ".sim_all_", "__pycache__")

def _is_generated_input_dir(path: Path) -> bool:
    try:
        rel_parts = path.relative_to(WORK_DIR).parts
    except Exception:
        rel_parts = path.parts
    return any(part.startswith(_EXCLUDED_INPUT_DIR_PREFIXES) or part.startswith(".")
               for part in rel_parts)

def discover_input_cases():
    entries = []
    for p in WORK_DIR.rglob("*_in.xlsx"):
        if p.parent == WORK_DIR:
            continue
        if p.stem.startswith(".~") or p.stem.startswith("ua_"):
            continue
        if _is_generated_input_dir(p.parent):
            continue
        case = p.stem[:-3] if p.stem.endswith("_in") else p.stem.replace("_in", "")
        rel = p.parent.relative_to(WORK_DIR).as_posix()
        entries.append({"case": case, "path": p, "rel": rel, "label": f"{case}  —  {rel}"})
    entries.sort(key=lambda e: (e["case"].lower(), e["rel"].lower()))
    return entries

def find_cases():
    return [e["label"] for e in discover_input_cases()]


def sample_values(dist, p1, p2, base, n, rng):
    """Sample n values from the specified distribution.

    Invalid parameters produce a descriptive ValueError so the UI can surface
    them before a run is launched.  This is the UI-side copy; the worker has its
    own defensive version that degrades gracefully instead of crashing.
    """
    if dist == "uniform":
        if p1 >= p2:
            raise ValueError(
                f"Uniform distribution requires Lower < Upper, got {p1} >= {p2}")
        return rng.uniform(p1, p2, n)
    elif dist == "normal":
        if p2 <= 0:
            raise ValueError(
                f"Normal distribution requires Std deviation > 0, got {p2}")
        raw = rng.normal(p1, p2, n)
        return np.clip(raw, p1 - 4*p2, p1 + 4*p2)
    elif dist == "lognormal":
        if p2 <= 0:
            raise ValueError(
                f"Lognormal distribution requires ln(std) > 0, got {p2}")
        return np.exp(rng.normal(p1, p2, n))
    elif dist == "triangular":
        if p1 >= p2:
            raise ValueError(
                f"Triangular distribution requires Lower < Upper, got {p1} >= {p2}")
        if not (p1 <= base <= p2):
            raise ValueError(
                f"Triangular mode (base={base:.4g}) must be between Lower={p1} and Upper={p2}")
        return rng.triangular(p1, base, p2, n)
    return np.full(n, base)


def build_ua_input(base_case, overrides: dict, sample_id: int, run_dir: Path = None, input_path=None):
    """
    Copy base _in.xlsx to a temp UA input file.
    Appends override assignments just before the Time (s) data table header,
    so last-assignment-wins semantics apply without touching existing rows.
    Returns temp case name.
    """
    src      = Path(input_path) if input_path is not None else WORK_DIR / f"{base_case}_in.xlsx"
    tmp_name = f"ua_{base_case}_{sample_id}"
    # Always write input to WORK_DIR so simulation finds it;
    # outputs are moved to run_dir after the run completes.
    dst      = WORK_DIR / f"{tmp_name}_in.xlsx"

    wb = load_workbook(src)
    ws = wb[f"{base_case}_in"]
    if f"{base_case}_out" in wb.sheetnames:
        wb[f"{base_case}_out"].title = f"{tmp_name}_out"
    ws.title = f"{tmp_name}_in"

    # Find the Time (s) header row  -  insert overrides just before it.
    time_row = None
    for row in ws.iter_rows(max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.strip().startswith("Time"):
            time_row = row[0].row
            break

    if time_row is None:
        time_row = ws.max_row  # fallback

    n_ins = len(overrides) + 1  # +1 comment
    ws.insert_rows(time_row, amount=n_ins)

    ws.cell(row=time_row, column=1).value = "# UA overrides (last-assignment-wins)"
    for i, (var, val) in enumerate(overrides.items(), start=1):
        ws.cell(row=time_row + i, column=1).value = f"{var} = {val:.8g}"

    wb.save(dst)
    wb.close()
    return tmp_name


def extract_scalars(base_case, sample_id, run_dir: Path = None):
    """Extract key scalar results from sample output CSV."""
    _dir     = run_dir if run_dir is not None else WORK_DIR
    csv_path = _dir / f"ua_{base_case}_{sample_id}_out.csv"
    if not csv_path.exists():
        return {}
    try:
        df = pd.read_csv(csv_path)
        r  = {}
        if "RCS Pressure (kPa)" in df.columns:
            r["P_min_kPa"] = df["RCS Pressure (kPa)"].min()
        if "DNBR" in df.columns:
            dn = df["DNBR"].replace(0, np.nan)
            r["DNBR_min"]  = dn.min()
        if "Clad Surface Temp (K)" in df.columns:
            tw = df["Clad Surface Temp (K)"]
            r["avg_clad_peak_K"]  = tw.max()
            r["avg_clad_final_K"] = tw.dropna().iloc[-1]
        if "Hot Pin Clad Temp (K)" in df.columns:
            thc = df["Hot Pin Clad Temp (K)"]
            r["hot_pin_clad_peak_K"]  = thc.max()
            r["hot_pin_clad_final_K"] = thc.dropna().iloc[-1]
        if "Rod Failures DNB (est.)" in df.columns:
            r["N_fail_DNB"]  = int(df["Rod Failures DNB (est.)"].max())
        if "Rod Failures Gap (est.)" in df.columns:
            r["N_fail_gap"]  = int(df["Rod Failures Gap (est.)"].max())
        if "Rod Failures EarlyIV (est.)" in df.columns:
            r["N_fail_eiv"]  = int(df["Rod Failures EarlyIV (est.)"].max())
        return r
    except Exception as e:
        return {"error": str(e)}


def cleanup_sample(base_case, sample_id):
    for suffix in ("_in.xlsx", "_out.xlsx", "_out.csv", "_fail.csv"):
        p = WORK_DIR / f"ua_{base_case}_{sample_id}{suffix}"
        try:
            p.unlink()
        except Exception:
            pass


# ── Worker-process helpers ───────────────────────────────────────────────────
_UA_STATUS_FILE = "ua_status.json"
_UA_CONFIG_FILE = "ua_worker_config.json"
_UA_ABORT_FILE  = "ua_abort_requested.json"

def _json_read(path: Path, default=None):
    try:
        if path and path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default

def _tail_text(path: Path, max_chars=12000):
    try:
        if path and path.exists():
            txt = path.read_text(encoding="utf-8", errors="replace")
            return txt[-max_chars:]
    except Exception as e:
        return f"(Could not read log: {e})"
    return ""

def _restore_input_columns_for_stats(results_df, samples_df=None):
    """Ensure UA result tables contain in_<var> columns for statistics plots.

    The worker writes clean CSV headers for user-facing output, but the
    Results tab historically expects sampled inputs to be named in_<var>.
    Reconstruct those columns from the samples table when needed, so scatter
    plots and Pearson/tornado rankings remain available for old and new runs.
    """
    if results_df is None:
        return results_df
    df = results_df.copy()

    # Already in the historical/internal format.
    if any(str(c).startswith("in_") for c in df.columns):
        return df

    if samples_df is None or len(samples_df) == 0 or "sample" not in samples_df.columns:
        return df

    try:
        samp = samples_df.copy()
        input_cols = [c for c in samp.columns if c != "sample"]
        if not input_cols or "sample" not in df.columns:
            return df

        add_cols = samp[["sample"] + input_cols].copy()
        add_cols = add_cols.rename(columns={c: f"in_{c}" for c in input_cols})

        # Drop any accidental duplicate unprefixed input columns from the result
        # frame before merge; otherwise they appear as output choices.
        drop_unprefixed = [c for c in input_cols if c in df.columns]
        if drop_unprefixed:
            df = df.drop(columns=drop_unprefixed)

        df = df.merge(add_cols, on="sample", how="left")
    except Exception:
        return results_df
    return df


def _format_elapsed_from_status(status: dict) -> str:
    """Return total elapsed run time from ua_status.json as H:MM:SS.

    The worker keeps its PID in the status file for abort handling, but the UI
    displays elapsed time instead because that is more useful to the user.
    """
    try:
        started = status.get("started")
        if started:
            # Worker timestamps are local ISO strings from datetime.now().isoformat().
            t0 = datetime.fromisoformat(str(started))
            secs = max(0, int((datetime.now() - t0).total_seconds()))
        else:
            # Fallback: current-sample elapsed if a legacy worker/status file lacks
            # the total-run start timestamp.
            secs = max(0, int(float(status.get("elapsed_current_s", 0) or 0)))
        h, rem = divmod(secs, 3600)
        m, sec = divmod(rem, 60)
        return f"{h:d}:{m:02d}:{sec:02d}" if h else f"{m:d}:{sec:02d}"
    except Exception:
        return "—"

def _load_ua_run_from_dir(run_dir: Path, base_case: str):
    """Load worker-produced UA results, samples, and time-series into session state."""
    res_csv  = run_dir / f"ua_{base_case}_results.csv"
    samp_csv = run_dir / f"ua_{base_case}_samples.csv"
    _res_df = pd.read_csv(res_csv) if res_csv.exists() else None
    _samp_df = pd.read_csv(samp_csv) if samp_csv.exists() else None
    if _samp_df is not None:
        st.session_state.ua_samples = _samp_df
    if _res_df is not None:
        st.session_state.ua_results = _restore_input_columns_for_stats(_res_df, _samp_df)

    # FLARE RCS time-series: ua_{base_case}_{sample}_out.csv
    _ts_list = []
    import re as _re
    for _ts_f in sorted(run_dir.glob(f"ua_{base_case}_*_out.csv")):
        _m = _re.search(r"_(\d+)_out\.csv$", _ts_f.name)
        if _m:
            try:
                _ts_list.append({"sample": int(_m.group(1)), "df": pd.read_csv(_ts_f)})
            except Exception:
                pass
    st.session_state.ua_ts = _ts_list

    # FLARECON containment time-series: ua_{base_case}_{sample}-CON.csv
    _con_ts_list = []
    for _con_f in sorted(run_dir.glob(f"ua_{base_case}_*-CON.csv")):
        _m = _re.search(r"_(\d+)-CON\.csv$", _con_f.name)
        if _m:
            try:
                _con_ts_list.append({"sample": int(_m.group(1)), "df": pd.read_csv(_con_f)})
            except Exception:
                pass
    st.session_state.ua_con_ts = _con_ts_list

    st.session_state.ua_run_dir = run_dir
    st.session_state.ua_case = base_case

def _find_recent_active_ua_run():
    """Return most recent UA run folder whose status is running/starting."""
    runs = sorted(
        [d for d in WORK_DIR.iterdir()
         if d.is_dir() and d.name.startswith("ua_") and (d / _UA_STATUS_FILE).exists()],
        key=lambda d: d.stat().st_mtime,
        reverse=True,
    )
    for d in runs:
        s = _json_read(d / _UA_STATUS_FILE, {})
        if s.get("status") in ("starting", "running"):
            return d
    return None

def _start_ua_worker(base_case, active_vars, n_samples, fast_mode, input_path=None):
    """Create a UA run folder, write config, and launch detached worker."""
    _run_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = WORK_DIR / f"ua_{base_case}_{_run_tag}"
    run_dir.mkdir(exist_ok=True)

    cfg = {
        "work_dir": str(WORK_DIR),
        "run_dir": str(run_dir),
        "base_case": base_case,
        "base_input_path": str(input_path) if input_path is not None else str(WORK_DIR / f"{base_case}_in.xlsx"),
        "active_vars": active_vars,
        "n_samples": int(n_samples),
        "fast_mode": bool(fast_mode),
        "timeout_s": 600,
    }
    cfg_path = run_dir / _UA_CONFIG_FILE
    cfg_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")

    stdout_path = run_dir / "ua_worker_stdout.log"
    stderr_path = run_dir / "ua_worker_stderr.log"

    worker = WORK_DIR / "flare_ua_worker.py"
    if not worker.exists():
        raise FileNotFoundError(
            f"{worker.name} was not found in the FLARE folder. "
            "Place flare_ua_worker.py beside flare_ua.py."
        )

    # Use unbuffered Python so status/log output is available promptly.
    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env["MPLBACKEND"] = "Agg"
    flags = 0
    try:
        flags = subprocess.CREATE_NEW_PROCESS_GROUP  # Windows
    except AttributeError:
        flags = 0

    with open(stdout_path, "w", encoding="utf-8", errors="replace") as out, \
         open(stderr_path, "w", encoding="utf-8", errors="replace") as err:
        proc = subprocess.Popen(
            [sys.executable, "-u", str(worker), str(cfg_path)],
            cwd=str(WORK_DIR),
            stdout=out,
            stderr=err,
            env=env,
            creationflags=flags,
        )

    st.session_state.ua_run_dir = run_dir
    st.session_state.ua_case = base_case
    st.session_state.ua_status = "running"
    st.session_state.ua_results = None
    st.session_state.ua_samples = None
    st.session_state.ua_ts = []
    st.session_state.ua_con_ts = []
    return proc.pid, run_dir

def _request_ua_abort(run_dir: Path):
    """Ask worker to abort and, if available, kill the worker/child process tree.

    After signalling the processes, this function writes "aborted" directly to
    the status file.  This is necessary because the worker process is being
    killed and may never get a chance to write the final aborted status itself,
    leaving the status file stuck at "running" and the UI polling indefinitely.
    """
    abort_path = run_dir / _UA_ABORT_FILE
    abort_path.write_text(json.dumps({
        "abort_requested": True,
        "requested_at": datetime.now().isoformat(timespec="seconds"),
        "reason": "User clicked Abort UA Run in the FLARE UI.",
    }, indent=2), encoding="utf-8")

    status = _json_read(run_dir / _UA_STATUS_FILE, {})
    pids = [status.get("current_pid"), status.get("worker_pid")]
    all_killed = True
    for pid in [p for p in pids if p]:
        try:
            if os.name == "nt":
                result = subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"],
                                        capture_output=True, text=True, timeout=10)
                if result.returncode != 0:
                    all_killed = False
            else:
                os.kill(int(pid), 15)
                # Give the process a moment to terminate, then confirm
                _time.sleep(0.5)
                try:
                    os.kill(int(pid), 0)   # signal 0 = existence check
                    all_killed = False     # still alive
                except (ProcessLookupError, PermissionError):
                    pass                   # gone — good
        except Exception:
            pass

    # Write aborted status directly so the UI sees the correct state on the
    # next rerun even if the worker never wrote its own final status.
    done  = int(status.get("completed_samples", 0) or 0)
    total = int(status.get("total_samples", 0) or 0)
    status.update({
        "status":    "aborted",
        "message":   f"UA aborted by user: {done}/{total} samples completed.",
        "last_update": datetime.now().isoformat(timespec="seconds"),
        "current_pid": None,
    })
    try:
        status_path = run_dir / _UA_STATUS_FILE
        tmp = status_path.with_suffix(".tmp")
        tmp.write_text(json.dumps(status, indent=2), encoding="utf-8")
        tmp.replace(status_path)
    except Exception:
        pass


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    # ── Return to FLARE home ──────────────────────────────────────────────────
    st.markdown(
        f'''<a href="?page=home" target="_self" style="text-decoration:none;display:flex;
            align-items:center;gap:0.55rem;padding:0.35rem 0.6rem;
            border:1px solid #e8530a;border-radius:4px;
            font-family:Share Tech Mono,monospace;font-size:1.6rem;
            font-weight:700;letter-spacing:0.08em;color:#f97316;"
            onmouseover="this.style.background='rgba(232,83,10,0.18)';this.style.boxShadow='0 0 14px rgba(232,83,10,0.45)';this.style.color='#ffffff';"
            onmouseout="this.style.background='transparent';this.style.boxShadow='none';this.style.color='#f97316';">
          {"<img src='" + _BUDDY_B64 + "' style='height:4.5rem;width:auto;object-fit:contain;flex-shrink:0;'/>" if _BUDDY_B64 else "🔥"}
          FLARE Home
        </a>''',
        unsafe_allow_html=True,
    )
    st.divider()
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:1.4rem;'
                'color:#e6edf3;font-weight:600;margin-bottom:2px">⚛ FLARE UA</div>',
                unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.8rem;color:#8b949e;'
                'margin-bottom:1rem">Uncertainty Analysis</div>',
                unsafe_allow_html=True)

    # Number of samples  -  first and most prominent
    n_samples = st.number_input("Number of samples", min_value=1,
                                max_value=5000, value=50, step=10,
                                help=(
                                    "Number of Monte Carlo simulation runs. "
                                    "Under Wilks' order-statistic method, the required sample size "
                                    "depends only on the desired confidence and probability level — "
                                    "not on the number of input variables. "
                                    "59 samples are sufficient for a one-sided 95th-percentile / "
                                    "95% confidence (95/95) statement; 93 for two-sided 95/95. "
                                    "For a 95th-percentile / 50% confidence (95/50) statement, "
                                    "14 samples suffice. "
                                    "Runtime scales linearly — allow ~1 min per sample."
                                ))

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

    # Case selector
    _case_entries = discover_input_cases()
    if not _case_entries:
        st.error("No *_in.xlsx files found in FLARE subfolders. Root-level input files are intentionally ignored.")
        st.stop()
    _case_labels = [e["label"] for e in _case_entries]
    _selected_label = st.selectbox("Base case", _case_labels,
                             help=(
                                 "The nominal input case to perturb. Input decks are discovered recursively in "
                                 "subfolders below the FLARE root; root-level inputs and generated output folders are ignored. "
                                 "All sampled parameters are varied around their base-case values; unselected parameters "
                                 "keep their base-case value exactly."
                             ))
    _selected_entry = next(e for e in _case_entries if e["label"] == _selected_label)
    selected = _selected_entry["case"]
    selected_input_path = _selected_entry["path"]
    st.caption(f"Input file: `{selected_input_path.relative_to(WORK_DIR)}`")

    # Reset run state when the user picks a different base case so the Run tab
    # shows IDLE rather than the previous run's status/folder/results.
    # Exception: if the user just loaded a previous run whose case happens to
    # differ from the dropdown selection (e.g. they loaded a CaseLBLOCA run
    # while the dropdown was on CaseATWS), do not reset on this render cycle —
    # the loaded results need to survive into the Results tab.  We detect this
    # by checking whether the loaded run folder matches the currently-selected
    # run in the "Load previous run" dropdown.
    _prev_case     = st.session_state.get("ua_case")
    _loaded_run    = st.session_state.get("_ua_loaded_run")
    _sel_run_check = st.session_state.get("ua_prev_run_sel", "— current session —")
    _just_loaded   = (_loaded_run is not None and _loaded_run == _sel_run_check)
    if _prev_case is not None and _prev_case != selected and not _just_loaded:
        _cur_status = st.session_state.get("ua_status", "idle")
        if _cur_status not in ("starting", "running"):
            st.session_state.ua_status  = "idle"
            st.session_state.ua_run_dir = None
            st.session_state.ua_results = None
            st.session_state.ua_case    = selected
            st.session_state.ua_ts      = []
            st.session_state.ua_con_ts  = []
            for _k in ("_ua_loaded_worker_run", "_ua_loaded_run"):
                if _k in st.session_state:
                    del st.session_state[_k]

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

    # Detect whether the selected input workbook has a FLARECON worksheet.
    # FLARECON UA variables are only shown when the sheet exists — there is no
    # point sampling containment parameters for a case that has no containment
    # model configured.
    def _has_flarecon_sheet(path: Path) -> bool:
        try:
            import shutil as _shutil, tempfile as _tempfile
            tmp = Path(_tempfile.mktemp(suffix=".xlsx"))
            _shutil.copy2(path, tmp)
            try:
                from openpyxl import load_workbook as _lwb
                wb = _lwb(str(tmp), read_only=True, data_only=True)
                result = any(n.strip().casefold() == "flarecon" for n in wb.sheetnames)
                wb.close()
                return result
            finally:
                try: tmp.unlink(missing_ok=True)
                except Exception: pass
        except Exception:
            return False

    _case_has_flarecon = _has_flarecon_sheet(selected_input_path)

    # Per-variable distribution editor
    active_vars = {}
    st.caption(
        "Enable each variable you want to treat as uncertain and set its "
        "probability distribution. Variables left unchecked are held at "
        "their base-case value for every sample. "
        "Hover over any variable name for its physical description and typical range."
    )
    with st.expander("Edit variable distributions", expanded=True):
        for var, _ucfg in UA_VARIABLES.items():
            label, def_dist, base, def_p1, def_p2, var_help = _ucfg[:6]
            var_sheet = _ucfg[6] if len(_ucfg) > 6 else ""
            # Skip FLARECON variables when the selected case has no FLARECON sheet
            if var_sheet.strip().casefold() == "flarecon" and not _case_has_flarecon:
                continue
            enabled = st.checkbox(label, value=False, key=f"en_{var}",
                                  help=var_help)
            if enabled:
                c1, c2, c3 = st.columns([2, 1.5, 1.5])
                with c1:
                    dist = st.selectbox("Distribution", DIST_OPTIONS,
                                        index=DIST_OPTIONS.index(def_dist),
                                        key=f"dist_{var}", label_visibility="collapsed",
                                        help=(
                                            "uniform — equal probability between lower and upper bound.\n"
                                            "normal — Gaussian; sampled values clipped to ±4σ.\n"
                                            "lognormal — log-space normal; use for strictly positive "
                                            "quantities with multiplicative uncertainty.\n"
                                            "triangular — peaked at the base-case value with "
                                            "user-specified lower and upper bounds."
                                        ))
                lbl1, lbl2 = DIST_PARAM_LABELS[dist]
                with c2:
                    p1 = st.number_input(lbl1, value=float(def_p1),
                                         format="%.6g",
                                         key=f"p1_{var}", label_visibility="collapsed",
                                         help=lbl1)
                with c3:
                    p2 = st.number_input(lbl2, value=float(def_p2),
                                         format="%.6g",
                                         key=f"p2_{var}", label_visibility="collapsed",
                                         help=lbl2)
                active_vars[var] = {"dist": dist, "base": base,
                                    "p1": p1, "p2": p2, "label": label,
                                    "sheet": var_sheet}

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

    # ── Previous-run selector ─────────────────────────────────────────────
    import re as _re2
    _prev_runs = sorted(
        [d for d in WORK_DIR.iterdir()
         if d.is_dir()
         and d.name.startswith("ua_")
         and any(d.glob("*_results.csv"))],
        key=lambda d: d.stat().st_mtime,
        reverse=True,
    )
    if _prev_runs:
        _run_labels = ["— current session —"] + [d.name for d in _prev_runs]
        _sel_run = st.selectbox(
            "Load previous run",
            _run_labels,
            key="ua_prev_run_sel",
            help="Select a completed UA run folder to review its results.",
        )
        if _sel_run != "— current session —":
            _sel_dir = WORK_DIR / _sel_run
            _case_from_dir = _re2.sub(r"_\d{8}_\d{6}$", "", _sel_run[3:])
            _res_csv  = _sel_dir / f"ua_{_case_from_dir}_results.csv"
            _samp_csv = _sel_dir / f"ua_{_case_from_dir}_samples.csv"
            if st.session_state.get("_ua_loaded_run") != _sel_run:
                if st.button("📥  Load", key="ua_load_prev"):
                    try:
                        _res_df = pd.read_csv(_res_csv)
                        _samp_df = pd.read_csv(_samp_csv) if _samp_csv.exists() else None
                        st.session_state.ua_results  = _restore_input_columns_for_stats(_res_df, _samp_df)
                        st.session_state.ua_run_dir  = _sel_dir
                        st.session_state.ua_case     = _case_from_dir
                        st.session_state.ua_status   = "done"   # so Run tab shows terminal state
                        if _samp_df is not None:
                            st.session_state.ua_samples = _samp_df
                        _ts_list = []
                        for _ts_f in sorted(_sel_dir.glob(
                                f"ua_{_case_from_dir}_*_out.csv")):
                            _m = _re2.search(r"_(\d+)_out\.csv$", _ts_f.name)
                            if _m:
                                try:
                                    _ts_list.append({"sample": int(_m.group(1)),
                                                     "df": pd.read_csv(_ts_f)})
                                except Exception:
                                    pass
                        st.session_state.ua_ts = _ts_list
                        _con_ts_list2 = []
                        for _con_f in sorted(_sel_dir.glob(
                                f"ua_{_case_from_dir}_*-CON.csv")):
                            _mc = _re2.search(r"_(\d+)-CON\.csv$", _con_f.name)
                            if _mc:
                                try:
                                    _con_ts_list2.append({"sample": int(_mc.group(1)),
                                                          "df": pd.read_csv(_con_f)})
                                except Exception:
                                    pass
                        st.session_state.ua_con_ts = _con_ts_list2
                        st.session_state._ua_loaded_run = _sel_run
                        st.rerun()
                    except Exception as _le:
                        st.error(f"Could not load run: {_le}")

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
    _fast_mode = st.checkbox(
        "Fast mode (no figures)",
        value=True,
        key="ua_fast_mode",
        help="Skip figure generation for each sample run. Only CSV and XLSX are written. Recommended for UA batch runs.",
    )

    # ── Pre-flight validation: check all active variable distributions ────
    def _validate_active_vars(av: dict) -> list:
        """Return list of (var_label, error_message) for any invalid distributions."""
        issues = []
        for _var, _cfg in av.items():
            _d  = _cfg.get("dist", "uniform")
            _p1 = float(_cfg.get("p1", 0))
            _p2 = float(_cfg.get("p2", 0))
            _b  = float(_cfg.get("base", 0))
            _lbl = _cfg.get("label", _var)
            if _d in ("uniform", "triangular") and _p1 >= _p2:
                issues.append((_lbl,
                    f"{_d}: Lower bound ({_p1:.6g}) must be < Upper bound ({_p2:.6g})"))
            elif _d in ("normal", "lognormal") and _p2 <= 0:
                issues.append((_lbl,
                    f"{_d}: Std deviation ({_p2:.6g}) must be > 0"))
            elif _d == "triangular" and _p1 < _p2 and not (_p1 <= _b <= _p2):
                issues.append((_lbl,
                    f"triangular: base/mode ({_b:.6g}) must be between "
                    f"Lower ({_p1:.6g}) and Upper ({_p2:.6g})"))
        return issues

    _preflight_issues = _validate_active_vars(active_vars) if active_vars else []

    run_btn = st.button("▶  Run UA", type="primary",
                        width="stretch",
                        disabled=(len(active_vars) == 0 or len(_preflight_issues) > 0))
    if len(active_vars) == 0:
        st.caption("Enable at least one variable above.")
    elif _preflight_issues:
        for _lbl, _msg in _preflight_issues:
            st.error(f"**{_lbl}**: {_msg}", icon="⚠️")
        st.caption("Fix the distribution parameters above before running.")

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
    st.caption("flare_sim.py")
    st.caption(f"Working dir: `{WORK_DIR}`")


# ── Main panel ────────────────────────────────────────────────────────────────

_display_case = st.session_state.get("ua_case") or selected
st.markdown(f"## Uncertainty Analysis &mdash; {_display_case}")

tab_run, tab_results, tab_samples = st.tabs(
    ["▶  Run", "📊  Results", "📋  Samples"]
)

# ── Session state ─────────────────────────────────────────────────────────────
if "ua_log"     not in st.session_state: st.session_state.ua_log     = ""
if "ua_status"  not in st.session_state: st.session_state.ua_status  = "idle"
if "ua_results" not in st.session_state: st.session_state.ua_results = None
if "ua_samples" not in st.session_state: st.session_state.ua_samples = None
if "ua_case"    not in st.session_state: st.session_state.ua_case    = None
if "ua_ts"      not in st.session_state: st.session_state.ua_ts      = []   # time-series per sample
if "ua_con_ts"  not in st.session_state: st.session_state.ua_con_ts  = []   # containment time-series per sample
if "ua_run_dir" not in st.session_state: st.session_state.ua_run_dir = None # output subfolder


# ── Run tab ───────────────────────────────────────────────────────────────────
with tab_run:
    # Recover an active worker run after browser refresh/reconnect.
    if st.session_state.ua_status in ("idle", "running") and st.session_state.ua_run_dir is None:
        _active = _find_recent_active_ua_run()
        if _active is not None:
            _st = _json_read(_active / _UA_STATUS_FILE, {})
            st.session_state.ua_run_dir = _active
            st.session_state.ua_case = _st.get("base_case") or selected
            st.session_state.ua_status = "running"

    _run_dir = st.session_state.get("ua_run_dir")
    _status_json = _json_read(_run_dir / _UA_STATUS_FILE, {}) if _run_dir else {}
    status = _status_json.get("status", st.session_state.ua_status)

    # Keep session status synchronized with worker status.
    if status in ("starting", "running"):
        st.session_state.ua_status = "running"
    elif status in ("complete", "done"):
        st.session_state.ua_status = "done"
    elif status in ("failed", "error"):
        st.session_state.ua_status = "error"
    elif status == "aborted":
        st.session_state.ua_status = "aborted"

    badge_map = {
        "idle":     '<span style="background:#9a6700;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">IDLE</span>',
        "starting": '<span style="background:#0969da;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">STARTING</span>',
        "running":  '<span style="background:#0969da;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">RUNNING</span>',
        "complete": '<span style="background:#1a7f37;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">COMPLETE</span>',
        "done":     '<span style="background:#1a7f37;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">COMPLETE</span>',
        "failed":   '<span style="background:#cf222e;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">FAILED</span>',
        "error":    '<span style="background:#cf222e;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">FAILED</span>',
        "aborted":  '<span style="background:#cf222e;color:white;padding:2px 10px;border-radius:4px;font-family:IBM Plex Mono;font-size:0.8rem">ABORTED</span>',
    }
    st.markdown(badge_map.get(status, badge_map["idle"]), unsafe_allow_html=True)

    if run_btn and status not in ("starting", "running") and active_vars:
        try:
            _pid, _rd = _start_ua_worker(selected, active_vars, int(n_samples), _fast_mode)
            st.success(f"Started UA run in `{_rd.name}`.")
            st.rerun()
        except Exception as _e:
            st.session_state.ua_status = "error"
            st.session_state.ua_log = f"Could not start UA worker: {_e}"
            st.error(st.session_state.ua_log)

    # Active worker status / progress display.
    _run_dir = st.session_state.get("ua_run_dir")
    _status_json = _json_read(_run_dir / _UA_STATUS_FILE, {}) if _run_dir else {}
    _status = _status_json.get("status", st.session_state.ua_status)

    if _run_dir:
        st.caption(f"Run folder: `{_run_dir.name}`")

    if _status in ("starting", "running"):
        total = int(_status_json.get("total_samples", n_samples) or n_samples or 1)
        done  = int(_status_json.get("completed_samples", 0) or 0)
        failed = int(_status_json.get("failed_samples", 0) or 0)
        current = _status_json.get("current_sample", "")
        frac = min(1.0, max(0.0, done / max(total, 1)))

        st.progress(frac, text=f"Completed {done}/{total} samples · Failed {failed} · Current sample {current}")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Completed", f"{done}/{total}")
        m2.metric("Failed", f"{failed}")
        m3.metric("Current sample", str(current or "—"))
        m4.metric("Elapsed", _format_elapsed_from_status(_status_json))

        if st.button("⛔ Abort UA Run", type="primary", width="content"):
            _request_ua_abort(_run_dir)
            st.warning("Abort requested — stopping the current simulation process.")
            # Reset UI state so the Run tab returns to the idle configuration
            # table rather than continuing to poll as if a run is in progress.
            st.session_state.ua_status  = "aborted"
            _time.sleep(1.5)
            st.rerun()

        log_path = Path(_status_json.get("current_log", "")) if _status_json.get("current_log") else None
        if log_path:
            with st.expander("Current sample console log", expanded=True):
                st.code(_tail_text(log_path, max_chars=12000), language="text")

        # Poll while browser remains connected. If browser disconnects, worker continues.
        # Safety net: if the status file has not been updated for > 30 seconds while
        # still showing "running", the worker is likely dead (e.g. killed externally
        # or crashed before writing its final status).  Treat it as failed so the
        # UI stops polling and shows an actionable terminal state.
        _last_update_str = _status_json.get("last_update", "")
        if _last_update_str:
            try:
                _last_update_dt = datetime.fromisoformat(str(_last_update_str))
                _stale_s = (datetime.now() - _last_update_dt).total_seconds()
                if _stale_s > 30:
                    # Worker has gone silent — write a failed status so the UI exits
                    # the polling branch on the next rerun.
                    _dead_status = dict(_status_json)
                    _dead_status.update({
                        "status":      "failed",
                        "message":     f"Worker process appears to have stopped (no update for {_stale_s:.0f} s).",
                        "last_update": datetime.now().isoformat(timespec="seconds"),
                        "current_pid": None,
                    })
                    try:
                        _sp = _run_dir / _UA_STATUS_FILE
                        _tp = _sp.with_suffix(".tmp")
                        _tp.write_text(json.dumps(_dead_status, indent=2), encoding="utf-8")
                        _tp.replace(_sp)
                    except Exception:
                        pass
            except (ValueError, TypeError):
                pass
        _time.sleep(2.0)
        st.rerun()

    elif _status in ("complete", "done", "failed", "error", "aborted"):
        base_case = _status_json.get("base_case") or st.session_state.get("ua_case") or selected
        if _run_dir and (st.session_state.ua_results is None or st.session_state.get("_ua_loaded_worker_run") != str(_run_dir)):
            try:
                _load_ua_run_from_dir(_run_dir, base_case)
                st.session_state._ua_loaded_worker_run = str(_run_dir)
            except Exception as _le:
                st.warning(f"Could not load worker output yet: {_le}")

        total = int(_status_json.get("total_samples", 0) or 0)
        done  = int(_status_json.get("completed_samples", 0) or 0)
        failed = int(_status_json.get("failed_samples", 0) or 0)
        msg = _status_json.get("message", "")

        if _status in ("complete", "done"):
            st.success(msg or f"UA complete: {done}/{total} samples completed.")
        elif _status == "aborted":
            st.warning(msg or f"UA aborted: {done}/{total} samples completed before abort.")
        else:
            st.error(msg or f"UA failed: {done}/{total} samples completed; {failed} failed.")

        if _run_dir:
            worker_log = _run_dir / "ua_worker_stdout.log"
            if worker_log.exists():
                with st.expander("Worker log", expanded=False):
                    st.code(_tail_text(worker_log, max_chars=16000), language="text")

        if st.session_state.ua_results is not None:
            st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                _res_path = (_run_dir / f"ua_{base_case}_results.csv"
                             if _run_dir else None)
                if _res_path and _res_path.exists():
                    with open(_res_path, "rb") as _f:
                        st.download_button("⬇  Download results CSV",
                                           data=_f.read(),
                                           file_name=_res_path.name,
                                           mime="text/csv", width="stretch")
            with col2:
                _samp_path = (_run_dir / f"ua_{base_case}_samples.csv"
                              if _run_dir else None)
                if _samp_path and _samp_path.exists():
                    with open(_samp_path, "rb") as _f:
                        st.download_button("⬇  Download samples CSV",
                                           data=_f.read(),
                                           file_name=_samp_path.name,
                                           mime="text/csv", width="stretch")

    elif st.session_state.ua_log:
        st.markdown("**Console output**")
        st.markdown(f'<div class="console">{st.session_state.ua_log}</div>', unsafe_allow_html=True)
    else:
        st.info("Configure uncertain variables in the sidebar and click **Run UA**.")
        st.caption("Long UA runs now execute in a separate worker process. If the browser disconnects, the worker continues and the run can be reloaded from its output folder.")


# ── Results tab ───────────────────────────────────────────────────────────────
with tab_results:
    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

    df = _restore_input_columns_for_stats(
        st.session_state.ua_results,
        st.session_state.get("ua_samples"),
    )
    if df is not None:
        st.session_state.ua_results = df

    if df is None or len(df) == 0:
        st.info("No results yet. Run the uncertainty analysis first.")
    else:
        df_ok = df[df["status"] == "OK"].copy()
        if len(df_ok) == 0:
            st.error("All samples failed.")
        else:
            # ── Unit system selector ───────────────────────────────────────
            unit_sys = st.radio(
                "Units",
                ["Metric  (°C, kPa, kg/s)", "English  (°F, psia, lb/s)"],
                horizontal=True, key="unit_sys",
            )
            _use_english = unit_sys.startswith("English")

            def convert_col(values, col_name):
                for suffix, conv in UNIT_CONV.items():
                    if col_name.endswith(suffix):
                        si_lbl, eng_lbl, si_fn, eng_fn = conv[:4]
                        fn = eng_fn if _use_english else si_fn
                        return fn(values), (eng_lbl if _use_english else si_lbl)
                return values, ""

            def convert_scalar(val, col_name):
                suffix = SCALAR_UNIT.get(col_name)
                if suffix and suffix in UNIT_CONV:
                    si_lbl, eng_lbl, si_fn, eng_fn, eng_scale = UNIT_CONV[suffix]
                    fn  = eng_fn if _use_english else si_fn
                    lbl = eng_lbl if _use_english else si_lbl
                    return fn(val), lbl
                return val, ""

            def convert_std(val, col_name):
                suffix = SCALAR_UNIT.get(col_name)
                if suffix and suffix in UNIT_CONV:
                    si_lbl, eng_lbl, si_fn, eng_fn, eng_scale = UNIT_CONV[suffix]
                    scale = eng_scale if _use_english else 1.0
                    lbl   = eng_lbl if _use_english else si_lbl
                    return val * scale, lbl
                return val, ""

            def fmt_general(v):
                try:
                    v = float(v)
                    if v == int(v) and abs(v) < 1e9:
                        return f"{int(v):,}"
                    elif abs(v) >= 1000:
                        return f"{v:,.2f}"
                    elif abs(v) >= 1:
                        return f"{v:.4f}"
                    else:
                        return f"{v:.6f}"
                except Exception:
                    return str(v)

            def sample_color(i, n):
                t = i / max(n - 1, 1)
                r = int(9   + t * (207 -   9))
                g = int(105 - t *  80)
                b = int(218 - t * 218)
                return f"rgba({r},{g},{b},0.6)"

            # ── Output variable selectors ──────────────────────────────────
            # CON scalar keys produced by the worker from -CON.csv files
            _CON_SCALAR_KEYS = {
                "CON_P_peak_kPa", "CON_T_peak_C",
                "CON_H2_peak_volpct", "CON_sump_level_peak_m",
            }
            # Human-readable labels for CON scalars, honoring the current unit system.
            _CON_DISPLAY_LABELS = {
                "CON_P_peak_kPa":        _ua_con_scalar_display_label("CON_P_peak_kPa", _use_english),
                "CON_T_peak_C":          _ua_con_scalar_display_label("CON_T_peak_C", _use_english),
                "CON_H2_peak_volpct":    _ua_con_scalar_display_label("CON_H2_peak_volpct", _use_english),
                "CON_sump_level_peak_m": _ua_con_scalar_display_label("CON_sump_level_peak_m", _use_english),
            }

            out_cols = [c for c in df_ok.columns
                        if not c.startswith("in_")
                        and c not in ("sample", "status", "error", "CON_error")]
            ts_out_cols     = [c for c in out_cols if c in TS_MAP]
            con_scalar_cols = [c for c in out_cols if c in _CON_SCALAR_KEYS]
            # Scalar selector: TS-mapped outputs first, then CON scalars, then anything else
            scalar_out_cols = (ts_out_cols + con_scalar_cols) if (ts_out_cols or con_scalar_cols) else out_cols

            in_cols = [c for c in df_ok.columns if c.startswith("in_")]

            ts_list_now = st.session_state.ua_ts
            if ts_list_now:
                _available_ts_cols = set()
                for _entry in ts_list_now:
                    try:
                        _available_ts_cols.update(
                            c for c in _entry["df"].columns if c != "Time (s)"
                        )
                    except Exception:
                        pass
                ts_csv_cols = [
                    c for c in UA_PLOTTED_TS_COLUMNS if c in _available_ts_cols
                ]
            else:
                ts_csv_cols = []

            def _col_display_label(col, _use_eng=_use_english):
                for sfx, (si_lbl, eng_lbl, *_) in UNIT_CONV.items():
                    if col.endswith(sfx):
                        base = col[:-len(sfx)].rstrip()
                        unit = eng_lbl if _use_eng else si_lbl
                        return f"{base} [{unit}]" if unit else base
                return col

            def _scalar_display_label(col, _use_eng=_use_english):
                if col in _CON_DISPLAY_LABELS:
                    return _CON_DISPLAY_LABELS[col]
                if col in TS_MAP:
                    base, _, unit = TS_MAP[col]
                    if _use_eng and unit == "°C":
                        unit = "°F"
                    return f"{base} [{unit}]" if unit else base
                return col.replace("_", " ")

            # Order scalar_out_cols so peak outputs precede final/valley outputs,
            # preventing a stale session-state index from landing on e.g.
            # hot_pin_clad_final_K instead of hot_pin_clad_peak_K.
            def _sort_scalar_cols(cols):
                def _priority(c):
                    c_low = c.lower()
                    if "peak" in c_low or "max" in c_low or "min" in c_low:
                        return 0
                    if "final" in c_low or "end" in c_low:
                        return 2
                    return 1
                return sorted(cols, key=_priority)

            scalar_out_cols = _sort_scalar_cols(scalar_out_cols)

            # Preferred default: hot_pin_clad_peak_K if present, else first peak col
            _preferred_sc = next(
                (c for c in ["hot_pin_clad_peak_K", "avg_clad_peak_K",
                             "hot_pin_fuel_peak_K", "P_min_kPa"]
                 if c in scalar_out_cols),
                scalar_out_cols[0] if scalar_out_cols else None,
            )
            # Initialise session state with the preferred column name on first load
            # or when the run changes (so stale index can never select the wrong col).
            _sc_key_name = "sel_sc_name"
            _cur_case = st.session_state.get("ua_case", "")
            _prev_sc_case = st.session_state.get("_sel_sc_case", "")
            if (_sc_key_name not in st.session_state
                    or _prev_sc_case != _cur_case
                    or st.session_state[_sc_key_name] not in scalar_out_cols):
                st.session_state[_sc_key_name] = _preferred_sc
                st.session_state["_sel_sc_case"] = _cur_case

            def _scalar_from_time_series(ts_col, available_scalar_cols):
                """Map the selected time-series variable to its scalar FOM.

                The UI now uses one user-selected variable for the time-series,
                histogram, CDF, scatter, importance analysis, and ordered-statistic
                views.  For quantities that have both peak and final scalar results,
                the peak / limiting value is preferred.
                """
                if not available_scalar_cols:
                    return None
                ts_to_scalar = {
                    "Hot Pin Clad Temp (K)": ["hot_pin_clad_peak_K", "hot_pin_clad_final_K"],
                    "Clad Surface Temp (K)": ["avg_clad_peak_K", "avg_clad_final_K"],
                    "Hot Pin Fuel Temp (K)": ["hot_pin_fuel_peak_K"],
                    "RCS Pressure (kPa)": ["P_max_kPa", "P_min_kPa"],
                    "RCS Temperature (K)": ["T_max_K"],
                    "DNBR": ["DNBR_min"],
                    "RK Total Power (MW)": ["P_peak_MW"],
                }
                # Also support TS_MAP labels that omit the unit suffix.
                for scalar_name, (mapped_ts, _offset, _unit) in TS_MAP.items():
                    if scalar_name not in available_scalar_cols:
                        continue
                    mapped_ts_base = str(mapped_ts).strip().casefold()
                    ts_base = str(ts_col or "").strip().casefold()
                    ts_no_unit = re.sub(r"\s*\([^)]*\)\s*$", "", str(ts_col or "")).strip().casefold()
                    if ts_base == mapped_ts_base or ts_no_unit == mapped_ts_base:
                        ts_to_scalar.setdefault(ts_col, []).append(scalar_name)

                for candidate in ts_to_scalar.get(ts_col, []):
                    if candidate in available_scalar_cols:
                        return candidate
                return (_preferred_sc if _preferred_sc in available_scalar_cols
                        else available_scalar_cols[0])

            sel_col1, sel_col2 = st.columns([2, 1])
            with sel_col1:
                default_ts = next(
                    (c for c in ["Hot Pin Clad Temp (K)", "Clad Surface Temp (K)",
                                  "RCS Pressure (kPa)"] if c in ts_csv_cols),
                    ts_csv_cols[0] if ts_csv_cols else None,
                )
                selected_ts = st.selectbox(
                    "Time-series variable",
                    ts_csv_cols if ts_csv_cols else ["(no plotted variables available)"],
                    index=ts_csv_cols.index(default_ts) if default_ts in ts_csv_cols else 0,
                    key="sel_ts",
                    format_func=_col_display_label,
                    help=("This selection also controls the scalar FOM used for the "
                          "distribution, CDF, scatter, importance, and ordered-statistic plots."),
                ) if ts_csv_cols else None

            selected_out = _scalar_from_time_series(selected_ts, scalar_out_cols)
            st.session_state[_sc_key_name] = selected_out
            st.session_state["_sel_sc_case"] = _cur_case

            with sel_col2:
                st.markdown(
                    '<div style="font-size:0.78rem;color:#57606a;'
                    'text-transform:uppercase;letter-spacing:0.1em;">'
                    'Derived scalar FOM</div>',
                    unsafe_allow_html=True,
                )
                st.caption(_scalar_display_label(selected_out) if selected_out else "—")

            # Separate FLARECON containment FOM selector.  The main scalar selector
            # remains available for core / PWR-system outputs; this selector drives
            # the containment-only distribution, ordered-statistic, and importance
            # analysis panels so only one containment FOM is shown at a time.
            selected_con_out = None
            if con_scalar_cols:
                _preferred_con = next(
                    (c for c in ["CON_P_peak_kPa", "CON_T_peak_C",
                                 "CON_H2_peak_volpct", "CON_sump_level_peak_m"]
                     if c in con_scalar_cols),
                    con_scalar_cols[0],
                )
                _con_fom_key = "sel_con_fom_name"
                _prev_con_case = st.session_state.get("_sel_con_fom_case", "")
                if (_con_fom_key not in st.session_state
                        or _prev_con_case != _cur_case
                        or st.session_state[_con_fom_key] not in con_scalar_cols):
                    st.session_state[_con_fom_key] = _preferred_con
                    st.session_state["_sel_con_fom_case"] = _cur_case

                selected_con_out = st.selectbox(
                    "Containment FOM",
                    con_scalar_cols,
                    index=(con_scalar_cols.index(st.session_state[_con_fom_key])
                           if st.session_state[_con_fom_key] in con_scalar_cols else 0),
                    key=_con_fom_key,
                    format_func=lambda c: _CON_DISPLAY_LABELS.get(c, c.replace("_", " ")),
                    help=("Select the FLARECON containment figure-of-merit to use in "
                          "the containment distribution, ordered-statistic, and "
                          "importance-analysis panels."),
                )

            y = df_ok[selected_out].dropna()
            y_arr, y_unit_disp = _ua_convert_scalar_for_plot(
                pd.to_numeric(df_ok[selected_out], errors="coerce"),
                selected_out, _use_english
            )
            # _scalar_display_label already embeds the unit for TS_MAP columns
            # (e.g. "Hot Pin Clad Temp [°F]").  Only append y_unit_disp when the
            # column is NOT in TS_MAP (e.g. CON scalars, DNBR, rod-failure counts).
            _lbl_base = _scalar_display_label(selected_out)
            if selected_out in TS_MAP or selected_out in _CON_DISPLAY_LABELS:
                y_label = _lbl_base   # unit already embedded
            else:
                y_label = _lbl_base + (f" [{y_unit_disp}]" if y_unit_disp else "")


            def _flare_result_peak_tiles_html():
                """Return HTML tiles summarizing selected FLARE/RCS scalar outcomes.

                These are independent of the selected FOM and provide a compact
                run-level summary of limiting core / RCS responses before the
                FLARECON containment section.
                """
                def _series_max_from_ts(col_name, use_english=False):
                    vals = []
                    for _entry in (ts_list_now or []):
                        try:
                            _df_ts = _entry.get("df") if isinstance(_entry, dict) else None
                            if _df_ts is None or col_name not in _df_ts.columns:
                                continue
                            _v, _u = _ua_convert_series_for_plot(_df_ts[col_name], col_name, use_english)
                            _finite = np.asarray(_v, dtype=float)
                            _finite = _finite[np.isfinite(_finite)]
                            if len(_finite):
                                vals.append(float(np.max(_finite)))
                        except Exception:
                            continue
                    return (max(vals), _u) if vals else (None, "")

                def _results_extreme(col_name, mode="max", use_english=False):
                    if col_name not in df_ok.columns:
                        return (None, "")
                    try:
                        _v, _u = _ua_convert_scalar_for_plot(df_ok[col_name], col_name, use_english)
                        _finite = np.asarray(_v, dtype=float)
                        _finite = _finite[np.isfinite(_finite)]
                        if not len(_finite):
                            return (None, _u)
                        if mode == "min":
                            return (float(np.min(_finite)), _u)
                        return (float(np.max(_finite)), _u)
                    except Exception:
                        return (None, "")

                _dnbr_val, _dnbr_unit = _results_extreme("DNBR_min", mode="min")
                _p_val, _p_unit = _results_extreme("P_max_kPa", mode="max", use_english=_use_english)
                if _p_val is None:
                    _p_val, _p_unit = _series_max_from_ts("RCS Pressure (kPa)", use_english=_use_english)
                _pct_val, _pct_unit = _results_extreme("hot_pin_clad_peak_K", mode="max", use_english=_use_english)
                if _pct_val is None:
                    _pct_val, _pct_unit = _series_max_from_ts("Hot Pin Clad Temp (K)", use_english=_use_english)
                _h2_val, _h2_unit = _series_max_from_ts("H2 Generated (kg)", use_english=False)

                def _fmt_tile_value(val, unit=""):
                    if val is None or not np.isfinite(val):
                        return "—"
                    return f"{fmt_general(val)} {unit}".strip()

                _dnbr_cls = "danger" if (_dnbr_val is not None and _dnbr_val < 1.0) else "ok"
                _pct_cls = "warn" if _pct_val is not None else ""
                return (
                    '<div class="metric-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:0.75rem;">'
                    f'<div class="metric-tile {_dnbr_cls}"><div class="val">{_fmt_tile_value(_dnbr_val, _dnbr_unit)}</div>'
                    '<div class="lbl">Min DNBR / dryout margin</div></div>'
                    f'<div class="metric-tile"><div class="val">{_fmt_tile_value(_p_val, _p_unit)}</div>'
                    '<div class="lbl">Peak RCS Pressure</div></div>'
                    f'<div class="metric-tile {_pct_cls}"><div class="val">{_fmt_tile_value(_pct_val, _pct_unit)}</div>'
                    '<div class="lbl">Peak Hot Pin Clad Temperature</div></div>'
                    f'<div class="metric-tile"><div class="val">{_fmt_tile_value(_h2_val, _h2_unit or "kg")}</div>'
                    '<div class="lbl">Peak H₂ Generation</div></div>'
                    '</div>'
                )

            # ── Student-t threshold for importance analysis ────────────────
            T_TABLE = {
                80: 1.282, 85: 1.440, 90: 1.645, 95: 1.960, 97.5: 2.000, 99: 2.576,
            }

            # ══════════════════════════════════════════════════════════════
            # Sub-tabs: FOM Distribution | Importance Analysis | Ordered Statistic
            # ══════════════════════════════════════════════════════════════
            sub_dist, sub_imp, sub_ord = st.tabs([
                "📊  FOM Distribution",
                "📋  Importance Analysis",
                "📈  Ordered Statistic",
            ])

            # ── SUB-TAB 1: FOM Distribution ───────────────────────────────
            with sub_dist:
                st.markdown("#### 🔥 FLARE Results")
                st.markdown(
                    '<div style="font-size:0.78rem;color:#57606a;text-transform:uppercase;'
                    'letter-spacing:0.1em;margin-bottom:0.5rem;">'
                    'RCS Scalar Outcomes — Peak Values</div>',
                    unsafe_allow_html=True,
                )
                st.markdown(_flare_result_peak_tiles_html(), unsafe_allow_html=True)
                st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

                # Time-series overlay
                if ts_list_now and selected_ts and selected_ts not in (
                        "(no data yet)", "(no plotted variables available)"):
                    n_ts = len(ts_list_now)
                    fig_ts = go.Figure()
                    _ts_peaks   = []
                    _ts_valleys = []
                    for j, entry in enumerate(ts_list_now):
                        _df_ts = entry["df"]
                        if selected_ts in _df_ts.columns:
                            _yvals, _unit = convert_col(
                                _df_ts[selected_ts].values, selected_ts)
                            fig_ts.add_trace(go.Scatter(
                                x=_df_ts["Time (s)"],
                                y=_yvals,
                                mode="lines",
                                line=dict(color=sample_color(j, n_ts), width=1.5),
                                name=f"S{entry['sample']}",
                                showlegend=(n_ts <= 20),
                            ))
                            _finite = _yvals[np.isfinite(_yvals)]
                            if len(_finite):
                                _ts_peaks.append(float(np.max(_finite)))
                                _ts_valleys.append(float(np.min(_finite)))
                    _, _unit_ts = convert_col(np.array([0.0]), selected_ts)
                    _base_lbl = selected_ts
                    for _sfx in UNIT_CONV:
                        if selected_ts.endswith(_sfx):
                            _base_lbl = selected_ts[:-len(_sfx)].rstrip()
                            break
                    fig_ts.update_layout(
                        title=dict(
                            text=f"Time-series — {_col_display_label(selected_ts)}  (n={n_ts})",
                            font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                        plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                        font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                        xaxis=dict(title="Time [s]", gridcolor=PLOT_GRID),
                        yaxis=dict(title=f"{_base_lbl} [{_unit_ts}]" if _unit_ts else _base_lbl,
                                   gridcolor=PLOT_GRID),
                        margin=dict(l=60, r=20, t=45, b=45),
                        height=380,
                        legend=dict(bgcolor="rgba(255,255,255,0.8)", font=dict(size=9)),
                    )
                    st.plotly_chart(fig_ts, use_container_width=True,
                                    config={"displayModeBar": False},
                                    key=_ua_element_key("pc_fig_ts", selected, selected_ts, selected_out))
                    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)

                # Histogram + ordered table
                c_hist, c_tbl = st.columns([3, 2])

                valid_mask = np.isfinite(y_arr)
                y_sorted_full = np.sort(y_arr[valid_mask])

                with c_hist:
                    st.markdown('<div class="section-label" style="font-size:0.8rem;'
                                'color:#57606a;text-transform:uppercase;'
                                'letter-spacing:0.1em;">Output Distribution</div>',
                                unsafe_allow_html=True)
                    if valid_mask.any():
                        counts_h, edges_h = np.histogram(y_sorted_full, bins=25)
                        mids_h = 0.5 * (edges_h[:-1] + edges_h[1:])
                        fig_dist = go.Figure()
                        fig_dist.add_trace(go.Bar(
                            x=mids_h, y=counts_h,
                            marker_color=C[0], marker_line_width=0,
                            opacity=0.8, name="Count",
                        ))
                        fig_dist.update_layout(
                            title=dict(
                                text=f"Distribution — {_scalar_display_label(selected_out)}  (n={valid_mask.sum()})",
                                font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                            font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                            xaxis=dict(title=y_label, gridcolor=PLOT_GRID),
                            yaxis=dict(title="Count", gridcolor=PLOT_GRID),
                            showlegend=False, height=320,
                            margin=dict(l=55, r=20, t=40, b=45),
                        )
                        st.plotly_chart(fig_dist, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=_ua_element_key("pc_fig_dist", selected, selected_out))

                with c_tbl:
                    st.markdown('<div class="section-label" style="font-size:0.8rem;'
                                'color:#57606a;text-transform:uppercase;'
                                'letter-spacing:0.1em;">Top-10 Ordered Values</div>',
                                unsafe_allow_html=True)
                    if valid_mask.any():
                        _top10_raw = y_sorted_full[-10:][::-1]
                        _top10_df  = pd.DataFrame({
                            "Rank (m)": range(1, len(_top10_raw) + 1),
                            y_label:   [fmt_general(v) for v in _top10_raw],
                        }).set_index("Rank (m)")
                        st.dataframe(_top10_df, use_container_width=True, height=260)


                # CDF
                st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
                if valid_mask.any():
                    cdf_vals = np.arange(1, len(y_sorted_full) + 1) / len(y_sorted_full)
                    fig_cdf = go.Figure()
                    fig_cdf.add_trace(go.Scatter(
                        x=y_sorted_full, y=cdf_vals,
                        line=dict(color=C[0], width=2), name="CDF",
                    ))
                    fig_cdf.update_layout(
                        title=dict(
                            text=f"Empirical CDF — {_scalar_display_label(selected_out)}",
                            font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                        plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                        font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                        xaxis=dict(title=y_label, gridcolor=PLOT_GRID),
                        yaxis=dict(title="Cumulative probability",
                                   gridcolor=PLOT_GRID, range=[0, 1]),
                        margin=dict(l=55, r=20, t=40, b=45), height=300,
                    )
                    st.plotly_chart(fig_cdf, use_container_width=True,
                                    config={"displayModeBar": False},
                                    key=_ua_element_key("pc_fig_cdf", selected, selected_out))

                # ── FLARECON scalar summary tiles and selected FOM distribution ──
                if con_scalar_cols:
                    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
                    st.markdown("#### ⚛ FLARECON Results")
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#57606a;text-transform:uppercase;'
                        'letter-spacing:0.1em;margin-bottom:0.5rem;">'
                        'Containment Scalar Outcomes — Peak Values</div>',
                        unsafe_allow_html=True,
                    )
                    _H2_LOWER_FLAMM = 4.0
                    _H2_DETONABLE   = 8.0
                    _con_tile_html = '<div class="metric-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:0.75rem;">'
                    for _ck in con_scalar_cols:
                        _cv_raw = pd.to_numeric(df_ok[_ck], errors="coerce").dropna()
                        if len(_cv_raw) == 0:
                            continue
                        _cv, _cu = _ua_convert_con_scalar_for_plot(_cv_raw, _ck, _use_english)
                        _cv = pd.Series(_cv).dropna()
                        if len(_cv) == 0:
                            continue
                        _c_max  = float(_cv.max())
                        _c_mean = float(_cv.mean())
                        _c_min  = float(_cv.min())
                        _clbl   = _CON_DISPLAY_LABELS.get(_ck, _ck.replace("_", " "))
                        if _ck == "CON_H2_peak_volpct":
                            _cls = "danger" if _c_max >= _H2_DETONABLE else (
                                   "warn"   if _c_max >= _H2_LOWER_FLAMM else "ok")
                        else:
                            _cls = "ok"
                        _con_tile_html += (
                            f'<div class="metric-tile {_cls}">'
                            f'<div class="val">{fmt_general(_c_max)}</div>'
                            f'<div class="lbl">{_clbl}<br>'
                            f'<small>mean {fmt_general(_c_mean)} · min {fmt_general(_c_min)}</small>'
                            f'</div></div>'
                        )
                    _con_tile_html += '</div>'
                    st.markdown(_con_tile_html, unsafe_allow_html=True)
                    if any(_ck == "CON_H2_peak_volpct" for _ck in con_scalar_cols):
                        st.caption(
                            "H₂ colour coding: green < 4 vol% (below LFL) · "
                            "orange ≥ 4 vol% (flammable) · red ≥ 8 vol% (detonable)"
                        )
                    st.caption("See the **Importance Analysis** and **Ordered Statistic** tabs below for full FLARECON statistical synthesis.")
                    if selected_con_out:
                        _con_lbl = _CON_DISPLAY_LABELS.get(selected_con_out, selected_con_out.replace("_", " "))
                        _con_y_raw = pd.to_numeric(df_ok[selected_con_out], errors="coerce").to_numpy(dtype=float)
                        _con_y, _con_unit = _ua_convert_con_scalar_for_plot(_con_y_raw, selected_con_out, _use_english)
                        _con_valid = np.isfinite(_con_y)
                        _con_sorted = np.sort(_con_y[_con_valid])

                        st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
                        st.markdown(f"##### FLARECON FOM Distribution — {_con_lbl}")

                        _con_d1, _con_d2 = st.columns([3, 2])
                        with _con_d1:
                            st.markdown('<div class="section-label" style="font-size:0.8rem;'
                                        'color:#57606a;text-transform:uppercase;'
                                        'letter-spacing:0.1em;">Output Distribution</div>',
                                        unsafe_allow_html=True)
                            if _con_valid.any():
                                _con_counts, _con_edges = np.histogram(_con_sorted, bins=25)
                                _con_mids = 0.5 * (_con_edges[:-1] + _con_edges[1:])
                                fig_con_dist = go.Figure()
                                fig_con_dist.add_trace(go.Bar(
                                    x=_con_mids, y=_con_counts,
                                    marker_color=C[0], marker_line_width=0,
                                    opacity=0.8, name="Count",
                                ))
                                if selected_con_out == "CON_H2_peak_volpct":
                                    fig_con_dist.add_vline(x=4.0, line_dash="dash", line_color=C[3],
                                                           annotation_text="LFL 4%",
                                                           annotation_font_color=C[3])
                                    fig_con_dist.add_vline(x=8.0, line_dash="dash", line_color=C[4],
                                                           annotation_text="Det. 8%",
                                                           annotation_font_color=C[4])
                                fig_con_dist.update_layout(
                                    title=dict(text=f"Distribution — {_con_lbl}  (n={_con_valid.sum()})",
                                               font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                                    plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                    font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                                    xaxis=dict(title=_con_lbl, gridcolor=PLOT_GRID),
                                    yaxis=dict(title="Count", gridcolor=PLOT_GRID),
                                    showlegend=False, height=320,
                                    margin=dict(l=55, r=20, t=40, b=45),
                                )
                                st.plotly_chart(fig_con_dist, use_container_width=True,
                                                config={"displayModeBar": False},
                                                key=_ua_element_key("con_fig_dist", selected, selected_con_out))

                        with _con_d2:
                            st.markdown('<div class="section-label" style="font-size:0.8rem;'
                                        'color:#57606a;text-transform:uppercase;'
                                        'letter-spacing:0.1em;">Top-10 Ordered Values</div>',
                                        unsafe_allow_html=True)
                            if _con_valid.any():
                                _con_top10 = _con_sorted[-10:][::-1]
                                _con_top10_df = pd.DataFrame({
                                    "Rank (m)": range(1, len(_con_top10) + 1),
                                    _con_lbl: [fmt_general(v) for v in _con_top10],
                                }).set_index("Rank (m)")
                                st.dataframe(_con_top10_df, use_container_width=True, height=260)


                        if _con_valid.any():
                            st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
                            _con_cdf = np.arange(1, len(_con_sorted) + 1) / len(_con_sorted)
                            fig_con_cdf = go.Figure()
                            fig_con_cdf.add_trace(go.Scatter(
                                x=_con_sorted, y=_con_cdf,
                                line=dict(color=C[0], width=2), name="CDF",
                            ))
                            if selected_con_out == "CON_H2_peak_volpct":
                                fig_con_cdf.add_vline(x=4.0, line_dash="dash", line_color=C[3],
                                                      annotation_text="LFL 4%",
                                                      annotation_font_color=C[3])
                                fig_con_cdf.add_vline(x=8.0, line_dash="dash", line_color=C[4],
                                                      annotation_text="Det. 8%",
                                                      annotation_font_color=C[4])
                            fig_con_cdf.update_layout(
                                title=dict(text=f"Empirical CDF — {_con_lbl}",
                                           font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                                plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                                xaxis=dict(title=_con_lbl, gridcolor=PLOT_GRID),
                                yaxis=dict(title="Cumulative probability",
                                           gridcolor=PLOT_GRID, range=[0, 1]),
                                margin=dict(l=55, r=20, t=40, b=45), height=300,
                            )
                            st.plotly_chart(fig_con_cdf, use_container_width=True,
                                            config={"displayModeBar": False},
                                            key=_ua_element_key("con_fig_cdf", selected, selected_con_out))

            # ── SUB-TAB 2: Importance Analysis ────────────────────────────
            with sub_imp:
                # Student-t threshold controls
                st.markdown("**Student-t Threshold Controls**")
                st.caption(
                    "Per Martin (Nucl. Technol., 175, 2011) Eq. (8): "
                    "r_min = t / √(n − 2 + t²). "
                    "Default: t₀.₉₇₅ = 2.00 (two-sided 97.5%, one-sided 95%). "
                    "The loop stops when max(|r|) < r_min."
                )
                tc1, tc2, tc3 = st.columns([2, 2, 3])
                with tc1:
                    confidence_pct_imp = st.select_slider(
                        "Confidence level (two-sided)",
                        options=[80, 85, 90, 95, 97.5, 99],
                        value=97.5,
                        format_func=lambda x: f"{x}%",
                        key="imp_conf_pct",
                    )
                with tc2:
                    t_crit_imp = T_TABLE[confidence_pct_imp]
                    n_ok_imp   = len(df_ok)
                    r_min_imp  = t_crit_imp / np.sqrt(n_ok_imp - 2 + t_crit_imp ** 2)
                    st.metric("t-critical value", f"{t_crit_imp:.3f}",
                              delta=f"two-sided {confidence_pct_imp}%, n={n_ok_imp}",
                              delta_color="off")
                with tc3:
                    st.metric("Minimum |r| to pass threshold", f"{r_min_imp:.4f}")

                st.markdown("---")
                st.markdown(f"#### 🔥 FLARE — Importance Analysis — {_scalar_display_label(selected_out)}")

                # Run importance analysis
                imp_df_ui, sqrt_sum_ui, ratio_ui, actual_std_ui, resid_std_ui = (
                    _build_importance_table(df_ok, selected_out, in_cols,
                                            r_min=r_min_imp)
                )
                actual_std_sc = float(
                    np.std(
                        pd.to_numeric(df_ok[selected_out], errors="coerce").dropna(),
                        ddof=1
                    )
                )

                # Key metrics strip — wide custom tiles matching the Ordered Statistic tab
                _ratio_cls = "ok" if ratio_ui >= 0.90 else "warn"
                _std_disp  = f"{fmt_general(actual_std_sc)} {y_unit_disp}".strip()
                _conv_disp = f"{fmt_general(sqrt_sum_ui)} {y_unit_disp}".strip()
                _resid_disp = f"{fmt_general(resid_std_ui)} {y_unit_disp}".strip()
                _resid_pct  = (f"{resid_std_ui / actual_std_sc * 100:.1f}% unexplained"
                               if actual_std_sc > 0 else "")
                st.markdown(
                    f'<div class="metric-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:0.75rem;">'
                    f'<div class="metric-tile">'
                    f'<div class="val">{_std_disp}</div>'
                    f'<div class="lbl">Actual std dev σ<sub>o</sub></div>'
                    f'</div>'
                    f'<div class="metric-tile">'
                    f'<div class="val">{_conv_disp}</div>'
                    f'<div class="lbl">Convolution √(Σ Var) — Eq.(11)</div>'
                    f'</div>'
                    f'<div class="metric-tile {_ratio_cls}">'
                    f'<div class="val">{ratio_ui:.3f}</div>'
                    f'<div class="lbl">Ratio estimate / actual — Eq.(12)<br>'
                    f'<small>{"≥ 0.90 ✓" if ratio_ui >= 0.90 else "< 0.90 — increase n"}</small>'
                    f'</div>'
                    f'</div>'
                    f'<div class="metric-tile">'
                    f'<div class="val">{_resid_disp}</div>'
                    f'<div class="lbl">Residual std dev (unexplained)<br>'
                    f'<small>{_resid_pct}</small>'
                    f'</div>'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                st.markdown("---")

                if len(imp_df_ui) == 0:
                    st.info(
                        f"No parameters exceed the Student-t threshold "
                        f"|r| = {r_min_imp:.4f} at n = {n_ok_imp}. "
                        "Try increasing the sample size or lowering the confidence level."
                    )
                else:
                    c_tbl_imp, c_bar_imp = st.columns([3, 2])

                    with c_tbl_imp:
                        st.markdown(
                            f'<div style="font-size:0.78rem;color:#57606a;'
                            f'text-transform:uppercase;letter-spacing:0.1em;">'
                            f'Stepwise Regression Table — '
                            f'{len(imp_df_ui)} parameter(s) above threshold '
                            f'(|r| ≥ {r_min_imp:.4f})</div>',
                            unsafe_allow_html=True,
                        )
                        disp_df_imp = _stepwise_display_table(
                            df_ok, selected_out, in_cols, imp_df_ui, r_min_imp
                        )

                        styled_imp = (
                            disp_df_imp.style
                            .map(lambda v: _corr_threshold_style(v, r_min_imp), subset=["Correlation"])
                            .format({
                                "Variance": "{:.2f}",
                                "ΔStd Dev": "{:.2f}",
                                "Correlation": "{:+.4f}",
                                "Adj R²": "{:.4f}",
                            }, na_rep="—")
                        )
                        st.dataframe(styled_imp, use_container_width=True, height=410)
                        st.caption(
                            "Terminal row: residual data set after the last accepted parameter; "
                            "Correlation color is black when |r| is >10% above the threshold, "
                            "yellow/amber when it passes but is within 10% above the threshold, "
                            "and red when it fails the threshold."
                        )

                    with c_bar_imp:
                        st.markdown(
                            '<div style="font-size:0.78rem;color:#57606a;'
                            'text-transform:uppercase;letter-spacing:0.1em;">'
                            'Importance Ranking (ΔStd Dev)</div>',
                            unsafe_allow_html=True,
                        )
                        bar_df_imp = (imp_df_ui[["Parameter", "ΔStd Dev"]]
                                      .sort_values("ΔStd Dev"))
                        fig_bar_imp = go.Figure(go.Bar(
                            y=bar_df_imp["Parameter"],
                            x=bar_df_imp["ΔStd Dev"],
                            orientation="h",
                            marker_color=C[0],
                            opacity=0.85,
                        ))
                        fig_bar_imp.update_layout(
                            title=dict(
                                text=f"ΔStd Dev per parameter  |threshold |r| ≥ {r_min_imp:.4f}",
                                font=dict(family="IBM Plex Mono", size=11, color=PLOT_TEXT)),
                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                            font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                            xaxis=dict(title="ΔStd Dev", gridcolor=PLOT_GRID),
                            yaxis=dict(gridcolor=PLOT_GRID),
                            height=min(380, 60 * len(bar_df_imp) + 80),
                            margin=dict(l=10, r=20, t=40, b=40),
                        )
                        st.plotly_chart(fig_bar_imp, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=_ua_element_key("pc_fig_bar_imp", selected, selected_out))

                        # Correlation |r| vs threshold bar
                        cor_df_imp = imp_df_ui[["Parameter", "Correlation"]].copy()
                        cor_df_imp["abs_r"] = cor_df_imp["Correlation"].abs()
                        cor_df_imp["pass"]  = cor_df_imp["abs_r"] >= r_min_imp
                        cor_df_imp = cor_df_imp.sort_values("abs_r")
                        cor_colors_imp = [C[0] if p else "#57606a"
                                          for p in cor_df_imp["pass"]]
                        fig_cor_imp = go.Figure(go.Bar(
                            y=cor_df_imp["Parameter"],
                            x=cor_df_imp["abs_r"],
                            orientation="h",
                            marker_color=cor_colors_imp,
                            opacity=0.85,
                            text=[f"{v:.3f}" for v in cor_df_imp["abs_r"]],
                            textposition="outside",
                        ))
                        fig_cor_imp.add_vline(
                            x=r_min_imp, line_dash="dash",
                            line_color="#9a6700", line_width=2,
                            annotation_text=f"|r_min|={r_min_imp:.4f}",
                            annotation_font_color="#9a6700",
                        )
                        fig_cor_imp.update_layout(
                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                            font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                            margin=dict(l=10, r=60, t=10, b=30),
                            xaxis=dict(gridcolor=PLOT_GRID,
                                       range=[0, min(1.05,
                                                     cor_df_imp["abs_r"].max() * 1.15)]),
                            yaxis=dict(gridcolor=PLOT_GRID),
                            showlegend=False, height=280,
                        )
                        st.plotly_chart(fig_cor_imp, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=_ua_element_key("pc_fig_cor_imp", selected, selected_out))

                    st.markdown("---")
                    st.markdown(
                        f'<div style="font-size:0.78rem;color:#57606a;'
                        f'text-transform:uppercase;letter-spacing:0.1em;">'
                        f'Scatter plots — parameters above threshold '
                        f'(|r| ≥ {r_min_imp:.4f})</div>',
                        unsafe_allow_html=True,
                    )

                    n_imp_cols = min(3, len(imp_df_ui))
                    _imp_grid = [
                        list(imp_df_ui.iterrows())[i:i + n_imp_cols]
                        for i in range(0, len(imp_df_ui), n_imp_cols)
                    ]
                    for _row_params in _imp_grid:
                        _scols = st.columns(n_imp_cols)
                        for _col_w, (_, _row) in zip(_scols, _row_params):
                            with _col_w:
                                _xv  = np.array(_row["_xv"])
                                _yv  = np.array(_row["_yv"])
                                _cx  = np.array(_row["_coeffs"])
                                _xf  = np.linspace(_xv.min(), _xv.max(), 80)
                                _yf  = np.polyval(_cx, _xf)
                                _r_p = float(_row["Correlation"])
                                _deg = int(_row["Fit Degree"])
                                _r_color = C[1] if _r_p < 0 else C[0]

                                _fig_sc = go.Figure()
                                _fig_sc.add_trace(go.Scatter(
                                    x=list(_xv), y=list(_yv), mode="markers",
                                    marker=dict(size=4, color=C[0], opacity=0.55),
                                    showlegend=False,
                                ))
                                _fig_sc.add_trace(go.Scatter(
                                    x=list(_xf), y=list(_yf), mode="lines",
                                    line=dict(color=C[1], width=2),
                                    showlegend=False,
                                ))
                                _fig_sc.add_hline(y=0 if _row.name > 0 else _yv.mean(),
                                                  line_dash="dot",
                                                  line_color="#57606a", line_width=1)
                                _fig_sc.update_layout(
                                    plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                    font_color=PLOT_TEXT, font_family="IBM Plex Sans",
                                    margin=dict(l=36, r=10, t=36, b=10),
                                    xaxis=dict(gridcolor=PLOT_GRID,
                                               title=dict(text=_row["Parameter"],
                                                          font_size=9)),
                                    yaxis=dict(gridcolor=PLOT_GRID,
                                               title=dict(
                                                   text=(_scalar_display_label(selected_out)
                                                         if _row.name == 0
                                                         else f"Residual after {_row['Response']} fit"),
                                                   font_size=9)),
                                    title=dict(
                                        text=(f"<b>{_row['Parameter']}</b>  "
                                              f"<span style='color:{_r_color}'>"
                                              f"r = {_r_p:+.3f}</span>"),
                                        font_color=PLOT_TEXT, font_size=10,
                                    ),
                                    showlegend=False, height=260,
                                )
                                # Polynomial equation string
                                _pwrs = list(range(_deg, -1, -1))
                                _parts = []
                                for _pw, _c in zip(_pwrs, list(_cx)):
                                    if _pw == 0:
                                        _parts.append(f"{_c:+.4g}")
                                    elif _pw == 1:
                                        _parts.append(f"{_c:+.4g}·x")
                                    else:
                                        _parts.append(f"{_c:+.4g}·x^{_pw}")
                                _eq_str = ("PCT ≈ " if _row.name == 0 else "Δ ≈ ") + "  ".join(_parts)
                                _deg_lbl = f"deg {_deg}  adj R²={_row['Adj R²']:.3f}"
                                st.plotly_chart(_fig_sc, use_container_width=True,
                                                config={"displayModeBar": False},
                                                key=_ua_element_key("pc_fig_sc", selected, selected_out, _row.name))
                                st.caption(f"{_eq_str}   [{_deg_lbl}]")


                # ── FLARECON Importance Analysis (appended when CON outputs present) ──
                if con_scalar_cols:
                    st.markdown("---")
                    st.markdown("#### ⚛ FLARECON — Importance Analysis")
                    con_tc1b, con_tc2b, con_tc3b = st.columns([2, 2, 3])
                    with con_tc1b:
                        _con_conf_pct = st.select_slider(
                            "Confidence level (two-sided)",
                            options=[80, 85, 90, 95, 97.5, 99],
                            value=97.5,
                            format_func=lambda x: f"{x}%",
                            key="con_imp_conf_pct",
                        )
                    with con_tc2b:
                        _con_t_crit = T_TABLE[_con_conf_pct]
                        _con_n_ok_imp = len(df_ok)
                        _con_r_min  = (_con_t_crit / np.sqrt(_con_n_ok_imp - 2 + _con_t_crit ** 2)
                                       if _con_n_ok_imp > 2 else 1.0)
                        st.metric("t-critical", f"{_con_t_crit:.3f}",
                                  delta=f"two-sided {_con_conf_pct}%, n={_con_n_ok_imp}",
                                  delta_color="off")
                    with con_tc3b:
                        st.metric("Minimum |r| threshold", f"{_con_r_min:.4f}")

                    for _csk_idx, _csk in enumerate([selected_con_out] if selected_con_out else []):
                        _csk_lbl = _CON_DISPLAY_LABELS.get(_csk, _csk.replace("_", " "))
                        _csk_y_raw = pd.to_numeric(df_ok[_csk], errors="coerce")
                        _csk_arr, _csk_unit = _ua_convert_con_scalar_for_plot(_csk_y_raw, _csk, _use_english)
                        _csk_y = pd.Series(_csk_arr, index=df_ok.index)
                        _csk_valid = np.isfinite(_csk_arr)
                        _csk_n_ok  = int(_csk_valid.sum())
                        if _csk_n_ok < 4:
                            st.info(f"{_csk_lbl}: insufficient valid samples for importance analysis.")
                            continue

                        # Build the stepwise table using display-unit values so
                        # variances, ΔStd Dev, residual standard deviations, and
                        # scatter axes are consistent with the selected unit system.
                        _df_ok_con_disp = df_ok.copy()
                        _df_ok_con_disp[_csk] = _csk_y
                        _csk_imp_df, _csk_sqrt_sum, _csk_ratio, _csk_actual_std, _csk_resid_std = (
                            _build_importance_table(_df_ok_con_disp, _csk, in_cols, r_min=_con_r_min)
                        )
                        _csk_actual_std_raw = float(np.std(_csk_y.dropna(), ddof=1))

                        st.markdown(f"**{_csk_lbl}**")
                        _csk_ratio_cls = "ok" if _csk_ratio >= 0.90 else "warn"
                        _csk_resid_pct = (f"{_csk_resid_std / _csk_actual_std_raw * 100:.1f}% unexplained"
                                          if _csk_actual_std_raw > 0 else "")
                        st.markdown(
                            f'<div class="metric-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:0.75rem;">'                            f'<div class="metric-tile"><div class="val">{fmt_general(_csk_actual_std_raw)}</div>'                            f'<div class="lbl">Actual std dev σ<sub>o</sub></div></div>'                            f'<div class="metric-tile"><div class="val">{fmt_general(_csk_sqrt_sum)}</div>'                            f'<div class="lbl">Convolution √(Σ Var) — Eq.(11)</div></div>'                            f'<div class="metric-tile {_csk_ratio_cls}"><div class="val">{_csk_ratio:.3f}</div>'                            f'<div class="lbl">Ratio estimate / actual — Eq.(12)<br>'                            f'<small>{"≥ 0.90 ✓" if _csk_ratio >= 0.90 else "< 0.90 — increase n"}</small>'                            f'</div></div>'                            f'<div class="metric-tile"><div class="val">{fmt_general(_csk_resid_std)}</div>'                            f'<div class="lbl">Residual std dev<br><small>{_csk_resid_pct}</small></div></div>'                            f'</div>',
                            unsafe_allow_html=True,
                        )

                        if len(_csk_imp_df) == 0:
                            st.info(f"No parameters exceed |r| = {_con_r_min:.4f} for {_csk_lbl}.")
                        else:
                            _csk_c1, _csk_c2 = st.columns([3, 2])
                            with _csk_c1:
                                st.markdown(
                                    f'<div style="font-size:0.78rem;color:#57606a;text-transform:uppercase;'                                    f'letter-spacing:0.1em;">Stepwise Regression — {_csk_lbl} '                                    f'({len(_csk_imp_df)} parameter(s) |r| ≥ {_con_r_min:.4f})</div>',
                                    unsafe_allow_html=True,
                                )
                                _csk_disp = _stepwise_display_table(
                                    _df_ok_con_disp, _csk, in_cols, _csk_imp_df, _con_r_min
                                )
                                _csk_styled = (
                                    _csk_disp.style
                                    .map(lambda v: _corr_threshold_style(v, _con_r_min),
                                         subset=["Correlation"])
                                    .format({"Variance": "{:.3f}", "ΔStd Dev": "{:.3f}",
                                             "Correlation": "{:+.4f}", "Adj R²": "{:.4f}"},
                                            na_rep="—")
                                )
                                st.dataframe(_csk_styled, use_container_width=True, height=320)

                            with _csk_c2:
                                _csk_bar = _csk_imp_df[["Parameter", "ΔStd Dev"]].sort_values("ΔStd Dev")
                                fig_csk_bar = go.Figure(go.Bar(
                                    y=_csk_bar["Parameter"], x=_csk_bar["ΔStd Dev"],
                                    orientation="h", marker_color=C[0], opacity=0.85,
                                ))
                                fig_csk_bar.update_layout(
                                    title=dict(text=f"ΔStd Dev — {_csk_lbl}",
                                               font=dict(family="IBM Plex Mono", size=10, color=PLOT_TEXT)),
                                    plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                    font=dict(family="IBM Plex Sans", size=10, color=PLOT_TEXT),
                                    xaxis=dict(title="ΔStd Dev", gridcolor=PLOT_GRID),
                                    yaxis=dict(gridcolor=PLOT_GRID),
                                    height=min(280, 55 * len(_csk_bar) + 70),
                                    margin=dict(l=10, r=20, t=35, b=35),
                                )
                                st.plotly_chart(fig_csk_bar, use_container_width=True,
                                                config={"displayModeBar": False},
                                                key=_ua_element_key("con_imp_bar", selected, _csk, _csk_idx))

                            # Scatter plots
                            _csk_n_cols = min(3, len(_csk_imp_df))
                            _csk_grid = [
                                list(_csk_imp_df.iterrows())[i:i + _csk_n_cols]
                                for i in range(0, len(_csk_imp_df), _csk_n_cols)
                            ]
                            for _csk_row_params in _csk_grid:
                                _csk_scols = st.columns(_csk_n_cols)
                                for _csk_col_w, (_csk_ri, _csk_row) in zip(_csk_scols, _csk_row_params):
                                    with _csk_col_w:
                                        _csk_xv  = np.array(_csk_row["_xv"])
                                        _csk_yv  = np.array(_csk_row["_yv"])
                                        _csk_cx  = np.array(_csk_row["_coeffs"])
                                        _csk_xf  = np.linspace(_csk_xv.min(), _csk_xv.max(), 80)
                                        _csk_yf  = np.polyval(_csk_cx, _csk_xf)
                                        _csk_r   = float(_csk_row["Correlation"])
                                        _csk_deg = int(_csk_row["Fit Degree"])
                                        _csk_r_color = C[1] if _csk_r < 0 else C[0]
                                        fig_csk_sc = go.Figure()
                                        fig_csk_sc.add_trace(go.Scatter(
                                            x=list(_csk_xv), y=list(_csk_yv), mode="markers",
                                            marker=dict(size=4, color=C[0], opacity=0.55),
                                            showlegend=False,
                                        ))
                                        fig_csk_sc.add_trace(go.Scatter(
                                            x=list(_csk_xf), y=list(_csk_yf), mode="lines",
                                            line=dict(color=C[1], width=2), showlegend=False,
                                        ))
                                        fig_csk_sc.update_layout(
                                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                            font_color=PLOT_TEXT,
                                            margin=dict(l=36, r=10, t=36, b=10),
                                            xaxis=dict(gridcolor=PLOT_GRID,
                                                       title=dict(text=_csk_row["Parameter"], font_size=9)),
                                            yaxis=dict(gridcolor=PLOT_GRID,
                                                       title=dict(
                                                           text=(_csk_lbl if _csk_row.name == 0
                                                                 else f"Residual after {_csk_row['Response']} fit"),
                                                           font_size=9)),
                                            title=dict(
                                                text=(f"<b>{_csk_row['Parameter']}</b>  "
                                                      f"<span style='color:{_csk_r_color}'>"
                                                      f"r = {_csk_r:+.3f}</span>"),
                                                font_color=PLOT_TEXT, font_size=10),
                                            showlegend=False, height=240,
                                        )
                                        _csk_pwrs = list(range(_csk_deg, -1, -1))
                                        _csk_parts = []
                                        for _cpw, _cc in zip(_csk_pwrs, list(_csk_cx)):
                                            if _cpw == 0: _csk_parts.append(f"{_cc:+.4g}")
                                            elif _cpw == 1: _csk_parts.append(f"{_cc:+.4g}·x")
                                            else: _csk_parts.append(f"{_cc:+.4g}·x^{_cpw}")
                                        _csk_eq = "y ≈ " + "  ".join(_csk_parts)
                                        st.plotly_chart(fig_csk_sc, use_container_width=True,
                                                        config={"displayModeBar": False},
                                                        key=_ua_element_key("con_imp_sc", selected, _csk, _csk_idx, _csk_ri))
                                        st.caption(f"{_csk_eq}   [deg {_csk_deg}  adj R²={_csk_row['Adj R²']:.3f}]")

            # ── SUB-TAB 3: Ordered Statistic ──────────────────────────────
            with sub_ord:
                st.markdown(f"#### 🔥 FLARE — Ordered Statistics — {_scalar_display_label(selected_out)}")
                c_ord, c_conf_ord = st.columns(2)

                n_ok_ord = int(np.isfinite(y_arr).sum())
                y_sorted_ord = np.sort(y_arr[np.isfinite(y_arr)])

                with c_ord:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#57606a;'
                        'text-transform:uppercase;letter-spacing:0.1em;">'
                        'Top-20 Ordered Values</div>',
                        unsafe_allow_html=True,
                    )
                    _m_sel = st.number_input(
                        "Order statistic (m)",
                        min_value=1, max_value=max(1, n_ok_ord),
                        value=1, step=1, key="ord_m_sel",
                        help="m=1 gives the maximum (most conservative) value.",
                    )
                    _pct_m = float(y_sorted_ord[-_m_sel]) if n_ok_ord >= _m_sel else None

                    if n_ok_ord > 0:
                        ord_n_disp = min(20, n_ok_ord)
                        _top_vals = y_sorted_ord[-ord_n_disp:][::-1]
                        dot_c_ord = [
                            C[1] if m == _m_sel else C[0]
                            for m in range(1, ord_n_disp + 1)
                        ]
                        fig_ord_ui = go.Figure()
                        fig_ord_ui.add_trace(go.Scatter(
                            x=list(range(1, ord_n_disp + 1)),
                            y=list(_top_vals),
                            mode="lines+markers",
                            line=dict(color=C[0], width=2),
                            marker=dict(color=dot_c_ord, size=8),
                        ))
                        if _pct_m is not None:
                            fig_ord_ui.add_scatter(
                                x=[_m_sel], y=[_pct_m],
                                mode="markers+text",
                                marker=dict(size=12, color="#9a6700", symbol="star"),
                                text=[f"m={_m_sel}  {fmt_general(_pct_m)}"],
                                textposition="top right",
                                textfont=dict(color="#9a6700", size=11),
                                name=f"m={_m_sel} order stat",
                            )
                        fig_ord_ui.update_layout(
                            title=dict(
                                text=f"Ordered Values — {_scalar_display_label(selected_out)}",
                                font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                            font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                            xaxis=dict(title="Order statistic (m)", gridcolor=PLOT_GRID),
                            yaxis=dict(title=y_label, gridcolor=PLOT_GRID),
                            height=340,
                        )
                        st.plotly_chart(fig_ord_ui, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=_ua_element_key("pc_fig_ord_ui", selected, selected_out, _m_sel))

                        # KPI strip — wide custom tiles to prevent value truncation
                        _beta_disp = _wilks_confidence_beta(n_ok_ord, _m_sel, 0.95)
                        _beta_cls  = "ok" if _beta_disp >= 0.95 else "danger"
                        _m_val_str = (f"{fmt_general(_pct_m)} {y_unit_disp}".strip()
                                      if _pct_m is not None else "—")
                        _max_val_str = (f"{fmt_general(float(y_sorted_ord[-1]))} {y_unit_disp}".strip()
                                        if n_ok_ord > 0 else "—")
                        _med_val_str = (f"{fmt_general(float(np.median(y_sorted_ord)))} {y_unit_disp}".strip()
                                        if n_ok_ord > 0 else "—")
                        st.markdown(
                            f'<div class="metric-grid" style="grid-template-columns:repeat(3,1fr);margin-top:0.75rem;">'
                            f'<div class="metric-tile">'
                            f'<div class="val">{_m_val_str}</div>'
                            f'<div class="lbl">m={_m_sel} order-stat value</div>'
                            f'</div>'
                            f'<div class="metric-tile {_beta_cls}">'
                            f'<div class="val">{_beta_disp*100:.2f}%</div>'
                            f'<div class="lbl">Wilks β (n={n_ok_ord}, m={_m_sel}, γ=95%) '
                            f'{"✓ ≥95%" if _beta_disp>=0.95 else "✗ <95%"}</div>'
                            f'</div>'
                            f'<div class="metric-tile">'
                            f'<div class="val">{_max_val_str}</div>'
                            f'<div class="lbl">Maximum value (m=1)</div>'
                            f'</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                with c_conf_ord:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#57606a;'
                        'text-transform:uppercase;letter-spacing:0.1em;">'
                        'Confidence β vs. Order m</div>',
                        unsafe_allow_html=True,
                    )
                    if n_ok_ord >= 2:
                        _ms_range = list(range(1, min(11, n_ok_ord + 1)))
                        _betas_95_ui  = [_wilks_confidence_beta(n_ok_ord, m, 0.95) * 100
                                         for m in _ms_range]
                        fig_conf_ui = go.Figure()
                        fig_conf_ui.add_trace(go.Scatter(
                            x=_ms_range, y=_betas_95_ui,
                            mode="lines+markers",
                            line=dict(color=C[1], width=2),
                            marker=dict(size=6), name="γ = 95%",
                        ))
                        fig_conf_ui.add_hline(y=95, line_dash="dot",
                                              line_color="#9a6700",
                                              annotation_text="95% confidence",
                                              annotation_font_color="#9a6700")
                        fig_conf_ui.update_layout(
                            title=dict(
                                text=f"β vs. m  (n = {n_ok_ord})",
                                font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                            plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                            font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                            xaxis=dict(title="Order statistic (m)", gridcolor=PLOT_GRID),
                            yaxis=dict(title="Confidence β (%)", gridcolor=PLOT_GRID,
                                       range=[85, 101]),
                            height=340,
                        )
                        st.plotly_chart(fig_conf_ui, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=_ua_element_key("pc_fig_conf_ui", selected, selected_out, n_ok_ord))


                # ── FLARECON Ordered Statistics (appended when CON outputs present) ──
                if con_scalar_cols:
                    st.markdown("---")
                    st.markdown("#### ⚛ FLARECON — Ordered Statistics")

                    for _cok_idx, _cok in enumerate([selected_con_out] if selected_con_out else []):
                        _cok_lbl = _CON_DISPLAY_LABELS.get(_cok, _cok.replace("_", " "))
                        _cok_y_raw = pd.to_numeric(df_ok[_cok], errors="coerce").to_numpy(dtype=float)
                        _cok_y, _cok_unit = _ua_convert_con_scalar_for_plot(_cok_y_raw, _cok, _use_english)
                        _cok_valid = np.isfinite(_cok_y)
                        _cok_n_ok  = int(_cok_valid.sum())
                        _cok_sorted = np.sort(_cok_y[_cok_valid])
                        if _cok_n_ok == 0:
                            continue

                        st.markdown(f"**{_cok_lbl}**")
                        _H2_LFL = 4.0
                        _H2_DET = 8.0
                        _cok_c_ord, _cok_c_conf = st.columns(2)

                        with _cok_c_ord:
                            _cok_m_sel = st.number_input(
                                f"Order statistic m — {_cok_lbl}",
                                min_value=1, max_value=max(1, _cok_n_ok),
                                value=1, step=1,
                                key=f"con_ord_m_{_cok_idx}",
                                help="m=1 gives the maximum (most conservative) value.",
                            )
                            _cok_pct_m = (float(_cok_sorted[-_cok_m_sel])
                                          if _cok_n_ok >= _cok_m_sel else None)
                            if _cok_n_ok > 0:
                                _cok_ord_n = min(20, _cok_n_ok)
                                _cok_top   = _cok_sorted[-_cok_ord_n:][::-1]
                                _cok_dot_c = [C[1] if m == _cok_m_sel else C[0]
                                              for m in range(1, _cok_ord_n + 1)]
                                fig_cok_ord = go.Figure()
                                fig_cok_ord.add_trace(go.Scatter(
                                    x=list(range(1, _cok_ord_n + 1)),
                                    y=list(_cok_top),
                                    mode="lines+markers",
                                    line=dict(color=C[0], width=2),
                                    marker=dict(color=_cok_dot_c, size=8),
                                ))
                                if _cok_pct_m is not None:
                                    fig_cok_ord.add_scatter(
                                        x=[_cok_m_sel], y=[_cok_pct_m],
                                        mode="markers+text",
                                        marker=dict(size=12, color="#9a6700", symbol="star"),
                                        text=[f"m={_cok_m_sel}  {fmt_general(_cok_pct_m)}"],
                                        textposition="top right",
                                        textfont=dict(color="#9a6700", size=11),
                                        name=f"m={_cok_m_sel} order stat",
                                    )
                                if _cok == "CON_H2_peak_volpct":
                                    fig_cok_ord.add_hline(y=_H2_LFL, line_dash="dash",
                                                          line_color=C[3], annotation_text="LFL 4%",
                                                          annotation_font_color=C[3])
                                    fig_cok_ord.add_hline(y=_H2_DET, line_dash="dash",
                                                          line_color=C[4], annotation_text="Det. 8%",
                                                          annotation_font_color=C[4])
                                fig_cok_ord.update_layout(
                                    title=dict(
                                        text=f"Ordered Values — {_cok_lbl}",
                                        font=dict(family="IBM Plex Mono", size=12, color=PLOT_TEXT)),
                                    plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                    font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                                    xaxis=dict(title="Order statistic (m)", gridcolor=PLOT_GRID),
                                    yaxis=dict(title=_cok_lbl, gridcolor=PLOT_GRID),
                                    height=300,
                                )
                                st.plotly_chart(fig_cok_ord, use_container_width=True,
                                                config={"displayModeBar": False},
                                                key=_ua_element_key("con_ord_plot", selected, _cok, _cok_idx, _cok_m_sel))

                                _cok_beta = _wilks_confidence_beta(_cok_n_ok, _cok_m_sel, 0.95)
                                _cok_beta_cls = "ok" if _cok_beta >= 0.95 else "danger"
                                _cok_m_str  = fmt_general(_cok_pct_m) if _cok_pct_m is not None else "—"
                                _cok_max_str = fmt_general(float(_cok_sorted[-1]))
                                st.markdown(
                                    f'<div class="metric-grid" style="grid-template-columns:repeat(3,1fr);margin-top:0.6rem;margin-bottom:1rem;">'                                    f'<div class="metric-tile"><div class="val">{_cok_m_str}</div>'                                    f'<div class="lbl">m={_cok_m_sel} order-stat value<br><small>{_cok_lbl}</small></div></div>'                                    f'<div class="metric-tile {_cok_beta_cls}"><div class="val">{_cok_beta*100:.2f}%</div>'                                    f'<div class="lbl">Wilks β (n={_cok_n_ok}, m={_cok_m_sel}, γ=95%)<br>'                                    f'<small>{"✓ ≥95%" if _cok_beta>=0.95 else "✗ <95%"}</small></div></div>'                                    f'<div class="metric-tile"><div class="val">{_cok_max_str}</div>'                                    f'<div class="lbl">Maximum value (m=1)</div></div>'                                    f'</div>',
                                    unsafe_allow_html=True,
                                )

                        with _cok_c_conf:
                            if _cok_n_ok >= 2:
                                _cok_ms_r  = list(range(1, min(11, _cok_n_ok + 1)))
                                _cok_betas = [_wilks_confidence_beta(_cok_n_ok, m, 0.95) * 100
                                              for m in _cok_ms_r]
                                fig_cok_conf = go.Figure()
                                fig_cok_conf.add_trace(go.Scatter(
                                    x=_cok_ms_r, y=_cok_betas,
                                    mode="lines+markers",
                                    line=dict(color=C[1], width=2),
                                    marker=dict(size=6), name="γ = 95%",
                                ))
                                fig_cok_conf.add_hline(y=95, line_dash="dot",
                                                       line_color="#9a6700",
                                                       annotation_text="95% confidence",
                                                       annotation_font_color="#9a6700")
                                fig_cok_conf.update_layout(
                                    title=dict(
                                        text=f"β vs. m  —  {_cok_lbl}  (n={_cok_n_ok})",
                                        font=dict(family="IBM Plex Mono", size=11, color=PLOT_TEXT)),
                                    plot_bgcolor=PLOT_BG, paper_bgcolor=PLOT_PAPER,
                                    font=dict(family="IBM Plex Sans", size=11, color=PLOT_TEXT),
                                    xaxis=dict(title="Order statistic (m)", gridcolor=PLOT_GRID),
                                    yaxis=dict(title="Confidence β (%)", gridcolor=PLOT_GRID,
                                               range=[85, 101]),
                                    height=300,
                                )
                                st.plotly_chart(fig_cok_conf, use_container_width=True,
                                                config={"displayModeBar": False},
                                                key=_ua_element_key("con_ord_conf", selected, _cok, _cok_idx, _cok_n_ok))

                # Minimum-n reference table
                st.markdown("---")
                st.markdown(
                    '<div style="font-size:0.78rem;color:#57606a;'
                    'text-transform:uppercase;letter-spacing:0.1em;">'
                    'Minimum n for 95/95 (various order statistics)</div>',
                    unsafe_allow_html=True,
                )
                _ref_rows = []
                for _m_ref in [1, 2, 3, 10, 40, 100]:
                    for _n_ref in range(_m_ref, 2500):
                        if _wilks_confidence_beta(_n_ref, _m_ref, 0.95) >= 0.95:
                            _ref_rows.append({
                                "m (failures allowed)": _m_ref,
                                "Min n for 95/95": _n_ref,
                                f"β at n={_n_ref}": f"{_wilks_confidence_beta(_n_ref, _m_ref, 0.95)*100:.2f}%",
                            })
                            break
                if _ref_rows:
                    st.dataframe(
                        pd.DataFrame(_ref_rows).set_index("m (failures allowed)"),
                        use_container_width=True,
                    )

                st.markdown("---")
                st.markdown(
                    '<div style="font-size:0.78rem;color:#57606a;'
                    'text-transform:uppercase;letter-spacing:0.1em;">'
                    'Confidence Summary (γ = 95%)</div>',
                    unsafe_allow_html=True,
                )
                st.caption(
                    "This Wilks confidence summary depends only on the number of "
                    "successful samples and the selected order statistic m; it is "
                    "therefore independent of the selected FLARE or FLARECON FOM."
                )

                def _confidence_summary_rows(n_ok_value):
                    rows = []
                    for m_val in [1, 2, 3, 5, 10]:
                        if m_val > n_ok_value:
                            break
                        beta_val = _wilks_confidence_beta(n_ok_value, m_val, 0.95)
                        rows.append({
                            "m": m_val,
                            "β (%)": round(beta_val * 100, 3),
                            "Status": "✓ ≥ 95%" if beta_val >= 0.95 else "✗ < 95%",
                        })
                    return rows

                _primary_conf_rows = _confidence_summary_rows(n_ok_ord)
                if _primary_conf_rows:
                    st.dataframe(
                        pd.DataFrame(_primary_conf_rows).set_index("m"),
                        use_container_width=True,
                    )
                else:
                    st.caption("Not enough successful samples for a confidence summary.")

            # ══════════════════════════════════════════════════════════════
            # PDF Export — builds cover + all result figures
            # ══════════════════════════════════════════════════════════════
            st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
            st.markdown("**PDF Export**")
            st.caption(
                "Generate a portrait-letter PDF with a cover page (run parameters + "
                "stepwise regression table) and all result plots two-up per page."
            )
            if st.button("📄  Export Results to PDF", key="ua_export_pdf_btn"):
                try:
                    _run_dir_pdf = Path(st.session_state.get("ua_run_dir") or WORK_DIR)
                    _pdf_tag = (
                        f"{_safe_plot_token(selected_out)}_"
                        f"{_datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    )
                    _created = _make_ua_plot_set(
                        df_ok,
                        ts_list_now,
                        _run_dir_pdf,
                        st.session_state.get("ua_case") or selected,
                        selected_ts,
                        selected_out,
                        use_english=_use_english,
                        tag=_pdf_tag,
                    )
                    st.success(
                        f"PDF created in `{_created['out_dir'].name}`."
                    )
                    with open(_created["pdf"], "rb") as _pf:
                        st.download_button(
                            "⬇  Download Results PDF",
                            data=_pf.read(),
                            file_name=_created["pdf"].name,
                            mime="application/pdf",
                            key="ua_download_results_pdf",
                        )
                except Exception as _pdf_e:
                    st.error(f"Could not create PDF: {_pdf_e}")

    # ── AI Uncertainty Narrative ──────────────────────────────────────────────
    if df is not None and len(df) > 0 and "status" in df.columns:
        with st.expander("🤖  AI Uncertainty Narrative", expanded=True):
            if _BUDDY_B64:
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:0.6rem;margin-bottom:0.5rem;'>"
                    f"<img src='{_BUDDY_B64}' style='height:2.2rem;width:2.2rem;border-radius:50%;object-fit:cover;flex-shrink:0;'/>"
                    f"<span style='font-weight:700;font-size:1rem;'>FLARE Buddy — AI Uncertainty Narrative</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            st.caption(
                "Generate a technical narrative of the uncertainty-analysis results, "
                "including key output distributions, sensitivity rankings, and interpretation."
            )
            # Put the expensive AI action in a form so the detail slider and
            # Generate click are submitted as one coherent snapshot.  This avoids
            # Streamlit rerun timing issues where a rapid click immediately after
            # changing a control can launch the narrative with stale context.
            with st.form("ua_ai_generate_form", clear_on_submit=False):
                _ua_detail = st.slider(
                    "Narrative detail", 0.0, 1.0, 0.55, 0.05,
                    key="ua_ai_detail",
                    help=("21 discrete levels in 0.05 increments. Each detent maps to a distinct "
                         "target length, from about 300 words at 0.00 to about 5300 words at 1.00."),
                )
                _ua_detent = _ua_detail_level(_ua_detail)
                _ua_word_target = _ua_detail_word_target(_ua_detail)
                _ua_word_lo, _ua_word_hi = _ua_detail_word_range(_ua_detail)
                _ua_detail_label = _ua_detail_descriptor(_ua_detail)
                st.caption(
                    f"**{_ua_detail_label}** — detent {_ua_detent:.2f}, "
                    f"target ~{_ua_word_target:,} words "
                    f"(acceptable range {_ua_word_lo:,}–{_ua_word_hi:,})."
                )
                _gen_ua_ai = st.form_submit_button(
                    "Generate narrative",
                    type="primary",
                    help="Submits the current narrative settings as a stable snapshot.",
                )

            # Use the submitted form value directly. Do not defer the API call
            # to a second rerun: that previously allowed the slider to fall
            # back to its default value on some deployments. The form submission
            # itself is the atomic snapshot.
            if _gen_ua_ai:
                try:
                    _ua_detail = float(st.session_state.get("ua_ai_detail", _ua_detail))
                except Exception:
                    _ua_detail = float(_ua_detail)
                st.session_state["ua_ai_last_submitted_detail"] = _ua_detail
                _ua_snap_level = _ua_detail_level(_ua_detail)
                _ua_snap_target = _ua_detail_word_target(_ua_detail)
                st.caption(
                    f"AI request snapshot: detail {_ua_snap_level:.2f}, "
                    f"target ~{_ua_snap_target:,} words."
                )

            _ua_cols = st.columns([1, 3])
            with _ua_cols[0]:
                if st.button("Clear", key="ua_ai_clear"):
                    st.session_state.pop("ua_ai_narrative", None)
                    st.rerun()

            if _gen_ua_ai:
                try:
                    # Place progress and the live narrative outside st.status so Streamlit
                    # renders section-by-section updates visibly.  After completion we
                    # rerun once to clear the live placeholder and display only the stored
                    # final report/download controls.
                    _ai_progress = st.progress(0, text="Preparing AI narrative request…")
                    _ua_ai_live_box = st.empty()
                    _ai_status = st.status("Preparing AI uncertainty narrative…", expanded=True)
                    _ai_progress.progress(10, text="Collecting UA results and selected FOM context…")
                    _ai_status.update(label="Collecting UA results and selected FOM context…", state="running")
                    _time.sleep(0.05)
                    _df_ai = _restore_input_columns_for_stats(
                        st.session_state.ua_results,
                        st.session_state.get("ua_samples"),
                    )
                    _df_ok_ai = _df_ai[_df_ai["status"] == "OK"].copy()
                    _in_cols_ai  = [c for c in _df_ok_ai.columns if c.startswith("in_")]
                    _out_cols_ai = [c for c in _df_ok_ai.columns
                                    if not c.startswith("in_") and c not in ("sample", "status", "error")]
                    _numeric_out = [c for c in _out_cols_ai if pd.api.types.is_numeric_dtype(_df_ok_ai[c])]
                    _CON_SCALAR_KEYS_AI = {"CON_P_peak_kPa", "CON_T_peak_C",
                                           "CON_H2_peak_volpct", "CON_sump_level_peak_m"}
                    _con_out   = [c for c in _numeric_out if c in _CON_SCALAR_KEYS_AI]
                    _flare_out = [c for c in _numeric_out if c not in _CON_SCALAR_KEYS_AI]

                    def _ua_ai_unit_policy_text(use_english=False):
                        return (
                            "NARRATIVE UNIT POLICY: The user selected English display units. Report dimensional values in English display units: °F, psia, lbm/lb/s, ft/ft/s, and Btu/hr where applicable. Use the converted values supplied in this payload; do not revert to SI/metric values except where a regulatory limit is conventionally stated in SI."
                            if use_english else
                            "NARRATIVE UNIT POLICY: The user selected Metric display units. Report dimensional values in metric display units: °C, kPa, kg/kg/s, m/m/s, and W/MW where applicable. Use the converted values supplied in this payload; do not add English equivalents except where a regulatory limit is conventionally stated in English."
                        )

                    def _ua_ai_convert_output_values(col_name, values):
                        if not col_name:
                            return np.asarray(values, dtype=float), ""
                        try:
                            if col_name in _CON_SCALAR_KEYS_AI:
                                return _ua_convert_con_scalar_for_plot(values, col_name, _use_english)
                            return _ua_convert_scalar_for_plot(values, col_name, _use_english)
                        except Exception:
                            return np.asarray(values, dtype=float), ""

                    def _ua_ai_sigma_scale(col_name):
                        try:
                            conv, unit = _ua_ai_convert_output_values(col_name, np.array([0.0, 1.0]))
                            scale = abs(float(conv[1]) - float(conv[0]))
                            if not np.isfinite(scale) or scale == 0.0:
                                scale = 1.0
                            return scale, unit
                        except Exception:
                            return 1.0, ""

                    # Scale the prompt payload, token budget, and read timeout together.
                    # Small narrative-detail values should receive a compact prompt and
                    # enough output tokens to finish; high-detail values may send more
                    # cross-check data and get a longer API read timeout.
                    _ua_detail = float(_ua_detail)
                    _detail_level_for_payload = _ua_detail_level(_ua_detail)
                    # Prompt payload also scales by the 0.05 detents.  Keep the
                    # low end compact enough to support abstract-length output,
                    # while exposing progressively more cross-check context as
                    # the requested narrative gets longer.
                    _summary_cap = int(round(4 + _detail_level_for_payload * 16))
                    _con_summary_cap = int(round(2 + _detail_level_for_payload * 6))
                    _corr_cap = int(round(4 + _detail_level_for_payload * 21))
                    _sample_head_n = 0 if _detail_level_for_payload < 0.30 else int(round(3 + _detail_level_for_payload * 9))
                    _include_full_stepwise_rows = _detail_level_for_payload >= 0.65

                    def _scalar_summary(cols, df_s, cap=20):
                        rows = []
                        for _c in cols[:cap]:
                            _s_raw = pd.to_numeric(df_s[_c], errors="coerce").dropna()
                            if len(_s_raw):
                                _vals, _unit = _ua_ai_convert_output_values(_c, _s_raw.values)
                                _s = pd.Series(_vals).dropna()
                                if len(_s):
                                    _label = (_CON_DISPLAY_LABELS.get(_c, _c.replace("_", " "))
                                              if _c in _CON_SCALAR_KEYS_AI else _scalar_display_label(_c))
                                    rows.append({
                                        "output": _c,
                                        "display_label": _label,
                                        "unit": _unit,
                                        "n": int(len(_s)),
                                        "min": float(_s.min()), "mean": float(_s.mean()),
                                        "max": float(_s.max()),
                                        "std": float(_s.std()) if len(_s) > 1 else 0.0,
                                        "p95": float(_s.quantile(0.95)),
                                    })
                        return rows

                    _flare_summary = _scalar_summary(_flare_out, _df_ok_ai, cap=_summary_cap)
                    _con_summary   = _scalar_summary(_con_out,   _df_ok_ai, cap=_con_summary_cap)

                    # ── Stepwise importance analysis for the selected UI FOMs ──
                    # Keep the AI narrative aligned with the visible UI: use the
                    # time-series-derived FLARE scalar FOM and the selected
                    # containment FOM, apply the same Student-t thresholds, and
                    # include the terminal stopping row in the supplied data.
                    _primary_out_ai = (selected_out if selected_out in _numeric_out
                                       else (_flare_out[0] if _flare_out else None))
                    _primary_out_label_ai = (_scalar_display_label(_primary_out_ai)
                                             if _primary_out_ai else "")
                    _con_out_ai = (selected_con_out if selected_con_out in _numeric_out
                                   else (_con_out[0] if _con_out else None))
                    _con_out_label_ai = (_CON_DISPLAY_LABELS.get(_con_out_ai, _con_out_ai.replace("_", " "))
                                         if _con_out_ai else "")
                    _r_min_ai = float(globals().get("r_min_imp", locals().get("r_min_imp", 0.0)) or 0.0)
                    _con_r_min_ai = float(globals().get("_con_r_min", locals().get("_con_r_min", _r_min_ai)) or 0.0)

                    def _threshold_status_ai(r_val, r_min):
                        try:
                            _ar = abs(float(r_val))
                            _r0 = abs(float(r_min))
                        except Exception:
                            return "not available"
                        if not np.isfinite(_ar) or not np.isfinite(_r0):
                            return "not available"
                        if _ar < _r0:
                            return "fails Student-t threshold"
                        if _ar <= 1.10 * _r0:
                            return "passes but is within 10% above threshold"
                        return "passes comfortably above threshold"

                    def _stepwise_payload_ai(out_col, out_label, r_min):
                        if not out_col or out_col not in _df_ok_ai.columns or not _in_cols_ai:
                            return None
                        try:
                            _imp_df, _sqrt_sum, _ratio, _actual_std, _resid_std = (
                                _build_importance_table(_df_ok_ai, out_col,
                                                        _in_cols_ai, r_min=float(r_min))
                            )
                            _disp = _stepwise_display_table(
                                _df_ok_ai, out_col, _in_cols_ai, _imp_df, float(r_min)
                            )
                            _sigma_scale, _sigma_unit = _ua_ai_sigma_scale(out_col)
                            _rows = []
                            for _, _ir in _disp.iterrows():
                                _corr = _ir.get("Correlation", np.nan)
                                _rows.append({
                                    "row_type": str(_ir.get("Row Type", "accepted")),
                                    "response_dataset": str(_ir.get("Response", "")),
                                    "variance": (None if pd.isna(_ir.get("Variance", np.nan))
                                                 else round(float(_ir.get("Variance")) * (_sigma_scale ** 2), 6)),
                                    "estimated_delta_std_dev": (None if pd.isna(_ir.get("ΔStd Dev", np.nan))
                                                                else round(float(_ir.get("ΔStd Dev")) * _sigma_scale, 6)),
                                    "std_dev_unit": _sigma_unit,
                                    "parameter": str(_ir.get("Parameter", "")),
                                    "correlation": (None if pd.isna(_corr) else round(float(_corr), 6)),
                                    "abs_correlation": (None if pd.isna(_corr) else round(abs(float(_corr)), 6)),
                                    "fit_degree": str(_ir.get("Fit Degree", "")),
                                    "adjusted_r_squared": (None if pd.isna(_ir.get("Adj R²", np.nan))
                                                           else round(float(_ir.get("Adj R²")), 6)),
                                    "threshold_status": _threshold_status_ai(_corr, r_min),
                                })
                            _accepted = [r for r in _rows if r.get("row_type") == "accepted"]
                            _terminal = next((r for r in _rows if r.get("row_type") == "terminal"), None)
                            _payload = {
                                "output": out_col,
                                "display_label": out_label or out_col,
                                "display_unit": _sigma_unit,
                                "student_t_min_abs_r": round(float(r_min), 6),
                                "accepted_parameter_count": len(_accepted),
                                "actual_sigma": round(float(_actual_std) * _sigma_scale, 6) if np.isfinite(_actual_std) else None,
                                "convolution_sqrt_sum_variance_eq11": round(float(_sqrt_sum) * _sigma_scale, 6) if np.isfinite(_sqrt_sum) else None,
                                "ratio_estimate_to_actual_eq12": round(float(_ratio), 6) if np.isfinite(_ratio) else None,
                                "residual_sigma_unexplained": round(float(_resid_std) * _sigma_scale, 6) if np.isfinite(_resid_std) else None,
                                "accepted_rows": _accepted,
                                "terminal_stopping_row": _terminal,
                            }
                            if _include_full_stepwise_rows:
                                _payload["full_display_rows"] = _rows
                            return _payload
                        except Exception as _sw_e:
                            return {"output": out_col, "display_label": out_label or out_col,
                                    "error": str(_sw_e)}

                    _flare_importance_ai = _stepwise_payload_ai(
                        _primary_out_ai, _primary_out_label_ai, _r_min_ai
                    )
                    _con_importance_ai = (_stepwise_payload_ai(
                        _con_out_ai, _con_out_label_ai, _con_r_min_ai
                    ) if _con_out_ai else None)

                    # Wilks order-stat summary for the same selected FLARE FOM.
                    _wilks_ai_rows = []
                    if _primary_out_ai:
                        _y_ai = pd.to_numeric(_df_ok_ai[_primary_out_ai], errors="coerce").dropna().values
                        if len(_y_ai) > 0:
                            _y_ai_s = np.sort(_y_ai)
                            for _m_ai in [1, 2, 3]:
                                if _m_ai > len(_y_ai_s):
                                    break
                                _b_ai = _wilks_confidence_beta(len(_y_ai_s), _m_ai, 0.95)
                                _wy, _wu = _ua_ai_convert_output_values(_primary_out_ai, np.array([float(_y_ai_s[-_m_ai])]))
                                _wilks_ai_rows.append({
                                    "m": _m_ai,
                                    "value": round(float(_wy[0]), 3),
                                    "unit": _wu,
                                    "beta_95pct": round(_b_ai * 100, 2),
                                    "meets_95_95": _b_ai >= 0.95,
                                })

                    # Pearson cross-check limited to the visible selected FOMs.
                    _corr_cols = [c for c in [_primary_out_ai, _con_out_ai] if c]
                    _corr_rows = []
                    for _out_ai in _corr_cols:
                        for _iv in _in_cols_ai:
                            _common = _df_ok_ai[[_iv, _out_ai]].apply(
                                pd.to_numeric, errors="coerce").dropna()
                            if len(_common) > 2:
                                _r = float(np.corrcoef(_common[_iv], _common[_out_ai])[0, 1])
                                if np.isfinite(_r):
                                    _rmin_for_out = _con_r_min_ai if _out_ai == _con_out_ai else _r_min_ai
                                    _corr_rows.append({
                                        "output": _out_ai,
                                        "display_label": (_CON_DISPLAY_LABELS.get(_out_ai, _out_ai.replace("_", " "))
                                                          if _out_ai in _CON_SCALAR_KEYS_AI
                                                          else _scalar_display_label(_out_ai)),
                                        "input": _iv.replace("in_", ""),
                                        "r": round(_r, 6),
                                        "abs_r": round(abs(_r), 6),
                                        "student_t_min_abs_r": round(float(_rmin_for_out), 6),
                                        "threshold_status": _threshold_status_ai(_r, _rmin_for_out),
                                    })
                    _corr_rows = sorted(_corr_rows, key=lambda x: x["abs_r"], reverse=True)[:_corr_cap]
                    _samples_csv = ""
                    try:
                        if _sample_head_n > 0 and st.session_state.get("ua_samples") is not None:
                            _samples_csv = st.session_state.ua_samples.head(_sample_head_n).to_csv(index=False)
                    except Exception:
                        _samples_csv = ""
                    _has_con = len(_con_summary) > 0
                    _ai_progress.progress(35, text="Building uncertainty/regulatory-FOM prompt…")
                    _ai_status.update(label="Building uncertainty/regulatory-FOM prompt…", state="running")
                    _system = (
                        "You are FLARE's uncertainty-analysis narrative writer for nuclear safety analysis. "
                        "Write continuous, flowing prose — not bullet points, not numbered lists. "
                        + _ua_ai_unit_policy_text(_use_english) + " "
                        "Every response must read as a cohesive technical narrative in the style "
                        "of a safety analysis report: full sentences, logical transitions between ideas, "
                        "conclusions drawn from evidence. Balance statistical interpretation with accident-physics prose: "
                        "explain how sampled inputs change the accident progression and therefore drive the regulatory figures of merit. "
                        "Structure the narrative with a concise title heading at the top, then the following "
                        "## sections in order: Sample Adequacy, Accident Progression and Regulatory Figures of Merit, Output Distributions, Importance Analysis "
                        "(discuss the selected FLARE FOM and, when available, the selected FLARECON FOM; "
                        "use the supplied Student-t threshold, accepted rows, and terminal stopping row; "
                        "state whether the terminal candidate fails, barely passes, or comfortably passes the threshold; "
                        "discuss ΔStd Dev per parameter, fit degree, Adj R², and the convolution estimate √(Σ Var) versus the actual σ), "
                        "Ordered-Statistic Results (state the m-th order-statistic value and the Wilks β "
                        "for each m provided — explicitly call out whether the 95/95 criterion is met), "
                        "and ## Conclusions / GDC Support. "
                        "Evaluate regulatory FOMs broadly, not only 10 CFR 50.46 LOCA metrics. Discuss, as data permit, "
                        "PCT, ECR, hydrogen generation, DNBR/dryout or other SAFDL-style non-LOCA fuel-design limits, "
                        "RCS pressure/inventory/control response, containment pressure and combustible-gas behavior under 10 CFR 50.44, "
                        "and dose limits associated with 10 CFR 50.34(a)(1), 10 CFR 50.67, 10 CFR 100.11, and 10 CFR 100.21(c)(2). "
                        "If containment results are present, add a ## Containment Response section before Conclusions "
                        "noting peak pressure vs design if available, H₂ vs flammability limits (4 vol% LFL, 8 vol% detonable), "
                        "and which inputs drive the selected containment FOM. "
                        "For stepwise regression, the terminal_stopping_row is a diagnostic row, not an accepted contributor; "
                        "use it to explain why the regression stopped and whether additional variables are statistically screened out. "
                        "Conclusions must state whether the uncertainty results support or challenge applicable General Design Criteria functions: "
                        "reactivity control, RCPB protection, ECCS/core cooling, containment integrity, combustible-gas control, and dose control. "
                        "Keep the length consistent with the requested narrative detail and do not exceed the requested target length. "
                        "All numeric measures must be stated explicitly with their value and unit where a unit is supplied. "
                        "Do not invent data; use only the supplied statistics. "
                        "Wilks' method: the m-th largest of N independent samples is a one-sided "
                        "95th-percentile tolerance limit at confidence β(N,m,γ=0.95). "
                        "N=59 gives β≥95% for m=1; N=93 for m=1 two-sided; N=14 gives β≥50% for m=1."
                    )
                    # Scale the response budget with the Narrative detail slider.
                    # At the minimum setting, target an abstract-length narrative
                    # rather than a short report.  The token budget intentionally
                    # leaves headroom above the requested word count so the model
                    # can finish cleanly without encouraging unnecessary length.
                    _detail01 = float(np.clip(_ua_detail, 0.0, 1.0))
                    _detail_level = _ua_detail_level(_detail01)
                    _word_target = _ua_detail_word_target(_detail_level)
                    _word_min, _word_max = _ua_detail_word_range(_detail_level)
                    _detail_desc = _ua_detail_descriptor(_detail_level)

                    # Token budget follows the discrete target length rather than
                    # broad detail bands.  Approximate 1 word ~= 1.35 tokens and
                    # add headroom so the model can finish cleanly without making
                    # each low-end detent unnecessarily long.
                    # The word range is the length control.  The API token
                    # ceiling is intentionally generous so the model stops by
                    # instruction rather than by hard max_tokens exhaustion.
                    _max_tokens = int(min(12000, max(2500, round(_word_max * 2.8 + 1000))))
                    _api_timeout = int(120 + _detail_level * 300)
                    _length = (
                        f"{_detail_desc}, target about {_word_target} words "
                        f"(acceptable range {_word_min}–{_word_max} words) for slider detail "
                        f"{_detail_level:.2f}. Do not exceed {_word_max} words."
                    )
                    _con_block = (
                        f"\n## FLARECON containment scalar outcomes:\n"
                        f"{json.dumps(_con_summary, indent=2)}\n"
                        f"H₂ flammability reference: LFL=4 vol%, detonable≥8 vol%\n"
                    ) if _has_con else ""
                    _imp_block = ""
                    if _flare_importance_ai:
                        _imp_block += (
                            f"\n## FLARE stepwise importance analysis for selected FOM "
                            f"({_primary_out_label_ai or _primary_out_ai}):\n"
                            f"{json.dumps(_flare_importance_ai, indent=2)}\n"
                        )
                    if _con_importance_ai:
                        _imp_block += (
                            f"\n## FLARECON stepwise importance analysis for selected containment FOM "
                            f"({_con_out_label_ai or _con_out_ai}):\n"
                            f"{json.dumps(_con_importance_ai, indent=2)}\n"
                        )
                    _wilks_block = ""
                    if _wilks_ai_rows:
                        _wilks_block = (
                            f"\n## Wilks order-statistic results for primary output ({_primary_out_ai}):\n"
                            f"{json.dumps(_wilks_ai_rows, indent=2)}\n"
                        )
                    _user = f"""
Write a {_length} uncertainty-analysis narrative for the following FLARE UA run.
{_ua_ai_unit_policy_text(_use_english)}
Case: {st.session_state.get('ua_case') or selected}
Samples requested / successful: {len(_df_ai)} total rows; {len(_df_ok_ai)} successful rows.
Input variables sampled: {[c.replace('in_', '') for c in _in_cols_ai]}

## FLARE RCS/core output summary statistics:
{json.dumps(_flare_summary, indent=2)}
{_con_block}
{_imp_block}
{_wilks_block}
## Regulatory FOM framework for narrative interpretation:
- LOCA fuel-cladding criteria: 10 CFR 50.46 PCT, local oxidation/ECR, and hydrogen generation.
- Non-LOCA fuel-design criteria: specified acceptable fuel design limits, including DNBR/dryout margin where applicable, fuel temperature, RCS pressure response, inventory, and reactivity/control response.
- Containment / combustible-gas criteria: 10 CFR 50.44 hydrogen control, flammability, and containment pressure/temperature response.
- Dose criteria: 10 CFR 50.34(a)(1), 10 CFR 50.67, 10 CFR 100.11, and 10 CFR 100.21(c)(2), as applicable to source-term/dose outputs.
- Conclusions should connect the probabilistic FOM results to applicable GDC safety functions rather than only restating distribution statistics.

## Pearson correlations by absolute value for the selected FOMs (cross-check for importance analysis):
{json.dumps(_corr_rows, indent=2)}

## First rows of sample matrix (for context; omitted at low narrative detail):
{_samples_csv if _samples_csv else "(omitted)"}
"""
                    def _ua_ai_section_plan(detail_level: float, has_con: bool):
                        """Return ordered UA narrative sections and relative word weights.

                        The section plan is deterministic so the UI can show useful
                        progress before the first long API call.  Each section is
                        then generated by a separate API call and displayed as it is
                        completed.
                        """
                        d = float(np.clip(detail_level, 0.0, 1.0))
                        if d < 0.15:
                            sections = [
                                {
                                    "title": "Abstract-Level UA Summary",
                                    "focus": (
                                        "Summarize the selected FLARE and FLARECON figures of merit, the main distribution/order-statistic result, "
                                        "and the principal uncertainty driver without detailed statistical exposition."
                                    ),
                                    "weight": 0.55,
                                },
                                {
                                    "title": "Regulatory FOM and GDC Bottom Line",
                                    "focus": (
                                        "State how the uncertainty results support or challenge the applicable regulatory figures of merit and GDC functions."
                                    ),
                                    "weight": 0.45,
                                },
                            ]
                        else:
                            sections = [
                                {
                                    "title": "UA Overview and Regulatory FOM Context",
                                    "focus": (
                                        "Identify the selected FLARE and FLARECON figures of merit, sample count, sampled input space, and regulatory criteria context."
                                    ),
                                    "weight": 0.18,
                                },
                                {
                                    "title": "FLARE RCS/Core FOM Distribution and Ordered Statistics",
                                    "focus": (
                                        "Interpret the selected FLARE FOM distribution, high-order values, Wilks confidence, and safety meaning."
                                    ),
                                    "weight": 0.24,
                                },
                                {
                                    "title": "Importance Analysis and Uncertainty Drivers",
                                    "focus": (
                                        "Explain accepted stepwise contributors, the terminal stopping row, Pearson cross-checks, and how sampled inputs drive the FOM."
                                    ),
                                    "weight": 0.26,
                                },
                                {
                                    "title": "Conclusions / GDC Support",
                                    "focus": (
                                        "Conclude whether the uncertainty results support or challenge applicable safety functions and regulatory FOMs."
                                    ),
                                    "weight": 0.20,
                                },
                            ]
                            if has_con and d >= 0.20:
                                sections.insert(-1, {
                                    "title": "FLARECON Containment FOM Distribution and Drivers",
                                    "focus": (
                                        "Interpret the selected containment FOM, containment peak values, hydrogen/flammability context, and containment importance drivers."
                                    ),
                                    "weight": 0.18,
                                })
                            if d >= 0.65:
                                sections.insert(-1, {
                                    "title": "Statistical Confidence and Review Observations",
                                    "focus": (
                                        "Discuss Wilks/order-statistic confidence, convergence limitations, and review caveats from the supplied data only."
                                    ),
                                    "weight": 0.12,
                                })
                        total = sum(s["weight"] for s in sections) or 1.0
                        for s in sections:
                            s["weight"] = s["weight"] / total
                        return sections

                    def _ua_section_prompt(section_title, section_focus, section_words, completed_titles):
                        section_hi = max(90, int(round(section_words * 1.20)))
                        section_lo = max(60, int(round(section_words * 0.80)))
                        prior = "; ".join(completed_titles) if completed_titles else "none"
                        return f"""
Generate one section of a larger FLARE uncertainty-analysis narrative.
Begin with exactly this heading on its own line: ### {section_title}
Then insert one blank line before the body text.
Write only this section. Do not write the other planned sections.
Section target: about {section_words} words. Hard maximum: {section_hi} words. Stop at or before the hard maximum even if additional details could be discussed. Body text must be ordinary paragraph text only; do not put body text on the same line as the heading; do not use markdown heading markers or full-line bold for body paragraphs.
Completed prior section titles: {prior}. Avoid repeating prior-section material except where needed for continuity.
This section focus: {section_focus}
Overall narrative setting: {_detail_desc}, detent {_detail_level:.2f}, total target about {_word_target} words and total hard maximum {_word_max} words.
{_ua_ai_unit_policy_text(_use_english)}

Use the same regulatory priorities as the full narrative: connect sampled inputs to accident progression, then to regulatory figures of merit, then to GDC support. The terminal_stopping_row is diagnostic, not an accepted contributor.
Use only the supplied data. If a data category is absent, say only that it was not available and move on.

Full data payload for context:
{_user}
""".strip()

                    try:
                        _ai_status.update(label="Planning narrative sections from requested detail level…", state="running")
                        _sections = _ua_ai_section_plan(_detail_level, _has_con)
                        _n_sections = max(1, len(_sections))
                        _ai_progress.progress(8, text=f"Planned {_n_sections} narrative sections.")
                        _time.sleep(0.05)
                        _live_box = _ua_ai_live_box
                        _section_texts = []
                        _completed_titles = []
                        st.session_state.ua_ai_narrative = ""

                        for _idx, _section in enumerate(_sections, start=1):
                            _pct_start = int(10 + ((_idx - 1) / _n_sections) * 85)
                            _pct_done = int(10 + (_idx / _n_sections) * 85)
                            _sec_words = max(80, int(round(_word_target * float(_section.get("weight", 1.0 / _n_sections)))))
                            # Keep section calls aligned with the selected total narrative detail.
                            # Use a modest per-section hard maximum and a relatively tight token budget;
                            # the post-call clamp below is the final backstop against length drift.
                            _sec_hi = max(90, int(round(_sec_words * 1.10)))
                            _sec_tokens = int(min(3500, max(500, round(_sec_hi * 1.55 + 250))))
                            _sec_timeout = int(min(_api_timeout, max(90, round(60 + _sec_hi * 0.18))))
                            _title = _section["title"]

                            _ai_progress.progress(_pct_start, text=f"Generating section {_idx}/{_n_sections}: {_title}…")
                            _ai_status.update(label=f"Generating section {_idx}/{_n_sections}: {_title}…", state="running")
                            _time.sleep(0.05)
                            _section_user = _ua_section_prompt(
                                _title,
                                _section.get("focus", "Write the requested uncertainty-analysis section."),
                                _sec_words,
                                _completed_titles,
                            )
                            _section_text, _section_stop = _anthropic_text(
                                _system, _section_user, max_tokens=_sec_tokens, timeout=_sec_timeout)
                            if _section_stop == "max_tokens":
                                _section_text += (
                                    f"\n\n⚠️ Section may be incomplete because it reached its per-section budget ({_sec_tokens} tokens)."
                                )
                            elif _section_stop and _section_stop not in {"end_turn", "stop_sequence"}:
                                _section_text += f"\n\n⚠️ Section stopped with API stop_reason: {_section_stop}."
                            _section_text = _ua_trim_to_word_limit(_section_text.strip(), _sec_hi)
                            _section_text = _ua_normalize_ai_section_markdown(_section_text, _title)

                            _section_texts.append(_section_text.strip())
                            _completed_titles.append(_title)
                            _partial = "\n\n".join(_section_texts).strip()
                            st.session_state.ua_ai_narrative = _partial
                            _live_box.markdown(_partial)
                            _ai_progress.progress(_pct_done, text=f"Completed section {_idx}/{_n_sections}: {_title}.")
                            _time.sleep(0.05)

                        st.session_state.ua_ai_narrative = _ua_normalize_ai_section_markdown(
                            _ua_trim_to_word_limit("\n\n".join(_section_texts).strip(), _word_max)
                        )
                        _live_box.empty()
                        _ai_progress.progress(100, text="AI narrative complete.")
                        _ai_status.update(label="AI narrative complete.", state="complete")
                        _time.sleep(0.25)
                        st.rerun()
                    except Exception:
                        raise

                except Exception as _ai_e:
                    try:
                        _ai_progress.progress(100, text="AI narrative failed.")
                        _ai_status.update(label=f"AI narrative failed: {_ai_e}", state="error")
                    except Exception:
                        pass
                    st.error(f"AI narrative failed: {_ai_e}")

            if st.session_state.get("ua_ai_narrative"):
                st.markdown(st.session_state.ua_ai_narrative)
                st.download_button(
                    "⬇  Download narrative (Markdown)",
                    data=st.session_state.ua_ai_narrative.encode("utf-8"),
                    file_name=f"ua_{st.session_state.get('ua_case') or selected}_narrative.md",
                    mime="text/markdown",
                    key="ua_ai_download",
                )

# ── Samples tab ───────────────────────────────────────────────────────────────
with tab_samples:
    if st.session_state.ua_samples is not None:
        st.markdown("**Sample input values**")
        st.dataframe(st.session_state.ua_samples, width="stretch",
                     hide_index=True)
    else:
        st.info("No samples generated yet.")
