"""
flare_risk_worker.py — durable FLARE risk batch worker

Launched by flare_risk.py. Runs outside the Streamlit browser session so a
browser/WebSocket disconnect does not stop the risk calculation.

Important implementation note:
The child flare_sim.py process writes stdout/stderr directly to a per-case
console log. Do not use subprocess.PIPE without continuously draining it; on
Windows that can deadlock a verbose simulation when the pipe buffer fills.
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path

import pandas as pd

_DEFAULT_FREQ  = 1e-3
NEI_FREQ_AOO   = 1e-2
NEI_FREQ_DBE   = 1e-4
NEI_FREQ_SCREEN= 5e-7
NEI_DOSE_AOO   = 1.0
NEI_DOSE_DBE   = 25.0
NEI_DOSE_BDBE  = 1000.0


def classify(freq: float) -> str:
    if freq >= NEI_FREQ_AOO:    return "AOO"
    if freq >= NEI_FREQ_DBE:    return "DBE"
    if freq >= NEI_FREQ_SCREEN: return "BDBE"
    return "Screened"


def dose_limit(freq: float) -> float:
    if freq >= NEI_FREQ_AOO:    return NEI_DOSE_AOO
    if freq >= NEI_FREQ_DBE:    return NEI_DOSE_DBE
    return NEI_DOSE_BDBE


def read_json(path: Path, default):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default


def write_json_atomic(path: Path, payload) -> bool:
    """Write JSON without letting transient Windows/OneDrive locks kill the run.

    The original implementation used a fixed `risk_status.json.tmp` followed by
    `Path.replace()`.  On Windows, OneDrive/Defender/Streamlit can briefly hold
    either the tmp file or the destination file, causing `PermissionError` and
    terminating the worker even though the simulations are healthy.

    This version:
      * uses a unique tmp filename for each write, avoiding stale locked tmp files;
      * retries atomic replace with short backoff;
      * falls back to a direct write if replace remains locked;
      * returns False rather than raising, because status updates are diagnostic
        and must not stop the risk calculation.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=2)

    last_err = None
    for attempt in range(25):
        tmp = path.with_name(
            f".{path.name}.{os.getpid()}.{int(time.time() * 1000)}.{attempt}.tmp"
        )
        try:
            tmp.write_text(text, encoding="utf-8")
            os.replace(str(tmp), str(path))
            return True
        except PermissionError as e:
            last_err = e
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
            time.sleep(min(0.05 * (attempt + 1), 1.0))
        except OSError as e:
            last_err = e
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
            time.sleep(min(0.05 * (attempt + 1), 1.0))

    # Last-resort non-atomic write.  This may briefly expose a partially written
    # file, but the UI already tolerates JSON read failures and will retry on the
    # next refresh.  Keeping the worker alive is more important than perfect
    # atomicity for a progress/status file.
    for attempt in range(10):
        try:
            path.write_text(text, encoding="utf-8")
            return True
        except PermissionError as e:
            last_err = e
            time.sleep(min(0.1 * (attempt + 1), 1.0))
        except OSError as e:
            last_err = e
            time.sleep(min(0.1 * (attempt + 1), 1.0))

    try:
        warn = path.parent / "risk_status_write_warnings.log"
        with warn.open("a", encoding="utf-8") as f:
            f.write(f"{time.strftime('%Y-%m-%dT%H:%M:%S')} status write failed: {last_err}\n")
    except Exception:
        pass
    print(f"[risk-worker] WARNING: could not update {path.name}: {last_err}", flush=True)
    return False


def write_summary_csv(run_dir: Path, results: dict) -> None:
    rows = [
        {
            "Case": k,
            "Freq (/yr)": v.get("freq"),
            "Category": v.get("category"),
            "EAB TEDE Total (rem)": v.get("dose"),
            "Accident EAB TEDE (rem)": v.get("accident_dose", ""),
            "Iodine Spike EAB TEDE (rem)": v.get("iodine_spike_dose", ""),
            "Limit (rem)": v.get("limit"),
            "Status": v.get("status"),
            "Error": v.get("run_error") or "",
        }
        for k, v in sorted(results.items())
    ]
    pd.DataFrame(rows).to_csv(run_dir / "FLARE_risk_results.csv", index=False)


def _extract_eab_from_sheet(wb, sheet_name: str):
    """Return the EAB TEDE value from a workbook sheet, or None if absent.

    FLARE writes both the accident source-term dose and the optional iodine-spike
    dose as small summary tables with a row whose first cell is `EAB` and whose
    second cell is the TEDE in rem.  The risk consequence metric must include
    both contributions when both sheets are present.
    """
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if row and str(row[0]).strip() == "EAB":
            v = row[1] if len(row) > 1 else None
            if v is None:
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
    return None


def extract_eab_dose_components(out_xlsx: Path):
    """Return accident, iodine-spike, and total EAB TEDE values from output xlsx.

    The risk chart/table consequence is the total EAB TEDE:

        total = accident dose from `Dose` sheet
              + pre-existing coolant iodine-spike dose from `Iodine Spike` sheet

    Missing sheets are treated as zero contribution.  If neither contribution is
    present, total is returned as None so the caller can preserve the existing
    no-release behavior.
    """
    if not out_xlsx.exists():
        return None, None, None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(out_xlsx, read_only=True, data_only=True)
        try:
            accident = _extract_eab_from_sheet(wb, "Dose")
            iodine   = _extract_eab_from_sheet(wb, "Iodine Spike")
        finally:
            wb.close()
    except Exception:
        return None, None, None

    if accident is None and iodine is None:
        return None, None, None
    total = (accident or 0.0) + (iodine or 0.0)
    return accident, iodine, total


def extract_eab_dose(out_xlsx: Path):
    """Backward-compatible wrapper returning total EAB TEDE."""
    _accident, _iodine, total = extract_eab_dose_components(out_xlsx)
    return total


def move_outputs(work_dir: Path, run_dir: Path, case: str) -> None:
    for ext in ["_out.xlsx", "_out.csv", "_fail.csv", "_diag.csv"]:
        src = work_dir / f"{case}{ext}"
        if src.exists():
            dst = run_dir / src.name
            if dst.exists():
                dst.unlink()
            shutil.move(str(src), str(dst))
    for fig in list(work_dir.glob(f"{case}_*.png")) + \
               list(work_dir.glob(f"{case}_*.pdf")) + \
               list(work_dir.glob("figure_*.png")):
        try:
            dst = run_dir / fig.name
            if dst.exists():
                dst.unlink()
            shutil.move(str(fig), str(dst))
        except Exception:
            pass



def apply_source_term_override(input_xlsx: Path, case: str, override: str | None) -> None:
    """Inject a run-level source_term_model override into a copied input workbook.

    The override is written into the copied per-run workbook, not the user's
    source input deck. It is inserted immediately before the time-series table
    so normal last-assignment-wins parsing applies.
    """
    if not override:
        return
    allowed = {"thermal_failure", "licensing_auto"}
    if override not in allowed:
        raise ValueError(f"Unsupported source_term_model override: {override}")

    from openpyxl import load_workbook

    wb = load_workbook(input_xlsx)
    try:
        sheet_name = f"{case}_in"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        insert_row = None
        for row in ws.iter_rows(max_col=1):
            v = row[0].value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                insert_row = row[0].row
                break
            if isinstance(v, str) and v.strip().lower().startswith("time"):
                insert_row = row[0].row
                break
        if insert_row is None:
            insert_row = ws.max_row + 1

        ws.insert_rows(insert_row, amount=2)
        ws.cell(row=insert_row, column=1).value = "# Risk tool source-term model override"
        ws.cell(row=insert_row + 1, column=1).value = f'source_term_model = "{override}"'
        wb.save(input_xlsx)
    finally:
        wb.close()

def first_error_line_from_file(path: Path):
    try:
        for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
            if "Traceback" in line or "Error" in line or "error" in line:
                return line
    except Exception:
        return None
    return None


def file_contains_error(path: Path) -> bool:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        return ("Traceback" in text) or ("Error" in text) or ("error" in text)
    except Exception:
        return False


def abort_requested(run_dir: Path) -> bool:
    return (run_dir / "risk_abort_requested.json").exists()


def terminate_process_tree(proc: subprocess.Popen) -> None:
    if proc is None or proc.poll() is not None:
        return
    try:
        if sys.platform.startswith("win"):
            subprocess.run(
                ["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        else:
            proc.terminate()
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


def run_case_with_abort(cmd, work_dir: Path, timeout_s: int, env: dict,
                        run_dir: Path, case: str, update_status):
    """Run one flare_sim case while polling for UI abort requests.

    stdout/stderr are written directly to <case>_console.log so the child process
    cannot block on a full PIPE buffer and the UI can tail the log during the run.
    """
    start = time.time()
    console_log = run_dir / f"{case}_console.log"
    with open(console_log, "w", encoding="utf-8", errors="replace", buffering=1) as log:
        log.write(f"[risk-worker] Command: {' '.join(str(x) for x in cmd)}\n")
        log.write(f"[risk-worker] Work dir: {work_dir}\n")
        log.write(f"[risk-worker] Started: {time.strftime('%Y-%m-%dT%H:%M:%S')}\n")
        log.flush()

        proc = subprocess.Popen(
            cmd,
            cwd=str(work_dir),
            stdout=log,
            stderr=subprocess.STDOUT,
            stdin=subprocess.DEVNULL,
            env=env,
            text=True,
            bufsize=1,
        )

        last_status = 0.0
        while True:
            elapsed = time.time() - start
            if abort_requested(run_dir):
                log.write(f"\n[risk-worker] Abort requested at elapsed={elapsed:.1f} s.\n")
                log.flush()
                terminate_process_tree(proc)
                return proc.returncode if proc.returncode is not None else -999, "aborted", console_log

            rc = proc.poll()
            if rc is not None:
                log.write(f"\n[risk-worker] Finished with return code {rc} after {elapsed:.1f} s.\n")
                log.flush()
                return rc, None, console_log

            if timeout_s and elapsed > timeout_s:
                log.write(f"\n[risk-worker] Timeout after {timeout_s} s.\n")
                log.flush()
                terminate_process_tree(proc)
                return proc.returncode if proc.returncode is not None else -998, "timeout", console_log

            # Heartbeat for UI.
            now = time.time()
            if now - last_status >= 2.0:
                last_status = now
                update_status(
                    message=f"Running {case} — elapsed {elapsed:.0f} s",
                    current_case=case,
                    current_case_pid=proc.pid,
                    current_case_elapsed_s=round(elapsed, 1),
                    current_case_console_log=str(console_log),
                    abort_requested=abort_requested(run_dir),
                )
            time.sleep(0.5)


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python flare_risk_worker.py risk_worker_config.json", file=sys.stderr)
        return 2

    cfg_path = Path(sys.argv[1]).resolve()
    cfg = read_json(cfg_path, {})
    work_dir = Path(cfg.get("work_dir", cfg_path.parent)).resolve()
    run_dir = Path(cfg.get("run_dir", work_dir / f"risk_{time.strftime('%Y%m%d_%H%M%S')}")).resolve()
    case_entries = list(cfg.get("case_entries") or [])
    if case_entries:
        cases = [str(e.get("case")) for e in case_entries]
    else:
        cases = list(cfg.get("cases", []))
        case_entries = [{"case": c, "input_path": str(work_dir / f"{c}_in.xlsx")} for c in cases]
    freqs = {str(k): float(v) for k, v in dict(cfg.get("freqs", {})).items()}
    fast_mode = bool(cfg.get("fast_mode", True))
    source_term_override = cfg.get("source_term_override")
    if source_term_override not in (None, "thermal_failure", "licensing_auto"):
        raise ValueError(f"Unsupported source_term_override in config: {source_term_override}")
    timeout_s = int(cfg.get("timeout_s", 600))

    run_dir.mkdir(parents=True, exist_ok=True)
    status_path = run_dir / "risk_status.json"
    local_results_path = run_dir / "flare_risk_results.json"
    local_pra_path = run_dir / "flare_pra_table.csv"
    try:
        pd.DataFrame([{"CaseName": k, "Frequency": v}
                      for k, v in sorted(freqs.items())]).to_csv(local_pra_path, index=False)
    except Exception as _pra_err:
        log_warning(run_dir, f"Could not write run-local flare_pra_table.csv: {_pra_err}")

    # Prevent stale abort requests from a reused run folder from killing a new run.
    try:
        (run_dir / "risk_abort_requested.json").unlink(missing_ok=True)
    except Exception:
        pass

    results = {}
    failed_runs = 0
    total = len(cases)

    def update_status(**kw):
        payload = {
            "status": "running",
            "message": "Running…",
            "total_runs": total,
            "completed_runs": len(results),
            "failed_runs": failed_runs,
            "current_case": None,
            "current_case_pid": None,
            "current_case_elapsed_s": None,
            "current_case_console_log": None,
            "run_dir": str(run_dir),
            "worker_pid": os.getpid(),
            "abort_requested": abort_requested(run_dir),
            "source_term_override": source_term_override,
            "last_update": time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        payload.update(kw)
        write_json_atomic(status_path, payload)

    update_status(status="running", message="Risk worker started.", worker_pid=os.getpid(), abort_requested=False)
    print(f"[risk-worker] Started: {run_dir}", flush=True)

    env = {
        **os.environ,
        "PYTHONUTF8": "1",
        "PYTHONUNBUFFERED": "1",
        "MPLBACKEND": "Agg",
    }

    for idx, case_entry in enumerate(case_entries, start=1):
        case = str(case_entry.get("case"))
        input_path = Path(case_entry.get("input_path") or (work_dir / f"{case}_in.xlsx"))
        if abort_requested(run_dir):
            update_status(
                status="aborted",
                message=f"Risk run aborted before starting next case. {len(results)}/{total} cases processed.",
                current_case=None,
                completed_runs=len(results),
                failed_runs=failed_runs,
                abort_requested=True,
                worker_pid=os.getpid(),
            )
            print("[risk-worker] Abort requested before next case.", flush=True)
            return 130

        freq = float(freqs.get(case, _DEFAULT_FREQ))
        update_status(
            status="running",
            message=f"Launching {case} ({idx}/{total})…",
            current_case=case,
            completed_runs=len(results),
            failed_runs=failed_runs,
        )
        print(f"[risk-worker] Running {case} ({idx}/{total})", flush=True)

        try:
            # Copy the source input deck into the risk run folder and run flare_sim
            # with cwd=run_dir so no input/output files are created in the FLARE root.
            if not input_path.exists():
                raise FileNotFoundError(f"Input workbook not found: {input_path}")
            _copied_input = run_dir / f"{case}_in.xlsx"
            shutil.copy2(str(input_path), str(_copied_input))
            apply_source_term_override(_copied_input, case, source_term_override)
            cmd = [sys.executable, "-u", str(work_dir / "flare_sim.py"), case]
            if fast_mode:
                cmd.append("--no-figures")

            # The Risk workflow is dose-consequence focused and intentionally
            # bypasses the optional FLARECON containment-response calculation.
            env["FLARE_SKIP_FLARECON"] = "1"

            returncode, special, console_log = run_case_with_abort(
                cmd, run_dir, timeout_s, env, run_dir, case, update_status
            )

            # Outputs are written directly in run_dir. The old move step is left
            # harmless for compatibility but should find nothing in work_dir.
            move_outputs(work_dir, run_dir, case)

            if special == "aborted":
                update_status(
                    status="aborted",
                    message=f"Risk run aborted while running {case}. {len(results)}/{total} completed before abort.",
                    current_case=case,
                    completed_runs=len(results),
                    failed_runs=failed_runs,
                    abort_requested=True,
                    worker_pid=os.getpid(),
                    current_case_console_log=str(console_log),
                )
                print(f"[risk-worker] Abort requested during {case}.", flush=True)
                return 130

            if special == "timeout":
                raise subprocess.TimeoutExpired(cmd, timeout_s)

            accident_dose, iodine_spike_dose, dose = extract_eab_dose_components(run_dir / f"{case}_out.xlsx")
            if dose is None:
                dose = 0.0

            limit = dose_limit(freq)
            cat = classify(freq)
            status = "PASS" if dose > 0 and dose <= limit else ("FAIL" if dose > limit else "no release")

            output_exists = (run_dir / f"{case}_out.xlsx").exists() or (run_dir / f"{case}_out.csv").exists()
            has_error = file_contains_error(console_log) or returncode != 0 or not output_exists
            if has_error:
                failed_runs += 1
                if not output_exists and returncode == 0:
                    run_error = "Simulation finished without expected output file."
                else:
                    run_error = first_error_line_from_file(console_log) or f"Return code {returncode}"
            else:
                run_error = None

            results[case] = {
                "freq": freq,
                "source_term_override": source_term_override,
                "dose": dose,
                "accident_dose": accident_dose or 0.0,
                "iodine_spike_dose": iodine_spike_dose or 0.0,
                "category": cat,
                "limit": limit,
                "status": status if not has_error else "error",
                "error": has_error,
                "run_error": run_error,
            }

        except subprocess.TimeoutExpired:
            failed_runs += 1
            results[case] = {
                "freq": freq,
                "dose": None,
                "category": classify(freq),
                "limit": dose_limit(freq),
                "status": "timeout",
                "error": True,
                "run_error": f"Timeout after {timeout_s} s",
            }
        except Exception as e:
            failed_runs += 1
            results[case] = {
                "freq": freq,
                "dose": None,
                "category": classify(freq),
                "limit": dose_limit(freq),
                "status": "error",
                "error": True,
                "run_error": str(e),
            }

        write_json_atomic(local_results_path, results)
        write_summary_csv(run_dir, results)
        update_status(
            status="running",
            message=f"Completed {case} ({idx}/{total}).",
            current_case=case,
            completed_runs=len(results),
            failed_runs=failed_runs,
        )

    update_status(
        status="complete",
        message=f"Risk run complete. {len(results)}/{total} cases processed; {failed_runs} failed/error cases.",
        current_case=None,
        completed_runs=len(results),
        failed_runs=failed_runs,
    )
    print(f"[risk-worker] Complete: {len(results)}/{total}; failed={failed_runs}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
