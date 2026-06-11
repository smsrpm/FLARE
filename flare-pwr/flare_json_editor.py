"""
flare_json_editor.py - shared Streamlit catalogue editor utilities for FLARE
============================================================================

This module provides a form-based Runtime JSON catalogue editor for FLARE.  It
is intentionally dependency-light and safe to import from both flare_ui.py and
flare_ua.py.
"""

from __future__ import annotations

import importlib
import json
import re
import sys
from copy import deepcopy
from pathlib import Path
from typing import Any

import streamlit as st

_DIST_OPTIONS = ["uniform", "normal", "lognormal", "triangular"]

_ADDITIONAL_SECTION_LABEL = "Additional parameters"
_ADDITIONAL_SECTION_KEY = "_section_additional_parameters"


def _module_mtime_ns() -> int | None:
    """Return this module file's modification time with nanosecond precision."""
    try:
        return Path(__file__).stat().st_mtime_ns
    except Exception:
        return None


_MODULE_MTIME_NS = _module_mtime_ns()


def _maybe_reload_this_module() -> bool:
    """Reload flare_json_editor.py if the source file changed on disk.

    Streamlit reruns the active script on browser refresh, but imported helper
    modules can remain cached in sys.modules.  This development hook lets the
    already-imported editor notice that its own source file has changed, reload
    itself, and request a clean Streamlit rerun.  It is intentionally silent in
    production and only triggers when the file timestamp changes.
    """
    current_mtime = _module_mtime_ns()
    previous_mtime = globals().get("_MODULE_MTIME_NS")
    if current_mtime is None or previous_mtime is None or current_mtime == previous_mtime:
        return False

    module = sys.modules.get(__name__)
    if module is None:
        globals()["_MODULE_MTIME_NS"] = current_mtime
        return False

    try:
        importlib.invalidate_caches()
        importlib.reload(module)
        return True
    except Exception:
        globals()["_MODULE_MTIME_NS"] = current_mtime
        return False


def _inject_catalogue_editor_css() -> None:
    """Apply compact styling used by the catalogue editor."""
    st.markdown(
        """
        <style>
        /* The editor uses primary buttons only for per-row delete actions. */
        button[data-testid="stBaseButton-primary"] {
            background-color: #c62828 !important;
            border-color: #c62828 !important;
            color: #ffffff !important;
        }
        button[data-testid="stBaseButton-primary"] p {
            color: #ffffff !important;
        }
        button[data-testid="stBaseButton-primary"]:hover {
            background-color: #a51f1f !important;
            border-color: #a51f1f !important;
            color: #ffffff !important;
        }
        button[data-testid="stBaseButton-primary"]:focus {
            box-shadow: 0 0 0 0.2rem rgba(198, 40, 40, 0.25) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )




def _inject_editor_dialog_css() -> None:
    """Style the JSON editor dialog and hide Streamlit's native X close control.

    Streamlit does not provide a server-side callback for the built-in dialog X.
    Hiding that control forces users to use the editor's own X/Close buttons,
    which clear the editor-open session state before rerunning the page.
    """
    st.markdown(
        """
        <style>
        /* Keep the JSON editor dialog on an explicit opaque light surface.
           Do not use inherit here: on some Streamlit themes/containers that
           can behave like transparency and let the page show through. */
        div[data-testid="stDialog"],
        div[data-testid="stDialog"] [data-testid="stDialogContent"],
        div[data-testid="stDialog"] [data-testid="stVerticalBlock"],
        div[role="dialog"],
        div[role="dialog"] [data-testid="stVerticalBlock"] {
            background-color: #ffffff !important;
            color: #111111 !important;
        }
        div[data-testid="stDialog"] p,
        div[data-testid="stDialog"] span,
        div[data-testid="stDialog"] label,
        div[data-testid="stDialog"] h1,
        div[data-testid="stDialog"] h2,
        div[data-testid="stDialog"] h3,
        div[data-testid="stDialog"] h4,
        div[data-testid="stDialog"] h5,
        div[data-testid="stDialog"] h6,
        div[role="dialog"] p,
        div[role="dialog"] span,
        div[role="dialog"] label,
        div[role="dialog"] h1,
        div[role="dialog"] h2,
        div[role="dialog"] h3,
        div[role="dialog"] h4,
        div[role="dialog"] h5,
        div[role="dialog"] h6 {
            color: #111111 !important;
        }
        div[data-testid="stDialog"] input,
        div[data-testid="stDialog"] textarea,
        div[role="dialog"] input,
        div[role="dialog"] textarea {
            background-color: #ffffff !important;
            color: #111111 !important;
        }

        /* Hide Streamlit's native dialog close control.  It does not invoke
           the server-side cleanup path used by the FLARE editor close buttons. */
        div[data-testid="stDialog"] button[aria-label="Close"],
        div[data-testid="stDialog"] button[title="Close"],
        div[data-testid="stDialog"] [data-testid="stModalCloseButton"],
        div[data-testid="stDialog"] [data-testid="stDialogCloseButton"],
        div[data-testid="stDialog"] [aria-label="Close dialog"],
        div[role="dialog"] button[aria-label="Close"],
        div[role="dialog"] button[title="Close"],
        div[role="dialog"] [aria-label="Close dialog"] {
            display: none !important;
            visibility: hidden !important;
            pointer-events: none !important;
        }

        /* Reduce the vertical gap between the native dialog title/header and
           the FLARE-rendered editor content.  These selectors intentionally
           target only the dialog container so normal page layout is unaffected. */
        div[data-testid="stDialog"] [data-testid="stVerticalBlock"],
        div[role="dialog"] [data-testid="stVerticalBlock"] {
            gap: 0.35rem !important;
        }
        div[data-testid="stDialog"] [data-testid="stVerticalBlock"] > div:first-child,
        div[role="dialog"] [data-testid="stVerticalBlock"] > div:first-child {
            margin-top: -0.75rem !important;
            padding-top: 0 !important;
        }
        div[data-testid="stDialog"] [data-testid="stCaptionContainer"],
        div[role="dialog"] [data-testid="stCaptionContainer"] {
            padding-top: 0 !important;
            margin-top: 0 !important;
        }

        /* Make the FLARE-controlled upper-right close button visibly separate
           from the hidden native close control. */
        div[data-testid="stDialog"] div[data-testid="column"]:has(button[kind="secondary"] p:only-child),
        div[role="dialog"] div[data-testid="column"]:has(button[kind="secondary"] p:only-child) {
            text-align: right;
        }
        div[data-testid="stDialog"] button:has(p:only-child):hover,
        div[role="dialog"] button:has(p:only-child):hover {
            border-color: #c62828 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )



def _inject_top_close_button_css(state_prefix: str) -> None:
    """Style the FLARE-controlled top-right close button as a compact square."""
    # Streamlit exposes widget-key wrappers as st-key-<key> on current releases.
    # The wrapper is targeted so this styling applies only to the editor's
    # application-controlled upper-right close button.
    key = _state_key(state_prefix, "top_close")
    key_class = "st-key-" + re.sub(r"[^A-Za-z0-9_-]", "-", key)
    st.markdown(
        f"""
        <style>
        .{key_class} {{
            display: flex !important;
            justify-content: flex-end !important;
            align-items: flex-start !important;
            margin-top: -0.25rem !important;
        }}
        .{key_class} button {{
            width: 2.15rem !important;
            height: 2.15rem !important;
            min-width: 2.15rem !important;
            min-height: 2.15rem !important;
            max-width: 2.15rem !important;
            max-height: 2.15rem !important;
            padding: 0 !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            line-height: 1 !important;
            font-size: 1.05rem !important;
            font-weight: 700 !important;
            border-radius: 0.25rem !important;
            margin-left: auto !important;
        }}
        .{key_class} button p {{
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1 !important;
        }}
        .{key_class} button:hover {{
            background-color: #c62828 !important;
            border-color: #c62828 !important;
            color: #ffffff !important;
        }}
        .{key_class} button:hover p {{
            color: #ffffff !important;
        }}
        .{key_class} button:focus {{
            box-shadow: 0 0 0 0.2rem rgba(198, 40, 40, 0.25) !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

def _catalogue_entry_columns():
    """Return columns for a catalogue row, with delete aligned to the row header."""
    try:
        return st.columns([7.0, 1.15], vertical_alignment="center")
    except TypeError:
        return st.columns([7.0, 1.15])


def _is_section_label(text: Any, label: str = _ADDITIONAL_SECTION_LABEL) -> bool:
    return str(text or "").strip().casefold() == label.casefold()


def _pwr_note_text(entry: Any) -> str:
    if isinstance(entry, dict) and "key" not in entry:
        return str(entry.get("_note") or entry.get("note") or "").strip()
    return ""


def _ensure_additional_section(obj: Any, editor_kind: str) -> Any:
    """Ensure the editor has an Additional parameters section marker.

    The marker is appended after the existing catalogue entries on first load.
    Newly added parameters are then appended after this marker and remain in that
    section unless the user reorders them.
    """
    if editor_kind == "pwr_params" and isinstance(obj, list):
        if not any(_is_section_label(_pwr_note_text(e)) for e in obj):
            obj.append({"_note": _ADDITIONAL_SECTION_LABEL})
    elif editor_kind == "ua_variables" and isinstance(obj, dict):
        if _ADDITIONAL_SECTION_KEY not in obj:
            obj[_ADDITIONAL_SECTION_KEY] = {"_note": _ADDITIONAL_SECTION_LABEL}
    return obj


def _move_list_item(seq: list[Any], index: int, delta: int) -> bool:
    new_index = index + delta
    if new_index < 0 or new_index >= len(seq):
        return False
    seq[index], seq[new_index] = seq[new_index], seq[index]
    return True


def _move_dict_key(obj: dict[str, Any], key: str, delta: int) -> bool:
    keys = list(obj.keys())
    if key not in keys:
        return False
    index = keys.index(key)
    new_index = index + delta
    if new_index < 0 or new_index >= len(keys):
        return False
    keys[index], keys[new_index] = keys[new_index], keys[index]
    reordered = {k: obj[k] for k in keys}
    obj.clear()
    obj.update(reordered)
    return True


def _move_list_item_relative(seq: list[Any], move_index: int, ref_index: int, where: str) -> bool:
    """Move one list item before or after another list item."""
    if move_index == ref_index:
        return False
    if move_index < 0 or move_index >= len(seq) or ref_index < 0 or ref_index >= len(seq):
        return False
    item = seq.pop(move_index)
    if move_index < ref_index:
        ref_index -= 1
    insert_at = ref_index if where == "before" else ref_index + 1
    insert_at = max(0, min(insert_at, len(seq)))
    seq.insert(insert_at, item)
    return True


def _move_dict_key_relative(obj: dict[str, Any], move_key: str, ref_key: str, where: str) -> bool:
    """Move one dictionary key before or after another key, preserving insertion order."""
    if move_key == ref_key or move_key not in obj or ref_key not in obj:
        return False
    keys = list(obj.keys())
    move_value = obj[move_key]
    keys.remove(move_key)
    ref_index = keys.index(ref_key)
    insert_at = ref_index if where == "before" else ref_index + 1
    keys.insert(insert_at, move_key)
    reordered: dict[str, Any] = {}
    for k in keys:
        reordered[k] = move_value if k == move_key else obj[k]
    obj.clear()
    obj.update(reordered)
    return True


def _pwr_entry_label(entry: dict[str, Any], index: int, *, include_section: bool = True) -> str:
    """Return a compact label for a PWR catalogue entry."""
    if "key" not in entry:
        note = _pwr_note_text(entry)
        return f"[{note}]" if include_section and note else f"Section {index + 1}"
    key = str(entry.get("key", ""))
    label = str(entry.get("label") or key)
    sheet = str(entry.get("sheet", "")).strip()
    prefix = "[FLARECON] " if sheet.casefold() == "flarecon" else "[FLARE] "
    return f"{prefix}{label} ({key})"


def _ua_entry_label(key: str, cfg: Any, *, include_section: bool = True) -> str:
    """Return a compact label for a UA catalogue entry."""
    if str(key).startswith("_"):
        note = ""
        if isinstance(cfg, dict):
            note = str(cfg.get("_note") or cfg.get("note") or "").strip()
        elif isinstance(cfg, str):
            note = cfg.strip()
        return f"[{note or key}]" if include_section else str(key)
    label = str(cfg.get("label", key)) if isinstance(cfg, dict) else str(key)
    sheet = str(cfg.get("sheet", "")).strip() if isinstance(cfg, dict) else ""
    prefix = "[FLARECON] " if sheet.casefold() == "flarecon" else "[FLARE] "
    return f"{prefix}{label} ({key})"




def _editor_spacer(height_rem: float) -> None:
    """Insert a small vertical spacer in the Streamlit editor layout."""
    st.markdown(f'<div style="height: {height_rem:.2f}rem;"></div>', unsafe_allow_html=True)


def _render_section_break(label: str) -> None:
    """Render a visually separated catalogue section label."""
    safe_label = str(label or "").strip()
    if not safe_label:
        return
    html = (
        '<div style="'
        'margin: 1.45rem 0 0.85rem 0;'
        'padding: 0.48rem 0.75rem;'
        'border-top: 1px solid #cfd6df;'
        'border-bottom: 1px solid #cfd6df;'
        'background: #f6f8fb;'
        'font-weight: 700;'
        'line-height: 1.25;'
        'color: #111111;'
        '">' + safe_label + '</div>'
    )
    st.markdown(html, unsafe_allow_html=True)


def _render_pwr_reorder_panel(state_prefix: str, obj: list[Any]) -> None:
    """Render a single-operation reorder panel for PWR list catalogues."""
    movable = [i for i, e in enumerate(obj) if isinstance(e, dict) and e.get("key")]
    refs = [i for i, e in enumerate(obj) if isinstance(e, dict) and (e.get("key") or _pwr_note_text(e))]
    if len(movable) < 2 or len(refs) < 2:
        return

    with st.expander("Reorder parameter", expanded=False):
        # Explicit two-row layout avoids Streamlit's selectbox labels crowding
        # the expander header and keeps the Move button aligned with the input row.
        _editor_spacer(1.15)
        label_c1, label_c2, label_c3, label_c4 = st.columns([2.4, 1.0, 2.4, 1.15])
        with label_c1:
            st.markdown("**Parameter to move**")
        with label_c2:
            st.markdown("**Place**")
        with label_c3:
            st.markdown("**Reference**")
        with label_c4:
            st.markdown("&nbsp;", unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns([2.4, 1.0, 2.4, 1.15])
        move_key = _state_key(state_prefix, "pwr_move_item")
        ref_key = _state_key(state_prefix, "pwr_ref_item")
        where_key = _state_key(state_prefix, "pwr_move_where")
        with c1:
            move_index = st.selectbox(
                "Parameter to move",
                options=movable,
                format_func=lambda i: _pwr_entry_label(obj[i], i, include_section=False),
                key=move_key,
                label_visibility="collapsed",
            )
        ref_options = [i for i in refs if i != move_index]
        if not ref_options:
            return
        with c2:
            where = st.selectbox(
                "Place",
                options=["before", "after"],
                key=where_key,
                label_visibility="collapsed",
            )
        with c3:
            ref_index = st.selectbox(
                "Reference",
                options=ref_options,
                format_func=lambda i: _pwr_entry_label(obj[i], i),
                key=ref_key,
                label_visibility="collapsed",
            )
        with c4:
            if st.button("Move", key=_state_key(state_prefix, "pwr_move_button"), use_container_width=True):
                if _move_list_item_relative(obj, int(move_index), int(ref_index), str(where)):
                    st.session_state[_state_key(state_prefix, "obj")] = obj
                    st.session_state[_state_key(state_prefix, "pwr_expanded_row")] = ""
                    st.rerun()


def _render_ua_reorder_panel(state_prefix: str, obj: dict[str, Any]) -> None:
    """Render a single-operation reorder panel for UA dictionary catalogues."""
    movable = [k for k, cfg in obj.items() if not str(k).startswith("_") and isinstance(cfg, dict)]
    refs = [k for k, cfg in obj.items() if (not str(k).startswith("_") and isinstance(cfg, dict)) or str(k).startswith("_")]
    if len(movable) < 2 or len(refs) < 2:
        return

    with st.expander("Reorder parameter", expanded=False):
        # Explicit two-row layout avoids Streamlit's selectbox labels crowding
        # the expander header and keeps the Move button aligned with the input row.
        _editor_spacer(1.15)
        label_c1, label_c2, label_c3, label_c4 = st.columns([2.4, 1.0, 2.4, 1.15])
        with label_c1:
            st.markdown("**Parameter to move**")
        with label_c2:
            st.markdown("**Place**")
        with label_c3:
            st.markdown("**Reference**")
        with label_c4:
            st.markdown("&nbsp;", unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns([2.4, 1.0, 2.4, 1.15])
        move_key = _state_key(state_prefix, "ua_move_item")
        ref_key = _state_key(state_prefix, "ua_ref_item")
        where_key = _state_key(state_prefix, "ua_move_where")
        with c1:
            move_name = st.selectbox(
                "Parameter to move",
                options=movable,
                format_func=lambda k: _ua_entry_label(str(k), obj.get(k), include_section=False),
                key=move_key,
                label_visibility="collapsed",
            )
        ref_options = [k for k in refs if k != move_name]
        if not ref_options:
            return
        with c2:
            where = st.selectbox(
                "Place",
                options=["before", "after"],
                key=where_key,
                label_visibility="collapsed",
            )
        with c3:
            ref_name = st.selectbox(
                "Reference",
                options=ref_options,
                format_func=lambda k: _ua_entry_label(str(k), obj.get(k)),
                key=ref_key,
                label_visibility="collapsed",
            )
        with c4:
            if st.button("Move", key=_state_key(state_prefix, "ua_move_button"), use_container_width=True):
                if _move_dict_key_relative(obj, str(move_name), str(ref_name), str(where)):
                    st.session_state[_state_key(state_prefix, "obj")] = obj
                    st.session_state[_state_key(state_prefix, "ua_expanded_row")] = ""
                    st.rerun()


def _state_key(prefix: str, suffix: str) -> str:
    return f"{prefix}__json_editor__{suffix}"


def _read_json_obj(path: Path, expected_top_level: str | None = None) -> Any:
    """Read a JSON file and return a valid top-level object."""
    raw = path.read_text(encoding="utf-8") if path.exists() else ""
    if not raw.strip():
        return [] if expected_top_level == "list" else {}
    obj = json.loads(raw)
    _validate_json_obj(obj, expected_top_level)
    return obj


def _validate_json_obj(obj: Any, expected_top_level: str | None = None) -> None:
    """Validate a JSON object and optionally enforce the top-level type."""
    if expected_top_level == "list" and not isinstance(obj, list):
        raise ValueError("The top-level JSON object must be a list.")
    if expected_top_level == "dict" and not isinstance(obj, dict):
        raise ValueError("The top-level JSON object must be a dictionary/object.")


def _pretty_json(obj: Any) -> str:
    return json.dumps(obj, indent=2, ensure_ascii=False) + "\n"


def _is_number(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _to_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _label_from_key(key: str) -> str:
    label = str(key).replace("_", " ").strip()
    return label[:1].upper() + label[1:] if label else "Parameter"


def _parse_command_value(text: str) -> tuple[str, Any] | None:
    """Parse a FLARE command-block assignment from a worksheet cell."""
    if not isinstance(text, str):
        return None
    if text.strip().startswith("#"):
        return None
    m = re.match(r"^\s*(\w+)\s*=\s*([+-]?\d+\.?\d*(?:[eE][+-]?\d+)?)", text)
    if m:
        try:
            return m.group(1), float(m.group(2))
        except ValueError:
            return None
    m = re.match(r'''^\s*(\w+)\s*=\s*["']([^"']+)["']''', text)
    if m:
        return m.group(1), m.group(2)
    return None


def discover_workbook_parameters(input_path: Path | str, case_name: str | None = None) -> list[dict[str, Any]]:
    """Return editable scalar parameters found in a FLARE input workbook.

    The search reads the first column of the main <case>_in worksheet and, when
    present, the FLARECON worksheet.  It stops at the first numeric time-history
    row in each sheet, mirroring the existing FLARE command-block convention.
    """
    path = Path(input_path)
    if not path.exists():
        return []

    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        candidate_sheets: list[str] = []
        if case_name:
            main = f"{case_name}_in"
            if main in wb.sheetnames:
                candidate_sheets.append(main)
        for name in wb.sheetnames:
            if name.strip().casefold() == "flarecon" and name not in candidate_sheets:
                candidate_sheets.append(name)
        if not candidate_sheets and wb.sheetnames:
            candidate_sheets.append(wb.sheetnames[0])

        for sheet_name in candidate_sheets:
            ws = wb[sheet_name]
            sheet_tag = "FLARECON" if sheet_name.strip().casefold() == "flarecon" else ""
            for i, row in enumerate(ws.iter_rows(max_col=1, values_only=True), 1):
                v = row[0]
                if _is_number(v):
                    break
                parsed = _parse_command_value(v) if isinstance(v, str) else None
                if parsed:
                    key, value = parsed
                    dedupe = (sheet_tag, key)
                    if dedupe not in seen:
                        seen.add(dedupe)
                        out.append({
                            "key": key,
                            "value": value,
                            "sheet": sheet_tag,
                            "label": _label_from_key(key),
                            "is_string": isinstance(value, str),
                        })
                if i > 1000:
                    break
    except Exception:
        return []
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass

    out.sort(key=lambda d: (str(d.get("sheet", "")), str(d.get("key", "")).lower()))
    return out


def _candidate_label(c: dict[str, Any]) -> str:
    sheet = str(c.get("sheet", "")).strip()
    key = str(c.get("key", ""))
    value = c.get("value", "")
    val = f" = {value:g}" if _is_number(value) else (f" = {value}" if value not in (None, "") else "")
    prefix = f"[{sheet}] " if sheet else "[FLARE] "
    return f"{prefix}{key}{val}"


def _ensure_editor_obj(path: Path, state_prefix: str, expected_top_level: str | None) -> None:
    obj_key = _state_key(state_prefix, "obj")
    mtime_key = _state_key(state_prefix, "mtime")
    error_key = _state_key(state_prefix, "error")
    current_mtime = path.stat().st_mtime if path.exists() else None
    if obj_key not in st.session_state or st.session_state.get(mtime_key) != current_mtime:
        try:
            st.session_state[obj_key] = _read_json_obj(path, expected_top_level)
            st.session_state[mtime_key] = current_mtime
            st.session_state[error_key] = ""
        except Exception as exc:
            st.session_state[obj_key] = [] if expected_top_level == "list" else {}
            st.session_state[mtime_key] = current_mtime
            st.session_state[error_key] = f"Existing file is not valid JSON: {exc}"


def _pwr_existing_keys(obj: Any) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    if isinstance(obj, list):
        for e in obj:
            if isinstance(e, dict) and e.get("key"):
                keys.add((str(e.get("sheet", "")).strip(), str(e["key"])))
    return keys


def _ua_existing_keys(obj: Any) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    if isinstance(obj, dict):
        for k, cfg in obj.items():
            if str(k).startswith("_"):
                continue
            sheet = str(cfg.get("sheet", "")).strip() if isinstance(cfg, dict) else ""
            keys.add((sheet, str(k)))
    return keys


def _render_add_parameter(*, obj: Any, state_prefix: str, editor_kind: str, candidates: list[dict[str, Any]]) -> None:
    """Render the add-parameter selector and insert a default entry when requested."""
    obj_key = _state_key(state_prefix, "obj")
    if editor_kind == "pwr_params":
        existing = _pwr_existing_keys(obj)
    else:
        existing = _ua_existing_keys(obj)

    available = [c for c in candidates if (str(c.get("sheet", "")).strip(), str(c.get("key", ""))) not in existing]
    st.markdown("##### Add parameter")
    if not available:
        st.caption("All parameters found in the selected input workbook are already listed in this catalogue.")
        return

    source_values = sorted({"FLARECON" if str(c.get("sheet", "")).strip().casefold() == "flarecon" else "FLARE" for c in available})
    if len(source_values) > 1:
        source = st.selectbox(
            "Parameter source",
            options=["FLARE", "FLARECON"],
            index=0,
            key=_state_key(state_prefix, "add_source"),
        )
        available = [
            c for c in available
            if ("FLARECON" if str(c.get("sheet", "")).strip().casefold() == "flarecon" else "FLARE") == source
        ]
    labels = [_candidate_label(c) for c in available]
    pick_key = _state_key(state_prefix, "add_pick")

    # Manual label row keeps the Add button aligned with the selectbox input
    # rather than aligned with Streamlit's selectbox label line.
    _editor_spacer(0.35)
    label_c1, label_c2 = st.columns([4.8, 0.9])
    with label_c1:
        st.markdown("**Parameter to add**")
    with label_c2:
        st.markdown("&nbsp;", unsafe_allow_html=True)

    c1, c2 = st.columns([4.8, 0.9])
    with c1:
        idx = st.selectbox(
            "Parameter to add",
            options=list(range(len(available))),
            format_func=lambda i: labels[i],
            key=pick_key,
            label_visibility="collapsed",
        )
    with c2:
        if st.button("Add parameter", key=_state_key(state_prefix, "add_button"), use_container_width=True):
            cand = available[int(idx)]
            key = str(cand.get("key", "")).strip()
            sheet = str(cand.get("sheet", "")).strip()
            value = cand.get("value")
            if editor_kind == "pwr_params":
                entry = {"key": key, "label": cand.get("label") or _label_from_key(key), "help": ""}
                if sheet:
                    entry["sheet"] = sheet
                if isinstance(value, str):
                    entry["options"] = [value]
                    entry["default"] = value
                if not isinstance(st.session_state[obj_key], list):
                    st.session_state[obj_key] = []
                st.session_state[obj_key].append(entry)
            else:
                base = _to_float(value, 0.0)
                if base == 0.0:
                    p1, p2 = 0.0, 1.0
                else:
                    lo = base * 0.9
                    hi = base * 1.1
                    p1, p2 = (min(lo, hi), max(lo, hi))
                entry = {
                    "label": cand.get("label") or _label_from_key(key),
                    "distribution": "uniform",
                    "base": base,
                    "p1": p1,
                    "p2": p2,
                    "help": "",
                }
                if sheet:
                    entry["sheet"] = sheet
                if not isinstance(st.session_state[obj_key], dict):
                    st.session_state[obj_key] = {}
                st.session_state[obj_key][key] = entry
            st.toast(f"Added {key}", icon="➕")
            st.rerun()


def _render_pwr_catalogue_editor(state_prefix: str) -> None:
    """Render form controls for Runtime/flare_ui_params.json."""
    obj_key = _state_key(state_prefix, "obj")
    expanded_key = _state_key(state_prefix, "pwr_expanded_row")
    obj = st.session_state.get(obj_key, [])
    if not isinstance(obj, list):
        st.error("This catalogue must be a JSON list. Use Advanced JSON to repair it.")
        return

    st.markdown("##### Parameter editor")
    st.caption("Entries in this list control which workbook parameters appear in the PWR Simulator's Edit parameters panel.")
    _render_pwr_reorder_panel(state_prefix, obj)

    remove_idx: int | None = None
    for i, entry in enumerate(obj):
        if not isinstance(entry, dict):
            continue
        if "key" not in entry:
            note = entry.get("_note") or entry.get("note")
            if note:
                _render_section_break(note)
            continue

        key = str(entry.get("key", ""))
        title = str(entry.get("label") or key)
        row_id = f"pwr::{i}::{key}"
        is_open = st.session_state.get(expanded_key) == row_id

        row_c1, row_c2 = _catalogue_entry_columns()
        with row_c1:
            arrow = "▾" if is_open else "▸"
            if st.button(
                f"{arrow}  {title}  ({key})",
                key=_state_key(state_prefix, f"pwr_toggle_{i}"),
                use_container_width=True
            ):
                st.session_state[expanded_key] = "" if is_open else row_id
                st.rerun()
        with row_c2:
            if st.button(
                "delete",
                type="primary",
                key=_state_key(state_prefix, f"pwr_delete_{i}"),
                use_container_width=True,
            ):
                remove_idx = i

        if is_open:
            with st.container(border=True):
                st.text_input("Parameter name", value=key, disabled=True, key=_state_key(state_prefix, f"pwr_key_{i}"))
                entry["label"] = st.text_input("Display label", value=str(entry.get("label", key)), key=_state_key(state_prefix, f"pwr_label_{i}"))
                sheet_val = str(entry.get("sheet", "")).strip()
                sheet_choice = st.selectbox(
                    "Worksheet",
                    options=["", "FLARECON"],
                    index=1 if sheet_val.strip().casefold() == "flarecon" else 0,
                    format_func=lambda v: "Main case sheet" if v == "" else v,
                    key=_state_key(state_prefix, f"pwr_sheet_{i}"),
                )
                if sheet_choice:
                    entry["sheet"] = sheet_choice
                else:
                    entry.pop("sheet", None)

                use_options = st.checkbox(
                    "Use a dropdown list for this parameter",
                    value=isinstance(entry.get("options"), list),
                    key=_state_key(state_prefix, f"pwr_use_options_{i}"),
                )
                if use_options:
                    opt_text = ", ".join(str(x) for x in entry.get("options", []))
                    opt_text = st.text_input(
                        "Dropdown options, comma separated",
                        value=opt_text,
                        key=_state_key(state_prefix, f"pwr_options_{i}"),
                    )
                    options = [x.strip() for x in opt_text.split(",") if x.strip()]
                    entry["options"] = options
                    default = str(entry.get("default", options[0] if options else ""))
                    entry["default"] = st.text_input("Default option", value=default, key=_state_key(state_prefix, f"pwr_default_{i}"))
                else:
                    entry.pop("options", None)
                    entry.pop("default", None)

                entry["help"] = st.text_area(
                    "Help text",
                    value=str(entry.get("help", "")),
                    height=95,
                    key=_state_key(state_prefix, f"pwr_help_{i}"),
                )

    if remove_idx is not None:
        deleted_key = str(obj[remove_idx].get("key", "parameter")) if isinstance(obj[remove_idx], dict) else "parameter"
        del obj[remove_idx]
        st.session_state[obj_key] = obj
        st.session_state[expanded_key] = ""
        st.toast(f"Deleted {deleted_key}", icon="🗑️")
        st.rerun()


def _render_ua_catalogue_editor(state_prefix: str) -> None:
    """Render form controls for Runtime/flare_ua_variables.json."""
    obj_key = _state_key(state_prefix, "obj")
    expanded_key = _state_key(state_prefix, "ua_expanded_row")
    obj = st.session_state.get(obj_key, {})
    if not isinstance(obj, dict):
        st.error("This catalogue must be a JSON object. Use Advanced JSON to repair it.")
        return

    st.markdown("##### Parameter editor")
    st.caption("Entries in this catalogue control the parameters available in the UA distribution panel.")
    _render_ua_reorder_panel(state_prefix, obj)

    remove_key: str | None = None
    for var in list(obj.keys()):
        cfg = obj.get(var)
        if str(var).startswith("_"):
            note = ""
            if isinstance(cfg, dict):
                note = str(cfg.get("_note") or cfg.get("note") or "").strip()
            elif isinstance(cfg, str):
                note = cfg.strip()
            if note:
                _render_section_break(note)
            continue
        if not isinstance(cfg, dict):
            continue
        label = str(cfg.get("label", var))
        row_id = f"ua::{var}"
        is_open = st.session_state.get(expanded_key) == row_id

        row_c1, row_c2 = _catalogue_entry_columns()
        with row_c1:
            arrow = "▾" if is_open else "▸"
            if st.button(
                f"{arrow}  {label}  ({var})",
                key=_state_key(state_prefix, f"ua_toggle_{var}"),
                use_container_width=True
            ):
                st.session_state[expanded_key] = "" if is_open else row_id
                st.rerun()
        with row_c2:
            if st.button(
                "delete",
                type="primary",
                key=_state_key(state_prefix, f"ua_delete_{var}"),
                use_container_width=True,
            ):
                remove_key = str(var)

        if is_open:
            with st.container(border=True):
                st.text_input("Parameter name", value=str(var), disabled=True, key=_state_key(state_prefix, f"ua_key_{var}"))
                cfg["label"] = st.text_input("Display label", value=label, key=_state_key(state_prefix, f"ua_label_{var}"))
                sheet_val = str(cfg.get("sheet", "")).strip()
                sheet_choice = st.selectbox(
                    "Worksheet",
                    options=["", "FLARECON"],
                    index=1 if sheet_val.strip().casefold() == "flarecon" else 0,
                    format_func=lambda v: "Main case sheet" if v == "" else v,
                    key=_state_key(state_prefix, f"ua_sheet_{var}"),
                )
                if sheet_choice:
                    cfg["sheet"] = sheet_choice
                else:
                    cfg.pop("sheet", None)

                dist = str(cfg.get("distribution", "uniform")).lower()
                if dist not in _DIST_OPTIONS:
                    dist = "uniform"
                cfg["distribution"] = st.selectbox(
                    "Distribution",
                    options=_DIST_OPTIONS,
                    index=_DIST_OPTIONS.index(dist),
                    key=_state_key(state_prefix, f"ua_dist_{var}"),
                )
                c1, c2, c3 = st.columns(3)
                with c1:
                    cfg["base"] = st.number_input("Base", value=_to_float(cfg.get("base"), 0.0), format="%.6g", key=_state_key(state_prefix, f"ua_base_{var}"))
                with c2:
                    cfg["p1"] = st.number_input("p1", value=_to_float(cfg.get("p1"), 0.0), format="%.6g", key=_state_key(state_prefix, f"ua_p1_{var}"))
                with c3:
                    cfg["p2"] = st.number_input("p2", value=_to_float(cfg.get("p2"), 0.0), format="%.6g", key=_state_key(state_prefix, f"ua_p2_{var}"))
                st.caption("For uniform/triangular: p1 = lower bound and p2 = upper bound. For normal: p1 = mean and p2 = standard deviation. For lognormal: p1 = ln(mean) and p2 = ln(std).")
                cfg["help"] = st.text_area(
                    "Help text",
                    value=str(cfg.get("help", "")),
                    height=95,
                    key=_state_key(state_prefix, f"ua_help_{var}"),
                )

    if remove_key is not None:
        obj.pop(remove_key, None)
        st.session_state[obj_key] = obj
        st.session_state[expanded_key] = ""
        st.toast(f"Deleted {remove_key}", icon="🗑️")
        st.rerun()

def _validate_catalogue_before_save(obj: Any, editor_kind: str, expected_top_level: str | None) -> None:
    _validate_json_obj(obj, expected_top_level)
    if editor_kind == "pwr_params":
        seen = set()
        for i, entry in enumerate(obj):
            if not isinstance(entry, dict):
                raise ValueError(f"Entry {i + 1} must be an object.")
            if "key" not in entry:
                continue
            key = str(entry.get("key", "")).strip()
            if not key:
                raise ValueError(f"Entry {i + 1} has a blank key.")
            sheet = str(entry.get("sheet", "")).strip()
            dedupe = (sheet, key)
            if dedupe in seen:
                raise ValueError(f"Duplicate parameter '{key}' in worksheet '{sheet or 'main'}'.")
            seen.add(dedupe)
            options = entry.get("options")
            if options is not None and (not isinstance(options, list) or not all(isinstance(x, str) for x in options)):
                raise ValueError(f"Parameter '{key}' options must be a list of strings.")
    elif editor_kind == "ua_variables":
        for var, cfg in obj.items():
            if str(var).startswith("_"):
                continue
            if not isinstance(cfg, dict):
                raise ValueError(f"Parameter '{var}' must be an object.")
            dist = str(cfg.get("distribution", "")).lower()
            if dist not in _DIST_OPTIONS:
                raise ValueError(f"Parameter '{var}' has invalid distribution '{dist}'.")
            for f in ("base", "p1", "p2"):
                _to_float(cfg.get(f))
                if cfg.get(f) is None:
                    raise ValueError(f"Parameter '{var}' must define numeric {f}.")


def _render_advanced_json(obj: Any, state_prefix: str, expected_top_level: str | None) -> None:
    """Optional raw JSON editor for repair or bulk edits."""
    obj_key = _state_key(state_prefix, "obj")
    raw_key = _state_key(state_prefix, "raw_json")
    err_key = _state_key(state_prefix, "raw_error")
    with st.expander("Advanced JSON", expanded=False):
        if st.button("Refresh JSON from form", key=_state_key(state_prefix, "raw_refresh")) or raw_key not in st.session_state:
            st.session_state[raw_key] = _pretty_json(obj)
            st.session_state[err_key] = ""
        st.text_area("Raw JSON", key=raw_key, height=260)
        c1, c2 = st.columns([1.0, 2.0])
        with c1:
            if st.button("Apply JSON to form", key=_state_key(state_prefix, "raw_apply")):
                try:
                    new_obj = json.loads(st.session_state.get(raw_key, ""))
                    _validate_json_obj(new_obj, expected_top_level)
                    st.session_state[obj_key] = new_obj
                    st.session_state[err_key] = ""
                    st.rerun()
                except Exception as exc:
                    st.session_state[err_key] = f"Could not apply JSON: {exc}"
        if st.session_state.get(err_key):
            st.error(st.session_state[err_key])


def _render_editor_body(
    *,
    path: Path,
    title: str,
    state_prefix: str,
    expected_top_level: str | None,
    editor_kind: str,
    candidates: list[dict[str, Any]] | None,
    compact_header: bool = False,
) -> None:
    """Render the catalogue editor and save/close controls."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    _ensure_editor_obj(path, state_prefix, expected_top_level)
    obj_key_for_section = _state_key(state_prefix, "obj")
    st.session_state[obj_key_for_section] = _ensure_additional_section(
        st.session_state.get(obj_key_for_section), editor_kind
    )
    _inject_catalogue_editor_css()

    obj_key = _state_key(state_prefix, "obj")
    error_key = _state_key(state_prefix, "error")
    open_key = _state_key(state_prefix, "open")
    mtime_key = _state_key(state_prefix, "mtime")

    header_c1, header_c2 = st.columns([12.0, 0.55])
    with header_c1:
        if not compact_header:
            st.markdown(f"**{title}**")
        st.caption(f"Editing `{path.name}` in `{path.parent}`")
    with header_c2:
        _inject_top_close_button_css(state_prefix)
        if st.button("X", key=_state_key(state_prefix, "top_close"), use_container_width=True):
            st.session_state[open_key] = False
            st.session_state[error_key] = ""
            st.rerun()

    if st.session_state.get(error_key):
        st.error(st.session_state[error_key])

    obj = st.session_state.get(obj_key)
    _render_add_parameter(
        obj=obj,
        state_prefix=state_prefix,
        editor_kind=editor_kind,
        candidates=candidates or [],
    )
    st.markdown("---")

    if editor_kind == "pwr_params":
        _render_pwr_catalogue_editor(state_prefix)
    elif editor_kind == "ua_variables":
        _render_ua_catalogue_editor(state_prefix)
    else:
        _render_advanced_json(obj, state_prefix, expected_top_level)

    _render_advanced_json(st.session_state.get(obj_key), state_prefix, expected_top_level)

    st.markdown("---")
    c1, c2, c3 = st.columns([1.2, 1.0, 2.0])
    with c1:
        if st.button("Save and Close", key=_state_key(state_prefix, "save")):
            try:
                obj = deepcopy(st.session_state.get(obj_key))
                _validate_catalogue_before_save(obj, editor_kind, expected_top_level)
                path.write_text(_pretty_json(obj), encoding="utf-8")
                st.session_state[mtime_key] = path.stat().st_mtime
                st.session_state[error_key] = ""
                st.session_state[open_key] = False
                st.toast(f"Saved {path.name}", icon="✅")
                st.rerun()
            except Exception as exc:
                st.session_state[error_key] = f"Save failed: {exc}"
                st.rerun()
    with c2:
        if st.button("Close", key=_state_key(state_prefix, "close")):
            st.session_state[open_key] = False
            st.session_state[error_key] = ""
            st.rerun()
    with c3:
        if st.button("Reload from file", key=_state_key(state_prefix, "reload")):
            try:
                st.session_state[obj_key] = _read_json_obj(path, expected_top_level)
                st.session_state[mtime_key] = path.stat().st_mtime if path.exists() else None
                st.session_state[error_key] = ""
            except Exception as exc:
                st.session_state[error_key] = f"Reload failed: {exc}"
            st.rerun()


def render_json_editor_button(
    *,
    path: Path,
    button_label: str,
    title: str,
    state_prefix: str,
    help_text: str = "Edit the Runtime JSON catalogue.",
    expected_top_level: str | None = None,
    editor_kind: str = "pwr_params",
    candidates: list[dict[str, Any]] | None = None,
) -> None:
    """Render a button that opens the form-based Runtime catalogue editor.

    The editor uses Streamlit's modal dialog when available, but it does not
    rely on Streamlit's native dialog close X.  That native X does not expose a
    server-side callback, so it can leave stale open-state in ``st.session_state``
    and reopen the editor after later sidebar events.  Instead, the native X is
    hidden and the editor supplies its own upper-right ``×`` button plus the
    normal Close button; both execute the same state reset before rerunning.

    Older Streamlit versions without a dialog API use an inline fallback with
    the same explicit close-state behavior.
    """
    if _maybe_reload_this_module():
        st.rerun()

    path = Path(path)
    open_key = _state_key(state_prefix, "open")

    if st.button(button_label, key=_state_key(state_prefix, "button"), use_container_width=True):
        st.session_state[open_key] = True
        st.rerun()

    if not st.session_state.get(open_key, False):
        return

    dialog_fn = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)
    if dialog_fn is not None:
        _inject_editor_dialog_css()
        try:
            decorator = dialog_fn(title, width="large")
        except TypeError:
            decorator = dialog_fn(title)

        @decorator
        def _catalogue_editor_modal():
            _inject_editor_dialog_css()
            _render_editor_body(
                path=path,
                title=title,
                state_prefix=state_prefix,
                expected_top_level=expected_top_level,
                editor_kind=editor_kind,
                candidates=candidates,
                compact_header=True,
            )

        _catalogue_editor_modal()
        return

    # Inline fallback for older Streamlit versions with no dialog API.
    with st.expander(title, expanded=True):
        st.info(
            "This Streamlit version does not support modal dialogs; "
            "the catalogue editor is shown inline instead."
        )
        _render_editor_body(
            path=path,
            title=title,
            state_prefix=state_prefix,
            expected_top_level=expected_top_level,
            editor_kind=editor_kind,
            candidates=candidates,
        )
