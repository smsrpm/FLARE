"""
PWR Dry Containment Response Simulator
=======================================
Author  : R. P. Martin  (structure)  —  containment physics extension
Version : 1.0  May 2026

Physics
-------
Single control-volume (CV) dry containment model.  The CV mixture is a
three-component system:

    1. Steam  (water vapour)
    2. Liquid water (condensate on walls + sump)
    3. Non-condensable gas (air, initially present; may include H₂ if later
       extended)

State variables (two ODEs, same 2×2 matrix structure as FLARE/pwr_sim):
    P   — containment total pressure  [Pa]
    h   — mixture specific enthalpy   [J/kg]  (steam + non-condensable mixture)

Derived figures of merit:
    T_cont   — bulk gas/steam temperature  [°C or K]
    V_liq    — accumulated liquid volume   [m³]
    x_nc     — non-condensable mass fraction (quality analogue)  [-]

Wall condensation model
-----------------------
The principal condensation mechanism is film condensation on the steel liner /
concrete wall in the presence of non-condensable gas (Dehbi correlation for
laminar-to-turbulent forced-convection condensation).  The rate is limited by
both heat transfer through the condensate film and mass-transfer resistance of
the non-condensable boundary layer.

Governing equations  (per FLARE 2×2 matrix form)
-------------------------------------------------
The lumped-parameter CV energy and volume constraints give:

    A * [dP/dt, dh/dt]^T  =  b

where A is the 2×2 Jacobian of internal energy and specific volume with
respect to P and h (identical in structure to the RCS solver in FLARE), and
the RHS vector b captures:

    b[0]  = Σ ṁ_in·h_in  −  ṁ_cond·h_fg  −  Q_wall  +  Q_src
    b[1]  = −v_mix · Σ(net mass flows)

v_mix and the partial derivatives are evaluated at each timestep from XSteam
(steam table) and the ideal-gas law for the non-condensable component.

References
----------
[1] Dehbi, A.A. (1991) – condensation in presence of non-condensables.
[2] Peterson, P.F. (1996) – Theoretical basis for the Prandtl number scaling.
[3] ANS/ASME containment analysis guidelines.
[4] FLARE pwr_sim solution method (FLARE v1.0, R.P. Martin 2026).

Units:  SI throughout unless noted.
"""

from __future__ import annotations

import math
import os
import shutil
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import numpy as np

# ── Optional XSteamPython ─────────────────────────────────────────────────────
try:
    import XSteamPython as XSteam
    _XSTEAM = True
except ImportError:
    _XSTEAM = False
    print("WARNING: XSteamPython not found. Install with: pip install XSteamPython")
    print("         Falling back to ideal-gas approximation (reduced accuracy).")

# ── Optional matplotlib ───────────────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use("Agg")          # non-interactive backend safe on all platforms
    import matplotlib.pyplot as plt
    _MPL = True
except ImportError:
    _MPL = False
    print("WARNING: matplotlib not found. Plots will be skipped.")

# ── Optional openpyxl ─────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                 Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    print("WARNING: openpyxl not found. Install with: pip install openpyxl")
    print("         Excel input/output will be unavailable.")

# ─────────────────────────────────────────────────────────────────────────────
# Physical constants
# ─────────────────────────────────────────────────────────────────────────────
R_AIR    = 287.05      # J/(kg·K)  — specific gas constant for air
R_STEAM  = 461.52      # J/(kg·K)  — specific gas constant for steam
R_H2     = 4157.0      # J/(kg·K)  — specific gas constant for hydrogen (R/M, M=2.016e-3)
CP_AIR   = 1005.0      # J/(kg·K)  — specific heat of air (constant pressure)
CV_AIR   = 718.0       # J/(kg·K)  — specific heat of air (constant volume)
CP_STEAM = 1872.0      # J/(kg·K)  — specific heat of steam (approximate)
CP_H2    = 14_307.0    # J/(kg·K)  — specific heat of hydrogen (constant pressure)
CV_H2    = 10_183.0    # J/(kg·K)  — specific heat of hydrogen (constant volume)
MW_AIR   = 28.97e-3    # kg/mol    — molar mass of air
MW_H2O   = 18.015e-3   # kg/mol    — molar mass of water / steam
MW_H2    = 2.016e-3    # kg/mol    — molar mass of hydrogen
K_STEEL  = 50.0        # W/(m·K)   — thermal conductivity of carbon steel liner
K_CONC   = 1.5         # W/(m·K)   — thermal conductivity of concrete
RHO_WATER = 958.0      # kg/m³     — liquid water density (near saturation, approx)
G_GRAV   = 9.81        # m/s²
SIGMA_SB = 5.67e-8     # W/(m²·K⁴) — Stefan-Boltzmann constant
PATM     = 101_325.0   # Pa        — standard atmosphere
T_REF    = 273.15      # K         — 0 °C in Kelvin

# ── Hydrogen flammability limits (volumetric / mole fractions in steam-air-H2) ─
# Source: NUREG/CR-6509, EPRI TR-106325, Shapiro & Moffette (1957)
# Lower flammability limit (LFL): ~4 vol% H2 in air (no steam)
# Upper flammability limit (UFL): ~75 vol% H2 in air (no steam)
# Steam inerting limit: ~55 vol% steam (balance air) — no combustion above this
# Detonation peninsula (approximate): H2 15–59 vol% at low steam
H2_LFL_MOLE   = 0.040   # lower flammability limit, H2 mole fraction in air
H2_UFL_MOLE   = 0.750   # upper flammability limit, H2 mole fraction in air
STEAM_INERT_MOLE = 0.55  # steam inerting limit, steam mole fraction
# Detonation region approximate bounds (Shapiro diagram)
H2_DET_LO_MOLE  = 0.18  # lower H2 mole fraction for detonation
H2_DET_HI_MOLE  = 0.59  # upper H2 mole fraction for detonation


# ─────────────────────────────────────────────────────────────────────────────
# Steam property helpers (XSteam wrapper with ideal-gas fallback)
# ─────────────────────────────────────────────────────────────────────────────

def _psat_kPa(T_K: float) -> float:
    """Saturation pressure [kPa] at temperature T_K [K]."""
    T_C = T_K - T_REF
    if _XSTEAM:
        try:
            return float(XSteam.Psat_T(T_C))      # kPa
        except Exception:
            pass
    # Antoine fallback (valid ~60–300 °C)
    T_C = max(1.0, min(T_C, 370.0))
    return math.exp(16.3872 - 3885.70 / (T_C + 230.170))


def _tsat_K(P_kPa: float) -> float:
    """Saturation temperature [K] at pressure P_kPa [kPa]."""
    if _XSTEAM:
        try:
            return float(XSteam.Tsat_p(max(P_kPa, 0.1))) + T_REF
        except Exception:
            pass
    # Antoine inversion
    P_kPa = max(P_kPa, 0.1)
    T_C = 3885.70 / (16.3872 - math.log(P_kPa)) - 230.170
    return T_C + T_REF


def _hfg_Jkg(P_kPa: float) -> float:
    """Latent heat of vaporisation [J/kg] at P_kPa."""
    if _XSTEAM:
        try:
            return 1000.0 * (XSteam.hV_p(P_kPa) - XSteam.hL_p(P_kPa))
        except Exception:
            pass
    # Simplified fit (Watson equation)
    T_K = _tsat_K(P_kPa)
    return max(0.0, 2.501e6 * (1.0 - T_K / 647.1) ** 0.38)


def _hg_Jkg(P_kPa: float) -> float:
    """Specific enthalpy of saturated steam [J/kg]."""
    if _XSTEAM:
        try:
            return 1000.0 * XSteam.hV_p(P_kPa)
        except Exception:
            pass
    T_K = _tsat_K(P_kPa)
    return CP_STEAM * (T_K - T_REF) + _hfg_Jkg(P_kPa)


def _hL_Jkg(P_kPa: float) -> float:
    """Specific enthalpy of saturated liquid [J/kg]."""
    if _XSTEAM:
        try:
            return 1000.0 * XSteam.hL_p(P_kPa)
        except Exception:
            pass
    T_K = _tsat_K(P_kPa)
    return 4186.0 * (T_K - T_REF)


def _cp_steam_Jkg(P_kPa: float, T_K: float) -> float:
    """Specific heat of superheated steam [J/(kg·K)] (approximate)."""
    if _XSTEAM:
        try:
            h_kJkg = XSteam.h_pT(P_kPa, T_K - T_REF)
            dp = 0.1 * P_kPa
            h1 = XSteam.h_pT(P_kPa, T_K - T_REF + 1.0)
            h2 = XSteam.h_pT(P_kPa, T_K - T_REF - 1.0)
            return max(1000.0, 1000.0 * (h1 - h2) / 2.0)
        except Exception:
            pass
    return CP_STEAM


def _rho_steam(P_kPa: float, T_K: float) -> float:
    """Density of steam [kg/m³] at (P_kPa, T_K)."""
    if _XSTEAM:
        try:
            h_kJkg = XSteam.h_pT(P_kPa, T_K - T_REF)
            return float(XSteam.rho_ph(P_kPa, h_kJkg))
        except Exception:
            pass
    return P_kPa * 1e3 / (R_STEAM * T_K)


# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# Wall condensation models — Uchida (1965) and DLM-FM
# ─────────────────────────────────────────────────────────────────────────────

def _uchida(
        T_bulk_K:   float,
        T_wall_K:   float,
        P_total_Pa: float,
        x_nc:       float,
        multiplier: float = 1.0,
) -> float:
    """
    Uchida (1965) condensation heat flux [W/m²].

    Empirical correlation:  h = 380 · (W_nc/W_stm)^{-0.707}  [W/(m²·K)]
    Driving ΔT = T_sat(P_steam) − T_wall.

    Uncertainty multiplier range 0.5–1.5 (GRS / AREVA study).
    """
    if T_wall_K >= T_bulk_K:
        return 0.0
    x_steam = max(1.0 - x_nc, 1e-6)
    if x_steam < 1e-4:
        return 0.0
    P_stm_kPa = P_total_Pa * 1e-3 * x_steam
    T_sat_K   = _tsat_K(P_stm_kPa)
    if T_sat_K <= T_wall_K:
        return 0.0
    ratio  = float(np.clip(x_nc / x_steam, 0.01, 1e4))
    h_cond = 380.0 / (ratio ** 0.707) * multiplier
    return float(max(h_cond, 1.0) * (T_sat_K - T_wall_K))


def _dlm_fm(
        T_bulk_K:   float,
        T_wall_K:   float,
        P_total_Pa: float,
        x_nc:       float,
        L_wall:     float = 10.0,
        multiplier: float = 1.0,
) -> float:
    """
    Diffusion Layer Model with Film and Mist corrections (DLM-FM) condensation
    heat flux [W/m²].

    Physical model
    --------------
    Condensation is controlled by steam diffusion through the air/steam
    boundary layer at the wall.  The approach follows the classical
    film-model mass transfer analogy:

    1.  Steam-air binary diffusivity  D_sa  [m²/s]:
            D_sa = 2.56e-5 · (T_bulk / 298)^1.8 / (P / P_atm)
        (Chapman–Enskog kinetic theory; agrees with GOTHIC qualification data.)

    2.  Convective mass transfer coefficient  h_m  [m/s] via the natural-
        convection Sherwood–Nusselt analogy on a vertical plate:
            Sh = 0.59 · Ra_m^{0.25}    (laminar,  10^4  < Ra < 10^9)
            Sh = 0.10 · Ra_m^{0.333}   (turbulent, Ra > 10^9)
        where Ra_m uses the mass-fraction driving force for buoyancy:
            Ra_m = g · β_c · Δω · L³ / (ν · D_sa)
            β_c  ≈ (MW_air/MW_steam − 1) · x_steam   (concentration expansion)
            Δω   = max steam mass-fraction difference driving diffusion
        h_m = Sh · D_sa / L

    3.  Spalding mass transfer number  B_m:
            ω_s  = steam mass fraction at wall  ≈ 0  (wall fully cold,
                   condensate drains — standard DLM assumption)
            ω_∞  = steam mass fraction in bulk  = 1 − x_nc
            B_m  = (ω_∞ − ω_s) / (1 − ω_∞)  =  (1 − x_nc) / x_nc

    4.  Steam mass flux  ṁ"  [kg/(m²·s)]:
            ṁ" = h_m · ρ_mix · ln(1 + B_m)

    5.  Heat flux  q  [W/m²]:
            q = ṁ" · h_fg(T_sat)

    Film correction (FM):
        The condensate film at the wall enhances turbulent mixing and effective
        conductance.  This is captured through the multiplier parameter, which
        absorbs both the ±20% DLM-FM model uncertainty (B&W mPower uncertainty
        study, Table 5-1) and any film-roughening enhancement.  Default = 1.0.

    Single-volume adaptation
    ------------------------
    In a single-volume model the local flow field is not resolved.  The model
    uses bulk-average gas state (T_bulk, P_total, x_nc) as the outer boundary
    condition for the diffusion layer, consistent with the B&W mPower GOTHIC
    SEM approach (MPWR-TECR-005062, Sec. 3.1.2).  The relative velocity is
    estimated via the natural-convection correlation on characteristic length
    L_wall, which is the only velocity information available in a lumped model.

    Parameters
    ----------
    T_bulk_K   : bulk gas temperature          [K]
    T_wall_K   : wall inner surface temperature [K]
    P_total_Pa : total containment pressure    [Pa]
    x_nc       : bulk non-condensable mass fraction [-]
    L_wall     : characteristic height / length [m]  (default 10 m)
    multiplier : DLM-FM model multiplier        [-]  (uncertainty range 0.8–1.2)

    Returns
    -------
    q_cond : condensation heat flux [W/m²]  (positive = heat into wall)

    References
    ----------
    Lienhard & Lienhard (2008), "A Heat Transfer Textbook," 4th ed.
    Peterson (1996), GOTHIC qualification report, NAI 8907-09.
    B&W mPower MPWR-TECR-005062 Rev 001, Sec. 3.1.2 (DLM-FM ±20%).
    """
    if T_wall_K >= T_bulk_K:
        return 0.0
    x_steam = max(1.0 - x_nc, 1e-6)
    if x_steam < 1e-4:
        return 0.0

    # Steam partial pressure and saturation temperature
    P_stm_Pa  = P_total_Pa * x_steam
    T_sat_K   = _tsat_K(P_stm_Pa * 1e-3)
    if T_sat_K <= T_wall_K:
        return 0.0

    # ── Mixture properties at bulk conditions ─────────────────────────────────
    T_film    = 0.5 * (T_bulk_K + T_wall_K)   # film temperature [K]
    P_atm     = 101325.0                        # Pa

    # Steam-air binary diffusivity [m²/s]  (Chapman–Enskog, pressure-corrected)
    D_sa = 2.56e-5 * (T_film / 298.0)**1.8 / (P_total_Pa / P_atm)

    # Mixture density (ideal gas, MW_mix weighted)
    MW_air    = 28.97e-3   # kg/mol
    MW_steam  = 18.015e-3  # kg/mol
    # molar fractions from mass fractions
    y_steam   = (x_steam / MW_steam) / (x_steam / MW_steam + x_nc / MW_air)
    MW_mix    = y_steam * MW_steam + (1.0 - y_steam) * MW_air
    R_univ    = 8.314
    rho_mix   = P_total_Pa * MW_mix / (R_univ * T_bulk_K)   # kg/m³

    # Dynamic viscosity of air/steam mixture (Wilke mixing rule simplified)
    mu_air    = 1.81e-5 * (T_film / 300.0)**0.7    # Pa·s
    mu_steam  = 1.22e-5 * (T_film / 380.0)**0.9
    mu_mix    = x_nc * mu_air + x_steam * mu_steam  # rough linear mix

    nu_mix    = mu_mix / max(rho_mix, 0.01)         # kinematic viscosity [m²/s]

    # ── Natural-convection Sherwood number (vertical plate analogy) ───────────
    # Concentration buoyancy: lighter steam-rich region rises along cold wall
    # β_c = solutal expansion coefficient
    beta_c    = (MW_air / MW_steam - 1.0) * x_steam   # ≈ 0.608 · x_steam
    beta_c    = max(beta_c, 0.0)

    # Concentration driving force Δω (bulk steam fraction - wall ≈ 0)
    delta_omega = x_steam   # ω_wall ≈ 0 (cold wall, condensate drains)

    g = 9.81
    Ra_m = (g * beta_c * delta_omega * L_wall**3) / (nu_mix * D_sa)
    Ra_m = max(Ra_m, 1.0)

    if Ra_m < 1e9:
        Sh = 0.59 * Ra_m**0.25   # laminar
    else:
        Sh = 0.10 * Ra_m**(1.0/3.0)  # turbulent

    h_m = Sh * D_sa / L_wall   # mass transfer coefficient [m/s]

    # ── Spalding transfer number and steam mass flux ──────────────────────────
    # B_m = (ω_∞ - ω_wall) / (1 - ω_∞)  with ω_wall = 0
    B_m    = x_steam / max(x_nc, 1e-6)
    B_m    = max(B_m, 1e-6)

    mdot_pp = h_m * rho_mix * math.log(1.0 + B_m)   # kg/(m²·s)

    # ── Heat flux via latent heat ─────────────────────────────────────────────
    h_fg  = max(_hfg_Jkg(P_stm_Pa * 1e-3), 1e4)   # J/kg
    q_cond = mdot_pp * h_fg * multiplier

    return float(max(q_cond, 0.0))


def wall_condensation_heat_flux(
        T_bulk_K:   float,
        T_wall_K:   float,
        P_total_Pa: float,
        x_nc:       float,
        L_wall:     float = 10.0,
        uchida_multiplier: float = 1.0,
        model: str = 'uchida',
) -> float:
    """
    Wall condensation heat flux [W/m²] — dispatcher.

    Selects between two models via the ``model`` parameter:

    'uchida'  (default)
        Uchida (1965) empirical correlation.  Standard for NRC-approved
        single-volume containment analysis (GOTHIC, CONTAIN, MELCOR).
        Uncertainty multiplier range 0.5–1.5.

    'dlm_fm'
        Diffusion Layer Model with Film and Mist corrections (DLM-FM).
        Mass-transfer based; explicitly models the steam/air diffusion layer.
        Used in the B&W mPower GOTHIC SEM (MPWR-TECR-005062).
        Uncertainty multiplier range 0.8–1.2.

    The ``uchida_multiplier`` parameter applies to whichever model is selected.
    """
    if model == 'dlm_fm':
        return _dlm_fm(T_bulk_K, T_wall_K, P_total_Pa, x_nc,
                       L_wall=L_wall, multiplier=uchida_multiplier)
    else:
        return _uchida(T_bulk_K, T_wall_K, P_total_Pa, x_nc,
                       multiplier=uchida_multiplier)


# ─────────────────────────────────────────────────────────────────────────────
# Wall heat conduction (1-D transient slab, lumped two-node approximation)
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Multi-layer wall conduction model
# steel liner  |  air gap  |  concrete backing (semi-infinite)
# ─────────────────────────────────────────────────────────────────────────────

class MultiLayerWall:
    """
    1-D transient conduction through the containment wall cross-section:

        [gas] → steel liner → air gap → concrete/wall backing → [outer BC]

    Layer layout
    ------------
    Layer 0  : steel liner    — 2 nodes, explicit FD, Neumann inner BC
    Interface: liner/concrete air gap — modelled as a contact resistance
               R_gap = gap_thickness / k_air  [m²·K/W]
    Layer 1  : concrete       — N_conc nodes, explicit FD

    Outer boundary condition
    ------------------------
    Two options, linearly blended by ``pcc_fraction`` [0–1]:

    **Adiabatic** (pcc_fraction = 0, default):
        dT/dx = 0 at the outer face.  Standard conservative assumption
        for dry concrete-lined containments (10 CFR 50 Appendix K).
        Heat can only accumulate in the wall — never leave.

    **Isothermal / active cooling** (pcc_fraction = 1):
        T = T_outer_K at the outer face (Dirichlet BC).
        Models a steel containment with passive containment cooling (PCC)
        where the outer surface is held at a fixed temperature by external
        air circulation and/or water film evaporation (AP1000/AP300 design).
        Heat flows continuously through the wall and is removed at the
        outer surface at whatever rate conduction allows.

    **Mixed** (0 < pcc_fraction < 1):
        The outer node update is a weighted blend:
            dT_last = (1 − f) × [adiabatic update]
                    +      f  × [Dirichlet restore toward T_outer]
        where f = pcc_fraction.  This handles cases where only part of
        the wall area participates in active cooling (e.g. a steel dome
        with PCC above a concrete cylindrical wall without).

    Parameters
    ----------
    T_init_K        : initial uniform temperature [K]
    liner_thick_m   : steel liner thickness [m]
    liner_k         : liner thermal conductivity [W/(m·K)]
    liner_rho_cp    : liner volumetric heat capacity [J/(m³·K)]
    gap_thick_m     : liner/concrete air gap thickness [m]
    k_air_gap       : gap effective conductivity [W/(m·K)]  (air ≈ 0.026)
    conc_thick_m    : concrete modelled depth [m]
    conc_k          : concrete thermal conductivity [W/(m·K)]
    conc_rho_cp     : concrete volumetric heat capacity [J/(m³·K)]
    N_conc          : number of concrete nodes
    T_outer_K       : outer surface temperature for isothermal BC [K]
                      (only used when pcc_fraction > 0; default = T_init_K)
    pcc_fraction    : fraction of wall area with active outer cooling [0–1]
                      0 = fully adiabatic (default, conservative)
                      1 = fully isothermal outer surface (PCC active)
    """

    def __init__(self,
                 T_init_K:      float,
                 liner_thick_m: float = 0.044,
                 liner_k:       float = K_STEEL,
                 liner_rho_cp:  float = 7800.0 * 500.0,
                 gap_thick_m:   float = 0.001,
                 k_air_gap:     float = 0.026,
                 conc_thick_m:  float = 1.5,
                 conc_k:        float = K_CONC,
                 conc_rho_cp:   float = 2300.0 * 900.0,
                 N_conc:        int   = 20,
                 T_outer_K:     float = None,
                 pcc_fraction:  float = 0.0):

        self.liner_k      = liner_k
        self.liner_rho_cp = liner_rho_cp
        self.dx_liner     = liner_thick_m / 2.0

        # Gap thermal resistance [m²·K/W]
        self.R_gap = gap_thick_m / max(k_air_gap, 1e-6)

        self.conc_k      = conc_k
        self.conc_rho_cp = conc_rho_cp
        self.N_conc      = N_conc
        self.dx_conc     = conc_thick_m / N_conc
        self.alpha_conc  = conc_k / conc_rho_cp
        self.alpha_liner = liner_k / liner_rho_cp

        # Outer BC parameters
        self.T_outer     = T_init_K if T_outer_K is None else T_outer_K
        self.pcc_frac    = float(np.clip(pcc_fraction, 0.0, 1.0))

        # State arrays
        self.T_liner = np.full(2,       T_init_K)
        self.T_conc  = np.full(N_conc,  T_init_K)

    @property
    def T_inner(self) -> float:
        """Inner liner surface temperature [K] — drives condensation."""
        return float(self.T_liner[0])

    @property
    def T_conc_surface(self) -> float:
        """Concrete surface temperature (gap side) [K]."""
        return float(self.T_conc[0])

    @property
    def T_outer_surface(self) -> float:
        """Outer wall surface temperature [K]."""
        return float(self.T_conc[-1])

    def q_outer_W_m2(self) -> float:
        """
        Heat flux leaving the outer surface [W/m²] (positive = heat out).
        Zero for fully adiabatic BC; non-zero when pcc_fraction > 0.
        """
        if self.pcc_frac <= 0.0:
            return 0.0
        # Flux from last concrete node to the fixed outer temperature
        dxC = self.dx_conc
        kC  = self.conc_k
        return kC / dxC * (self.T_conc[-1] - self.T_outer) * self.pcc_frac

    def step(self, q_inner: float, dt: float) -> float:
        """
        Advance wall temperatures by dt [s].

        q_inner [W/m²]: heat flux into the wall inner surface (positive inward).

        Outer BC is a weighted blend of adiabatic and isothermal, controlled
        by pcc_fraction:
          - Adiabatic  : last concrete node uses ghost-node zero-flux condition
          - Isothermal : last concrete node is restored toward T_outer at the
                         rate that pure Dirichlet conduction would imply
          - Mixed      : linear blend of both updates

        Returns inner liner surface temperature [K].
        """
        dt_liner = 0.5 * self.dx_liner**2 / max(self.alpha_liner, 1e-12)
        dt_conc  = 0.5 * self.dx_conc**2  / max(self.alpha_conc,  1e-12)
        dt_s     = min(dt_liner, dt_conc) * 0.45
        dt_s     = max(dt_s, 1e-6)
        n_sub    = max(1, min(int(math.ceil(dt / dt_s)), 50000))
        dt_s     = dt / n_sub

        alpha_L  = self.alpha_liner
        alpha_C  = self.alpha_conc
        dxL      = self.dx_liner
        dxC      = self.dx_conc
        rcpL     = self.liner_rho_cp
        rcpC     = self.conc_rho_cp
        R_gap    = self.R_gap
        f_pcc    = self.pcc_frac
        T_outer  = self.T_outer

        for _ in range(n_sub):
            TL = self.T_liner
            TC = self.T_conc

            # Heat flux across gap
            q_gap = (TL[1] - TC[0]) / max(R_gap, 1e-9)

            # ── Liner nodes ───────────────────────────────────────────────────
            dTL0 = (q_inner / (rcpL * dxL)
                    - alpha_L / dxL**2 * (TL[0] - TL[1])) * dt_s
            dTL1 = (alpha_L / dxL**2 * (TL[0] - TL[1])
                    - q_gap  / (rcpL * dxL)) * dt_s

            # ── Concrete nodes ────────────────────────────────────────────────
            dTC = np.zeros(self.N_conc)
            dTC[0] = (q_gap / (rcpC * dxC)
                      - alpha_C / dxC**2 * (TC[0] - TC[1])) * dt_s
            for i in range(1, self.N_conc - 1):
                dTC[i] = alpha_C / dxC**2 * (TC[i-1] - 2*TC[i] + TC[i+1]) * dt_s

            # ── Outer BC — blended adiabatic / isothermal ─────────────────────
            # Adiabatic update: ghost node = T_last → Laplacian = T[-2] - T[-1]
            dTC_adiabatic = alpha_C / dxC**2 * (TC[-2] - TC[-1]) * dt_s
            # Isothermal update: ghost node = T_outer → Laplacian = T[-2] - 2T[-1] + T_outer
            dTC_isothermal = alpha_C / dxC**2 * (TC[-2] - 2*TC[-1] + T_outer) * dt_s
            # Blend
            dTC[-1] = (1.0 - f_pcc) * dTC_adiabatic + f_pcc * dTC_isothermal

            self.T_liner += np.array([dTL0, dTL1])
            self.T_conc  += dTC

        return self.T_inner


# ─────────────────────────────────────────────────────────────────────────────
# Internal floor/sump treatment
# ─────────────────────────────────────────────────────────────────────────────
# The previous standalone SumpPool class was removed.  The reduced-order model
# now treats the full containment floor as an active condensing/convecting
# surface, partitioned into dry floor and wet floor/sump-film area inside the
# main mass/energy integrator.  RWST behavior is represented only through
# injection/source boundary conditions unless explicitly modeled as exposed to
# containment atmosphere.


# ─────────────────────────────────────────────────────────────────────────────
# Design parameters dataclass
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ContainmentDesign:
    """
    Geometric and material design parameters for the dry containment.

    All parameters have physically representative PWR defaults.
    """
    # Geometry
    air_volume_m3:       float = 50_000.0   # free air volume (gas space)  [m³]
    floor_area_m2:       float = 600.0      # projected containment floor area [m²]

    # Engineered sump / floor flooding geometry
    # Condensate is assumed to wet the containment floor early while draining
    # to the sump.  Standing liquid accumulates in the sump first; only after
    # the sump is full does a nonzero containment-floor flood depth appear.
    sump_area_m2:        float = 25.0       # engineered sump free-surface / plan area [m²]
    sump_depth_m:        float = 2.0        # sump depth to overflow onto floor [m]

    wall_surf_area_m2:   float = 8_000.0    # total condensing wall area   [m²]
    wall_thickness_m:    float = 0.044      # steel liner thickness        [m]
    wall_char_height_m:  float = 10.0       # characteristic wall height   [m]

    # Steel liner material
    wall_k_Wm_K:     float = K_STEEL   # thermal conductivity [W/(m·K)]
    wall_rho_kg_m3:  float = 7800.0    # density              [kg/m³]
    wall_cp_J_kg_K:  float = 500.0     # specific heat        [J/(kg·K)]

    # Liner / concrete air gap
    gap_thick_m:     float = 0.001     # gap thickness [m]  (Table V: 0.225–3 mm)
    gap_k_Wm_K:      float = 0.026     # gap conductivity [W/(m·K)]  (air)

    # Concrete backing
    conc_k_Wm_K:     float = K_CONC    # thermal conductivity [W/(m·K)]
    conc_rho_kg_m3:  float = 2300.0    # density              [kg/m³]
    conc_cp_J_kg_K:  float = 900.0     # specific heat        [J/(kg·K)]
    conc_depth_m:    float = 1.5       # modelled concrete depth [m]
    conc_nodes:      int   = 20        # FD nodes in concrete

    # Design pressure / temperature limits (informational, not enforced)
    P_design_kPa:   float = 520.0    # design pressure  [kPa]
    T_design_K:     float = 450.0    # design temperature [K]

    # Condensation model
    # 'uchida'  — Uchida (1965) empirical HTC; uncertainty multiplier 0.5–1.5
    # 'dlm_fm'  — Diffusion Layer Model with Film/Mist; multiplier 0.8–1.2
    condensation_model: str   = 'uchida'
    uchida_multiplier:  float = 1.0    # applies to whichever model is selected

    # Outer wall boundary condition
    # pcc_fraction = 0  : fully adiabatic (default — conservative dry containment)
    # pcc_fraction = 1  : fully isothermal outer surface (AP1000/AP300 PCC)
    # 0 < f < 1         : fraction of wall area with active outer cooling
    pcc_fraction:   float = 0.0    # fraction of wall with active outer cooling [-]
    T_outer_C:      float = 15.0   # fixed outer surface temperature when PCC active [°C]

    # Containment spray model
    # Sprays are modeled as liquid water injected into the containment gas space.
    # The water removes heat from the atmosphere by direct sensible cooling and,
    # when the atmosphere is supersaturated, by condensing steam on droplets.
    spray_enabled: bool = False
    spray_effectiveness: float = 0.80       # residual calibration/uncertainty factor [-]
    spray_sauter_mean_diameter_um: float = 500.0  # Sauter mean droplet diameter d32 [micron]
    spray_flow_time_s: np.ndarray = None
    spray_flow_kg_s: np.ndarray = None
    spray_temp_C: np.ndarray = None

    def __post_init__(self):
        if self.spray_flow_time_s is None:
            self.spray_flow_time_s = np.array([0.0, 86400.0])
        if self.spray_flow_kg_s is None:
            self.spray_flow_kg_s = np.zeros_like(self.spray_flow_time_s, dtype=float)
        if self.spray_temp_C is None:
            self.spray_temp_C = np.full_like(self.spray_flow_time_s, 27.0, dtype=float)

    @property
    def sump_capacity_m3(self) -> float:
        """Liquid volume [m³] held below floor level before sump overflow."""
        return max(float(self.sump_area_m2), 0.0) * max(float(self.sump_depth_m), 0.0)

    def liquid_distribution(self, V_liq_m3: float) -> dict:
        """Return reduced-order sump/floor flooding state for a total liquid volume.

        Definitions
        -----------
        sump_level_m
            Water level measured from the sump bottom.  It rises from zero to
            ``sump_depth_m`` while the sump fills.  After overflow, it is
            ``sump_depth_m + floor_flood_level_m`` because the floor film and
            sump are hydraulically connected.

        floor_flood_level_m
            Equivalent water depth on the containment floor outside the sump,
            measured above the floor elevation.  It remains zero until the sump
            capacity is exceeded.

        A_floor_wet_m2
            Active wet horizontal surface area exposed to the containment
            atmosphere.  Because wall condensate drains across the floor toward
            the sump, floor wetness is decoupled from standing flood depth: once
            liquid is present, the full projected floor is treated as wetted for
            heat-transfer purposes.
        """
        V_total = max(float(V_liq_m3), 0.0)
        A_floor = max(float(self.floor_area_m2), 1.0)
        A_sump = float(np.clip(float(self.sump_area_m2), 0.0, A_floor))
        depth_sump = max(float(self.sump_depth_m), 0.0)
        V_sump_cap = A_sump * depth_sump
        A_floor_outside_sump = max(A_floor - A_sump, 1.0e-9)

        V_sump = min(V_total, V_sump_cap)
        V_floor = max(V_total - V_sump_cap, 0.0)

        if A_sump > 0.0:
            if V_total <= V_sump_cap or V_floor <= 0.0:
                sump_level = V_sump / A_sump
                floor_level = 0.0
            else:
                floor_level = V_floor / A_floor_outside_sump
                sump_level = depth_sump + floor_level
        else:
            # Degenerate case: no engineered sump; all liquid spreads on floor.
            V_sump = 0.0
            V_floor = V_total
            floor_level = V_floor / A_floor
            sump_level = 0.0

        # Wetness is not the same as standing liquid depth.  Condensate
        # films draining from the wall wet the floor early even while the
        # accumulated inventory is still stored hydraulically in the sump.
        # Standing floor flooding remains zero until sump overflow.
        A_wet = A_floor if V_total > 1.0e-9 else 0.0
        A_dry = max(A_floor - A_wet, 0.0)
        return {
            'V_sump_m3': V_sump,
            'V_floor_m3': V_floor,
            'sump_level_m': sump_level,
            'floor_flood_level_m': floor_level,
            'A_floor_wet_m2': A_wet,
            'A_floor_dry_m2': A_dry,
        }

    def spray_at(self, t: float) -> tuple[float, float]:
        """Return containment spray flow [kg/s] and water temperature [K]."""
        if not self.spray_enabled:
            return 0.0, T_REF + 27.0
        mdot = float(np.interp(t, self.spray_flow_time_s, self.spray_flow_kg_s))
        temp_C = float(np.interp(t, self.spray_flow_time_s, self.spray_temp_C))
        return max(mdot, 0.0), temp_C + T_REF

    def spray_thermal_effectiveness(self, Tgas: float, Pgas: float) -> float:
        """Reduced-order droplet thermal effectiveness based on Sauter mean diameter.

        The model estimates a droplet residence-time heat-transfer effectiveness
        using A/m = 6/(rho_l*d32), a Ranz-Marshall Nusselt number, and a
        characteristic fall height approximated by wall_char_height_m.  The
        user-facing physical parameter is d32; spray_effectiveness remains a
        secondary calibration/uncertainty multiplier.
        """
        d = max(float(self.spray_sauter_mean_diameter_um), 1.0) * 1.0e-6
        rho_l = 1000.0
        cp_l = 4180.0
        g = 9.80665
        # Conservative dry-gas properties representative of steam/air at containment conditions.
        rho_g = max(Pgas / (287.0 * max(Tgas, 250.0)), 0.05)
        mu_g = 2.0e-5
        k_g = 0.030
        Pr = 0.70
        # Approximate terminal/fall velocity; bounded for robustness.
        Cd = 0.8
        v = (4.0 * g * d * max(rho_l - rho_g, 1.0) / (3.0 * Cd * rho_g)) ** 0.5
        v = float(np.clip(v, 0.25, 8.0))
        tau = max(float(self.wall_char_height_m), 1.0) / v
        Re = max(rho_g * v * d / mu_g, 1.0e-6)
        Nu = 2.0 + 0.6 * (Re ** 0.5) * (Pr ** (1.0/3.0))
        h = Nu * k_g / d
        eps = 1.0 - np.exp(-6.0 * h * tau / (rho_l * cp_l * d))
        return float(np.clip(self.spray_effectiveness, 0.0, 2.0) * np.clip(eps, 0.0, 1.0))


@dataclass
class ContainmentIC:
    """
    Initial conditions for the containment atmosphere.

    Typical pre-accident state: air at ~PATM, slightly above ambient.
    """
    P0_Pa:      float = PATM           # total initial pressure         [Pa]
    T0_K:       float = 300.0          # initial gas temperature        [K]
    T_wall0_K:  float = 300.0          # initial wall temperature       [K]
    RH0:        float = 0.0            # initial relative humidity [-]  (0 = dry)


# ─────────────────────────────────────────────────────────────────────────────
# Source term specification
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SourceTable:
    """
    Time-dependent mass and energy source into containment.

    Arrays must be the same length.  Linear interpolation is used between
    table points.  Negative mass flows are not physical and are clamped to 0.

    Attributes
    ----------
    time_s        : time  [s]  (monotone increasing, starting at 0)
    mdot_kg_s     : steam/water mass flow into containment  [kg/s]
    Qdot_W        : direct volumetric energy source         [W]  (normally 0)
    h_src_J_kg    : specific enthalpy of the incoming steam/water mass  [J/kg]
    mdot_h2_kg_s  : hydrogen mass flow into containment     [kg/s]
                    (from zircaloy oxidation or radiolysis; zero if not applicable)
    """
    time_s:       np.ndarray
    mdot_kg_s:    np.ndarray
    Qdot_W:       np.ndarray
    h_src_J_kg:   np.ndarray
    mdot_h2_kg_s: np.ndarray = None   # optional; defaults to zero if not supplied

    def __post_init__(self):
        if self.mdot_h2_kg_s is None:
            self.mdot_h2_kg_s = np.zeros_like(self.time_s)

    def at(self, t: float) -> Tuple[float, float, float, float]:
        """Interpolate source at time t → (mdot_steam, Qdot, h_src, mdot_h2)."""
        mdot   = float(np.interp(t, self.time_s, self.mdot_kg_s, left=0.0, right=0.0))
        Qdot   = float(np.interp(t, self.time_s, self.Qdot_W, left=0.0, right=0.0))
        h_src  = float(np.interp(t, self.time_s, self.h_src_J_kg, left=0.0, right=0.0))
        mdot_h2= float(np.interp(t, self.time_s, self.mdot_h2_kg_s, left=0.0, right=0.0))
        return max(0.0, mdot), Qdot, h_src, max(0.0, mdot_h2)


# ─────────────────────────────────────────────────────────────────────────────
# Helper: mixture property partial derivatives  (2×2 Jacobian)
# ─────────────────────────────────────────────────────────────────────────────

def _mixture_partials(
        P_Pa: float, T_K: float,
        x_nc: float,
        M_total: float, V_gas: float,
        x_h2: float = 0.0,
) -> dict:
    """
    Compute partial derivatives of internal energy and specific volume
    with respect to pressure and enthalpy for the steam + air + H2 mixture.

    Parameters
    ----------
    P_Pa   : total pressure  [Pa]
    T_K    : gas temperature [K]
    x_nc   : total non-condensable mass fraction (air + H2)  [-]
    M_total: total gas mass  [kg]
    V_gas  : gas-space volume [m³]
    x_h2   : H2 mass fraction of total gas mixture  [-]
    """
    x_air   = max(x_nc - x_h2, 0.0)
    x_steam = 1.0 - x_nc

    dP  = max(P_Pa * 1e-4, 100.0)
    dh  = max(abs(T_K - T_REF) * 4.0, 100.0)

    def _v_mix(Pp, Tp):
        v_air = R_AIR   * Tp / Pp if Pp > 0 else 1.0
        v_h2  = R_H2    * Tp / Pp if Pp > 0 else 1.0
        v_stm = R_STEAM * Tp / Pp if Pp > 0 else 1.0
        if _XSTEAM and x_steam > 1e-6:
            try:
                v_stm = 1.0 / _rho_steam(Pp * 1e-3, Tp)
            except Exception:
                pass
        return x_air * v_air + x_h2 * v_h2 + x_steam * v_stm

    def _u_mix(Pp, Tp):
        u_air = CV_AIR * (Tp - T_REF)
        u_h2  = CV_H2  * (Tp - T_REF)
        if _XSTEAM and x_steam > 1e-6:
            try:
                h_stm = 1000.0 * XSteam.h_pT(Pp * 1e-3, Tp - T_REF)
                v_stm = 1.0 / _rho_steam(Pp * 1e-3, Tp)
                u_stm = h_stm - Pp * v_stm
            except Exception:
                u_stm = CP_STEAM * (Tp - T_REF) - R_STEAM * Tp
        else:
            u_stm = CP_STEAM * (Tp - T_REF) - R_STEAM * Tp
        return x_air * u_air + x_h2 * u_h2 + x_steam * u_stm

    dv_dP = (_v_mix(P_Pa + dP, T_K) - _v_mix(P_Pa - dP, T_K)) / (2.0 * dP)
    du_dP = (_u_mix(P_Pa + dP, T_K) - _u_mix(P_Pa - dP, T_K)) / (2.0 * dP)

    cp_mix = x_air * CP_AIR + x_h2 * CP_H2 + x_steam * _cp_steam_Jkg(P_Pa * 1e-3, T_K)
    cp_mix = max(cp_mix, 500.0)
    dT     = dh / cp_mix

    dv_dh = (_v_mix(P_Pa, T_K + dT) - _v_mix(P_Pa, T_K - dT)) / (2.0 * dh)
    du_dh = (_u_mix(P_Pa, T_K + dT) - _u_mix(P_Pa, T_K - dT)) / (2.0 * dh)

    du_dh = float(np.clip(du_dh, 0.01, 1.0))
    du_dP = float(du_dP)
    dv_dh = float(np.clip(abs(dv_dh), 1e-10, 1.0)) * np.sign(dv_dh + 1e-30)
    dv_dP = float(np.clip(dv_dP, -1e-3, -1e-9))

    return dict(du_dP=du_dP, du_dh=du_dh, dv_dP=dv_dP, dv_dh=dv_dh)


# ─────────────────────────────────────────────────────────────────────────────
# Containment simulator
# ─────────────────────────────────────────────────────────────────────────────

class ContainmentSimulator:
    """
    Single-node dry containment pressure/temperature transient simulator.

    Solution method
    ---------------
    Mirrors FLARE pwr_sim:  at each timestep form the 2×2 linear system

        A · [ΔP, Δh]ᵀ = b · Δt

    then advance:  [P, h]_(n+1) = [P, h]_n + [ΔP, Δh]ᵀ

    The matrix A encodes the volume constraint and energy/mass conservation;
    b is the RHS source/sink vector.  Mass and energy tracking follows the
    gas-phase mixture (steam + non-condensables).  Liquid accumulates in the
    sump as condensation and incoming liquid mass are tracked separately.

    Usage
    -----
    >>> design  = ContainmentDesign(air_volume_m3=50_000, wall_surf_area_m2=8000)
    >>> ic      = ContainmentIC(P0_Pa=101325, T0_K=300)
    >>> src     = SourceTable(...)
    >>> sim     = ContainmentSimulator(design, ic, src)
    >>> results = sim.run(t_end=3600, dt=1.0)
    """

    def __init__(self,
                 design: ContainmentDesign,
                 ic:     ContainmentIC,
                 source: SourceTable,
                 dt_s:   float = 1.0,
                 verbose: bool = True):

        self.design  = design
        self.ic      = ic
        self.source  = source
        self.dt      = dt_s
        self.verbose = verbose

        # ── Initial state ─────────────────────────────────────────────────────
        # Partial pressure of steam = RH * P_sat(T0)
        P_sat0_kPa = _psat_kPa(ic.T0_K)
        P_steam0   = ic.RH0 * P_sat0_kPa * 1e3
        P_air0     = ic.P0_Pa - P_steam0

        # Mass of dry air (ideal gas)
        self.M_air0  = P_air0 * design.air_volume_m3 / (R_AIR * ic.T0_K)

        # Mass of initial steam
        rho_stm0  = P_steam0 / (R_STEAM * ic.T0_K) if P_steam0 > 0 else 0.0
        self.M_stm0 = rho_stm0 * design.air_volume_m3

        # Total initial gas mass
        M_gas0   = self.M_air0 + self.M_stm0

        # Initial specific enthalpy of mixture [J/kg]
        h_air0  = CP_AIR * (ic.T0_K - T_REF)
        h_stm0  = (CP_STEAM * (ic.T0_K - T_REF)  # superheated steam approx
                   if _XSTEAM is False
                   else 1000.0 * XSteam.h_pT(max(P_steam0 * 1e-3, 0.01), ic.T0_K - T_REF))
        x_nc0   = self.M_air0 / M_gas0 if M_gas0 > 0 else 1.0
        self._h0 = x_nc0 * h_air0 + (1.0 - x_nc0) * h_stm0

        # Multi-layer wall (steel liner + gap + concrete) — vertical surfaces
        self._wall = MultiLayerWall(
            T_init_K      = ic.T_wall0_K,
            liner_thick_m = design.wall_thickness_m,
            liner_k       = design.wall_k_Wm_K,
            liner_rho_cp  = design.wall_rho_kg_m3 * design.wall_cp_J_kg_K,
            gap_thick_m   = design.gap_thick_m,
            k_air_gap     = design.gap_k_Wm_K,
            conc_thick_m  = design.conc_depth_m,
            conc_k        = design.conc_k_Wm_K,
            conc_rho_cp   = design.conc_rho_kg_m3 * design.conc_cp_J_kg_K,
            N_conc        = design.conc_nodes,
            T_outer_K     = design.T_outer_C + T_REF,
            pcc_fraction  = design.pcc_fraction,
        )

        # Basemat — always adiabatic outer BC regardless of pcc_fraction
        # (PCC acts on the vertical steel shell, not the basemat)
        self._basemat = MultiLayerWall(
            T_init_K      = ic.T_wall0_K,
            liner_thick_m = design.conc_depth_m / design.conc_nodes,
            liner_k       = design.conc_k_Wm_K,
            liner_rho_cp  = design.conc_rho_kg_m3 * design.conc_cp_J_kg_K,
            gap_thick_m   = 1e-4,
            k_air_gap     = design.conc_k_Wm_K,
            conc_thick_m  = design.conc_depth_m,
            conc_k        = design.conc_k_Wm_K,
            conc_rho_cp   = design.conc_rho_kg_m3 * design.conc_cp_J_kg_K,
            N_conc        = design.conc_nodes,
            pcc_fraction  = 0.0,   # basemat always adiabatic
        )

        # Internal liquid inventory is handled directly in the main integrator.

        if verbose:
            print(f"\n{'='*65}")
            print("  PWR Dry Containment Simulator")
            print(f"{'='*65}")
            print(f"  Volume          : {design.air_volume_m3:,.0f} m³")
            print(f"  Wall area       : {design.wall_surf_area_m2:,.0f} m²")
            print(f"  Liner δ         : {design.wall_thickness_m*100:.1f} cm  "
                  f"(k={design.wall_k_Wm_K:.0f} W/m·K)")
            print(f"  Gap             : {design.gap_thick_m*1000:.2f} mm")
            print(f"  Concrete        : {design.conc_depth_m:.1f} m  "
                  f"(k={design.conc_k_Wm_K:.2f} W/m·K, "
                  f"ρcp={design.conc_rho_kg_m3*design.conc_cp_J_kg_K/1e6:.2f} MJ/m³·K)")
            print(f"  Floor area      : {design.floor_area_m2:.0f} m²")
            print(f"  Sump geometry   : {design.sump_area_m2:.1f} m² × {design.sump_depth_m:.2f} m  "
                  f"(capacity={design.sump_capacity_m3:.2f} m³)")
            print(f"  P₀              : {ic.P0_Pa/1e3:.1f} kPa")
            print(f"  T₀              : {ic.T0_K - T_REF:.1f} °C")
            print(f"  M_air           : {self.M_air0:,.1f} kg")
            print(f"  M_steam₀        : {self.M_stm0:.2f} kg  (RH₀ = {ic.RH0*100:.0f}%)")
            print(f"  Condensation    : {design.condensation_model.upper()}"
                  f"  (multiplier={design.uchida_multiplier:.2f})")
            if design.pcc_fraction > 0:
                print(f"  PCC fraction    : {design.pcc_fraction:.2f}  "
                      f"(outer T = {design.T_outer_C:.1f} °C)")
            else:
                print(f"  Outer BC        : adiabatic  (pcc_fraction = 0)")
            print(f"{'='*65}\n")

    # ─────────────────────────────────────────────────────────────────────────
    # Main integration loop
    # ─────────────────────────────────────────────────────────────────────────

    def run(self, t_end: float) -> dict:
        """
        Integrate the containment transient from t=0 to t_end [s].

        Returns
        -------
        results : dict with numpy arrays indexed by time step:
            't'           — time                                [s]
            'P_kPa'       — total pressure                     [kPa]
            'T_C'         — gas temperature                    [°C]
            'T_wall_C'    — inner wall temperature             [°C]
            'V_liq_m3'    — accumulated liquid volume          [m³]
            'x_nc'        — non-condensable mass fraction      [-]
            'M_air_kg'    — air mass (conserved constant)      [kg]
            'M_steam_kg'  — steam mass remaining in gas phase  [kg]
            'M_cond_kg'   — total condensed mass in sump       [kg]
            'mdot_cond'   — instantaneous condensation rate    [kg/s]
            'q_cond'      — wall condensation heat flux        [W/m²]
            'Q_wall_W'    — total wall heat removal            [W]
            'mdot_in'     — source mass flow rate              [kg/s]
            'Hdot_in_W'   — source enthalpy flux = ṁ·h_src    [W]
        """
        dt   = self.dt
        N    = int(math.ceil(t_end / dt)) + 1
        time = np.linspace(0.0, t_end, N)
        dt   = float(time[1] - time[0]) if N > 1 else dt

        design = self.design
        ic     = self.ic

        # ── allocate output arrays ────────────────────────────────────────────
        P        = np.zeros(N)    # Pa
        T_gas    = np.zeros(N)    # K
        T_wall   = np.zeros(N)    # K
        h_mix    = np.zeros(N)    # J/kg  (mixture specific enthalpy, gas phase)
        V_liq    = np.zeros(N)    # m³    (liquid volume in sump)
        x_nc     = np.zeros(N)    # –     (non-condensable mass fraction in gas)
        M_gas    = np.zeros(N)    # kg    (total gas-phase mass)
        M_cond   = np.zeros(N)    # kg    (total condensed mass accumulated)
        mdot_cond= np.zeros(N)    # kg/s
        q_cond   = np.zeros(N)    # W/m²
        Q_wall   = np.zeros(N)    # W
        mdot_in  = np.zeros(N)    # kg/s
        Hdot_in  = np.zeros(N)    # W  — source enthalpy flux = ṁ_in · h_src
        mdot_h2  = np.zeros(N)    # kg/s — H2 source rate
        M_steam  = np.zeros(N)    # kg — steam mass remaining in gas phase
        M_air_arr= np.zeros(N)    # kg — air mass (conserved, constant)
        M_h2_arr = np.zeros(N)    # kg — H2 mass in gas phase (accumulates)
        T_pool   = np.zeros(N)    # K  — sump pool temperature
        Q_pool_W = np.zeros(N)    # W  — pool surface condensation heat removal
        Q_lhsi_W = np.zeros(N)    # W  — kept as zeros for backward-compatible results dict
        T_conc   = np.zeros(N)    # K  — concrete surface temperature (near gap)
        T_bsmt   = np.zeros(N)    # K  — basemat surface temperature

        # ── set initial conditions ────────────────────────────────────────────
        P[0]        = ic.P0_Pa
        T_gas[0]    = ic.T0_K
        T_wall[0]   = ic.T_wall0_K
        h_mix[0]    = self._h0
        M_gas[0]    = self.M_air0 + self.M_stm0
        x_nc[0]     = self.M_air0 / max(M_gas[0], 1e-9)
        V_liq[0]    = 0.0
        M_cond[0]   = 0.0
        M_steam[0]  = self.M_stm0
        M_air_arr[0]= self.M_air0
        M_h2_arr[0] = 0.0
        T_pool[0]   = ic.T_wall0_K
        T_conc[0]   = ic.T_wall0_K
        T_bsmt[0]   = ic.T_wall0_K

        # ── working state ──────────────────────────────────────────────────────
        _M_air   = self.M_air0    # conserved: air never condenses
        _M_H2    = 0.0            # cumulative H2 mass in gas space [kg]
        _M_cond  = 0.0            # cumulative condensed mass
        _V_liq   = 0.0            # cumulative liquid volume

        _last_print = -30.0

        # ── time march ────────────────────────────────────────────────────────
        for t in range(N - 1):

            # Current state
            _P  = float(P[t])
            _h  = float(h_mix[t])
            _Mg = float(M_gas[t])
            _xn = float(x_nc[t])
            _Tk = float(T_gas[t])
            _Tw = float(T_wall[t])

            # Guard: keep within physical range.
            # The upper pressure bound is set to 10 MPa — well above any PWR
            # containment design pressure (~0.5 MPa) but below the region where
            # steam table lookups become unreliable.  This allows the calculation
            # to capture above-design-pressure transients without clamping.
            _P  = float(np.clip(_P, 1e3, 10e6))      # 1 kPa – 10 MPa
            _h  = float(np.clip(_h, 1e3, 5e6))        # 1 – 5000 kJ/kg
            _xn = float(np.clip(_xn, 0.01, 0.9999))   # always some air

            Dt  = time[t+1] - time[t]

            # ── Sub-step during high-flux blowdown ────────────────────────────
            # When the source mass flow adds more than 5% of the current gas
            # mass in a single timestep the explicit solver can diverge.
            # Automatically sub-step to keep Δm/m < 5% per sub-step.
            _mdot_chk, _, _, _ = self.source.at(time[t])
            if not np.isfinite(_mdot_chk):
                _mdot_chk = 0.0
            _n_sub = max(1, int(math.ceil(_mdot_chk * Dt
                                          / max(_Mg * 0.05, 1.0))))
            _n_sub = min(_n_sub, 50)   # cap sub-steps for performance
            Dt_sub = Dt / _n_sub

            # Accumulate sub-step results into working variables
            _P_sub  = _P
            _h_sub  = _h
            _Mg_sub = _Mg
            _xn_sub = _xn
            _Tw_sub = _Tw
            _M_H2_sub   = _M_H2
            _M_cond_sub = _M_cond
            _V_liq_sub  = _V_liq

            for _ss in range(_n_sub):
                t_ss = time[t] + _ss * Dt_sub

                # ── source term ───────────────────────────────────────────────
                _mdot, _Qdot, _h_src, _mdot_h2_ss = self.source.at(t_ss)

                # ── wall condensation ─────────────────────────────────────────
                _Tk_ss = float(T_gas[t]) if _ss == 0 else _T_est_ss
                _q = wall_condensation_heat_flux(
                    T_bulk_K          = _Tk_ss,
                    T_wall_K          = _Tw_sub,
                    P_total_Pa        = _P_sub,
                    x_nc              = _xn_sub,
                    L_wall            = design.wall_char_height_m,
                    uchida_multiplier = design.uchida_multiplier,
                    model             = design.condensation_model,
                )
                _Q_wall    = _q * design.wall_surf_area_m2
                _hfg       = max(_hfg_Jkg(_P_sub * 1e-3 * (1.0 - _xn_sub)), 1e4)
                _mdot_cond = _Q_wall / _hfg

                # ── basemat condensation/convection (always active) ───────────
                # Gas-to-floor heat transfer: use same Uchida HTC but with
                # floor characteristic length (1 m, horizontal surface).
                # Basemat removes heat from the gas regardless of liquid level.
                _q_base = wall_condensation_heat_flux(
                    T_bulk_K          = _Tk_ss,
                    T_wall_K          = self._basemat.T_inner,
                    P_total_Pa        = _P_sub,
                    x_nc              = _xn_sub,
                    L_wall            = 1.0,
                    uchida_multiplier = design.uchida_multiplier,
                    model             = design.condensation_model,
                )
                _Q_base    = _q_base * design.floor_area_m2
                _mdot_cond_base = _Q_base / _hfg

                # Advance basemat temperature
                self._basemat.step(q_inner=_q_base, dt=Dt_sub)

                # ── wall temperature advance (multi-layer liner) ──────────────
                _Tw_sub = self._wall.step(q_inner=_q, dt=Dt_sub)

                # PCC outer heat removal — heat that has conducted all the way
                # through the wall and is removed at the outer surface.
                # This is additional heat removal beyond what enters the wall
                # from the gas side; it only matters once the wall heats up
                # and a temperature gradient is established to the cooled surface.
                _Q_pcc = (self._wall.q_outer_W_m2()
                          * design.wall_surf_area_m2
                          * design.pcc_fraction)

                # ── pool surface condensation ─────────────────────────────────
                _T_sat_ss = _tsat_K(_P_sub * (1.0 - _xn_sub) * 1e-3)
                (_Q_pool_ss, _mdot_pool_ss, _T_pool_ss) = self._pool.step(
                    dt            = Dt_sub,
                    mdot_cond_in  = _mdot_cond + _mdot_cond_base,
                    T_sat_K       = _T_sat_ss,
                    T_bulk_K      = _Tk_ss,
                    P_total_Pa    = _P_sub,
                    x_nc          = _xn_sub,
                    floor_area_m2 = design.floor_area_m2,
                    V_liq_m3      = _V_liq_sub,
                    uchida_mult   = design.uchida_multiplier,
                    model         = design.condensation_model,
                )

                # Total condensation = wall + basemat + pool surface
                _mdot_cond_total = _mdot_cond + _mdot_cond_base + _mdot_pool_ss
                # Total heat removal from gas space (wall condensation +
                # basemat condensation + pool surface condensation + PCC outer)
                _Q_total_removal = _Q_wall + _Q_base + _Q_pool_ss + _Q_pcc

                # ── 2×2 matrix solve ──────────────────────────────────────────
                _x_h2  = _M_H2_sub / max(_Mg_sub, 1e-9)
                ppd    = _mixture_partials(_P_sub, _Tk_ss, _xn_sub, _Mg_sub,
                                           design.air_volume_m3, x_h2=_x_h2)
                A = np.array([
                    [_Mg_sub * ppd['du_dP'],  _Mg_sub * ppd['du_dh']],
                    [_Mg_sub * ppd['dv_dP'],  _Mg_sub * ppd['dv_dh']],
                ])
                _v_mix    = design.air_volume_m3 / max(_Mg_sub, 1e-6)
                _u_mix    = _h_sub - _P_sub * _v_mix
                _h_h2_src = CP_H2 * (_Tk_ss - T_REF)
                _net_mdot = _mdot - _mdot_cond_total + _mdot_h2_ss
                b = np.array([
                    _mdot             * (_h_src    - _u_mix)
                    + _mdot_h2_ss     * (_h_h2_src - _u_mix)
                    - _mdot_cond_total * _hfg
                    - _Q_total_removal
                    + _Qdot,
                    -_v_mix * _net_mdot,
                ])
                try:
                    delta  = np.linalg.solve(A, b * Dt_sub)
                    P_ss   = float(_P_sub + np.real(delta[0]))
                    h_ss   = float(_h_sub + np.real(delta[1]))
                    if not (np.isfinite(P_ss) and np.isfinite(h_ss)
                            and P_ss > 0 and h_ss > 0):
                        raise ValueError("unphysical")
                except (np.linalg.LinAlgError, ValueError):
                    P_ss = _P_sub
                    h_ss = _h_sub + Dt_sub * b[0] / max(_Mg_sub, 1.0)
                    h_ss = float(np.clip(h_ss, 1e3, 5e6))

                # ── mass balance ──────────────────────────────────────────────
                _M_H2_sub    += _mdot_h2_ss * Dt_sub
                _Mg_sub_new   = _Mg_sub + Dt_sub * _net_mdot
                _Mg_sub_new   = max(_Mg_sub_new, _M_air + _M_H2_sub)
                _M_cond_sub  += _mdot_cond_total * Dt_sub
                _V_liq_sub   += _mdot_cond_total * Dt_sub / RHO_WATER

                # ── temperature Newton solve ──────────────────────────────────
                P_ss    = float(np.clip(P_ss, 1e3, 10e6))
                h_ss    = float(np.clip(h_ss, 1e3, 5e6))
                _xn_new = (_M_air + _M_H2_sub) / max(_Mg_sub_new, _M_air + _M_H2_sub)
                _NC_coeff = (_M_air * R_AIR + _M_H2_sub * R_H2) / design.air_volume_m3
                _T_est_ss = _Tk_ss
                for _nit in range(30):
                    _Psat_est = _psat_kPa(_T_est_ss) * 1e3
                    _Pnc_est  = _NC_coeff * _T_est_ss
                    _f        = P_ss - _Psat_est - _Pnc_est
                    _dPsat_dT = (_psat_kPa(_T_est_ss + 0.5) -
                                 _psat_kPa(_T_est_ss - 0.5)) * 1e3
                    _df       = -_dPsat_dT - _NC_coeff
                    _dT_ss    = -_f / (_df if abs(_df) > 1e-6 else 1e-6)
                    _T_est_ss += _dT_ss
                    _T_est_ss  = float(np.clip(_T_est_ss, 273.16, 700.0))
                    if abs(_dT_ss) < 0.01:
                        break
                if _psat_kPa(_T_est_ss) * 1e3 >= P_ss:
                    _T_est_ss = P_ss / max(_NC_coeff, 1e-6)
                    _T_est_ss = float(np.clip(_T_est_ss, 273.16, 700.0))

                # advance sub-step state
                _P_sub  = P_ss
                _h_sub  = h_ss
                _Mg_sub = _Mg_sub_new
                _xn_sub = _xn_new

            # ── propagate sub-stepped state to next outer step ────────────────
            P_new     = _P_sub
            h_new     = _h_sub
            T_new     = _T_est_ss
            M_gas_new = _Mg_sub
            _M_H2     = _M_H2_sub
            _M_cond   = _M_cond_sub
            _V_liq    = _V_liq_sub
            _xn_new   = _xn_sub
            _Tw_new   = _Tw_sub

            # ── record diagnostics at the outer step midpoint (first sub-step) ─
            _mdot_rec, _, _h_src_rec, _mdot_h2_rec = self.source.at(time[t])
            mdot_in[t]   = _mdot_rec
            Hdot_in[t]   = _mdot_rec * _h_src_rec
            mdot_h2[t]   = _mdot_h2_rec
            q_cond[t]    = _q
            Q_wall[t]    = _Q_wall + _Q_base
            mdot_cond[t] = _mdot_cond_total
            Q_pool_W[t]  = _Q_pool_ss

            # ── store to arrays ────────────────────────────────────────────────
            P[t+1]        = P_new
            h_mix[t+1]    = h_new
            T_gas[t+1]    = T_new
            T_wall[t+1]   = _Tw_new
            x_nc[t+1]     = _xn_new
            M_gas[t+1]    = M_gas_new
            V_liq[t+1]    = _V_liq
            M_cond[t+1]   = _M_cond
            M_steam[t+1]  = max(0.0, M_gas_new - _M_air - _M_H2)
            M_air_arr[t+1]= _M_air
            M_h2_arr[t+1] = _M_H2
            T_pool[t+1]   = float(self._pool.T_pool)
            T_conc[t+1]   = float(self._wall.T_conc_surface)
            T_bsmt[t+1]   = float(self._basemat.T_inner)

            # Progress print
            if self.verbose and (time[t] - _last_print) >= 60.0:
                _last_print = time[t]
                print(f"  t={time[t]:7.1f} s  |  P={P_new/1e3:6.1f} kPa  "
                      f"T={T_new-T_REF:5.1f} °C  "
                      f"T_wall={_Tw_new-T_REF:5.1f} °C  "
                      f"V_liq={_V_liq:.2f} m³  "
                      f"x_nc={_xn_new:.3f}")

            # Design-pressure warning
            if P_new > self.design.P_design_kPa * 1e3 * 0.95:
                if self.verbose:
                    print(f"  *** WARNING: P={P_new/1e3:.1f} kPa approaching "
                          f"design pressure {self.design.P_design_kPa:.1f} kPa ***")

        # fill last step
        mdot_in[-1]   = mdot_in[-2]
        Hdot_in[-1]   = Hdot_in[-2]
        mdot_h2[-1]   = mdot_h2[-2]
        mdot_cond[-1] = mdot_cond[-2]
        Q_wall[-1]    = Q_wall[-2]
        q_cond[-1]    = q_cond[-2]
        Q_pool_W[-1]  = Q_pool_W[-2]
        Q_lhsi_W[-1]  = Q_lhsi_W[-2]

        results = {
            't':           time,
            'P_kPa':       P / 1e3,
            'T_C':         T_gas - T_REF,
            'T_wall_C':    T_wall - T_REF,
            'T_conc_C':    T_conc - T_REF,
            'T_pool_C':    T_pool - T_REF,
            'T_bsmt_C':    T_bsmt - T_REF,
            'V_liq_m3':    V_liq,
            'x_nc':        x_nc,
            'M_air_kg':    M_air_arr,
            'M_steam_kg':  M_steam,
            'M_h2_kg':     M_h2_arr,
            'M_cond_kg':   M_cond,
            'mdot_cond':   mdot_cond,
            'q_cond':      q_cond,
            'Q_wall_W':    Q_wall,
            'Q_pool_W':    Q_pool_W,
            'Q_lhsi_W':    Q_lhsi_W,   # zeros — LHSI not modelled in containment
            'mdot_in':     mdot_in,
            'mdot_h2':     mdot_h2,
            'Hdot_in_W':   Hdot_in,
        }
        return results




# ─────────────────────────────────────────────────────────────────────────────
# Physics-audit V4 solver override
# ─────────────────────────────────────────────────────────────────────────────

def _gas_U_lumped(T_K, M_air, M_steam, M_h2):
    """Approximate absolute gas internal energy [J]."""
    L0 = 2.50e6
    return (M_air * CV_AIR * (T_K - T_REF)
            + M_h2 * CV_H2 * (T_K - T_REF)
            + M_steam * (L0 + CV_AIR * 0.0 + CP_STEAM * (T_K - T_REF)))

def _gas_T_from_U_lumped(U, M_air, M_steam, M_h2):
    L0 = 2.50e6
    denom = M_air * CV_AIR + M_h2 * CV_H2 + M_steam * CP_STEAM
    if denom <= 1e-9:
        return T_REF + 25.0
    T = T_REF + (U - M_steam * L0) / denom
    return float(np.clip(T, 250.0, 900.0))

def _run_v4(self, t_end: float) -> dict:
    """
    Replacement lumped-parameter containment integrator.

    Key changes relative to the original prototype:
      * pressure is calculated from species partial pressures, not saturation closure;
      * condensation is limited to actual supersaturation inventory;
      * source mass is flashed into vapor/liquid using source enthalpy;
      * source extrapolation beyond the final table point is zero;
      * sensible heat transfer between gas and liner/floor/liquid is bidirectional;
      * late-time condensation is deadbanded/relaxed to reduce saturation chatter.
    """
    dt0 = self.dt
    N = int(math.ceil(t_end / dt0)) + 1
    time = np.linspace(0.0, t_end, N)
    dt = float(time[1] - time[0]) if N > 1 else dt0
    design, ic = self.design, self.ic
    V = design.air_volume_m3

    P=np.zeros(N); T=np.zeros(N); T_wall=np.zeros(N); T_conc=np.zeros(N); T_pool=np.zeros(N); T_bsmt=np.zeros(N)
    V_liq=np.zeros(N); x_nc=np.zeros(N); M_air_arr=np.zeros(N); M_steam_arr=np.zeros(N); M_h2_arr=np.zeros(N); M_cond=np.zeros(N)
    mdot_cond=np.zeros(N); q_cond=np.zeros(N); Q_wall=np.zeros(N); Q_pool_W=np.zeros(N); mdot_in=np.zeros(N); mdot_h2_arr=np.zeros(N); Hdot_in=np.zeros(N)
    Q_floor_dry_W=np.zeros(N); A_floor_wet_arr=np.zeros(N); A_floor_dry_arr=np.zeros(N)
    V_sump_arr=np.zeros(N); V_floor_arr=np.zeros(N); sump_level_arr=np.zeros(N); floor_flood_level_arr=np.zeros(N)
    Q_spray_W=np.zeros(N); mdot_spray_arr=np.zeros(N); mdot_spray_cond_arr=np.zeros(N)

    # Initial vapor from RH; air mass from dry partial pressure.
    # Sump/floor model: liquid fills the engineered sump first, then overflows
    # and spreads across the containment floor.  The sump free surface and the
    # wetted floor film participate in condensation/convective heat transfer.
    cond_relax_tau_s = 30.0        # late-time condensation relaxation time constant
    cond_deadband = 0.005          # 0.5% supersaturation deadband
    cond_relax_start_s = 1200.0    # preserve early blowdown response; relax after peak

    Ps0 = ic.RH0 * _psat_kPa(ic.T0_K) * 1e3
    M_air = max((ic.P0_Pa - Ps0) * V / (R_AIR * ic.T0_K), 0.0)
    M_steam = max(Ps0 * V / (R_STEAM * ic.T0_K), 0.0)
    M_h2 = 0.0
    Mliq = 0.0
    Mcond_total = 0.0
    Tgas = ic.T0_K
    Ugas = _gas_U_lumped(Tgas, M_air, M_steam, M_h2)
    Tpool = ic.T_wall0_K
    Upool = 0.0
    cpw = 4186.0

    def pressure(TK, Ms, Mh):
        return (M_air*R_AIR + Ms*R_STEAM + Mh*R_H2) * TK / V

    def steam_sat_mass(TK):
        return max(_psat_kPa(TK)*1e3 * V/(R_STEAM*TK), 0.0)

    # store initial
    P[0]=pressure(Tgas,M_steam,M_h2); T[0]=Tgas; T_wall[0]=ic.T_wall0_K; T_conc[0]=ic.T_wall0_K; T_pool[0]=Tpool; T_bsmt[0]=ic.T_wall0_K
    _dist0 = design.liquid_distribution(0.0)
    V_sump_arr[0]=_dist0['V_sump_m3']; V_floor_arr[0]=_dist0['V_floor_m3']
    sump_level_arr[0]=_dist0['sump_level_m']; floor_flood_level_arr[0]=_dist0['floor_flood_level_m']
    A_floor_wet_arr[0]=_dist0['A_floor_wet_m2']; A_floor_dry_arr[0]=_dist0['A_floor_dry_m2']
    M_air_arr[0]=M_air; M_steam_arr[0]=M_steam; M_h2_arr[0]=M_h2; x_nc[0]=(M_air+M_h2)/max(M_air+M_h2+M_steam,1e-9)

    _last_print=-1e9
    for i in range(N-1):
        t = time[i]
        mdot, Qdot, hsrc, mdoth2 = self.source.at(t)
        mdot_in[i]=mdot; mdot_h2_arr[i]=mdoth2; Hdot_in[i]=mdot*hsrc
        Pnow = pressure(Tgas,M_steam,M_h2)
        # Flash incoming water at containment pressure.  Enthalpy values are relative to sat-liquid/vapor helpers.
        hf = _hL_Jkg(max(Pnow/1e3, 0.1)); hg = _hg_Jkg(max(Pnow/1e3, 0.1)); hfg=max(hg-hf,1e5)
        xflash = float(np.clip((hsrc-hf)/hfg if mdot>0 else 0.0, 0.0, 1.0))
        m_in = mdot*dt; m_v = m_in*xflash; m_l = m_in-m_v
        # Add source species.
        M_steam += m_v
        M_h2 += mdoth2*dt
        # Energy split: vapor fraction carries saturated vapor enthalpy hg into
        # the gas space; liquid fraction carries saturated liquid enthalpy hf into
        # the sump.  Using the bulk mixture enthalpy hsrc for the vapor fraction
        # underestimates the gas energy when the source is a wet mixture (hsrc < hg),
        # driving the gas temperature unphysically low.
        Ugas += m_v * hg + Qdot*dt
        # Liquid source goes to sump/pool, not directly to gas pressure.
        if m_l > 0:
            Mliq += m_l
            Upool += m_l * hf

        # Containment sprays: liquid injected into the gas volume from RWST or
        # recirculation.  Sprays are not a boundary surface; they are a direct
        # atmospheric heat/mass transfer path.  The injected liquid is added to
        # the sump after droplet-atmosphere interaction.  Spray cooling is
        # limited by the finite heat capacity of the injected water and by
        # available supersaturated steam inventory.
        Qspray_step = 0.0
        mspray_cond_step = 0.0
        mdot_spray, Tspray = design.spray_at(t)
        mdot_spray_arr[i] = mdot_spray
        m_spray = mdot_spray * dt
        if m_spray > 0.0:
            eff_spray = float(np.clip(design.spray_thermal_effectiveness(Tgas, Pnow), 0.0, 1.0))
            Tspray = float(np.clip(Tspray, T_REF, max(Tgas, T_REF)))
            # Add injected water sensible enthalpy to the internal liquid inventory.
            Mliq += m_spray
            Upool += m_spray * cpw * (Tspray - T_REF)

            # Total droplet cooling capacity as it approaches current gas temperature.
            Qcap = m_spray * cpw * max(Tgas - Tspray, 0.0) * eff_spray
            if Qcap > 0.0:
                hfg_sp = max(_hfg_Jkg(max(_psat_kPa(Tgas), 0.1)), 1e5)
                Mexcess_sp = max(M_steam - steam_sat_mass(Tgas) * (1.0 + cond_deadband), 0.0)
                # Do not let droplet condensation alone drive the gas below
                # the injected-water temperature.  This keeps the reduced-order
                # droplet model bounded without solving droplet transport.
                U_floor = _gas_U_lumped(Tspray, M_air, M_steam, M_h2)
                dm_energy_cap = max((Ugas - U_floor) / hfg_sp, 0.0)
                dm_sp = min(Mexcess_sp, Qcap / hfg_sp, dm_energy_cap)
                if dm_sp > 0.0:
                    M_steam -= dm_sp
                    Mliq += dm_sp
                    Mcond_total += dm_sp
                    mspray_cond_step = dm_sp
                    Ugas -= dm_sp * hfg_sp
                    Upool += dm_sp * _hL_Jkg(max(_psat_kPa(Tgas), 0.1))
                    Qspray_step += dm_sp * hfg_sp / dt
                    Qcap -= dm_sp * hfg_sp
                    Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)

                # Remaining capacity provides direct sensible cooling of the gas,
                # again bounded so the gas cannot undershoot the spray water.
                if Qcap > 0.0:
                    U_floor = _gas_U_lumped(Tspray, M_air, M_steam, M_h2)
                    Qsens = min(Qcap, max(Ugas - U_floor, 0.0))
                    Ugas -= Qsens
                    Upool += Qsens
                    Qspray_step += Qsens / dt

        # Recompute gas temperature after source/spray before heat removal.
        Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)
        Pnow = pressure(Tgas,M_steam,M_h2)
        Mtot = M_air+M_steam+M_h2
        xnc = (M_air+M_h2)/max(Mtot,1e-9)

        # Determine engineered sump / containment-floor flooding state.
        # Liquid fills the sump first; only after the sump capacity is exceeded
        # does liquid spread onto the surrounding containment floor.
        Vliq_now = Mliq / RHO_WATER
        _liq_dist = design.liquid_distribution(Vliq_now)
        A_floor_wet = _liq_dist['A_floor_wet_m2']
        A_floor_dry = _liq_dist['A_floor_dry_m2']

        # Condensation only if steam exceeds saturation at current gas temperature.
        # Active condensing surfaces are: walls, dry floor/basemat, and wet floor
        # / sump film.  RWST free-surface condensation is not included here; RWST
        # enters only as an injection/source boundary unless exposed by design.
        mcond_step = 0.0; Qwall_step=0.0; qwall=0.0; Qdry_step=0.0; Qwet_step=0.0
        for _ in range(5):
            Msat = steam_sat_mass(Tgas)
            Mexcess = max(M_steam - Msat * (1.0 + cond_deadband), 0.0)
            if Mexcess <= 1e-9:
                break
            hfg = max(_hfg_Jkg(max(_psat_kPa(Tgas),0.1)), 1e5)

            candidates = []
            if Tgas > self._wall.T_inner:
                q = wall_condensation_heat_flux(Tgas, self._wall.T_inner, Pnow, xnc, design.wall_char_height_m, design.uchida_multiplier, design.condensation_model)
                if q > 0: candidates.append(('wall', design.wall_surf_area_m2, q))
            if A_floor_dry > 0 and Tgas > self._basemat.T_inner:
                q = wall_condensation_heat_flux(Tgas, self._basemat.T_inner, Pnow, xnc, 1.0, design.uchida_multiplier, design.condensation_model)
                if q > 0: candidates.append(('dry_floor', A_floor_dry, q))
            if A_floor_wet > 0 and Tgas > Tpool:
                q = wall_condensation_heat_flux(Tgas, Tpool, Pnow, xnc, 1.0, design.uchida_multiplier, design.condensation_model)
                if q > 0: candidates.append(('wet_floor', A_floor_wet, q))

            mdot_cap = sum(max(q*A/hfg, 0.0) for _,A,q in candidates)
            if t >= cond_relax_start_s:
                Mexcess_limited = Mexcess * (1.0 - math.exp(-dt / max(cond_relax_tau_s, 1e-6)))
            else:
                Mexcess_limited = Mexcess
            dm_total = min(Mexcess_limited, mdot_cap*dt)
            if dm_total <= 0 or mdot_cap <= 0: break

            for name, A, q in candidates:
                frac = max(q*A/hfg, 0.0) / mdot_cap
                dm = dm_total * frac
                q_eff = dm*hfg/(dt*A) if A > 0 else 0.0
                if name == 'wall':
                    self._wall.step(q_eff, dt)
                    Qwall_step += dm*hfg/dt; qwall += q_eff
                elif name == 'dry_floor':
                    self._basemat.step(q_eff, dt)
                    Qdry_step += dm*hfg/dt
                else:
                    Qwet_step += dm*hfg/dt
                    Upool += dm * _hL_Jkg(max(_psat_kPa(Tgas),0.1))
                Mliq += dm; Mcond_total += dm; mcond_step += dm

            Ugas -= dm_total*hfg
            M_steam -= dm_total
            Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)
            Pnow = pressure(Tgas,M_steam,M_h2)
            Mtot = M_air+M_steam+M_h2; xnc=(M_air+M_h2)/max(Mtot,1e-9)

        # Sensible gas-to-surface convection is bidirectional.
        # q > 0 means heat leaves the gas and enters the surface;
        # q < 0 means stored heat in the liner/floor/liquid returns to gas.
        # Condensation above remains one-way and supersaturation-limited.
        hconv_wall = 5.0
        qcv = hconv_wall * (Tgas - self._wall.T_inner)
        if abs(qcv) > 1e-12:
            self._wall.step(qcv, dt)
            Qcv_wall = qcv * design.wall_surf_area_m2
            Ugas -= Qcv_wall * dt
            Qwall_step += Qcv_wall
            qwall += qcv
            Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)

        hconv_floor = 2.0
        if A_floor_dry > 0:
            qbase = hconv_floor * (Tgas - self._basemat.T_inner)
            if abs(qbase) > 1e-12:
                self._basemat.step(qbase, dt)
                Qbase = qbase * A_floor_dry
                Ugas -= Qbase * dt
                Qdry_step += Qbase
                Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)

        if A_floor_wet > 0 and Mliq > 1e-9:
            qwet_cv = hconv_floor * (Tgas - Tpool)
            Qwet_cv = qwet_cv * A_floor_wet

            # If heat is returning from the liquid to the gas, prevent the liquid
            # inventory from being pulled below its reference sensible energy.
            if Qwet_cv < 0.0:
                Umin_pool = Mliq * cpw * (ic.T_wall0_K - T_REF)
                Qwet_cv = -min(-Qwet_cv, max((Upool - Umin_pool) / dt, 0.0))
                qwet_cv = Qwet_cv / max(A_floor_wet, 1e-9)

            if abs(Qwet_cv) > 1e-9:
                Ugas -= Qwet_cv * dt
                Upool += Qwet_cv * dt
                Qwet_step += Qwet_cv
                Tgas = _gas_T_from_U_lumped(Ugas, M_air, M_steam, M_h2)

        # Update pool temperature from energy balance
        if Mliq > 1e-9:
            Tpool = T_REF + Upool/max(Mliq*cpw,1e-9)
            Tpool = float(np.clip(Tpool, ic.T_wall0_K, 650.0))
        else:
            Tpool = ic.T_wall0_K

        Pnow = pressure(Tgas,M_steam,M_h2)
        j=i+1
        P[j]=Pnow; T[j]=Tgas; T_wall[j]=self._wall.T_inner; T_conc[j]=self._wall.T_conc_surface; T_bsmt[j]=self._basemat.T_inner; T_pool[j]=Tpool
        V_liq[j]=Mliq/RHO_WATER; x_nc[j]=(M_air+M_h2)/max(M_air+M_steam+M_h2,1e-9)
        _liq_dist = design.liquid_distribution(V_liq[j])
        V_sump_arr[j]=_liq_dist['V_sump_m3']; V_floor_arr[j]=_liq_dist['V_floor_m3']
        sump_level_arr[j]=_liq_dist['sump_level_m']; floor_flood_level_arr[j]=_liq_dist['floor_flood_level_m']
        M_air_arr[j]=M_air; M_steam_arr[j]=M_steam; M_h2_arr[j]=M_h2; M_cond[j]=Mcond_total
        mdot_cond[i]=mcond_step/dt; q_cond[i]=qwall; Q_wall[i]=Qwall_step; Q_pool_W[i]=Qwet_step
        Q_floor_dry_W[i]=Qdry_step; A_floor_wet_arr[i]=A_floor_wet; A_floor_dry_arr[i]=A_floor_dry
        Q_spray_W[i]=Qspray_step; mdot_spray_cond_arr[i]=mspray_cond_step/dt
        if self.verbose and (t-_last_print)>=600:
            _last_print=t
            print(f"FLARECON_SIMTIME {t:.1f}", flush=True)
            print(f"  t={t:7.1f} s  |  P={Pnow/1e3:6.1f} kPa  T={Tgas-T_REF:6.1f} °C  T_wall={self._wall.T_inner-T_REF:5.1f} °C  V_liq={V_liq[j]:.2f} m³  x_nc={x_nc[j]:.3f}", flush=True)

    for arr in (mdot_in,Hdot_in,mdot_h2_arr,mdot_cond,q_cond,Q_wall,Q_pool_W,Q_floor_dry_W,A_floor_wet_arr,A_floor_dry_arr,Q_spray_W,mdot_spray_arr,mdot_spray_cond_arr):
        arr[-1]=arr[-2]
    return {'t':time,'P_kPa':P/1e3,'T_C':T-T_REF,'T_wall_C':T_wall-T_REF,'T_conc_C':T_conc-T_REF,'T_pool_C':T_pool-T_REF,'T_bsmt_C':T_bsmt-T_REF,'V_liq_m3':V_liq,'V_sump_m3':V_sump_arr,'V_floor_m3':V_floor_arr,'sump_level_m':sump_level_arr,'floor_flood_level_m':floor_flood_level_arr,'x_nc':x_nc,'M_air_kg':M_air_arr,'M_steam_kg':M_steam_arr,'M_h2_kg':M_h2_arr,'M_cond_kg':M_cond,'mdot_cond':mdot_cond,'q_cond':q_cond,'Q_wall_W':Q_wall,'Q_pool_W':Q_pool_W,'Q_floor_dry_W':Q_floor_dry_W,'A_floor_wet_m2':A_floor_wet_arr,'A_floor_dry_m2':A_floor_dry_arr,'Q_lhsi_W':np.zeros(N),'Q_spray_W':Q_spray_W,'mdot_spray':mdot_spray_arr,'mdot_spray_cond':mdot_spray_cond_arr,'mdot_in':mdot_in,'mdot_h2':mdot_h2_arr,'Hdot_in_W':Hdot_in}

# Use the V5 physics-audit solver for command-line runs below.
ContainmentSimulator.run = _run_v4

# ─────────────────────────────────────────────────────────────────────────────
# Post-processing / plotting
# ─────────────────────────────────────────────────────────────────────────────

def plot_results(results: dict, title: str = "PWR Dry Containment Response",
                 outfile: str = "containment_out.png",
                 floor_area_m2: float = 600.0) -> None:
    """
    Generate a 3×2 summary plot of the containment response.

    Figures of merit (primary):
        1. Containment pressure [kPa]
        2. Gas and wall temperature [°C]
    Supporting diagnostics:
        3. Liquid volume accumulated [m³]
        4. Non-condensable quality x_nc [-]
        5. Wall condensation heat flux [W/m²]
        6. Mass inventory: gas / condensed [kg]
    """
    if not _MPL:
        print("matplotlib not available — skipping plots.")
        return

    t   = results['t'] / 3600.0   # convert to hours for readability
    lw  = 1.8
    clr = ['#1f77b4', '#d62728', '#2ca02c', '#ff7f0e']

    fig, axes = plt.subplots(3, 2, figsize=(13, 11))
    fig.suptitle(title, fontsize=13, fontweight='bold')

    ax = axes[0, 0]
    ax.plot(t, results['P_kPa'], color=clr[0], lw=lw)
    ax.set_xlabel('Time [hr]');  ax.set_ylabel('Pressure [kPa]')
    ax.set_title('Containment Pressure');  ax.grid(True)

    ax = axes[0, 1]
    ax.plot(t, results['T_C'],      color=clr[0], lw=lw, label='Gas')
    ax.plot(t, results['T_wall_C'], color=clr[1], lw=lw, linestyle='--',
            label='Liner inner')
    ax.plot(t, results['T_conc_C'], color=clr[2], lw=lw, linestyle=':',
            label='Concrete surface')
    ax.plot(t, results['T_bsmt_C'], color=clr[3], lw=lw, linestyle='-.',
            label='Basemat surface')
    ax.plot(t, results['T_pool_C'], color='purple', lw=lw, linestyle=(0,(3,1,1,1)),
            label='Wet floor / sump liquid')
    ax.set_xlabel('Time [hr]');  ax.set_ylabel('Temperature [°C]')
    ax.set_title('Temperature');  ax.legend(fontsize=7);  ax.grid(True)

    ax = axes[1, 0]
    liq_level = results.get('floor_flood_level_m', results['V_liq_m3'] / max(floor_area_m2, 1.0))   # m
    ax.plot(t, results['V_liq_m3'], color=clr[2], lw=lw, label='Volume')
    ax.set_xlabel('Time [hr]')
    ax.set_ylabel('Volume [m³]', color=clr[2])
    ax.tick_params(axis='y', labelcolor=clr[2])
    ax.set_title('Accumulated Liquid: Volume & Sump Level')
    ax.grid(True)
    ax2 = ax.twinx()
    ax2.plot(t, liq_level, color=clr[3], lw=lw, linestyle='--', label='Level')
    ax2.set_ylabel('Sump Level [m]', color=clr[3])
    ax2.tick_params(axis='y', labelcolor=clr[3])
    # combined legend
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc='lower right')

    ax = axes[1, 1]
    ax.plot(t, results['x_nc'], color=clr[3], lw=lw)
    ax.set_xlabel('Time [hr]');  ax.set_ylabel('x_nc  [—]')
    ax.set_title('Non-condensable Quality');  ax.grid(True)
    ax.set_ylim(0, 1.05)

    ax = axes[2, 0]
    ax.plot(t, results['q_cond'] / 1e3, color=clr[0], lw=lw)
    ax.set_xlabel('Time [hr]');  ax.set_ylabel('Heat flux [kW/m²]')
    ax.set_title('Wall Condensation Heat Flux');  ax.grid(True)

    ax = axes[2, 1]
    ax.plot(t, results['M_air_kg'],   color=clr[0], lw=lw, label='Air (conserved)')
    ax.plot(t, results['M_steam_kg'], color=clr[1], lw=lw, linestyle='--',
            label='Steam in gas phase')
    ax.plot(t, results['M_h2_kg'],    color=clr[3], lw=lw, linestyle=':',
            label='Hydrogen')
    ax.plot(t, results['M_cond_kg'],  color=clr[2], lw=lw, linestyle='-.',
            label='Condensed (sump)')
    ax.set_xlabel('Time [hr]');  ax.set_ylabel('Mass [kg]')
    ax.set_title('Mass Inventory');  ax.legend(fontsize=7);  ax.grid(True)

    plt.tight_layout()
    plt.savefig(outfile, dpi=150)
    plt.close(fig)
    print(f"  Plot saved → {outfile}")


def plot_heat_removal(results: dict,
                      title:   str = "Containment Heat Removal",
                      outfile: str = "containment_heat.png") -> None:
    """
    Plot the three heat removal mechanisms vs time:
      1. Wall condensation  (liner surface)
      2. Pool surface condensation
    plus total heat removal and containment pressure on a second axis.
    """
    if not _MPL:
        return

    t   = results['t'] / 3600.0
    lw  = 1.8

    Q_wall  = results['Q_wall_W']  / 1e6   # MW
    Q_pool  = results['Q_pool_W']  / 1e6
    Q_spray = results.get('Q_spray_W', np.zeros_like(results['t'])) / 1e6
    Q_total = Q_wall + Q_pool + Q_spray

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8), sharex=True)
    fig.suptitle(title, fontsize=12, fontweight='bold')

    # Heat removal
    ax1.plot(t, Q_wall,  color='steelblue',   lw=lw, label='Wall condensation')
    ax1.plot(t, Q_pool,  color='seagreen',    lw=lw, label='Wet-floor condensation')
    ax1.plot(t, Q_spray, color='purple',      lw=lw, label='Containment sprays')
    ax1.plot(t, Q_total, color='black',       lw=lw, linestyle='--', label='Total')
    ax1.set_ylabel('Heat removal [MW]')
    ax1.set_title('Heat Removal Mechanisms')
    ax1.legend(fontsize=9);  ax1.grid(True)

    # Pressure
    ax2.plot(t, results['P_kPa'], color='crimson', lw=lw)
    ax2.set_xlabel('Time [hr]')
    ax2.set_ylabel('Pressure [kPa]')
    ax2.set_title('Containment Pressure')
    ax2.grid(True)

    plt.tight_layout()
    plt.savefig(outfile, dpi=150)
    plt.close(fig)
    print(f"  Heat removal plot saved → {outfile}")


def plot_shapiro(results: dict,
                 title:  str = "Shapiro Steam–Air–H₂ Diagram",
                 outfile: str = "containment_shapiro.png") -> None:
    """
    Generate a Shapiro (1957) ternary composition diagram for the
    steam – air – hydrogen mixture in the containment gas space.

    Axes (mole fractions, summing to 1):
        Bottom  : H₂  mole fraction  (left→right)
        Left    : Steam mole fraction (bottom→top)
        Right   : Air mole fraction   (top→bottom, derived)

    Flammability regions (per NUREG/CR-6509 / Shapiro & Moffette):
        • LFL boundary  : H₂ = 4 vol% (below this: too lean to burn)
        • UFL boundary  : H₂ = 75 vol%
        • Steam inerting: steam ≥ 55 vol% (mixture non-flammable)
        • Detonation peninsula: approximate H₂ 18–59 vol% at low steam

    The transient trajectory of the containment atmosphere is overlaid,
    colour-coded by time, with start (○) and end (★) markers.
    """
    if not _MPL:
        print("matplotlib not available — skipping Shapiro plot.")
        return

    # ── Convert mass fractions → mole fractions ───────────────────────────────
    M_air   = results['M_air_kg']
    M_steam = results['M_steam_kg']
    M_h2    = results['M_h2_kg']
    M_total = M_air + M_steam + M_h2
    M_total = np.where(M_total < 1e-9, 1.0, M_total)

    # Moles of each species
    n_air   = M_air   / MW_AIR
    n_steam = M_steam / MW_H2O
    n_h2    = M_h2    / MW_H2
    n_total = n_air + n_steam + n_h2
    n_total = np.where(n_total < 1e-12, 1.0, n_total)

    y_h2    = n_h2    / n_total   # H₂ mole fraction
    y_steam = n_steam / n_total   # steam mole fraction
    y_air   = n_air   / n_total   # air mole fraction  (= 1 − y_h2 − y_steam)

    # ── Ternary → Cartesian projection ────────────────────────────────────────
    # Standard equilateral ternary: vertices at
    #   H2    (bottom-left)  = (0, 0)
    #   Air   (bottom-right) = (1, 0)
    #   Steam (top)          = (0.5, √3/2)
    def _tern2cart(y_h2_, y_air_, y_stm_):
        x = y_air_ + 0.5 * y_stm_
        y = (3.0 ** 0.5) / 2.0 * y_stm_
        return x, y

    # ── Build flammability boundary curves ────────────────────────────────────
    # The flammable region is a closed polygon in ternary space with five
    # distinct boundary segments.  Key geometric constraint: mole fractions
    # must always sum to 1 and all be ≥ 0.  The UFL line (H₂=75%) hits the
    # Air=0 binary axis at Steam=25% — it cannot extend to the 55% steam
    # inerting limit because H₂+Steam would exceed 100%.
    #
    # Correct polygon vertices (mole fractions):
    #   A  LFL base          H₂=4%,  Steam=0%,   Air=96%
    #   B  LFL @ inerting    H₂=4%,  Steam=55%,  Air=41%
    #   C  Inerting, Air=0   H₂=45%, Steam=55%,  Air=0%
    #   D  UFL on Air=0 axis H₂=75%, Steam=25%,  Air=0%
    #   E  UFL base          H₂=75%, Steam=0%,   Air=25%
    #   (close back to A along Steam=0 bottom edge)

    n_pts = 200

    # Segment A→B: LFL vertical (constant H₂=4%), steam 0→55%
    stm_AB  = np.linspace(0.0, STEAM_INERT_MOLE, n_pts)
    h2_AB   = np.full_like(stm_AB, H2_LFL_MOLE)
    air_AB  = 1.0 - h2_AB - stm_AB

    # Segment B→C: steam inerting horizontal (constant steam=55%),
    # H₂ from 4% up to 45% (= 1 - 55%), Air from 41% down to 0%
    h2_BC   = np.linspace(H2_LFL_MOLE, 1.0 - STEAM_INERT_MOLE, n_pts)
    stm_BC  = np.full_like(h2_BC, STEAM_INERT_MOLE)
    air_BC  = 1.0 - h2_BC - stm_BC

    # Segment C→D: Air=0 binary axis (H₂+Steam=1),
    # H₂ from 45% down to 75% (steam falls from 55% to 25%)
    h2_CD   = np.linspace(1.0 - STEAM_INERT_MOLE, H2_UFL_MOLE, n_pts)
    stm_CD  = 1.0 - h2_CD          # Air=0
    air_CD  = np.zeros_like(h2_CD)

    # Segment D→E: UFL vertical (constant H₂=75%), steam 25%→0%
    stm_DE  = np.linspace(1.0 - H2_UFL_MOLE, 0.0, n_pts)
    h2_DE   = np.full_like(stm_DE, H2_UFL_MOLE)
    air_DE  = 1.0 - h2_DE - stm_DE

    # Segment E→A: bottom edge (constant steam=0), H₂ from 75%→4%
    h2_EA   = np.linspace(H2_UFL_MOLE, H2_LFL_MOLE, n_pts)
    stm_EA  = np.zeros_like(h2_EA)
    air_EA  = 1.0 - h2_EA

    # Full polygon
    poly_h2  = np.concatenate([h2_AB,  h2_BC,  h2_CD,  h2_DE,  h2_EA])
    poly_stm = np.concatenate([stm_AB, stm_BC, stm_CD, stm_DE, stm_EA])
    poly_air = np.concatenate([air_AB, air_BC, air_CD, air_DE, air_EA])

    # Detonation peninsula (approximate): H₂ 18–59 vol%, low steam, Air>0
    # Upper steam boundary of detonation is ~30% (well below inerting limit)
    det_stm_max = 0.30
    det_h2_lo   = H2_DET_LO_MOLE
    det_h2_hi   = min(H2_DET_HI_MOLE, 1.0 - det_stm_max)
    det_h2  = [det_h2_lo, det_h2_hi, det_h2_hi, det_h2_lo, det_h2_lo]
    det_stm = [0.0,        0.0,       det_stm_max, det_stm_max, 0.0]
    det_air = [max(0.0, 1.0 - h - s) for h, s in zip(det_h2, det_stm)]

    # ── Draw ──────────────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9, 8))
    ax.set_aspect('equal')
    ax.axis('off')

    sqrt3_2 = 3.0 ** 0.5 / 2.0

    # Triangle outline
    tri_x = [0.0, 1.0, 0.5, 0.0]
    tri_y = [0.0, 0.0, sqrt3_2, 0.0]
    ax.plot(tri_x, tri_y, 'k-', lw=1.5)

    # Gridlines (every 10%)
    for frac in np.arange(0.1, 1.0, 0.1):
        # Constant H2 lines (vertical-ish)
        x0, y0 = _tern2cart(frac, 1.0 - frac, 0.0)
        x1, y1 = _tern2cart(frac, 0.0, 1.0 - frac)
        ax.plot([x0, x1], [y0, y1], color='gray', lw=0.4, ls='--', alpha=0.5)
        # Constant steam lines (horizontal-ish)
        x0, y0 = _tern2cart(0.0, 1.0 - frac, frac)
        x1, y1 = _tern2cart(1.0 - frac, 0.0, frac)
        ax.plot([x0, x1], [y0, y1], color='gray', lw=0.4, ls='--', alpha=0.5)
        # Constant air lines (diagonal-ish)
        x0, y0 = _tern2cart(frac, frac, 1.0 - 2.0*frac) if 1.0 - 2.0*frac >= 0 else (None, None)
        # simpler: connect steam=0 to H2=0 at fixed air
        _stm_r = np.linspace(0.0, 1.0 - frac, 20)
        _h2_r  = np.clip(1.0 - frac - _stm_r, 0, 1)
        _xr, _yr = _tern2cart(_h2_r, np.full_like(_h2_r, frac), _stm_r)
        ax.plot(_xr, _yr, color='gray', lw=0.4, ls='--', alpha=0.5)

    # Axis tick labels
    for frac in np.arange(0.0, 1.01, 0.2):
        # H2 axis (bottom)
        xp, yp = _tern2cart(frac, 1.0 - frac, 0.0)
        ax.text(xp, yp - 0.04, f'{frac:.0%}', ha='center', va='top', fontsize=7)
        # Steam axis (left side)
        xp, yp = _tern2cart(0.0, 1.0 - frac, frac)
        ax.text(xp - 0.04, yp, f'{frac:.0%}', ha='right', va='center', fontsize=7,
                rotation=60)
        # Air axis (right side)
        xp, yp = _tern2cart(frac, 0.0, 1.0 - frac)
        ax.text(xp + 0.03, yp, f'{1.0-frac:.0%}', ha='left', va='center', fontsize=7,
                rotation=-60)

    # Axis labels — placed at the midpoint of each side, nudged perpendicularly
    # outward by a small fixed offset so they clear the tick labels.
    # Left side midpoint (H2=0, Steam=0.5, Air=0.5) in Cartesian: (0.25, sqrt3_2/2)
    # Right side midpoint (H2=0.5, Steam=0, Air=0.5) in Cartesian: (0.75, 0)  — no, wrong
    # Left edge:  H2 vertex (0,0) → Steam vertex (0.5, sqrt3_2)  midpoint = (0.25, sqrt3_2/2)
    # Right edge: Air vertex (1,0) → Steam vertex (0.5, sqrt3_2) midpoint = (0.75, sqrt3_2/2)
    # Bottom edge: H2 (0,0) → Air (1,0)  midpoint = (0.5, 0)
    _off = 0.03   # perpendicular nudge distance
    # Bottom (H2 axis): nudge downward
    ax.text(0.5, -_off, 'H₂  mole fraction →',
            ha='center', va='top', fontsize=11, fontweight='bold')
    # Left side (Steam axis): nudge left-perpendicular to the 60° edge
    ax.text(0.25 - _off * 0.87, sqrt3_2 / 2 + _off * 0.5,
            '← Steam  mole fraction',
            ha='right', va='center', fontsize=11, fontweight='bold', rotation=60)
    # Right side (Air axis): nudge right-perpendicular to the −60° edge
    ax.text(0.75 + _off * 0.87, sqrt3_2 / 2 + _off * 0.5,
            'Air  mole fraction →',
            ha='left', va='center', fontsize=11, fontweight='bold', rotation=-60)

    # Vertex labels
    ax.text(0.0, -0.07, 'H₂\n(100%)', ha='center', fontsize=10, color='navy')
    ax.text(1.0, -0.07, 'Air\n(100%)', ha='center', fontsize=10, color='saddlebrown')
    ax.text(0.5,  sqrt3_2 + 0.04, 'Steam\n(100%)', ha='center', fontsize=10, color='steelblue')

    # Flammable region fill
    poly_x, poly_y = _tern2cart(np.array(poly_h2), np.array(poly_air), np.array(poly_stm))
    ax.fill(poly_x, poly_y, color='salmon', alpha=0.35, label='Deflagration region')
    ax.plot(poly_x, poly_y, color='red', lw=1.2)

    # Detonation region fill
    det_x, det_y = _tern2cart(np.array(det_h2), np.array(det_air), np.array(det_stm))
    ax.fill(det_x, det_y, color='darkred', alpha=0.45, label='Detonation peninsula')
    ax.plot(det_x, det_y, color='darkred', lw=1.2, ls='--')

    # LFL boundary line
    lfl_x, lfl_y = _tern2cart(h2_AB, air_AB, stm_AB)
    ax.plot(lfl_x, lfl_y, 'r-', lw=1.5)
    ax.text(lfl_x[0], lfl_y[0] - 0.02, 'LFL 4%', color='red', fontsize=8, ha='center')

    # UFL boundary line (vertical segment D→E only — the Air=0 segment is the triangle edge)
    ufl_x, ufl_y = _tern2cart(h2_DE, air_DE, stm_DE)
    ax.plot(ufl_x, ufl_y, 'r-', lw=1.5)
    ax.text(ufl_x[-1], ufl_y[-1] - 0.02, 'UFL 75%', color='red', fontsize=8, ha='center')

    # Steam inerting line (B→C)
    in_x, in_y = _tern2cart(h2_BC, air_BC, stm_BC)
    ax.plot(in_x, in_y, 'b-', lw=1.5)
    mid = len(in_x) // 2
    ax.text(in_x[mid], in_y[mid] + 0.03,
            'Steam inerting 55%', color='blue', fontsize=8, ha='center')

    # Air=0 boundary segment (C→D) — label the H2+Steam binary edge
    cd_x, cd_y = _tern2cart(h2_CD, air_CD, stm_CD)
    ax.plot(cd_x, cd_y, 'r-', lw=1.5)

    # Transient trajectory coloured by time.
    # Downsample to at most 2000 points — the ternary diagram doesn't
    # benefit from sub-second resolution and scatter with 40k+ individually
    # coloured markers is slow to render and save.
    traj_x, traj_y = _tern2cart(y_h2, y_air, y_steam)
    t_hr = results['t'] / 3600.0
    _stride = max(1, len(t_hr) // 2000)
    sc = ax.scatter(traj_x[::_stride], traj_y[::_stride],
                    c=t_hr[::_stride], cmap='plasma',
                    s=6, zorder=5, label='Transient trajectory')
    cbar = plt.colorbar(sc, ax=ax, fraction=0.03, pad=0.02)
    cbar.set_label('Time [hr]', fontsize=9)

    # Start and end markers
    ax.plot(traj_x[0],  traj_y[0],  'go', ms=10, zorder=6, label='Start')
    ax.plot(traj_x[-1], traj_y[-1], 'g*', ms=14, zorder=6, label='End')

    ax.legend(loc='upper right', fontsize=8, framealpha=0.8)
    ax.set_title(title, fontsize=12, fontweight='bold', pad=15)

    plt.tight_layout()
    plt.savefig(outfile, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Shapiro plot saved → {outfile}")


def print_summary(results: dict) -> None:
    """Print a concise scalar summary of key figures of merit."""
    t   = results['t']
    i_end = -1
    print("\n" + "="*65)
    print("  CONTAINMENT RESPONSE SUMMARY")
    print("="*65)
    print(f"  Peak pressure        : {results['P_kPa'].max():.1f} kPa"
          f"  @ t={t[results['P_kPa'].argmax()]/3600:.2f} hr")
    print(f"  Peak gas temperature : {results['T_C'].max():.1f} °C"
          f"  @ t={t[results['T_C'].argmax()]/3600:.2f} hr")
    print(f"  End pressure         : {results['P_kPa'][i_end]:.1f} kPa")
    print(f"  End temperature      : {results['T_C'][i_end]:.1f} °C")
    print(f"  Liquid volume (end)  : {results['V_liq_m3'][i_end]:.2f} m³")
    print(f"  NC quality (end)     : {results['x_nc'][i_end]:.3f}")
    print(f"  Total condensed mass : {results['M_cond_kg'][i_end]:.1f} kg")
    peak_q = results['q_cond'].max()
    print(f"  Peak cond. heat flux : {peak_q/1e3:.2f} kW/m²"
          f"  @ t={t[results['q_cond'].argmax()]/3600:.2f} hr")
    print("="*65 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# Built-in test case — large dry PWR containment LOCA blowdown
# ─────────────────────────────────────────────────────────────────────────────

def _build_loca_source(t_end: float = 7200.0) -> SourceTable:
    """
    Physically representative large-break LOCA source table for a 3411 MWt PWR.

    Three distinct phases:

    1. Blowdown  (0 – ~30 s)
       Critical (choked) flow from the double-ended guillotine break.  The RCS
       (~250,000 kg inventory at ~15 MPa) empties through the break at peak
       rates of 600–800 kg/s as choked critical flow.  The incoming mixture is
       at near-RCS saturation enthalpy (~2596 kJ/kg).

    2. Refill / reflood  (~30 – 120 s)
       RCS pressure falls below containment back-pressure; critical flow ends.
       ECCS water floods the vessel.  Mass flow to containment drops sharply as
       the break transitions from steam/mixture to cold ECCS spill.  Enthalpy
       falls to near-saturated liquid at containment pressure (~500 kJ/kg).
       By ~120 s the RCS is essentially empty and the break is flowing only
       ECCS water that immediately pools in the sump — negligible steam release.

    3. Long-term recirculation  (>120 s)
       The only steam source is decay-heat boiloff of sump water.  For a
       3411 MWt plant the ANS-1979 decay heat is ~100 MW at 30 s, declining
       as t^(-0.2).  Of this, only a fraction (~10–15%) actually flashes to
       steam in the containment gas space; the rest heats the sump liquid or
       is removed by the ECCS heat exchangers.  This gives a realistic long-
       term steam source of order 1–3 kg/s — well within the condensation
       sink capacity of the liner (~50–100 kg/s at design pressure).

    Enthalpy of source:
       Blowdown: saturated steam at RCS pressure (~2596 kJ/kg, 15 MPa).
       Refill:   declining from steam toward saturated liquid at containment
                 pressure as the break switches from flashing mixture to
                 ECCS spill (~500–800 kJ/kg).
       Long-term: saturated steam at containment pressure (~2675 kJ/kg at
                  ~400 kPa) — the small decay-heat fraction that boils off.

    Replace this table with actual RELAP5 / MAAP / MELCOR output for
    licensing or design-basis calculations.
    """
    # Time points [s]
    t_pts = np.array([
        0,    3,   10,   20,   30,   45,   60,   90,
        120,  180,  300,  600,  1800, 3600, 7200
    ], dtype=float)

    # Steam/water mass flow into containment [kg/s]
    # ── Blowdown: rises to ~800 kg/s peak at ~10 s (choked critical flow),
    #    then declines as RCS inventory depletes.
    # ── Refill: sharp drop after ~30 s as critical flow ends and ECCS fills
    #    the vessel; only cold ECCS spill reaches containment gas space as steam.
    # ── Long-term: ~1–3 kg/s decay-heat boiloff, decaying slowly with time.
    mdot_pts = np.array([
        0,   500,  800,  650,  350,  120,   40,   10,
        3.5,  2.8,  2.2,  1.7,   1.2,  0.9,  0.7
    ], dtype=float)

    # Enthalpy of incoming mass [J/kg]
    # Blowdown: near-RCS saturation enthalpy at 15 MPa (hg ≈ 2596 kJ/kg)
    # Refill transition: drops as break shifts from steam/mixture to ECCS spill
    # Long-term: saturated steam at containment pressure (~400 kPa → ~2738 kJ/kg)
    h_src_pts = np.array([
        2.596e6, 2.596e6, 2.60e6, 2.58e6, 2.50e6, 2.20e6, 1.50e6, 0.80e6,
        2.738e6, 2.738e6, 2.738e6, 2.738e6, 2.738e6, 2.738e6, 2.738e6
    ], dtype=float)

    # Direct volumetric energy source [W]
    # Small contribution from activated-water gamma heating of containment gas.
    # Conservatively set to zero; real value is <1% of decay heat.
    Qdot_pts = np.zeros_like(t_pts)

    # Hydrogen source [kg/s] from zircaloy-steam oxidation:
    #   Zr + 2H₂O → ZrO₂ + 2H₂   (ΔH ≈ −586 kJ/mol Zr)
    # For a 100% oxidised core (severe accident) ~800 kg H₂ total.
    # Oxidation rate peaks during blowdown/reflood, then decays sharply.
    # This is a representative severe-accident profile — for a design-basis
    # LOCA (limited cladding damage) H₂ generation is much smaller.
    # Radiolysis adds a small long-term background (~0.01 kg/s).
    mdot_h2_pts = np.array([
        0.0,  0.5,  2.0,  3.0,  2.5,  1.5,  0.5,  0.2,
        0.05, 0.02, 0.02, 0.01, 0.01, 0.01, 0.01
    ], dtype=float)

    return SourceTable(
        time_s       = t_pts,
        mdot_kg_s    = mdot_pts,
        Qdot_W       = Qdot_pts,
        h_src_J_kg   = h_src_pts,
        mdot_h2_kg_s = mdot_h2_pts,
    )



def extend_source_with_decay_steam(
        base_source: SourceTable,
        t_extend_s: float,
        rated_power_MWt: float = 575.0,
        decay_heat_fraction: float = 1.0,
        late_steam_enthalpy_Jkg: float = 2.738e6,
        transition_s: float = 300.0,
) -> SourceTable:
    """
    Extend a tabulated early LOCA source with a decay-heat-driven steam tail.

    Intended use: the source table or FLARE export supplies the early blowdown /
    refill response up to its final time.  Beyond that time, the containment
    mass-and-energy source is estimated as

        mdot_steam = decay_heat_fraction * Q_decay(t) / h_fg

    using the same simplified ANS-style decay heat curve used elsewhere in this
    prototype.  The final tabulated source value is not held constant; the
    continuation is explicit and user controlled.
    """
    t0 = float(np.max(base_source.time_s))
    if t_extend_s <= t0 + 1e-9:
        return base_source

    f = float(np.clip(decay_heat_fraction, 0.0, 1.0e3))
    P0_W = float(rated_power_MWt) * 1e6
    h_fg_est = 2.2e6

    def _decay_heat_W(t_s: float) -> float:
        return P0_W * 0.066 * max(float(t_s), 1.0) ** (-0.2)

    # Adaptive sparse extension grid; SourceTable interpolation handles run steps.
    t_pts = []
    t = t0 + min(10.0, max(1.0, transition_s / 30.0))
    while t <= t_extend_s + 1e-9:
        t_pts.append(t)
        if t < 3600.0:
            t += 60.0
        elif t < 14400.0:
            t += 300.0
        else:
            t += 1800.0
    if not t_pts or t_pts[-1] < t_extend_s:
        t_pts.append(t_extend_s)
    t_ext = np.array(t_pts, dtype=float)

    mdot_decay = np.array([f * _decay_heat_W(ti) / h_fg_est for ti in t_ext], dtype=float)

    # Optional short linear handoff prevents a discontinuity if the final FLARE/table
    # value differs substantially from the decay-heat estimate.
    mdot0 = float(base_source.mdot_kg_s[-1])
    if transition_s > 0.0:
        blend = np.clip((t_ext - t0) / transition_s, 0.0, 1.0)
        mdot_ext = (1.0 - blend) * mdot0 + blend * mdot_decay
    else:
        mdot_ext = mdot_decay

    h_ext = np.full_like(t_ext, float(late_steam_enthalpy_Jkg), dtype=float)
    q_ext = np.zeros_like(t_ext)
    h2_ext = np.zeros_like(t_ext)

    return SourceTable(
        time_s       = np.concatenate([base_source.time_s,       t_ext]),
        mdot_kg_s    = np.concatenate([base_source.mdot_kg_s,    np.clip(mdot_ext, 0.0, None)]),
        Qdot_W       = np.concatenate([base_source.Qdot_W,       q_ext]),
        h_src_J_kg   = np.concatenate([base_source.h_src_J_kg,   h_ext]),
        mdot_h2_kg_s = np.concatenate([base_source.mdot_h2_kg_s, h2_ext]),
    )


def build_source_from_flare(
        flare_csv: str,
        t_extend_s: float = 86400.0,
        decay_heat_fraction: float = 1.0,
        source_multiplier: float = 1.0,
        rated_power_MWt: float = None,
        late_steam_enthalpy_Jkg: float = 2.738e6,
) -> SourceTable:
    """
    Build a SourceTable directly from a FLARE RCS/core simulation output CSV.

    This function extracts only the **mass and energy source term** — it does
    not infer or return containment geometry or initial conditions.  Those must
    be supplied independently via a containment-specific Excel input file
    (<case>_in.xlsx) or programmatically via ContainmentDesign / ContainmentIC.

    Rationale
    ---------
    FLARE input files contain a dose-screening parameter
    ``nbt_containment_volume_ft3`` which is a placeholder sized for a large PWR
    (typically ~2.74×10⁶ ft³).  This value is NOT the actual containment design
    volume for the plant being simulated and must never be used as such.
    Containment geometry is an independent design input that belongs in the
    containment simulator's own input file.

    Source term construction
    ------------------------
    Three contributions are read from the FLARE output CSV:

    1. Break flow and enthalpy  ('Break Flow (kg/s)', 'Break Enthalpy (kJ/kg)')
       h > 800 kJ/kg → steam/mixture enters the gas space directly.
       h < 800 kJ/kg → liquid discharge to sump; gas source replaced by
       decay-heat-driven boiloff estimate (decay_heat_fraction × P_decay).

    2. Hydrogen generation rate  ('H2 Generated (kg)' differentiated to kg/s).

    3. Long-term extension beyond FLARE end time using ANS-1979 decay heat
       with the supplied decay_heat_fraction.

    Parameters
    ----------
    flare_csv          : path to FLARE output CSV  (<case>_out.csv)
    t_extend_s         : total time to extend the source table to  [s]
    decay_heat_fraction: fraction of decay heat that becomes gas-space steam
                         source in long-term phase (default 1.0 = 100%,
                         i.e. all decay heat drives steam boiloff)
    source_multiplier  : uniform scale factor applied to all mass and energy
                         flows (default 1.0).  Use to represent a break size
                         different from the one simulated in FLARE — e.g. 0.56
                         scales the source term to 56% of the FLARE break flow
                         while preserving enthalpy (break fluid quality unchanged).

    Returns
    -------
    source : SourceTable  (time_s, mdot_kg_s, Qdot_W, h_src_J_kg, mdot_h2_kg_s)
    """
    import pandas as pd

    df = pd.read_csv(flare_csv)

    # ── Extract key columns ────────────────────────────────────────────────────
    t_arr   = df['Time (s)'].values
    mdot_bk = df['Break Flow (kg/s)'].values
    h_bk    = df['Break Enthalpy (kJ/kg)'].values * 1e3    # → J/kg
    h2_kg   = df['H2 Generated (kg)'].values

    # Differentiate H2 cumulative → rate [kg/s]
    dh2     = np.gradient(h2_kg, t_arr)
    mdot_h2 = np.clip(dh2, 0.0, None)

    # ── Gas-space steam source ─────────────────────────────────────────────────
    H_STEAM_THRESH = 800e3   # J/kg — above this: steam/mixture to gas space
    t_flare_end    = float(t_arr[-1])

    mdot_gas = np.where(h_bk >= H_STEAM_THRESH, mdot_bk, 0.0)
    h_src    = np.where(h_bk >= H_STEAM_THRESH, h_bk,    2.738e6)

    # Decay heat for low-enthalpy and long-term phases
    P0_MW = 575.0 if rated_power_MWt is None else float(rated_power_MWt)
    if rated_power_MWt is None:
        try:
            P0_MW = float(df['Core Power (MW)'].iloc[0])
        except Exception:
            pass

    def _decay_heat_W(t_s):
        return P0_MW * 1e6 * 0.066 * max(t_s, 1.0) ** (-0.2)

    h_fg_est = 2.2e6
    for i in range(len(t_arr)):
        if h_bk[i] < H_STEAM_THRESH and mdot_bk[i] > 0:
            mdot_gas[i] = _decay_heat_W(t_arr[i]) * decay_heat_fraction / h_fg_est
            h_src[i]    = late_steam_enthalpy_Jkg

    # ── Long-term extension ────────────────────────────────────────────────────
    if t_extend_s > t_flare_end:
        t_ext_pts = []
        t = t_flare_end + 10.0
        while t <= t_extend_s:
            t_ext_pts.append(t)
            if   t < 3600:  t += 60.0
            elif t < 14400: t += 300.0
            else:           t += 1800.0
        t_ext = np.array(t_ext_pts)
        mdot_ext    = np.array([_decay_heat_W(ti) * decay_heat_fraction / h_fg_est
                                 for ti in t_ext])
        h_ext       = np.full_like(t_ext, late_steam_enthalpy_Jkg)
        mdot_h2_ext = np.zeros_like(t_ext)

        t_arr    = np.concatenate([t_arr,    t_ext])
        mdot_gas = np.concatenate([mdot_gas, mdot_ext])
        h_src    = np.concatenate([h_src,    h_ext])
        mdot_h2  = np.concatenate([mdot_h2,  mdot_h2_ext])


    # ── Apply source multiplier ───────────────────────────────────────────────
    # Scale mass flows only; enthalpy is a fluid property independent of flow rate.
    if source_multiplier != 1.0:
        mdot_gas = mdot_gas * source_multiplier
        mdot_h2  = mdot_h2  * source_multiplier

    # ── Build source table by integrating FLARE output ────────────────────────
    # Rather than interpolating instantaneous flow rates (which misrepresents
    # the rapidly-falling blowdown peak when the containment timestep is coarser
    # than the FLARE timestep), integrate the FLARE arrays to get cumulative
    # mass and energy, then difference at each containment table point to get
    # the correct interval-average flux.  This is exact regardless of the
    # relative timestep sizes.

    dt_flare  = np.diff(t_arr)
    m_avg     = 0.5 * (mdot_gas[:-1] + mdot_gas[1:])
    E_avg     = 0.5 * (mdot_gas[:-1] * h_src[:-1] + mdot_gas[1:] * h_src[1:])
    h2_avg    = 0.5 * (mdot_h2[:-1]  + mdot_h2[1:])

    cum_mass   = np.concatenate([[0.0], np.cumsum(m_avg  * dt_flare)])
    cum_energy = np.concatenate([[0.0], np.cumsum(E_avg  * dt_flare)])
    cum_h2     = np.concatenate([[0.0], np.cumsum(h2_avg * dt_flare)])

    # Containment table breakpoints
    t_tbl = np.unique(np.concatenate([
        np.arange(0,    min(30,   t_arr[-1]) + 1, 2.0),
        np.arange(30,   min(120,  t_arr[-1]),      5.0),
        np.arange(120,  min(600,  t_arr[-1]),      10.0),
        np.arange(600,  min(3600, t_arr[-1]),      60.0),
        np.arange(3600, t_arr[-1] + 1,             300.0),
    ]))
    t_tbl = t_tbl[t_tbl <= t_arr[-1]]

    # Interpolate cumulative integrals at breakpoints, then difference
    cum_m_tbl  = np.interp(t_tbl, t_arr, cum_mass)
    cum_E_tbl  = np.interp(t_tbl, t_arr, cum_energy)
    cum_h2_tbl = np.interp(t_tbl, t_arr, cum_h2)

    dt_int      = np.diff(t_tbl)
    mdot_int    = np.diff(cum_m_tbl)  / dt_int
    E_int       = np.diff(cum_E_tbl)  / dt_int
    h2_int      = np.diff(cum_h2_tbl) / dt_int
    h_src_int   = np.where(mdot_int > 1e-6, E_int / mdot_int, 2.738e6)

    # Represent each interval at its midpoint; bookend for clean interpolation
    t_mid = 0.5 * (t_tbl[:-1] + t_tbl[1:])
    t_out    = np.concatenate([[0.0],         t_mid,    [t_arr[-1]]])
    mdot_out = np.concatenate([[mdot_int[0]], mdot_int, [mdot_int[-1]]])
    hsrc_out = np.concatenate([[h_src_int[0]],h_src_int,[h_src_int[-1]]])
    h2_out   = np.concatenate([[h2_int[0]],   h2_int,   [h2_int[-1]]])

    source = SourceTable(
        time_s       = t_out,
        mdot_kg_s    = np.clip(mdot_out, 0.0, None),
        Qdot_W       = np.zeros_like(t_out),
        h_src_J_kg   = hsrc_out,
        mdot_h2_kg_s = np.clip(h2_out,  0.0, None),
    )

    print(f"\n  FLARE source term loaded : {flare_csv}")
    print(f"  FLARE time range         : 0 – {t_flare_end:.0f} s")
    print(f"  Extended to              : {t_arr[-1]:.0f} s  ({t_arr[-1]/3600:.1f} hr)")
    print(f"  Table points             : {len(t_out)}")
    print(f"  Source multiplier        : {source_multiplier:.3f}"
          + ("  (unscaled)" if source_multiplier == 1.0 else f"  ({source_multiplier*100:.0f}% of FLARE flows)"))
    print(f"  Peak break flow (inst.)  : {mdot_bk.max():.0f} kg/s  @ t=0 s  (FLARE)")
    print(f"  Avg flux, first 2s       : {mdot_int[0]:.0f} kg/s  (interval average, scaled)")
    print(f"  Total H2 generated       : {h2_kg.max() * source_multiplier:.3f} kg  (scaled)")
    print(f"  Rated core power         : {P0_MW:.0f} MWt")
    print(f"\n  NOTE: Containment geometry must be supplied via a separate")
    print(f"        <case>_in.xlsx input file or ContainmentDesign object.")
    print(f"        Do NOT use FLARE's nbt_containment_volume_ft3 — it is a")
    print(f"        dose-screening placeholder, not the actual design volume.")

    return source




def run_loca_example(t_end: float = 7200.0, dt: float = 2.0,
                     outfile_prefix: str = "containment") -> dict:
    """
    Run a representative large-break LOCA dry containment example.

    Parameters
    ----------
    t_end           : simulation end time  [s]  (default 2 hr)
    dt              : integration timestep [s]  (default 2 s)
    outfile_prefix  : prefix for output PNG

    Returns
    -------
    results dict from ContainmentSimulator.run()
    """

    # ── Design parameters — large dry PWR (Westinghouse-class) ───────────────
    design = ContainmentDesign(
        air_volume_m3      = 50_667.0,
        floor_area_m2      = 600.0,
        wall_surf_area_m2  = 8_000.0,
        wall_thickness_m   = 0.044,
        wall_char_height_m = 12.0,
        wall_k_Wm_K        = K_STEEL,
        wall_rho_kg_m3     = 7800.0,
        wall_cp_J_kg_K     = 500.0,
        P_design_kPa       = 517.0,
        T_design_K         = 450.0,
        uchida_multiplier  = 1.0,
    )

    # ── Initial conditions ────────────────────────────────────────────────────
    ic = ContainmentIC(
        P0_Pa     = PATM,      # 101.325 kPa (pre-accident)
        T0_K      = 300.15,    # 27 °C (pre-accident ambient)
        T_wall0_K = 300.15,
        RH0       = 0.20,      # 20 % relative humidity (conservative)
    )

    # ── Source table ──────────────────────────────────────────────────────────
    source = _build_loca_source(t_end)

    # ── Run ───────────────────────────────────────────────────────────────────
    print("Building containment simulator …")
    sim = ContainmentSimulator(design, ic, source, dt_s=dt, verbose=True)

    print(f"Running LOCA transient to t = {t_end:.0f} s ({t_end/3600:.2f} hr) …")
    results = sim.run(t_end=t_end)

    print_summary(results)
    plot_results(results,
                 title="PWR Large-Break LOCA — Dry Containment Response",
                 outfile=f"{outfile_prefix}_loca.png",
                 floor_area_m2=design.floor_area_m2)
    plot_shapiro(results,
                 title="PWR Dry Containment — Steam / Air / H₂ (Shapiro Diagram)",
                 outfile=f"{outfile_prefix}_shapiro.png")
    save_csv(results, f"{outfile_prefix}_out.csv",
             floor_area_m2=design.floor_area_m2)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# CSV / tabular output helper
# ─────────────────────────────────────────────────────────────────────────────

def save_csv(results: dict, filename: str = "containment_out.csv",
             floor_area_m2: float = 600.0) -> None:
    """Write time-series results to a CSV file."""
    import pandas as pd

    liq_level = results.get('floor_flood_level_m', results['V_liq_m3'] / max(floor_area_m2, 1.0))

    keys_ordered = [
        ('t',             'Time [s]'),
        ('P_kPa',         'Pressure [kPa]'),
        ('T_C',           'Gas Temp [C]'),
        ('T_wall_C',      'Wall Temp Inner [C]'),
        ('T_conc_C',      'Concrete Surface Temp [C]'),
        ('T_pool_C',      'Wet Floor/Sump Temp [C]'),
        ('T_bsmt_C',      'Basemat Surface Temp [C]'),
        ('V_liq_m3',      'Liquid Volume [m3]'),
        ('_level',        'Sump Level [m]'),
        ('x_nc',          'NC Quality [-]'),
        ('M_air_kg',      'Air Mass [kg]'),
        ('M_steam_kg',    'Steam in Gas [kg]'),
        ('M_h2_kg',       'H2 Mass [kg]'),
        ('M_cond_kg',     'Condensed Mass [kg]'),
        ('mdot_cond',     'Cond Rate [kg/s]'),
        ('q_cond',        'Wall HF [W/m2]'),
        ('Q_wall_W',      'Wall Heat Removal [W]'),
        ('Q_floor_dry_W', 'Dry Floor Heat Removal [W]'),
        ('Q_pool_W',      'Wet Floor Heat Removal [W]'),
        ('A_floor_dry_m2','Dry Floor Area [m2]'),
        ('A_floor_wet_m2','Wet Floor Area [m2]'),
        ('Q_spray_W',     'Spray Heat Removal [W]'),
        ('mdot_spray',    'Spray Flow [kg/s]'),
        ('mdot_spray_cond','Spray Cond Rate [kg/s]'),
        ('mdot_in',       'Source mdot [kg/s]'),
        ('Hdot_in_W',     'Source Enthalpy Flux [W]'),
    ]

    data = {}
    for k, label in keys_ordered:
        data[label] = liq_level if k == '_level' else results[k]

    pd.DataFrame(data).to_csv(filename, index=False, float_format='%.6g')
    print(f"  CSV saved → {filename}")


# ─────────────────────────────────────────────────────────────────────────────
# Excel input parsing  (mirrors FLARE pwr_sim pattern)
# ─────────────────────────────────────────────────────────────────────────────

def _sanitise_command(cmd: str) -> str:
    """
    Normalise common input-file formatting issues before exec().

    1. Unicode minus/dash characters → ASCII hyphen-minus.
    2. Trailing unit annotations stripped from RHS.
    3. Thousands-separator commas in plain numeric RHS (e.g. 50,000 → 50000).
    """
    import re as _re
    for _ch in ('\u2212', '\u2013', '\u2014'):
        cmd = cmd.replace(_ch, '-')
    if '=' in cmd:
        _lhs, _, _rhs_raw = cmd.partition('=')
        _comment_split = _rhs_raw.split('#', 1)
        _rhs = _comment_split[0].strip()
        _comment = (' #' + _comment_split[1]) if len(_comment_split) > 1 else ''
        _rhs_no_unit = _re.sub(r'\s+[A-Za-z][A-Za-z0-9/°·\s]*$', '', _rhs).strip()
        if _re.match(r'^[+-]?[\d,]+\.?\d*([eE][+-]?\d+)?\s*$', _rhs_no_unit):
            cmd = _lhs + '= ' + _rhs_no_unit.replace(',', '') + _comment
    return cmd.rstrip()


def _get(ns: dict, name: str, default, echo_log: list):
    """Retrieve name from namespace or use default; append to echo log."""
    value  = ns[name] if name in ns else default
    source = 'user' if name in ns else 'default'
    echo_log.append({'Variable': name, 'Value': value,
                     'Default': default, 'Source': source})
    return value


def _to_bool(value, default=False) -> bool:
    """Robust bool conversion for Excel command values."""
    if value is None:
        return bool(default)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = str(value).strip().lower()
    if s in ('1', 'true', 'yes', 'y', 'on'):
        return True
    if s in ('0', 'false', 'no', 'n', 'off'):
        return False
    return bool(default)


def _read_flare_nbt_inputs(xlsx_file: str) -> dict:
    """Read selected NOTBADTRAD/NBT-style inputs from a FLARE input workbook.

    This is intentionally permissive.  FLARE input workbooks store commands in
    column A as text assignments.  The containment code only imports selected
    spray-related flags/parameters and does not use FLARE's NBT containment
    volume as a geometry source.
    """
    data = {}
    path = Path(xlsx_file)
    if not path.exists():
        return data
    if not _OPENPYXL:
        return data
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        ns = {}
        for row in ws.iter_rows(values_only=True):
            val = row[0] if row else None
            if isinstance(val, str) and '=' in val:
                cmd = _sanitise_command(val.split('#')[0].strip())
                lhs = cmd.split('=', 1)[0].strip()
                if lhs.startswith('nbt_') or 'spray' in lhs.lower():
                    try:
                        exec(cmd, {}, ns)
                    except Exception:
                        pass
        wb.close()
        data.update(ns)
    except Exception:
        pass
    return data


def read_excel_input(case: str) -> tuple[ContainmentDesign,
                                         ContainmentIC,
                                         SourceTable,
                                         float, float, list]:
    """
    Read a FLARE-style Excel input workbook  <case>_in.xlsx
    and return  (design, ic, source, t_end, dt, echo_log).

    Workbook layout
    ---------------
    Sheet name : <case>_in

    COMMAND BLOCK  (rows above the time-series header)
    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    Each row in column A may contain a  key = value  assignment.
    Comments after '#' are stripped.  All keys are optional; defaults
    match the ContainmentDesign / ContainmentIC defaults documented below.

    Design geometry
        air_volume_m3       = 50667      # free gas volume [m³]
        floor_area_m2       = 600        # projected containment floor area [m²]
        sump_area_m2        = 25         # engineered sump plan/free-surface area [m²]
        sump_depth_m        = 2.0        # sump depth before floor overflow [m]
        wall_surf_area_m2   = 8000       # condensing wall area [m²]
        wall_thickness_m    = 0.044      # steel liner thickness [m]
        wall_char_height_m  = 12.0       # Dehbi correlation height [m]
        wall_k              = 50.0       # wall thermal conductivity [W/(m·K)]
        wall_rho            = 7800       # wall density [kg/m³]
        wall_cp             = 500        # wall specific heat [J/(kg·K)]
        P_design_kPa        = 517        # design pressure [kPa]
        T_design_C          = 177        # design temperature [°C]

    Initial conditions
        P0_kPa              = 101.325    # initial total pressure [kPa]
        T0_C                = 27.0       # initial gas temperature [°C]
        T_wall0_C           = 27.0       # initial wall temperature [°C]
        RH0                 = 0.20       # initial relative humidity [-]

    Run control
        endtime             = 7200       # simulation end time [s]
        timestep            = 2          # integration timestep [s]

    TIME-SERIES TABLE  (rows at and below the header row)
    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    Header row: column A must start with "time" (case-insensitive).
    Data columns (A–F):

        A  time        [s]
        B  mdot_steam  [kg/s]   — steam/water mass flow into containment
        C  h_src       [kJ/kg]  — specific enthalpy of source mass
        D  mdot_h2     [kg/s]   — hydrogen mass source (0 if absent)
        E  Qdot        [W]      — direct volumetric energy source (normally 0)
        F  (reserved / unused)

    Blank or missing cells default to 0.  Up to 3 consecutive blank rows
    in column A are tolerated before the table is considered ended.
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl is required for Excel input. "
                          "Install with: pip install openpyxl")

    import csv as _csv, uuid as _uuid, shutil as _shutil, tempfile as _tmp

    fn_in = f"{case}_in.xlsx"
    sheet = f"{case}_in"

    if not Path(fn_in).exists():
        raise FileNotFoundError(
            f"Input file not found: {fn_in}\n"
            f"  Expected location : {Path.cwd() / fn_in}\n"
            f"  Run  containment_sim.py --template {case}  to create a template."
        )

    # ── snapshot (mirrors FLARE: copy first so original can stay open in Excel)
    _tmp_path = (Path(tempfile.gettempdir())
                 / f"CONT_{case}_{os.getpid()}_{_uuid.uuid4().hex[:8]}_in.xlsx")
    shutil.copy2(fn_in, _tmp_path)
    print(f"  Input snapshot : {_tmp_path}")

    # ── load workbook from private snapshot
    _wb = openpyxl.load_workbook(str(_tmp_path), read_only=True, data_only=True)
    if sheet not in _wb.sheetnames:
        available = ', '.join(_wb.sheetnames)
        _wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found in {fn_in}.\n"
            f"  Available sheets: {available}"
        )
    _ws = _wb[sheet]

    # ── export sheet to in-memory CSV rows (same as FLARE)
    _csv_rows = []
    for _row in _ws.iter_rows(values_only=True):
        _csv_rows.append(["" if v is None else v for v in _row])
    _wb.close()

    # ── locate time-series header (strict: col-A starts with "time", numeric follows)
    _col_a = [r[0] if r else None for r in _csv_rows]
    _time_hdr = None
    for _i, _v in enumerate(_col_a):
        if not isinstance(_v, str):
            continue
        _vs = _v.strip().lower().replace(' ', '')
        # Identify the main LOCA source table header row.
        _b = _csv_rows[_i][1] if len(_csv_rows[_i]) > 1 else ''
        _bs = str(_b).strip().lower().replace(' ', '')
        if (_vs.startswith('time[') or _vs == 'time') and ('mdot' in _bs) and '=' not in _vs:
            for _j in range(_i + 1, min(_i + 4, len(_col_a))):
                _nxt = _col_a[_j]
                if _nxt is None or _nxt == '':
                    continue
                try:
                    float(_nxt); _time_hdr = _i; break
                except (TypeError, ValueError):
                    break
        if _time_hdr is not None:
            break

    # ── command block = everything above the time header
    _cmd_end = _time_hdr if _time_hdr is not None else len(_col_a)
    _comdata = _col_a[:_cmd_end]

    if _time_hdr is None:
        print("  WARNING: No time-series table found — using built-in LOCA defaults.")

    # ── parse command block via exec (same as FLARE)
    _ns  = {}
    _log = []
    for _cmd in _comdata:
        if isinstance(_cmd, str) and '=' in _cmd:
            try:
                _c = _sanitise_command(_cmd.split('#')[0].strip())
                _parts = _c.split('=', 1)
                if len(_parts) == 2 and _parts[0].strip().isidentifier():
                    exec(_c, {}, _ns)
            except Exception as _e:
                print(f"  WARNING: could not parse command '{_cmd}': {_e}")

    # ── read time-series table (columns A–F, tolerates up to 3 blank rows)
    _tbl_rows = []
    if _time_hdr is not None:
        _blanks = 0
        for _raw in _csv_rows[_time_hdr + 1:]:
            _row = list(_raw[:6]) + [''] * max(0, 6 - len(_raw))
            _row = [None if v == '' else v for v in _row[:6]]
            _t = _row[0]
            if _t is None:
                _blanks += 1
                if _blanks > 3: break
                continue
            try:
                float(_t)
            except (TypeError, ValueError):
                _blanks += 1
                if _blanks > 3: break
                continue
            _blanks = 0
            _tbl_rows.append(_row)

    # ── convert table to numpy (6 columns, pad missing with 0)
    if _tbl_rows:
        _arr = np.array(
            [[float(c) if c is not None else 0.0 for c in r] +
             [0.0] * (6 - len(r))
             for r in _tbl_rows], dtype=float)
    else:
        # fallback: built-in LOCA source
        src = _build_loca_source()
        _arr = np.column_stack([
            src.time_s, src.mdot_kg_s,
            src.h_src_J_kg / 1e3,    # stored in kJ/kg in Excel
            src.mdot_h2_kg_s,
            src.Qdot_W,
            np.zeros(len(src.time_s)),
        ])

    time_s     = _arr[:, 0]
    mdot_s     = _arr[:, 1]
    h_src_kJkg = _arr[:, 2]
    mdot_h2    = _arr[:, 3]
    Qdot       = _arr[:, 4]
    # column F reserved

    # h_src in Excel is kJ/kg (human-friendly) → convert to J/kg
    h_src_Jkg = h_src_kJkg * 1e3

    # ── assemble objects from parsed namespace ─────────────────────────────────
    G = lambda name, default: _get(_ns, name, default, _log)

    # Optional link to FLARE input/NBT spray flag.  Direct containment spray
    # inputs still control the actual flow/temperature; the NBT flag is used
    # only as an enable/disable default when requested.
    spray_from_flare_nbt = _to_bool(G('spray_from_flare_nbt', False), False)
    flare_input_file = str(G('flare_input_file', '') or '').strip()
    _flare_nbt = {}
    if flare_input_file:
        _flare_in_path = Path(flare_input_file)
        if not _flare_in_path.exists():
            _candidate = Path(fn_in).resolve().parent / flare_input_file
            if _candidate.exists():
                _flare_in_path = _candidate
        _flare_nbt = _read_flare_nbt_inputs(str(_flare_in_path))

    spray_enabled_default = False
    if spray_from_flare_nbt and _flare_nbt:
        spray_enabled_default = _to_bool(_flare_nbt.get('nbt_sprays_on', False), False)
    spray_enabled = _to_bool(G('spray_enabled', spray_enabled_default), spray_enabled_default)

    # Flow can be supplied either as kg/s or as U.S. gpm.  If neither is
    # supplied, sprays may be enabled but with zero flow; this is intentional to
    # avoid silently crediting an unspecified system.
    if 'spray_flow_kg_s' in _ns:
        spray_flow_kg_s_default = float(_ns['spray_flow_kg_s'])
    elif 'spray_flow_gpm' in _ns:
        spray_flow_kg_s_default = float(_ns['spray_flow_gpm']) * 0.06309
    else:
        spray_flow_kg_s_default = 0.0

    design = ContainmentDesign(
        air_volume_m3      = G('air_volume_m3',     50_667.0),
        floor_area_m2      = G('floor_area_m2',       600.0),
        sump_area_m2       = G('sump_area_m2',         25.0),
        sump_depth_m       = G('sump_depth_m',          2.0),
        wall_surf_area_m2  = G('wall_surf_area_m2',  8_000.0),
        wall_thickness_m   = G('wall_thickness_m',     0.044),
        wall_char_height_m = G('wall_char_height_m',  12.0),
        wall_k_Wm_K        = G('wall_k',              K_STEEL),
        wall_rho_kg_m3     = G('wall_rho',           7800.0),
        wall_cp_J_kg_K     = G('wall_cp',             500.0),
        gap_thick_m        = G('gap_thick_m',          0.001),
        gap_k_Wm_K         = G('gap_k',               0.026),
        conc_k_Wm_K        = G('conc_k',              K_CONC),
        conc_rho_kg_m3     = G('conc_rho',           2300.0),
        conc_cp_J_kg_K     = G('conc_cp',             900.0),
        conc_depth_m       = G('conc_depth_m',         1.5),
        P_design_kPa       = G('P_design_kPa',        517.0),
        T_design_K         = G('T_design_C',          177.0) + T_REF,
        uchida_multiplier  = G('uchida_multiplier',  1.0),
        condensation_model = str(G('condensation_model', 'uchida') or 'uchida').strip().lower(),
        pcc_fraction       = G('pcc_fraction',         0.0),
        T_outer_C          = G('T_outer_C',            15.0),
        spray_enabled      = spray_enabled,
        spray_effectiveness= G('spray_effectiveness',  0.80),
        spray_sauter_mean_diameter_um = G('spray_sauter_mean_diameter_um', 500.0),
    )

    ic = ContainmentIC(
        P0_Pa     = G('P0_kPa',    101.325) * 1e3,
        T0_K      = G('T0_C',       27.0)   + T_REF,
        T_wall0_K = G('T_wall0_C',  27.0)   + T_REF,
        RH0       = G('RH0',         0.20),
    )

    source = SourceTable(
        time_s       = time_s,
        mdot_kg_s    = mdot_s,
        Qdot_W       = Qdot,
        h_src_J_kg   = h_src_Jkg,
        mdot_h2_kg_s = mdot_h2,
    )

    t_end = float(G('endtime',  time_s[-1]))
    dt    = float(G('timestep', 2.0))

    # Source-term mode.  Default is the literal table.  The intended coupled
    # workflow is ``flare_then_decay_steam``: use FLARE mass/energy data through
    # the end of the FLARE CSV, then derive late-time steam from decay heat.
    source_mode = str(G('source_mode', 'table') or 'table').strip().lower()
    source_multiplier = float(G('source_multiplier', 1.0))
    late_steam_fraction = float(G('late_steam_fraction', G('decay_heat_fraction', 1.0)))
    rated_power_MWt = float(G('rated_power_MWt', 575.0))
    late_steam_enthalpy_kJkg = float(G('late_steam_enthalpy_kJkg', 2738.0))
    late_steam_transition_s = float(G('late_steam_transition_s', 300.0))
    flare_source_file = str(G('flare_source_file', '') or '').strip()

    if source_multiplier != 1.0 and source_mode.startswith('table'):
        source.mdot_kg_s = source.mdot_kg_s * source_multiplier
        source.mdot_h2_kg_s = source.mdot_h2_kg_s * source_multiplier

    if source_mode in ('flare_then_decay_steam', 'flare_decay', 'flare_then_decay'):
        if not flare_source_file:
            raise ValueError("source_mode='flare_then_decay_steam' requires flare_source_file")
        _flare_path = Path(flare_source_file)
        if not _flare_path.exists():
            _candidate = Path(fn_in).resolve().parent / flare_source_file
            if _candidate.exists():
                _flare_path = _candidate
        if not _flare_path.exists():
            raise FileNotFoundError(f"FLARE source file not found: {flare_source_file}")
        source = build_source_from_flare(
            str(_flare_path),
            t_extend_s=t_end,
            decay_heat_fraction=late_steam_fraction,
            source_multiplier=source_multiplier,
            rated_power_MWt=rated_power_MWt,
            late_steam_enthalpy_Jkg=late_steam_enthalpy_kJkg * 1e3,
        )
    elif source_mode in ('table_then_decay_steam', 'table_decay', 'table_then_decay'):
        source = extend_source_with_decay_steam(
            source,
            t_extend_s=t_end,
            rated_power_MWt=rated_power_MWt,
            decay_heat_fraction=late_steam_fraction,
            late_steam_enthalpy_Jkg=late_steam_enthalpy_kJkg * 1e3,
            transition_s=late_steam_transition_s,
        )

    # ── Parse optional containment spray table (time_spray / flow / temp) ────
    _spray_hdr = None
    for _i, _v in enumerate(_col_a):
        if isinstance(_v, str) and _v.strip().lower().startswith('time_spray'):
            _spray_hdr = _i; break

    _spray_start = float(G('spray_start_s', 0.0))
    _spray_end   = float(G('spray_end_s', 0.0))
    _spray_temp  = float(G('spray_temp_C', 27.0))
    if _spray_end > _spray_start and spray_flow_kg_s_default > 0.0:
        _spray_t = [0.0, _spray_start, _spray_start + 1.0, _spray_end, _spray_end + 1.0, t_end]
        _spray_f = [0.0, 0.0, spray_flow_kg_s_default, spray_flow_kg_s_default, 0.0, 0.0]
        _spray_T = [_spray_temp] * len(_spray_t)
    else:
        _spray_t = [0.0, t_end]
        _spray_f = [0.0, 0.0]
        _spray_T = [_spray_temp, _spray_temp]

    if _spray_hdr is not None:
        _st, _sf, _sT = [], [], []
        for _raw in _csv_rows[_spray_hdr + 1:]:
            _row3 = list(_raw[:3]) + [''] * max(0, 3 - len(_raw))
            _tv, _fv, _Tval = _row3[0], _row3[1], _row3[2]
            if _tv is None or _tv == '':
                break
            try:
                _st.append(float(_tv))
                _sf.append(float(_fv) if _fv != '' else 0.0)
                _sT.append(float(_Tval) if _Tval != '' else _spray_temp)
            except (TypeError, ValueError):
                break
        if len(_st) >= 2:
            _spray_t, _spray_f, _spray_T = _st, _sf, _sT

    design.spray_flow_time_s = np.array(_spray_t, dtype=float)
    design.spray_flow_kg_s   = np.array(_spray_f, dtype=float)
    design.spray_temp_C      = np.array(_spray_T, dtype=float)

    if flare_input_file and _flare_nbt:
        print(f"  FLARE NBT spray flag : nbt_sprays_on={_flare_nbt.get('nbt_sprays_on', 'not found')}")

    # ── print echo log ─────────────────────────────────────────────────────────
    print(f"\n  {'Variable':<22}  {'Value':>12}  {'Source'}")
    print(f"  {'-'*22}  {'-'*12}  {'-'*7}")
    for e in _log:
        print(f"  {e['Variable']:<22}  {str(e['Value']):>12}  {e['Source']}")
    print()

    return design, ic, source, t_end, dt, _log


# ─────────────────────────────────────────────────────────────────────────────
# Excel template generator
# ─────────────────────────────────────────────────────────────────────────────

def write_excel_template(case: str = "containment") -> None:
    """
    Write a ready-to-edit Excel template  <case>_in.xlsx  pre-populated
    with all supported command-block variables and a representative LOCA
    time-series table.

    The template mirrors the FLARE input workbook style:
      • Sheet name : <case>_in
      • Coloured header sections for Design, Initial Conditions, Run Control
      • Time-series table with labelled columns and units row
      • Column widths and number formats set for readability
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl required. pip install openpyxl")

    fn = f"{case}_in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{case}_in"

    # ── Style helpers ──────────────────────────────────────────────────────────
    def _hdr_fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _side():
        return Side(style='thin', color='999999')

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_BLUE  = _hdr_fill('BDD7EE')   # design parameters
    HDR_GREEN = _hdr_fill('C6EFCE')   # initial conditions
    HDR_AMBER = _hdr_fill('FFEB9C')   # run control
    HDR_GREY  = _hdr_fill('D9D9D9')   # table header
    BOLD      = Font(bold=True)
    TITLE     = Font(bold=True, size=12)

    def _write(row, col, value, font=None, fill=None, align=None, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font      = font
        if fill:  c.fill      = fill
        if align: c.alignment = align
        if fmt:   c.number_format = fmt
        return c

    center = Alignment(horizontal='center')
    left   = Alignment(horizontal='left')

    # ── Title ──────────────────────────────────────────────────────────────────
    _write(1, 1, 'PWR Dry Containment Response Simulator — Input File',
           font=TITLE)
    _write(2, 1, f'Case: {case}', font=Font(italic=True, color='595959'))
    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')

    row = 4

    # ── Section writer ─────────────────────────────────────────────────────────
    def _section(title, fill_color, items):
        """items: list of (key, value, units, description)"""
        nonlocal row
        _write(row, 1, title, font=BOLD, fill=fill_color)
        _write(row, 2, 'Value', font=BOLD, fill=fill_color, align=center)
        _write(row, 3, 'Units', font=BOLD, fill=fill_color, align=center)
        _write(row, 4, 'Description', font=BOLD, fill=fill_color)
        row += 1
        for key, val, units, desc in items:
            assignment = f"{key} = {val}"
            _write(row, 1, assignment, font=Font(name='Courier New', size=10))
            _write(row, 2, val,   align=center)
            _write(row, 3, units, align=center,
                   font=Font(italic=True, color='595959'))
            _write(row, 4, desc,  font=Font(color='595959'))
            row += 1
        row += 1   # blank spacer

    _section('DESIGN PARAMETERS', HDR_BLUE, [
        ('air_volume_m3',     50667,  'm³',      'Free gas volume inside containment'),
        ('floor_area_m2',       600,  'm²',      'Projected containment floor area'),
        ('sump_area_m2',         25,  'm²',      'Engineered sump plan/free-surface area exposed before overflow'),
        ('sump_depth_m',        2.0,  'm',       'Sump depth to overflow onto containment floor'),
        ('wall_surf_area_m2',  8000,  'm²',      'Total condensing wall + liner area'),
        ('wall_thickness_m',  0.044,  'm',       'Steel liner thickness'),
        ('wall_char_height_m', 12.0,  'm',       'Characteristic height for Uchida correlation'),
        ('wall_k',             50.0,  'W/(m·K)', 'Liner thermal conductivity'),
        ('wall_rho',           7800,  'kg/m³',   'Liner density'),
        ('wall_cp',             500,  'J/(kg·K)','Liner specific heat'),
        ('gap_thick_m',       0.001,  'm',       'Liner/concrete air gap thickness (Table V: 0.225–3 mm)'),
        ('gap_k',             0.026,  'W/(m·K)', 'Gap effective conductivity (air)'),
        ('conc_k',              1.5,  'W/(m·K)', 'Concrete thermal conductivity (±30% uncertainty)'),
        ('conc_rho',           2300,  'kg/m³',   'Concrete density (±20% uncertainty)'),
        ('conc_cp',             900,  'J/(kg·K)','Concrete specific heat (±30% uncertainty)'),
        ('conc_depth_m',        1.5,  'm',       'Concrete modelled depth (semi-infinite approx)'),
        ('P_design_kPa',        517,  'kPa',     'Design pressure (informational)'),
        ('T_design_C',          177,  '°C',      'Design temperature (informational)'),
        ('uchida_multiplier',   1.0,  '—',       'Condensation model multiplier (Uchida: 0.5–1.5; DLM-FM: 0.8–1.2)'),
        ('condensation_model', 'uchida', '—',   'Condensation model: uchida or dlm_fm'),
        ('pcc_fraction',        0.0,  '—',       'Fraction of wall with active outer cooling (0=adiabatic, 1=PCC/AP1000)'),
        ('T_outer_C',          15.0,  '°C',      'Fixed outer surface temperature when pcc_fraction > 0'),
    ])

    _section('INITIAL CONDITIONS', HDR_GREEN, [
        ('P0_kPa',    101.325,  'kPa', 'Initial total pressure'),
        ('T0_C',         27.0,  '°C',  'Initial gas temperature'),
        ('T_wall0_C',    27.0,  '°C',  'Initial wall (liner) temperature'),
        ('RH0',          0.20,  '—',   'Initial relative humidity (0–1)'),
    ])

    _section('RUN CONTROL', HDR_AMBER, [
        ('endtime',  86400, 's', 'Simulation end time'),
        ('timestep',     2, 's', 'Integration timestep'),
    ])

    # ── Time-series table ──────────────────────────────────────────────────────
    _write(row, 1, 'TIME-SERIES SOURCE TABLE', font=BOLD)
    row += 1

    col_hdrs = [
        ('time [s]',         'A'),
        ('mdot_steam [kg/s]','B'),
        ('h_src [kJ/kg]',    'C'),
        ('mdot_h2 [kg/s]',   'D'),
        ('Qdot [W]',         'E'),
        ('(reserved)',        'F'),
    ]
    for ci, (label, _) in enumerate(col_hdrs, start=1):
        _write(row, ci, label, font=BOLD, fill=HDR_GREY, align=center)
    row += 1

    # Representative LOCA data (same as _build_loca_source)
    src = _build_loca_source()
    for i in range(len(src.time_s)):
        ws.cell(row=row, column=1).value = src.time_s[i]
        ws.cell(row=row, column=2).value = round(float(src.mdot_kg_s[i]),    3)
        ws.cell(row=row, column=3).value = round(float(src.h_src_J_kg[i]) / 1e3, 2)
        ws.cell(row=row, column=4).value = round(float(src.mdot_h2_kg_s[i]), 4)
        ws.cell(row=row, column=5).value = float(src.Qdot_W[i])
        ws.cell(row=row, column=6).value = 0.0
        row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    # ── Freeze panes below title rows ─────────────────────────────────────────
    ws.freeze_panes = 'A3'

    wb.save(fn)
    print(f"  Template written → {fn}")
    print(f"  Edit the file, then run:  python containment_sim.py {case}")


# ─────────────────────────────────────────────────────────────────────────────
# Main simulation runner from Excel input
# ─────────────────────────────────────────────────────────────────────────────

def run_from_flare_csv(flare_csv: str,
                       cont_case: str = None,
                       t_end_s:   float = 86400.0,
                       dt_s:      float = 2.0,
                       source_multiplier: float = 1.0,
                       decay_heat_fraction: float = 1.0,
                       rated_power_MWt: float = None) -> dict:
    """
    Full containment simulation workflow driven by a FLARE output CSV
    plus a containment-specific Excel input file.

    The FLARE CSV supplies the mass and energy source term only.
    All containment geometry, initial conditions, and run control parameters
    must be provided via a containment Excel input file (<cont_case>_in.xlsx).

    Parameters
    ----------
    flare_csv  : path to FLARE output CSV  (<case>_out.csv)
    cont_case  : base name of the containment input workbook
                 (default: same stem as flare_csv with '_out' removed)
    t_end_s    : simulation end time [s]  (default 24 hr)
    dt_s       : integration timestep [s]

    Usage
    -----
        # Generate a containment input template first:
        python containment_sim.py --template MyContainment

        # Then run with FLARE source term:
        python containment_sim.py --flare CaseLBLOCA_lpsi_out.csv MyContainment
    """
    flare_stem = Path(flare_csv).stem.replace('_out', '')
    if cont_case is None:
        cont_case = flare_stem

    # ── Load source term from FLARE ────────────────────────────────────────────
    source = build_source_from_flare(flare_csv, t_extend_s=t_end_s,
                                     source_multiplier=source_multiplier,
                                     decay_heat_fraction=decay_heat_fraction,
                                     rated_power_MWt=rated_power_MWt)

    # ── Load containment design and IC from its own Excel input ───────────────
    cont_fn = f"{cont_case}_in.xlsx"
    if not Path(cont_fn).exists():
        raise FileNotFoundError(
            f"\n  Containment input file not found: {cont_fn}"
            f"\n  Generate a template with:"
            f"\n    python containment_sim.py --template {cont_case}"
            f"\n  Then edit the geometry/IC to match the actual plant design."
        )

    print(f"\n  Reading containment input: {cont_fn}")
    design, ic, _src_from_xl, t_end_xl, dt_xl, _ = read_excel_input(cont_case)

    # Source term always comes from FLARE; run control from Excel unless overridden
    if t_end_s == 86400.0:   # user did not override → use Excel endtime
        t_end_s = t_end_xl
    if dt_s == 2.0:
        dt_s = dt_xl

    print(f"\n{'='*65}")
    print(f"  Containment simulation driven by FLARE output")
    print(f"  FLARE case       : {flare_stem}")
    print(f"  Containment input: {cont_fn}")
    print(f"{'='*65}")

    sim     = ContainmentSimulator(design, ic, source, dt_s=dt_s, verbose=True)
    results = sim.run(t_end=t_end_s)

    out_prefix = f"{flare_stem}_{cont_case}"
    print_summary(results)
    plot_results(results,
                 title=f"Containment Response — {flare_stem} / {cont_case}",
                 outfile=f"{out_prefix}_loca.png",
                 floor_area_m2=design.floor_area_m2)
    plot_heat_removal(results,
                      title=f"Heat Removal — {flare_stem} / {cont_case}",
                      outfile=f"{out_prefix}_heat.png")
    plot_shapiro(results,
                 title=f"Steam / Air / H₂  Shapiro — {flare_stem} / {cont_case}",
                 outfile=f"{out_prefix}_shapiro.png")
    save_csv(results, f"{out_prefix}_out.csv",
             floor_area_m2=design.floor_area_m2)

    return results


def run_from_excel(case: str) -> dict:
    """
    Full simulation workflow driven by  <case>_in.xlsx:
      1. Read and parse Excel input
      2. Run ContainmentSimulator
      3. Save plots (time-series + Shapiro) and CSV
      4. Return results dict

    Usage
    -----
        python containment_sim.py <case>
        python containment_sim.py --template <case>   # generate blank template
    """
    print(f"\n{'='*65}")
    print(f"  Reading input: {case}_in.xlsx")
    print(f"{'='*65}")

    design, ic, source, t_end, dt, echo_log = read_excel_input(case)

    print(f"\n  t_end    = {t_end:.0f} s  ({t_end/3600:.2f} hr)")
    print(f"  timestep = {dt:.2f} s")

    sim     = ContainmentSimulator(design, ic, source, dt_s=dt, verbose=True)
    results = sim.run(t_end=t_end)

    print_summary(results)
    plot_results(results,
                 title=f"PWR Dry Containment Response — {case}",
                 outfile=f"{case}_loca.png",
                 floor_area_m2=design.floor_area_m2)
    plot_shapiro(results,
                 title=f"Steam / Air / H₂  Shapiro Diagram — {case}",
                 outfile=f"{case}_shapiro.png")
    save_csv(results, f"{case}_out.csv",
             floor_area_m2=design.floor_area_m2)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os, shutil

    args = sys.argv[1:]

    # ── --flare <csv> [cont_case] [t_end] [dt]  : run from FLARE output CSV ───
    if args and args[0] == '--flare':
        if len(args) < 2:
            print("Usage: python containment_sim.py --flare <case>_out.csv"
                  " [cont_case] [t_end_s] [dt_s] [--scale N]")
            print("  --scale N : multiply all mass/energy flows by N  (e.g. 0.56)")
            sys.exit(1)
        _flare_csv  = args[1]
        _cont_case  = None
        _t_end      = 86400.0
        _dt         = 2.0
        _scale      = 1.0
        _idx        = 2
        # Parse optional --scale flag anywhere in remaining args
        _args_rest = args[_idx:]
        if '--scale' in _args_rest:
            _si = _args_rest.index('--scale')
            try: _scale = float(_args_rest[_si + 1])
            except (IndexError, ValueError): pass
            _args_rest = [a for i, a in enumerate(_args_rest)
                          if i != _si and i != _si + 1]
        # Remaining positional args: [cont_case] [t_end] [dt]
        _pos = [a for a in _args_rest if a]
        if _pos and not _pos[0].replace('.','').isdigit():
            _cont_case = _pos.pop(0)
        if _pos:
            try: _t_end = float(_pos.pop(0))
            except ValueError: pass
        if _pos:
            try: _dt = float(_pos.pop(0))
            except ValueError: pass
        run_from_flare_csv(_flare_csv, cont_case=_cont_case,
                           t_end_s=_t_end, dt_s=_dt,
                           source_multiplier=_scale)
        sys.exit(0)

    # ── --template <case>  : generate a blank input workbook ─────────────────
    if args and args[0] == '--template':
        case = args[1] if len(args) > 1 else 'containment'
        write_excel_template(case)
        sys.exit(0)

    # ── <case>  : run from Excel input file ───────────────────────────────────
    if args and not args[0].lstrip('-').isdigit():
        case = args[0]
        run_from_excel(case)
        sys.exit(0)

    # ── legacy numeric arguments: python containment_sim.py [t_end] [dt] ─────
    t_end = 7200.0
    dt    = 2.0
    if len(args) > 0:
        try: t_end = float(args[0])
        except ValueError: pass
    if len(args) > 1:
        try: dt = float(args[1])
        except ValueError: pass

    run_loca_example(t_end=t_end, dt=dt, outfile_prefix="containment")
    print("Done.")

# ─────────────────────────────────────────────────────────────────────────────
# FLARE integrated workflow helpers
# ─────────────────────────────────────────────────────────────────────────────

def find_flarecon_sheet(workbook_path: str) -> str | None:
    """Return the actual worksheet name for case-insensitive FLARECON, or None."""
    if not _OPENPYXL:
        return None
    p = Path(workbook_path)
    if not p.exists():
        return None
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            if str(name).strip().casefold() == "flarecon":
                return name
    finally:
        wb.close()
    return None


def _workbook_with_single_flarecon_sheet(workbook_path: str, sheet_name: str, cont_case: str) -> Path:
    """Copy the FLARECON sheet values into a temporary <cont_case>_in workbook."""
    tmp_dir = Path(tempfile.gettempdir()) / f"FLARECON_{os.getpid()}_{cont_case}"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out = tmp_dir / f"{cont_case}_in.xlsx"
    src_wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        src_ws = src_wb[sheet_name]
        dst_wb = openpyxl.Workbook()
        dst_ws = dst_wb.active
        dst_ws.title = f"{cont_case}_in"
        for r_idx, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                if c_idx == 1 and isinstance(val, str) and val.strip().lower().startswith("source_mode") and "=" in val:
                    val = 'source_mode = "table"  # overridden internally; FLARE CSV is coupled by run_integrated_flarecon'
                dst_ws.cell(row=r_idx, column=c_idx).value = val
        dst_wb.save(str(out))
        dst_wb.close()
        # Keep the original FLARE input workbook beside the extracted sheet so
        # relative flare_input_file references and NBT spray flags can resolve.
        try:
            shutil.copy2(str(workbook_path), str(tmp_dir / Path(workbook_path).name))
        except Exception:
            pass
    finally:
        src_wb.close()
    return out


def _echo_value(echo_log: list, name: str, default=None):
    for e in reversed(echo_log or []):
        if e.get("Variable") == name:
            return e.get("Value")
    return default


def composition_mole_fractions(results: dict) -> dict:
    """Return air/steam/H2 mole fractions from result mass inventories."""
    m_air = np.asarray(results.get('M_air_kg', []), dtype=float)
    m_stm = np.asarray(results.get('M_steam_kg', []), dtype=float)
    m_h2  = np.asarray(results.get('M_h2_kg', []), dtype=float)
    n_air = np.maximum(m_air, 0.0) / MW_AIR
    n_stm = np.maximum(m_stm, 0.0) / MW_H2O
    n_h2  = np.maximum(m_h2,  0.0) / MW_H2
    ntot = np.maximum(n_air + n_stm + n_h2, 1.0e-30)
    return {"x_air_mole": n_air/ntot, "x_steam_mole": n_stm/ntot, "x_h2_mole": n_h2/ntot}


def classify_hydrogen_flammability(results: dict) -> dict:
    """Approximate Shapiro-diagram classification for UI/header reporting."""
    comp = composition_mole_fractions(results)
    xh2 = comp["x_h2_mole"]
    xs  = comp["x_steam_mole"]
    deflag = (xs < STEAM_INERT_MOLE) & (xh2 >= H2_LFL_MOLE) & (xh2 <= H2_UFL_MOLE)
    deton  = deflag & (xh2 >= H2_DET_LO_MOLE) & (xh2 <= H2_DET_HI_MOLE) & (xs < 0.35)
    status = "Non-flammable / steam-inerted"
    if bool(np.any(deton)):
        status = "Detonation-region entry"
    elif bool(np.any(deflag)):
        status = "Deflagration-region entry"
    return {
        "status": status,
        "deflagration": bool(np.any(deflag)),
        "detonation": bool(np.any(deton)),
        "peak_h2_mole_fraction": float(np.nanmax(xh2)) if len(xh2) else 0.0,
        "peak_steam_mole_fraction": float(np.nanmax(xs)) if len(xs) else 0.0,
    }


def containment_scalar_summary(results: dict, floor_area_m2: float = 600.0) -> dict:
    comp = composition_mole_fractions(results)
    flm = classify_hydrogen_flammability(results)
    sump_level = np.asarray(results.get('sump_level_m', []), dtype=float)
    floor_level = np.asarray(results.get('floor_flood_level_m', []), dtype=float)
    if sump_level.size == 0:
        sump_level = np.asarray(results.get('V_liq_m3', []), dtype=float) / max(float(floor_area_m2), 1.0)
    return {
        "Peak Containment Pressure [kPa]": float(np.nanmax(results['P_kPa'])),
        "Peak Containment Temperature [C]": float(np.nanmax(results['T_C'])),
        "Peak H2 Concentration [vol%]": 100.0 * flm["peak_h2_mole_fraction"],
        "Peak Steam Concentration [vol%]": 100.0 * flm["peak_steam_mole_fraction"],
        "Peak Sump Level [m]": float(np.nanmax(sump_level)) if len(sump_level) else 0.0,
        "Peak Containment Floor Flood Level [m]": float(np.nanmax(floor_level)) if len(floor_level) else 0.0,
        "Hydrogen Flammability Status": flm["status"],
    }


def results_to_dataframe(results: dict, floor_area_m2: float = 600.0):
    import pandas as pd
    comp = composition_mole_fractions(results)
    fallback_level = np.asarray(results['V_liq_m3'], dtype=float) / max(float(floor_area_m2), 1.0)
    sump_level = np.asarray(results.get('sump_level_m', fallback_level), dtype=float)
    floor_level = np.asarray(results.get('floor_flood_level_m', np.zeros_like(fallback_level)), dtype=float)
    V_sump = np.asarray(results.get('V_sump_m3', np.zeros_like(fallback_level)), dtype=float)
    V_floor = np.asarray(results.get('V_floor_m3', results['V_liq_m3']), dtype=float)
    data = {
        'Time [s]': results['t'],
        'Pressure [kPa]': results['P_kPa'],
        'Gas Temp [C]': results['T_C'],
        'Wall Temp Inner [C]': results['T_wall_C'],
        'Concrete Surface Temp [C]': results['T_conc_C'],
        'Wet Floor/Sump Temp [C]': results['T_pool_C'],
        'Basemat Surface Temp [C]': results['T_bsmt_C'],
        'Liquid Volume [m3]': results['V_liq_m3'],
        'Sump Liquid Volume [m3]': V_sump,
        'Floor Liquid Volume [m3]': V_floor,
        'Sump Level [m]': sump_level,
        'Containment Floor Flood Level [m]': floor_level,
        'NC Quality [-]': results['x_nc'],
        'Air Mass [kg]': results['M_air_kg'],
        'Steam in Gas [kg]': results['M_steam_kg'],
        'H2 Mass [kg]': results['M_h2_kg'],
        'H2 Mole Fraction [-]': comp['x_h2_mole'],
        'H2 Concentration [vol%]': 100.0 * comp['x_h2_mole'],
        'Steam Mole Fraction [-]': comp['x_steam_mole'],
        'Steam Concentration [vol%]': 100.0 * comp['x_steam_mole'],
        'Air Mole Fraction [-]': comp['x_air_mole'],
        'Condensed Mass [kg]': results['M_cond_kg'],
        'Cond Rate [kg/s]': results['mdot_cond'],
        'Wall HF [W/m2]': results['q_cond'],
        'Wall Heat Removal [W]': results['Q_wall_W'],
        'Dry Floor Heat Removal [W]': results['Q_floor_dry_W'],
        'Wet Floor Heat Removal [W]': results['Q_pool_W'],
        'Spray Heat Removal [W]': results['Q_spray_W'],
        'Spray Flow [kg/s]': results['mdot_spray'],
        'Spray Cond Rate [kg/s]': results['mdot_spray_cond'],
        'Source mdot [kg/s]': results['mdot_in'],
        'Source Enthalpy Flux [W]': results['Hdot_in_W'],
    }
    return pd.DataFrame(data)


def save_integrated_outputs(results: dict, flare_case: str, output_dir: str | Path, flare_out_xlsx: str | Path, floor_area_m2: float = 600.0) -> dict:
    import pandas as pd
    output_dir = Path(output_dir)
    con_csv = output_dir / f"{flare_case}-CON.csv"
    df = results_to_dataframe(results, floor_area_m2=floor_area_m2)
    df.to_csv(con_csv, index=False)

    try:
        from openpyxl import load_workbook
        xlsx_path = Path(flare_out_xlsx)
        if xlsx_path.exists():
            wb = load_workbook(str(xlsx_path))
            if "FLARECON" in wb.sheetnames:
                del wb["FLARECON"]
            ws = wb.create_sheet("FLARECON")
            summary = containment_scalar_summary(results, floor_area_m2=floor_area_m2)
            ws.append(["FLARECON Scalar Summary"])
            for k, v in summary.items():
                ws.append([k, v])
            ws.append([])
            ws.append(list(df.columns))
            for row in df.itertuples(index=False, name=None):
                ws.append(list(row))
            wb.save(str(xlsx_path))
            wb.close()
    except Exception as e:
        print(f"WARNING: could not append FLARECON worksheet: {e}", flush=True)

    return {"csv": str(con_csv), "summary": containment_scalar_summary(results, floor_area_m2=floor_area_m2)}


def run_integrated_flarecon(flare_case: str, flare_input_xlsx: str, flare_output_csv: str, output_dir: str | Path = ".", make_plots: bool = True) -> dict | None:
    """Run FLARECON after a FLARE run if a case-insensitive FLARECON sheet exists."""
    sheet = find_flarecon_sheet(flare_input_xlsx)
    if not sheet:
        return None
    print("FLARECON_PRE_PROCESSING", flush=True)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cont_case = f"{flare_case}_FLARECON"
    tmp_wb = _workbook_with_single_flarecon_sheet(flare_input_xlsx, sheet, cont_case)
    old_cwd = Path.cwd()
    try:
        os.chdir(tmp_wb.parent)
        design, ic, _src, t_end, dt, echo = read_excel_input(cont_case)
    finally:
        os.chdir(old_cwd)

    # FLARE output is the authoritative source term in the integrated workflow.
    late_frac = float(_echo_value(echo, 'late_steam_fraction', _echo_value(echo, 'decay_heat_fraction', 1.0)))
    src_mult  = float(_echo_value(echo, 'source_multiplier', 1.0))
    rated     = float(_echo_value(echo, 'rated_power_MWt', 575.0))
    late_h    = float(_echo_value(echo, 'late_steam_enthalpy_kJkg', 2738.0)) * 1e3
    source = build_source_from_flare(
        str(flare_output_csv),
        t_extend_s=float(t_end),
        decay_heat_fraction=late_frac,
        source_multiplier=src_mult,
        rated_power_MWt=rated,
        late_steam_enthalpy_Jkg=late_h,
    )
    sim = ContainmentSimulator(design, ic, source, dt_s=float(dt), verbose=True)
    print("FLARECON_SIMULATION", flush=True)
    results = sim.run(t_end=float(t_end))
    print("FLARECON_POST_PROCESSING", flush=True)
    print_summary(results)

    if make_plots:
        prefix = output_dir / f"{flare_case}-CON"
        plot_results(results, title=f"Containment Response — {flare_case}", outfile=str(prefix) + "_response.png", floor_area_m2=design.floor_area_m2)
        plot_heat_removal(results, title=f"Containment Heat Removal — {flare_case}", outfile=str(prefix) + "_heat.png")
        plot_shapiro(results, title=f"Steam / Air / H₂ Shapiro — {flare_case}", outfile=str(prefix) + "_shapiro.png")

    flare_out_xlsx = output_dir / f"{flare_case}_out.xlsx"
    saved = save_integrated_outputs(results, flare_case, output_dir, flare_out_xlsx, floor_area_m2=design.floor_area_m2)
    status = saved.get('summary', {}).get('Hydrogen Flammability Status', '')
    print(f"FLARECON: wrote {saved.get('csv')}", flush=True)
    print(f"FLARECON: hydrogen flammability status = {status}", flush=True)
    return {"results": results, "design": design, "summary": saved.get("summary", {}), "csv": saved.get("csv")}
