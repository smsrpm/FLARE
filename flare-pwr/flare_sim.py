"""
Fast Licensing Accident Response Engine - FLARE
Original author: R. P. Martin  Version 1.0  May 20 2026

"""

from __future__ import annotations
import sys, shutil, subprocess, importlib, importlib.util, os, functools, tempfile
from typing import Dict, Iterable, Optional, Tuple
from pathlib import Path

os.environ['PYTHONUNBUFFERED'] = '1'
print = functools.partial(print, flush=True)


def _runtime_dir(base_dir: Path | None = None) -> Path:
    """Return the FLARE runtime folder, preferring runtime/ then Runtime/."""
    base = Path(base_dir) if base_dir is not None else Path(__file__).parent
    lower = base / "runtime"
    upper = base / "Runtime"
    if lower.exists():
        rt = lower
    elif upper.exists():
        rt = upper
    else:
        rt = lower
        rt.mkdir(parents=True, exist_ok=True)
    return rt

# ── preflight ────────────────────────────────────────────────────────────────
def detect_platform() -> dict:
    """
    Detect the runtime platform and return a dict of capability flags.

    Keys
    ----
    'is_ios'        — iPhone/iPad (Pyto or Pythonista)
    'is_android'    — Android (QPython / Termux)
    'is_windows'    — Windows desktop
    'is_macos'      — macOS desktop
    'is_linux'      — Linux desktop / server
    'runtime'       — human-readable label ('Pyto', 'Windows-CPython', …)
    'has_display'   — GUI display likely available
    'has_network'   — assumed True (no reliable cross-platform check)
    'can_bind_port' — local port binding (Streamlit / Flask) likely allowed
    """
    info = dict(is_ios=False, is_android=False, is_windows=False,
                is_macos=False,  is_linux=False,   runtime='CPython',
                has_display=True, has_network=True, can_bind_port=True)

    exe  = sys.executable.lower()
    argv0 = (sys.argv[0].lower() if sys.argv else "")
    plat  = sys.platform.lower()

    # Pyto (iOS)
    if "pyto" in exe or "pyto" in argv0:
        info.update(is_ios=True, runtime='Pyto',
                    has_display=False, can_bind_port=False)
        return info

    # Pythonista (iOS)
    try:
        import console                          # Pythonista-specific built-in
        info.update(is_ios=True, runtime='Pythonista',
                    has_display=False, can_bind_port=False)
        return info
    except ImportError:
        pass

    # Generic iOS fallback: no /proc, no /usr, not Windows, not macOS
    if (not os.path.exists('/proc') and not os.path.exists('/usr')
            and not plat.startswith('win') and not plat.startswith('darwin')):
        info.update(is_ios=True, runtime='iOS-unknown',
                    has_display=False, can_bind_port=False)
        return info

    # Android (Termux / QPython)
    if os.path.exists('/data/data') or 'com.termux' in exe or 'qpython' in exe:
        info.update(is_android=True, runtime='Android',
                    has_display=False, can_bind_port=False)
        return info

    # Standard desktop platforms
    if plat.startswith('win'):
        info.update(is_windows=True, runtime='Windows-CPython')
    elif plat.startswith('darwin'):
        info.update(is_macos=True, runtime='macOS-CPython')
    elif plat.startswith('linux'):
        info.update(is_linux=True, runtime='Linux-CPython')
        if not os.environ.get('DISPLAY') and not os.environ.get('WAYLAND_DISPLAY'):
            info['has_display'] = False   # headless server / CI

    return info


def _has_module(modname):
    try:
        importlib.import_module(modname); return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _pip_available_for_this_python():
    if importlib.util.find_spec("pip") is None:
        return False, "pip module not found."
    try:
        cp = subprocess.run([sys.executable, "-m", "pip", "--version"],
                            capture_output=True, text=True, check=False)
        if cp.returncode != 0:
            return False, f"`pip` failed: {(cp.stderr or cp.stdout or '').strip()}"
        return True, (cp.stdout or "").strip()
    except Exception as e:
        return False, f"Unable to run pip: {e}"


def preflight(min_python=(3,10), required_modules=None,
              required_commands=None, platform=None) -> dict:
    """
    Run environment preflight checks, skipping packages incompatible with
    the detected platform (e.g. Streamlit/plotly/PyQt5 on iOS).

    Returns the platform dict so callers can branch on capabilities.
    """
    required_modules  = required_modules  or {}
    required_commands = list(required_commands or [])
    if platform is None:
        platform = detect_platform()

    # Packages that require local port binding
    _PORT_PKG    = {'streamlit', 'flask', 'fastapi', 'uvicorn', 'tornado'}
    # Packages that require a GUI display
    _DISPLAY_PKG = {'PyQt5', 'PyQt6', 'tkinter', 'wx', 'plotly'}

    skipped, filtered = [], {}
    for imp, pip in required_modules.items():
        reason = None
        if not platform['can_bind_port'] and imp in _PORT_PKG:
            reason = "requires local port binding (not supported on this platform)"
        elif not platform['has_display'] and imp in _DISPLAY_PKG:
            reason = "requires a GUI display (not available on this platform)"
        if reason:
            skipped.append((imp, reason))
        else:
            filtered[imp] = pip

    problems, warnings = [], []

    if sys.version_info < min_python:
        problems.append(f"Python {min_python[0]}.{min_python[1]}+ required.")

    pip_ok, pip_info = _pip_available_for_this_python()
    if not pip_ok:
        warnings.append(f"pip unavailable ({pip_info}). Manual install may be needed.")

    missing, broken = [], []
    for imp, pip in filtered.items():
        ok, err = _has_module(imp)
        if not ok:
            if imp == "matplotlib":
                warnings.append(f"matplotlib import issue ({err}); will retry with Agg backend.")
            else:
                (missing if (err or "").startswith("ModuleNotFoundError")
                 else broken).append((imp, pip, err))

    if [c for c in required_commands if shutil.which(c) is None]:
        problems.append("Commands not on PATH.")
    if missing:
        problems.append("Missing: " + ", ".join(f"{i} (pip install {p})" for i,p,*_ in missing))
    if broken:
        problems.append("Broken: "  + ", ".join(f"{i}" for i,*_ in broken))

    if skipped:
        print(f"\n[Preflight] Platform detected : {platform['runtime']}")
        print("[Preflight] Skipping incompatible packages (not an error):")
        for name, reason in skipped:
            print(f"  • {name}: {reason}")

    if warnings:
        for w in warnings: print(f"[Preflight] Warning: {w}")

    if problems:
        print("\n=== Preflight FAILED ===")
        for p in problems: print(f"\n• {p}")
        print(f"\n  Python  : {sys.executable}")
        print(f"  Version : {sys.version.split()[0]}")
        print(f"  Run     : python -m pip install <package>")
        print(f"  Tip     : If pip install fails, try creating a fresh")
        print(f"            virtual environment:")
        print(f"            python -m venv flare_env")
        print(f"            flare_env\\Scripts\\activate  (Windows)")
        print(f"            source flare_env/bin/activate  (Mac/Linux)")
        print(f"            python -m pip install numpy pandas scipy")
        print(f"              openpyxl matplotlib XSteamPython plotly streamlit reportlab")
        raise SystemExit(1)

    print(f"[Preflight] OK — {platform['runtime']}  Python {sys.version.split()[0]}")
    return platform


# Detect platform at import time; available everywhere in the module as _platform
_platform = detect_platform()

if __name__ == "__main__":
    # ── Fast preflight ────────────────────────────────────────────────────────
    # The full preflight (pip subprocess + 7 importlib calls) costs 2-5 s on
    # Windows with a cold Python process.  We skip it if a sentinel file from
    # a previous successful check exists in the same folder.  The sentinel is
    # invalidated whenever the Python executable path changes (e.g. new venv).
    import hashlib as _hashlib
    _sentinel = _runtime_dir(Path(__file__).parent) / ".preflight_ok"
    _exe_hash = _hashlib.md5(sys.executable.encode()).hexdigest()[:12]
    _sentinel_tag = f"{_exe_hash}\n"

    _skip_preflight = (
        _sentinel.exists() and
        _sentinel.read_text(encoding="utf-8").strip() == _exe_hash
    )

    if _skip_preflight:
        print(f"[Preflight] OK (cached) — {_platform['runtime']}  Python {sys.version.split()[0]}")
    else:
        # First run or new environment: do the full check but skip the slow
        # pip subprocess — if packages are missing we'll get a clear ImportError
        # below anyway.
        _REQUIRED = {
            "numpy":        "numpy",
            "pandas":       "pandas",
            "scipy":        "scipy",
            "openpyxl":     "openpyxl",
            "matplotlib":   "matplotlib",
            "XSteamPython": "XSteamPython",
            "reportlab":    "reportlab",
        }
        _missing = []
        for _mod, _pip in _REQUIRED.items():
            _ok, _err = _has_module(_mod)
            if not _ok:
                _missing.append(f"{_mod}  (pip install {_pip})")
        if _missing:
            print("\n=== Preflight FAILED ===")
            for _m in _missing:
                print(f"  Missing: {_m}")
            raise SystemExit(1)
        # Write sentinel so subsequent launches skip this block
        _sentinel.write_text(_sentinel_tag, encoding="utf-8")
        print(f"[Preflight] OK — {_platform['runtime']}  Python {sys.version.split()[0]}")

# ── standard imports ─────────────────────────────────────────────────────────
import XSteamPython as XSteam
from scipy.optimize import brentq
from scipy.stats   import norm as _norm_dist
import pandas as pd
import numpy as np
import openpyxl
import math as _math
import time as timer

SCRIPT_DIR = Path(__file__).resolve().parent

# Do not force the process cwd back to SCRIPT_DIR here.
# flare_ui.py launches flare_sim.py with cwd set to the run folder:
#     sim_<case>_<timestamp>
# so outputs, input CSV snapshots, and archived inputs should be written there.
# When flare_sim.py is run directly from the command line, cwd remains whatever
# the user selected, preserving normal standalone behavior.

mpldir = Path(tempfile.gettempdir()) / "mpl"
mpldir.mkdir(parents=True, exist_ok=True)
os.environ["MPLCONFIGDIR"] = str(mpldir)
os.environ["MATPLOTLIBRC"] = ""

# Select matplotlib backend using _platform (already detected above).
#   - GUI backend when a display is available (Windows, macOS, Linux+DISPLAY)
#   - Agg on iOS (Pyto), headless servers, or any no-display environment
def _running_in_spyder():
    """Return True when executed inside the Spyder IDE."""
    return (os.environ.get("SPY_PYTHONPATH") is not None or
            os.environ.get("SPYDER_ARGS")    is not None or
            "spyder" in os.environ.get("PYTHONSTARTUP", "").lower())

# _platform is already set by detect_platform() above; derive the two
# legacy flags from it so the rest of the file is unchanged.
_IN_SPYDER     = _running_in_spyder()
_GUI_AVAILABLE = _platform['has_display']

if not _GUI_AVAILABLE:
    os.environ["MPLBACKEND"] = "Agg"
elif _IN_SPYDER:
    # Switch the IPython kernel to Qt5 so figures open in separate windows
    # rather than appearing in the Plots pane.  Equivalent to typing
    # %matplotlib qt5 in the Spyder console.
    try:
        from IPython import get_ipython
        _ipy = get_ipython()
        if _ipy is not None:
            _ipy.run_line_magic("matplotlib", "qt5")
    except Exception:
        pass   # not in an IPython kernel — no action needed

try:
    import matplotlib
    if not _GUI_AVAILABLE:
        matplotlib.use("Agg")
    elif _IN_SPYDER:
        pass   # backend already set by %matplotlib qt5 above
    # On macOS outside Spyder, prefer the native backend unless already set
    elif matplotlib.get_backend() == "":
        try:
            matplotlib.use("MacOSX")
        except Exception:
            pass
except Exception:
    pass
# matplotlib.pyplot is imported lazily inside post_processing() to avoid
# adding ~1.4 s to every simulation startup.  Do not import it here.

# ── PATCH 1: pump_module integrated inline ───────────────────────────────────
# PUMP MODULE  (formerly pump_module.py — integrated)
# RCP coastdown + SG model, all-SI interface
# Homologous curves: Westinghouse / RELAP5-3D format
# Author: R. P. Martin
# ═══════════════════════════════════════════════════════════════════════════════
# ── Optional XSteamPython for fluid properties ─────────────────────────────
# XSteamPython is imported unconditionally above; set flag used by pump functions
_XSTEAM = True


# ============================================================================
# RATED CONDITIONS  (SI/metric user-facing defaults; internal pump basis unchanged)
# ============================================================================
# These defaults preserve the original FLARE pump behavior.  They are exposed
# to input decks in SI/metric units and converted back to the original
# English-unit internal basis by configure_pump_model().
_DEFAULT_PUMP_RATED_SPEED_RPM      = 3550.0        # rev/min
_DEFAULT_PUMP_RATED_FLOW_M3S       = 5.747086      # m^3/s  (91121.7 gpm)
_DEFAULT_PUMP_RATED_HEAD_M         = 41.51376      # m      (136.2 ft)
_DEFAULT_PUMP_RATED_TORQUE_NM      = 5661.909968   # N-m    (4176 lbf-ft)
_DEFAULT_PUMP_RATED_DENSITY_KG_M3  = 682.386396    # kg/m^3 (42.6 lbm/ft^3)
_DEFAULT_PUMP_INERTIA_KG_M2        = 32.339143     # kg-m^2 (767 lbm-ft^2)
_DEFAULT_PUMP_FLOW_AREA_M2         = 0.464516      # m^2    (5.0 ft^2)
_DEFAULT_PUMP_FRICTION_TORQUE_NM   = 56.619100     # N-m    (41.76 lbf-ft)
_DEFAULT_PUMP_FRICTION_SPEED2_NM   = 56.619100     # N-m    (41.76 lbf-ft)

# Active values are retained internally in the English-unit basis used by the
# original homologous-curve pump implementation.
_RPM_R      = _DEFAULT_PUMP_RATED_SPEED_RPM
_Q_R_FT3S   = _DEFAULT_PUMP_RATED_FLOW_M3S / 0.02831685
_H_R_FT     = _DEFAULT_PUMP_RATED_HEAD_M / 0.3048
_TAU_R      = _DEFAULT_PUMP_RATED_TORQUE_NM / 1.355818
_RHO_R_BU   = _DEFAULT_PUMP_RATED_DENSITY_KG_M3 / 16.01846
_I_BU       = _DEFAULT_PUMP_INERTIA_KG_M2 / (0.453592 * 0.3048**2)
_A_BU       = _DEFAULT_PUMP_FLOW_AREA_M2 / (0.3048**2)
_TF0        = _DEFAULT_PUMP_FRICTION_TORQUE_NM / 1.355818
_TF2        = _DEFAULT_PUMP_FRICTION_SPEED2_NM / 1.355818

# Unit conversions (British → SI)
_LBM_FT3_SI = 16.01846        # kg/m³  per  lbm/ft³
_FT3S_SI    = 0.02831685      # m³/s   per  ft³/s
_LBFFT_SI   = 1.355818        # N·m    per  lbf·ft
_FT_SI      = 0.3048          # m      per  ft
_LBM_SI     = 0.453592        # kg     per  lbm
_FT2_SI     = 0.3048**2       # m²     per  ft²
# 1 lbm·ft²  =  0.453592 kg × (0.3048 m)²  =  0.04214 kg·m²
_LBMFT2_SI  = 0.453592 * 0.3048**2

def _refresh_pump_derived_constants():
    """Refresh SI values and rated loop resistance after pump scaling changes."""
    global _Q_R_SI, _RHO_R_SI, _I_SI, _TAU_R_SI, _A_SI, _R_RATED_BU
    _Q_R_SI     = _Q_R_FT3S * _FT3S_SI           # m³/s
    _RHO_R_SI   = _RHO_R_BU * _LBM_FT3_SI        # kg/m³
    _I_SI       = _I_BU * _LBMFT2_SI             # kg·m²
    _TAU_R_SI   = _TAU_R * _LBFFT_SI             # N·m
    _A_SI       = _A_BU * _FT2_SI                # m²
    _R_RATED_BU = 2.0 * 32.174 * _H_R_FT * _A_BU**2 / _Q_R_FT3S**2

def configure_pump_model(rated_speed_rpm=None, rated_flow_m3s=None,
                         rated_head_m=None, rated_torque_Nm=None,
                         rated_density_kg_m3=None, inertia_kg_m2=None,
                         flow_area_m2=None, friction_torque_Nm=None,
                         friction_speed2_Nm=None) -> dict:
    """
    Configure the scalar pump scaling constants from SI/metric user inputs.

    The homologous curve shapes remain the built-in Westinghouse/RELAP-style
    tables.  Only the dimensional anchors are user-configurable.  Inputs are
    exposed in SI/metric units and converted here to the English-unit internal
    basis used by the original pump equations.
    """
    global _RPM_R, _Q_R_FT3S, _H_R_FT, _TAU_R, _RHO_R_BU, _I_BU, _A_BU, _TF0, _TF2

    def _pos(value, default, name):
        try:
            v = float(default if value is None else value)
        except Exception:
            print(f"WARNING: invalid {name}; using default {default}")
            return float(default)
        if v <= 0.0:
            print(f"WARNING: {name} must be positive; using default {default}")
            return float(default)
        return v

    def _nonneg(value, default, name):
        try:
            v = float(default if value is None else value)
        except Exception:
            print(f"WARNING: invalid {name}; using default {default}")
            return float(default)
        if v < 0.0:
            print(f"WARNING: {name} must be nonnegative; using default {default}")
            return float(default)
        return v

    # Validate in user-facing SI/metric units.
    _speed_rpm  = _pos(rated_speed_rpm,      _DEFAULT_PUMP_RATED_SPEED_RPM,     "pump_rated_speed_rpm")
    _flow_m3s   = _pos(rated_flow_m3s,       _DEFAULT_PUMP_RATED_FLOW_M3S,      "pump_rated_flow_m3s")
    _head_m     = _pos(rated_head_m,         _DEFAULT_PUMP_RATED_HEAD_M,        "pump_rated_head_m")
    _torque_Nm  = _pos(rated_torque_Nm,      _DEFAULT_PUMP_RATED_TORQUE_NM,     "pump_rated_torque_Nm")
    _rho_kg_m3  = _pos(rated_density_kg_m3,  _DEFAULT_PUMP_RATED_DENSITY_KG_M3, "pump_rated_density_kg_m3")
    _inertia    = _pos(inertia_kg_m2,        _DEFAULT_PUMP_INERTIA_KG_M2,       "pump_inertia_kg_m2")
    _area_m2    = _pos(flow_area_m2,         _DEFAULT_PUMP_FLOW_AREA_M2,        "pump_flow_area_m2")
    _tf0_Nm     = _nonneg(friction_torque_Nm, _DEFAULT_PUMP_FRICTION_TORQUE_NM, "pump_friction_torque_Nm")
    _tf2_Nm     = _nonneg(friction_speed2_Nm, _DEFAULT_PUMP_FRICTION_SPEED2_NM, "pump_friction_speed2_Nm")

    # Convert SI/metric user inputs to the internal English-unit basis.
    _RPM_R    = _speed_rpm
    _Q_R_FT3S = _flow_m3s / _FT3S_SI
    _H_R_FT   = _head_m / _FT_SI
    _TAU_R    = _torque_Nm / _LBFFT_SI
    _RHO_R_BU = _rho_kg_m3 / _LBM_FT3_SI
    _I_BU     = _inertia / _LBMFT2_SI
    _A_BU     = _area_m2 / _FT2_SI
    _TF0      = _tf0_Nm / _LBFFT_SI
    _TF2      = _tf2_Nm / _LBFFT_SI

    _refresh_pump_derived_constants()
    return {
        "pump_rated_speed_rpm": _speed_rpm,
        "pump_rated_flow_m3s": _Q_R_SI,
        "pump_rated_head_m": _H_R_FT * _FT_SI,
        "pump_rated_torque_Nm": _TAU_R * _LBFFT_SI,
        "pump_rated_density_kg_m3": _RHO_R_SI,
        "pump_inertia_kg_m2": _I_SI,
        "pump_flow_area_m2": _A_SI,
        "pump_friction_torque_Nm": _TF0 * _LBFFT_SI,
        "pump_friction_speed2_Nm": _TF2 * _LBFFT_SI,
        "pump_rated_loop_resistance": _R_RATED_BU,
    }

# Derived values for the default configuration at import time.
_refresh_pump_derived_constants()


# ============================================================================
# HOMOLOGOUS CURVES  (Westinghouse, RELAP5-3D format, dimensionless)
# ============================================================================

# HAN : h/α²  vs  v/α       (already ascending in v/α → ready for np.interp)
_HAN = np.array([
    [0.00, 1.73],
    [0.20, 1.50],
    [0.46, 1.24],
    [0.52, 1.23],
    [0.60, 1.24],
    [0.66, 1.24],
    [0.80, 1.17],
    [0.90, 1.10],
    [1.00, 1.00],
])

# HVN : h/v²  vs  α/v       (original is descending in α/v; reversed here)
_HVN = np.array([
    [1.00,  1.00],
    [0.80,  0.37],
    [0.65,  0.00],
    [0.53, -0.30],
    [0.40, -0.54],
    [0.30, -0.70],
    [0.20, -0.81],
    [0.10, -0.90],
    [0.00, -0.96],
])[::-1]   # flip → ascending α/v for np.interp; h/v² now also ascending

# BAN : β/α²  vs  v/α       (ascending → ready for np.interp)
_BAN = np.array([
    [0.00, 1.01],
    [0.10, 0.96],
    [0.20, 0.92],
    [0.30, 0.90],
    [0.40, 0.89],
    [0.50, 0.91],
    [0.70, 0.99],
    [0.80, 1.02],
    [0.90, 1.02],
    [1.00, 1.00],
])

# BVN : β/v²  vs  α/v       (original is descending in α/v; reversed here)
_BVN = np.array([
    [1.00,  1.00],
    [0.74,  0.40],
    [0.40, -0.31],
    [0.30, -0.48],
    [0.20, -0.63],
    [0.10, -0.76],
    [0.00, -0.87],
])[::-1]   # flip → ascending α/v


# ============================================================================
# INTERNAL HELPERS
# ============================================================================

def _yh(orifice_area: float) -> float:
    """
    Dimensionless head ratio yh = loop_R / rated_loop_R.
    Equation 12 from 11111-000-JAC-YA-00002.
    """
    loop_R = ((1.0 / (0.62 + 0.38 * orifice_area**3) - orifice_area)**2
              / orifice_area**2)
    return loop_R / _R_RATED_BU


def _han_residual(voa: float, yh_val: float) -> float:
    """Residual for HAN root-find:  h/α² − yh·(v/α)² = 0."""
    h_alpha2 = np.interp(voa, _HAN[:, 0], _HAN[:, 1])
    return h_alpha2 - yh_val * voa**2


def _rho_bu(pressure_Pa: float, temperature_K: float) -> float:
    """
    Subcooled liquid density [lbm/ft³] from current RCS conditions.

    Uses saturated-liquid density as the lookup basis.  XSteam's rho_pT
    returns near-zero when the (P, T) point is interpreted as superheated
    steam, which can happen when the fluid temperature approaches Tsat
    during rapid depressurisation.  Clamping T to 0.9999·Tsat keeps the
    lookup in the liquid region and is physically accurate for a subcooled
    or marginally saturated primary coolant.

    Falls back to rated density (_RHO_R_BU) if XSteamPython is unavailable.
    """
    if _XSTEAM:
        try:
            P_kPa  = pressure_Pa * 1e-3
            T_C    = temperature_K - 273.15
            Tsat_C = XSteam.Tsat_p(P_kPa)
            T_lkp  = min(T_C, 0.9999 * Tsat_C)   # clamp to liquid region
            rho_si = XSteam.rhoL_T(T_lkp)         # saturated liquid density
            if rho_si > 1.0:                       # sanity check (> 1 kg/m³)
                return rho_si / _LBM_FT3_SI
        except Exception:
            pass
    return _RHO_R_BU


def _operating_point(alpha: float, yh_val: float, rho_bu_val: float) -> dict:
    """
    Solve for the pump operating point given speed ratio, head ratio, density.

    Parameters
    ----------
    alpha     : ω / ω_rated  (dimensionless speed ratio)
    yh_val    : loop_R / rated_loop_R  (dimensionless head ratio)
    rho_bu_val: fluid density [lbm/ft³]

    Returns
    -------
    dict with all quantities in British units (converted by callers to SI).
    """
    hvn_max = float(np.max(_HVN[:, 1]))   # maximum h/v² on HVN curve = 1.00

    if yh_val <= hvn_max:
        # ── HVN / BVN quadrant (low speed or high flow resistance) ──────────
        # Given h/v², find α/v by interpolation on the (now ascending) HVN table.
        aov = np.interp(yh_val, _HVN[:, 1], _HVN[:, 0])
        aov = max(aov, 1e-9)           # guard against divide-by-zero
        v   = alpha / aov
        yt  = np.interp(aov, _BVN[:, 0], _BVN[:, 1])
        torque_lbfft = yt * v**2 * _TAU_R
        voa = v / alpha if abs(alpha) > 1e-12 else 1.0

    else:
        # ── HAN / BAN quadrant (normal pumping operation) ───────────────────
        # Root-find v/α on HAN curve.
        # At voa = 0:  residual = 1.73 − 0  > 0
        # At voa = 1:  residual = 1.00 − yh < 0  (because yh > hvn_max = 1.00)
        # → guaranteed sign change on [0, 1].
        try:
            voa = brentq(_han_residual, 0.0, 1.0, args=(yh_val,), maxiter=200)
        except ValueError:
            voa = 1.0      # fallback: rated flow ratio
        v = voa * alpha
        yt = np.interp(voa, _BAN[:, 0], _BAN[:, 1])
        torque_lbfft = yt * alpha**2 * _TAU_R

    # Head [ft of fluid]
    head_ft = yh_val * v**2 * _H_R_FT

    # Friction torque [lbf·ft]  — speed-dependent
    tau_fric = _TF0 + _TF2 * alpha**2

    # Net shaft torque (negative ⟹ decelerating)
    tau_shaft = -(torque_lbfft * rho_bu_val / _RHO_R_BU + tau_fric)

    return {
        "v":           v,              # volumetric flow ratio
        "alpha":       alpha,          # speed ratio
        "head_ft":     head_ft,
        "torque_lbfft": torque_lbfft,
        "tau_shaft_lbfft": tau_shaft,
        "Q_ft3s":      _Q_R_FT3S * v,
        "mdot_lbms":   _Q_R_FT3S * v * rho_bu_val,
    }


# ============================================================================
# PUBLIC SI INTERFACE
# ============================================================================

def init(orifice_area: float,
         speed_rpm:    float,
         pressure_Pa:  float,
         temperature_K: float) -> dict:
    """
    Compute the steady-state pump operating point at t = 0.

    Parameters
    ----------
    orifice_area  : normalised orifice area  (0 < A ≤ 1.0)
    speed_rpm     : initial pump speed  [rpm]
    pressure_Pa   : RCS pressure  [Pa]
    temperature_K : RCS temperature  [K]

    Returns
    -------
    pump_state : dict
        omega_rpm      – pump speed  [rpm]
        omega_rads     – pump speed  [rad/s]
        tau_Nm         – net shaft torque  [N·m]
        mass_flow_kgs  – coolant mass flow rate  [kg/s]
        vol_flow_m3s   – coolant volumetric flow rate  [m³/s]
        velocity_ms    – mean coolant velocity  Q / A_flow  [m/s]
        head_Pa        – pump head  [Pa]
        yh             – dimensionless head ratio (fixed by loop geometry)
        coasted_down   – False
    """
    yh_val    = _yh(orifice_area)
    alpha     = speed_rpm / _RPM_R
    rho_bu_v  = _rho_bu(pressure_Pa, temperature_K)
    rho_si_v  = rho_bu_v * _LBM_FT3_SI

    op = _operating_point(alpha, yh_val, rho_bu_v)

    return {
        "omega_rpm":       speed_rpm,
        "omega_rads":      speed_rpm * 2.0 * np.pi / 60.0,
        "omega_rpm_trip":  speed_rpm,   # rated speed at moment of trip — fixed
        "tau_Nm":          op["tau_shaft_lbfft"] * _LBFFT_SI,
        "mass_flow_kgs":   op["mdot_lbms"]       * _LBM_SI,
        "vol_flow_m3s":    op["Q_ft3s"]          * _FT3S_SI,
        "velocity_ms":     op["Q_ft3s"] * _FT3S_SI / _A_SI,
        "head_Pa":         op["head_ft"] * _FT_SI * rho_si_v * 9.81,
        "yh":              yh_val,
        "coasted_down":    False,
        "t_coast_elapsed": 0.0,       # time elapsed since pump trip  [s]
    }


def step(state:         dict,
         pressure_Pa:   float,
         temperature_K: float,
         delta_t:       float,
         tau_s:         float = 30.0) -> dict:
    """
    Advance pump state by one timestep using an exponential coastdown model.

    Speed law:  ω(t) = ω_trip · exp(−t_elapsed / τ_s)

    This gives a pure exponential decay with time constant τ_s, which is
    the physically expected profile for an inertia-dominated RCP coastdown.
    It is also timestep-independent (no Euler error accumulation), making
    it robust for mixed coarse/fine time grids.

    Two-way coupling is preserved: fluid density at the current RCS state
    is evaluated each step and fed into the homologous-curve operating point
    so that flow, head, and torque respond to changing system conditions.

    Parameters
    ----------
    state         : dict returned by init() or a previous step()
    pressure_Pa   : RCS pressure at the *current* timestep  [Pa]
    temperature_K : RCS temperature at the *current* timestep  [K]
    delta_t       : timestep size  [s]  (used only to advance t_coast_elapsed)
    tau_s         : exponential time constant  [s]  (default 30 s)

    Returns
    -------
    new_state : updated dict with the same keys as init()
    """
    if state["coasted_down"]:
        return {**state,
                "mass_flow_kgs": 0.0, "vol_flow_m3s": 0.0,
                "velocity_ms":   0.0, "tau_Nm":        0.0,
                "head_Pa":       0.0}

    # Accumulate elapsed coastdown time
    t_elapsed_new = state.get("t_coast_elapsed", 0.0) + delta_t

    # Exponential speed decay
    omega_rpm_trip = state.get("omega_rpm_trip", _RPM_R)
    omega_rpm_new  = omega_rpm_trip * np.exp(-t_elapsed_new / tau_s)

    if omega_rpm_new <= 0.5:   # effectively stopped
        return {**state,
                "omega_rpm":       0.0, "omega_rads":      0.0,
                "mass_flow_kgs":   0.0, "vol_flow_m3s":    0.0,
                "velocity_ms":     0.0, "tau_Nm":           0.0,
                "head_Pa":         0.0, "coasted_down":    True,
                "t_coast_elapsed": t_elapsed_new}

    # Two-way coupling: density from current RCS conditions
    rho_bu_v = _rho_bu(pressure_Pa, temperature_K)
    rho_si_v = rho_bu_v * _LBM_FT3_SI
    alpha    = omega_rpm_new / _RPM_R

    op = _operating_point(alpha, state["yh"], rho_bu_v)

    return {
        "omega_rpm":       omega_rpm_new,
        "omega_rads":      omega_rpm_new * 2.0 * np.pi / 60.0,
        "omega_rpm_trip":  omega_rpm_trip,
        "tau_Nm":          op["tau_shaft_lbfft"] * _LBFFT_SI,
        "mass_flow_kgs":   op["mdot_lbms"]       * _LBM_SI,
        "vol_flow_m3s":    op["Q_ft3s"]          * _FT3S_SI,
        "velocity_ms":     op["Q_ft3s"] * _FT3S_SI / _A_SI,
        "head_Pa":         op["head_ft"] * _FT_SI * rho_si_v * 9.81,
        "yh":              state["yh"],
        "coasted_down":    False,
        "t_coast_elapsed": t_elapsed_new,
    }


def htc(state:         dict,
        pressure_Pa:   float,
        temperature_K: float,
        D_h:           float = 0.012,
        L_heat:        float = 1.6) -> float:
    """
    Return the appropriate convection HTC for the heat-structure BC.

    While the pump is running   → Dittus-Boelter forced-convection correlation.
    When the pump has stopped   → returns 0.0 (caller falls back to NatConvHTC).

    Dittus-Boelter (heating case):
        Nu = 0.023 · Re^0.8 · Pr^0.4
        HTC = Nu · k / D_h

    Parameters
    ----------
    state         : current pump_state dict
    pressure_Pa   : RCS pressure  [Pa]
    temperature_K : bulk coolant temperature  [K]
    D_h           : hydraulic diameter  [m]   (default 0.012 m, rod-bundle estimate)
    L_heat        : heated length  [m]         (default 1.6 m)

    Returns
    -------
    htc_val : heat transfer coefficient  [W/m²·K]
              0.0 if pump coasted down or XSteam unavailable.
    """
    velocity = state.get("velocity_ms", 0.0)

    if state.get("coasted_down", True) or velocity < 0.01 or not _XSTEAM:
        return 0.0

    P_kPa = pressure_Pa * 1e-3
    T_C   = temperature_K - 273.15

    try:
        Tsat  = XSteam.Tsat_p(P_kPa)
        T_lkp = min(T_C, 0.9999 * Tsat)        # keep subcooled for properties

        hf_kJ = XSteam.hL_T(T_lkp)             # kJ/kg
        rho   = XSteam.rho_ph(P_kPa, hf_kJ)    # kg/m³
        cp    = XSteam.cp_ph(P_kPa, hf_kJ) * 1e3   # J/(kg·K)
        k     = XSteam.tc_ph(P_kPa, hf_kJ)     # W/(m·K)
        mu    = XSteam.my_ph(P_kPa, hf_kJ)     # Pa·s

        Re = rho * velocity * D_h / mu
        Pr = mu * cp / k

        if Re < 2300.0:
            Nu = 3.66          # laminar, fully developed, constant heat flux
        else:
            Nu = 0.023 * Re**0.8 * Pr**0.4    # Dittus-Boelter

        return Nu * k / D_h

    except Exception:
        return 0.0


# ============================================================================
# STEAM GENERATOR MODEL
# ============================================================================

def sg_heat(pump_state:    dict,
            T_vessel_K:    float,
            T_sec_K:       float,
            UA_rated:      float,
            cp_primary:    float = 5500.0) -> float:
    """
    Lumped UA steam generator model with Dittus-Boelter flow scaling.

    Heat removal from the primary side (enters b[0] as a *negative* term):

        UA_eff  = UA_rated · (v / v_rated)^0.8      [W/K]
        Q_cap   = ṁ_pump · cp · (T_vessel − T_sec)  [W]   (ε = 1 upper bound)
        Q_SG    = min(UA_eff, ṁ_pump·cp) · max(T_vessel − T_sec, 0)

    When the pump has coasted down UA_eff = 0, so Q_SG = 0 (no natural
    circulation — per user specification).

    Parameters
    ----------
    pump_state  : current dict returned by init() or step()
    T_vessel_K  : vessel (primary) bulk temperature  [K]
    T_sec_K     : SG secondary saturation temperature  [K]  (fixed boundary)
    UA_rated    : rated overall heat-transfer conductance  [W/K]
    cp_primary  : primary-side specific heat used for the flow cap  [J/(kg·K)]
                  Default 5500 J/(kg·K) ≈ liquid water at ~300 °C, 15 MPa.
                  Override from XSteam at the call site for higher accuracy.

    Returns
    -------
    Q_SG : heat removed from the primary side  [W]  (always ≥ 0)
           Caller subtracts this from b[0]:  b[0] -= Q_SG
    """
    if pump_state.get("coasted_down", True):
        return 0.0

    # Dimensionless flow ratio v = Q / Q_rated
    # pump_state carries vol_flow_m3s; rated volumetric flow in SI
    v_ratio = pump_state["vol_flow_m3s"] / _Q_R_SI
    v_ratio = max(v_ratio, 0.0)

    # Effective UA with Dittus-Boelter (Re ~ v) scaling on primary side
    UA_eff = UA_rated * v_ratio ** 0.8

    # Primary-side flow heat-capacity rate  [W/K]
    mdot = pump_state["mass_flow_kgs"]
    C_primary = mdot * cp_primary        # ṁ · cp  [W/K]

    # Effective conductance capped at C_primary (ε = 1 limit)
    UA_use = min(UA_eff, C_primary)

    # Temperature difference (floor at zero — SG cannot add heat to primary)
    delta_T = max(T_vessel_K - T_sec_K, 0.0)

    return UA_use * delta_T


def sg_secondary_temperature(P_sec_kPa: float) -> float:
    """
    Return secondary saturation temperature [K] from secondary pressure [kPa].
    Used once at initialisation to set the fixed T_sec boundary condition.

    Parameters
    ----------
    P_sec_kPa : secondary steam pressure  [kPa]

    Returns
    -------
    T_sec_K : saturation temperature  [K]
    """
    if _XSTEAM:
        try:
            return XSteam.Tsat_p(P_sec_kPa) + 273.15
        except Exception:
            pass
    # Fallback: Antoine approximation valid for 100–300 °C
    # ln(P/kPa) ≈ 16.3872 − 3885.70/(T_C + 230.170)  → solve for T_C
    import math
    T_C = 3885.70 / (16.3872 - math.log(P_sec_kPa)) - 230.170
    return T_C + 273.15

# ═══════════════════════════════════════════════════════════════════════════════
# END PUMP MODULE
# ═══════════════════════════════════════════════════════════════════════════════


# global sentinel initialisations
time, tmp2, x_eq, Total_Mass_scaled = None, None, None, None
acc_pres, acc_wdot, acc_liqvol      = None, None, None
Pressure, dp_head, rel_delz         = None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# post_processing  — PATCH 5 / 8c: accepts pump + SG arrays
# ─────────────────────────────────────────────────────────────────────────────
def post_processing(time, Pressure, Temperature, massflow_break,
                    acc_pres, acc_tgas, massflow_in,
                    acc_liqvol, Total_Mass_scaled, ves_ll, net_heat_total,
                    R5_p, R5_t, R5_vessel_mass_scaled, R5_mdot,
                    R5_accp, R5_acct, R5_massflow_in, R5_power,
                    x_eq, wkstbase, title_name_include_line_1="",
                    # PATCH 5b / 8c — optional pump + SG diagnostic arrays
                    pump_omega=None, pump_mdot=None,
                    pump_velocity=None, pump_head=None,
                    steam_vel_arr=None,
                    Q_sg=None,
                    alpha_void_out=None,
                    TTwall_out=None, alpha_out=None,
                    rkpower_total_out=None, core_flag=False,
                    DNBR_out=None, q_chf_out=None, q_hot_out=None,
                    F_r=1.0, F_z=1.0,
                    massflow_PORV_out=None,
                    h_break_out=None,
                    T_fuel_out=None,
                    T_hot_clad_out=None,
                    T_hot_fuel_out=None,
                    rho_ext_out=None, rho_boron_out=None, rho_D_out=None,
                    rcs_boron_ppm_out=None,
                    rho_M_out=None,   rho_net_out=None,   rho_scram_out=None,
                    N_fail_DNB_out=None,
                    N_fail_clad_out=None,
                    N_fail_gap_out=None,
                    N_fail_eiv_out=None,
                    source_term_out=None,
                    pzr_level_out=None,
                    pzr_level_norm_out=None,
                    pzr_mdot_surge_out=None,
                    minimum_output=0,
                    cvcs_mdot_out=None,
                    cvcs_makeup_out=None,
                    cvcs_letdown_out=None,
                    acc_level_out=None,
                    hpsi_mdot_out=None,
                    lpsi_mdot_out=None,
                    si_pumped_out=None,
                    dose_out=None,
                    iodine_spike_dose_out=None,
                    input_echo_out=None,
                    N_pins_out=18200,
                    d_pin_out=0.0095,
                    L_heated_out=2.4,
                    delta_clad_out=0.00057,
                    k_sigma_out=3.0,
                    Pressure_report_out=None):

    # Lazy matplotlib import — deferred from module level to save ~1.4 s on
    # every simulation startup (matplotlib is only needed for plotting here).
    import matplotlib.pyplot as plt

    # Diagnostic: confirm boron array received
    if rho_boron_out is not None:
        _boron_max = float(np.nanmin(rho_boron_out))
        print(f"POST-PROC: rho_boron_out received, peak = {_boron_max:.1f} pcm, "
              f"non-zero = {np.count_nonzero(rho_boron_out)}")
    else:
        print("POST-PROC: rho_boron_out is None")

    filename   = f"{wkstbase}_out.xlsx"
    line_width = 1.5
    colors     = ["m", "b", "g", "r"]
    _zeros     = np.zeros(len(time))
    _Pressure_output = Pressure_report_out if Pressure_report_out is not None else Pressure

    # ── Zircaloy oxidation / hydrogen generation diagnostic ─────────────────
    # Baker-Just parabolic oxidation using oxygen weight-gain W [mg O2/cm²].
    # Maximum local ECR is based on hot-pin clad temperature. Core-wide H2 uses
    # a "mean oxidizing rod" temperature: the conditional mean of the same
    # Gaussian clad-temperature distribution used for rod-failure estimates,
    # truncated over 800 °C < T < T_hot_clad.
    _zirc_ecr_hot_pct      = np.zeros(len(time))
    _zirc_ecr_oxid_pct     = np.zeros(len(time))
    _zirc_h2_kg            = np.zeros(len(time))
    _zirc_h2_full_core_kg  = np.zeros(len(time))
    _zirc_w_hot_mg_cm2     = np.zeros(len(time))
    _zirc_w_oxid_mg_cm2    = np.zeros(len(time))
    _zirc_n_oxid_rods      = np.zeros(len(time))
    _zirc_T_oxid_mean_K    = np.full(len(time), np.nan)

    try:
        _A_BJ = 33.3e6       # mg²/cm⁴/s, oxygen weight-gain basis
        # Baker-Just is commonly written as exp(-45500/T) with T in °R.
        # FLARE uses Kelvin internally, so convert the activation temperature:
        #   T_R = 1.8*T_K  ->  exp(-45500/T_R) = exp(-(45500/1.8)/T_K)
        _B_BJ = 45500.0 / 1.8  # K-equivalent activation temperature
        _T_OX_MIN_K = 800.0 + 273.15

        _rho_zr_g_cm3 = 6.56
        _M_Zr = 91.224
        _M_O2 = 31.998
        _M_H2 = 2.01588

        _clad_thick_cm = max(float(delta_clad_out), 0.0) * 100.0
        _d_pin_cm      = max(float(d_pin_out), 0.0) * 100.0
        _L_heated_cm   = max(float(L_heated_out), 0.0) * 100.0
        _N_pins_ox     = max(float(N_pins_out), 0.0)
        _k_sigma_ox    = max(float(k_sigma_out), 0.1)

        # Oxygen uptake if the full clad wall thickness were converted to ZrO2.
        # W_O_full = Zr areal mass × oxygen/Zr mass ratio.
        _W_full_O_mg_cm2 = (
            _rho_zr_g_cm3 * _clad_thick_cm * (_M_O2 / _M_Zr) * 1000.0
            if _clad_thick_cm > 0 else np.inf
        )
        _area_per_rod_cm2 = np.pi * _d_pin_cm * _L_heated_cm

        # Hypothetical H2 if all cladding in the core fully reacts.
        # Based on full-wall oxygen uptake and stoichiometry:
        #   O2 + 2 Zr -> 2 ZrO2;  O2 mass -> 2 H2 mass.
        _mO2_full_core_kg = (
            _W_full_O_mg_cm2 * _area_per_rod_cm2 * _N_pins_ox * 1.0e-6
            if np.isfinite(_W_full_O_mg_cm2) else 0.0
        )
        _H2_full_core_kg_const = _mO2_full_core_kg * (2.0 * _M_H2 / _M_O2)

        def _phi(_z):
            return np.exp(-0.5 * _z * _z) / np.sqrt(2.0 * np.pi)

        def _Phi(_z):
            return 0.5 * (1.0 + _math.erf(_z / np.sqrt(2.0)))

        def _bj_increment(_W_old, _T_K, _dt_s):
            if not np.isfinite(_T_K) or _T_K <= _T_OX_MIN_K or _dt_s <= 0.0:
                return _W_old
            _rate_w2 = _A_BJ * np.exp(-_B_BJ / _T_K)
            return np.sqrt(max(_W_old*_W_old + _rate_w2 * _dt_s, 0.0))

        _W_hot = 0.0
        _W_oxid = 0.0
        _H2_kg = 0.0

        _T_hot_arr = T_hot_clad_out if T_hot_clad_out is not None else _zeros
        _T_avg_arr = TTwall_out if TTwall_out is not None else _zeros
        _N_gap_arr = N_fail_gap_out if N_fail_gap_out is not None else _zeros

        for _i in range(len(time)):
            _dt_s = 0.0 if _i == 0 else max(float(time[_i] - time[_i-1]), 0.0)
            _T_hot = float(_T_hot_arr[_i]) if np.isfinite(_T_hot_arr[_i]) else np.nan
            _T_avg = float(_T_avg_arr[_i]) if np.isfinite(_T_avg_arr[_i]) else np.nan
            _N_oxid = max(float(_N_gap_arr[_i]), 0.0)
            # Always preserve the oxidizing-rod diagnostic from the existing
            # T(clad)>800°C failed-rod estimate, even if the downstream
            # Baker-Just/truncated-Gaussian calculation encounters a problem.
            _zirc_n_oxid_rods[_i] = _N_oxid

            # Hot-pin maximum local oxidation/ECR
            _W_hot_new = min(_bj_increment(_W_hot, _T_hot, _dt_s), _W_full_O_mg_cm2)

            # Mean oxidizing rod temperature for rods above 800 °C.
            if _N_oxid > 0.0 and np.isfinite(_T_hot) and np.isfinite(_T_avg) and _T_hot > _T_OX_MIN_K:
                _sigma_T = max((_T_hot - _T_avg) / _k_sigma_ox, 0.1)
                _a = (_T_OX_MIN_K - _T_avg) / _sigma_T
                _b = (_T_hot      - _T_avg) / _sigma_T
                _den = max(_Phi(_b) - _Phi(_a), 1e-15)
                _T_oxid_mean = _T_avg + _sigma_T * ((_phi(_a) - _phi(_b)) / _den)
                # The truncated-normal formula is mathematically bounded, but
                # roundoff protection keeps the diagnostic robust.
                _T_oxid_mean = min(
                    min(max(_T_oxid_mean, _T_OX_MIN_K), _T_hot),
                    2500.0 + 273.15
                )

                _W_oxid_new = min(_bj_increment(_W_oxid, _T_oxid_mean, _dt_s), _W_full_O_mg_cm2)

                # Incremental O2 uptake for the oxidizing population.
                _dW_oxid = max(_W_oxid_new - _W_oxid, 0.0)  # mg O2/cm²
                _mO2_kg = _dW_oxid * _area_per_rod_cm2 * _N_oxid * 1.0e-6
                _H2_kg += _mO2_kg * (2.0 * _M_H2 / _M_O2)
                _H2_kg = min(_H2_kg, _H2_full_core_kg_const)

                _W_oxid = _W_oxid_new
                _zirc_T_oxid_mean_K[_i] = _T_oxid_mean
            else:
                _T_oxid_mean = np.nan

            _W_hot = _W_hot_new

            _zirc_w_hot_mg_cm2[_i]  = _W_hot
            _zirc_w_oxid_mg_cm2[_i] = _W_oxid
            _zirc_n_oxid_rods[_i]   = _zirc_n_oxid_rods[_i]
            _zirc_h2_kg[_i]         = min(_H2_kg, _H2_full_core_kg_const)
            _zirc_h2_full_core_kg[_i] = _H2_full_core_kg_const

            if np.isfinite(_W_full_O_mg_cm2) and _W_full_O_mg_cm2 > 0:
                _zirc_ecr_hot_pct[_i]  = min(100.0, 100.0 * _W_hot / _W_full_O_mg_cm2)
                _zirc_ecr_oxid_pct[_i] = min(100.0, 100.0 * _W_oxid / _W_full_O_mg_cm2)

    except Exception as _ox_err:
        print(f"WARNING: Zr oxidation/H2 diagnostic failed; output set to zero: {_ox_err}")
        _zirc_ecr_hot_pct[:]      = 0.0
        _zirc_ecr_oxid_pct[:]     = 0.0
        _zirc_h2_kg[:]            = 0.0
        _zirc_h2_full_core_kg[:]  = 0.0
        _zirc_n_oxid_rods[:]      = 0.0
        _zirc_T_oxid_mean_K[:] = np.nan

    data = {
        "Time (s)":                    time,
        "RCS Pressure (kPa)":          _Pressure_output / 1e3,
        "RCS Temperature (K)":         Temperature,
        "Break Flow (kg/s)":           massflow_break,
        "Break Enthalpy (kJ/kg)":      h_break_out if h_break_out is not None else _zeros,
        "Accumulator Pressure (kPa)":  acc_pres / 1e3,
        "Accumulator Temperature (K)": acc_tgas,
        "Accumulator Flow (kg/s)":     massflow_in,
        "CVCS Flow (kg/s)":        cvcs_mdot_out     if cvcs_mdot_out     is not None else _zeros,
        "CVCS Makeup (kg/s)":      cvcs_makeup_out   if cvcs_makeup_out   is not None else _zeros,
        "CVCS Letdown (kg/s)":     cvcs_letdown_out  if cvcs_letdown_out  is not None else _zeros,
        "HPSI Flow (kg/s)":            hpsi_mdot_out   if hpsi_mdot_out   is not None else _zeros,
        "LPSI Flow (kg/s)":            lpsi_mdot_out   if lpsi_mdot_out   is not None else _zeros,
        "SI Pumped Total (kg/s)":      si_pumped_out   if si_pumped_out   is not None else _zeros,
        "Accumulator Liquid Volume (m3)": acc_liqvol,
        "Accumulator Level (m)":          acc_level_out if acc_level_out is not None else _zeros,
        "Total Mass Scaled":           Total_Mass_scaled,
        "Vessel Level (m)":            ves_ll,
        "Core Power (MW)":             net_heat_total / 1e6,
        # PATCH 5b/8c additions
        # Steam Cooling Velocity: upward steam velocity in the core channel
        # driven by decay-heat boil-off.  v = (Q_decay/h_fg) / (rhoV * A_flow_core).
        # Computed in the time loop only while film boiling is active; zero otherwise.
        # This is the velocity actually used in core_htc_db_steam post-CHF.

        "Pump Speed (rpm)":            pump_omega    if pump_omega    is not None else _zeros,
        "Pump Mass Flow (kg/s)":       pump_mdot     if pump_mdot     is not None else _zeros,
        "Core Coolant Velocity (m/s)": pump_velocity if pump_velocity is not None else _zeros,
        "Pump Velocity (m/s)":         pump_velocity if pump_velocity is not None else _zeros,
        "Steam Cooling Velocity (m/s)": (
            steam_vel_arr if steam_vel_arr is not None else _zeros
        ),
        "Pump Head (Pa)":              pump_head     if pump_head     is not None else _zeros,
        "SG Heat Removal (MW)":        (Q_sg         if Q_sg         is not None else _zeros) / 1e6,
        "Equilibrium Quality (-)":     x_eq                if x_eq             is not None else _zeros,
        "Void Fraction (-)":           alpha_void_out      if alpha_void_out    is not None else _zeros,
        "Clad Surface Temp (K)":       TTwall_out          if TTwall_out       is not None else _zeros,
        "Clad HTC (W/m2-K)":          alpha_out           if alpha_out        is not None else _zeros,
        "RK Total Power (MW)":         (rkpower_total_out  if rkpower_total_out is not None else _zeros) / 1e6,
        "DNBR":                        DNBR_out   if DNBR_out   is not None else _zeros,
        "CHF Heat Flux (MW/m2)":       (q_chf_out if q_chf_out  is not None else _zeros) / 1e6,
        "Hot Pin Heat Flux (MW/m2)":   (q_hot_out if q_hot_out  is not None else _zeros) / 1e6,
        "PORV Mass Flow (kg/s)":        massflow_PORV_out if massflow_PORV_out is not None else _zeros,
        "Pressurizer Level (m)":        pzr_level_out      if pzr_level_out      is not None else _zeros,
        "Pressurizer Level (norm)":     pzr_level_norm_out if pzr_level_norm_out is not None else _zeros,
        "Pressurizer Surge (kg/s)":     pzr_mdot_surge_out if pzr_mdot_surge_out is not None else _zeros,
        "Fuel Avg Temp (K)":            T_fuel_out      if T_fuel_out      is not None else _zeros,
        "Hot Pin Clad Temp (K)":        T_hot_clad_out  if T_hot_clad_out  is not None else _zeros,
        "Hot Pin Fuel Temp (K)":        T_hot_fuel_out  if T_hot_fuel_out  is not None else _zeros,
        "Reactivity scram (pcm)":       rho_scram_out if rho_scram_out is not None else _zeros,
        "Reactivity Boron (pcm)":       rho_boron_out      if rho_boron_out      is not None else _zeros,
        "RCS Boron (ppm)":              rcs_boron_ppm_out  if rcs_boron_ppm_out  is not None else _zeros,
        "Reactivity Doppler (pcm)":     rho_D_out    if rho_D_out   is not None else _zeros,
        "Reactivity Moderator (pcm)":   rho_M_out    if rho_M_out   is not None else _zeros,
        "Reactivity net (pcm)":         rho_net_out  if rho_net_out is not None else _zeros,
        "Rod Failures DNB (est.)": N_fail_DNB_out if N_fail_DNB_out is not None else _zeros,
        "Rod Failures PCT (est.)": N_fail_clad_out if N_fail_clad_out is not None else _zeros,
        "Rod Failures Gap (est.)": N_fail_gap_out if N_fail_gap_out is not None else _zeros,
        "Rod Failures EarlyIV (est.)": N_fail_eiv_out if N_fail_eiv_out is not None else _zeros,
        "Zr Oxidation Hot Pin ECR (%)": _zirc_ecr_hot_pct,
        "Zr Oxidation Mean Oxidizing Rod ECR (%)": _zirc_ecr_oxid_pct,
        "Zr Oxidizing Rods (est.)": _zirc_n_oxid_rods,
        "Zr Mean Oxidizing Rod Temp (K)": _zirc_T_oxid_mean_K,
        "Zr Hot Pin O2 Uptake (mg/cm2)": _zirc_w_hot_mg_cm2,
        "Zr Mean Oxidizing Rod O2 Uptake (mg/cm2)": _zirc_w_oxid_mg_cm2,
        "H2 Generated (kg)": _zirc_h2_kg,
        "H2 Full Core Cladding Reaction (kg)": _zirc_h2_full_core_kg,
    }
    df = pd.DataFrame(data)

    # Write the time-series CSV immediately — before Excel writing and all
    # matplotlib work — so a crash in either of those paths still leaves a
    # readable output CSV on disk.
    csv_filename = filename.replace("_out.xlsx", "_out.csv")
    try:
        df.to_csv(csv_filename, index=False)
        print(f"Output CSV written: {csv_filename}", flush=True)
    except Exception as _csv_err:
        print(f"WARNING: could not write output CSV: {_csv_err}", flush=True)

    correl_df = pd.DataFrame([], columns=["Correlation", "TS 95% Coverage"])
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        correl_df.to_excel(writer, sheet_name="Sheet2")
        if source_term_out is not None:
            source_term_out.to_excel(writer, sheet_name="Source Term",
                                     index=False)
        if dose_out is not None:
            # Write NOTBADTRAD results to a Dose sheet
            _dose_ws = writer.book.create_sheet("Dose")
            _dose_ws.append(["NOTBADTRAD Radiological Screening (RG 1.183 / 10 CFR 50.67)"])
            _dose_ws.append([])
            # Summary table
            _dose_ws.append(["Location", "TEDE (rem)", "Limit (rem)", "Margin (rem)", "Result"])
            for _row in dose_out["summary"].to_dict("records"):
                _dose_ws.append([_row["Location"], round(_row["TEDE (rem)"], 12),
                                  _row["Limit (rem)"], round(_row["Margin"], 12),
                                  _row["Result"]])
            _dose_ws.append([])
            # Group contributions
            _dose_ws.append(["Group", "Released (Ci)", "EAB (rem)", "LPZ (rem)", "CR (rem)", "Release frac"])
            for _row in dose_out["groups"].to_dict("records"):
                _dose_ws.append([_row["Group"],
                                  _row["Released (Ci)"],
                                  round(_row["EAB (rem)"], 14),
                                  round(_row["LPZ (rem)"], 14),
                                  round(_row["CR (rem)"], 14),
                                  round(_row["Release frac"], 8)])
            _dose_ws.append([])
            # chi/Q and key inputs
            _xq = dose_out.get("chi_q", {})
            _dose_ws.append(["chi/Q EAB (s/m3)", _xq.get("eab", "")])
            _dose_ws.append(["chi/Q LPZ (s/m3)", _xq.get("lpz", "")])
            _dose_ws.append(["chi/Q CR  (s/m3)", _xq.get("cr",  "")])
            _dose_ws.append(["Runtime (ms)", round(dose_out.get("runtime_ms", 0), 2)])
            _inv_meta = dose_out.get("inputs", {}).get("inventory", {})
            if _inv_meta:
                _dose_ws.append(["Inventory model", _inv_meta.get("model", "")])
                _dose_ws.append(["Estimated fissions", _inv_meta.get("n_fissions", "")])
            # Iodine spike summary
            _isp = dose_out.get("iodine_spike", {})
            if _isp:
                _dose_ws.append([])
                _dose_ws.append(["--- Iodine Spike (Pre-existing Coolant Activity) ---"])
                _dose_ws.append(["Coolant activity (uCi/g)",  _isp.get("coolant_act_uci_g", "")])
                _dose_ws.append(["Primary mass (kg)",          round(_isp.get("primary_mass_kg", 0), 1)])
                _dose_ws.append(["Equilibrium spike (Ci) 500x", round(_isp.get("equil_spike_ci", 0), 2)])
                _dose_ws.append(["Accident spike (Ci) 335x",    round(_isp.get("accid_spike_ci", 0), 2)])
                _dose_ws.append(["Equilibrium frac (500x)",     round(_isp.get("f_eq_spike", 0), 8)])
                _dose_ws.append(["Accident frac (335x)",        round(_isp.get("f_accid_spike", 0), 8)])
                _dose_ws.append(["Combined halogen frac",       round(_isp.get("f_combined", 0), 8)])
                _dose_ws.append(["Source term model",           _isp.get("model_used", "")])
                _dose_ws.append(["Is LOCA",                     str(_isp.get("is_loca", False))])
                _dose_ws.append(["Scram occurred",              str(_isp.get("scram", False))])
                _dose_ws.append(["PORV opened",                 str(_isp.get("porv", False))])
                _dose_ws.append(["DNB occurred",                str(_isp.get("dnb", False))])
                _dose_ws.append(["Fuel damage",                 str(_isp.get("fuel_damage", False))])
                _dose_ws.append(["N rods DNB/dryout",           _isp.get("N_dnb", 0)])
            # Distance fit and table
            _fit = dose_out.get("dist_fit")
            if _fit:
                _dose_ws.append([])
                _dose_ws.append(["Power-law fit", _fit.get("formula", "")])
                _dose_ws.append(["Fit error (%)", round(_fit.get("fit_error_pct", 0), 6)])
                _dose_ws.append([])
                _dose_ws.append(["Distance (m)",
                                  "chi/Q (s/m3)",
                                  "TEDE EAB interval (rem)",
                                  "TEDE LPZ interval (rem)"])
                _dt = dose_out.get("dist_table")
                if _dt is not None and not _dt.empty:
                    for _, row in _dt.iterrows():
                        _dose_ws.append([int(row["distance_m"]),
                                          row.get("chi_q_s_m3", ""),
                                          row.get("eab_integration_tede_rem", row.get("tede_rem", 0.0)),
                                          row.get("lpz_integration_tede_rem", row.get("tede_rem", 0.0))])

        # ── Iodine Spike sheet (separate from Dose sheet) ────────────────────
        _isp_out = iodine_spike_dose_out
        if _isp_out is not None:
            _isp_ws = writer.book.create_sheet("Iodine Spike")
            _isp = _isp_out.get("iodine_spike", {})
            _isp_ws.append(["Iodine Spike Dose — Pre-existing Coolant Activity"])
            _isp_ws.append([])
            _isp_ws.append(["Model",                        _isp.get("model_used", "")])
            _isp_ws.append(["Coolant activity (uCi/g)",     _isp.get("coolant_act_uci_g", "")])
            _isp_ws.append(["Primary mass (kg)",             round(_isp.get("primary_mass_kg", 0), 1)])
            _isp_ws.append(["Equilibrium spike (Ci) 500x",  round(_isp.get("equil_spike_ci", 0), 2)])
            _isp_ws.append(["Accident spike (Ci) 335x",     round(_isp.get("accid_spike_ci", 0), 2)])
            _isp_ws.append(["Equilibrium frac (500x)",      round(_isp.get("f_eq_spike", 0), 8)])
            _isp_ws.append(["Accident frac (335x)",         round(_isp.get("f_accid_spike", 0), 8)])
            _isp_ws.append(["Combined halogen frac",        round(_isp.get("f_combined", 0), 8)])
            _isp_ws.append(["Is LOCA",                      str(_isp.get("is_loca", False))])
            _isp_ws.append(["Scram occurred",               str(_isp.get("scram", False))])
            _isp_ws.append(["PORV opened",                  str(_isp.get("porv", False))])
            _isp_ws.append(["DNB occurred",                 str(_isp.get("dnb", False))])
            _isp_ws.append(["Fuel damage",                  str(_isp.get("fuel_damage", False))])
            _isp_ws.append(["N rods DNB/dryout",            _isp.get("N_dnb", 0)])
            _isp_ws.append([])
            # TEDE summary
            _isp_ws.append(["Location", "TEDE (rem)", "Limit (rem)", "Margin (rem)", "Result"])
            if "summary" in _isp_out:
                for _row in _isp_out["summary"].to_dict("records"):
                    _isp_ws.append([_row["Location"], _row["TEDE (rem)"],
                                     _row["Limit (rem)"], _row["Margin"], _row["Result"]])
            _isp_ws.append([])
            # Group contributions
            _isp_ws.append(["Group", "Released (Ci)", "EAB (rem)", "LPZ (rem)", "CR (rem)", "Release frac"])
            if "groups" in _isp_out and not _isp_out["groups"].empty:
                for _row in _isp_out["groups"].to_dict("records"):
                    _isp_ws.append([_row["Group"], _row["Released (Ci)"],
                                     _row["EAB (rem)"], _row["LPZ (rem)"],
                                     _row["CR (rem)"], _row["Release frac"]])
            _xq = _isp_out.get("chi_q", {})
            _isp_ws.append([])
            _isp_ws.append(["chi/Q EAB (s/m3)", _xq.get("eab", "")])
            _isp_ws.append(["chi/Q LPZ (s/m3)", _xq.get("lpz", "")])
            _isp_ws.append(["chi/Q CR  (s/m3)", _xq.get("cr",  "")])
            _isp_ws.append(["Runtime (ms)", round(_isp_out.get("runtime_ms", 0), 2)])
            _fit = _isp_out.get("dist_fit")
            if _fit:
                _isp_ws.append([])
                _isp_ws.append(["Power-law fit", _fit.get("formula", "")])
                _isp_ws.append(["Fit error (%)", round(_fit.get("fit_error_pct", 0), 6)])
                _isp_ws.append([])
                _isp_ws.append(["Distance (m)",
                                  "chi/Q (s/m3)",
                                  "TEDE EAB interval (rem)",
                                  "TEDE LPZ interval (rem)"])
                _dt = _isp_out.get("dist_table")
                if _dt is not None and not _dt.empty:
                    for _, row in _dt.iterrows():
                        _isp_ws.append([int(row["distance_m"]),
                                         row.get("chi_q_s_m3", ""),
                                         row.get("eab_integration_tede_rem", row.get("tede_rem", 0.0)),
                                         row.get("lpz_integration_tede_rem", row.get("tede_rem", 0.0))])

    plt.close('all')

    # ── Input Echo sheet ──────────────────────────────────────────────────────
    # Write all simulation parameters — user-supplied and defaults — to a
    # dedicated sheet so the output is fully self-documenting.
    if input_echo_out:
        try:
            import openpyxl as _opxl
            _wb = _opxl.load_workbook(filename)
            _ews = _wb.create_sheet("Input Echo")
            # Header row
            _hdr_font = _opxl.styles.Font(bold=True)
            _hdr_fill = _opxl.styles.PatternFill("solid", fgColor="1F3864")
            _hdr_color = _opxl.styles.Font(bold=True, color="FFFFFF")
            for col, hdr in enumerate(["Variable", "Value", "Default", "Source"], 1):
                cell = _ews.cell(row=1, column=col, value=hdr)
                cell.font = _hdr_color
                cell.fill = _hdr_fill
            # Data rows — deduplicate keeping last value (latest assignment wins)
            _seen = {}
            for entry in input_echo_out:
                _seen[entry["Variable"]] = entry
            _user_fill    = _opxl.styles.PatternFill("solid", fgColor="E8F4FD")
            _default_fill = _opxl.styles.PatternFill("solid", fgColor="F5F5F5")
            for row_idx, entry in enumerate(
                    sorted(_seen.values(), key=lambda x: x["Variable"]), 2):
                _fill = _user_fill if entry["Source"] == "user" else _default_fill
                for col, key in enumerate(["Variable", "Value", "Default", "Source"], 1):
                    cell = _ews.cell(row=row_idx, column=col, value=str(entry[key]))
                    cell.fill = _fill
            # Column widths
            _ews.column_dimensions["A"].width = 32
            _ews.column_dimensions["B"].width = 20
            _ews.column_dimensions["C"].width = 20
            _ews.column_dimensions["D"].width = 10
            _wb.save(filename)
        except Exception as _ee:
            print(f"Warning: could not write Input Echo sheet: {_ee}")

    # Skip all figure generation when minimum_output = 1 (UA / Risk batch mode)
    if not minimum_output:
        # Remove any stale figure PNGs from previous runs so the PDF
        # only contains figures generated by this case.
        for _old_png in Path(__file__).parent.glob("figure_*.png"):
            _old_png.unlink(missing_ok=True)

        plt.figure(1)
        plt.plot(time, _Pressure_output/1e3, color=colors[0], linewidth=line_width, label="Pressure")
        plt.xlabel("Time [s]"); plt.ylabel("Pressure [kPa]")
        plt.title(f"RCS Pressure: {title_name_include_line_1}"); plt.legend(); plt.grid()
        plt.savefig("figure_1.png")

        plt.figure(2)
        plt.plot(time, x_eq,        color=colors[0], linewidth=line_width, label='Equilibrium Quality')
        if alpha_void_out is not None and np.any(alpha_void_out > 0):
            plt.plot(time, alpha_void_out, color=colors[1], linewidth=line_width,
                     linestyle='--', label='Void Fraction')
        plt.xlabel("Time [s]"); plt.ylabel("Quality / Void Fraction [-]")
        plt.title(f"X_eq & Void Fraction: {title_name_include_line_1}")
        plt.legend(); plt.grid()
        plt.savefig("figure_2.png")

        plt.figure(3)
        plt.plot(time, Total_Mass_scaled,    color=colors[0], linewidth=line_width, label="Scaled Mass")
        plt.xlabel("Time [s]"); plt.ylabel("Mass (normalized)")
        plt.title(f"RCS Mass: {title_name_include_line_1}"); plt.legend(); plt.grid()
        plt.savefig("figure_3.png")

        plt.figure(4)
        plt.plot(time, ves_ll, color=colors[0], linewidth=line_width, label='Vessel Level')
        plt.xlabel('Time [s]'); plt.ylabel('Vessel Level [m]')
        plt.title(f'Vessel Level: {title_name_include_line_1}')
        plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
        plt.savefig("figure_4.png")

        plt.figure(5)
        plt.plot(time, massflow_break, color=colors[0], linewidth=line_width, label='Break flow')
        if massflow_PORV_out is not None and np.any(massflow_PORV_out > 0):
            plt.plot(time, massflow_PORV_out, color=colors[3], linewidth=line_width,
                     linestyle='--', label='PORV flow')
        plt.xlabel('Time [s]'); plt.ylabel('Mass flow [kg/s]')
        plt.title('Break / PORV Flow Rate'); plt.xlim(0, max(time)); plt.legend(); plt.grid()
        plt.savefig("figure_5.png")

        plt.figure(6)
        plt.plot(time, acc_pres,  color=colors[0], linewidth=line_width, label="Accumulator Pressure")
        plt.xlabel("Time [s]"); plt.ylabel("Pressure [kPa]")
        plt.title(f"Accumulator Pressure: {title_name_include_line_1}"); plt.legend(); plt.grid()
        plt.savefig("figure_6.png")

        plt.figure(7)
        plt.plot(time, acc_tgas, color=colors[0], linewidth=line_width, label='Simulation')
        plt.xlabel('Time [s]'); plt.ylabel('Temperature [K]')
        plt.title('Accumulator Temperature: ' + title_name_include_line_1)
        plt.xlim(0, np.max(time)); plt.legend(); plt.grid(True)
        plt.savefig("figure_7.png")

        plt.figure(8)
        plt.plot(time, massflow_in, color=colors[0], linewidth=line_width, label='Accumulator Flow')
        if cvcs_makeup_out is not None and np.any(cvcs_makeup_out > 0):
            plt.plot(time, cvcs_makeup_out,   color=colors[2], linewidth=line_width,
                     label='CVCS Makeup')
        if cvcs_letdown_out is not None and np.any(cvcs_letdown_out > 0):
            plt.plot(time, -cvcs_letdown_out, color=colors[2], linewidth=line_width,
                     linestyle=':', label='CVCS Letdown')
        if hpsi_mdot_out is not None and np.any(hpsi_mdot_out > 0):
            plt.plot(time, hpsi_mdot_out, color='darkorange', linewidth=line_width, label='HPSI Flow')
        if lpsi_mdot_out is not None and np.any(lpsi_mdot_out > 0):
            plt.plot(time, lpsi_mdot_out, color='purple', linewidth=line_width, label='LPSI Flow')
        _cvcs_net   = ((cvcs_makeup_out - cvcs_letdown_out)
                       if (cvcs_makeup_out is not None and cvcs_letdown_out is not None)
                       else (cvcs_makeup_out if cvcs_makeup_out is not None else 0))
        _total_eccs = (massflow_in
                       + _cvcs_net
                       + (si_pumped_out if si_pumped_out is not None else 0))
        if np.any(_total_eccs > massflow_in + 1e-3):
            plt.plot(time, _total_eccs, color=colors[3], linewidth=line_width,
                     linestyle='--', label='Total ECCS Flow')
        plt.xlabel('Time [s]'); plt.ylabel(r'$\dot{m}\ [kg/s]$')
        plt.title('Accumulator & SI / CVCS Flow: ' + title_name_include_line_1)
        plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
        plt.savefig("figure_8.png")

        plt.figure(9)
        if core_flag and rkpower_total_out is not None:
            # With core model: show heat SOURCE (fission+decay) and heat DELIVERED to coolant
            plt.plot(time, rkpower_total_out*1e-6, color=colors[3], linewidth=line_width,
                     label='Fission+Decay Heat Source')
            plt.plot(time, net_heat_total*1e-6,    color=colors[0], linewidth=line_width,
                     label='Convected to Coolant')
        else:
            plt.plot(time, net_heat_total*1e-6, color=colors[0], linewidth=line_width, label='Simulation')
        plt.xlabel('Time [s]'); plt.ylabel('Power [MW]')
        plt.title('Core Power: ' + title_name_include_line_1)
        plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
        plt.savefig("figure_9.png")

        # PATCH 5b / 8c — pump and SG figures
        _pump_omega_plot = pump_omega if pump_omega is not None else _zeros
        _pump_vel_plot   = pump_velocity if pump_velocity is not None else _zeros
        _Q_sg_plot       = Q_sg if Q_sg is not None else _zeros

        plt.figure(10)
        plt.plot(time, _pump_omega_plot, color=colors[0], linewidth=line_width, label="RCP Speed")
        plt.xlabel("Time [s]"); plt.ylabel("Speed [rpm]")
        plt.title("RCP Coastdown: " + title_name_include_line_1); plt.grid(); plt.legend()
        plt.savefig("figure_10.png")

        plt.figure(11)
        plt.plot(time, _pump_vel_plot, color=colors[1], linewidth=line_width, label="Coolant Velocity")
        plt.xlabel("Time [s]"); plt.ylabel("Velocity [m/s]")
        plt.title("RCP Coolant Velocity: " + title_name_include_line_1); plt.grid(); plt.legend()
        plt.savefig("figure_11.png")

        plt.figure(12)
        plt.plot(time, _Q_sg_plot/1e6, color=colors[3], linewidth=line_width, label="SG Removal")
        plt.xlabel("Time [s]"); plt.ylabel("Heat Removal [MW]")
        plt.title("Steam Generator Heat Removal: " + title_name_include_line_1); plt.grid(); plt.legend()
        plt.savefig("figure_12.png")

        # Core heat transfer figures — only generated when core_flag=1
        if core_flag:
            _T_cool_C  = Temperature - 273.15
            _T_clad_C  = TTwall_out - 273.15 if TTwall_out is not None else _zeros
            _htc_plot  = alpha_out           if alpha_out  is not None else _zeros
            _rkpow_MW  = rkpower_total_out*1e-6 if rkpower_total_out is not None else _zeros
            _netpow_MW = net_heat_total*1e-6

            plt.figure(13)
            plt.plot(time, _T_clad_C, color=colors[3], linewidth=line_width, label='Clad surface')
            plt.plot(time, _T_cool_C, color=colors[0], linewidth=line_width, label='Coolant bulk')
            if T_fuel_out is not None and np.any(~np.isnan(T_fuel_out)):
                _T_fuel_C = T_fuel_out - 273.15
                plt.plot(time, _T_fuel_C, color=colors[2], linewidth=line_width,
                         linestyle='--', label='Fuel avg centreline')
            if T_hot_clad_out is not None and np.any(~np.isnan(T_hot_clad_out)):
                _T_hc_C = T_hot_clad_out - 273.15
                plt.plot(time, _T_hc_C, color=colors[1], linewidth=line_width,
                         linestyle=':', label='Hot pin clad')
            if T_hot_fuel_out is not None and np.any(~np.isnan(T_hot_fuel_out)):
                _T_hf_C = T_hot_fuel_out - 273.15
                plt.plot(time, _T_hf_C, color=colors[3], linewidth=line_width,
                         linestyle=':', label='Hot pin fuel centreline')
            plt.xlabel('Time [s]'); plt.ylabel('Temperature [°C]')
            plt.title('Core Temperature: ' + title_name_include_line_1)
            plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
            plt.savefig("figure_13.png")

            plt.figure(14)
            ax1 = plt.gca()
            ax2 = ax1.twinx()
            ax1.plot(time, _htc_plot,  color=colors[0], linewidth=line_width, label='HTC (left)')
            ax2.plot(time, _rkpow_MW,  color=colors[3], linewidth=line_width,
                     linestyle='--', label='Heat source MW (right)')
            ax2.plot(time, _netpow_MW, color=colors[2], linewidth=line_width,
                     linestyle='--', label='Convected MW (right)')
            ax1.set_xlabel('Time [s]'); ax1.set_ylabel('HTC [W/m²·K]', color=colors[0])
            ax2.set_ylabel('Power [MW]', color=colors[3])
            plt.title('Core HTC and Power: ' + title_name_include_line_1)
            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1+lines2, labels1+labels2, loc='upper right')
            plt.grid(True); plt.savefig("figure_14.png")

        # ── Figure 15: DNBR vs time ──────────────────────────────────────────────
        if DNBR_out is not None and not np.all(np.isnan(DNBR_out)):
            _dnbr_valid = np.where(np.isfinite(DNBR_out), DNBR_out, np.nan)
            _min_dnbr   = float(np.nanmin(_dnbr_valid))
            _t_min      = time[np.nanargmin(_dnbr_valid)]
            plt.figure(15)
            plt.plot(time, _dnbr_valid, color=colors[0], linewidth=line_width,
                     label=f"DNBR  (F_r={F_r:.2f}, F_z={F_z:.2f})")
            plt.axhline(y=1.0, color="red", linewidth=1.0, linestyle="--",
                        label="DNBR = 1.0  (CHF limit)")
            plt.axhline(y=1.3, color="orange", linewidth=0.8, linestyle=":",
                        label="DNBR = 1.3  (typical design limit)")
            plt.xlabel("Time [s]"); plt.ylabel("DNBR [-]")
            plt.title("DNBR (Biasi/Zuber/Bowring): " + title_name_include_line_1)
            plt.legend(loc="upper right"); plt.grid(True)
            plt.annotate(f"Min DNBR = {_min_dnbr:.2f} at t = {_t_min:.1f} s",
                         xy=(_t_min, _min_dnbr),
                         xytext=(_t_min + 0.02*(time[-1]-time[0]), _min_dnbr + 0.5),
                         arrowprops=dict(arrowstyle="->", color="black"), fontsize=8)
            plt.savefig("figure_15.png")
            print(f"DNBR:  min = {_min_dnbr:.3f}  at t = {_t_min:.1f} s"
                  f"  (Biasi/Zuber/Bowring blend, F_r={F_r:.2f}, F_z={F_z:.2f})")

        # ── Figure 16: reactivity vs time ────────────────────────────────────────
        _has_rho = any([
            rho_net_out   is not None and np.nanmax(np.abs(rho_net_out))   > 0,
            rho_ext_out   is not None and np.nanmax(np.abs(rho_ext_out))   > 0,
            rho_scram_out is not None and np.nanmax(np.abs(rho_scram_out)) > 0,
            rho_boron_out is not None and np.nanmax(np.abs(rho_boron_out)) > 0,
            rho_D_out     is not None and np.nanmax(np.abs(rho_D_out))     > 0,
            rho_M_out     is not None and np.nanmax(np.abs(rho_M_out))     > 0,
        ])
        if _has_rho:
            plt.figure(16)
            plt.clf()
            if rho_scram_out is not None and np.nanmax(np.abs(rho_scram_out)) > 0:
                plt.plot(time, rho_scram_out, color=colors[1], linewidth=line_width,
                         label='ρ_scram')
            if rho_ext_out is not None and np.nanmax(np.abs(rho_ext_out)) > 0:
                plt.plot(time, rho_ext_out, color='grey', linewidth=line_width,
                         linestyle='--', label='ρ_ext (rod withdrawal)')
            if rho_boron_out is not None and np.nanmax(np.abs(rho_boron_out)) > 0:
                plt.plot(time, rho_boron_out, color='purple', linewidth=line_width,
                         linestyle='-.', label='ρ_boron (SLCS)')
            if rho_D_out is not None and np.nanmax(np.abs(rho_D_out)) > 0:
                plt.plot(time, rho_D_out, color=colors[3], linewidth=line_width,
                         linestyle='--', label='ρ_Doppler')
            if rho_M_out is not None and np.nanmax(np.abs(rho_M_out)) > 0:
                plt.plot(time, rho_M_out, color=colors[0], linewidth=line_width,
                         linestyle=':', label='ρ_moderator')
            if rho_net_out is not None:
                plt.plot(time, rho_net_out, color='k', linewidth=line_width+0.5,
                         label='ρ_net')
            plt.axhline(y=0, color='grey', linewidth=0.8, linestyle='-')
            plt.xlabel('Time [s]'); plt.ylabel('Reactivity [pcm]')
            plt.title('Reactivity: ' + title_name_include_line_1)
            plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
            plt.savefig("figure_16.png")

        # ── Fig 17: Accumulator liquid level ─────────────────────────────────
        plt.figure(17)
        _acc_level = acc_level_out if acc_level_out is not None else np.zeros_like(acc_liqvol)
        plt.plot(time, _acc_level, color=colors[0], linewidth=line_width)
        plt.xlabel("Time [s]"); plt.ylabel("Accumulator Level [m]")
        plt.title(f"Accumulator Level: {title_name_include_line_1}")
        plt.xlim(0, np.max(time)); plt.grid(True)
        plt.savefig("figure_17.png")

        # ── Zr oxidation / H2 diagnostic figures — only when non-zero ───────
        if np.nanmax(_zirc_ecr_hot_pct) > 0 or np.nanmax(_zirc_h2_kg) > 0:
            plt.figure(18)
            plt.plot(time, _zirc_ecr_hot_pct, color=colors[3], linewidth=line_width,
                     label="Hot-pin ECR")
            plt.plot(time, _zirc_ecr_oxid_pct, color=colors[0], linewidth=line_width,
                     linestyle="--", label="Mean oxidizing rod ECR")
            plt.axhline(y=17.0, color="red", linewidth=1.0, linestyle=":",
                        label="17% ECR criterion")
            plt.xlabel("Time [s]"); plt.ylabel("ECR [%]")
            plt.title("Zircaloy Oxidation / ECR: " + title_name_include_line_1)
            plt.xlim(0, np.max(time)); plt.grid(True); plt.legend()
            plt.savefig("figure_18.png")

            plt.figure(19)
            ax1 = plt.gca()
            ax2 = ax1.twinx()
            ax1.plot(time, _zirc_h2_kg, color=colors[1], linewidth=line_width,
                     label="H2 generated")
            ax1.plot(time, _zirc_h2_full_core_kg, color=colors[3], linewidth=line_width,
                     linestyle=":", label="Full-core clad reaction H2")
            ax2.plot(time, _zirc_n_oxid_rods, color=colors[0], linewidth=line_width,
                     linestyle="--", label="Oxidizing rods")
            ax1.set_xlabel("Time [s]")
            ax1.set_ylabel("H2 generated [kg]", color=colors[1])
            ax2.set_ylabel("Oxidizing rods [count]", color=colors[0])
            plt.title("Hydrogen Generation: " + title_name_include_line_1)
            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")
            ax1.grid(True)
            plt.savefig("figure_19.png")


        # ── Pressurizer level — always shown when pzr_vol > 0 ────────────────
        if pzr_level_out is not None:
            fig_pzr, ax1 = plt.subplots(figsize=(8, 4))
            ax2 = ax1.twinx()
            valid = ~np.isnan(pzr_level_out)
            ax1.plot(time[valid], pzr_level_out[valid],
                     color="steelblue", linewidth=line_width, label="Level (m)")
            if pzr_level_norm_out is not None:
                ax2.plot(time[valid], pzr_level_norm_out[valid],
                         color="darkorange", linewidth=line_width,
                         linestyle="--", label="Level (norm)")
            ax1.set_xlabel("Time [s]")
            ax1.set_ylabel("Pressurizer Level [m]", color="steelblue")
            ax2.set_ylabel("Normalised Level [-]",   color="darkorange")
            ax1.set_title(f"Pressurizer Level: {title_name_include_line_1}")
            ax1.grid(True)
            lines1, labs1 = ax1.get_legend_handles_labels()
            lines2, labs2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labs1 + labs2)
            plt.tight_layout()
            plt.savefig("figure_pzr.png")
            plt.close(fig_pzr)

        print(wkstbase)

        try:
            backend = matplotlib.get_backend().lower()
            if "agg" in backend:
                print("Non-interactive backend; plots saved as figure_*.png")
            else:
                print("Figures saved as figure_*.png  |  displaying interactive plots...")
                plt.show()
        except Exception as e:
            print(f"Plot display not available ({e}); plots saved as figure_*.png")

        # Bundle PNGs into PDF
        import re
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.utils import ImageReader

        def natural_key(p):
            m = re.search(r"(\d+)", p.stem)
            return int(m.group(1)) if m else 9999

        folder     = Path(__file__).parent
        img_paths  = sorted(folder.glob("figure_*.png"), key=natural_key)

        if img_paths:
            out_pdf = folder / f"{wkstbase}_figures.pdf"
            c = rl_canvas.Canvas(str(out_pdf), pagesize=letter)
            page_w, page_h = letter
            margin = 36
            for p in img_paths:
                img  = ImageReader(str(p))
                iw, ih = img.getSize()
                scale = min((page_w-2*margin)/iw, (page_h-2*margin)/ih)
                w, h  = iw*scale, ih*scale
                c.drawImage(img, (page_w-w)/2, (page_h-h)/2, width=w, height=h,
                            preserveAspectRatio=True)
                c.showPage()
            c.save()
            print(f"Created PDF: {out_pdf}  ({len(img_paths)} images)")


# ─────────────────────────────────────────────────────────────────────────────
# Helper / physics functions  (unchanged from original)
# ─────────────────────────────────────────────────────────────────────────────
def compute_scaling_ratios(prop_partial_derivatives, mix_properties, Pressure,
                           Total_Mass, massflow_out, massflow_in_avg,
                           stagnation_enthalpy_out, enthalpy_in, net_heat_total):
    dedv = (prop_partial_derivatives["internal_energy"]["wrt_enthalpy"] /
            prop_partial_derivatives["specific_volume"]["wrt_enthalpy"])
    dedp = (-prop_partial_derivatives["specific_volume"]["wrt_pressure"] /
            prop_partial_derivatives["specific_volume"]["wrt_enthalpy"])
    scaling_ratios = {}
    # Guard against zero break flow (pre-break SS phase): return NaN rather
    # than triggering RuntimeWarning.  These ratios are diagnostic only and
    # do not affect the solution.
    _mdot = massflow_out if abs(massflow_out) > 1e-10 else float("nan")
    scaling_ratios["residence_time"]  = Total_Mass / _mdot
    scaling_ratios["massflow_ratio"]  = massflow_in_avg / _mdot
    energy_out_term = (stagnation_enthalpy_out - mix_properties["internal_energy"] +
                       mix_properties["specific_volume"] * dedv)
    energy_in_term  = (enthalpy_in - mix_properties["internal_energy"] +
                       mix_properties["specific_volume"] * dedv)
    _denom = (_mdot * energy_out_term) if abs(energy_out_term) > 1e-10 else float("nan")
    scaling_ratios["dilation"]        = Pressure * dedp / energy_out_term
    scaling_ratios["energyflow_ratio"]= massflow_in_avg * energy_in_term / _denom
    scaling_ratios["power_ratio"]     = net_heat_total / _denom
    return scaling_ratios


def NatConvHTC(Tw, Tb, P, L, hflag=0):
    Tsat = 1.00000001 * XSteam.Tsat_p(P)
    Pw   = XSteam.Psat_T(Tw)
    hf   = 0.999999999999 * XSteam.hL_T(min(Tb, Tsat))
    rhof = XSteam.rho_ph(P, hf)
    cp   = XSteam.cp_ph(P, hf) * 1000
    lambda_ = XSteam.tc_ph(P, hf)
    mu   = XSteam.my_ph(P, hf)
    g    = 9.8
    sigma= XSteam.st_p(P)
    hfg  = 1000 * (XSteam.hV_p(P) - XSteam.hL_p(P))
    rhog = XSteam.rhoV_p(P)
    Tbb  = Tb + 273.15
    beta = 10**-4 * (1.9607E-06*Tbb**3 - 2.2615E-03*Tbb**2 + 9.2045E-01*Tbb - 1.2301E+02)
    if (Tw <= Tsat and Tb <= Tsat) and hflag == 0:
        Gr   = (rhof**2 * g * beta * abs(Tw-Tb) * L**3) / mu**2
        Pr   = mu * cp / lambda_
        Ra   = Gr * Pr
        Nu   = (0.825 + (0.387*Ra**(1/6)) / ((1+(0.492/Pr)**(9/16))**(8/27)))**2
        htc  = Nu * lambda_ / L
    else:
        nhtc = 0.00122 * lambda_**0.79 * cp**0.45 * rhof**0.49
        dhtc = sigma**0.5 * mu**0.29 * hfg**0.24 * rhog**0.25
        htc  = (nhtc/dhtc) * (Tw-Tsat)**0.24 * (1e3*(Pw-P))**0.75
    return htc


def core_htc_db(pressure_Pa, temperature_K, velocity_ms, D_h):
    """
    Dittus-Boelter HTC for rod-bundle core coolant channel.
    Nu = 0.023 * Re^0.8 * Pr^0.4  (heating, subcooled liquid)
    Falls back to laminar Nu=3.66 if Re < 2300.
    Returns HTC in W/m2-K; 0.0 if velocity is negligible.
    """
    if velocity_ms < 1e-6:
        return 0.0
    P_kPa = pressure_Pa * 1e-3
    T_C   = temperature_K - 273.15
    try:
        Tsat  = XSteam.Tsat_p(P_kPa)
        T_lkp = min(T_C, 0.9999 * Tsat)
        hf    = XSteam.hL_T(T_lkp)
        rho   = XSteam.rho_ph(P_kPa, hf)
        cp    = XSteam.cp_ph(P_kPa, hf) * 1e3   # J/kg-K
        k     = XSteam.tc_ph(P_kPa, hf)          # W/m-K
        mu    = XSteam.my_ph(P_kPa, hf)          # Pa-s
    except Exception:
        return 0.0
    Re = rho * velocity_ms * D_h / mu
    Pr = mu * cp / k
    # Always use turbulent D-B — no laminar fallback to Nu=3.66.
    # At low Re the D-B extrapolation still exceeds Churchill-Chu natural
    # convection, correctly maintaining forced-flow dominance until flow
    # is truly negligible. The Nu=3.66 laminar value (~174 W/m²K) would
    # incorrectly allow Churchill-Chu (~877 W/m²K) to dominate while the
    # pump is still spinning.
    Nu = 0.023 * max(Re, 1.0)**0.8 * Pr**0.4
    return Nu * k / D_h


def core_htc_db_steam(pressure_Pa, velocity_ms, D_h):
    """
    Dittus-Boelter HTC for steam-only cooling of fuel rods (post-CHF, high void).
    Nu = 0.023 * Re^0.8 * Pr^0.4  using saturated vapour properties.
    Used during reflood (cold-leg LOCA) and throughout hot-leg LOCA post-CHF.
    Returns HTC in W/m2-K; floor of 50 W/m2-K.
    """
    P_kPa = pressure_Pa * 1e-3
    v_eff = max(velocity_ms, 0.1)   # floor: natural-circulation steam flow
    try:
        hV   = XSteam.hV_p(P_kPa)
        rhoV = XSteam.rhoV_p(P_kPa)
        cpV  = XSteam.cp_ph(P_kPa, hV) * 1e3          # J/kg-K
        kV   = XSteam.tcV_p(P_kPa)                     # W/m-K
        muV  = np.clip(XSteam.my_ph(P_kPa, hV), 1e-6, 1e-3)  # Pa-s
    except Exception:
        return 50.0
    Re = rhoV * v_eff * D_h / muV
    Pr = muV * cpV / kV
    Nu = 3.66 if Re < 2300.0 else 0.023 * Re**0.8 * Pr**0.4
    return max(Nu * kV / D_h, 50.0)


def chf_biasi_zuber(pressure_Pa, G_kgm2s, x_eq, D_h_m, T_K):
    """
    Critical heat flux [W/m²] using the Biasi (1967) / Zuber (1959) model
    with a linear transition, following the approach in RELAP5/MOD2 and TRACE.

    Biasi correlation  (Energia Nucleare 14(9), 530-536, 1967)
    ─────────────────────────────────────────────────────────────
    Input units:  D_h in cm, G in g/cm²s, P in bar.
    Two sub-forms; for G > 300 g/cm²s take the max of both.

      F(P) = 0.7249 + 0.099·P·exp(−0.032·P)
      H(P) = −1.159 + 0.149·P·exp(−0.019·P) + 9·P/(10 + P²)

      Low-flux  (always evaluated):
          q"_lo = 1.5027×10⁷ / D_cm^0.6 · G_cgs^(−1/6) · (H − x)
      High-flux (only when G > 300 g/cm²s; take max):
          q"_hi = 2.764×10⁷  / D_cm^0.6 · G_cgs^(−0.4) · (F − x)

    Validity range (flags set but value still returned):
      P : 0.27–14 MPa;  G : 100–6000 kg/m²s;  x : −0.1–1.0

    Modified Zuber pool-boiling CHF  (RELAP5/MOD3 Eqs. 4.2-78, 4.2-83)
    ─────────────────────────────────────────────────────────────────────
      q"_pb  = K · h_fg · ρ_g · [σ·g·(ρ_f−ρ_g)/ρ_g²]^0.25    K = 0.14 (Folkin-Goldberg)
      F_sub  = 1 + 0.1·(ρ_f/ρ_g)^0.75 · (c_p,f · ΔT_sub / h_fg)
      q"_Zuber = q"_pb · F_sub   (evaluated at saturation props)

    Transition (RELAP5/MOD2 / TRACE convention):
      G < 100 kg/m²s  → pure Zuber
      100 ≤ G ≤ 200   → linear blend
      G > 200 kg/m²s  → pure Biasi

    Returns: q"_cr [W/m²]  (always ≥ 0; negative intermediate results → 0)
    """
    G_LO   = 100.0    # kg/m²s — below this: pure Zuber
    G_HI   = 200.0    # kg/m²s — above this: pure Biasi
    G_CGS_THRESH = 300.0  # g/cm²s = 3000 kg/m²s — Biasi low/high-flux boundary

    P_kPa  = pressure_Pa * 1e-3
    P_bar  = pressure_Pa * 1e-5

    # ── Zuber pool-boiling CHF ──────────────────────────────────────────
    try:
        T_sat_C = XSteam.Tsat_p(P_kPa)
        h_fg    = 1e3 * (XSteam.hV_p(P_kPa) - XSteam.hL_p(P_kPa))   # J/kg
        rho_g   = XSteam.rhoV_p(P_kPa)                                 # kg/m³
        rho_f   = XSteam.rhoL_p(P_kPa)                                 # kg/m³
        sigma   = XSteam.st_p(P_kPa)                                   # N/m
        cp_f    = 1e3 * XSteam.cpL_p(P_kPa)                            # J/kg-K
        dT_sub  = max(T_sat_C - (T_K - 273.15), 0.0)                   # K, ≥ 0
        K_ZUB   = 0.14
        q_pb    = K_ZUB * h_fg * rho_g * (
                      sigma * 9.81 * (rho_f - rho_g) / rho_g**2
                  ) ** 0.25
        F_sub   = 1.0 + 0.1 * (rho_f / rho_g) ** 0.75 * (cp_f * dT_sub / h_fg)
        q_zuber = max(q_pb * F_sub, 0.0)
    except Exception:
        q_zuber = 0.0

    # Pure Zuber regime
    if G_kgm2s <= G_LO:
        return q_zuber

    # ── Biasi CHF ──────────────────────────────────────────────────────
    D_cm  = D_h_m * 100.0                 # m  → cm
    G_cgs = G_kgm2s / 10.0               # kg/m²s → g/cm²s  (1 g/cm²s = 10 kg/m²s)

    F_p = 0.7249 + 0.099 * P_bar * np.exp(-0.032 * P_bar)
    H_p = (-1.159 + 0.149 * P_bar * np.exp(-0.019 * P_bar)
           + 9.0 * P_bar / (10.0 + P_bar**2))

    # Low-flux form — return NaN when H(P) ≤ x_eq (outside Biasi validity)
    if H_p <= x_eq:
        return np.nan          # caller treats NaN DNBR as "correlation invalid"
    q_lo = (1.5027e7 / D_cm**0.6) * G_cgs**(-1.0/6.0) * (H_p - x_eq)
    q_lo = max(q_lo, 0.0)

    # High-flux form (only when G > 300 g/cm²s)
    if G_cgs > G_CGS_THRESH:
        q_hi = (2.764e7 / D_cm**0.6) * G_cgs**(-0.4) * (F_p - x_eq)
        q_biasi = max(q_lo, max(q_hi, 0.0))
    else:
        q_biasi = q_lo

    # Pure Biasi regime
    if G_kgm2s >= G_HI:
        return q_biasi

    # ── Linear transition 100–200 kg/m²s ──────────────────────────────
    w = (G_kgm2s - G_LO) / (G_HI - G_LO)   # 0→1 as G goes 100→200
    return (1.0 - w) * q_zuber + w * q_biasi



def chf_bowring(pressure_Pa, G_kgm2s, x_eq, D_h_m, L_m, T_K):
    """
    Bowring (1972) CHF correlation — AEEW-R-789, Winfrith.
    Approximate quality-explicit form following Tong & Tang (1997).
    Valid range: P = 7–17 MPa, G = 200–6000 kg/m²s, D = 2–45 mm.

    Specifically used here for the HIGH-PRESSURE regime (P > ~13 MPa)
    where the Biasi H(P) function approaches zero.

    NOTE: coefficients require verification against the original paper;
    conservative for scoping use in the 14–17 MPa range.
    """
    P_kPa = pressure_Pa * 1e-3
    p_r   = pressure_Pa / 6.895e6   # normalised to 1000 psia

    if p_r <= 0 or p_r > 2.5:       # outside valid domain
        return np.nan

    n = 2.0 - 0.5 * p_r

    try:
        h_fg  = 1e3 * (XSteam.hV_p(P_kPa) - XSteam.hL_p(P_kPa))
        T_sat = XSteam.Tsat_p(P_kPa)
        cp_f  = 1e3 * XSteam.cpL_p(P_kPa)
    except Exception:
        return np.nan

    if h_fg <= 0:
        return np.nan

    dT_sub = max(T_sat - (T_K - 273.15), 0.0)
    dh_sub = cp_f * dT_sub           # subcooling enthalpy  [J/kg]

    F1 = p_r ** 0.0143
    F2 = p_r ** 0.134
    F3 = p_r ** 0.333
    F4 = p_r ** 1.649

    # Bulk-flow CHF factor (numerator): (H_bowring - x_eq + subcooling_quality)
    H_bow  = F1 / (1.0 + 0.0143 * F2 * np.sqrt(max(G_kgm2s * D_h_m, 0.0)))
    B_sub  = dh_sub / h_fg           # subcooling as equivalent quality increment

    # Length correction factor (denominator)
    C_len  = F3 / (1.0 + 0.347 * F4 * (G_kgm2s / 1356.0) ** n)

    # Explicit CHF flux  [W/m²]
    num = h_fg * G_kgm2s * (H_bow - x_eq + B_sub)
    den = C_len * 4.0 * L_m / D_h_m + 1.0
    q_cr = num / (4.0 * den)

    return max(q_cr, 0.0)


def fission_scram(t, rho_dollars=-20):
    """Fission power fraction after scram.

    rho_dollars: shutdown reactivity in dollars (negative).
                 Default -20 (deeply subcritical generic assumption).
                 Pass trip_rod_worth_pcm / beta_eff_pcm for case-specific value.
    """
    pow, beta_lam = 0, 485.738
    rho = max(rho_dollars, -200)   # floor to prevent numerical issues
    yld_gam = np.array([[0.038,0.0127],[0.213,0.0317],[0.188,0.115],
                         [0.407,0.311],[0.128,1.4],[0.026,3.87]])
    for i in range(len(yld_gam)):
        pow += (beta_lam/(beta_lam-rho*beta_lam) *
                yld_gam[i,0]*np.exp(-yld_gam[i,1]*t)/np.sum(yld_gam[:,0]))
    return pow


def dhp(t, T, R, fr35, fr38, fr39, stdd):
    pow  = fr35*dhp_u235(t,T,stdd) + fr38*dhp_u238(t,T,stdd) + fr39*dhp_pu239(t,T,stdd)
    Gtab = np.array([
        [0e+00,1.02],[1e+00,1.020],[1.5e+00,1.020],[2e+00,1.020],[4e+00,1.021],
        [6e+00,1.022],[8e+00,1.022],[1e+01,1.022],[1.5e+01,1.022],[2e+01,1.022],
        [4e+01,1.022],[6e+01,1.022],[8e+01,1.022],[1e+02,1.023],[1.5e+02,1.024],
        [2e+02,1.025],[4e+02,1.028],[6e+02,1.030],[8e+02,1.032],[1e+03,1.033],
        [1.5e+03,1.037],[2e+03,1.039],[4e+03,1.048],[6e+03,1.054],[8e+03,1.060],
        [1e+04,1.064],[1.5e+04,1.074],[2e+04,1.081],[4e+04,1.098],[6e+04,1.111],
        [8e+04,1.119],[1e+05,1.124],[1.5e+05,1.130],[2e+05,1.131],[4e+05,1.126],
        [6e+05,1.124],[8e+05,1.123],[1e+06,1.124],[1.5e+06,1.125],[2e+06,1.127],
        [4e+06,1.134],[6e+06,1.146],[8e+06,1.162],[1e+07,1.181],[1.5e+07,1.233],
        [2e+07,1.284],[4e+07,1.444],[6e+07,1.535],[8e+07,1.586],[1e+08,1.598],
        [1.5e+08,1.498],[2e+08,1.343],[4e+08,1.065],[6e+08,1.021],[8e+08,1.012],
        [1e+09,1.007]])
    G    = np.interp(t, Gtab[:,0], Gtab[:,1])
    pow *= G
    pow += dhp_act(t, T, R)
    return pow


def dhp_u235(t, T, stdd):
    if stdd == 1979:
        alplam = np.array([
            [6.5057E-01,2.2138E+01],[5.1264E-01,5.1587E-01],[2.4384E-01,1.9594E-01],
            [1.3850E-01,1.0314E-01],[5.5440E-02,3.3656E-02],[2.2225E-02,1.1681E-02],
            [3.3088E-03,3.5870E-03],[9.3015E-04,1.3930E-03],[8.0943E-04,6.2630E-04],
            [1.9567E-04,1.8906E-04],[3.2535E-05,5.4988E-05],[7.5595E-06,2.0958E-05],
            [2.5232E-06,1.0010E-05],[4.9948E-07,2.5438E-06],[1.8531E-07,6.6361E-07],
            [2.6608E-08,1.2290E-07],[2.2398E-09,2.7213E-08],[8.1641E-12,4.3714E-09],
            [8.7797E-11,7.5780E-10],[2.5131E-14,2.4786E-10],[3.2176E-16,2.2384E-13],
            [4.5038E-17,2.4600E-14],[7.4791E-17,1.5699E-14]])
    else:
        alplam = np.array([
            [5.2800E-04,2.7216E+00],[6.8588E-01,1.0256E+00],[4.0752E-01,3.1419E-01],
            [2.1937E-01,1.1788E-01],[5.7701E-02,3.4365E-02],[2.2530E-02,1.1762E-02],
            [3.3392E-03,3.6065E-03],[9.3667E-04,1.3963E-03],[8.0899E-04,6.2608E-04],
            [1.9572E-04,1.8924E-04],[3.2609E-05,5.5074E-05],[7.5827E-06,2.0971E-05],
            [2.5189E-06,9.9940E-06],[4.9836E-07,2.5401E-06],[1.8523E-07,6.6332E-07],
            [2.6592E-08,1.2281E-07],[2.2356E-09,2.7163E-08],[8.9582E-12,3.2956E-09],
            [8.5968E-11,7.4225E-10],[2.1072E-14,2.4681E-10],[7.1219E-16,1.5596E-13],
            [8.1126E-17,2.2573E-14],[9.4678E-17,2.0503E-14]])
    pow = 0
    for a,l in alplam:
        pow += (a/l)*np.exp(-l*t)*(1-np.exp(-l*T))
    return pow


def dhp_u238(t, T, stdd):
    if stdd == 1979:
        alplam = np.array([
            [1.23E+00,3.29E+00],[1.15E+00,9.38E-01],[7.07E-01,3.71E-01],
            [2.52E-01,1.11E-01],[7.19E-02,3.61E-02],[2.83E-02,1.33E-02],
            [6.84E-03,5.01E-03],[1.23E-03,1.37E-03],[6.84E-04,5.52E-04],
            [1.70E-04,1.79E-04],[2.42E-05,4.90E-05],[6.64E-06,1.71E-05],
            [1.01E-06,7.05E-06],[4.99E-07,2.32E-06],[1.64E-07,6.45E-07],
            [2.34E-08,1.26E-07],[2.81E-09,2.55E-08],[3.62E-11,8.48E-09],
            [6.46E-11,7.51E-10],[4.50E-14,2.42E-10],[3.67E-16,2.27E-13],
            [5.63E-17,9.05E-14],[7.16E-17,5.61E-15]])
    else:
        alplam = np.array([
            [3.9368E-01,4.3427E+00],[7.4588E-01,1.7114E+00],[1.2169E+00,6.0572E-01],
            [5.2820E-01,1.9429E-01],[1.4805E-01,6.9788E-02],[4.5980E-02,1.8809E-02],
            [1.0406E-02,6.1265E-03],[1.6991E-03,1.3799E-03],[6.9102E-04,5.2799E-04],
            [1.4736E-04,1.6145E-04],[2.4049E-05,4.8419E-05],[6.9288E-06,1.5644E-05],
            [6.4927E-07,5.3610E-06],[4.3556E-07,2.1689E-06],[1.6020E-07,6.3343E-07],
            [2.3089E-08,1.2879E-07],[2.5481E-09,2.5604E-08],[3.5071E-11,9.1544E-09],
            [6.3399E-11,7.3940E-10],[4.1599E-14,2.4731E-10],[5.3295E-16,1.9594E-13],
            [1.6695E-18,6.4303E-14],[4.1058E-16,6.4229E-14]])
    pow = 0
    for a,l in alplam:
        pow += (a/l)*np.exp(-l*t)*(1-np.exp(-l*T))
    return pow


def dhp_pu239(t, T, stdd):
    if stdd == 1979:
        alplam = np.array([
            [2.083E-01,1.002E+01],[3.853E-01,6.433E-01],[2.213E-01,2.186E-01],
            [9.460E-02,1.004E-01],[3.531E-02,3.728E-02],[2.292E-02,1.435E-02],
            [3.946E-03,4.549E-03],[1.317E-03,1.328E-03],[7.052E-04,5.356E-04],
            [1.432E-04,1.730E-04],[1.765E-05,4.881E-05],[7.347E-06,2.006E-05],
            [1.747E-06,8.319E-06],[5.481E-07,2.358E-06],[1.671E-07,6.450E-07],
            [2.112E-08,1.278E-07],[2.996E-09,2.466E-08],[5.107E-11,9.378E-09],
            [5.703E-11,7.450E-10],[4.138E-14,2.426E-10],[1.088E-15,2.210E-13],
            [2.454E-17,2.640E-14],[7.557E-17,1.380E-14]])
    else:
        alplam = np.array([
            [3.0934E-01,2.9049E+00],[5.4434E-01,6.4911E-01],[4.0782E-01,2.5569E-01],
            [1.5828E-01,8.7123E-02],[4.1577E-02,2.5068E-02],[1.4818E-02,1.3323E-02],
            [5.8176E-03,6.3772E-03],[1.9482E-03,2.0221E-03],[9.5196E-04,6.2933E-04],
            [1.8208E-04,1.7462E-04],[1.5310E-05,4.0172E-05],[4.5039E-06,1.5289E-05],
            [9.8277E-07,7.6118E-06],[5.1832E-07,2.5083E-06],[2.3018E-08,1.1312E-06],
            [1.5817E-07,6.2987E-07],[1.8074E-08,1.3149E-07],[3.6922E-09,2.4237E-08],
            [5.3843E-11,9.6433E-09],[5.3003E-11,7.3467E-10],[4.8358E-14,2.4827E-10],
            [9.8516E-16,1.6873E-13],[1.3076E-16,8.3639E-15]])
    pow = 0
    for a,l in alplam:
        pow += (a/l)*np.exp(-l*t)*(1-np.exp(-l*T))
    return pow


def dhp_act(t, T, R):
    U239  = 0.474*R*(1-np.exp(-0.000491*T))*np.exp(-0.000491*t)
    NP239 = (0.419*R*(0.000491/(0.000491-0.00000341)*
             (1-np.exp(-0.00000341*T))*np.exp(-0.00000341*t) -
             0.00000341/(0.000491-0.00000341)*
             (1-np.exp(-0.000491*T))*np.exp(-0.000491*t)))
    return U239 + NP239


def get_saturation_prop(pressure):
    pressure_kPa = pressure * 1e-3
    sat_liquid = {"specific_volume": XSteam.vL_p(pressure_kPa),
                  "internal_energy": 1000*XSteam.uL_p(pressure_kPa),
                  "entropy":         1000*XSteam.sL_p(pressure_kPa),
                  "enthalpy":        1000*XSteam.hL_p(pressure_kPa)}
    sat_vapor  = {"specific_volume": XSteam.vV_p(pressure_kPa),
                  "internal_energy": 1000*XSteam.uV_p(pressure_kPa),
                  "entropy":         1000*XSteam.sV_p(pressure_kPa),
                  "enthalpy":        1000*XSteam.hV_p(pressure_kPa)}
    return sat_liquid, sat_vapor


def critical_flow_newton(stagnation_pressure, stagnation_enthalpy, slip_type,
                         critical_pressure_guess_init, solver_settings=None):
    if solver_settings is None:
        solver_settings = {"Pressure_increment_psia":1e-2,
                           "convergence_tolerance":1e-5,
                           "max_iterations":200,"min_iterations":20}
    pressure_increment = 1e5*0.06895*solver_settings["Pressure_increment_psia"]
    convergence_tolerance = solver_settings["convergence_tolerance"]
    max_iterations  = solver_settings["max_iterations"]
    min_iterations  = solver_settings["min_iterations"]
    diff_value = 10; number_iter = 1
    critical_pressure_guess = critical_pressure_guess_init
    critical_pressure_guess_vector = np.full(max_iterations, np.nan)
    while diff_value >= convergence_tolerance and number_iter <= max_iterations:
        critical_pressure_guess_vector[number_iter-1] = critical_pressure_guess
        cpu = critical_pressure_guess + pressure_increment
        cpd = critical_pressure_guess - pressure_increment
        G0,_ = two_phase_critical_flow_model(stagnation_pressure,stagnation_enthalpy,slip_type,critical_pressure_guess)
        Gu,_ = two_phase_critical_flow_model(stagnation_pressure,stagnation_enthalpy,slip_type,cpu)
        Gd,_ = two_phase_critical_flow_model(stagnation_pressure,stagnation_enthalpy,slip_type,cpd)
        fd = (Gu-Gd)/(2*pressure_increment)
        sd = (Gu-2*G0+Gd)/pressure_increment**2
        if sd == 0:
            if number_iter < max_iterations-min_iterations:
                critical_pressure_guess = (1+np.random.randn())*critical_pressure_guess_init
                diff_value = 10; number_iter += 1
            else:
                critical_pressure_guess = critical_pressure_guess_vector[number_iter-2]; break
        else:
            cp_new = critical_pressure_guess - fd/sd
            diff_value = abs(fd)
            if diff_value < convergence_tolerance and sd >= 0:
                critical_pressure_guess = (1+np.random.randn())*critical_pressure_guess_init
                diff_value = 10; number_iter += 1
            else:
                number_iter += 1; critical_pressure_guess = cp_new
    critical_pressure = critical_pressure_guess
    solver_error = diff_value
    if np.isnan(critical_pressure):
        critical_massflux = 0; critical_enthalpy = 0
    else:
        critical_massflux, critical_enthalpy = two_phase_critical_flow_model(
            stagnation_pressure, stagnation_enthalpy, slip_type, critical_pressure_guess)
    return critical_massflux, critical_enthalpy, critical_pressure, solver_error


def two_phase_critical_flow_model(Po, ho, slip_type, Pc):
    sat_liquid_o, sat_vapor_o = get_saturation_prop(Po)
    sat_liquid_c, sat_vapor_c = get_saturation_prop(Pc)
    sfg_o = sat_vapor_o["entropy"] - sat_liquid_o["entropy"]
    sfg_c = sat_vapor_c["entropy"] - sat_liquid_c["entropy"]
    if sfg_o <= 0 or sfg_c <= 0:   # above critical pressure — not applicable
        return 0.0, ho
    xo    = (ho - sat_liquid_o["enthalpy"]) / (sat_vapor_o["enthalpy"] - sat_liquid_o["enthalpy"])
    xc    = xo*sfg_o/sfg_c + (sat_liquid_o["entropy"]-sat_liquid_c["entropy"])/sfg_c
    hc    = xc*(sat_vapor_c["enthalpy"]-sat_liquid_c["enthalpy"]) + sat_liquid_c["enthalpy"]
    if   slip_type == 'HEM':    S = 1
    elif slip_type == 'Moody':  S = (sat_vapor_c["specific_volume"]/sat_liquid_c["specific_volume"])**(1/3)
    elif slip_type == 'Fauske': S = np.sqrt(sat_vapor_c["specific_volume"]/sat_liquid_c["specific_volume"])
    else: raise ValueError("Invalid slip type.")
    rhoa  = sat_vapor_c["specific_volume"]*xc + (1-xc)*S*sat_liquid_c["specific_volume"]
    rhob  = xc + (1-xc)/S**2
    rho_3 = 1/(rhoa**2*rhob)
    _arg  = rho_3 * 2 * (ho - hc)
    G     = np.sqrt(max(0.0, _arg))   # guard against negative from floating point at low dP
    return G, hc


def critical_flow_ERM(pressure, enthalpy_mix, sat_liquid, sat_vapor):
    """
    Critical mass flux [kg/m²s] — ERM blended to subcooled Bernoulli at high P.

    The Equilibrium Rate Model (Moody 1965) assumes full thermal equilibrium
    flash in the nozzle throat.  This is valid at PWR operating pressures
    (~15-16 MPa) where hfg is large, but becomes increasingly non-physical
    above ~16.5 MPa: the fluid is near its critical point, hfg collapses,
    and the short PORV throat residence time precludes equilibrium flashing.
    The flow approaches incompressible subcooled liquid orifice discharge.

    Blend thresholds (based on hfg, which correlates with distance from
    the thermodynamic critical point):
      hfg >= HFG_ERM  (≈ 16.4 MPa): pure ERM  — no change to LOCA cases
      hfg <= HFG_BERN (≈ 22.0 MPa): pure Bernoulli subcooled orifice
      between: linear blend by hfg fraction

    P_back for Bernoulli is containment pressure (101 kPa).
    """
    HFG_ERM  = 900.0e3   # J/kg — pure ERM above this  (≈ 16.4 MPa)
    HFG_BERN = 150.0e3   # J/kg — pure Bernoulli below this (≈ 22.0 MPa)

    hfg = sat_vapor["enthalpy"] - sat_liquid["enthalpy"]
    vfg = sat_vapor["specific_volume"] - sat_liquid["specific_volume"]

    # Bernoulli subcooled orifice (always computed as the high-P fallback)
    try:
        rho = 1.0 / XSteam.v_ph(1e-3*pressure, enthalpy_mix/1000)
        G_bern = float(np.sqrt(max(2.0*rho*(pressure - 101.0e3), 0.0)))
    except Exception:
        G_bern = 0.0

    if vfg <= 0 or hfg <= 0:
        # Above critical pressure — Bernoulli only
        return G_bern

    To      = XSteam.T_ph(1e-3*pressure, enthalpy_mix/1000)
    cp_f    = 1000*XSteam.cpL_p(1e-3*pressure)
    G_ERM_0 = (hfg/vfg)*((0.64/(To+273.15)/cp_f)**0.5)
    rhof    = 1/XSteam.v_ph(1e-3*pressure, enthalpy_mix/1000)
    Psat_To = 1e3*XSteam.Psat_T(To)
    G_erm   = float(np.sqrt(max(2*(pressure-Psat_To)*rhof + G_ERM_0**2, 0.0)))

    if hfg >= HFG_ERM:
        return G_erm                      # pure ERM — LOCA/normal range
    elif hfg <= HFG_BERN:
        return G_bern                     # pure Bernoulli — near critical
    else:
        # Linear blend: w=0 at HFG_ERM → w=1 at HFG_BERN
        w = (HFG_ERM - hfg) / (HFG_ERM - HFG_BERN)
        return (1.0 - w)*G_erm + w*G_bern


def compute_property_partials(x_eq, pressure, enthalpy_mix, sat_vapor, sat_liquid,
                               pressure_increment, enthalpy_increment):
    ppi = pressure + 0.5*pressure_increment
    pmi = pressure - 0.5*pressure_increment
    epi = enthalpy_mix + 0.5*enthalpy_increment
    emi = enthalpy_mix - 0.5*enthalpy_increment
    ppd = {"internal_energy":{"wrt_enthalpy":0,"wrt_pressure":0},
           "specific_volume":{"wrt_enthalpy":0,"wrt_pressure":0}}
    if 0 <= x_eq <= 1:
        hfg = sat_vapor["enthalpy"]       - sat_liquid["enthalpy"]
        vfg = sat_vapor["specific_volume"] - sat_liquid["specific_volume"]
        ppd["internal_energy"]["wrt_enthalpy"] = 1 - pressure*vfg/hfg
        ppd["specific_volume"]["wrt_enthalpy"] = vfg/hfg
        dhfdP = 1000*(XSteam.hL_p(1e-3*ppi)-XSteam.hL_p(1e-3*pmi))/(ppi-pmi)
        dhgdP = 1000*(XSteam.hV_p(1e-3*ppi)-XSteam.hV_p(1e-3*pmi))/(ppi-pmi)
        dvfdP = (XSteam.vL_p(1e-3*ppi)-XSteam.vL_p(1e-3*pmi))/(ppi-pmi)
        dvgdP = (XSteam.vV_p(1e-3*ppi)-XSteam.vV_p(1e-3*pmi))/(ppi-pmi)
        dhfgdP= dhgdP-dhfdP; dvfgdP=dvgdP-dvfdP
        ppd["internal_energy"]["wrt_pressure"] = (-x_eq*dvfgdP-dvfdP+(vfg/hfg)*(x_eq*dhfgdP+dhfdP))
        ppd["specific_volume"]["wrt_pressure"] = ( x_eq*dvfgdP+dvfdP-(vfg/hfg)*(x_eq*dhfgdP+dhfdP))
    else:
        ppd["internal_energy"]["wrt_pressure"] = 1000*(XSteam.u_ph(1e-3*ppi,enthalpy_mix/1000)-XSteam.u_ph(1e-3*pmi,enthalpy_mix/1000))/(ppi-pmi)
        ppd["internal_energy"]["wrt_enthalpy"] = 1000*(XSteam.u_ph(1e-3*pressure,epi/1000)-XSteam.u_ph(1e-3*pressure,emi/1000))/(epi-emi)
        ppd["specific_volume"]["wrt_pressure"] = (XSteam.v_ph(1e-3*ppi,enthalpy_mix/1000)-XSteam.v_ph(1e-3*pmi,enthalpy_mix/1000))/(ppi-pmi)
        ppd["specific_volume"]["wrt_enthalpy"] = (XSteam.v_ph(1e-3*pressure,epi/1000)-XSteam.v_ph(1e-3*pressure,emi/1000))/(epi-emi)
    return ppd


def compute_mixture_properties(pressure, enthalpy_mix):
    px = 1e-3*pressure
    sat_liquid = {"enthalpy":1000*XSteam.hL_p(px), "specific_volume":XSteam.vL_p(px)}
    sat_vapor  = {"enthalpy":1000*XSteam.hV_p(px), "specific_volume":XSteam.vV_p(px)}
    hfg  = sat_vapor["enthalpy"] - sat_liquid["enthalpy"]
    if hfg <= 0:
        # At or above the critical pressure hfg → 0; treat as single-phase
        x_eq = -1.0
    else:
        x_eq = (enthalpy_mix - sat_liquid["enthalpy"]) / hfg
    mix  = {}
    if 0 <= x_eq <= 1:
        sat_liquid["internal_energy"] = 1000*XSteam.uL_p(px)
        sat_vapor["internal_energy"]  = 1000*XSteam.uV_p(px)
        efg = sat_vapor["internal_energy"] - sat_liquid["internal_energy"]
        vfg = sat_vapor["specific_volume"] - sat_liquid["specific_volume"]
        mix["internal_energy"] = x_eq*efg + sat_liquid["internal_energy"]
        mix["specific_volume"] = x_eq*vfg + sat_liquid["specific_volume"]
    else:
        mix["internal_energy"] = 1000*XSteam.u_ph(px, enthalpy_mix/1000)
        mix["specific_volume"] = XSteam.v_ph(px, enthalpy_mix/1000)
    return x_eq, mix, sat_vapor, sat_liquid


def get_variable(namespace, var_name, default):
    value = namespace[var_name] if var_name in namespace else default
    source = "user" if var_name in namespace else "default"
    # Record to the module-level input echo log
    _input_echo_log.append({
        "Variable": var_name,
        "Value": value,
        "Default": default,
        "Source": source,
    })
    return value


# Module-level log — reset at the start of each pwr_sim call
_input_echo_log = []


# ─────────────────────────────────────────────────────────────────────────────
# pwr_sim  — main simulation
# ─────────────────────────────────────────────────────────────────────────────
def pwr_sim(wkstbase, minimum_output=0):
    global _input_echo_log
    _input_echo_log = []   # reset for this run
    # ── Validate input file before touching anything else ─────────────────
    fn_in  = f"{wkstbase}_in.xlsx"
    sheet  = wkstbase + '_in'
    if not Path(fn_in).exists():
        print(f"Error: input file not found.")
        print(f"  Expected : {fn_in}")
        print(f"  Folder   : {Path.cwd()}")
        print(f"  Check that the filename matches the case name exactly.")
        return

    filename = f"{wkstbase}_out.xlsx"
    try:
        pd.DataFrame([0]).to_excel(filename, index=False, header=False)
        print(f"File '{filename}' initialized.")
    except Exception as e:
        print(f"Error initializing '{filename}': {e}")
        print("  Ensure the output file is not open in Excel.")
        return

    # ── Input snapshot pipeline ──────────────────────────────────────────────
    # Excel is treated as the user-facing editing format only. At run start:
    #   1. Copy the user's workbook to a private temporary snapshot.
    #   2. Archive that exact snapshot into the case output folder using the
    #      original workbook filename.
    #   3. Export the input worksheet to CSV.
    #   4. Parse the CSV snapshot for command/table data.
    #
    # This lets the user keep the original workbook open in Excel. After the
    # initial copy, the original file is never opened again.
    import csv as _csv
    import uuid as _uuid

    _input_original = Path(fn_in)

    # Use the current working directory as the run archive directory.
    # When launched from flare_ui.py, cwd is already the run folder:
    #     sim_<case>_<timestamp>
    # When run standalone, cwd is the FLARE folder, preserving legacy behavior.
    _run_archive_dir = Path.cwd()

    _tmp_in = Path(tempfile.gettempdir()) / f"FLARE_{wkstbase}_{os.getpid()}_{_uuid.uuid4().hex[:8]}_in.xlsx"
    shutil.copy2(_input_original, _tmp_in)

    _archived_input_xlsx = _run_archive_dir / _input_original.name

    # If the input workbook already lives in the run folder, do not overwrite it
    # with itself.  Otherwise archive the exact private snapshot used by this run.
    try:
        _same_input_archive = _input_original.resolve() == _archived_input_xlsx.resolve()
    except Exception:
        _same_input_archive = False

    if not _same_input_archive:
        shutil.copy2(_tmp_in, _archived_input_xlsx)

    _csv_in = _run_archive_dir / f"{wkstbase}_in.csv"

    # Open only the private workbook snapshot. The original workbook may remain
    # open in Excel without interfering with the simulation.
    _wb_all = openpyxl.load_workbook(str(_tmp_in), read_only=True, data_only=True)
    _ws_in  = _wb_all[sheet]

    # Export the input worksheet to CSV for all command/table parsing.
    _csv_rows = []
    for _row in _ws_in.iter_rows(values_only=True):
        _row_vals = ["" if _v is None else _v for _v in _row]
        _csv_rows.append(_row_vals)

    with _csv_in.open("w", newline="", encoding="utf-8") as _f_csv:
        _writer = _csv.writer(_f_csv)
        _writer.writerows(_csv_rows)

    print(f"Input snapshot archived: {_archived_input_xlsx}")
    print(f"Input CSV snapshot written: {_csv_in}")

    # ── Single-pass CSV parse ────────────────────────────────────────────────
    # Robust parsing strategy:
    #   - No row cap: scan entire column A regardless of file length
    #   - Time header: strict match — cell must start with "time" (no "="),
    #     AND the next non-blank cell must be numeric. This guards against
    #     "endtime = ...", "# ... Endtime ...", and other false matches.
    #   - Table: tolerates up to 3 consecutive blank/non-numeric rows before
    #     stopping (handles visual separator rows the user may have inserted).
    #   - Command block: all rows before the Time header.
    _all_col_a = [_row[0] if len(_row) > 0 else None for _row in _csv_rows]

    # Locate the Time header using strict criteria
    _time_header_row = None   # 0-indexed position in _all_col_a
    for _i, _v in enumerate(_all_col_a):
        if not isinstance(_v, str):
            continue
        _vs = _v.strip().lower().replace(" ", "").replace("\t", "")
        if _vs.startswith("time") and "=" not in _vs:
            # Confirm: a numeric value must follow within the next 3 rows
            for _j in range(_i + 1, min(_i + 4, len(_all_col_a))):
                _nxt = _all_col_a[_j]
                if _nxt is None:
                    continue
                try:
                    float(_nxt)
                    _time_header_row = _i
                    break
                except (TypeError, ValueError):
                    break
        if _time_header_row is not None:
            break

    # Command block = everything before the Time header
    _cmd_end = _time_header_row if _time_header_row is not None else len(_all_col_a)
    comdata  = _all_col_a[:_cmd_end]

    if _time_header_row is None:
        print(f"WARNING: No time-series table header found in sheet '{sheet}'. "
              "Using a default zero table.")

    start_row = _time_header_row + 1 if _time_header_row is not None else 50

    # Read table columns A-F for rows after the Time header from the CSV snapshot
    _table_rows = []
    if _time_header_row is not None:
        _blank_run = 0
        for _raw_row in _csv_rows[_time_header_row + 1:]:
            _row = list(_raw_row[:6]) + [""] * max(0, 6 - len(_raw_row))
            _row = tuple(None if _v == "" else _v for _v in _row[:6])
            _t = _row[0]
            if _t is None:
                _blank_run += 1
                if _blank_run > 3:
                    break
                continue
            try:
                float(_t)
            except (TypeError, ValueError):
                _blank_run += 1
                if _blank_run > 3:
                    break
                continue
            _blank_run = 0
            _table_rows.append(_row)

    # R5 reference data from _out sheet if it exists
    _sheet_out = wkstbase + '_out'
    _r5_rows = []
    if _sheet_out in _wb_all.sheetnames:
        _ws_out = _wb_all[_sheet_out]
        _found_data = False
        for _row in _ws_out.iter_rows(min_col=1, max_col=12, values_only=True):
            if not _found_data:
                if _row[0] is not None:
                    try:
                        float(_row[0]); _found_data = True
                    except (TypeError, ValueError):
                        continue
            if _found_data:
                if _row[0] is None:
                    break
                _r5_rows.append(_row)

    _wb_all.close()



    # Convert table to numpy arrays — pad to 6 columns for backward compatibility
    # with older input files that only have columns A:E (no rho_ext column F).
    if _table_rows:
        _ncols = max(len(r) for r in _table_rows)
        _tbl = np.array(
            [[float(c) if c is not None else 0.0 for c in r] +
             [0.0] * (6 - len(r))
             for r in _table_rows], dtype=float)
        if _tbl.shape[1] < 6:
            _pad = np.zeros((_tbl.shape[0], 6 - _tbl.shape[1]))
            _tbl = np.hstack([_tbl, _pad])
    else:
        _tbl = np.zeros((2, 6))
        _tbl[:, 0] = [0.0, 36000.0]

    time_xls = _tbl[:, 0]
    _col_b   = _tbl[:, 1]   # Structure Q [MW]
    _col_c   = _tbl[:, 2]   # Decay heat [MW]
    _col_d   = _tbl[:, 3]   # SG table [MW]
    _col_e   = _tbl[:, 4]   # SI flow [kg/s]
    _col_f   = _tbl[:, 5]   # rho_ext [pcm]

    nrows = len(time_xls)

    # R5 reference arrays
    if _r5_rows:
        _r5 = np.array([[float(c) if c is not None else 0.0 for c in r]
                        for r in _r5_rows], dtype=float)
    else:
        _r5 = None

    def _sanitise_command(cmd):
        """Normalise common input-file formatting issues before exec.

        1. Unicode minus/dash characters → ASCII hyphen-minus.
        2. Trailing unit annotations stripped from RHS before numeric check
           (e.g. '= -3,000 pcm' — units removed first, then comma stripped).
        3. Thousands-separator commas in plain numeric RHS (e.g. -3,000 → -3000).
           Only fires when the entire RHS (sans units) is a bare numeric literal —
           dicts, lists, strings, and expressions are left untouched.
        """
        import re as _re

        # 1. Unicode minus variants → ASCII hyphen-minus (must be first)
        for _ch in ('\u2212', '\u2013', '\u2014'):
            cmd = cmd.replace(_ch, '-')

        if '=' in cmd:
            _lhs, _, _rhs_raw = cmd.partition('=')
            _comment_split = _rhs_raw.split('#', 1)
            _rhs = _comment_split[0].strip()
            _comment = (' #' + _comment_split[1]) if len(_comment_split) > 1 else ''

            # 2. Strip trailing unit annotation from RHS before numeric check
            _rhs_no_unit = _re.sub(
                r'\s+[A-Za-z][A-Za-z0-9/°·\s]*$', '', _rhs).strip()

            # 3. Comma stripping — only when RHS (sans units) is plain numeric
            if _re.match(r'^[+-]?[\d,]+\.?\d*([eE][+-]?\d+)?\s*$', _rhs_no_unit):
                _rhs_clean = _rhs_no_unit.replace(',', '').strip()
                cmd = _lhs + '= ' + _rhs_clean + _comment

        return cmd.rstrip()

    local_namespace = {}
    for command in comdata:
        if isinstance(command, str) and "=" in command:
            try:
                # Strip inline comments, then sanitise formatting issues
                _cmd = _sanitise_command(command.split("#")[0].strip())
                parts = _cmd.split("=", 1)
                if len(parts) == 2 and parts[0].strip().isidentifier():
                    exec(_cmd, {}, local_namespace)
            except Exception as e:
                print(f"Warning: Could not execute '{command}': {e}")

    # ── time / geometry defaults ────────────────────────────────────────────
    timestep   = 1
    endtime       = get_variable(local_namespace, "endtime", 36000)
    R5_lowCV_height  = get_variable(local_namespace, "R5_lowCV_height",  70*0.3048)
    R5_highCV_height = get_variable(local_namespace, "R5_highCV_height",  7*0.3048)
    R5_area          = get_variable(local_namespace, "R5_area",          33.0*0.3048**2)

    # ── break location / elevation model ─────────────────────────────────────
    # break_location is a user-facing descriptor used to set the elevation at
    # which the break has access to liquid inventory.  The default remains the
    # top of the reactor vessel, preserving prior behavior.  If a user needs a
    # plant-specific elevation, break_elevation_m overrides the named location.
    break_location = str(get_variable(local_namespace,
                                      "break_location",
                                      "top_of_vessel")).strip("'\"").lower()
    break_location = break_location.replace(" ", "_").replace("-", "_")
    break_elevation_m_user = get_variable(local_namespace,
                                          "break_elevation_m",
                                          None)

    def _resolve_break_elevation_m(_location, _user_elev=None):
        _vessel_top = R5_lowCV_height + R5_highCV_height
        if _user_elev is not None:
            try:
                _elev = float(_user_elev)
                return float(np.clip(_elev, 0.0, _vessel_top))
            except Exception:
                print(f"Warning: invalid break_elevation_m '{_user_elev}'; using break_location.")

        _aliases = {
            "top_of_vessel": _vessel_top,
            "vessel_top":    _vessel_top,
            "rpv_top":       _vessel_top,
            "rv_top":        _vessel_top,
            "upper_vessel":  _vessel_top,
            "upper_head":    _vessel_top,
            "top":           _vessel_top,
            "core_exit":     R5_lowCV_height,
            "upper_plenum":  R5_lowCV_height,
            "vessel_bottom": 0.0,
            "bottom_of_vessel": 0.0,
            "bottom":        0.0,
        }
        if _location not in _aliases:
            print(f"Warning: unrecognised break_location '{_location}'; using top_of_vessel. "
                  f"Use break_elevation_m for plant-specific break elevations.")
        return float(_aliases.get(_location, _vessel_top))

    break_elevation_m = _resolve_break_elevation_m(break_location,
                                                   break_elevation_m_user)
    # loop_vol_m3: total primary loop volume OUTSIDE the RPV (hot/cold legs,
    # pump bowl, pressurizer, SG primary side).  Added to ves_vol so that
    # Total_Mass and the 2x2 pressure-rise equation use the full RCS inventory.
    # Default 0.0 preserves backward compatibility.
    loop_vol_m3      = float(get_variable(local_namespace, "loop_vol_m3", 0.0))
    # Total_Mass_init_kg: override the initial RCS inventory.
    # Default (0.0) fills the vessel volume completely at the initial P and T,
    # which is correct for blowdown/accident cases but wrong for a startup from
    # cold conditions — the cold-filled inventory is too heavy and causes
    # continuous PORV relief as the coolant expands during heatup.
    # For startup cases, set this to the mass that fills the vessel at rated
    # operating temperature and pressure (ves_vol / v(P_rated, T_rated)).
    Total_Mass_init_kg = float(get_variable(local_namespace, "Total_Mass_init_kg", 0.0))

    # ── pressurizer ──────────────────────────────────────────────────────────
    # Reduced-order pressurizer diagnostics.  Pressure is solved by the lumped
    # RCS mass/energy/volume model; the pressurizer no longer clamps pressure.
    # Pressurizer inputs are used only for level/steam-space diagnostics and
    # for selecting a physically consistent PORV discharge enthalpy.
    pzr_area        = float(get_variable(local_namespace, "pzr_area",       0.0))  # m²
    pzr_height      = float(get_variable(local_namespace, "pzr_height",     0.0))  # m
    pzr_level_init  = float(get_variable(local_namespace, "pzr_level_init", 0.0))  # m
    pzr_level_hi    = float(get_variable(local_namespace,
                            "pzr_level_hi", 0.9 * pzr_height))                     # m
    # Deprecated/ignored legacy pressure-hold inputs.  Retained only so older
    # input files do not fail if they still contain these variables.
    pzr_hold_kPa = get_variable(local_namespace, "pzr_hold_kPa", None)
    pzr_pressure_hold = 0
    pzr_vol    = pzr_area * pzr_height   # total pressurizer volume [m³]
    # Algebraic level diagnostic parameters (from reduced-order model doc)
    pzr_K_vol_T = float(get_variable(local_namespace, "pzr_K_vol_T", 0.12))  # m³/K thermal expansion coefficient
    pzr_T0      = float(get_variable(local_namespace, "pzr_T0",      580.0)) # K  reference temperature

    # ── accumulator ─────────────────────────────────────────────────────────
    acc_area   = get_variable(local_namespace, "acc_area",   66.63*0.3048**2)
    acc_length = get_variable(local_namespace, "acc_length", 10.5*0.3048)
    acc_totvol = get_variable(local_namespace, "acc_totvol", acc_area*acc_length)
    acc_gasvol = get_variable(local_namespace, "acc_gasvol", acc_totvol-630.0*0.3048**3)
    acc_narea  = get_variable(local_namespace, "acc_narea",  0.032*0.3048**2)
    inflow_info= get_variable(local_namespace, "inflow_info",
                              {'pressure':600*6.895,'temp':(120-32)*5/9})
    tgas       = get_variable(local_namespace, "tgas", inflow_info['temp']+273.15)

    # ── initial conditions ──────────────────────────────────────────────────
    pressure_kPa          = get_variable(local_namespace, "pressure_kPa",         14820)
    pressure_containment  = get_variable(local_namespace, "pressure_containment",  101)
    temp_core_exit        = get_variable(local_namespace, "temp_core_exit",        312.8)


    # ── break model ─────────────────────────────────────────────────────────
    diameter_break                = get_variable(local_namespace, "diameter_break",   0.0)
    choked                        = get_variable(local_namespace, "choked",            1)
    Cd                            = get_variable(local_namespace, "Cd",               0.9)
    Cd_sub                        = get_variable(local_namespace, "Cd_sub",           Cd)
    Cd_sat                        = get_variable(local_namespace, "Cd_sat",           Cd)
    critical_pressure_ratio_guess = get_variable(local_namespace, "crticial_pressure_ratio_guess", 0.55)
    flag_vent_vapor               = get_variable(local_namespace, "flag_vent_vapor",  0)
    initial_flag_stagnation_break = 'mixture'
    flag_stagnation_break         = initial_flag_stagnation_break
    transition_mixture_mass       = get_variable(local_namespace, "transition_mixture_mass", 0.41)
    slip_type   = get_variable(local_namespace, "Slip_model",   "Fauske")
    if slip_type not in ("HEM", "Moody", "Fauske"):
        print(f"Warning: unrecognised Slip_model '{slip_type}', defaulting to 'Fauske'")
        slip_type = "Fauske"

    # ── ADV ─────────────────────────────────────────────────────────────────
    setpoint_ADV_open = get_variable(local_namespace, "setpoint_ADV_open", 13.8)
    area_per_ADV      = get_variable(local_namespace, "area_per_ADV",      0.01266)
    ADV_number        = get_variable(local_namespace, "ADV_number",        0)

    # ── PORV (high-pressure relief valve) ────────────────────────────────────
    # Hysteresis/stroke model. Opens at PORV_setpoint_kPa (sensed on predicted
    # unrelieved pressure P_none), reseats at PORV_reseat_kPa. Stroke dynamics
    # control fractional opening x_porv.  PORV discharge enthalpy is selected
    # from the modeled source region: steam-space enthalpy when a pressurizer
    # steam space exists, otherwise bulk RCS enthalpy.
    # Set PORV_area_m2 = 0 (default) to disable entirely.
    PORV_setpoint_kPa = float(get_variable(local_namespace, "PORV_setpoint_kPa", 22064.0))  # critical point → disabled by default
    PORV_reseat_kPa   = float(get_variable(local_namespace, "PORV_reseat_kPa",
                              PORV_setpoint_kPa - 500.0))   # 500 kPa deadband
    PORV_area_m2      = float(get_variable(local_namespace, "PORV_area_m2",   0.0))
    PORV_Cd           = float(get_variable(local_namespace, "PORV_Cd",        0.84))
    PORV_tau_open     = float(get_variable(local_namespace, "PORV_tau_open",  0.5))   # s to full open
    PORV_tau_close    = float(get_variable(local_namespace, "PORV_tau_close", 1.0))   # s to full close
    # PORV numerical/physical safeguards.  The inventory cap prevents the
    # single-node RCS solve from removing an excessive fraction of system mass
    # in one time step.  The steam margin avoids treating a completely filled
    # pressurizer as a steam source.
    PORV_max_frac_mass_per_step = float(get_variable(local_namespace, "PORV_max_frac_mass_per_step", 0.005))
    PORV_steam_margin           = float(get_variable(local_namespace, "PORV_steam_margin", 0.05))
    PORV_min_liquid_level       = float(get_variable(local_namespace, "PORV_min_liquid_level", 0.05))
    PORV_max_frac_mass_per_step = max(0.001, min(0.25, PORV_max_frac_mass_per_step))
    PORV_steam_margin           = max(0.0, min(0.25, PORV_steam_margin))
    PORV_min_liquid_level       = max(0.0, min(0.50, PORV_min_liquid_level))
    # P_relax: under-relaxation applied only when PORV is active (PORV_area_m2 > 0).
    # P_relax_dP_ref: reference rate [kPa/s]; damping increases above this.
    P_relax        = float(get_variable(local_namespace, "P_relax",        0.5))
    P_relax        = max(0.05, min(1.0, P_relax))
    P_relax_dP_ref = float(get_variable(local_namespace, "P_relax_dP_ref", 500.0))

    # ── Normal pressurizer pressure-control surrogate ────────────────────────
    # Optional startup/power-ascension pressure-control model.  This is separate
    # from the PORV accident relief model and is intended to represent normal
    # pressurizer pressure control (spray/heater/charging-letdown aggregate)
    # without explicitly modeling those systems.  Default is disabled so accident
    # cases retain the existing pressure response unless the user enables it.
    pzr_control_enabled      = int(get_variable(local_namespace, "pzr_control_enabled",      0))
    pzr_control_setpoint_kPa = float(get_variable(local_namespace, "pzr_control_setpoint_kPa", 15500.0))
    pzr_control_band_kPa     = float(get_variable(local_namespace, "pzr_control_band_kPa",      100.0))
    pzr_control_tau_s        = float(get_variable(local_namespace, "pzr_control_tau_s",          30.0))
    # Optional explicit heater/spray capacity limits for the normal pressure
    # controller.  Defaults of 0.0 preserve the older pressure-relaxation-only
    # behavior unless the user specifies capacities.
    pzr_control_heater_MW    = float(get_variable(local_namespace, "pzr_control_heater_MW",       0.0))
    pzr_control_spray_MW     = float(get_variable(local_namespace, "pzr_control_spray_MW",        0.0))
    pzr_control_enabled      = 1 if pzr_control_enabled else 0
    pzr_control_band_kPa     = max(1.0e-9, pzr_control_band_kPa)
    pzr_control_tau_s        = max(1.0e-6, pzr_control_tau_s)
    pzr_control_heater_MW    = max(0.0, pzr_control_heater_MW)
    pzr_control_spray_MW     = max(0.0, pzr_control_spray_MW)
    _pzr_control_Q_W         = 0.0

    # ── Pressurizer/RCS compliance surrogate ────────────────────────────────
    # Optional normal-transient pressure slew limiter. This represents the
    # cushioning effect of pressurizer steam space, surge flow, saturation
    # buffering, fluid compressibility, and structural compliance that are not
    # explicitly resolved in FLARE's single-node hydraulic model.
    #
    # Intended use: intact-RCS transients such as boron dilution or load changes.
    # Default is disabled.
    pzr_compliance_enabled       = int(get_variable(local_namespace, "pzr_compliance_enabled", 0))
    pzr_max_depress_rate_kPa_s   = float(get_variable(local_namespace, "pzr_max_depress_rate_kPa_s", 150.0))
    pzr_max_press_rate_kPa_s     = float(get_variable(local_namespace, "pzr_max_press_rate_kPa_s",   300.0))
    pzr_compliance_intact_only   = int(get_variable(local_namespace, "pzr_compliance_intact_only", 1))
    pzr_compliance_enabled       = 1 if pzr_compliance_enabled else 0
    pzr_compliance_intact_only   = 1 if pzr_compliance_intact_only else 0
    pzr_max_depress_rate_kPa_s   = max(0.0, pzr_max_depress_rate_kPa_s)
    pzr_max_press_rate_kPa_s     = max(0.0, pzr_max_press_rate_kPa_s)

    # ── RCS enthalpy relaxation / thermal inertia surrogate ─────────────────
    # Optional intact-transient smoothing of the 2x2 solve enthalpy update.
    # This represents omitted distributed thermal inertia in metal structures,
    # pressurizer/surge volume, and nonuniform primary inventory.  It should not
    # be used for LOCA blowdown unless intentionally exploring a sensitivity.
    rcs_enthalpy_relax_tau_s       = float(get_variable(local_namespace, "rcs_enthalpy_relax_tau_s", 0.0))
    rcs_enthalpy_relax_intact_only = int(get_variable(local_namespace, "rcs_enthalpy_relax_intact_only", 1))
    rcs_enthalpy_relax_tau_s       = max(0.0, rcs_enthalpy_relax_tau_s)
    rcs_enthalpy_relax_intact_only = 1 if rcs_enthalpy_relax_intact_only else 0

    # ── Reactor trip setpoints ────────────────────────────────────────────────
    # All five setpoints are evaluated each timestep when full_power_flag=1 AND
    # the reactor has not yet scrammed AND not in the pre-break phase.
    # A value of 0 disables that setpoint (trip_time_s and trip_power_frac excepted).
    # NOTE: trip_P_hi_kPa and PORV_setpoint_kPa are INDEPENDENT parameters.
    #   trip_P_hi_kPa   : high-pressure scram [kPa]       default 0 (disabled)
    #   trip_P_lo_kPa   : low-pressure scram [kPa]        default 0.9 x rated pressure
    #   trip_power_frac : high-power scram, fraction rated default 1.10 (110%)
    #   trip_time_s     : manual trip at simulation time   default 1e6 s (never)
    #   trip_flow_frac  : low-flow scram, fraction rated   default 0.0 (disabled)
    trip_P_lo_kPa   = float(get_variable(local_namespace, "trip_P_lo_kPa",
                            0.9 * pressure_kPa))
    trip_P_hi_kPa   = float(get_variable(local_namespace, "trip_P_hi_kPa", 0.0))
    trip_power_frac = float(get_variable(local_namespace, "trip_power_frac", 1.10))
    trip_time_s     = float(get_variable(local_namespace, "trip_time_s",     1e6))
    trip_flow_frac  = float(get_variable(local_namespace, "trip_flow_frac",  0.0))
    # trip_delay: signal processing and relay delay [s] applied to every RPS trip
    # before the scram actually fires.  A value of 1.5 s is representative of
    # conventional PWR RPS with hard-wired relays.  Set to 0 for instantaneous
    # scram (original behaviour).  The delay is applied to all five setpoints,
    # the internal PK high-power check, and the PORV-triggered scram.
    trip_delay = float(get_variable(local_namespace, "trip_delay", 1.5))
    # Turbine under-speed trip (ATWS detection) ───────────────────────────────
    # trip_turbine_underspeed_frac: pump speed fraction below which turbine trips.
    #   Default 0.0 = disabled (preserves startup, coastdown, low-power cases).
    #   Typical ATWS value: 0.95 (trip at 95% of rated speed).
    # trip_turbine_delay_s: time from turbine trip detection to diverse reactor
    #   scram (proxy for SLCS actuation). Default 10.0 s per ATWS design basis.
    trip_turbine_underspeed_frac = float(get_variable(local_namespace,
                                         "trip_turbine_underspeed_frac", 0.0))
    trip_turbine_delay_s         = float(get_variable(local_namespace,
                                         "trip_turbine_delay_s", 10.0))
    # trip_sgiv_stroke_s: SG isolation valve stroke time from trip/AMSAC signal
    # to full closure. Default 30s (typical MOV stroke time).
    trip_sgiv_stroke_s           = float(get_variable(local_namespace,
                                         "trip_sgiv_stroke_s", 30.0))
    # sg_trip_isolation_enabled:
    #   1 = isolate the forced-flow SG heat sink on reactor scram/trip,
    #       while retaining the AFW/natural-circulation SG cooling model.
    #   0 = preserve older behavior; SG forced heat removal remains available
    #       after reactor trip unless otherwise disabled.
    sg_trip_isolation_enabled    = int(get_variable(local_namespace,
                                         "sg_trip_isolation_enabled", 1))
    # SLCS boron injection model ───────────────────────────────────────────────
    # trip_slcs_transport_s: time from SLCS actuation to first boron reaching
    #   the core (transport delay through cold leg and reactor vessel).
    #   Default 240s — conservative for SMR integral RCS.
    # trip_slcs_worth_pcm: total negative reactivity worth of the SLCS boron
    #   injection. Default -5000 pcm (sufficient for hot shutdown with margin).
    # trip_slcs_ramp_s: duration over which boron worth ramps from 0 to full
    #   worth once transport delay has elapsed. Default 300s.
    trip_slcs_transport_s = float(get_variable(local_namespace,
                                  "trip_slcs_transport_s", 240.0))
    trip_slcs_worth_pcm   = float(get_variable(local_namespace,
                                  "trip_slcs_worth_pcm",  -5000.0))
    trip_slcs_ramp_s      = float(get_variable(local_namespace,
                                  "trip_slcs_ramp_s",     300.0))
    # Accumulator boron model ─────────────────────────────────────────────────
    # acc_boron_ppm:      boron concentration in accumulator water [ppm].
    #                     Default 2300 ppm (typical PWR accumulator).
    # si_boron_ppm:       boron concentration in pumped SI water (HPSI/LPSI)
    #                     drawn from RWST. Default 2300 ppm (same as accumulator).
    #                     Set to 0 to simulate an unborated SI source (dilution).
    # alpha_boron_pcm_ppm: differential boron worth [pcm/ppm].
    #                     Default -10 pcm/ppm (typical PWR at EOC).
    # rcs_boron_ppm_init: initial RCS boron concentration [ppm].
    #                     Default 0 (EOC, all boron depleted).
    acc_boron_ppm        = float(get_variable(local_namespace,
                                 "acc_boron_ppm",         2300.0))
    si_boron_ppm         = float(get_variable(local_namespace,
                                 "si_boron_ppm",          2300.0))
    alpha_boron_pcm_ppm  = float(get_variable(local_namespace,
                                 "alpha_boron_pcm_ppm",    -10.0))
    rcs_boron_ppm_init   = float(get_variable(local_namespace,
                                 "rcs_boron_ppm_init",       0.0))
    # full_power_flag = 1 : hold power at rated; ALL automatic trips inhibited.
    #   Use for pure SS or ATWS cases.  scram_on_PORV can still override.
    # t_break [s]         : time at which the RCS pressure boundary break opens.
    #   At t_break the break opens; reactor scram and pump trip also fire unless
    #   full_power_flag=1.  t_break=0: break open from t=0.
    # scram_on_PORV   = 1 : fire scram the first time PORV flow > 0 (legacy;
    #   equivalent to setting trip_P_hi_kPa = PORV_setpoint_kPa).
    # pump_trip_time: time [s] at which the RCP loses power independently of scram.
    # Default None = pump trips coincident with reactor scram (normal behavior).
    # Set to a positive value to simulate a pump trip independent of the scram,
    # e.g. loss of power to RCP as a non-LOCA initiating event.
    _pump_trip_time_raw = get_variable(local_namespace, "pump_trip_time", None)
    pump_trip_time = float(_pump_trip_time_raw) if _pump_trip_time_raw is not None else None
    t_break         = float(get_variable(local_namespace, "t_break",       0.0))
    scram_on_PORV   = int(get_variable(local_namespace, "scram_on_PORV",   0))

    # ── Reactor trip setpoints ────────────────────────────────────────────────
    # All five setpoints are evaluated each timestep when full_power_flag=0 and
    # the reactor has not yet scrammed.  Defaults are set AFTER pressure_kPa and
    # PORV_setpoint_kPa are known so they can reference those values.

    # ── PATCH 2: RCP parameters ──────────────────────────────────────────────
    # User-facing pump controls.  pump_oriface_area is accepted as a permissive
    # alias for the misspelled spelling sometimes used in notes/input decks.
    _pump_orifice_default = get_variable(local_namespace, "pump_oriface_area", 0.46)
    pump_orifice_area = float(get_variable(local_namespace, "pump_orifice_area", _pump_orifice_default))
    pump_speed_rpm    = float(get_variable(local_namespace, "pump_speed_rpm",    3550.0))
    pump_D_h          = float(get_variable(local_namespace, "pump_D_h",          0.012))
    pump_L_heat       = float(get_variable(local_namespace, "pump_L_heat",       1.6))
    # Exponential coastdown time constant [s].  Typical PWR RCP: 25-40 s.
    pump_tau          = float(get_variable(local_namespace, "pump_tau",    30.0))
    pump_flag         = int(get_variable(local_namespace, "pump_flag",     1))

    # Pump scaling constants.  These were formerly hard-coded inside the pump
    # module.  Defaults preserve the original model, but user-facing inputs are
    # now SI/metric; configure_pump_model() converts them to the internal basis.
    pump_rated_speed_rpm      = float(get_variable(local_namespace, "pump_rated_speed_rpm",      _DEFAULT_PUMP_RATED_SPEED_RPM))
    pump_rated_flow_m3s       = float(get_variable(local_namespace, "pump_rated_flow_m3s",       _DEFAULT_PUMP_RATED_FLOW_M3S))
    pump_rated_head_m         = float(get_variable(local_namespace, "pump_rated_head_m",         _DEFAULT_PUMP_RATED_HEAD_M))
    pump_rated_torque_Nm      = float(get_variable(local_namespace, "pump_rated_torque_Nm",      _DEFAULT_PUMP_RATED_TORQUE_NM))
    pump_rated_density_kg_m3  = float(get_variable(local_namespace, "pump_rated_density_kg_m3",  _DEFAULT_PUMP_RATED_DENSITY_KG_M3))
    pump_inertia_kg_m2        = float(get_variable(local_namespace, "pump_inertia_kg_m2",        _DEFAULT_PUMP_INERTIA_KG_M2))
    pump_flow_area_m2         = float(get_variable(local_namespace, "pump_flow_area_m2",         _DEFAULT_PUMP_FLOW_AREA_M2))
    pump_friction_torque_Nm   = float(get_variable(local_namespace, "pump_friction_torque_Nm",   _DEFAULT_PUMP_FRICTION_TORQUE_NM))
    pump_friction_speed2_Nm   = float(get_variable(local_namespace, "pump_friction_speed2_Nm",   _DEFAULT_PUMP_FRICTION_SPEED2_NM))

    pump_model_config = configure_pump_model(
        rated_speed_rpm=pump_rated_speed_rpm,
        rated_flow_m3s=pump_rated_flow_m3s,
        rated_head_m=pump_rated_head_m,
        rated_torque_Nm=pump_rated_torque_Nm,
        rated_density_kg_m3=pump_rated_density_kg_m3,
        inertia_kg_m2=pump_inertia_kg_m2,
        flow_area_m2=pump_flow_area_m2,
        friction_torque_Nm=pump_friction_torque_Nm,
        friction_speed2_Nm=pump_friction_speed2_Nm,
    )
    print("Pump model scale: "
          f"rated_speed={pump_model_config['pump_rated_speed_rpm']:.1f} rpm, "
          f"rated_flow={pump_model_config['pump_rated_flow_m3s']:.4f} m3/s, "
          f"rated_head={pump_model_config['pump_rated_head_m']:.2f} m")

    total_power = get_variable(local_namespace, "total_power", 575)  # MW

    # ── PATCH 6: SG parameters ───────────────────────────────────────────────
    P_sec_kPa     = get_variable(local_namespace, "P_sec_kPa",   6895.0)  # kPa
    sg_flag       = int(get_variable(local_namespace, "sg_flag",  1))
    # sg_table_flag: when 1, column D of the time table is used as a prescribed
    # Q_sg boundary condition [MW vs time], bypassing the UA model entirely.
    # Positive = SG removes heat from primary (normal); negative = SG adds heat
    # to primary (overcooling reversal, e.g. MSLB).  sg_flag must also be 1.
    sg_table_flag = int(get_variable(local_namespace, "sg_table_flag", 0))
    # T_sec_K       = sg_secondary_temperature(P_sec_kPa)
    T_sec_K       = get_variable(local_namespace, "T_sec_K", sg_secondary_temperature(P_sec_kPa))  # K
    # UA_sg_rated sizing modes:
    #   sg_dynamic_ua_flag = 0 (default): preserve the normal rated-condition
    #       sizing based on power_target and thot_target. This is the production
    #       behavior and is appropriate for startup cases where the initial RCS
    #       temperature/power may not represent full-power SG design conditions.
    #   sg_dynamic_ua_flag = 1: diagnostic-only sizing from the instantaneous
    #       current heat balance:
    #           UA_sg_rated = RK Total Power / (RCS_Temperature - T_sec_K)
    #       where RK Total Power is the current timestep reactor-power term [W]
    #       reported in the normal output as "RK Total Power (MW)".
    #
    # A user-supplied UA_sg_rated always takes precedence over either default.
    sg_dynamic_ua_flag = int(get_variable(local_namespace, "sg_dynamic_ua_flag", 0))

    _power_target    = float(get_variable(local_namespace, "power_target", total_power))
    _t_hot           = float(get_variable(local_namespace, "thot_target",  temp_core_exit))
    _T_primary_rated = _t_hot + 273.15  # K
    _dT_rated        = max(_T_primary_rated - T_sec_K, 1.0)  # avoid div/0
    _UA_sg_default   = (_power_target * 1e6) / _dT_rated       # W/K

    _RCS_Temperature_ref_K = float(temp_core_exit) + 273.15
    _dT_sg_dynamic = _RCS_Temperature_ref_K - float(T_sec_K)
    if not np.isfinite(_dT_sg_dynamic) or _dT_sg_dynamic <= 0.0:
        print("WARNING: SG diagnostic dynamic-UA reference ΔT is nonpositive; "
              "using 1.0 K to avoid division by zero.")
        _dT_sg_dynamic = 1.0
    _UA_sg_dynamic = (float(total_power) * 1e6) / _dT_sg_dynamic  # initial W/K candidate

    _UA_sg_selected_default = _UA_sg_dynamic if sg_dynamic_ua_flag else _UA_sg_default
    _UA_sg_user_supplied = "UA_sg_rated" in local_namespace
    UA_sg_rated = float(get_variable(local_namespace, "UA_sg_rated", _UA_sg_selected_default))
    sg_nat_frac = float(get_variable(local_namespace, "sg_nat_frac", 0.05))  # natural-circulation UA fraction

    def _compute_sg_dynamic_ua(_T_primary_K: float, _rk_power_W: float) -> tuple[float, float]:
        """Return diagnostic SG UA [W/K] and primary-secondary ΔT [K].

        Diagnostic definition:
            UA_sg_rated = RK Total Power [W] / (RCS Temperature - T_sec_K) [K]

        The same value is optionally used by the SG model when
        sg_dynamic_ua_flag=1 and UA_sg_rated was not supplied explicitly.
        """
        _dT = float(_T_primary_K) - float(T_sec_K)
        _P = float(_rk_power_W)
        if (not np.isfinite(_dT)) or _dT <= 0.0 or (not np.isfinite(_P)):
            return np.nan, _dT
        return max(_P, 0.0) / _dT, _dT

    if sg_dynamic_ua_flag and _UA_sg_user_supplied:
        print("SG: sg_dynamic_ua_flag=1 but UA_sg_rated was supplied explicitly; "
              "the supplied UA is used for the SG model and the dynamic UA is "
              "reported in the diagnostics CSV only.")

    if sg_dynamic_ua_flag:
        print(f"SG: UA_rated={UA_sg_rated/1e6:.2f} MW/K  "
              f"(diagnostic dynamic-UA enabled, ΔT_init={_dT_sg_dynamic:.1f} K "
              f"from RCS_T={_RCS_Temperature_ref_K-273.15:.1f} C), "
              f"P_sec={P_sec_kPa:.0f} kPa, T_sec={T_sec_K-273.15:.1f} C")
    else:
        print(f"SG: UA_rated={UA_sg_rated/1e6:.2f} MW/K  "
              f"(ΔT_rated={_dT_rated:.1f} K), "
              f"P_sec={P_sec_kPa:.0f} kPa, T_sec={T_sec_K-273.15:.1f} C")

    # ── decay heat / power ───────────────────────────────────────────────────
    Tinf        = get_variable(local_namespace, "Tinf",        365*24*3600)
    U239yield   = get_variable(local_namespace, "U239yield",   1.0)
    fr35        = get_variable(local_namespace, "fr35",        1.0)
    fr38        = get_variable(local_namespace, "fr38",        0.0)
    fr39        = get_variable(local_namespace, "fr39",        0.0)
    Efis        = get_variable(local_namespace, "Efis",        200)
    stdd        = 1979

    # ── heat conduction ──────────────────────────────────────────────────────
    L          = get_variable(local_namespace, "L",          0.1)
    lambda_ss  = get_variable(local_namespace, "lambda_ss",  13.25)
    rho_ss     = get_variable(local_namespace, "rho_ss",     8000)
    cp_ss      = get_variable(local_namespace, "cp_ss",      500)
    Twall_ini  = get_variable(local_namespace, "Twall_ini",  323.15)
    N          = get_variable(local_namespace, "N",          10)
    DELTA_x    = L / N
    x          = np.linspace(0, L, N+1)
    cond_flag  = get_variable(local_namespace, "cond_flag",  0)
    SurfArea   = get_variable(local_namespace, "SurfArea",   10.0)
    hflag      = get_variable(local_namespace, "hflag",      0)
    R5_plot        = get_variable(local_namespace, "R5_plot",        1)

    # ── Core heat transfer model (core_flag = 1) ─────────────────────────────
    # Slab = fuel cladding; fission+decay heat in at inner surface, convection out.
    # Default geometry: typPWR 575 MWt — 18200 Zircaloy-clad pins.
    core_flag    = int(get_variable(local_namespace, "core_flag",    1))
    N_pins       = get_variable(local_namespace, "N_pins",       18200)
    d_pin        = get_variable(local_namespace, "d_pin",        0.0095)  # [m] pin OD
    pitch_pin    = get_variable(local_namespace, "pitch_pin",    0.0126)  # [m] square pitch
    L_heated     = get_variable(local_namespace, "L_heated",     2.4)     # [m]
    delta_clad   = get_variable(local_namespace, "delta_clad",   0.00057) # [m]
    lambda_clad  = get_variable(local_namespace, "lambda_clad",  13.0)    # [W/m-K] Zirc-4
    rho_clad     = get_variable(local_namespace, "rho_clad",     6500.0)  # [kg/m3]
    cp_clad      = get_variable(local_namespace, "cp_clad",      330.0)   # [J/kg-K]
    core_nc_frac = get_variable(local_namespace, "core_nc_frac", 0.02)    # nat-conv threshold
    # HTC under-relaxation factor.  At each step:
    #   alpha_new = blend of D-B and NatConv across a transition band
    #   alpha[t+1] = (1-htc_relax)*alpha[t] + htc_relax*alpha_new
    # Default 0.3 damps correlation-switching oscillations without hiding
    # the underlying physics.  Set to 1.0 to disable relaxation.
    htc_relax = float(get_variable(local_namespace, "htc_relax", 0.3))
    # htc_core_mult: multiplicative bias factor on the core forced-convection HTC.
    # Applied to _htc_target (D-B / Churchill-Chu) before under-relaxation.
    # Does NOT affect the post-CHF Bromley or steam D-B regimes, so uncertainty
    # in nucleate/forced-conv HTC is varied independently of the film-boiling model.
    # Default 1.0 = no adjustment.  UA range typically 0.8-1.2 (+/-20%).
    htc_core_mult = float(get_variable(local_namespace, "htc_core_mult", 1.0))
    # htc_fb_mult: multiplicative bias factor on the film boiling HTC.
    # Applied to both the Bromley (IAFB) result and the steam Dittus-Boelter
    # (reflood / uncovery) result before the htc_post_chf cap is enforced.
    # Default 1.0 = no adjustment.  UA range typically 0.5-2.0 given the
    # large scatter in film boiling correlations (~factor of 2).
    htc_fb_mult = float(get_variable(local_namespace, "htc_fb_mult", 1.0))

    # ── Pumped Safety Injection (HPSI / LPSI) model ───────────────────────────
    # Implements the abbreviated HPSI/LPSI model from the FLARE SI model spec.
    # SI signal: when RCS pressure falls below p_si_setpoint_Pa.
    # Flow: normalized pump-head relation  mdot_j = a_j(t) * mdot_j,r
    #                                             * max(0, P_so,j - P_RCS)
    #                                             /     (P_so,j - P_r,j)
    # a_j(t): ramp from 0→1 over si_ramp_time after actuation delay.
    # RWST inventory depletes; flow goes to zero when RWST is empty.
    # Default rated flows are 0 kg/s (unavailable) — set in input deck.
    si_enabled          = bool(get_variable(local_namespace, "si_enabled",          True))
    p_si_setpoint_Pa    = float(get_variable(local_namespace, "p_si_setpoint_Pa",   12.5e6))
    hpsi_mdot_rated     = float(get_variable(local_namespace, "hpsi_mdot_rated",    0.0))
    hpsi_p_shutoff_Pa   = float(get_variable(local_namespace, "hpsi_p_shutoff_Pa",  15.5e6))
    hpsi_p_rated_Pa     = float(get_variable(local_namespace, "hpsi_p_rated_Pa",    7.0e6))
    hpsi_delay_s        = float(get_variable(local_namespace, "hpsi_delay_s",       20.0))
    lpsi_mdot_rated     = float(get_variable(local_namespace, "lpsi_mdot_rated",    0.0))
    lpsi_p_shutoff_Pa   = float(get_variable(local_namespace, "lpsi_p_shutoff_Pa",  1.5e6))
    lpsi_p_rated_Pa     = float(get_variable(local_namespace, "lpsi_p_rated_Pa",    0.3e6))
    lpsi_delay_s        = float(get_variable(local_namespace, "lpsi_delay_s",       30.0))
    si_ramp_time_s      = float(get_variable(local_namespace, "si_ramp_time_s",     5.0))
    rwst_mass_initial   = float(get_variable(local_namespace, "rwst_mass_initial",  1.5e6))
    # Injection temperature: default to accumulator temperature
    t_inj_K             = float(get_variable(local_namespace, "t_inj_K",
                                             inflow_info.get("temp", 49.85) + 273.15))
    # Maximum RCS mass as a fraction of initial inventory before SI is suppressed.
    # Default 1.60 = rhoL(~98°C, 1 atm) / rhoL(~343°C, 148 bar) ≈ 960/600:
    # the maximum mass the vessel can physically hold when filled with cold
    # near-atmospheric water relative to the initial hot pressurised inventory.
    si_mass_ceiling     = float(get_variable(local_namespace, "si_mass_ceiling",     1.60))
    # si_fill_threshold: normalised vessel fill fraction above which SI pump
    # flow begins to degrade.  Default 0.5 (LOCA reflood — full flow until
    # vessel is half full, then linear taper to zero at full).  Set to 1.0
    # to disable degradation (e.g. makeup/letdown dilution scenario where the
    # SI table represents normal-operation makeup flow in an intact RCS).
    si_fill_threshold   = float(get_variable(local_namespace, "si_fill_threshold",   0.50))
    # Chemical and Volume Control System (CVCS) makeup/letdown flow [kg/s].
    # Represents the steady-state balanced makeup/letdown loop:
    #   - Makeup injects cvcs_kgs at si_boron_ppm concentration
    #   - Letdown removes cvcs_kgs at the current RCS boron concentration
    # Net mass effect = 0.  Net boron effect = cvcs_kgs × (si_boron_ppm - C_RCS).
    # Column E (cvcs_mdot) is an adder/subtracter on top of this baseline:
    #   positive = extra makeup (adds mass and boron at si_boron_ppm)
    #   negative = extra letdown (removes mass and boron at RCS concentration)
    cvcs_kgs            = float(get_variable(local_namespace, "cvcs_kgs",            0.0))
    # cvcs_boron_ppm: boron concentration of CVCS makeup water [ppm].
    # Default 0 (unborated — worst-case dilution scenario).
    # Set to match RCS concentration for a neutral (no dilution) baseline.
    cvcs_boron_ppm      = float(get_variable(local_namespace, "cvcs_boron_ppm",      0.0))
    # cvcs_start_time_s: time at which the CVCS makeup/letdown loop activates [s].
    # Before this time cvcs_kgs has no effect (boron concentration unchanged).
    # Default 0 (active from the start of the simulation).
    cvcs_start_time_s   = float(get_variable(local_namespace, "cvcs_start_time_s",   0.0))

    # ── DNBR parameters (requires core_flag = 1) ─────────────────────────────
    # DNBR is computed at every timestep when core_flag=1 (Biasi/Zuber CHF)
    # F_r           : radial hot-pin peaking factor
    # F_z           : axial peaking factor (peak/average axial flux)
    #                 q"_hot = F_r × F_z × q"_avg
    #                 Typical F_z ≈ 1.55 (chopped cosine, fresh core).
    #                 Default 1.0 preserves backward compatibility.
    # T_fuel_limit_C: simulation terminates when the average outer cladding
    # surface temperature (TTwall) exceeds this value and core_flag=1.
    # 1500°C represents severe cladding damage (Zircaloy oxidation / clad breach).
    # The average outer clad is the direct slab-solve result; it is the
    # appropriate termination variable for DNBR/film-boiling scenarios.
    T_fuel_limit_C = float(get_variable(local_namespace,
                                         "T_fuel_limit_C", 1500.0))
    # RG 1.183 source-term failure thresholds:
    # T_gap_release_C: cladding rupture temperature triggering gap release.
    #   ~800 C for Zircaloy under internal fission-gas pressure.
    # T_early_iv_C: 10 CFR 50.46 PCT limit (1204 C) — rods above this
    #   are counted as Early In-Vessel failures for source term purposes.
    T_gap_release_C = float(get_variable(local_namespace, "T_gap_release_C",  800.0))
    T_early_iv_C    = float(get_variable(local_namespace, "T_early_iv_C",    1204.0))

    # Fuel-damage source-term model.
    #
    # thermal_failure:
    #     Existing FLARE best-estimate model. RG 1.183 group release fractions
    #     are scaled by the thermally predicted failed-rod fractions.
    #
    # licensing_auto:
    #     Reduced-order licensing-style model:
    #       LOCA     -> full RG 1.183 gap + early-in-vessel fractions
    #       non-LOCA -> full RG 1.183 gap fractions only when fuel damage is predicted
    #
    # RG1183_LOCA:
    #     Force full LOCA gap + early-in-vessel fractions.
    #
    # RG1183_nonLOCA:
    #     Force full non-LOCA fuel-damage gap fractions; no early-in-vessel release.

    # RG1183_nonLOCA_DNB:
    #     Non-LOCA DNB/fuel-failure source term using DNB failed-fuel fraction.
    #
    # RG1183_nonLOCA_DNB:
    #     Non-LOCA DNB/fuel-failure source term: gap release scaled by
    #     the DNBR-predicted failed-fuel fraction; no early-in-vessel release.
    #
    # The iodine spike / pre-existing coolant activity model is separate and is
    # not disabled or replaced by this option.
    source_term_model = str(get_variable(local_namespace, "source_term_model",
                                          "thermal_failure")).strip("'\"").lower()

    # Severe-event source-term augmentation.  When predicted mean oxidizing-rod
    # ECR exceeds the threshold, FLARE interpolates each group release fraction
    # from the base gap + early-in-vessel release toward the group-specific
    # BDBE target factor.  ECR is used as a fraction; e.g., 20% ECR -> 0.20.
    bdbE_ecr_release_flag = int(get_variable(local_namespace,
                                             "bdbE_ecr_release_flag", 1))
    bdbE_ecr_threshold_pct = float(get_variable(local_namespace,
                                                "bdbE_ecr_threshold_pct", 17.0))

    post_chf_model = str(get_variable(local_namespace, "post_chf_model", "bromley")).strip("'\"")
    # htc_post_chf: effective post-CHF HTC cap [W/m²K] used in the Bromley
    # film boiling correlation when DNBR < 1.  Bromley overestimates at high
    # pressure (P > 16 MPa) due to anomalously large steam k_v near the
    # critical point.  This cap represents the actual IAFB regime HTC from
    # empirical data in high-pressure PWR rod bundles.  Default 450 W/m²K.
    htc_post_chf = float(get_variable(local_namespace, "htc_post_chf", 1e6))
    # Cold-leg LOCA: cap Bromley at 50 W/m²K — models immediate film boiling
    # at flow stagnation point.  For all other models htc_post_chf is
    # effectively uncapped (default 1e6) unless set explicitly in input.
    if post_chf_model == "cold_leg":
        htc_post_chf = 50.0
    # post_chf_model: selects post-CHF heat transfer regime.
    #   'bromley'  (default) — Bromley inverted-annular film boiling throughout.
    #   'cold_leg' — Bromley during blowdown/falling level; switches to steam
    #                Dittus-Boelter once reflood is detected (level rising).
    #   'hot_leg'  — steam Dittus-Boelter throughout (no Bromley).
    # steam_velocity: fixed steam velocity [m/s] for core_htc_db_steam.
    # Represents ECCS-driven upflow or natural-circulation boil-off,
    # independent of pump coastdown.  At 0.5 m/s the steam HTC is
    # consistently at or below Bromley across the reflood pressure range,
    # making steam cooling the more conservative post-CHF choice.
    steam_velocity = float(get_variable(local_namespace, "steam_velocity", 0.5))
    # Rewetting / film boiling criteria.
    # Two independent triggers activate Bromley film boiling:
    #   (a) DNBR < 1 (onset of CHF, requires core_flag=1)
    #   (b) Core uncovery: ves_ll[t] < film_boiling_level_m
    # Each trigger deactivates independently when its own recovery condition
    # is met AND T_wall < T_sat + rewet_dT_K (Leidenfrost guard).
    #
    #   rewet_dnbr           : DNBR must recover above this to rewet from (a)
    #                          default 2.0.  Set 0 to disable rewetting.
    #   rewet_dT_K           : max T_wall superheat above T_sat for rewet [K]
    #                          default 200 K (applies to BOTH triggers).
    #   film_boiling_level_m : liquid level [m] below which core uncovery
    #                          triggers film boiling (trigger b).
    #                          Default = R5_lowCV_height (steam enters core CV).
    #   film_boiling_level_margin_m : level must recover this many metres above
    #                          film_boiling_level_m before trigger (b) resets.
    #                          Default 0.5 m.
    rewet_dnbr   = float(get_variable(local_namespace, "rewet_dnbr",  2.0))
    rewet_dT_K   = float(get_variable(local_namespace, "rewet_dT_K", 200.0))
    # film_boiling_level_m: set to a positive level [m] to enable the
    # uncovery trigger.  0.0 (default) disables it — ves_ll is always >= 0
    # so the condition (ves_ll < 0) never fires unless explicitly activated.
    # Typical value for a severe LOCA: 5–10 m (when the core is substantially
    # exposed).  The parameter must be set per-case via the input file.
    film_boiling_level_m = float(get_variable(local_namespace,
                                  "film_boiling_level_m", 0.0))
    film_boiling_level_margin_m = float(get_variable(local_namespace,
                                  "film_boiling_level_margin_m", 0.5))
    F_r       = float(get_variable(local_namespace, "F_r",       1.0))
    F_z       = float(get_variable(local_namespace, "F_z",       1.0))
    # k_sigma: number of sigma between core-average (f=1) and hot pin (f=F_r).
    # sigma_r = (F_r-1)/k_sigma. Default 3.0 places F_r at 99.87th percentile
    # of the Gaussian radial power distribution used in failure count estimates.
    k_sigma  = float(get_variable(local_namespace, "k_sigma", 3.0))
    _sigma_r = (F_r - 1.0) / max(k_sigma, 0.1)

    # ── RIA (reactivity insertion accident) model ─────────────────────────────
    # Single-group point kinetics:  dP/dt = (rho_net/l*) * P
    # rho_net = rho_ext(t) + rho_Doppler + rho_moderator
    # rho_ext tabulated in column B of the time/data block (pcm vs time).
    # ria_flag=0 (default): model inactive, power governed by existing logic.
    alpha_D_pcm    = float(get_variable(local_namespace, "alpha_D_pcm",  -3.0))   # pcm/°C fuel
    alpha_M_pcm    = float(get_variable(local_namespace, "alpha_M_pcm",   -35.0))  # pcm/°C coolant temperature
    alpha_M_void   = float(get_variable(local_namespace, "alpha_M_void", -150.0))  # pcm/% void fraction
    beta_eff_pcm   = float(get_variable(local_namespace, "beta_eff_pcm",  650.0)) # pcm
    Lambda_star    = float(get_variable(local_namespace, "Lambda_star",   0.083)) # s (=Λ/β)
    # Scram rod worth [pcm] — negative, representing total shutdown reactivity
    # inserted by control rods after trip.  Default −15000 pcm (conservative PWR).
    trip_rod_worth_pcm = float(get_variable(local_namespace, "trip_rod_worth_pcm", -15000.0))
    # Optional rod-insertion ramp [s].  Default 0 preserves the legacy
    # instantaneous insertion.  A value of ~2 s is useful for normal trip
    # transients where a step insertion overstates the immediate power collapse.
    trip_rod_insertion_time_s = float(get_variable(local_namespace, "trip_rod_insertion_time_s", 0.0))
    trip_rod_insertion_time_s = max(0.0, trip_rod_insertion_time_s)
    # Convert to dollars for fission_scram() — ensures the post-scram fission
    # power decay curve reflects the actual rod worth, not a generic -20$ assumption.
    _rho_scram_dollars = trip_rod_worth_pcm / max(beta_eff_pcm, 1.0)
    k_fuel         = float(get_variable(local_namespace, "k_fuel",        3.0))   # W/m-K UO2
    r_fuel         = float(get_variable(local_namespace, "r_fuel",        0.00411))# m pellet radius
    # h_gap: fuel-clad gap conductance [W/m²K].  Default 6000 W/m²K represents a
    # BOL He-filled gap with modest contact conductance (as-fabricated clearance
    # ~80-100 µm).  This adds ~50-100°C at rated power and scales with power.
    h_gap          = float(get_variable(local_namespace, "h_gap",         6000.0))
    # r_fuel_m: fuel pellet outer radius [m] for gap area calculation.
    # Default = pin outer radius minus clad thickness (zero-gap geometry).
    r_fuel_m       = float(get_variable(local_namespace, "r_fuel_m",
                                         d_pin / 2.0 - delta_clad))
    # UO2 thermal properties for fuel stored-energy time constant.
    # tau_fuel = rho_fuel*cp_fuel*r_fuel_m^2 / (4*k_fuel) ≈ 5 s
    rho_fuel       = float(get_variable(local_namespace, "rho_fuel", 10400.0))
    cp_fuel        = float(get_variable(local_namespace, "cp_fuel",    330.0))
    high_flux_trip = float(get_variable(local_namespace, "high_flux_trip", 1.18)) # fraction rated
    # pja_flag: 1 = apply prompt jump approximation to the PK power at each
    # timestep where rho_net changes.  For rapid reactivity insertions (rod
    # ejection), PJA correctly captures the instantaneous power spike that the
    # Euler inhour step misses.  PJA factor = (β-rho_prev)/(β-rho_curr), valid
    # only when both are sub-critical.  Default 0 (off) preserves existing
    # behaviour for all cases that do not set it.
    pja_flag = int(get_variable(local_namespace, "pja_flag", 0))
    # rho_table_power_target_cutoff_flag:
    #   1 (default): freeze the time/reactivity-table contribution once RK
    #      core power first reaches power_target.  The frozen value is the
    #      table reactivity at the latch time; subsequent positive additions
    #      from the table are not allowed.  If the table later demands less
    #      reactivity than the frozen value, that lower value is honored so
    #      shutdown/hold-down reactivity still works.
    #   0: legacy behavior; use the table exactly as supplied.
    rho_table_power_target_cutoff_flag = int(get_variable(
        local_namespace, "rho_table_power_target_cutoff_flag", 1))
    # Nordheim-Fuchs prompt supercritical model.
    # Lambda_prompt: prompt neutron lifetime [s].  For a PWR ~ 5e-5 s.
    # When rho_net > beta_eff and Lambda_prompt > 0, the Euler step is
    # replaced by NF sub-integration.  Set Lambda_prompt=0 to disable.
    Lambda_prompt  = float(get_variable(local_namespace, "Lambda_prompt", 5e-5))
    # Active fuel height for total fuel heat capacity C_f [m].
    nf_H_active_m  = float(get_variable(local_namespace, "nf_H_active_m", 2.44))
    # Fuel enthalpy failure threshold [cal/g].  200 cal/g is the NRC
    # criterion for fresh fuel; ~150 cal/g for high-burnup irradiated fuel.
    H_fuel_limit_calg = float(get_variable(local_namespace, "H_fuel_limit_calg", 200.0))

    if core_flag:
        SurfArea    = N_pins * np.pi * d_pin * L_heated
        A_flow_core = N_pins * (pitch_pin**2 - np.pi * d_pin**2 / 4.0)
        D_h_core    = 4.0 * (pitch_pin**2 - np.pi * d_pin**2 / 4.0) / (np.pi * d_pin)
        L           = delta_clad
        lambda_ss   = lambda_clad
        rho_ss      = rho_clad
        cp_ss       = cp_clad
        DELTA_x     = L / N
        print(f"Core model: {N_pins} pins, SurfArea={SurfArea:.0f} m2, "
              f"A_flow={A_flow_core:.3f} m2, D_h={D_h_core*1000:.2f} mm, "
              f"clad={delta_clad*1000:.2f} mm")
    else:
        A_flow_core = None
        D_h_core    = None
    # Nordheim-Fuchs: total fuel heat capacity and fuel thermal time constant.
    # Computed outside core_flag so it is always available for PK block.
    _nf_Cf      = (N_pins * np.pi * r_fuel_m**2 * nf_H_active_m * rho_fuel * cp_fuel)
    _tau_nf     = max(rho_fuel * cp_fuel * r_fuel_m**2 / (2.0 * k_fuel), 0.1)
    _m_fuel_pin = np.pi * r_fuel_m**2 * nf_H_active_m * rho_fuel  # kg/pin

    # ── time vector ──────────────────────────────────────────────────────────
    # The 100-point fine grid (dt=0.01 s) is anchored at t_break so the rapid
    # initial depressurization is always resolved with small steps regardless of
    # when the break occurs.  When t_break=0 the structure is identical to the
    # original (fine grid at t=0, coarse grid thereafter).
    if t_break <= 0:
        # Original structure: fine grid 0→0.99, coarse 1→endtime
        time1 = np.linspace(0, 0.99, 100)
        time2 = np.linspace(1, endtime, int((endtime - 1) / timestep + 1))
        time  = np.concatenate((time1, time2))
    else:
        # Pre-break coarse steps (1 s), fine grid at t_break, post-break coarse
        tb = float(t_break)
        _pre   = np.arange(0, tb, timestep) if tb > 0 else np.array([])
        _fine  = np.linspace(tb, tb + 0.99, 100)
        _post_start = tb + 1.0
        _post  = np.linspace(_post_start, endtime,
                              max(1, int((endtime - _post_start) / timestep + 1)))                  if _post_start <= endtime else np.array([])
        time   = np.concatenate([_pre, _fine, _post])
        # Remove any duplicates that arise at join points
        time   = np.unique(np.round(time, 6))
    number_timesteps = len(time)
    Pressure_increment = 1000; enthalpy_increment = 1000

    # ── tabulated inputs (already read in single-pass block above) ──────────
    end_row = 10000

    if sg_table_flag:
        q_in     = _col_b + _col_c          # col D is Q_sg table, not power
        q_sg_xls = _col_d                   # [MW] Q_sg boundary condition
    else:
        q_in     = _col_b + _col_c + _col_d
        q_sg_xls = np.zeros(len(time_xls))  # unused when sg_table_flag=0
    power_input = 1e6 * np.interp(time, time_xls, q_in)

    # Column E: Safety Injection (SI) mass flow rate [kg/s] vs time.
    si_flow_xls = np.where(np.isnan(_col_e), 0.0, _col_e)

    # Column F: rho_ext [pcm] vs time for RIA cases.
    rho_ext_xls = np.where(np.isnan(_col_f), 0.0, _col_f)

    # R5 reference arrays (from single-pass _r5 read above)
    _zeros_t = np.zeros(number_timesteps)
    if _r5 is not None and _r5.shape[0] > 1 and _r5.shape[1] >= 8:
        R5_time               = _r5[:, 0]
        R5_p                  = np.interp(time, R5_time, _r5[:, 1])
        R5_t                  = np.zeros(number_timesteps) + 300
        R5_mdot               = np.interp(time, R5_time, _r5[:, 2])
        R5_accp               = np.interp(time, R5_time, _r5[:, 3])
        R5_acct               = np.interp(time, R5_time, _r5[:, 4]) if _r5.shape[1] > 4 else _zeros_t
        R5_massflow_in        = np.interp(time, R5_time, _r5[:, 5]) if _r5.shape[1] > 5 else _zeros_t
        R5_vessel_mass_scaled = np.interp(time, R5_time, _r5[:, 6]) if _r5.shape[1] > 6 else _zeros_t
        R5_power              = np.interp(time, R5_time, _r5[:, 7]) if _r5.shape[1] > 7 else _zeros_t
        if cond_flag == 1 and _r5.shape[1] > 10:
            R5_pow2fld = np.interp(time, R5_time, _r5[:, 8])
            R5_hstemp  = np.interp(time, R5_time, _r5[:, 9])
            R5_htc     = np.interp(time, R5_time, _r5[:, 10])
    else:
        R5_p = R5_t = R5_mdot = R5_accp = R5_acct = _zeros_t
        R5_massflow_in = R5_vessel_mass_scaled = R5_power = _zeros_t

    # ── output arrays ────────────────────────────────────────────────────────
    Pressure         = np.full(number_timesteps, np.nan)
    # Pressure_report stores the unrelaxed 2x2 predictor pressure for output only.
    # The simulation state continues to use Pressure, including any enthalpy/
    # pressure relaxation applied below.
    Pressure_report  = np.full(number_timesteps, np.nan)
    enthalpy_mix     = np.full(number_timesteps, np.nan)
    Temperature      = np.full(number_timesteps, np.nan)
    Total_Mass_scaled= np.full(number_timesteps, np.nan)
    x_eq             = np.full(number_timesteps, np.nan)
    alpha_void       = np.full(number_timesteps, np.nan)  # volumetric void fraction [-]
    massflow_break   = np.full(number_timesteps, np.nan)
    rkpower_total    = np.full(number_timesteps, np.nan)
    cond_heat        = np.full(number_timesteps, np.nan)
    Total_Mass       = np.zeros(number_timesteps)
    net_heat_total   = np.full(number_timesteps, np.nan)
    energy_outflow_break = np.full(number_timesteps, np.nan)
    h_break_arr          = np.full(number_timesteps, np.nan)  # stagnation enthalpy at break [kJ/kg]
    tmp2             = np.full(number_timesteps, np.nan)
    TT               = np.full((N+1, number_timesteps), np.nan)
    TTwall           = np.full(number_timesteps, np.nan)
    for i in range(N+1): TT[i,0] = Twall_ini
    TTwall[0] = TT[N,0]
    alpha = np.full(number_timesteps, np.nan)
    alpha[0] = 0

    # ── PORV array
    massflow_PORV = np.zeros(number_timesteps)

    # ── pressurizer arrays
    pzr_level_arr      = np.full(number_timesteps, np.nan)  # liquid level [m]
    pzr_level_norm_arr = np.full(number_timesteps, np.nan)  # normalised 0-1
    pzr_mdot_surge_arr = np.full(number_timesteps, np.nan)  # surge flow [kg/s]
    _pzr_hold_disabled = False  # latched True when PORV first fires; hold never re-engages
    _x_porv    = 0.0            # PORV fractional opening [0=closed, 1=fully open]
    _porv_mode = "closed"       # "closed" | "opening" | "open" | "closing"
    _M_porv_loss_cum = 0.0      # cumulative PORV discharge mass [kg]

    # ── RIA arrays
    T_fuel_arr      = np.full(number_timesteps, np.nan)  # avg fuel centrl temp [K]
    T_hot_clad_arr  = np.full(number_timesteps, np.nan)  # hot-pin outer clad [K]
    T_hot_fuel_arr  = np.full(number_timesteps, np.nan)  # hot-pin fuel centreline [K]
    rho_ext_arr   = np.full(number_timesteps, np.nan)    # external reactivity [pcm]
    rho_scram_arr = np.full(number_timesteps, np.nan)    # scram reactivity [pcm]
    rho_boron_arr = np.full(number_timesteps, np.nan)    # SLCS boron reactivity [pcm]
    rcs_boron_mass = np.full(number_timesteps, np.nan)   # RCS boron mass [kg]
    rcs_boron_ppm  = np.full(number_timesteps, np.nan)   # RCS boron concentration [ppm]
    rcs_boron_ppm[0] = rcs_boron_ppm_init               # consistent initial state
    alpha_void[0]    = 0.0                               # no void at t=0 (subcooled)
    rho_D_arr   = np.full(number_timesteps, np.nan)    # Doppler feedback [pcm]
    rho_M_arr   = np.full(number_timesteps, np.nan)    # moderator feedback [pcm]
    rho_net_arr = np.full(number_timesteps, np.nan)    # net reactivity [pcm]
    # _ria_scrammed / _ria_scram_t superseded by _reactor_scrammed / _scram_t
    _ria_scrammed = False   # kept for backward compat; not used in logic below
    _nf_delta_T   = 0.0     # fuel temperature increment from NF burst [K]
    _nf_E_total   = 0.0     # cumulative NF burst energy deposited [J]
    N_fail_DNB    = np.zeros(number_timesteps)  # estimated rods in DNB each step
    N_fail_clad   = np.zeros(number_timesteps)  # estimated rods above PCT limit
    N_fail_gap    = np.zeros(number_timesteps)  # RG1.183: gap release (rupture >800C)
    N_fail_eiv    = np.zeros(number_timesteps)  # RG1.183: early in-vessel (>1200C)

    # ── DNBR arrays ──────────────────────────────────────────────────────────
    DNBR      = np.full(number_timesteps, np.nan)   # DNBR at each timestep
    q_chf     = np.full(number_timesteps, np.nan)   # CHF heat flux [W/m²]
    q_hot     = np.full(number_timesteps, np.nan)   # Hot-pin heat flux [W/m²]

    # ── PATCH 3: pump arrays ─────────────────────────────────────────────────
    pump_omega    = np.zeros(number_timesteps)
    pump_mdot     = np.zeros(number_timesteps)
    pump_velocity = np.zeros(number_timesteps)
    pump_head     = np.zeros(number_timesteps)
    steam_vel_arr = np.zeros(number_timesteps)  # boil-off steam velocity in core [m/s]

    if pump_flag:
        _pump_state = init(pump_orifice_area, pump_speed_rpm,
                                       1e3*pressure_kPa,
                                       temp_core_exit+273.15)
        pump_omega[0]    = _pump_state["omega_rpm"]
        pump_mdot[0]     = _pump_state["mass_flow_kgs"]
        pump_velocity[0] = _pump_state["velocity_ms"]
        pump_head[0]     = _pump_state["head_Pa"]
        print(f"Pump init: {_pump_state['omega_rpm']:.0f} rpm, "
              f"{_pump_state['mass_flow_kgs']:.0f} kg/s, "
              f"velocity {_pump_state['velocity_ms']:.2f} m/s")
    else:
        _pump_state = {"coasted_down": True}

    # ── PATCH 7: SG array ────────────────────────────────────────────────────
    Q_sg = np.zeros(number_timesteps)

    # Separate SG diagnostic arrays written to <case>_diag.csv.  These make the
    # UA sizing path auditable without crowding the main results workbook.
    sg_UA_used        = np.full(number_timesteps, np.nan)  # W/K actually used by model
    sg_UA_dynamic     = np.full(number_timesteps, np.nan)  # W/K from rkpower_total/(T_RCS - T_sec)
    sg_dT_primary_sec = np.full(number_timesteps, np.nan)  # K
    sg_v_ratio        = np.full(number_timesteps, np.nan)  # pump volumetric-flow ratio
    sg_UA_eff         = np.full(number_timesteps, np.nan)  # W/K after flow scaling
    sg_C_primary      = np.full(number_timesteps, np.nan)  # W/K mdot*cp cap
    sg_flow_frac      = np.full(number_timesteps, np.nan)  # explicit flow degradation factor
    sg_open_frac      = np.full(number_timesteps, np.nan)  # SG isolation fraction
    sg_Q_forced_raw   = np.full(number_timesteps, np.nan)  # W before flow/isolation multipliers
    sg_Q_forced       = np.full(number_timesteps, np.nan)  # W forced-flow contribution
    sg_Q_nat          = np.full(number_timesteps, np.nan)  # W natural-circulation floor
    sg_flow_cap_W     = np.full(number_timesteps, np.nan)  # W = (mdot*cp)*primary-secondary dT
    sg_flow_cap_ratio = np.full(number_timesteps, np.nan)  # RK power / SG flow-capacity limit
    sg_flow_cap_warn  = np.zeros(number_timesteps, dtype=int)
    _sg_flow_cap_warned = False
    _sg_flow_cap_warn_count = 0
    sg_nat_active_arr = np.zeros(number_timesteps, dtype=int)
    sg_dynamic_active_arr = np.zeros(number_timesteps, dtype=int)

    # ── scaling / solver bookkeeping ─────────────────────────────────────────
    determine_scaling = {'max_number':4,'number':1,'reach_sat_flag':0,
                         'reach_Mtrans_flag':0,'open_ADV_flag':0}
    scaling_ratios = []; current_scaling_ratios = []
    dilation = np.full(number_timesteps, np.nan)
    pow_rat  = np.full(number_timesteps, np.nan)
    e_rat    = np.full(number_timesteps, np.nan)

    # ── vessel geometry ──────────────────────────────────────────────────────
    ves_delz = R5_lowCV_height + R5_highCV_height
    ves_vol  = ves_delz * R5_area + loop_vol_m3   # RPV + external loop volume
    ves_vol -= pzr_vol   # pressurizer volume excluded from liquid node
    ves_area = R5_area

    init_cond = {'label':'core exit', 'pressure':1e3*pressure_kPa}
    temp_core_exit += 273.15
    init_cond['temperature']   = temp_core_exit
    init_cond['enthalpy_mix']  = 1000*XSteam.h_pT(pressure_kPa, temp_core_exit-273.15)

    Pressure[0]        = init_cond['pressure']
    Pressure_report[0] = init_cond['pressure']
    enthalpy_mix[0]    = init_cond['enthalpy_mix']
    Temperature[0]  = init_cond['temperature']

    # ── Core: set initial clad surface temperature from rated-power film drop ─
    if core_flag:
        _q_flux_rated = 1e6 * total_power / SurfArea   # W/m2 average surface flux
        _alpha_guess  = 40000.0                         # W/m2-K typical PWR forced conv
        Twall_ini = init_cond['temperature'] + _q_flux_rated / _alpha_guess
        for i in range(N+1):
            TT[i, 0] = Twall_ini
        TTwall[0] = TT[N, 0]
        print(f"Core: rated q_flux={_q_flux_rated/1e3:.1f} kW/m2, "
              f"initial clad T={Twall_ini-273.15:.1f} C "
              f"(+{_q_flux_rated/_alpha_guess:.1f} K above coolant)")

    acc_liqvol = np.full(number_timesteps, np.nan)
    acc_wdot   = np.zeros(number_timesteps)
    acc_pres   = np.full(number_timesteps, np.nan)
    acc_tgas   = np.full(number_timesteps, np.nan)
    dp_head    = np.full(number_timesteps, np.nan)
    rel_delz   = np.full(number_timesteps, np.nan)
    ves_ll     = np.full(number_timesteps, np.nan)

    inflow_avg  = 0
    massflow_in = np.zeros(number_timesteps)
    cvcs_mdot        = np.zeros(number_timesteps)   # CVCS column E imbalance flow [kg/s]
    cvcs_makeup_arr  = np.zeros(number_timesteps)   # CVCS makeup flow [kg/s] (positive)
    cvcs_letdown_arr = np.zeros(number_timesteps)   # CVCS letdown flow [kg/s] (positive magnitude)
    # Pumped SI model arrays
    hpsi_mdot_arr    = np.zeros(number_timesteps)  # HPSI flow [kg/s]
    lpsi_mdot_arr    = np.zeros(number_timesteps)  # LPSI flow [kg/s]
    si_pumped_mdot   = np.zeros(number_timesteps)  # HPSI + LPSI total [kg/s]
    # SI signal and RWST state
    _si_signal      = False    # True once RCS pressure < p_si_setpoint_Pa
    _si_signal_t    = None     # time of SI signal
    _rwst_mass      = rwst_mass_initial  # remaining RWST inventory [kg]
    # Pre-compute injection enthalpy (constant T_inj approximation)
    _h_inj          = 1000.0 * XSteam.h_pT(p_si_setpoint_Pa * 1e-3,
                                             t_inj_K - 273.15)
    inflow_info['method']    = 'use_avg'
    inflow_info['mdot_avg']  = inflow_avg
    inflow_info['enthalpy']  = 1000*XSteam.h_pT(inflow_info['pressure'], inflow_info['temp'])
    enthalpy_in = inflow_info['enthalpy']

    pressure_cutoff          = pressure_containment
    mass_scaled_cutoff       = 0.01
    area_ADV_total           = ADV_number * area_per_ADV
    area_break               = 0.25*np.pi*diameter_break**2
    effective_area_break     = area_break
    effective_area_ADV       = area_ADV_total
    transition_mixture_mass_use = transition_mixture_mass
    flag_ADV_open            = 0
    effective_total_area     = effective_area_break

    # ── Fuel thermal time constant ─────────────────────────────────────────
    _tau_fuel = (rho_fuel * cp_fuel * r_fuel_m**2) / (4.0 * k_fuel)

    # ── Initialise alpha[0] from rated D-B HTC ─────────────────────────────
    # Replaces the default zero so t=0 hot-pin estimates are meaningful.
    if core_flag and pump_flag and A_flow_core > 0:
        try:
            _rhoL0 = XSteam.rhoL_p(pressure_kPa)
            _v0    = pump_mdot[0] / max(_rhoL0 * A_flow_core, 1e-9)
            alpha[0] = core_htc_db(pressure_kPa * 1e3,
                                    temp_core_exit + 273.15,
                                    _v0, D_h_core) * htc_core_mult
        except Exception:
            alpha[0] = 40000.0 * htc_core_mult

    print('Input processing complete. Beginning simulation.')

    # ── Unified reactor scram state ──────────────────────────────────────────
    _reactor_scrammed = False          # True once any trip setpoint fires
    _trip_detected_t    = None          # time when first trip setpoint was exceeded
    _trip_reason_pending = None          # trip reason string while delay is pending
    # ── Film boiling flag ────────────────────────────────────────────────────
    _fb_dnbr    = False   # film boiling from DNBR < 1  (resets on DNBR recovery)
    _fb_uncover = False   # film boiling from core uncovery (resets on reflooding)
    _film_boiling_active = False   # _fb_dnbr OR _fb_uncover
    _cc_dominant = False  # True when Churchill-Chu > D-B (natural convection active)
    # Latched SG natural-circulation availability.  This is set after the
    # core HTC block confirms Churchill-Chu dominance following scram, then
    # used by the SG block on subsequent timesteps to avoid order-dependence.
    _sg_nat_active = False
    _reflood_started = False  # latches True when level rises during uncovery

    # Dry-core heat-transfer handoff.  This is distinct from the later
    # dry-core continuation used when the lumped hydraulic model reaches
    # near-empty inventory.  The user-defined level trigger should reduce
    # core heat transfer while hydraulics, accumulator injection, and break
    # flow continue normally.
    dry_core_trigger_level_m = float(get_variable(local_namespace, "dry_core_trigger_level_m", -1.0))
    dry_core_htc_W_m2K       = float(get_variable(local_namespace, "dry_core_htc_W_m2K", 5.0))
    dry_core_emissivity      = float(get_variable(local_namespace, "dry_core_emissivity", 0.8))
    # Approximate UO2 melting-temperature cap used only to prevent nonphysical
    # dry-core temperature runaway in BDB diagnostic continuation.
    _T_melt_cap_K            = 2500.0 + 273.15
    _dry_core_ht_latched     = False
    _scram_t          = None           # simulation time of scram
    _decay_heat_W     = 0.0            # ANS-1979 decay heat at current timestep [W]
    _pump_mdot_rated  = (pump_mdot[0]  # rated flow for low-flow trip
                         if pump_flag and pump_mdot[0] > 0 else None)
    _pump_omega_rated = (pump_speed_rpm   # rated speed for turbine under-speed trip
                         if pump_flag and pump_speed_rpm > 0 else None)

    # ── Hybrid post-scram PK state ───────────────────────────────────────────
    # Effective single-group decay constant derived from the six-group data
    # already used in fission_scram() — yield-weighted average.
    _yld_gam_arr = np.array([[0.038,0.0127],[0.213,0.0317],[0.188,0.115],
                              [0.407,0.311],[0.128,1.4],[0.026,3.87]])
    lambda_eff       = float(np.sum(_yld_gam_arr[:,0] * _yld_gam_arr[:,1])
                             / np.sum(_yld_gam_arr[:,0]))  # ≈ 0.0765 s⁻¹
    _P_excess        = 0.0    # excess fission power above fission_scram baseline [W]
    _C_precursor     = 0.0    # single-group precursor concentration [W]
    _pk_scram_active = False  # latched True when scram fires and hybrid PK begins

    # ── t_break transition tracker ───────────────────────────────────────────
    _prev_pre_break  = (t_break > 0)   # True before break, False after
    _porv_scram_t    = None            # set when PORV-triggered scram fires
    _porv_scram_done     = False
    _clad_limit_warned   = False
    _choked_initial = choked          # save for reset at break opening

    # ══════════════════════════════════════════════════════════════════════════
    # MAIN SIMULATION LOOP
    # ══════════════════════════════════════════════════════════════════════════
    start_time = timer.time()
    counter    = 0

    _last_print_t = -10.0   # track last printed simulation time
    _is_loca      = (diameter_break > 0.0)   # used inside time loop
    _turbine_trip_detected = False            # ATWS: turbine under-speed detected
    _turbine_trip_t        = None             # time of turbine trip detection
    _rho_boron             = 0.0             # SLCS boron reactivity [pcm]
    _rcs_boron_mass        = 0.0   # set correctly at t=0 once Total_Mass is known
    # Startup table-reactivity latch.  Once the RK power first reaches the
    # user-defined power_target, the table contribution is frozen at its current
    # value.  Later lower/negative table values are still honored; later higher
    # positive additions are blocked.
    _rho_table_power_target_latched = False
    _rho_table_power_target_latch_t = None
    _rho_table_power_target_latch_value = None
    for t in range(number_timesteps - 1):

        # ── periodic progress print (parsed by flare_ui.py) ──────────────────
        if time[t] - _last_print_t >= 10.0:
            print(f"SIMTIME {time[t]:.1f}", flush=True)
            _last_print_t = time[t]

        # ── fluid properties ─────────────────────────────────────────────────
        # Clamp pressure and enthalpy to XSteam valid range before lookup
        Pressure[t]     = float(np.clip(Pressure[t],     1e3,    1e8))   # 1 kPa–100 MPa
        enthalpy_mix[t] = float(np.clip(enthalpy_mix[t], 1e3, 4.5e6))   # 1–4500 kJ/kg
        x_eq[t], mix_properties, sat_vapor, sat_liquid = \
            compute_mixture_properties(Pressure[t], enthalpy_mix[t])
        prop_partial_derivatives = compute_property_partials(
            x_eq[t], Pressure[t], enthalpy_mix[t], sat_vapor, sat_liquid,
            Pressure_increment, enthalpy_increment)

        # Volumetric void fraction: α = x·ρL / (x·ρL + (1-x)·ρV)
        _x_v = max(0.0, x_eq[t]) if not np.isnan(x_eq[t]) else 0.0
        if _x_v > 0.0:
            _rhoL_av = 1.0 / max(sat_liquid["specific_volume"], 1e-9)
            _rhoV_av = 1.0 / max(sat_vapor["specific_volume"],  1e-9)
            alpha_void[t] = (_x_v * _rhoL_av /
                             max(_x_v * _rhoL_av + (1.0 - _x_v) * _rhoV_av, 1e-9))
        else:
            alpha_void[t] = 0.0

        # ── SG UA diagnostic preliminary path ───────────────────────────────
        # The final dynamic-UA update is performed after rkpower_total[t] has
        # been computed for this timestep.  Here, only initialize the ΔT
        # diagnostic so it is populated even if an early failure occurs.
        _dT_dyn_t = float(Temperature[t]) - float(T_sec_K)
        sg_dT_primary_sec[t] = _dT_dyn_t

        # ── advance dependent variables ──────────────────────────────────────
        if t == 0:
            # Pressurizer: initialise level and track normalised level
            pzr_level_arr[0]      = pzr_level_init
            pzr_level_norm_arr[0] = (pzr_level_init / pzr_height
                                     if pzr_height > 0 else 0.0)
            pzr_mdot_surge_arr[0] = 0.0
            # Steam volume in pressurizer at t=0 (used in blowdown drain model)
            _pzr_steam_vol = max(0.0, pzr_vol - pzr_level_init * pzr_area)
            # Initial RCS liquid mass: fill ves_vol (pressurizer already excluded)
            if Total_Mass_init_kg > 0:
                Total_Mass = Total_Mass_init_kg
            else:
                Total_Mass = ves_vol / mix_properties["specific_volume"]
            acc_liqvol[t]        = acc_totvol - acc_gasvol
            acc_pres[t]          = 1e3*inflow_info["pressure"]
            acc_tgas[t]          = tgas
            init_cond["Total_Mass_RCS"] = Total_Mass
            Total_Mass_init = Total_Mass          # stored for si_mass_ceiling guard
            # Initialise RCS boron mass now that Total_Mass is known
            _rcs_boron_mass = rcs_boron_ppm_init * 1e-6 * Total_Mass
            Total_Mass_scaled[0] = 1
            rkpower_total[t]     = total_power * 1e6   # rated power at t=0
            cond_heat[t]         = 0
            net_heat_total[t]    = total_power * 1e6   # rated power at t=0
        else:
            timestep = time[t] - time[t-1]
            # Accumulator injection [kg/s]
            massflow_in[t] = acc_wdot[t-1]*XSteam.rho_pT(inflow_info["pressure"],
                                                           inflow_info["temp"])
            # Safety injection from table [kg/s] — column E, interpolated
            cvcs_mdot[t] = float(np.interp(time[t], time_xls, si_flow_xls))
            Total_Mass += (massflow_in[t-1] + cvcs_mdot[t-1] + si_pumped_mdot[t-1]
                           - pzr_mdot_surge_arr[t-1]          # outsurge (−ve) adds to RCS; insurge (+ve) removes
                           - massflow_break[t-1] - massflow_PORV[t-1])*timestep
            Total_Mass_scaled[t] = Total_Mass / init_cond["Total_Mass_RCS"]
            acc_liqvol[t]        = acc_liqvol[t-1] - acc_wdot[t-1]*timestep

        # ── fission + decay heat ─────────────────────────────────────────────
        # t_rel: time since the break/scram event (zero before t_break)
        _pre_break = (t_break > 0) and (time[t] < t_break)
        _break_just_opened = _prev_pre_break and (not _pre_break)
        if _scram_t is not None:
            t_rel = max(0.0, time[t] - _scram_t)
        elif _porv_scram_done and _porv_scram_t is not None:
            t_rel = max(0.0, time[t] - _porv_scram_t)
        elif t_break > 0:
            t_rel = max(0.0, time[t] - t_break)
        else:
            # No break, no scram — reactor is at full power; t_rel is undefined
            t_rel = None

        # ── Fire delayed scram when trip_delay has expired ─────────────────
        if (_trip_detected_t is not None and not _reactor_scrammed
                and time[t] >= _trip_detected_t + trip_delay):
            _reactor_scrammed = True
            _scram_t          = time[t]
            _delay_note = (f"  (trip detected t={_trip_detected_t:.2f} s,"
                           f" delay {trip_delay:.1f} s)")
            if "PORV" in (_trip_reason_pending or ""):
                print(f"PORV scram at t={time[t]:.2f} s,"
                      f"  P={Pressure[t]/1e3:.1f} kPa{_delay_note}")
            else:
                print(f"Reactor scram at t={time[t]:.2f} s"
                      f"  [{_trip_reason_pending}]{_delay_note}")
            # Initialise hybrid post-scram PK precursor to equilibrium at
            # the current fission power.  P_excess starts at zero — the
            # baseline fission_scram curve handles normal decay; the hybrid
            # PK integration only contributes when rho_net rises above
            # -beta_eff (re-criticalization scenario).
            _P_now       = float(rkpower_total[t]) if not np.isnan(rkpower_total[t]) else 1e6 * total_power
            _C_precursor = (beta_eff_pcm * 1e-5) / lambda_eff * _P_now / Lambda_star
            _P_excess    = 0.0
            _pk_scram_active = True
            # t_rel was computed before _scram_t was set at the top of this
            # timestep. Recompute it so any same-step post-scram logic sees
            # the correct start of the insertion ramp.
            t_rel = max(0.0, time[t] - _scram_t)

        # ── Turbine under-speed trip (ATWS detection) ─────────────────────
        if (trip_turbine_underspeed_frac > 0.0
                and pump_flag
                and _pump_omega_rated is not None
                and not _turbine_trip_detected):
            if pump_omega[t] < trip_turbine_underspeed_frac * _pump_omega_rated:
                _turbine_trip_detected = True
                _turbine_trip_t        = time[t]
                print(f"Turbine under-speed trip at t={time[t]:.2f} s"
                      f"  (pump speed {pump_omega[t]:.0f} rpm"
                      f" < {trip_turbine_underspeed_frac*100:.0f}% rated"
                      f" {_pump_omega_rated:.0f} rpm)"
                      f"  → diverse scram in {trip_turbine_delay_s:.1f} s")

        # Diverse scram: fires trip_turbine_delay_s after turbine under-speed
        # Rather than an instantaneous rod scram, model SLCS boron injection:
        #   - Transport delay before boron reaches core
        #   - Linear ramp of negative reactivity over injection period
        if (_turbine_trip_detected and _turbine_trip_t is not None):
            _t_slcs_start = _turbine_trip_t + trip_turbine_delay_s + trip_slcs_transport_s
            _t_slcs_end   = _t_slcs_start + trip_slcs_ramp_s
            if time[t] < _t_slcs_start:
                _rho_boron = 0.0
            elif time[t] < _t_slcs_end:
                _frac = (time[t] - _t_slcs_start) / trip_slcs_ramp_s
                _rho_boron = trip_slcs_worth_pcm * _frac
                if abs(_rho_boron - trip_slcs_worth_pcm * max(0.0,
                       (time[t] - 1.0 - _t_slcs_start) / trip_slcs_ramp_s)) > \
                       abs(trip_slcs_worth_pcm) * 0.01:
                    pass  # print progress every ~1% worth change (suppressed)
            else:
                _rho_boron = trip_slcs_worth_pcm
            # Print milestones
            if (time[t] == _t_slcs_start or
                    (time[t] > _t_slcs_start and
                     time[t] - time[max(0,t-1)] > 0 and
                     abs(_rho_boron) > 0 and
                     abs(_rho_boron - trip_slcs_worth_pcm) < 1.0)):
                if abs(_rho_boron - trip_slcs_worth_pcm) < 1.0:
                    print(f"SLCS boron injection complete at t={time[t]:.1f} s"
                          f"  total worth = {trip_slcs_worth_pcm:.0f} pcm")

        if _nf_delta_T > 0 and t > 0:
            _nf_delta_T *= np.exp(-timestep / _tau_nf)

        # Bug 1 fix: reset choked flag when break opens so critical flow
        # is evaluated correctly at the first post-break step.
        if _break_just_opened:
            choked = _choked_initial

        # ── Fuel / hot-pin temperatures ──────────────────────────────────────
        # Uses rkpower_total[t-1] (set by previous step's power block) so the
        # correct decay-heat power is used post-scram.  At t=0 falls back to
        # rated power (initial condition).
        #
        # Thermal resistance chain (average pin):
        #   1. Convection:    TTwall[t] already includes this via slab solve
        #   2. Cladding wall: dT = q"_avg * delta_clad / lambda_clad
        #   3. Fuel-clad gap: dT = Q_pin / (h_gap * 2π * r_fuel_m * L_heated)
        #   4. Fuel pellet:   dT = Q_pin / (8π * k_fuel * L_heated)   [avg]
        #                     dT = Q_pin / (4π * k_fuel * L_heated)   [centreline]
        #
        # Hot-pin algebraic correction (F = F_r * F_z):
        #   T_outer_hot = TTwall[t] + (F-1) * q"_avg / alpha[t]
        #   All internal resistances scale with F * Q_pin_avg.
        if core_flag and SurfArea > 0 and N_pins > 0 and not np.isnan(TTwall[t]):
            # ── Power: use previous step so decay-heat level is correct ──────
            if t > 0 and not np.isnan(rkpower_total[t-1]):
                _q_now = rkpower_total[t-1]
            else:
                _q_now = 1e6 * total_power  # t=0 IC: initial power

            _qavg_f   = _q_now / SurfArea       # avg surface heat flux [W/m²]
            _q_pin    = _q_now / N_pins          # avg pin power [W]
            # ── Quasi-steady-state fuel temperatures ──────────────────────────
            _dT_clad_avg  = _qavg_f * delta_clad / lambda_clad
            _dT_gap_avg   = _q_pin / (h_gap * 2.0 * np.pi * r_fuel_m * L_heated)
            _dT_fuel_avg  = _q_pin / (8.0 * np.pi * k_fuel * L_heated)
            _T_fuel_ss    = TTwall[t] + _dT_clad_avg + _dT_gap_avg + _dT_fuel_avg

            # ── Fuel thermal inertia (first-order exponential lag) ────────────
            # Captures stored energy in the pellet: pellet cools toward _T_fuel_ss
            # with time constant tau_fuel = rho*cp*r^2/(4k) ≈ 5 s.
            _Dt_step = (time[t] - time[t-1]) if t > 0 else 0.0
            _relax   = np.exp(-_Dt_step / _tau_fuel) if _tau_fuel > 0 else 0.0
            if t > 0 and not np.isnan(T_fuel_arr[t-1]):
                _T_fuel_arr_t = _T_fuel_ss + (T_fuel_arr[t-1] - _T_fuel_ss) * _relax
            else:
                _T_fuel_arr_t = _T_fuel_ss      # t=0: rated-power IC

            # ── Hot-pin combined peaking factor ───────────────────────────────
            _F_hot   = F_r * F_z
            # Divide-by-zero guard only — floor at 1 W/m²K.
            # No upper-threshold cutoff: hot-pin must remain active during
            # film boiling (alpha ~ 50 W/m²K) — that is precisely when it matters.
            _alpha_t = max(alpha[t], 1.0) if not np.isnan(alpha[t]) else 1.0
            # Hot-pin clad: quasi-steady offset above average, but relaxed
            # to avoid unphysical spikes when HTC drops suddenly (film boiling).
            _T_hot_clad_ss = TTwall[t] + (_F_hot - 1.0) * _qavg_f / _alpha_t
            if t > 0 and not np.isnan(T_hot_clad_arr[t-1]):
                _T_hot_clad_t = (_T_hot_clad_ss
                                 + (T_hot_clad_arr[t-1] - _T_hot_clad_ss) * _relax)
            else:
                _T_hot_clad_t = _T_hot_clad_ss
            _dT_clad_hot   = _F_hot * _qavg_f * delta_clad / lambda_clad
            _dT_gap_hot    = _F_hot * _q_pin / (h_gap * 2.0 * np.pi * r_fuel_m * L_heated)
            _dT_fuel_cl    = _F_hot * _q_pin / (4.0 * np.pi * k_fuel * L_heated)
            _T_hot_fuel_ss = _T_hot_clad_t + _dT_clad_hot + _dT_gap_hot + _dT_fuel_cl
            if t > 0 and not np.isnan(T_hot_fuel_arr[t-1]):
                _T_hot_fuel_t = (_T_hot_fuel_ss
                                + (T_hot_fuel_arr[t-1] - _T_hot_fuel_ss) * _relax)
            else:
                _T_hot_fuel_t = _T_hot_fuel_ss

        elif core_flag and not np.isnan(TTwall[t]):
            _T_fuel_arr_t  = TTwall[t]          # zero-power limit
            _T_hot_clad_t  = TTwall[t]
            _T_hot_fuel_t  = TTwall[t]
        else:
            _T_fuel_arr_t  = Temperature[t]
            _T_hot_clad_t  = Temperature[t]
            _T_hot_fuel_t  = Temperature[t]

        T_fuel_arr[t]     = _T_fuel_arr_t
        T_hot_clad_arr[t] = _T_hot_clad_t
        T_hot_fuel_arr[t] = _T_hot_fuel_t

        # Cap all reported/feedback fuel and clad temperatures at the BDB
        # diagnostic ceiling (~UO2 melt).  This applies to both average and
        # hot-pin values and prevents normal-mode dry-core heatup from running
        # away before the dry-continuation branch is reached.
        for _Tarr in (T_fuel_arr, T_hot_clad_arr, T_hot_fuel_arr):
            if not np.isnan(_Tarr[t]) and _Tarr[t] > _T_melt_cap_K:
                _Tarr[t] = _T_melt_cap_K
        if not np.isnan(TTwall[t]) and TTwall[t] > _T_melt_cap_K:
            TTwall[t] = _T_melt_cap_K

        # ── Unified reactor trip evaluation ──────────────────────────────────────
        # Evaluated for both full_power_flag=1 and ria_flag=1 cases.
        # Manual time trip fires regardless of core_flag
        if (not _reactor_scrammed and not _pre_break
                and _trip_detected_t is None
                and time[t] >= trip_time_s and trip_time_s < 1e6):
            _trip_detected_t     = time[t]
            _trip_reason_pending = f"manual trip (trip_time_s={trip_time_s:.0f} s)"

        # ── Reactor trip evaluation ───────────────────────────────────────────
        # Pressure, power, and time trips don't require pump initialisation.
        # Flow trip requires _pump_mdot_rated so is gated separately.
        if (core_flag and not _reactor_scrammed and not _pre_break):
            _trip_reason = None
            if trip_P_lo_kPa > 0 and Pressure[t] < trip_P_lo_kPa * 1e3:
                _trip_reason = (f"low-pressure trip  P={Pressure[t]/1e3:.0f} kPa"
                                f" < {trip_P_lo_kPa:.0f} kPa setpoint")
            elif trip_P_hi_kPa > 0 and Pressure[t] >= trip_P_hi_kPa * 1e3:
                _trip_reason = (f"high-pressure trip  P={Pressure[t]/1e3:.0f} kPa"
                                f" >= {trip_P_hi_kPa:.0f} kPa setpoint")
            elif (not np.isnan(rkpower_total[t])
                  and rkpower_total[t] >= trip_power_frac * 1e6 * total_power):
                _trip_reason = (f"high-power trip  {rkpower_total[t]/1e6:.1f} MW"
                                f" >= {trip_power_frac*100:.0f}% rated")
            elif (_pump_mdot_rated is not None and trip_flow_frac > 0
                  and pump_mdot[t] / _pump_mdot_rated < trip_flow_frac):
                _mdot_norm = pump_mdot[t] / _pump_mdot_rated
                _trip_reason = (f"low-flow trip  m_pump={pump_mdot[t]:.0f} kg/s"
                                f"  ({_mdot_norm*100:.1f}% < {trip_flow_frac*100:.0f}% rated)")
            elif _break_just_opened and t_break > 0:
                _trip_reason = f"break opening at t_break={t_break:.2f} s"

            if _trip_reason and _trip_detected_t is None:
                _trip_detected_t     = time[t]
                _trip_reason_pending = _trip_reason
                if trip_delay > 0:
                    print(f"Trip signal at t={time[t]:.2f} s:"
                          f" {_trip_reason}  (scram in {trip_delay:.1f} s)")

        # ── Accumulator boron mass balance (before reactivity block) ────────
        # Accumulator and pumped SI (HPSI/LPSI) may have different boron
        # concentrations — allows boron dilution scenarios (si_boron_ppm=0).
        # Boron leaves via break flow and letdown at current RCS concentration.
        # CVCS: balanced makeup/letdown — zero net mass, net boron =
        #   cvcs_kgs × (si_boron_ppm - C_RCS_ppm) × 1e-6
        # Column E (cvcs_mdot): imbalance adder — positive = extra makeup at
        #   si_boron_ppm; negative = extra letdown at RCS concentration.
        if t > 0:
            _Dt_b = time[t] - time[t-1]
            _mdot_acc  = massflow_in[t-1]
            _mdot_si   = si_pumped_mdot[t-1]       # HPSI + LPSI pumped
            _mdot_tbl  = cvcs_mdot[t-1]              # column E imbalance (may be negative)
            _mdot_brk  = max(0.0, massflow_break[t-1])
            _C_boron   = _rcs_boron_mass / max(Total_Mass, 1.0)
            # CVCS balanced loop: makeup adds cvcs_boron_ppm, letdown removes C_RCS
            if cvcs_kgs > 0.0 and time[t] >= cvcs_start_time_s:
                _boron_cvcs = cvcs_kgs * (cvcs_boron_ppm * 1e-6 - _C_boron)
            else:
                _boron_cvcs = 0.0
            # Column E imbalance: positive = extra makeup at cvcs_boron_ppm, negative = extra letdown at C_RCS
            if _mdot_tbl >= 0.0:
                _boron_tbl = _mdot_tbl * cvcs_boron_ppm * 1e-6
            else:
                _boron_tbl = _mdot_tbl * _C_boron
            _rcs_boron_mass += (_mdot_acc * acc_boron_ppm * 1e-6
                                + _mdot_si  * si_boron_ppm * 1e-6
                                + _boron_cvcs
                                + _boron_tbl
                                - _mdot_brk * _C_boron) * _Dt_b
            _rcs_boron_mass  = max(0.0, _rcs_boron_mass)
        rcs_boron_mass[t] = _rcs_boron_mass
        rcs_boron_ppm[t]  = (_rcs_boron_mass / max(Total_Mass, 1.0)) * 1e6
        # CVCS makeup and letdown flows for output/plotting
        # Makeup = cvcs_kgs (when active) + positive column E imbalance
        # Letdown = cvcs_kgs (when active) + magnitude of negative column E imbalance
        _cvcs_active = cvcs_kgs > 0.0 and time[t] >= cvcs_start_time_s
        _tbl_now = cvcs_mdot[t]
        cvcs_makeup_arr[t]  = (cvcs_kgs if _cvcs_active else 0.0) + max(0.0, _tbl_now)
        cvcs_letdown_arr[t] = (cvcs_kgs if _cvcs_active else 0.0) + abs(min(0.0, _tbl_now))
        # Accumulator boron reactivity — stored separately, added to rho_net below
        _rho_boron_conc = alpha_boron_pcm_ppm * (rcs_boron_ppm[t] - rcs_boron_ppm_init)

        if core_flag and not _pre_break:
            # ── Doppler + moderator feedback (unconditional — includes post-scram) ─
            # Removing the _reactor_scrammed gate so reactivity components
            # are computed throughout the transient for meaningful plots.
            if t == 0 or np.isnan(rkpower_total[t]):
                rkpower_total[t] = 1e6 * total_power

            # T_fuel_arr[t] is computed unconditionally before this block.
            _T_fuel_avg = T_fuel_arr[t]

            _rho_ext_table = float(np.interp(time[t], time_xls, rho_ext_xls))
            # Startup power-target latch for table-driven reactivity.  Once
            # current RK core power first reaches the user-defined power_target,
            # freeze the table reactivity at the value present at that time.
            # Subsequent higher table values are blocked (no additional
            # reactivity addition), but lower/negative table values are honored
            # so the table can still insert shutdown/hold-down reactivity.
            if (rho_table_power_target_cutoff_flag
                    and (not _rho_table_power_target_latched)
                    and rkpower_total[t] >= _power_target * 1.0e6):
                _rho_table_power_target_latched = True
                _rho_table_power_target_latch_t = time[t]
                _rho_table_power_target_latch_value = _rho_ext_table
                print(
                    f"INFO: power_target reached at t={time[t]:.3f} s; "
                    f"reactivity-table contribution frozen at "
                    f"{_rho_table_power_target_latch_value:.3f} pcm "
                    f"(RK power={rkpower_total[t]/1e6:.3f} MW, "
                    f"power_target={_power_target:.3f} MW).",
                    flush=True
                )

            if (_rho_table_power_target_latched
                    and _rho_table_power_target_latch_value is not None):
                _rho_ext = min(_rho_ext_table, _rho_table_power_target_latch_value)
            else:
                _rho_ext = _rho_ext_table
            _T_fuel_0 = (T_fuel_arr[0] if not np.isnan(T_fuel_arr[0])
                         else _T_fuel_avg)
            _rho_D   = alpha_D_pcm * (_T_fuel_avg + _nf_delta_T - _T_fuel_0)
            # Moderator reactivity: split temperature + void model.
            # Temperature term weighted by (1 - void_fraction): in a voided
            # core the cold liquid present is less effective as moderator.
            # Void term uses mass steam quality (avoids rhoL/rhoV amplification).
            _x_void   = max(0.0, x_eq[t]) if not np.isnan(x_eq[t]) else 0.0
            _liq_frac = (1.0 - (alpha_void[t] if not np.isnan(alpha_void[t]) else 0.0))**3
            _rho_M_T    = alpha_M_pcm  * (Temperature[t] - Temperature[0]) * _liq_frac
            _rho_M_void = alpha_M_void * _x_void * 100.0   # pcm per % steam quality
            _rho_M      = _rho_M_T + _rho_M_void
            _rho_net = _rho_ext + _rho_D + _rho_M + _rho_boron + _rho_boron_conc
            rho_ext_arr[t]   = _rho_ext
            rho_boron_arr[t] = _rho_boron
            rho_D_arr[t]     = _rho_D
            rho_M_arr[t]     = _rho_M
            rho_net_arr[t]   = _rho_net

            # ── Scram reactivity ─────────────────────────────────────────────
            # Rod worth insertion after trip delay.  By default, the full rod
            # worth is inserted instantaneously.  If trip_rod_insertion_time_s
            # is positive, the shutdown worth is linearly ramped over that
            # interval to avoid an unrealistically abrupt power collapse.
            if t_rel is not None and t_rel >= 0.0:
                if trip_rod_insertion_time_s > 0.0:
                    _scram_frac = min(max(t_rel / trip_rod_insertion_time_s, 0.0), 1.0)
                else:
                    _scram_frac = 1.0
                _rho_scram_now = _scram_frac * trip_rod_worth_pcm
                rho_scram_arr[t] = _rho_scram_now
                # Update rho_net to include the current inserted scram worth.
                rho_net_arr[t] = (_rho_scram_now + _rho_D + _rho_M
                                  + _rho_boron + _rho_boron_conc)
            else:
                rho_scram_arr[t] = 0.0

        if core_flag and not _pre_break and not _reactor_scrammed:
            # For a step change in reactivity from rho_prev to rho_net, the
            # power makes an instantaneous prompt jump:
            #   P_new = P_old * (beta_eff - rho_prev) / (beta_eff - rho_net)
            # This is the term the Euler inhour step cannot resolve when the
            # reactivity changes faster than the timestep (rod ejection).
            # Only applied when both rho values are sub-critical (< beta_eff).
            if (pja_flag and beta_eff_pcm > 1 and t > 0
                    and not np.isnan(rho_net_arr[t-1])):
                _D_prev = beta_eff_pcm - rho_net_arr[t-1]
                _D_curr = beta_eff_pcm - _rho_net
                if _D_prev > 0 and _D_curr > 0:
                    rkpower_total[t] = rkpower_total[t] * (_D_prev / _D_curr)

            # ── High-power trip (unified trip_power_frac setpoint) ─────────
            if (not _reactor_scrammed and _trip_detected_t is None
                    and rkpower_total[t] >= trip_power_frac * 1e6 * total_power):
                _trip_detected_t    = time[t]
                _trip_reason_pending = (f"high-power trip"
                                        f"  {rkpower_total[t]/1e6:.1f} MW"
                                        f" >= {trip_power_frac*100:.0f}% rated"
                                        f"  (trip setpoint)")
                if trip_delay > 0:
                    print(f"Trip signal at t={time[t]:.2f} s:"
                          f" {_trip_reason_pending}"
                          f"  (scram in {trip_delay:.1f} s)")

            if not _reactor_scrammed:
                _Dt = time[min(t+1, number_timesteps-1)] - time[t]
                if Lambda_prompt > 0 and _rho_net > beta_eff_pcm:
                    # ── Nordheim-Fuchs prompt supercritical burst ──────────────
                    # dP/dt = [(δρ - |αD|E/Cf) × 1e-5 / Λ_prompt] × P
                    # dE/dt = P
                    # δρ = rho_net - beta_eff [pcm].  Burst self-terminates
                    # when adiabatic Doppler (|αD|×E/Cf) consumes all δρ.
                    _P_nf   = float(rkpower_total[t])  # W
                    _E_nf   = 0.0                       # J deposited this step
                    _dR0    = _rho_net - beta_eff_pcm   # initial excess [pcm]
                    _k_nf   = 1e-5 / Lambda_prompt      # [1/(s·pcm)]
                    _A_nf   = abs(alpha_D_pcm) / _nf_Cf # [pcm/J]
                    _nsub   = 1000
                    _dtsub  = _Dt / _nsub
                    _P_peak = _P_nf
                    for _i in range(_nsub):
                        _dR = _dR0 - _A_nf * _E_nf     # remaining excess [pcm]
                        if _dR <= 0: break               # burst terminated
                        _dP    = _k_nf * _dR * _P_nf * _dtsub
                        _dE    = _P_nf * _dtsub
                        _P_nf += _dP
                        _E_nf += _dE
                        _P_nf  = max(_P_nf, 0.0)
                        if _P_nf > _P_peak: _P_peak = _P_nf
                    rkpower_total[t+1] = max(_P_nf, 0.0)
                    _nf_delta_T += _E_nf / _nf_Cf       # adiabatic fuel heating
                    _nf_E_total += _E_nf                 # cumulative burst energy
                    if _E_nf > 1e6:
                        print(f"N-F burst t={time[t]:.3f} s: "
                              f"δρ={_dR0:.0f} pcm, "
                              f"P_peak={_P_peak/1e6:.0f} MW, "
                              f"E={_E_nf/1e6:.0f} MJ, "
                              f"ΔT_fuel={_E_nf/_nf_Cf:.1f} K")
                    # Trip detection on NF peak (burst is sub-timestep;
                    # normal trip check sees only pre-burst rkpower_total[t])
                    if (_trip_detected_t is None and not _reactor_scrammed
                            and _P_peak >= trip_power_frac * 1e6 * total_power):
                        _trip_detected_t    = time[t]
                        _trip_reason_pending = (f"high-power trip"
                                                f"  {_P_peak/1e6:.1f} MW"
                                                f" >= {trip_power_frac*100:.0f}%"
                                                f" rated  (N-F burst)")
                        if trip_delay > 0:
                            print(f"Trip signal at t={time[t]:.3f} s"
                                  f" (N-F peak): {_trip_reason_pending}"
                                  f"  (scram in {trip_delay:.1f} s)")
                else:
                    # Standard single-group Euler (sub-critical)
                    # Clamp rho_net to avoid Euler overshoot with large negative reactivity
                    _rho_clamped = max(_rho_net*1e-5, -Lambda_star / _Dt)
                    _P_next = rkpower_total[t] * (1.0 + _Dt * _rho_clamped / Lambda_star)
                    rkpower_total[t+1] = max(_P_next, 0.0)

        elif _pre_break:
            rkpower_total[t] = 1e6 * total_power
        else:
            if t_rel is None:
                # No break, no scram — steady full power
                rkpower_total[t] = 1e6 * total_power
            else:
                _decay_heat_W = 1e6*total_power/Efis*dhp(t_rel,Tinf,U239yield,
                                                          fr35,fr38,fr39,stdd)

                # Use the current ramped rod worth in the post-scram fission
                # decay model. Earlier code used the final full rod worth
                # (_rho_scram_dollars), so trip_rod_insertion_time_s only
                # affected the output reactivity diagnostic, not the power.
                if trip_rod_insertion_time_s > 0.0:
                    _scram_frac_power = min(max(t_rel / trip_rod_insertion_time_s, 0.0), 1.0)
                else:
                    _scram_frac_power = 1.0
                _rho_scram_now_pcm = _scram_frac_power * trip_rod_worth_pcm
                _rho_scram_now_dollars = _rho_scram_now_pcm / max(beta_eff_pcm, 1.0)

                rkpower_total[t] = (
                    1e6*total_power*fission_scram(t_rel, _rho_scram_now_dollars)
                    + _decay_heat_W
                )

                # ── Hybrid post-scram PK: excess fission power ────────────────
                # Integrates the single-group precursor equation using rho_net
                # (which includes rod worth + Doppler + moderator).  Only
                # contributes when rho_net > -beta_eff (re-criticalization).
                # Baseline fission_scram handles normal decay; this adds the
                # excess when insufficient rod worth allows power recovery.
                if _pk_scram_active and core_flag:
                    _Dt_pk = time[t] - time[t-1] if t > 0 else 1.0
                    _rho_n = rho_net_arr[t]   # already includes rod worth + D + M
                    # Quasi-static inhour: net growth rate = rho_net / Lambda_star.
                    # Seed from P_base so re-criticalization can develop even
                    # when _P_excess has been near zero for a long time.
                    _P_base = (
                        1e6 * total_power * fission_scram(t_rel, _rho_scram_now_dollars)
                        if t_rel is not None else 0.0
                    )
                    _P_seed = max(_P_excess, _P_base)
                    _dP_ex  = (_rho_n * 1e-5 / Lambda_star) * _P_seed * _Dt_pk
                    _P_excess = max(0.0, min(_P_excess + _dP_ex,
                                            1e6 * total_power))
                    if _P_excess > 0.0:
                        rkpower_total[t] += _P_excess

            if not np.isnan(power_input[t]):
                rkpower_total[t] += power_input[t]

        # Bug 2 fix: when core_flag=1, cond_heat[t_break] is stale from SS.
        # At the first post-break step, use rkpower directly so the energy
        # equation sees scram power, not the pre-break 575 MW.
        if _break_just_opened and core_flag:
            net_heat_total[t] = rkpower_total[t]
        elif t == 0:
            # cond_heat[0] is not yet computed; use rated power directly
            net_heat_total[t] = total_power * 1e6
        else:
            net_heat_total[t] = cond_heat[t] if (cond_flag==1 or core_flag==1) else rkpower_total[t]

        # ── SG UA diagnostic/update path ─────────────────────────────────────
        # Dynamic diagnostic definition requested for this option:
        #     UA_sg_rated = RK Total Power / (RCS Temperature - T_sec_K)
        # This update is intentionally placed after rkpower_total[t] is finalized
        # for the current timestep and before the SG heat-removal model uses
        # UA_sg_rated.
        _UA_dyn_t, _dT_dyn_t = _compute_sg_dynamic_ua(Temperature[t], rkpower_total[t])
        if sg_dynamic_ua_flag and not _UA_sg_user_supplied and np.isfinite(_UA_dyn_t):
            UA_sg_rated = _UA_dyn_t
            sg_dynamic_active_arr[t] = 1
        sg_UA_used[t]        = UA_sg_rated
        sg_UA_dynamic[t]     = _UA_dyn_t
        sg_dT_primary_sec[t] = _dT_dyn_t

        # ── Energy deposition limiter (voided core in film boiling) ──────────
        # When the RCS is severely voided and in film boiling, cap the heat
        # input to prevent the 2×2 solver from receiving more energy than the
        # available coolant mass can absorb in one timestep.  Physically, in a
        # nearly-empty RCS the fission energy heats the fuel and clad; only
        # the convective portion (already captured by cond_heat via the low
        # film-boiling HTC) reaches the coolant.  The cap is:
        #   Q_max = Total_Mass * cp_steam * dT_max / dt
        # where dT_max limits the coolant temperature rise to 100 K/s.
        if (_film_boiling_active and Total_Mass > 0
                and Total_Mass_scaled[t] < 0.6):
            _Dt_lim  = time[min(t+1, number_timesteps-1)] - time[t]
            _cp_stm  = 2000.0    # J/kg·K  steam specific heat (conservative)
            _dT_max  = 100.0     # K/s max coolant temperature rise
            _Q_max   = Total_Mass * _cp_stm * _dT_max
            if net_heat_total[t] > _Q_max:
                net_heat_total[t] = _Q_max

        # rho_boron_arr: SLCS + accumulator boron combined
        rho_boron_arr[t] = _rho_boron + _rho_boron_conc

        _prev_pre_break = _pre_break  # update tracker for next step

        # ── dry-core continuation / early termination ────────────────────────
        # The dry-core continuation can be triggered either by the legacy
        # near-zero RCS inventory cutoff or by a user-defined vessel liquid level.
        #
        # New input:
        #   dry_core_trigger_level_m
        #       <0  : disabled; use mass_scaled_cutoff only
        #       >=0 : start dry-core continuation when ves_ll <= this level
        #
        # Recommended BDB sensitivity use:
        #   dry_core_trigger_level_m = top of active fuel or another user-defined
        #                              level where the lumped-water model is no
        #                              longer appropriate.
        # Full dry-core continuation is reserved for near-empty inventory where
        # the lumped hydraulic model is no longer numerically meaningful.
        # The user-defined level trigger above affects core heat transfer only;
        # it does not terminate the hydraulic transient.
        _dry_level_trigger_m = dry_core_trigger_level_m
        _dry_by_level = False
        _dry_by_mass  = (core_flag and Total_Mass_scaled[t] < mass_scaled_cutoff)
        _dry_triggered = _dry_by_mass

        if Pressure[t] <= 0*pressure_cutoff*1e3 or _dry_triggered:
            if core_flag and _dry_triggered:
                _dry_reason = (
                    f"vessel level {ves_ll[t]:.3f} m <= trigger {_dry_level_trigger_m:.3f} m"
                    if _dry_by_level else
                    f"RCS inventory below {100*mass_scaled_cutoff:.2f}% of initial"
                )
                print(f"Dry-core continuation started at t={time[t]:.2f} s: "
                      f"{_dry_reason}. Accumulator flow at handoff = "
                      f"{massflow_in[t]:.3f} kg/s. Hydraulics are outside the "
                      f"lumped-water model; continuing with ad-hoc dry-core "
                      f"heatup for oxidation/H2 diagnostics.")

                # Ad-hoc beyond-design-basis continuation:
                #   - hydraulic state is pinned to a dry/depressurized vessel;
                #   - decay heat continues to heat the fuel/clad;
                #   - residual dry gas/steam cooling uses a small HTC plus radiation.
                # This is not a best-estimate dry-core model; it is a robust
                # continuation model to support qualitative BDB oxidation/H2 trends.
                _dry_htc = dry_core_htc_W_m2K
                _dry_eps = dry_core_emissivity
                _sigma_sb = 5.670374419e-8
                _T_sink = max(float(Temperature[t]) if np.isfinite(Temperature[t]) else 373.15, 300.0)

                _r_o = max(0.5 * float(d_pin), 1.0e-5)
                _r_i = max(_r_o - float(delta_clad), 1.0e-5)
                _A_core = max(float(N_pins) * np.pi * float(d_pin) * float(L_heated), 1.0)
                _m_clad = max(float(N_pins) * np.pi * (_r_o**2 - _r_i**2) * float(L_heated) * float(rho_clad), 1.0)
                _r_fuel = _r_i
                _m_fuel = max(float(N_pins) * np.pi * _r_fuel**2 * float(L_heated) * 10400.0, 1.0)
                _C_core = max(_m_clad * float(cp_clad) + _m_fuel * 330.0, 1.0)
                _hot_mult = max(float(F_r) * float(F_z), 1.0)

                def _dry_fail_count(_T_hot_K, _T_avg_K, _T_lim_C, _prev):
                    _T_lim_K = _T_lim_C + 273.15
                    if not np.isfinite(_T_hot_K) or _T_hot_K <= _T_lim_K:
                        return _prev
                    _dT = max(_T_hot_K - _T_avg_K, 0.1)
                    _sig = _dT / max(k_sigma, 0.1)
                    _z = (_T_lim_K - _T_avg_K) / _sig
                    _n_inst = max(0, int(round(N_pins * (1.0 - float(_norm_dist.cdf(_z))))))
                    return max(_prev, _n_inst)

                _last_valid = t
                for _k in range(t+1, number_timesteps):
                    _dt_dry = max(float(time[_k] - time[_k-1]), 0.0)

                    # Use available decay/fission heat if already populated; otherwise
                    # carry forward the last heat source.  Most BDB dryout cases are
                    # post-scram, so this is principally decay heat.
                    _Q_decay = rkpower_total[_k-1] if (_k > 0 and np.isfinite(rkpower_total[_k-1]) and rkpower_total[_k-1] > 0) else rkpower_total[t]
                    if not np.isfinite(_Q_decay) or _Q_decay <= 0:
                        _Q_decay = max(net_heat_total[t], 0.0)

                    _T_clad_prev = TTwall[_k-1] if np.isfinite(TTwall[_k-1]) else TTwall[t]
                    _T_hot_prev  = T_hot_clad_arr[_k-1] if np.isfinite(T_hot_clad_arr[_k-1]) else T_hot_clad_arr[t]

                    _Q_conv = _dry_htc * _A_core * max(_T_clad_prev - _T_sink, 0.0)
                    _Q_rad  = _dry_eps * _sigma_sb * _A_core * max(_T_clad_prev**4 - _T_sink**4, 0.0)
                    _Q_net_dry = max(_Q_decay - _Q_conv - _Q_rad, 0.0)
                    _dT_avg = (_Q_net_dry / _C_core) * _dt_dry

                    TTwall[_k] = min(_T_clad_prev + _dT_avg, _T_melt_cap_K)

                    _T_fuel_next = (
                        (T_fuel_arr[_k-1] if np.isfinite(T_fuel_arr[_k-1]) else T_fuel_arr[t])
                        + _dT_avg
                    )
                    T_fuel_arr[_k] = min(_T_fuel_next, _T_melt_cap_K)

                    _T_hot_clad_next = _T_hot_prev + _dT_avg * _hot_mult
                    T_hot_clad_arr[_k] = min(_T_hot_clad_next, _T_melt_cap_K)

                    _T_hot_fuel_next = (
                        (T_hot_fuel_arr[_k-1] if np.isfinite(T_hot_fuel_arr[_k-1]) else T_hot_fuel_arr[t])
                        + _dT_avg * _hot_mult
                    )
                    T_hot_fuel_arr[_k] = min(_T_hot_fuel_next, _T_melt_cap_K)

                    Pressure[_k] = max(pressure_containment * 1e3, 101325.0)
                    Temperature[_k] = _T_sink
                    enthalpy_mix[_k] = enthalpy_mix[t] if np.isfinite(enthalpy_mix[t]) else enthalpy_mix[_k-1]
                    Total_Mass_scaled[_k] = 0.0
                    Total_Mass = 0.0
                    ves_ll[_k] = min(ves_ll[t], _dry_level_trigger_m) if _dry_by_level else 0.0
                    x_eq[_k] = 1.0
                    alpha_void[_k] = 1.0
                    alpha[_k] = _dry_htc
                    # In dry-core continuation, keep the reported core power as
                    # the decay/fission heat source rather than zeroing it.  The
                    # normal coolant-energy interpretation of net_heat_total is
                    # no longer applicable, but a zero "Core Power" column is
                    # misleading for BDB dry-core heatup.
                    rkpower_total[_k] = _Q_decay
                    net_heat_total[_k] = _Q_decay
                    cond_heat[_k] = max(_Q_conv + _Q_rad, 0.0)

                    massflow_break[_k] = 0.0
                    h_break_arr[_k] = 0.0
                    massflow_PORV[_k] = 0.0
                    pump_mdot[_k] = 0.0
                    pump_velocity[_k] = 0.0
                    pump_head[_k] = 0.0
                    Q_sg[_k] = 0.0
                    DNBR[_k] = DNBR[_k-1] if _k > 0 else np.nan
                    q_chf[_k] = 0.0
                    q_hot[_k] = 0.0

                    N_fail_gap[_k]  = _dry_fail_count(T_hot_clad_arr[_k], TTwall[_k], T_gap_release_C, N_fail_gap[_k-1])
                    N_fail_eiv[_k]  = _dry_fail_count(T_hot_clad_arr[_k], TTwall[_k], T_early_iv_C,    N_fail_eiv[_k-1])
                    N_fail_clad[_k] = N_fail_eiv[_k]
                    N_fail_DNB[_k]  = max(N_fail_DNB[_k-1], N_fail_gap[_k])

                    rho_ext_arr[_k] = rho_ext_arr[_k-1]
                    rho_scram_arr[_k] = rho_scram_arr[_k-1]
                    rho_boron_arr[_k] = rho_boron_arr[_k-1]
                    rcs_boron_ppm[_k] = rcs_boron_ppm[_k-1]
                    rho_D_arr[_k] = rho_D_arr[_k-1]
                    rho_M_arr[_k] = rho_M_arr[_k-1]
                    rho_net_arr[_k] = rho_net_arr[_k-1]

                    pzr_level_arr[_k] = 0.0
                    pzr_level_norm_arr[_k] = 0.0
                    pzr_mdot_surge_arr[_k] = 0.0
                    acc_liqvol[_k] = 0.0
                    acc_pres[_k] = acc_pres[_k-1]
                    acc_tgas[_k] = acc_tgas[_k-1]
                    massflow_in[_k] = 0.0
                    hpsi_mdot_arr[_k] = 0.0
                    lpsi_mdot_arr[_k] = 0.0
                    si_pumped_mdot[_k] = 0.0
                    cvcs_mdot[_k] = 0.0
                    cvcs_makeup_arr[_k] = 0.0
                    cvcs_letdown_arr[_k] = 0.0

                    _last_valid = _k

                print(f"Dry-core continuation completed to t={time[_last_valid]:.1f} s.")
                break
            else:
                print(f"Simulation stopped early at t={time[t]:.2f} s.")
                break

        # Fuel temperature termination: avg fuel temperature exceeds
        # Cladding temperature limit: log a one-time warning but continue.
        # The simulation holds fuel temperature at the UO2 melt ceiling and
        # drives net heat transfer to zero, so the solver remains stable.
        if (core_flag and not np.isnan(TTwall[t])
                and TTwall[t] - 273.15 > T_fuel_limit_C
                and not _clad_limit_warned):
            print(f"WARNING t={time[t]:.1f} s: avg clad surface temperature "
                  f"{TTwall[t]-273.15:.0f} C exceeds limit {T_fuel_limit_C:.0f} C "
                  f"-- continuing with fuel temperature held at ceiling.")
            _clad_limit_warned = True

        # Critical-pressure termination: XSteam two-phase properties and ERM
        # critical flow are not valid above 22.064 MPa (water critical point).
        # Terminate gracefully rather than continuing with unphysical state.
        _P_crit_kPa = 22064.0   # kPa — critical pressure of water
        if Pressure[t] >= _P_crit_kPa * 1e3:
            print(f"Simulation terminated at t={time[t]:.1f} s: "
                  f"RCS pressure {Pressure[t]/1e3:.0f} kPa reached "
                  f"critical point ({_P_crit_kPa:.0f} kPa). "
                  f"Two-phase model no longer valid.")
            break

        # PORV-depletion guard: when the RCS is drained by the PORV alone
        # (no break, no ECCS injection) to below 15% of initial inventory,
        # the single-node model is outside its valid regime (essentially dry
        # superheated steam).  Terminate before XSteam returns sentinel values.
        _porv_only_depletion = (
            PORV_area_m2 > 0
            and massflow_break[t] < 0.01          # no break flow
            and acc_wdot[t] < 0.01                # no ECCS injection
            and Total_Mass_scaled[t] < 0.15       # 85% inventory gone
        )
        if _porv_only_depletion:
            print(f"Terminating at t={time[t]:.1f} s: RCS inventory depleted to "                  f"{Total_Mass_scaled[t]*100:.1f}% by PORV — "                  f"single-node model not valid below 15% inventory.")
            break

        # ── ADV logic ────────────────────────────────────────────────────────
        if flag_ADV_open < 1:
            if Pressure[t] < setpoint_ADV_open*1e3:
                effective_total_area = effective_area_break + effective_area_ADV
                flag_ADV_open = 10

        # ── break area gating — zero before t_break ───────────────────────────
        if _pre_break:
            _active_area = 0.0
        else:
            _active_area = effective_total_area

        # ── break flow ───────────────────────────────────────────────────────
        # Reset vapor latch when the RCS transitions to subcooled (x_eq < 0),
        # indicating the vessel has genuinely refilled with liquid and the
        # vapor-venting phase is over.  ves_ll is an all-liquid equivalent
        # level (not the actual two-phase interface) and is not a reliable
        # indicator of submergence at the break plane, so it is not used here.
        if (x_eq[t] < 0 and flag_vent_vapor > 0):
            flag_vent_vapor       = 0
            flag_stagnation_break = initial_flag_stagnation_break

        if x_eq[t] < 0.005:
            # Subcooled bulk mixture — the RCS is liquid-filled.
            # Use vessel level to distinguish submerged break from a deeply
            # depressurised but still partially voided vessel.
            _ves_ll_now = ves_ll[t-1] if t > 0 and np.isfinite(ves_ll[t-1]) else 0.0
            _break_elevation = break_elevation_m
            if _ves_ll_now >= 0.95 * _break_elevation:
                # Vessel full enough that break is submerged
                stagnation_enthalpy_break = sat_liquid["enthalpy"]
                massflux_break = Cd_sub*critical_flow_ERM(Pressure[t],
                                                           stagnation_enthalpy_break,
                                                           sat_liquid, sat_vapor)
            else:
                # Subcooled but level not yet at break — mixture enthalpy
                massflux_break = Cd_sub*critical_flow_ERM(Pressure[t],enthalpy_mix[t],
                                                           sat_liquid,sat_vapor)
                stagnation_enthalpy_break = enthalpy_mix[t]
        else:
            # Two-phase or superheated bulk — break plane is not submerged.
            # Select stagnation enthalpy based on vapor latch state.
            #
            # The transition from "mixture" to "vapor" is driven by the
            # blowdown-phase transition_mixture_mass threshold — a one-way
            # latch.  Once the latch fires (flag_vent_vapor > 0) it must NOT
            # be re-evaluated using Total_Mass_scaled: that threshold was
            # derived from blowdown experiments and has no meaning during
            # reflood, where mass is being added back from ECCS rather than
            # draining from the RCS.  The only release is when the bulk
            # mixture becomes genuinely subcooled (x_eq < 0), handled above,
            # which indicates the vessel has refilled to the break elevation.
            if flag_vent_vapor < 1:
                # Latch not yet set — evaluate blowdown transition
                if Total_Mass_scaled[t] <= transition_mixture_mass_use:
                    flag_stagnation_break = "vapor"
                    flag_vent_vapor       = 10
                # else: stay on initial_flag_stagnation_break ("mixture")
            # Once flag_vent_vapor > 0 the latch is permanent until x_eq < 0
            if flag_stagnation_break == "vapor":
                stagnation_enthalpy_break = sat_vapor["enthalpy"]
            elif flag_stagnation_break == "mixture":
                af   = (Total_Mass_scaled[t]-transition_mixture_mass_use)/(1-transition_mixture_mass_use)
                ag   = 1 - af
                rhof = 1/XSteam.vL_p(Pressure[t]/1e3)
                rhog = 1/XSteam.vV_p(Pressure[t]/1e3)
                # Use current void fraction — not cumulative max — so stagnation
                # enthalpy correctly decreases as system refills with liquid.
                tmp2[t] = ag*rhog/(ag*rhog+af*rhof)
                stagnation_enthalpy_break = (tmp2[t]*(sat_vapor["enthalpy"]-sat_liquid["enthalpy"])
                                              + sat_liquid["enthalpy"])

            # ── Dynamic choking check ─────────────────────────────────────────
            # Critical pressure ratio for steam ≈ 0.55 (Fauske/Moody).
            # When P_RCS / P_containment falls below this the flow is no longer
            # choked and transitions to subcritical orifice.
            _P_cont_Pa = pressure_containment * 1e3
            _pr = _P_cont_Pa / max(Pressure[t], _P_cont_Pa + 1.0)
            _is_choked = choked and (_pr < critical_pressure_ratio_guess)

            if _is_choked:
                cp_guess = (critical_pressure_ratio_guess*(Pressure[t]-_P_cont_Pa)
                            + _P_cont_Pa)
                massflux_break,_,_,_ = critical_flow_newton(
                    Pressure[t], stagnation_enthalpy_break, slip_type, cp_guess)
            else:
                # Subcritical orifice: G = Cd * sqrt(2 * rho * dP)
                _dP = max(0.0, Pressure[t] - _P_cont_Pa)
                _rho_break = 1.0 / max(mix_properties["specific_volume"], 1e-6)
                massflux_break = Cd_sat * np.sqrt(2.0 * _rho_break * _dP)
            massflux_break = Cd_sat*np.real(massflux_break)
            if not np.isfinite(massflux_break):
                massflux_break = 0.0

        massflow_break[t]      = _active_area * massflux_break
        # Guard: critical flow model may return NaN near atmospheric pressure
        if not np.isfinite(massflow_break[t]) or massflow_break[t] < 0:
            massflow_break[t] = massflow_break[t-1] if t > 0 else 0.0
        _h_break = stagnation_enthalpy_break if np.isfinite(stagnation_enthalpy_break) else enthalpy_mix[t]
        energy_outflow_break[t]= massflow_break[t] * _h_break
        h_break_arr[t]         = _h_break * 1e-3   # convert J/kg → kJ/kg for output

        # ── Vessel level and PORV source enthalpy ───────────────────────────────
        # Compute level before the PORV block.  PORV discharge enthalpy must be
        # self-consistent with the modeled source region.  If a pressurizer with
        # a finite steam space exists, the PORV vents saturated steam.  If no
        # pressurizer is configured, or if the pressurizer is effectively solid
        # water, the PORV removes bulk RCS enthalpy.  This avoids the nonphysical
        # energy sink caused by removing steam enthalpy from an all-liquid lumped
        # RCS node.
        rhoL     = XSteam.rhoL_p(Pressure[t]/1e3)
        ves_ll[t]= (init_cond["Total_Mass_RCS"]*Total_Mass_scaled[t]/rhoL/ves_vol
                    *(R5_lowCV_height+R5_highCV_height))

        _pzr_level_norm = (pzr_level_arr[t] / pzr_height
                           if pzr_height > 0 else 1.0)
        # PORV steam discharge is only assumed when the diagnostic pressurizer
        # inventory is credible and two-phase: bounded away from both empty and
        # solid.  If the level diagnostic is empty/invalid, fall back to bulk
        # RCS enthalpy to keep the lumped RCS energy balance self-consistent.
        _pzr_has_valid_steam_space = (pzr_vol > 0.0 and
                                      pzr_height > 0.0 and
                                      np.isfinite(_pzr_level_norm) and
                                      _pzr_level_norm > PORV_min_liquid_level and
                                      _pzr_level_norm < (1.0 - PORV_steam_margin))
        _h_porv = sat_vapor["enthalpy"] if _pzr_has_valid_steam_space else enthalpy_mix[t]

        # The predictor-corrector block below computes the actual PORV flow.
        massflow_PORV[t] = 0.0

        # ── accumulator ──────────────────────────────────────────────────────
        acc_delz = acc_liqvol[t] / acc_area
        rho_acc  = XSteam.rho_pT(inflow_info["pressure"], inflow_info["temp"])
        # rhoL and ves_ll[t] already computed above

        if t > 0:
            v2   = acc_totvol - acc_liqvol[t]
            h    = (0.15*0.029*(9.8*0.73*0.0033*((acc_pres[t-1]**0.99)/1.29)**2)**(1/3)
                    *(tgas-acc_tgas[t-1])**(1/3))
            dq   = h*2*(1.3333*acc_area+v2*(np.pi/acc_area)**0.5)*(tgas-acc_tgas[t-1])
            a    = 1.4/v2*acc_wdot[t-1]
            b_acc= 0.4/v2*dq
            ba   = 0 if a==0 else b_acc/a
            acc_pres[t] = (acc_pres[t-1]-ba)*np.exp(-a*timestep)+ba
            acc_tgas[t] = acc_tgas[t-1]*np.exp(
                0.4*np.log((acc_totvol-acc_liqvol[t-1])/v2)
                +0.4*timestep*dq/acc_pres[t]/v2)
            if acc_liqvol[t] < 0.1 and a == 0:
                dq = h*(1.3333*acc_area+2*v2*(np.pi/acc_area)**0.5)*(tgas-acc_tgas[t-1])
                acc_tgas[t] = acc_tgas[t-1]*np.exp(
                    0.4*np.log((acc_totvol-acc_liqvol[t-1])/v2)
                    +0.4*timestep*dq/acc_pres[t]/v2)
                acc_pres[t] = acc_pres[t-1]*(acc_tgas[t]/acc_tgas[t-1])
            if Pressure[t] < acc_pres[t]+rho_acc*9.8*acc_delz and acc_liqvol[t] > 0:
                _acc_wdot_new = acc_narea*np.sqrt(
                    2*(acc_pres[t]+rho_acc*9.8*acc_delz-Pressure[t])/rho_acc)
                # Under-relax with previous timestep to eliminate the 2-step
                # on/off chatter caused by the one-timestep lag in the
                # accumulator pressure update.
                acc_wdot[t] = 0.7*acc_wdot[t-1] + 0.3*_acc_wdot_new

        # ── Pumped Safety Injection (HPSI / LPSI) ────────────────────────────
        # Abbreviated HPSI/LPSI model per FLARE SI specification.
        # Signal: fires when Pressure[t] first drops below p_si_setpoint_Pa.
        # Flow:   mdot_j = a_j(t) * mdot_j,r * max(0, P_so,j - P_RCS)
        #                                      /     (P_so,j - P_r,j)
        # a_j(t) ramps 0→1 over si_ramp_time_s after each class delay.
        # LPSI is inherently pressure-permissive: the (P_so,L - P_RCS) term
        # is zero until RCS depressurises below the LPSI shutoff.
        # RWST inventory depletes; flow stops when exhausted.
        if si_enabled and t > 0:
            if not _si_signal and Pressure[t] < p_si_setpoint_Pa:
                _si_signal   = True
                _si_signal_t = time[t]
                print(f"SI signal at t={time[t]:.2f} s: P={Pressure[t]/1e6:.3f} MPa"
                      f" < setpoint {p_si_setpoint_Pa/1e6:.3f} MPa")

            _m_hpsi = 0.0
            _m_lpsi = 0.0
            if _si_signal and _si_signal_t is not None and _rwst_mass > 0.0:
                _elapsed = time[t] - _si_signal_t

                # HPSI
                _a_H = float(np.clip(
                    (_elapsed - hpsi_delay_s) / max(si_ramp_time_s, 1e-6),
                    0.0, 1.0))
                if _a_H > 0.0 and hpsi_mdot_rated > 0.0:
                    _dp_H   = hpsi_p_shutoff_Pa - Pressure[t]
                    _den_H  = hpsi_p_shutoff_Pa - hpsi_p_rated_Pa
                    if _den_H > 0.0 and _dp_H > 0.0:
                        _m_hpsi = min(
                            _a_H * hpsi_mdot_rated * (_dp_H / _den_H),
                            hpsi_mdot_rated)

                # LPSI (pressure-permissive via (P_so,L - P_RCS) term)
                _a_L = float(np.clip(
                    (_elapsed - lpsi_delay_s) / max(si_ramp_time_s, 1e-6),
                    0.0, 1.0))
                if _a_L > 0.0 and lpsi_mdot_rated > 0.0:
                    _dp_L  = lpsi_p_shutoff_Pa - Pressure[t]
                    _den_L = lpsi_p_shutoff_Pa - lpsi_p_rated_Pa
                    if _den_L > 0.0 and _dp_L > 0.0:
                        _m_lpsi = min(
                            _a_L * lpsi_mdot_rated * (_dp_L / _den_L),
                            lpsi_mdot_rated)

                # RWST depletion
                _dt_si     = time[t] - time[t - 1]
                _rwst_draw = (_m_hpsi + _m_lpsi) * _dt_si
                if _rwst_draw > _rwst_mass:
                    _scale  = _rwst_mass / max(_rwst_draw, 1e-9)
                    _m_hpsi *= _scale
                    _m_lpsi *= _scale
                    _rwst_mass = 0.0
                    print(f"RWST depleted at t={time[t]:.1f} s — SI flow stops.")
                else:
                    _rwst_mass -= _rwst_draw

                # Overfill guard: suppress SI injection when RCS mass exceeds
                # si_mass_ceiling * initial inventory.  Prevents unbounded mass
                # accumulation when break flow cannot keep pace with injection.
                if Total_Mass > si_mass_ceiling * Total_Mass_init:
                    _m_hpsi = 0.0
                    _m_lpsi = 0.0

            # ── SI pump degradation with vessel fill level ────────────────────
            # As the vessel refills, static back-pressure reduces pump delivery.
            # Full flow until 50% fill; linearly degrades to zero at 100% fill.
            _ves_total_ht = R5_lowCV_height + R5_highCV_height
            if _ves_total_ht > 0 and not np.isnan(ves_ll[t]):
                _fill_frac   = min(1.0, max(0.0, ves_ll[t] / _ves_total_ht))
                _si_degrade  = max(0.0, 1.0 - max(0.0, _fill_frac - si_fill_threshold)
                                              / max(1.0 - si_fill_threshold, 1e-9))
                _m_hpsi *= _si_degrade
                _m_lpsi *= _si_degrade

            hpsi_mdot_arr[t]  = _m_hpsi
            lpsi_mdot_arr[t]  = _m_lpsi
            si_pumped_mdot[t] = _m_hpsi + _m_lpsi

        # ── Pump coastdown step ───────────────────────────────────────────────
        #   - pump_trip_time is set and time[t] >= pump_trip_time, OR
        #   - pump_trip_time is None and the reactor has scrammed
        # In both cases the pump holds at rated speed until the trip condition.
        if pump_trip_time is not None:
            _pump_tripped = (time[t] >= pump_trip_time)
        else:
            _pump_tripped = _reactor_scrammed
        _coast_ok = (pump_flag and not _pre_break and _pump_tripped)
        if _coast_ok:
            _pump_state = step(_pump_state, Pressure[t],
                               Temperature[t], timestep,
                               tau_s=pump_tau)
        # pre-break or pump not yet tripped: pump frozen at rated speed.
        # _pump_state frozen at rated speed
        pump_omega[t+1]    = _pump_state.get("omega_rpm",     0.0)
        pump_mdot[t+1]     = _pump_state.get("mass_flow_kgs", 0.0)
        pump_velocity[t+1] = _pump_state.get("velocity_ms",   0.0)
        pump_head[t+1]     = _pump_state.get("head_Pa",       0.0)

        # ── Steam boil-off velocity in core channel ───────────────────────────
        # Upward steam velocity driven by decay-heat boil-off.
        # mdot_steam = Q_decay / h_fg  ;  v_steam = mdot_steam / (rhoV * A_flow_core)
        # Used as the physically-based steam velocity for post-CHF D-B and
        # written to the output for transparency.  Guarded to zero outside
        # film-boiling conditions (no meaningful steam upflow when core is liquid-filled).
        if _film_boiling_active and A_flow_core is not None and A_flow_core > 0:
            try:
                _hfg_sv  = max((XSteam.hV_p(Pressure[t] * 1e-3)
                                - XSteam.hL_p(Pressure[t] * 1e-3)) * 1e3, 1.0)  # J/kg
                _rhoV_sv = max(XSteam.rhoV_p(Pressure[t] * 1e-3), 0.1)           # kg/m3
                _mdot_steam = max(rkpower_total[t], 0.0) / _hfg_sv                # kg/s
                steam_vel_arr[t+1] = _mdot_steam / (_rhoV_sv * A_flow_core)
            except Exception:
                steam_vel_arr[t+1] = steam_vel_arr[t]
        else:
            steam_vel_arr[t+1] = 0.0

        # ── PATCH 8a: SG heat removal ─────────────────────────────────────────
        # SG isolation: the forced-flow SG heat sink can be isolated on normal
        # reactor scram/trip and/or on AMSAC/turbine trip.  The isolation ramp
        # applies only to forced SG heat removal.  The AFW-supported natural
        # circulation model below is retained and can remove decay heat after
        # the forced-flow SG path is isolated.
        _sg_isolation_candidates = []
        if sg_trip_isolation_enabled and _scram_t is not None:
            _sg_isolation_candidates.append(_scram_t)
        if _turbine_trip_detected and _turbine_trip_t is not None:
            _sg_isolation_candidates.append(_turbine_trip_t + trip_turbine_delay_s)

        _sg_isolation_start = (min(_sg_isolation_candidates)
                               if _sg_isolation_candidates else None)
        if _sg_isolation_start is not None and time[t] >= _sg_isolation_start:
            _sg_open_frac = max(0.0, 1.0 - (time[t] - _sg_isolation_start)
                                / max(trip_sgiv_stroke_s, 1.0e-9))
        else:
            _sg_open_frac = 1.0
        _sg_isolated = (_sg_open_frac == 0.0)

        # Current-timestep Churchill-Chu dominance predictor for SG natural
        # circulation.  This is deliberately NOT latched.  It mirrors the core
        # HTC regime logic using the known state at t so that SG natural
        # circulation is available only while the core is actually in the
        # Churchill-Chu-dominant, non-film-boiling regime.
        _sg_nat_active = False
        if core_flag:
            try:
                _mdot_sg_cc = max(pump_mdot[t], 0.0)
                try:
                    _rho_sg_cc = XSteam.rhoL_p(Pressure[t] * 1e-3)
                except Exception:
                    _rho_sg_cc = 700.0
                _v_sg_cc = _mdot_sg_cc / max(_rho_sg_cc * A_flow_core, 1e-12)

                _htc_db_sg = core_htc_db(Pressure[t], Temperature[t],
                                         _v_sg_cc, D_h_core)

                _P_kPa_sg  = Pressure[t] * 1e-3
                _T_cool_sg = Temperature[t] - 273.15
                _T_wall_sg = TT[N, t] - 273.15
                try:
                    _Tsat_sg   = XSteam.Tsat_p(_P_kPa_sg)
                    _T_lkp_sg  = min(_T_cool_sg, 0.9999 * _Tsat_sg)
                    _hf_sg     = XSteam.hL_T(_T_lkp_sg)
                    _rho_nc_sg = XSteam.rho_ph(_P_kPa_sg, _hf_sg)
                    _cp_nc_sg  = XSteam.cp_ph(_P_kPa_sg, _hf_sg) * 1e3
                    _lam_nc_sg = XSteam.tc_ph(_P_kPa_sg, _hf_sg)
                    _mu_nc_sg  = XSteam.my_ph(_P_kPa_sg, _hf_sg)
                    _Tbb_sg    = _T_cool_sg + 273.15
                    _beta_sg   = 1e-4 * (1.9607e-6*_Tbb_sg**3
                                          - 2.2615e-3*_Tbb_sg**2
                                          + 9.2045e-1*_Tbb_sg
                                          - 1.2301e2)
                    _dT_nc_sg  = max(abs(_T_wall_sg - _T_cool_sg), 0.1)
                    _Gr_sg     = (_rho_nc_sg**2 * 9.81 * _beta_sg
                                  * _dT_nc_sg * D_h_core**3 / _mu_nc_sg**2)
                    _Pr_sg     = _mu_nc_sg * _cp_nc_sg / _lam_nc_sg
                    _Ra_sg     = _Gr_sg * _Pr_sg
                    _Nu_sg     = (0.825 + 0.387*_Ra_sg**(1/6)
                                  / (1 + (0.492/_Pr_sg)**(9/16))**(8/27))**2
                    _htc_cc_sg = _Nu_sg * _lam_nc_sg / D_h_core
                except Exception:
                    _htc_cc_sg = 200.0

                # For SG natural circulation, use the primary saturation
                # temperature implied by RCS pressure rather than the lumped
                # coolant temperature.  This represents a sustained saturated
                # primary-side mixture/reflux-boiling regime during post-trip
                # cooldown until RHR is available.
                try:
                    _T_primary_sg_nat = XSteam.Tsat_p(Pressure[t] * 1e-3) + 273.15
                except Exception:
                    _T_primary_sg_nat = Temperature[t]

                _sg_nat_active = (
                    (_htc_cc_sg > _htc_db_sg)
                    and _reactor_scrammed
                    and not _film_boiling_active
                    and _T_primary_sg_nat > T_sec_K
                )
            except Exception:
                _sg_nat_active = False

        _sg_diag_Q_forced_raw = 0.0
        _sg_diag_Q_forced     = 0.0
        _sg_diag_Q_nat        = 0.0
        _sg_diag_flow_frac    = 0.0
        _sg_diag_v_ratio      = 0.0
        _sg_diag_UA_eff       = 0.0
        _sg_diag_C_primary    = 0.0
        sg_open_frac[t]       = _sg_open_frac
        sg_nat_active_arr[t]  = 1 if _sg_nat_active else 0

        if sg_flag:
            if sg_table_flag:
                # Prescribed Q_sg boundary condition from column D of the time
                # table [MW vs time].  Positive = SG removes heat from primary
                # (normal direction); negative = SG adds heat to primary
                # (overcooling / MSLB reversal).  No UA model; no caps.
                #
                # Even though the table bypasses the UA model, still calculate
                # the primary-side pump-flow heat-capacity diagnostics.  These
                # drive the UI SG flow-capacity status tile for LOHS/feedwater
                # cases that prescribe SG duty from the input table.
                if pump_flag:
                    try:
                        _P_kPa = Pressure[t]*1e-3
                        _hf    = XSteam.hL_T(min(Temperature[t]-273.15,
                                                 XSteam.Tsat_p(_P_kPa)*0.9999))
                        _cp_si = XSteam.cp_ph(_P_kPa, _hf)*1e3
                    except Exception:
                        _cp_si = 5500.0
                    try:
                        _sg_diag_v_ratio = max(0.0, _pump_state.get("vol_flow_m3s", 0.0) / _Q_R_SI)
                        _sg_diag_UA_eff  = UA_sg_rated * _sg_diag_v_ratio**0.8
                        _sg_diag_C_primary = _pump_state.get("mass_flow_kgs", 0.0) * _cp_si
                    except Exception:
                        _sg_diag_v_ratio = _sg_diag_UA_eff = _sg_diag_C_primary = np.nan

                # Scale by relative RCP flow: reduced flow → reduced SG duty.
                Q_sg_forced = 1e6 * float(np.interp(time[t], time_xls, q_sg_xls))
                if pump_flag and _pump_mdot_rated is not None and _pump_mdot_rated > 0:
                    _flow_frac = max(0.0, pump_mdot[t] / _pump_mdot_rated)
                    _sg_diag_flow_frac = _flow_frac
                    Q_sg_forced *= _flow_frac
                else:
                    _sg_diag_flow_frac = 1.0
                _sg_diag_Q_forced_raw = Q_sg_forced
                # SG isolation affects the forced/table heat sink only.
                Q_sg_forced *= _sg_open_frac
                _sg_diag_Q_forced = Q_sg_forced
                Q_sg[t] = Q_sg_forced
                if _sg_nat_active:
                    try:
                        _T_primary_sg_nat = XSteam.Tsat_p(Pressure[t] * 1e-3) + 273.15
                    except Exception:
                        _T_primary_sg_nat = Temperature[t]
                    Q_sg_nat = min(
                        sg_nat_frac * UA_sg_rated
                        * max(_T_primary_sg_nat - T_sec_K, 0.0),
                        _decay_heat_W)
                    _sg_diag_Q_nat = Q_sg_nat
                    Q_sg[t] = max(Q_sg_forced, Q_sg_nat)
            elif pump_flag:
                try:
                    _P_kPa = Pressure[t]*1e-3
                    _hf    = XSteam.hL_T(min(Temperature[t]-273.15,
                                             XSteam.Tsat_p(_P_kPa)*0.9999))
                    _cp_si = XSteam.cp_ph(_P_kPa, _hf)*1e3
                except Exception:
                    _cp_si = 5500.0
                # Forced-circulation SG heat removal from the existing
                # pump-flow/UA model.  sg_heat() already scales UA with
                # volumetric flow, but retain an explicit pumped-flow
                # degradation factor so the forced SG term fades with RCP
                # coastdown in the same way as the spreadsheet SG option.
                try:
                    _sg_diag_v_ratio = max(0.0, _pump_state.get("vol_flow_m3s", 0.0) / _Q_R_SI)
                    _sg_diag_UA_eff  = UA_sg_rated * _sg_diag_v_ratio**0.8
                    _sg_diag_C_primary = _pump_state.get("mass_flow_kgs", 0.0) * _cp_si
                except Exception:
                    _sg_diag_v_ratio = _sg_diag_UA_eff = _sg_diag_C_primary = np.nan
                Q_sg_forced_raw = sg_heat(_pump_state, Temperature[t],
                                          T_sec_K, UA_sg_rated, _cp_si)
                _sg_diag_Q_forced_raw = Q_sg_forced_raw
                if _pump_mdot_rated is not None and _pump_mdot_rated > 0.0:
                    _flow_frac = max(0.0, min(1.0, pump_mdot[t] / _pump_mdot_rated))
                else:
                    _flow_frac = 1.0
                _sg_diag_flow_frac = _flow_frac
                # SG isolation affects the forced-flow SG heat sink only.
                Q_sg_forced = Q_sg_forced_raw * _flow_frac * _sg_open_frac
                _sg_diag_Q_forced = Q_sg_forced
                Q_sg[t] = Q_sg_forced

                # Natural-circulation SG heat-removal floor.  This is
                # a current-timestep condition, not a latch: it is available
                # only while the core heat-transfer regime is predicted to be
                # Churchill-Chu dominant, the reactor is scrammed, and the core
                # is not in film boiling.  The driving temperature is the
                # primary saturation temperature implied by RCS pressure, not
                # the lumped coolant temperature, consistent with a sustained
                # saturated-mixture/reflux-boiling assumption.
                if _sg_nat_active:
                    try:
                        _T_primary_sg_nat = XSteam.Tsat_p(Pressure[t] * 1e-3) + 273.15
                    except Exception:
                        _T_primary_sg_nat = Temperature[t]

                    Q_sg_nat = min(
                        sg_nat_frac * UA_sg_rated
                        * max(_T_primary_sg_nat - T_sec_K, 0.0),
                        _decay_heat_W)
                    _sg_diag_Q_nat = Q_sg_nat
                    Q_sg[t] = max(Q_sg_forced, Q_sg_nat)

                # Do not multiply the combined SG removal by _sg_open_frac here;
                # the forced term has already been isolated.  Natural
                # circulation remains available by design.
            else:
                Q_sg[t] = 0.0
        else:
            Q_sg[t] = 0.0

        sg_v_ratio[t]      = _sg_diag_v_ratio
        sg_UA_eff[t]       = _sg_diag_UA_eff
        sg_C_primary[t]    = _sg_diag_C_primary
        sg_flow_frac[t]    = _sg_diag_flow_frac
        sg_Q_forced_raw[t] = _sg_diag_Q_forced_raw
        sg_Q_forced[t]     = _sg_diag_Q_forced
        sg_Q_nat[t]        = _sg_diag_Q_nat

        # ── SG flow-capacity warning ────────────────────────────────────────
        # The primary-side flow-capacity limit is C_primary*dT.  If the core
        # power is more than 20% above this limit, increasing UA or SG surface
        # area cannot resolve the heat balance in this lumped model; the pump
        # flow scale or the requested power level is inconsistent with the run.
        try:
            _sg_cap_dT = max(float(sg_dT_primary_sec[t]), 0.0)
            _sg_cap_W = float(_sg_diag_C_primary) * _sg_cap_dT
            if np.isfinite(_sg_cap_W) and _sg_cap_W > 0.0:
                sg_flow_cap_W[t] = _sg_cap_W
                _sg_cap_ratio = float(rkpower_total[t]) / _sg_cap_W
                sg_flow_cap_ratio[t] = _sg_cap_ratio
                if np.isfinite(_sg_cap_ratio) and _sg_cap_ratio > 1.20:
                    sg_flow_cap_warn[t] = 1
                    _sg_flow_cap_warn_count += 1
                    if not _sg_flow_cap_warned:
                        _sg_flow_cap_warned = True
                        print(
                            "WARNING: Core power exceeds the SG primary-flow "
                            "capacity limit by more than 20%. "
                            f"t={time[t]:.3g} s, RK power={rkpower_total[t]/1.0e6:.3g} MW, "
                            f"SG flow-capacity limit={_sg_cap_W/1.0e6:.3g} MW, "
                            f"ratio={_sg_cap_ratio:.3f}. If the calculation fails or heats up, "
                            "revise the pump model/flow scaling or lower the power level.",
                            flush=True,
                        )
        except Exception:
            pass

        # ── normal pressurizer pressure-control surrogate: heater/spray ─────
        # Positive Q adds heat to the RCS/pressurizer surrogate when pressure is
        # below setpoint. Negative Q represents spray/letdown cooling when above
        # setpoint. The proportional command is limited by user-defined heater
        # and spray capacities, then first-order filtered by pzr_control_tau_s.
        if pzr_control_enabled and (pzr_control_heater_MW > 0.0 or pzr_control_spray_MW > 0.0):
            _p_err_kPa = pzr_control_setpoint_kPa - Pressure[t] * 1.0e-3
            _u_ctl = max(-1.0, min(1.0, _p_err_kPa / pzr_control_band_kPa))
            if _u_ctl >= 0.0:
                _pzr_Q_cmd_W = _u_ctl * pzr_control_heater_MW * 1.0e6
            else:
                _pzr_Q_cmd_W = _u_ctl * pzr_control_spray_MW * 1.0e6
            try:
                _Dt_ctl = max(float(time[min(t+1, number_timesteps-1)] - time[t]), 0.0)
            except Exception:
                _Dt_ctl = timestep
            _alpha_ctl = 1.0 - np.exp(-_Dt_ctl / pzr_control_tau_s)
            _alpha_ctl = max(0.0, min(1.0, _alpha_ctl))
            _pzr_control_Q_W = _pzr_control_Q_W + _alpha_ctl * (_pzr_Q_cmd_W - _pzr_control_Q_W)
            net_heat_total[t] += _pzr_control_Q_W

        # ── scaling ratios ────────────────────────────────────────────────────
        current_scaling_ratios = compute_scaling_ratios(
            prop_partial_derivatives, mix_properties, Pressure[t], Total_Mass,
            massflow_break[t], inflow_avg, stagnation_enthalpy_break,
            enthalpy_in, net_heat_total[t])
        dilation[t] = current_scaling_ratios["dilation"]
        pow_rat[t]  = current_scaling_ratios["power_ratio"]
        e_rat[t]    = current_scaling_ratios["energyflow_ratio"]

        if determine_scaling["number"] <= determine_scaling["max_number"]:
            if t == 0:
                scaling_ratios.append(current_scaling_ratios)
                determine_scaling["number"] += 1
            else:
                if determine_scaling["reach_sat_flag"] < 1 and x_eq[t] > 0:
                    scaling_ratios.append(current_scaling_ratios)
                    determine_scaling["reach_sat_flag"] = 10
                    determine_scaling["number"] += 1
                elif determine_scaling["reach_Mtrans_flag"] < 1 and flag_vent_vapor > 1:
                    scaling_ratios.append(current_scaling_ratios)
                    determine_scaling["reach_Mtrans_flag"] = 10
                    determine_scaling["number"] += 1
                elif determine_scaling["open_ADV_flag"] < 1 and flag_ADV_open > 1:
                    scaling_ratios.append(current_scaling_ratios)
                    determine_scaling["open_ADV_flag"] = 10
                    determine_scaling["number"] += 1

        # ── pressurizer level tracking ────────────────────────────────────────
        _pzr_level_prev = pzr_level_arr[t-1] if t > 0 else pzr_level_init
        _pzr_active = False   # pressure-hold removed; PORV hysteresis governs

        # ── 2×2 linear solve with PORV predictor-corrector ───────────────────
        if t < number_timesteps:
            A = np.array([
                [Total_Mass*prop_partial_derivatives["internal_energy"]["wrt_pressure"],
                 Total_Mass*prop_partial_derivatives["internal_energy"]["wrt_enthalpy"]],
                [Total_Mass*prop_partial_derivatives["specific_volume"]["wrt_pressure"],
                 Total_Mass*prop_partial_derivatives["specific_volume"]["wrt_enthalpy"]]
            ])
            _Dt   = time[t+1] - time[t]
            _e    = mix_properties["internal_energy"]
            _v    = mix_properties["specific_volume"]

            def _make_b(m_porv, h_porv=None):
                if h_porv is None:
                    h_porv = enthalpy_mix[t]
                _m_inj = massflow_in[t] + cvcs_mdot[t] + si_pumped_mdot[t]   # acc + SI table + HPSI/LPSI
                # Surge is internal redistribution within the RCS lumped volume;
                # treated as diagnostic only — not fed back into mass/energy balance.
                return np.array([
                    _m_inj*(enthalpy_in - _e) +
                    si_pumped_mdot[t]*(_h_inj - enthalpy_in) -
                    massflow_break[t]*(stagnation_enthalpy_break - _e) -
                    m_porv*(h_porv - _e) +
                    net_heat_total[t] - Q_sg[t],
                    -_v*(_m_inj - massflow_break[t] - m_porv)
                ])

            def _solve_b(b_vec):
                try:
                    u = (np.array([Pressure[t], enthalpy_mix[t]])
                         + _Dt * np.linalg.solve(A, b_vec))
                    P_new = float(np.real(u[0]))
                    h_new = float(np.real(u[1]))
                    # Guard: if solve produced unphysical values, fall back
                    if (not np.isfinite(P_new) or not np.isfinite(h_new)
                            or P_new < 0 or P_new > 1e9 or h_new < 0):
                        return Pressure[t], enthalpy_mix[t] + _Dt*b_vec[0]/max(Total_Mass, 1.0)
                    return P_new, h_new
                except np.linalg.LinAlgError:
                    return Pressure[t], enthalpy_mix[t] + _Dt*b_vec[0]/max(Total_Mass, 1.0)

            # Step 1: solve without PORV.
            # Always use the full 2x2 solver — even when b[1]=0 (no primary
            # mass flows) the volume constraint couples pressure to enthalpy:
            #   A[1,0]*dP/dt + A[1,1]*dh/dt = 0
            # The old shortcut (hold P constant when b[1]=0) suppressed the
            # pressure drop that occurs when cooling a sealed RCS, e.g. during
            # an MSLB.  For a truly static case (b=[0,0]) the solver returns
            # [0,0] so pressure correctly stays constant.
            b_none = _make_b(0.0)
            P_none, h_none = _solve_b(b_none)

            if _active_area > 0 and b_none[1] == 0:
                choked = 0

            # ── PORV hysteresis and stroke dynamics ──────────────────────
            # Architecture: RCS solver predicts P_none (unrelieved), PORV
            # senses P_sense = max(Pressure[t], P_none) so it responds to
            # where pressure is heading, not just where it was.  Flow uses
            # _h_porv, selected above from the represented PORV source region.
            # Final RCS solve repeats with mdot_PORV as external mass loss.
            # Under-relaxation only when PORV is configured (PORV_area_m2 > 0).
            if PORV_area_m2 > 0:
                _P_sense = max(Pressure[t], P_none)
                _setpt_Pa  = PORV_setpoint_kPa * 1e3
                _reseat_Pa = PORV_reseat_kPa   * 1e3
                # Hysteresis transitions
                if _porv_mode == "closed" and _P_sense >= _setpt_Pa:
                    _porv_mode = "opening"
                elif _porv_mode in ("open", "opening") and _P_sense <= _reseat_Pa:
                    _porv_mode = "closing"
                elif _porv_mode == "closing" and _P_sense >= _setpt_Pa:
                    _porv_mode = "opening"
                # Stroke dynamics
                if _porv_mode in ("opening", "open"):
                    _x_porv = min(1.0, _x_porv + _Dt / max(PORV_tau_open,  1e-6))
                else:
                    _x_porv = max(0.0, _x_porv - _Dt / max(PORV_tau_close, 1e-6))
                # Mode snap from stroke limits
                if _x_porv >= 1.0:
                    _porv_mode = "open"
                elif _x_porv <= 0.0:
                    _porv_mode = "closed"
                # Flow: use _P_sense as upstream pressure so flow responds
                # to predicted unrelieved pressure, not lagged Pressure[t].
                # Cap the one-step mass removal to keep the lumped RCS state
                # inside its physical validity range.
                if _x_porv > 0.0:
                    _pf_pc = PORV_Cd * critical_flow_ERM(
                                 _P_sense, _h_porv,
                                 sat_liquid, sat_vapor)
                    _m_porv_raw = _x_porv * PORV_area_m2 * _pf_pc
                    _m_porv_cap = PORV_max_frac_mass_per_step * max(Total_Mass, 0.0) / max(_Dt, 1e-9)
                    massflow_PORV[t] = min(max(_m_porv_raw, 0.0), _m_porv_cap)
                else:
                    massflow_PORV[t] = 0.0
            else:
                massflow_PORV[t] = 0.0

            # Final RCS solve with PORV as external mass loss
            if massflow_PORV[t] > 0.0:
                P_new, h_new = _solve_b(_make_b(massflow_PORV[t], _h_porv))
            else:
                P_new, h_new = P_none, h_none

            # Under-relaxation: only when PORV is configured
            if massflow_PORV[t] > 0.0 and P_relax < 1.0:
                _dP_now = abs(P_new - Pressure[t])
                if P_relax_dP_ref > 0.0 and _Dt > 0.0:
                    _dP_rate_kPas = _dP_now / (_Dt * 1e3)
                    _omega = P_relax / max(1.0, _dP_rate_kPas / P_relax_dP_ref)
                    _omega = max(0.05, min(1.0, _omega))
                else:
                    _omega = P_relax
                P_new = (1.0 - _omega) * Pressure[t] + _omega * P_new

            # ── optional normal pressurizer pressure-control surrogate ───
            # Applies only when explicitly enabled.  This is not PORV relief;
            # it softly damps pressure excursions above the normal-control band
            # toward the specified setpoint.  It is intended for startup and
            # normal power-ascension cases that previously relied on the removed
            # pressure-hold model.
            if (pzr_control_enabled
                    and pzr_control_heater_MW <= 0.0
                    and pzr_control_spray_MW <= 0.0):
                # Legacy pressure-relaxation-only controller, retained only
                # when the new explicit heater/spray capacities are not used.
                _pzr_ctl_target = pzr_control_setpoint_kPa * 1.0e3
                _pzr_ctl_upper  = (pzr_control_setpoint_kPa + pzr_control_band_kPa) * 1.0e3
                if P_new > _pzr_ctl_upper:
                    _pzr_ctl_alpha = 1.0 - np.exp(-_Dt / pzr_control_tau_s)
                    _pzr_ctl_alpha = max(0.0, min(1.0, _pzr_ctl_alpha))
                    P_new = P_new + _pzr_ctl_alpha * (_pzr_ctl_target - P_new)

            # ── optional pressurizer/RCS compliance slew limiter ───────
            # Limits abrupt pressure collapse/rise in intact-RCS normal
            # transients without replacing the slower heater/spray controller.
            # Disabled by default. If pzr_compliance_intact_only=1, the limiter
            # is bypassed for LOCA/break cases and whenever PORV is flowing.
            if pzr_compliance_enabled:
                _compliance_allowed = True
                if pzr_compliance_intact_only:
                    _compliance_allowed = (
                        diameter_break <= 0.0
                        and massflow_break[t] <= 1.0e-9
                        and massflow_PORV[t] <= 1.0e-9
                    )
                if _compliance_allowed and _Dt > 0.0:
                    _dP_raw = P_new - Pressure[t]
                    _dP_min = -pzr_max_depress_rate_kPa_s * 1.0e3 * _Dt
                    _dP_max =  pzr_max_press_rate_kPa_s   * 1.0e3 * _Dt
                    _dP_limited = max(_dP_min, min(_dP_raw, _dP_max))
                    P_new = Pressure[t] + _dP_limited

            # Preserve the unrelaxed predictor pressure for reporting only.
            # Subsequent relaxation may alter P_new for the simulation state,
            # but the CSV/figures can still show the raw 2x2 pressure response.
            _P_unrelaxed_report = P_new

            # ── optional RCS enthalpy relaxation / thermal inertia ───────
            # Smooths abrupt HEM enthalpy changes in intact transients.  This
            # acts upstream of the steam-table lookup, so it also smooths the
            # apparent saturation/quality transition without directly damping
            # diagnostic x_eq or void fraction.
            if rcs_enthalpy_relax_tau_s > 0.0:
                _h_relax_allowed = True
                if rcs_enthalpy_relax_intact_only:
                    _h_relax_allowed = (
                        diameter_break <= 0.0
                        and massflow_break[t] <= 1.0e-9
                        and massflow_PORV[t] <= 1.0e-9
                    )
                if _h_relax_allowed and _Dt > 0.0:
                    _alpha_h_relax = _Dt / (rcs_enthalpy_relax_tau_s + _Dt)
                    _alpha_h_relax = max(0.0, min(1.0, _alpha_h_relax))
                    h_new = enthalpy_mix[t] + _alpha_h_relax * (h_new - enthalpy_mix[t])

                    # Re-solve pressure using the same fractional movement for
                    # consistency with the semi-implicit P-h state update.
                    # This is intentionally a relaxation, not a thermodynamic
                    # constraint; it preserves direction while adding inertia.
                    P_new = Pressure[t] + _alpha_h_relax * (P_new - Pressure[t])

            # ── pressurizer level tracking (always when pzr exists) ──────
            if massflow_PORV[t] > 0:
                _pzr_hold_disabled = True

            Pressure[t+1]        = P_new
            Pressure_report[t+1] = _P_unrelaxed_report
            enthalpy_mix[t+1]    = h_new
            # Clamp to XSteam valid range before property lookup
            # P: 1 kPa (near vacuum) to 100 MPa;  h: 1 kJ/kg to 4500 kJ/kg
            _P_xst = float(np.clip(Pressure[t+1] * 1e-3, 1.0, 100000.0))
            _h_xst = float(np.clip(enthalpy_mix[t+1] * 1e-3, 1.0, 4500.0))
            _T_xst = XSteam.T_ph(_P_xst, _h_xst)
            Temperature[t+1] = (_T_xst + 273.15 if np.isfinite(_T_xst)
                                 else Temperature[t])

            # ── Pressurizer level diagnostic (algebraic) ─────────────────
            # Diagnostic only — not fed back into mass/energy balance.  This
            # must be evaluated after Temperature[t+1] is valid; otherwise the
            # uninitialized next-step temperature can spuriously collapse level
            # to zero on the first step.
            _M_porv_loss_cum += massflow_PORV[t] * _Dt
            if pzr_vol > 0:
                try:
                    _rhoL_pzr = XSteam.rhoL_p(max(P_new * 1e-3, 1.0))
                except Exception:
                    _rhoL_pzr = 700.0
                _T_avg     = Temperature[t+1]
                _dV_therm  = pzr_K_vol_T * (_T_avg - pzr_T0)
                _dV_loss   = _M_porv_loss_cum / max(_rhoL_pzr, 1.0)
                _new_level = pzr_level_init + (_dV_therm - _dV_loss) / max(pzr_area, 1e-9)
                _new_level = max(0.0, min(_new_level, pzr_height))
                _m_surge   = 0.0   # diagnostic only
            else:
                _m_surge   = 0.0
                _new_level = pzr_level_init

            _t_next = t + 1
            pzr_level_arr[_t_next]      = _new_level
            pzr_level_norm_arr[_t_next] = (_new_level / pzr_height
                                            if pzr_height > 0 else 0.0)
            pzr_mdot_surge_arr[_t_next] = _m_surge

        # ── PORV-triggered scram (evaluated after predictor-corrector) ─────────
        # Fires when PORV first provides flow AND scram_on_PORV=1.
        # Also fires via the unified high-pressure trip (trip_P_hi_kPa) which
        # is set to PORV_setpoint_kPa when scram_on_PORV=1 (see parsing block).
        # This path handles the post-predictor-corrector timing correctly.
        if (scram_on_PORV and not _porv_scram_done
                and PORV_area_m2 > 0 and massflow_PORV[t] > 0):
            _porv_scram_done = True
            if not _reactor_scrammed and _trip_detected_t is None:
                _trip_detected_t    = time[t]
                _trip_reason_pending = (f"PORV  P={Pressure[t]/1e3:.1f} kPa")
                if trip_delay > 0:
                    print(f"Trip signal at t={time[t]:.2f} s:"
                          f" PORV flow  P={Pressure[t]/1e3:.1f} kPa"
                          f"  (scram in {trip_delay:.1f} s)")

        # ── heat structure ────────────────────────────────────────────────────
        DELTA_t = time[t+1] - time[t]

        if core_flag:
            # ── Core model: fission+decay heat drives the clad slab ──────────
            QQQ = rkpower_total[t]   # W -- total fission scram + decay heat

            # Core HTC: D-B for all forced-flow levels; single-phase
            # Churchill-Chu natural convection when pump is stopped.
            # NatConvHTC (which fires Forster-Zuber when Tw>Tsat) is
            # excluded — the core is assumed covered throughout.
            _mdot = pump_mdot[t+1]
            try:
                _rho = XSteam.rhoL_p(Pressure[t+1] * 1e-3)
            except Exception:
                _rho = 700.0
            _v_core = _mdot / (_rho * A_flow_core)

            # Always compute both mechanisms and take the larger.
            # D-B and Churchill-Chu act simultaneously; whichever dominates
            # wins, giving a smooth natural handoff as flow decays with no
            # switch point or step change.

            # Dittus-Boelter (forced convection; laminar Nu=3.66 near zero flow)
            _htc_db = core_htc_db(Pressure[t+1], Temperature[t+1],
                                   _v_core, D_h_core)

            # Churchill-Chu single-phase natural convection (buoyancy-driven)
            _P_kPa  = Pressure[t+1] * 1e-3
            _T_cool = Temperature[t+1] - 273.15
            _T_wall = TT[N, t] - 273.15
            try:
                _Tsat   = XSteam.Tsat_p(_P_kPa)
                _T_lkp  = min(_T_cool, 0.9999 * _Tsat)
                _hf     = XSteam.hL_T(_T_lkp)
                _rho_nc = XSteam.rho_ph(_P_kPa, _hf)
                _cp_nc  = XSteam.cp_ph(_P_kPa, _hf) * 1e3
                _lam_nc = XSteam.tc_ph(_P_kPa, _hf)
                _mu_nc  = XSteam.my_ph(_P_kPa, _hf)
                _Tbb    = _T_cool + 273.15
                _beta   = 1e-4 * (1.9607e-6*_Tbb**3 - 2.2615e-3*_Tbb**2
                                  + 9.2045e-1*_Tbb - 1.2301e2)
                _dT_nc  = max(abs(_T_wall - _T_cool), 0.1)  # floor dT at 0.1 K
                _Gr     = _rho_nc**2 * 9.81 * _beta * _dT_nc * D_h_core**3 / _mu_nc**2
                _Pr     = _mu_nc * _cp_nc / _lam_nc
                _Ra     = _Gr * _Pr
                _Nu     = (0.825 + 0.387*_Ra**(1/6)
                           / (1 + (0.492/_Pr)**(9/16))**(8/27))**2
                _htc_cc = _Nu * _lam_nc / D_h_core
            except Exception:
                _htc_cc = 200.0   # conservative fallback [W/m²K]

            _htc_target = max(_htc_db, _htc_cc) * htc_core_mult
            # Track whether Churchill-Chu dominates for SG natural-circ logic.
            # Only true when CC > DB AND not in any film boiling regime.
            _cc_dominant = (_htc_cc > _htc_db) and not _film_boiling_active

            # ── Film boiling override (post-CHF — Bromley correlation) ─────────
            # Once DNBR[t] has dropped below 1.0, the core is in film boiling.
            # The Bromley (1950) pool film boiling correlation is used rather
            # than vapor-phase D-B because:
            #   1. Bromley gives physically representative HTCs (200-2000 W/m²K)
            #      vs vapor D-B which gives 5000-17000 W/m²K at near-critical
            #      conditions due to anomalously high steam k and cp.
            #   2. Bromley naturally collapses as h_fg → 0 near the critical
            #      point, producing increasing ΔT_wall → T_fuel → 1500°C.
            # Film boiling is irreversible — flag stays set once triggered.
            #
            # Bromley: h_fb = 0.62*[λ_v³*ρ_v*(ρ_L-ρ_v)*g*h_fg/(μ_v*D*ΔT)]^0.25
            # Implicit form (no ΔT needed):
            #   h_fb = C_B^(4/3) / q"_avg^(1/3)
            #   where C_B = 0.62 * [λ_v³*ρ_v*(ρ_L-ρ_v)*g*h_fg/(μ_v*D_h)]^0.25
            # ── Post-CHF regime selection ─────────────────────────────────
            # Detect reflood: level rising while uncovery active
            if _fb_uncover and t > 0 and not _reflood_started:
                if (not np.isnan(ves_ll[t]) and not np.isnan(ves_ll[t-1])
                        and ves_ll[t] > ves_ll[t-1]):
                    _reflood_started = True

            # Reflood velocity = rate of level rise [m/s]
            if (_reflood_started and t > 0
                    and not np.isnan(ves_ll[t]) and not np.isnan(ves_ll[t-1])):
                _dt_ll = time[t] - time[t-1] if t > 0 else 1.0
                _v_reflood = max((ves_ll[t] - ves_ll[t-1]) / max(_dt_ll, 1e-9),
                                  0.01)
            else:
                # Use boil-off steam velocity when computed; fall back to the
                # user-supplied steam_velocity input only if not yet available
                # (t=0) or outside film-boiling conditions.
                _v_reflood = (steam_vel_arr[t] if steam_vel_arr[t] > 0.0
                              else steam_velocity)

            if _film_boiling_active and not np.isnan(Pressure[t+1]):
                _P_kPa_fb = Pressure[t+1] * 1e-3

                # Average core heat flux [W/m²] used for correlation selection.
                # Bromley is a pool-boiling correlation validated only above
                # ~10 000 W/m².  Below that threshold the core is in the
                # low-flux steam-cooling regime and single-phase steam
                # Dittus-Boelter is the more appropriate model.
                _q_avg_W = max(rkpower_total[t] / SurfArea, 1.0)
                _bromley_flux_threshold = 1.0e4   # W/m²
                _blend = None   # set only in hot_leg transition band

                if post_chf_model == "hot_leg":
                    # Hot-leg post-CHF model - four sub-cases in priority order:
                    #   1. Reflood detected              -> steam Dittus-Boelter
                    #   2. q_avg < 10 000 W/m2           -> steam Dittus-Boelter
                    #   3. 10 000 <= q_avg < 11 000 W/m2 -> linear blend
                    #      (smooth transition; avoids step discontinuity at threshold)
                    #   4. q_avg >= 11 000 W/m2          -> Bromley IAFB
                    _bromley_flux_upper = 1.1e4   # W/m2 - top of blend band
                    if _reflood_started or _q_avg_W < _bromley_flux_threshold:
                        _htc_fb = core_htc_db_steam(
                                      Pressure[t+1], _v_reflood, D_h_core) * htc_fb_mult
                    elif _q_avg_W < _bromley_flux_upper:
                        # Blend fraction: 0 at lower edge (pure steam D-B),
                        #                 1 at upper edge (pure Bromley)
                        _blend = (_q_avg_W - _bromley_flux_threshold) / (
                                   _bromley_flux_upper - _bromley_flux_threshold)
                        _htc_db_blend = core_htc_db_steam(
                                      Pressure[t+1], _v_reflood, D_h_core) * htc_fb_mult
                        _use_bromley = True   # compute Bromley below, then blend
                        _htc_fb = _htc_db_blend   # stash DB value; overwritten after blend
                    else:
                        _use_bromley = True   # fall through to pure Bromley below
                        _htc_fb = None

                elif post_chf_model == "cold_leg":
                    # Cold-leg: Bromley while level >= film_boiling_level_m;
                    # steam DB once level drops below top of active fuel
                    if (not np.isnan(ves_ll[t])
                            and ves_ll[t] < film_boiling_level_m):
                        _htc_fb = core_htc_db_steam(
                                      Pressure[t+1], _v_reflood, D_h_core) * htc_fb_mult
                    else:
                        _use_bromley = True
                        _htc_fb = None

                else:  # "bromley" (default)
                    _use_bromley = True
                    _htc_fb = None

                # Bromley inverted-annular film boiling.
                # Entered for pure Bromley (_htc_fb is None) and for the
                # transition blend band (_htc_fb holds the D-B value, _blend set).
                if _htc_fb is None or _blend is not None:
                    try:
                        _rhoV_fb = XSteam.rhoV_p(_P_kPa_fb)
                        _rhoL_fb = XSteam.rhoL_p(_P_kPa_fb)
                        _kV_fb   = XSteam.tcV_p(_P_kPa_fb)
                        _hfg_fb  = max((XSteam.hV_p(_P_kPa_fb)
                                        - XSteam.hL_p(_P_kPa_fb)) * 1e3, 0.0)
                        _muV_raw = XSteam.my_ph(_P_kPa_fb,
                                                 XSteam.hV_p(_P_kPa_fb))
                        _muV_fb  = np.clip(_muV_raw, 1e-6, 1e-3)
                        _drho    = max(_rhoL_fb - _rhoV_fb, 0.0)
                        _numer   = (_kV_fb**3 * _rhoV_fb * _drho
                                    * 9.81 * _hfg_fb)
                        _denom   = _muV_fb * D_h_core
                        if _numer > 0 and _denom > 0:
                            _C_B        = 0.62 * (_numer / _denom)**0.25
                            # _q_avg_W already computed above for flux-based selector.
                            # hot_leg: htc_post_chf cap removed — Bromley result used
                            # directly (cap was suppressing physically valid high HTCs).
                            # cold_leg / bromley: cap retained — htc_post_chf=50 W/m2K
                            # for cold_leg is the design-basis blowdown model and is
                            # required for numerical stability of the slab solve.
                            _htc_bromley_raw = htc_fb_mult * _C_B**(4/3) / _q_avg_W**(1/3)
                            if post_chf_model == "hot_leg":
                                _htc_bromley = max(_htc_bromley_raw, 50.0)
                            else:
                                _htc_bromley = min(max(_htc_bromley_raw, 50.0),
                                                   htc_post_chf)
                        else:
                            _htc_bromley = 50.0
                        if _blend is not None:
                            # Transition band: linear blend between D-B and Bromley
                            _htc_fb = (1.0 - _blend) * _htc_fb + _blend * _htc_bromley
                        else:
                            _htc_fb = _htc_bromley
                    except Exception:
                        if _htc_fb is None:
                            _htc_fb = 50.0
                        # else: keep the D-B value already in _htc_fb

                # Cold-leg: under-relax the transition to 50 W/m²K for a
                # physically reasonable ad-hoc film boiling result.
                # All other models: apply Bromley HTC directly.
                _alpha_prev = alpha[t] if (t > 0 and not np.isnan(alpha[t])) else _htc_fb
                if post_chf_model == "cold_leg":
                    alpha[t+1] = (1.0 - htc_relax) * _alpha_prev + htc_relax * _htc_fb
                else:
                    alpha[t+1] = _htc_fb
                _htc_target = _htc_fb

            else:
                # Under-relax to damp property-driven step-to-step jumps
                _alpha_prev = alpha[t] if (t > 0 and not np.isnan(alpha[t])) else _htc_target
                alpha[t+1]  = (1.0 - htc_relax) * _alpha_prev + htc_relax * _htc_target
        else:
            # ── Original path: tabulated power, pump-loop or nat-conv HTC ────
            QQQ = power_input[t]

            _htc_forced = htc(_pump_state, Pressure[t+1], Temperature[t+1],
                                           D_h=pump_D_h, L_heat=pump_L_heat)
            if _htc_forced > 1.0:
                alpha[t+1] = _htc_forced
            else:
                alpha[t+1] = NatConvHTC(TT[N,t]-273.15, Temperature[t+1]-273.15,
                                         Pressure[t+1]/1e3, 1.6)

        # User-defined dry-core heat-transfer handoff.
        # This operates before vessel inventory is zero.  It deliberately does
        # NOT stop the hydraulic calculation; accumulator/ECCS delivery and
        # break flow continue, but the core-to-coolant HTC is reduced to the
        # dry-core value so decay heat primarily heats the fuel/clad.
        if core_flag and dry_core_trigger_level_m >= 0.0:
            try:
                if (not _dry_core_ht_latched and np.isfinite(ves_ll[t])
                        and ves_ll[t] <= dry_core_trigger_level_m):
                    _dry_core_ht_latched = True
                    print(f"Dry-core heat-transfer mode started at t={time[t]:.2f} s: "
                          f"vessel level {ves_ll[t]:.3f} m <= trigger "
                          f"{dry_core_trigger_level_m:.3f} m; accumulator flow "
                          f"{massflow_in[t]:.3f} kg/s.")
                elif (_dry_core_ht_latched and np.isfinite(ves_ll[t])
                        and ves_ll[t] >= dry_core_trigger_level_m + film_boiling_level_margin_m):
                    # Level has recovered above trigger + margin — liquid has
                    # returned to the core region; release dry-core override and
                    # return to the film-boiling / steam D-B heat transfer path.
                    _dry_core_ht_latched = False
                    print(f"Dry-core heat-transfer mode cleared at t={time[t]:.2f} s: "
                          f"vessel level {ves_ll[t]:.3f} m >= "
                          f"{dry_core_trigger_level_m + film_boiling_level_margin_m:.3f} m.")
            except Exception:
                pass

        if core_flag and _dry_core_ht_latched:
            alpha[t+1] = dry_core_htc_W_m2K
            _cc_dominant = False
            _sg_nat_active = False

        htcm = 1

        if cond_flag == 0 and not core_flag:
            # Short-circuit: zero thermal resistance (power direct to fluid)
            alpha[t+1] = 100000
            DELTA_x    = 0.001

        AT = np.zeros((N+1, N+1)); bT = np.zeros((N+1, 1))
        AT[0,1] = -lambda_ss*DELTA_t/DELTA_x
        AT[0,0] = rho_ss*cp_ss*DELTA_x - AT[0,1]
        bT[0,0] = rho_ss*cp_ss*DELTA_x*TT[0,t] + (QQQ/SurfArea)*DELTA_t
        for i in range(1, N):
            Gm = rho_ss*cp_ss*DELTA_x
            AT[i,i-1] = -lambda_ss*DELTA_t/(2*DELTA_x)
            AT[i,i+1] = -lambda_ss*DELTA_t/(2*DELTA_x)
            AT[i,i]   = Gm - AT[i,i-1] - AT[i,i+1]
            bT[i,0]   = (-AT[i,i-1]*TT[i-1,t]+(Gm+AT[i,i-1]+AT[i,i+1])*TT[i,t]
                         -AT[i,i+1]*TT[i+1,t])
        AT[N,N-1] = -lambda_ss*DELTA_t/DELTA_x
        AT[N,N]   = rho_ss*cp_ss*DELTA_x + htcm*alpha[t+1]*DELTA_t - AT[N,N-1]
        bT[N,0]   = rho_ss*cp_ss*DELTA_x*TT[N,t] + htcm*alpha[t+1]*Temperature[t+1]*DELTA_t
        TT[:,t+1] = np.linalg.solve(AT, bT).flatten()
        TTwall[t+1] = TT[N,t+1]
        # Clamp average clad surface temperature at the BDB diagnostic
        # temperature ceiling.  This is not a material model; it prevents
        # nonphysical temperature runaway after fuel/clad damage in dry-core
        # sensitivities.
        if TTwall[t+1] > _T_melt_cap_K:
            TTwall[t+1] = _T_melt_cap_K

        cond_heat[t+1] = SurfArea*alpha[t+1]*(TTwall[t+1]-Temperature[t+1])

        # When the fuel/clad temperatures have reached the BDB temperature
        # ceiling, avoid driving the coolant equation with an unphysical
        # high-flux term from capped damaged fuel.
        if (not np.isnan(T_fuel_arr[t]) and T_fuel_arr[t] >= _T_melt_cap_K - 5.0
                and cond_heat[t+1] > 0):
            cond_heat[t+1] = 0.0

        # ── DNBR (Biasi/Zuber)  ───────────────────────────────────────────────
        # All quantities indexed at [t]: x_eq, Pressure, Temperature, and
        # rkpower_total are all computed for the current timestep t.
        # pump_mdot[t+1] is available because the pump step runs before this.
        if core_flag:
            # Average core mass flux  [kg/m²s]
            _G_core = pump_mdot[t+1] / A_flow_core if A_flow_core > 0 else 0.0
            # Hot-pin heat flux = F_r × average  [W/m²]
            _q_avg  = rkpower_total[t] / SurfArea if SurfArea > 0 else 0.0
            _q_hot  = F_r * F_z * _q_avg
            q_hot[t] = _q_hot
            # CHF: Biasi/Zuber below 13 MPa, Bowring above 15 MPa,
            # linear blend in between.  Bowring is used where H(P)→0
            # renders Biasi invalid (high-pressure ATWS/PORV scenarios).
            # Minimum G guard: Biasi/Zuber/Bowring are only valid for
            # G > ~100 kg/m²s. Below this threshold CHF → 0 artificially
            # and DNBR becomes meaningless. Churchill-Chu natural convection
            # handles the low-flow HTC; set DNBR = NaN to indicate
            # the correlation is inapplicable rather than reporting false DNB.
            _G_MIN_CHF = 100.0   # [kg/m²s] minimum valid mass flux
            if _G_core < _G_MIN_CHF:
                q_chf[t] = np.nan
                DNBR[t]  = np.nan
            else:
                _P_lo = 13.0e6    # [Pa] pure Biasi below this
                _P_hi = 15.0e6    # [Pa] pure Bowring above this
                _q_biasi = chf_biasi_zuber(
                    Pressure[t], _G_core, x_eq[t],
                    D_h_core, Temperature[t]
                )
                _q_bowr  = chf_bowring(
                    Pressure[t], _G_core, x_eq[t],
                    D_h_core, L_heated, Temperature[t]
                )
                if Pressure[t] <= _P_lo:
                    _q_chf = _q_biasi
                elif Pressure[t] >= _P_hi:
                    _q_chf = _q_bowr
                else:
                    _w = (Pressure[t] - _P_lo) / (_P_hi - _P_lo)
                    _qa = _q_biasi if not np.isnan(_q_biasi) else _q_bowr
                    _qb = _q_bowr  if not np.isnan(_q_bowr)  else _q_biasi
                    _q_chf = (1.0 - _w) * _qa + _w * _qb
                q_chf[t] = _q_chf
                DNBR[t]  = (_q_chf / _q_hot) if _q_hot > 1.0 else np.nan
            # DNB failure count: Gaussian radial power distribution N(mu=1, sigma=sigma_r).
            # Interpret DNBR[t] as the representative/mean rod DNB margin.
            # A rod at local power factor f enters DNB when f > dnb_threshold_factor.
            #
            # Therefore:
            #   DNBR = 1.00 -> about 50% of rods exceed the threshold
            #   DNBR = 0.98 -> slightly more than 50% exceed the threshold
            #   DNBR = 1.02 -> slightly less than 50% exceed the threshold
            #
            # Avoid "critical power" terminology here because it has a specific
            # BWR meaning; this is a PWR DNB threshold factor.
            if not np.isnan(DNBR[t]) and _sigma_r and _sigma_r > 0:
                _dnb_threshold_factor = DNBR[t]
                _z_dnb = (_dnb_threshold_factor - 1.0) / _sigma_r
                N_fail_DNB[t] = N_pins * float(_norm_dist.sf(_z_dnb))

        # Cold-leg: all rods assumed failed DNB at subcooling loss.
        # _pump_tripped guards against firing during the trip-delay window
        # while the pump is still delivering rated flow at full power.
        if (post_chf_model == "cold_leg" and core_flag
                and _active_area > 0 and not _pre_break
                and not _reflood_started and x_eq[t] >= 0.0
                and _pump_tripped
                and core_flag):
            N_fail_DNB[t] = N_pins

        # When film boiling is active, rods in film boiling have by definition
        # gone through DNB.  The count depends on the trigger:
        #   _fb_uncover: all rods are uncovered -> N_pins.
        #   _fb_dnbr: count rods whose local power factor exceeds the DNB
        #             threshold factor, using the same DNBR-based radial power
        #             distribution model used before film boiling was activated.
        #
        # Previous code used sf(0)=50% here, which artificially pinned DNB
        # failures at half the core whenever film boiling was DNBR-triggered.
        # The corrected model keeps the statistical peaking-factor estimate and
        # responds to the actual DNBR margin.
        if _film_boiling_active and core_flag:
            if _fb_uncover:
                N_fail_DNB[t] = max(N_fail_DNB[t], N_pins)
            elif _fb_dnbr:
                if not np.isnan(DNBR[t]) and _sigma_r and _sigma_r > 0:
                    _dnb_threshold_factor_fb = DNBR[t]
                    _z_dnb_fb = (_dnb_threshold_factor_fb - 1.0) / _sigma_r
                    N_fb = N_pins * float(_norm_dist.sf(_z_dnb_fb))
                else:
                    # Fallback only: if DNBR is unavailable after transition,
                    # preserve the previously latched count rather than forcing
                    # an artificial 50% failed fraction.
                    N_fb = N_fail_DNB[t-1] if t > 0 else 0.0
                N_fail_DNB[t] = max(N_fail_DNB[t], N_fb)

        # Latch: N_fail_DNB is non-decreasing.
        if t > 0:
            N_fail_DNB[t] = max(N_fail_DNB[t], N_fail_DNB[t-1])

        # ── Cladding temperature failure counts (PCT / source term) ──────────
        # Temperature distribution model:
        #   T_hot is treated deterministically: if T_hot <= T_limit, zero rods fail.
        #   When T_hot > T_limit, rod temperatures are normally distributed with
        #   mean = T_avg and σ_T = (T_hot - T_avg) / 3  (hot rod = 3σ upper tail).
        #   N_survivors = N_pins * Φ((T_limit - T_avg) / σ_T)  (rods below limit)
        #   N_fail = N_pins - N_survivors
        if (core_flag and not np.isnan(TTwall[t])
                and not np.isnan(T_hot_clad_arr[t])):
            _T_hot_c  = T_hot_clad_arr[t]         # hot pin clad [K]
            _T_avg_c  = TTwall[t]                  # average clad [K]
            # Rod failure counting in temperature space.
            # The clad temperature distribution across pins is Gaussian:
            #   mean  = T_avg_clad  (from slab solve)
            #   sigma = (T_hot - T_avg) / k_sigma
            # This correctly reflects the power-peaking spread in clad temp.
            _dT      = max(_T_hot_c - _T_avg_c, 0.1)
            _sigma_T = _dT / max(k_sigma, 0.1)

            def _n_fail_T(T_lim_C):
                _T_lim_K = T_lim_C + 273.15
                if _T_hot_c <= _T_lim_K:    # hot pin survives — none fail
                    return 0
                _z = (_T_lim_K - _T_avg_c) / _sigma_T
                return max(0, N_pins - int(round(N_pins * float(_norm_dist.cdf(_z)))))

            # Instantaneous exceedance counts from the current temperature
            # distribution.  Once a rod has failed/ruptured it does not "heal"
            # when the cladding cools, so the reported/output counts are
            # latched cumulative counts.
            _N_fail_clad_inst = _n_fail_T(T_early_iv_C)
            _N_fail_gap_inst  = _n_fail_T(T_gap_release_C)
            _N_fail_eiv_inst  = _n_fail_T(T_early_iv_C)

            if t > 0:
                N_fail_clad[t] = max(_N_fail_clad_inst, N_fail_clad[t-1])
                N_fail_gap[t]  = max(_N_fail_gap_inst,  N_fail_gap[t-1])
                N_fail_eiv[t]  = max(_N_fail_eiv_inst,  N_fail_eiv[t-1])
            else:
                N_fail_clad[t] = _N_fail_clad_inst
                N_fail_gap[t]  = _N_fail_gap_inst
                N_fail_eiv[t]  = _N_fail_eiv_inst

            # NOTE: DNB does NOT independently trigger RG 1.183 gap release.
            # But any rod above 800°C has certainly gone through DNB first —
            # enforce N_fail_DNB >= cumulative N_fail_gap now that it is known.
            N_fail_DNB[t] = max(N_fail_DNB[t], N_fail_gap[t])
            # Gap release requires T_clad > 800°C (cladding rupture criterion).
            # DNB-induced dose is captured via the 335× accident iodine spike.
            if t > 0:
                N_fail_DNB[t] = max(N_fail_DNB[t], N_fail_DNB[t-1])

        # ── Film boiling activation and rewetting (two independent triggers) ──
        # Helper: wall-temperature Leidenfrost guard (shared by both triggers)
        try:
            _Tsat_fb = XSteam.Tsat_p(Pressure[t] * 1e-3) + 273.15  # K
            _wall_cool_fb = (TTwall[t] - _Tsat_fb) < rewet_dT_K
        except Exception:
            _Tsat_fb = 373.15; _wall_cool_fb = True

        # ── Trigger (a): DNBR < 1 ────────────────────────────────────────────
        # Guard: only activate if fluid is near saturation (x_eq > -0.05).
        # At deeply subcooled conditions the Biasi/Zuber CHF → 0 as G → 0
        # (correlation out of range), giving spurious DNBR < 1 without real
        # DNB. Churchill-Chu natural convection already handles low-flow HTC.
        _near_sat = x_eq[t] > -0.05 if not np.isnan(x_eq[t]) else False
        if core_flag and not np.isnan(DNBR[t]):
            if not _fb_dnbr and DNBR[t] < 1.0 and _near_sat:
                _fb_dnbr = True
                print(f"Film boiling (CHF) at t={time[t]:.1f} s: "
                      f"DNBR = {DNBR[t]:.3f} < 1.0 — Bromley HTC "
                      f"(cap {htc_post_chf:.0f} W/m2K).")
            elif _fb_dnbr and rewet_dnbr > 0 and DNBR[t] >= rewet_dnbr:
                if _wall_cool_fb:
                    _fb_dnbr = False
                    print(f"Rewetting (DNBR) at t={time[t]:.1f} s: "
                          f"DNBR = {DNBR[t]:.3f} >= {rewet_dnbr:.1f}, "
                          f"T_wall = {TTwall[t]-273.15:.0f} C — "
                          f"nucleate boiling restored.")

        # ── Trigger (a2): cold-leg — subcooling lost with break open ────────
        # Once subcooling is lost (x_eq >= 0) and pump has tripped, assume
        # immediate film boiling.  _pump_tripped guards against firing at
        # full flow before scram — film boiling at full rated flow would be
        # unphysical.  With default trip_delay=1.5s, fires at t~1.5s.
        if (post_chf_model == "cold_leg" and core_flag
                and _active_area > 0 and not _pre_break
                and not _fb_dnbr and not _reflood_started
                and x_eq[t] >= 0.0
                and _pump_tripped):
            _fb_dnbr = True
            print(f"Film boiling (cold-leg, subcooling lost) at t={time[t]:.1f} s: "
                  f"x_eq={x_eq[t]:.3f}, mdot={pump_mdot[t]:.0f} kg/s "
                  f"— Bromley HTC (cap {htc_post_chf:.0f} W/m2K).")

        # ── Trigger (b): core uncovery (ves_ll < film_boiling_level_m) ───────
        if core_flag and not np.isnan(ves_ll[t]):
            if not _fb_uncover and ves_ll[t] < film_boiling_level_m:
                _fb_uncover = True
                print(f"Film boiling (uncovery) at t={time[t]:.1f} s: "
                      f"level = {ves_ll[t]:.2f} m < "
                      f"{film_boiling_level_m:.2f} m — Bromley HTC "
                      f"(cap {htc_post_chf:.0f} W/m2K).")
            elif (_fb_uncover
                      and ves_ll[t] >= film_boiling_level_m + film_boiling_level_margin_m):
                # Level has recovered above critical + margin.
                # No Leidenfrost guard here: submergence-driven quenching is a
                # hydraulic event — cold liquid floods over the rod and forces
                # rewetting regardless of wall temperature.  The Leidenfrost
                # criterion is appropriate for trigger (a) (CHF/surface), not
                # for trigger (b) (bulk level recovery).
                _fb_uncover = False
                # Cold-leg: subcooling-loss trigger also resets on level recovery.
                # Physical basis: rod is now submerged — hydraulic quench overrides
                # Leidenfrost criterion that would otherwise prevent _fb_dnbr reset.
                if post_chf_model == "cold_leg":
                    _fb_dnbr = False
                print(f"Rewetting (reflooded) at t={time[t]:.1f} s: "
                      f"level = {ves_ll[t]:.2f} m >= "
                      f"{film_boiling_level_m + film_boiling_level_margin_m:.2f} m, "
                      f"T_wall = {TTwall[t]-273.15:.0f} C — "
                      f"nucleate boiling restored.")

        # Combined flag: either trigger keeps film boiling active
        _film_boiling_active = _fb_dnbr or _fb_uncover

        # ── status print ──────────────────────────────────────────────────────
        elapsed_time = timer.time() - start_time
        if int(elapsed_time) > counter + 9:
            print(f"Job time: {elapsed_time:.1f} s, Sim time: {time[t]:.1f} s.")
            counter = int(elapsed_time)

    end_time = timer.time()
    print(f"Total execution time: {end_time-start_time:.1f} s.")

    # ── Fuel failure summary ─────────────────────────────────────────────────
    _N_fail_DNB_peak = int(np.nanmax(N_fail_DNB)) if core_flag else 0
    # Enthalpy failures from NF burst energy
    if _nf_E_total > 0 and _m_fuel_pin > 0 and N_pins > 0:
        # Average enthalpy per pin = _nf_E_total / N_pins (J/pin)
        # Hot pin at power factor F_r gets F_r × average enthalpy.
        # Pin at factor f fails when:
        #   f × _nf_E_total/(N_pins × _m_fuel_pin × 4184) > H_fuel_limit_calg
        #   f > f_crit = H_fuel_limit_calg × N_pins × _m_fuel_pin × 4184 / _nf_E_total
        _f_crit = (H_fuel_limit_calg * N_pins * _m_fuel_pin * 4184.0) / _nf_E_total
        _f_fail = max(0.0, F_r - _f_crit) / F_r if F_r > 0 else 0.0
        _N_fail_enth = int(N_pins * _f_fail)
        _H_hot_calg  = (_nf_E_total * F_r * F_z
                        / (N_pins * _m_fuel_pin * 4184.0))
        print(f"Fuel enthalpy failures: {_N_fail_enth} / {N_pins} pins"
              f"  (hot-pin enthalpy {_H_hot_calg:.1f} cal/g,"
              f" limit {H_fuel_limit_calg:.0f} cal/g)")
    else:
        _N_fail_enth = 0
    _N_fail_clad_peak = int(np.nanmax(N_fail_clad)) if core_flag else 0
    if _N_fail_clad_peak > 0:
        print(f"Fuel clad failures PCT>{T_early_iv_C:.0f}C (peak):"
              f" {_N_fail_clad_peak} / {N_pins} pins"
              f"  ({100*_N_fail_clad_peak/N_pins:.1f}% of core)")
    if _N_fail_DNB_peak > 0:
        print(f"Fuel DNB failures (peak): {_N_fail_DNB_peak} / {N_pins} pins"
              f"  ({100*_N_fail_DNB_peak/N_pins:.1f}% of core)")

    # ── Source term model selection (auto-selected from calculated outcomes) ──
    # Decision tree:
    #   LOCA (break > 0)  → temperature-based only (T_clad > 800/1204°C)
    #   Non-LOCA          → iodine spike model, selected by outcomes:
    #     Scram OR PORV   → equilibrium spike (500× coolant activity)
    #     DNB occurs      → accident-initiated spike (gap release from DNB rods)
    #     T_clad > 800°C  → temperature gap release (overrides spike)
    #     T_clad > 1204°C → gap + early in-vessel
    #     Final ST        → MAX of all applicable contributions

    # Coolant activity parameter (optional input, conservative TS-limit default)
    _coolant_act_uci_g = float(get_variable(local_namespace,
                                            "coolant_activity_uci_g", 60.0))
    # Primary coolant mass [g] from initial conditions
    _primary_mass_g    = init_cond.get("Total_Mass_RCS", 0.0) * 1000.0
    if _primary_mass_g <= 0:
        _primary_mass_g = 1.5e8   # fallback: ~150,000 kg typical PWR primary [g]

    _is_loca         = (diameter_break > 0.0)
    _porv_opened     = np.any(massflow_PORV > 0)
    _scram_occurred  = _reactor_scrammed
    _dnb_occurred    = core_flag and (int(np.nanmax(N_fail_DNB)) > 0)
    _N_dnb_peak      = int(np.nanmax(N_fail_DNB)) if core_flag else 0

    # Equilibrium iodine spike: 500× coolant activity × primary mass
    # Applied to noble gases and halogens only (iodines partition to coolant)
    # Expressed as fraction of core inventory for each group
    _spike_ci         = 500.0 * _coolant_act_uci_g * 1e-6 * _primary_mass_g  # Ci total I equiv
    # Convert spike inventory to a fraction of the current NBT halogen inventory.
    # Use the same inventory model used by NOTBADTRAD, including typPWR geometry
    # calibration and any nbt_inv_Halogens_Ci user override.
    try:
        import copy as _copy_isp_inv
        _isp_inv_groups, _isp_inv_meta = _nbt_apply_inventory_model(
            _copy_isp_inv.deepcopy(_NBT_GROUPS), local_namespace)
        _nbt_halogen_inv = float(_isp_inv_groups.get("Halogens", {}).get("inv_ci", 0.0))
    except Exception:
        _nbt_halogen_inv = 1.110e8

    _f_eq_spike       = min(_spike_ci / _nbt_halogen_inv, 1.0) if _nbt_halogen_inv > 0 else 0.0

    # ── Iodine spike model — triggered by event outcomes, applies to ALL cases ──
    # Both LOCA and non-LOCA have pre-existing coolant activity.
    # A LOCA always causes a scram and often opens the PORV.
    _f_gap_equil  = 0.0   # equilibrium spike release fraction (halogens)
    _f_gap_accid  = 0.0   # accident-initiated spike release fraction (halogens)
    _spike_model_used = "None"

    # Equilibrium spike: 500× TS limit — scram or PORV opening (all cases)
    if _scram_occurred or _porv_opened:
        _f_gap_equil      = _f_eq_spike
        _spike_model_used = "Equilibrium spike (500× TS)"

    # Accident-initiated spike: 335× TS limit — fuel damage (all cases)
    # Triggered by DNB or T_clad > 800°C; additive with equilibrium spike
    _has_fuel_damage = (_dnb_occurred or
                        (core_flag and int(np.nanmax(N_fail_gap)) > 0))
    if _has_fuel_damage:
        _spike_accid_ci = 335.0 * _coolant_act_uci_g * 1e-6 * _primary_mass_g
        _f_gap_accid    = min(_spike_accid_ci / _nbt_halogen_inv, 1.0) \
                          if _nbt_halogen_inv > 0 else 0.0
        _spike_model_used = ("Equilibrium (500×) + Accident (335×)"
                             if (_scram_occurred or _porv_opened)
                             else "Accident spike (335× TS)")
    else:
        _spike_accid_ci = 0.0

    # Combined halogen spike fraction (additive, capped at 1.0)
    _f_isp_combined = min(_f_gap_equil + _f_gap_accid, 1.0)

    # ── Iodine spike dose — standalone NOTBADTRAD call for ALL cases ─────────
    # Computed independently of rod failures so it always appears in results.
    _iodine_spike_dose_df = None
    if _f_isp_combined > 0:
        try:
            # Iodine spike releases halogens only — explicitly zero all other groups
            _isp_rel = {
                "NG":            0.0,
                "Halogens":      _f_isp_combined,
                "Alkali_metals": 0.0,
                "Te_group":      0.0,
                "Ba_Sr":         0.0,
                "Noble_metals":  0.0,
                "Ce_group":      0.0,
                "Lanthanides":   0.0,
                "U_actinides":   0.0,
            }
            _isp_d = _notbadtrad(
                power_mwt               = float(get_variable(local_namespace, "total_power", 575.0)),
                containment_volume_ft3  = float(get_variable(local_namespace, "nbt_containment_volume_ft3", 2.74e6)),
                leak_rate_frac_per_day  = float(get_variable(local_namespace, "nbt_leak_rate_frac_per_day", 0.001)),
                sprays_on               = bool( get_variable(local_namespace, "nbt_sprays_on", True)),
                spray_start_hr          = float(get_variable(local_namespace, "nbt_spray_start_hr", 1/60)),
                spray_stop_hr           = float(get_variable(local_namespace, "nbt_spray_stop_hr", 24.0)),
                release_duration_hr     = float(get_variable(local_namespace, "nbt_gap_release_duration_hr", 0.5)),
                wind_speed_m_s          = float(get_variable(local_namespace, "nbt_wind_speed_m_s", 1.0)),
                stability_class         = str(  get_variable(local_namespace, "nbt_stability_class", "F")),
                release_height_m        = float(get_variable(local_namespace, "nbt_release_height_m", 0.0)),
                distance_eab_m          = float(get_variable(local_namespace, "nbt_distance_eab_m", 914.0)),
                distance_lpz_m          = float(get_variable(local_namespace, "nbt_distance_lpz_m", 4800.0)),
                distance_cr_intake_m    = float(get_variable(local_namespace, "nbt_distance_cr_intake_m", 100.0)),
                cr_flow_cfm             = float(get_variable(local_namespace, "nbt_cr_flow_cfm", 100.0)),
                cr_volume_ft3           = float(get_variable(local_namespace, "nbt_cr_volume_ft3", 20000.0)),
                cr_filter_on            = bool( get_variable(local_namespace, "nbt_cr_filter_on", True)),
                end_time_hr             = float(get_variable(local_namespace, "nbt_end_time_hr", 24.0)),
                breathing_rate_m3_s     = float(get_variable(local_namespace, "nbt_breathing_rate_m3_s", 3.47e-4)),
                eab_integration_hr      = float(get_variable(local_namespace, "nbt_eab_integration_hr", 2.0)),
                lpz_integration_hr      = float(get_variable(local_namespace, "nbt_lpz_integration_hr", 720.0)),
                release_overrides       = _isp_rel,
                credit_decay            = bool( get_variable(local_namespace, "nbt_credit_decay", False)),
                credit_deposition       = bool( get_variable(local_namespace, "nbt_credit_deposition", False)),
                local_namespace          = local_namespace,
            )
            _LIMIT_EAB = 25.0; _LIMIT_LPZ = 25.0; _LIMIT_CR = 5.0
            _iodine_spike_dose_df = {
                "summary": pd.DataFrame([
                    {"Location": "EAB", "TEDE (rem)": _isp_d["tede_eab_rem"],
                     "Limit (rem)": f"{_LIMIT_EAB:g} rem ({_isp_d.get('integration_times_hr', {}).get('eab', 2.0):g} hr)", "Margin": _LIMIT_EAB - _isp_d["tede_eab_rem"],
                     "Result": "PASS" if _isp_d["tede_eab_rem"] < _LIMIT_EAB else "EXCEEDS"},
                    {"Location": "LPZ", "TEDE (rem)": _isp_d["tede_lpz_rem"],
                     "Limit (rem)": f"{_LIMIT_LPZ:g} rem ({_isp_d.get('integration_times_hr', {}).get('lpz', 720.0):g} hr)", "Margin": _LIMIT_LPZ - _isp_d["tede_lpz_rem"],
                     "Result": "PASS" if _isp_d["tede_lpz_rem"] < _LIMIT_LPZ else "EXCEEDS"},
                    {"Location": "Control Room", "TEDE (rem)": _isp_d["tede_cr_rem"],
                     "Limit (rem)": _LIMIT_CR, "Margin": _LIMIT_CR - _isp_d["tede_cr_rem"],
                     "Result": "PASS" if _isp_d["tede_cr_rem"] < _LIMIT_CR else "EXCEEDS"},
                ]),
                "groups": pd.DataFrame([
                    {"Group": gn, "Released (Ci)": gd["released_ci"],
                     "EAB (rem)": gd["tede_eab"], "LPZ (rem)": gd["tede_lpz"],
                     "CR (rem)": gd["tede_cr"], "Release frac": _isp_rel.get(gn, 0.0)}
                    for gn, gd in _isp_d["group_doses"].items() if gd["released_ci"] > 0
                ]),
                "chi_q":       _isp_d.get("chi_q_used", {}),
                "runtime_ms":  _isp_d["runtime_ms"],
                "dist_fit":    _isp_d.get("dist_fit"),
                "dist_table":  pd.DataFrame(_isp_d.get("dist_table", [])),
                "iodine_spike": {
                    "coolant_act_uci_g":      _coolant_act_uci_g,
                    "primary_mass_kg":        _primary_mass_g / 1000.0,
                    "equil_spike_ci":         _spike_ci,
                    "accid_spike_ci":         _spike_accid_ci,
                    "spike_multiplier_equil": 500,
                    "spike_multiplier_accid": 335,
                    "f_eq_spike":             _f_gap_equil,
                    "f_accid_spike":          _f_gap_accid,
                    "f_combined":             _f_isp_combined,
                    "model_used":             _spike_model_used,
                    "is_loca":                _is_loca,
                    "scram":                  _scram_occurred,
                    "porv":                   _porv_opened,
                    "dnb":                    _dnb_occurred,
                    "fuel_damage":            _has_fuel_damage,
                    "N_dnb":                  _N_dnb_peak,
                },
            }
            print(f"\nIodine Spike Dose ({_spike_model_used}):")
            print(f"  Equil (500×): {_spike_ci:.3e} Ci  frac={_f_gap_equil:.4e}")
            print(f"  Accid (335×): {_spike_accid_ci:.3e} Ci  frac={_f_gap_accid:.4e}")
            print(f"  Combined halogen frac: {_f_isp_combined:.4e}")
            print(f"  EAB: {_isp_d['tede_eab_rem']:.4e} rem  "
                  f"LPZ: {_isp_d['tede_lpz_rem']:.4e} rem  "
                  f"CR: {_isp_d['tede_cr_rem']:.4e} rem")
        except Exception as _isp_err:
            print(f"WARNING: Iodine spike dose calculation failed: {_isp_err}")

    # ── RG 1.183 source term estimate ──────────────────────────────────────
    # Release fractions per RG 1.183 Table 3 (PWR, LOCA basis).
    # Gap release: noble gases 5%, halogens 5%, alkali metals 5%.
    # Early in-vessel: noble gases 95% additional, halogens 35%, alkali metals 25%,
    #   Te group 5%, Ba/Sr 2%.
    # LOCA: source term = f_gap_pk*gap_frac + f_eiv_pk*early_iv_frac
    # Non-LOCA: source term = max(temperature-based, spike-based) per group
    _source_term_df        = None   # populated only when failures > 0
    _dose_df               = None   # populated when release occurs
    # Do not reset _iodine_spike_dose_df here. It may already contain the
    # standalone pre-existing coolant activity / iodine-spike dose result.
    _is_loca               = (diameter_break > 0.0)
    _rg183 = {
        # Group                    (gap_frac, early_iv_frac)  — RG 1.183 Rev 0 Table 3, PWR LOCA
        # gap_frac    : fraction released from ALL failed rods (gap release)
        # early_iv_frac: ADDITIONAL fraction released from early in-vessel rods only
        "Noble gases (Kr, Xe)":    (0.05, 0.95),   # gap: 5%; early IV additional: 95% (total 100%)
        "Halogens (I, Br)":        (0.05, 0.35),   # gap: 5%; early IV additional: 35%
        "Alkali metals (Cs, Rb)":  (0.05, 0.25),   # gap: 5%; early IV additional: 25%
        "Te group (Te, Sb, Se)":   (0.00, 0.05),   # gap: 0%; early IV additional: 5%
        "Ba / Sr":                 (0.00, 0.02),    # gap: 0%; early IV additional: 2%
        "Ru group (Ru, Mo, Tc)":   (0.00, 0.025),  # gap: 0%; early IV additional: 2.5%
        "Lanthanides (La, Zr, Nd)":(0.00, 0.002),  # gap: 0%; early IV additional: 0.2%
        "Ce group (Ce, Pu, Np)":   (0.00, 0.0005), # gap: 0%; early IV additional: 0.05%
        "U / Actinides":           (0.00, 0.0000), # no RG 1.183 base release; severe-ECR only
    }

    # Group-specific severe-event release targets.  If the mean oxidizing-rod
    # ECR threshold is exceeded, the severe-event total release fraction is
    # interpolated from the existing gap + early-in-vessel total release
    # fraction toward the BDBE group target:
    #     Total_severe = Total_base + ECR_fraction * (F_BDBE - Total_base)
    # where ECR_fraction = peak_mean_oxidizing_rod_ECR_percent / 100.
    _bdbE_F = {
        "Noble gases (Kr, Xe)":     1.0,
        "Halogens (I, Br)":         0.75,
        "Alkali metals (Cs, Rb)":   0.75,
        "Te group (Te, Sb, Se)":    0.30,
        "Ba / Sr":                  0.12,
        "Ru group (Ru, Mo, Tc)":    0.005,
        "Lanthanides (La, Zr, Nd)": 0.005,
        "Ce group (Ce, Pu, Np)":    0.005,
        "U / Actinides":            0.001,
    }

    def _estimate_mean_oxidizing_rod_ecr_pct_for_source_term():
        """Return peak mean-oxidizing-rod ECR [%] using the same Baker-Just
        diagnostic basis used in post_processing().  This is a core-wide
        source-term severity indicator, not the local hot-pin ECR.

        The oxidizing-rod mean temperature is estimated from the same truncated
        Gaussian clad-temperature distribution used for rod-failure estimates,
        restricted to rods above 800 °C and bounded by the hot-pin clad
        temperature.  The ECR is accumulated on that mean oxidizing rod.
        """
        if not core_flag or len(time) == 0:
            return 0.0
        try:
            _A_BJ = 33.3e6             # mg²/cm⁴/s, oxygen weight-gain basis
            _B_BJ = 45500.0 / 1.8      # K-equivalent activation temperature
            _T_OX_MIN_K = 800.0 + 273.15
            _rho_zr_g_cm3 = 6.56
            _M_Zr = 91.224
            _M_O2 = 31.998
            _clad_thick_cm = max(float(delta_clad), 0.0) * 100.0
            if _clad_thick_cm <= 0.0:
                return 0.0
            _W_full_O_mg_cm2 = _rho_zr_g_cm3 * _clad_thick_cm * (_M_O2 / _M_Zr) * 1000.0
            if not np.isfinite(_W_full_O_mg_cm2) or _W_full_O_mg_cm2 <= 0.0:
                return 0.0

            _k_sigma_src = max(float(k_sigma), 0.1)

            def _phi(_z):
                return np.exp(-0.5 * _z * _z) / np.sqrt(2.0 * np.pi)

            def _Phi(_z):
                return 0.5 * (1.0 + _math.erf(_z / np.sqrt(2.0)))

            _W_oxid = 0.0
            _peak_ecr = 0.0
            for _i in range(len(time)):
                _dt_s = 0.0 if _i == 0 else max(float(time[_i] - time[_i-1]), 0.0)
                _T_hot = float(T_hot_clad_arr[_i]) if np.isfinite(T_hot_clad_arr[_i]) else np.nan
                _T_avg = float(TTwall[_i]) if np.isfinite(TTwall[_i]) else np.nan
                _N_oxid = max(float(N_fail_gap[_i]), 0.0) if np.isfinite(N_fail_gap[_i]) else 0.0

                if (_N_oxid > 0.0 and np.isfinite(_T_hot) and np.isfinite(_T_avg)
                        and _T_hot > _T_OX_MIN_K and _dt_s > 0.0):
                    _sigma_T = max((_T_hot - _T_avg) / _k_sigma_src, 0.1)
                    _a = (_T_OX_MIN_K - _T_avg) / _sigma_T
                    _b = (_T_hot      - _T_avg) / _sigma_T
                    _den = max(_Phi(_b) - _Phi(_a), 1e-15)
                    _T_oxid_mean = _T_avg + _sigma_T * ((_phi(_a) - _phi(_b)) / _den)
                    _T_oxid_mean = min(
                        min(max(_T_oxid_mean, _T_OX_MIN_K), _T_hot),
                        2500.0 + 273.15
                    )
                    _rate_w2 = _A_BJ * np.exp(-_B_BJ / _T_oxid_mean)
                    _W_oxid = min(np.sqrt(max(_W_oxid*_W_oxid + _rate_w2*_dt_s, 0.0)),
                                  _W_full_O_mg_cm2)

                _peak_ecr = max(_peak_ecr, 100.0 * _W_oxid / _W_full_O_mg_cm2)
            return float(min(max(_peak_ecr, 0.0), 100.0))
        except Exception as _ecr_err:
            print(f"WARNING: severe-event mean oxidizing-rod ECR source-term diagnostic failed: {_ecr_err}")
            return 0.0

    _bdbE_ecr_mean_oxid_peak_pct = _estimate_mean_oxidizing_rod_ecr_pct_for_source_term()
    _bdbE_ecr_active = (bool(bdbE_ecr_release_flag) and
                        _bdbE_ecr_mean_oxid_peak_pct > bdbE_ecr_threshold_pct)
    _bdbE_ecr_frac = (_bdbE_ecr_mean_oxid_peak_pct / 100.0) if _bdbE_ecr_active else 0.0

    if core_flag and N_pins > 0:
        _N_gap_pk  = int(np.nanmax(N_fail_gap))  if core_flag else 0
        _N_eiv_pk  = int(np.nanmax(N_fail_eiv))  if core_flag else 0
        _N_dnb_pk  = int(np.nanmax(N_fail_DNB))  if core_flag else 0

        # Best-estimate thermal-failure diagnostics.
        _f_gap_thermal = _N_gap_pk / N_pins   # fraction of rods with gap release (T>800°C)
        _f_eiv_thermal = _N_eiv_pk / N_pins   # fraction of rods with early in-vessel release
        _f_dnb_thermal = _N_dnb_pk / N_pins   # RIA/DNB failed-fuel diagnostic

        # Effective fractions used by the fuel-damage source term.
        # The iodine spike / pre-existing coolant activity term is independent.
        _stm = source_term_model.strip("'\"").lower()
        _stm = _stm.replace("-", "_").replace(" ", "_")
        _stm = _stm.replace("non_loca", "nonloca")

        if _stm == "rg1183_loca":
            _f_gap_pk = 1.0
            _f_eiv_pk = 1.0
            _source_model_label = "RG 1.183 licensing LOCA — full gap + early in-vessel"

        elif _stm in ("rg1183_nonloca", "rg1183_non_loca"):
            _f_gap_pk = 1.0
            _f_eiv_pk = 0.0
            _source_model_label = "RG 1.183 licensing non-LOCA fuel damage — full gap only"

        elif _stm == "rg1183_nonloca_dnb":
            _f_gap_pk = min(max(_f_dnb_thermal, 0.0), 1.0)
            _f_eiv_pk = 0.0
            _source_model_label = (
                "RG 1.183 non-LOCA DNB fuel damage — "
                "DNB failed fraction × gap only"
            )

        elif _stm == "licensing_auto":
            if _is_loca:
                _f_gap_pk = 1.0
                _f_eiv_pk = 1.0
                _source_model_label = "RG 1.183 licensing auto: LOCA — full gap + early in-vessel"
            else:
                _fuel_damage_predicted = (_N_gap_pk > 0 or _N_eiv_pk > 0 or _N_dnb_pk > 0)
                _f_gap_pk = 1.0 if _fuel_damage_predicted else 0.0
                _f_eiv_pk = 0.0
                _source_model_label = "RG 1.183 licensing auto: non-LOCA — full gap if fuel damage"

        else:
            _f_gap_pk = _f_gap_thermal
            _f_eiv_pk = _f_eiv_thermal
            _source_model_label = "FLARE thermal-failure best estimate — failed fraction × RG 1.183"

        _has_fuel_release = (_f_gap_pk > 0.0 or _f_eiv_pk > 0.0)
        _has_release = (_has_fuel_release or _N_fail_enth > 0 or
                        _f_gap_equil > 0 or _f_gap_accid > 0 or
                        _bdbE_ecr_frac > 0.0)

        if _has_release:
            print()
            print(f"Source term model: {_source_model_label}")
            if _f_gap_equil > 0 or _f_gap_accid > 0:
                print(f"  + Iodine spike / pre-existing coolant activity: {_spike_model_used}")
            if not _is_loca:
                print(f"  Coolant activity: {_coolant_act_uci_g:.3f} μCi/g")
                print(f"  Equilibrium spike fraction: {100*_f_gap_equil:.4f}%")
                if _dnb_occurred:
                    print(f"  Accident spike (DNB rods): {_N_dnb_peak} / {N_pins}"
                          f" = {100*_f_gap_accid:.1f}%")
            print(f"  Diagnostic gap release rods (T>{T_gap_release_C:.0f}C): "
                  f"{_N_gap_pk} / {N_pins} pins ({100*_f_gap_thermal:.1f}% of core)")
            print(f"  Diagnostic DNB rods: {_N_dnb_pk} / {N_pins} pins "
                  f"({100*_f_dnb_thermal:.1f}% of core)")
            if _N_eiv_pk > 0:
                print(f"  Diagnostic early in-vessel rods (T>{T_early_iv_C:.0f}C): "
                      f"{_N_eiv_pk} / {N_pins} pins ({100*_f_eiv_thermal:.1f}% of core)")
            print(f"  Effective gap fraction used: {100*_f_gap_pk:.3f}%")
            print(f"  Effective early-IV fraction used: {100*_f_eiv_pk:.3f}%")
            print(f"  Severe-event mean oxidizing-rod ECR diagnostic: {_bdbE_ecr_mean_oxid_peak_pct:.2f}% "
                  f"(threshold {bdbE_ecr_threshold_pct:.2f}%, "
                  f"active={_bdbE_ecr_active})")
            _hdr = f"  {'Group':<28} {'Gap frac':>10} {'EIV add.':>10} {'Total release':>15}"
            print(_hdr)
            print(f"  {'-'*56}")
            # Source Term table: include group inventories and released inventory.
            # The inventory basis is the same model used by NOTBADTRAD, including
            # geometry/typPWR calibration and any nbt_inv_<Group>_Ci overrides.
            import copy as _copy_st
            _st_inv_groups, _st_inv_meta = _nbt_apply_inventory_model(
                _copy_st.deepcopy(_NBT_GROUPS), local_namespace)

            _grp_map_st = {
                "Noble gases (Kr, Xe)":    "NG",
                "Halogens (I, Br)":        "Halogens",
                "Alkali metals (Cs, Rb)":  "Alkali_metals",
                "Te group (Te, Sb, Se)":   "Te_group",
                "Ba / Sr":                 "Ba_Sr",
                "Ru group (Ru, Mo, Tc)":   "Noble_metals",
                "Lanthanides (La, Zr, Nd)":"Lanthanides",
                "Ce group (Ce, Pu, Np)":   "Ce_group",
                "U / Actinides":           "U_actinides",
            }

            _st_rows = []
            _st_total_inventory_ci = 0.0
            _st_total_released_ci  = 0.0

            _st_release_by_group = {}
            _st_bdbE_add_by_group = {}
            for _grp, (_fg, _fe) in _rg183.items():
                # Base release = effective_gap_frac × gap_release_fraction
                #              + eiv_frac × early_iv_additional_fraction.
                _st_base = min(max(_f_gap_pk * _fg + _f_eiv_pk * _fe, 0.0), 1.0)

                # Severe-event adjustment for high predicted ECR.
                # Total_severe = Total_base + ECR_fraction*(F_BDBE - Total_base).
                # This treats F_BDBE as the severe-event target release fraction.
                _bdbE_F_grp = float(_bdbE_F.get(_grp, 0.0))
                _st_target = min(max(_bdbE_F_grp, 0.0), 1.0)
                _bdbE_add = (_bdbE_ecr_frac * (_st_target - _st_base)
                             if _bdbE_ecr_frac > 0.0 else 0.0)
                _st = min(max(_st_base + _bdbE_add, 0.0), 1.0)
                _bdbE_add = _st - _st_base
                _st_release_by_group[_grp] = _st
                _st_bdbE_add_by_group[_grp] = _bdbE_add

                _nbt_key = _grp_map_st.get(_grp)
                _inv_ci = float(_st_inv_groups.get(_nbt_key, {}).get("inv_ci", 0.0)) if _nbt_key else 0.0
                _rel_ci = _inv_ci * _st
                _st_total_inventory_ci += _inv_ci
                _st_total_released_ci  += _rel_ci

                print(f"  {_grp:<28} {100*_fg:>9.0f}%"
                      f" {100*_fe:>9.2f}%"
                      f" {100*_st:>13.3f}%"
                      f"  inv={_inv_ci:.3e} Ci rel={_rel_ci:.3e} Ci")
                _st_rows.append({
                    "Group":                   _grp,
                    "NBT key":                 _nbt_key,
                    "Core inventory (Ci)":     _inv_ci,
                    "Gap release frac":        _fg,
                    "Early IV frac":           _fe,
                    "BDBE F factor":           _bdbE_F_grp,
                    "BDBE ECR adjustment frac": _bdbE_add,
                    "Total release frac":      _st,
                    "Released inventory (Ci)": _rel_ci,
                    "Gap release %":           round(100*_fg, 3),
                    "Early IV %":              round(100*_fe, 3),
                    "BDBE ECR adjustment %":    round(100*_bdbE_add, 3),
                    "Total release %":         round(100*_st, 3),
                })

            # Include NBT groups that are not explicitly represented in the
            # current RG 1.183 release-fraction table, so the inventory total is complete.
            for _nbt_key, _g in _st_inv_groups.items():
                if _nbt_key in set(_grp_map_st.values()):
                    continue
                _inv_ci = float(_g.get("inv_ci", 0.0))
                _st_total_inventory_ci += _inv_ci
                _st_rows.append({
                    "Group":                 _g.get("label", _nbt_key),
                    "NBT key":               _nbt_key,
                    "Core inventory (Ci)":   _inv_ci,
                    "Gap release frac":      0.0,
                    "Early IV frac":         0.0,
                    "BDBE F factor":         0.0,
                    "BDBE ECR adjustment frac": 0.0,
                    "Total release frac":    0.0,
                    "Released inventory (Ci)": 0.0,
                    "Gap release %":         0.0,
                    "Early IV %":            0.0,
                    "BDBE ECR adjustment %":    0.0,
                    "Total release %":       0.0,
                })

            _st_rows.append({
                "Group":                   "TOTAL",
                "NBT key":                 "",
                "Core inventory (Ci)":     _st_total_inventory_ci,
                "Total release frac":      "",
                "Released inventory (Ci)": _st_total_released_ci,
            })
            _st_rows.append({})
            _st_rows.append({"Group": "--- Source term model ---"})
            _st_rows.append({"Group": "Model selected",
                             "Total release frac": _source_model_label})
            _st_rows.append({"Group": "Fuel source-term option",
                             "Total release frac": source_term_model})
            _st_rows.append({"Group": "Inventory model",
                             "Total release frac": _st_inv_meta.get("model", "")})
            _st_rows.append({"Group": "Estimated fissions",
                             "Total release frac": _st_inv_meta.get("n_fissions", "")})
            _st_rows.append({"Group": "NOTBADTRAD gap release duration (hr)",
                             "Total release frac": get_variable(local_namespace, "nbt_gap_release_duration_hr", 0.5)})
            _st_rows.append({"Group": "NOTBADTRAD early-IV release duration (hr)",
                             "Total release frac": get_variable(local_namespace, "nbt_early_iv_duration_hr", 1.3)})
            _st_rows.append({"Group": "NOTBADTRAD EAB integration time (hr)",
                             "Total release frac": get_variable(local_namespace, "nbt_eab_integration_hr", 2.0)})
            _st_rows.append({"Group": "NOTBADTRAD LPZ integration time (hr)",
                             "Total release frac": get_variable(local_namespace, "nbt_lpz_integration_hr", 720.0)})
            _st_rows.append({"Group": "Severe-event ECR release flag",
                             "Total release frac": bdbE_ecr_release_flag})
            _st_rows.append({"Group": "Severe-event ECR threshold (%)",
                             "Total release frac": bdbE_ecr_threshold_pct})
            _st_rows.append({"Group": "Predicted mean oxidizing-rod ECR (%)",
                             "Total release frac": round(_bdbE_ecr_mean_oxid_peak_pct, 4)})
            _st_rows.append({"Group": "Severe-event ECR active",
                             "Total release frac": str(bool(_bdbE_ecr_active))})
            if not _is_loca:
                _st_rows.append({"Group": "Coolant activity (μCi/g)",
                                 "Total release frac": _coolant_act_uci_g})
                _st_rows.append({"Group": "Equilibrium spike fraction",
                                 "Total release frac": round(_f_gap_equil, 6)})
                _st_rows.append({"Group": "Accident spike fraction (DNB)",
                                 "Total release frac": round(_f_gap_accid, 6)})
            _st_rows.append({})
            _st_rows.append({"Group": "--- Failed rod counts (peak) ---"})
            _st_rows.append({"Group": f"Diagnostic gap rods (T>{T_gap_release_C:.0f}C)",
                             "Total release %": _N_gap_pk,
                             "Total release frac": f"{100*_f_gap_thermal:.1f}% of {N_pins} pins"})
            _st_rows.append({"Group": "Diagnostic DNB rods",
                             "Total release %": _N_dnb_pk,
                             "Total release frac": f"{100*_f_dnb_thermal:.1f}% of {N_pins} pins"})
            _st_rows.append({"Group": f"Diagnostic early in-vessel rods (T>{T_early_iv_C:.0f}C)",
                             "Total release %": _N_eiv_pk,
                             "Total release frac": f"{100*_f_eiv_thermal:.1f}% of {N_pins} pins"})
            _st_rows.append({"Group": "Effective gap fraction used",
                             "Total release frac": f"{100*_f_gap_pk:.1f}% of core"})
            _st_rows.append({"Group": "Effective early-IV fraction used",
                             "Total release frac": f"{100*_f_eiv_pk:.1f}% of core"})
            _source_term_df = pd.DataFrame(_st_rows)

            # ── NOTBADTRAD dose screening ──────────────────────────────────
            # Called only when there is a non-zero source term.
            # Parameters are read from the input file with safe defaults.
            try:
                # Build per-group release overrides from FLARE source term.
                #
                # For phase-specific timing, split the fuel source term into
                # separate gap and early-in-vessel release vectors. NOTBADTRAD
                # itself accepts one uniform release duration per call, so FLARE
                # calls it once for the gap phase and once for the early-IV phase,
                # then linearly sums the resulting doses by location and group.
                _grp_map = {
                    "Noble gases (Kr, Xe)":    "NG",
                    "Halogens (I, Br)":        "Halogens",
                    "Alkali metals (Cs, Rb)":  "Alkali_metals",
                    "Te group (Te, Sb, Se)":   "Te_group",
                    "Ba / Sr":                 "Ba_Sr",
                    "Ru group (Ru, Mo, Tc)":   "Noble_metals",
                    "Lanthanides (La, Zr, Nd)":"Lanthanides",
                    "Ce group (Ce, Pu, Np)":   "Ce_group",
                    "U / Actinides":           "U_actinides",
                }

                _nbt_gap_releases = {}
                _nbt_eiv_releases = {}
                _nbt_releases     = {}

                for _grp, (_fg, _fe) in _rg183.items():
                    _nbt_key = _grp_map.get(_grp)
                    if not _nbt_key:
                        continue
                    _gap_rel = float(_f_gap_pk * _fg)
                    _eiv_base_rel = float(_f_eiv_pk * _fe)
                    _severe_adj = float(_st_bdbE_add_by_group.get(_grp, 0.0))

                    # Treat the severe-ECR adjustment as part of the later
                    # in-vessel/severe release phase for dose timing.  The
                    # group total itself is taken from _st_release_by_group,
                    # which already applies the ECR interpolation formula.
                    _eiv_rel = _eiv_base_rel + _severe_adj
                    _total_rel = float(_st_release_by_group.get(
                        _grp, min(max(_gap_rel + _eiv_rel, 0.0), 1.0)))
                    _eiv_rel = max(_total_rel - _gap_rel, 0.0)

                    _nbt_gap_releases[_nbt_key] = _gap_rel
                    _nbt_eiv_releases[_nbt_key] = _eiv_rel
                    _nbt_releases[_nbt_key]     = _total_rel

                # Read NOTBADTRAD parameters from input file (with defaults)
                # total_power comes from the FLARE model — no separate nbt_ variable
                _nbt_power    = float(get_variable(local_namespace, "total_power",  575.0))
                _nbt_vol      = float(get_variable(local_namespace, "nbt_containment_volume_ft3", 2.74e6))
                _nbt_leak     = float(get_variable(local_namespace, "nbt_leak_rate_frac_per_day",  0.001))
                _nbt_sprays   = bool( get_variable(local_namespace, "nbt_sprays_on",               True))
                _nbt_spray_t0 = float(get_variable(local_namespace, "nbt_spray_start_hr",          1/60))
                _nbt_spray_t1 = float(get_variable(local_namespace, "nbt_spray_stop_hr",           24.0))
                _nbt_wind     = float(get_variable(local_namespace, "nbt_wind_speed_m_s",           1.0))
                _nbt_stab     = str(  get_variable(local_namespace, "nbt_stability_class",         "F"))
                _nbt_ht       = float(get_variable(local_namespace, "nbt_release_height_m",         0.0))
                _nbt_eab      = float(get_variable(local_namespace, "nbt_distance_eab_m",         914.0))
                _nbt_lpz      = float(get_variable(local_namespace, "nbt_distance_lpz_m",        4800.0))
                _nbt_cr_d     = float(get_variable(local_namespace, "nbt_distance_cr_intake_m",   100.0))
                _nbt_cr_flow  = float(get_variable(local_namespace, "nbt_cr_flow_cfm",            100.0))
                _nbt_cr_vol   = float(get_variable(local_namespace, "nbt_cr_volume_ft3",        20000.0))
                _nbt_cr_filt  = bool( get_variable(local_namespace, "nbt_cr_filter_on",            True))
                _nbt_end      = float(get_variable(local_namespace, "nbt_end_time_hr",             24.0))
                _nbt_eab_int  = float(get_variable(local_namespace, "nbt_eab_integration_hr",       2.0))
                _nbt_lpz_int  = float(get_variable(local_namespace, "nbt_lpz_integration_hr",     720.0))
                _nbt_br       = float(get_variable(local_namespace, "nbt_breathing_rate_m3_s",   3.47e-4))

                # Phase-specific RG 1.183 / NUREG-1465-style release durations.
                _nbt_gap_dur  = float(get_variable(local_namespace, "nbt_gap_release_duration_hr", 0.5))
                _nbt_eiv_dur  = float(get_variable(local_namespace, "nbt_early_iv_duration_hr",    1.3))

                _nbt_credit_decay      = bool(get_variable(local_namespace, "nbt_credit_decay",       False))
                _nbt_credit_deposition = bool(get_variable(local_namespace, "nbt_credit_deposition",  False))

                def _nbt_has_release(_rels):
                    return any(float(_v) > 0.0 for _v in _rels.values())

                def _nbt_call(_rels, _duration_hr):
                    return _notbadtrad(
                        power_mwt               = _nbt_power,
                        containment_volume_ft3  = _nbt_vol,
                        leak_rate_frac_per_day  = _nbt_leak,
                        sprays_on               = _nbt_sprays,
                        spray_start_hr          = _nbt_spray_t0,
                        spray_stop_hr           = _nbt_spray_t1,
                        release_duration_hr     = _duration_hr,
                        wind_speed_m_s          = _nbt_wind,
                        stability_class         = _nbt_stab,
                        release_height_m        = _nbt_ht,
                        distance_eab_m          = _nbt_eab,
                        distance_lpz_m          = _nbt_lpz,
                        distance_cr_intake_m    = _nbt_cr_d,
                        cr_flow_cfm             = _nbt_cr_flow,
                        cr_volume_ft3           = _nbt_cr_vol,
                        cr_filter_on            = _nbt_cr_filt,
                        end_time_hr             = _nbt_end,
                        breathing_rate_m3_s     = _nbt_br,
                        eab_integration_hr      = _nbt_eab_int,
                        lpz_integration_hr      = _nbt_lpz_int,
                        release_overrides       = _rels,
                        credit_decay            = _nbt_credit_decay,
                        credit_deposition       = _nbt_credit_deposition,
                        local_namespace          = local_namespace,
                    )

                _dose_gap = _nbt_call(_nbt_gap_releases, _nbt_gap_dur) if _nbt_has_release(_nbt_gap_releases) else None
                _dose_eiv = _nbt_call(_nbt_eiv_releases, _nbt_eiv_dur) if _nbt_has_release(_nbt_eiv_releases) else None

                if _dose_gap is not None and _dose_eiv is not None:
                    _dose = dict(_dose_gap)
                    _dose["tede_eab_rem"] = _dose_gap["tede_eab_rem"] + _dose_eiv["tede_eab_rem"]
                    _dose["tede_lpz_rem"] = _dose_gap["tede_lpz_rem"] + _dose_eiv["tede_lpz_rem"]
                    _dose["tede_cr_rem"]  = _dose_gap["tede_cr_rem"]  + _dose_eiv["tede_cr_rem"]
                    _dose["runtime_ms"]   = _dose_gap.get("runtime_ms", 0.0) + _dose_eiv.get("runtime_ms", 0.0)

                    _gd = {}
                    _all_g = set(_dose_gap.get("group_doses", {}).keys()) | set(_dose_eiv.get("group_doses", {}).keys())
                    for _g in _all_g:
                        _gg = _dose_gap.get("group_doses", {}).get(_g, {})
                        _ee = _dose_eiv.get("group_doses", {}).get(_g, {})
                        _gd[_g] = {
                            "released_ci": float(_gg.get("released_ci", 0.0)) + float(_ee.get("released_ci", 0.0)),
                            "tede_eab":    float(_gg.get("tede_eab",    0.0)) + float(_ee.get("tede_eab",    0.0)),
                            "tede_lpz":    float(_gg.get("tede_lpz",    0.0)) + float(_ee.get("tede_lpz",    0.0)),
                            "tede_cr":     float(_gg.get("tede_cr",     0.0)) + float(_ee.get("tede_cr",     0.0)),
                        }
                    _dose["group_doses"] = _gd

                    # Combine phase-specific distance tables.  _dose_gap was
                    # used as the base dictionary, so without this block the
                    # TEDE-vs-distance table would show only the gap-phase dose,
                    # while the summary table correctly shows gap + early-IV.
                    _gap_dt = _dose_gap.get("dist_table", []) or []
                    _eiv_dt = _dose_eiv.get("dist_table", []) or []
                    _eiv_by_d = {
                        int(_r.get("distance_m")): _r
                        for _r in _eiv_dt
                        if _r.get("distance_m") is not None
                    }
                    _combined_dt = []
                    for _gr in _gap_dt:
                        _d = int(_gr.get("distance_m"))
                        _er = _eiv_by_d.get(_d, {})
                        _combined_dt.append({
                            "distance_m": _d,
                            "chi_q_s_m3": _gr.get("chi_q_s_m3", _er.get("chi_q_s_m3", "")),
                            "eab_integration_tede_rem": (
                                float(_gr.get("eab_integration_tede_rem", 0.0)) +
                                float(_er.get("eab_integration_tede_rem", 0.0))
                            ),
                            "lpz_integration_tede_rem": (
                                float(_gr.get("lpz_integration_tede_rem", 0.0)) +
                                float(_er.get("lpz_integration_tede_rem", 0.0))
                            ),
                        })

                    # Include any early-IV distance rows not present in the gap table.
                    _gap_ds = {int(_r.get("distance_m")) for _r in _gap_dt if _r.get("distance_m") is not None}
                    for _er in _eiv_dt:
                        _d = int(_er.get("distance_m"))
                        if _d not in _gap_ds:
                            _combined_dt.append({
                                "distance_m": _d,
                                "chi_q_s_m3": _er.get("chi_q_s_m3", ""),
                                "eab_integration_tede_rem": float(_er.get("eab_integration_tede_rem", 0.0)),
                                "lpz_integration_tede_rem": float(_er.get("lpz_integration_tede_rem", 0.0)),
                            })

                    _combined_dt.sort(key=lambda _r: _r["distance_m"])
                    _dose["dist_table"] = _combined_dt

                    _dose["inputs"] = dict(_dose_gap.get("inputs", {}))
                    _dose["inputs"]["gap_release_duration_hr"] = _nbt_gap_dur
                    _dose["inputs"]["early_iv_duration_hr"] = _nbt_eiv_dur
                    _dose["inputs"]["release_duration_model"] = "phase-specific: gap + early in-vessel"

                elif _dose_gap is not None:
                    _dose = _dose_gap
                    _dose["inputs"]["gap_release_duration_hr"] = _nbt_gap_dur
                    _dose["inputs"]["release_duration_model"] = "gap phase only"

                elif _dose_eiv is not None:
                    _dose = _dose_eiv
                    _dose["inputs"]["early_iv_duration_hr"] = _nbt_eiv_dur
                    _dose["inputs"]["release_duration_model"] = "early in-vessel phase only"

                else:
                    raise RuntimeError(
                        "No NOTBADTRAD release phases were generated despite nonzero source term."
                    )
                # Regulatory limits (10 CFR 50.67 / RG 1.183)
                _LIMIT_EAB = 25.0  # rem TEDE at EAB
                _LIMIT_LPZ = 25.0  # rem TEDE at LPZ
                _LIMIT_CR  = 5.0   # rem TEDE at Control Room

                # Print summary
                print()
                print("NOTBADTRAD Dose Screening (RG 1.183 release fractions from FLARE):")
                print(f"  EAB TEDE : {_dose['tede_eab_rem']:.4e} rem"
                      f"  ({'PASS' if _dose['tede_eab_rem'] < _LIMIT_EAB else 'EXCEEDS LIMIT'},"
                      f" limit {_LIMIT_EAB} rem)")
                print(f"  LPZ TEDE : {_dose['tede_lpz_rem']:.4e} rem"
                      f"  ({'PASS' if _dose['tede_lpz_rem'] < _LIMIT_LPZ else 'EXCEEDS LIMIT'},"
                      f" limit {_LIMIT_LPZ} rem)")
                print(f"  CR  TEDE : {_dose['tede_cr_rem']:.4e} rem"
                      f"  ({'PASS' if _dose['tede_cr_rem'] < _LIMIT_CR  else 'EXCEEDS LIMIT'},"
                      f" limit {_LIMIT_CR} rem)")
                print(f"  Runtime  : {_dose['runtime_ms']:.1f} ms")

                # Build dose DataFrame for output
                _it = _dose.get("integration_times_hr", {})
                _eab_int = float(_it.get("eab", _nbt_eab_int))
                _lpz_int = float(_it.get("lpz", _nbt_lpz_int))
                _cr_int  = float(_it.get("cr",  _nbt_end))
                _dose_summary = [
                    {"Location": "EAB", "TEDE (rem)": _dose["tede_eab_rem"],
                     "Limit (rem)": f"{_LIMIT_EAB:g} rem ({_eab_int:g} hr)",
                     "Margin": _LIMIT_EAB - _dose["tede_eab_rem"],
                     "Result": "PASS" if _dose["tede_eab_rem"] < _LIMIT_EAB else "EXCEEDS"},
                    {"Location": "LPZ", "TEDE (rem)": _dose["tede_lpz_rem"],
                     "Limit (rem)": f"{_LIMIT_LPZ:g} rem ({_lpz_int:g} hr)",
                     "Margin": _LIMIT_LPZ - _dose["tede_lpz_rem"],
                     "Result": "PASS" if _dose["tede_lpz_rem"] < _LIMIT_LPZ else "EXCEEDS"},
                    {"Location": "Control Room", "TEDE (rem)": _dose["tede_cr_rem"],
                     "Limit (rem)": f"{_LIMIT_CR:g} rem ({_cr_int:g} hr)",
                     "Margin": _LIMIT_CR - _dose["tede_cr_rem"],
                     "Result": "PASS" if _dose["tede_cr_rem"] < _LIMIT_CR else "EXCEEDS"},
                ]
                _dose_groups = [
                    {"Group":       gn,
                     "Released (Ci)": gd["released_ci"],
                     "EAB (rem)":  gd["tede_eab"],
                     "LPZ (rem)":  gd["tede_lpz"],
                     "CR (rem)":   gd["tede_cr"],
                     "Release frac": _nbt_releases.get(gn, 0.0)}
                    for gn, gd in _dose["group_doses"].items()
                ]
                _dose_df = {
                    "summary":      pd.DataFrame(_dose_summary),
                    "groups":       pd.DataFrame(_dose_groups),
                    "chi_q":        _dose.get("chi_q_used", {}),
                    "inputs":       _dose.get("inputs", {}),
                    "runtime_ms":   _dose["runtime_ms"],
                    "dist_fit":     _dose.get("dist_fit"),
                    "dist_table":   pd.DataFrame(_dose.get("dist_table", [])),
                    "iodine_spike": {
                        "coolant_act_uci_g":       _coolant_act_uci_g,
                        "primary_mass_kg":         _primary_mass_g / 1000.0,
                        "equil_spike_ci":          _spike_ci,
                        "accid_spike_ci":          _spike_accid_ci,
                        "spike_multiplier_equil":  500,
                        "spike_multiplier_accid":  335,
                        "f_eq_spike":              _f_gap_equil,
                        "f_accid_spike":           _f_gap_accid,
                        "f_combined":              _f_isp_combined,
                        "model_used":              _spike_model_used,
                        "is_loca":                 _is_loca,
                        "scram":                   _scram_occurred,
                        "porv":                    _porv_opened,
                        "dnb":                     _dnb_occurred,
                        "fuel_damage":             _has_fuel_damage,
                        "N_dnb":                   _N_dnb_peak,
                    },
                }
            except Exception as _nbt_err:
                print(f"WARNING: NOTBADTRAD dose calculation failed: {_nbt_err}")
                _dose_df = None
        else:
            _source_term_df = None
            # Even with no rod failures, compute iodine spike dose from
            # pre-existing coolant activity (applies to all transients)
            _dose_df = None
            try:
                _isp_releases = {"Halogens": _f_gap_equil if _f_gap_equil > 0 else 0.0}
                if any(v > 0 for v in _isp_releases.values()):
                    _nbt_power    = float(get_variable(local_namespace, "total_power", 575.0))
                    _nbt_vol      = float(get_variable(local_namespace, "nbt_containment_volume_ft3", 2.74e6))
                    _nbt_leak     = float(get_variable(local_namespace, "nbt_leak_rate_frac_per_day", 0.001))
                    _nbt_sprays   = bool( get_variable(local_namespace, "nbt_sprays_on", True))
                    _nbt_spray_t0 = float(get_variable(local_namespace, "nbt_spray_start_hr", 1/60))
                    _nbt_spray_t1 = float(get_variable(local_namespace, "nbt_spray_stop_hr", 24.0))
                    _nbt_wind     = float(get_variable(local_namespace, "nbt_wind_speed_m_s", 1.0))
                    _nbt_stab     = str(  get_variable(local_namespace, "nbt_stability_class", "F"))
                    _nbt_ht       = float(get_variable(local_namespace, "nbt_release_height_m", 0.0))
                    _nbt_eab      = float(get_variable(local_namespace, "nbt_distance_eab_m", 914.0))
                    _nbt_lpz      = float(get_variable(local_namespace, "nbt_distance_lpz_m", 4800.0))
                    _nbt_cr_d     = float(get_variable(local_namespace, "nbt_distance_cr_intake_m", 100.0))
                    _nbt_cr_flow  = float(get_variable(local_namespace, "nbt_cr_flow_cfm", 100.0))
                    _nbt_cr_vol   = float(get_variable(local_namespace, "nbt_cr_volume_ft3", 20000.0))
                    _nbt_cr_filt  = bool( get_variable(local_namespace, "nbt_cr_filter_on", True))
                    _nbt_end      = float(get_variable(local_namespace, "nbt_end_time_hr", 24.0))
                    _nbt_eab_int  = float(get_variable(local_namespace, "nbt_eab_integration_hr", 2.0))
                    _nbt_lpz_int  = float(get_variable(local_namespace, "nbt_lpz_integration_hr", 720.0))
                    _nbt_br       = float(get_variable(local_namespace, "nbt_breathing_rate_m3_s", 3.47e-4))
                    _nbt_dur      = float(get_variable(local_namespace, "nbt_gap_release_duration_hr", 0.5))
                    _nbt_credit_decay      = bool(get_variable(local_namespace, "nbt_credit_decay", False))
                    _nbt_credit_deposition = bool(get_variable(local_namespace, "nbt_credit_deposition", False))
                    _dose_isp = _notbadtrad(
                        power_mwt=_nbt_power, containment_volume_ft3=_nbt_vol,
                        leak_rate_frac_per_day=_nbt_leak, sprays_on=_nbt_sprays,
                        spray_start_hr=_nbt_spray_t0, spray_stop_hr=_nbt_spray_t1,
                        release_duration_hr=_nbt_dur, wind_speed_m_s=_nbt_wind,
                        stability_class=_nbt_stab, release_height_m=_nbt_ht,
                        distance_eab_m=_nbt_eab, distance_lpz_m=_nbt_lpz,
                        distance_cr_intake_m=_nbt_cr_d, cr_flow_cfm=_nbt_cr_flow,
                        cr_volume_ft3=_nbt_cr_vol, cr_filter_on=_nbt_cr_filt,
                        end_time_hr=_nbt_end, breathing_rate_m3_s=_nbt_br,
                        eab_integration_hr=_nbt_eab_int, lpz_integration_hr=_nbt_lpz_int,
                        release_overrides=_isp_releases,
                        credit_decay=_nbt_credit_decay,
                        credit_deposition=_nbt_credit_deposition,
                        local_namespace=local_namespace,
                    )
                    _LIMIT_EAB = 25.0; _LIMIT_LPZ = 25.0; _LIMIT_CR = 5.0
                    _dose_summary_isp = [
                        {"Location":"EAB","TEDE (rem)":_dose_isp["tede_eab_rem"],
                         "Limit (rem)":f"{_LIMIT_EAB:g} rem ({_dose_isp.get('integration_times_hr', {}).get('eab', 2.0):g} hr)","Margin":_LIMIT_EAB-_dose_isp["tede_eab_rem"],
                         "Result":"PASS" if _dose_isp["tede_eab_rem"]<_LIMIT_EAB else "EXCEEDS"},
                        {"Location":"LPZ","TEDE (rem)":_dose_isp["tede_lpz_rem"],
                         "Limit (rem)":f"{_LIMIT_LPZ:g} rem ({_dose_isp.get('integration_times_hr', {}).get('lpz', 720.0):g} hr)","Margin":_LIMIT_LPZ-_dose_isp["tede_lpz_rem"],
                         "Result":"PASS" if _dose_isp["tede_lpz_rem"]<_LIMIT_LPZ else "EXCEEDS"},
                        {"Location":"Control Room","TEDE (rem)":_dose_isp["tede_cr_rem"],
                         "Limit (rem)":f"{_LIMIT_CR:g} rem ({_dose_isp.get('integration_times_hr', {}).get('cr', 24.0):g} hr)","Margin":_LIMIT_CR-_dose_isp["tede_cr_rem"],
                         "Result":"PASS" if _dose_isp["tede_cr_rem"]<_LIMIT_CR else "EXCEEDS"},
                    ]
                    _dose_df = {
                        "summary":    pd.DataFrame(_dose_summary_isp),
                        "groups":     pd.DataFrame([
                            {"Group": gn, "Released (Ci)": gd["released_ci"],
                             "EAB (rem)": gd["tede_eab"], "LPZ (rem)": gd["tede_lpz"],
                             "CR (rem)": gd["tede_cr"], "Release frac": _isp_releases.get(gn,0.0)}
                            for gn, gd in _dose_isp["group_doses"].items()
                        ]),
                        "chi_q":      _dose_isp.get("chi_q_used", {}),
                        "runtime_ms": _dose_isp["runtime_ms"],
                        "dist_fit":   _dose_isp.get("dist_fit"),
                        "dist_table": pd.DataFrame(_dose_isp.get("dist_table", [])),
                        "iodine_spike": {
                            "coolant_act_uci_g":       _coolant_act_uci_g,
                            "primary_mass_kg":         _primary_mass_g / 1000.0,
                            "equil_spike_ci":          _spike_ci,
                            "accid_spike_ci":          _spike_accid_ci,
                            "spike_multiplier_equil":  500,
                            "spike_multiplier_accid":  335,
                            "f_eq_spike":              _f_gap_equil,
                            "f_accid_spike":           _f_gap_accid,
                            "f_combined":              _f_isp_combined,
                            "model_used":              _spike_model_used,
                            "is_loca":                 _is_loca,
                            "scram":                   _scram_occurred,
                            "porv":                    _porv_opened,
                            "dnb":                     _dnb_occurred,
                            "fuel_damage":             _has_fuel_damage,
                            "N_dnb":                   _N_dnb_peak,
                        },
                    }
            except Exception as _isp_err:
                print(f"WARNING: Iodine spike dose calculation failed: {_isp_err}")
                _dose_df = None



    # Forward-fill the last timestep index — the loop runs range(N-1)
    # so index N-1 is never written. Copy from N-2 so the output reaches
    # the requested end time without a spurious NaN or zero final row.
    _last = number_timesteps - 1
    for _arr in [
            Pressure, Temperature, enthalpy_mix,
            massflow_break, h_break_arr, acc_pres, acc_tgas, massflow_in,
            acc_liqvol, Total_Mass_scaled, ves_ll, net_heat_total,
            x_eq, alpha_void, pump_omega, pump_mdot, pump_velocity, pump_head,
            steam_vel_arr,
            Q_sg, sg_UA_used, sg_UA_dynamic, sg_dT_primary_sec, sg_v_ratio,
            sg_UA_eff, sg_C_primary, sg_flow_frac, sg_open_frac,
            sg_Q_forced_raw, sg_Q_forced, sg_Q_nat,
            sg_flow_cap_W, sg_flow_cap_ratio, sg_flow_cap_warn,
            TTwall, alpha, rkpower_total, DNBR, q_chf, q_hot,
            massflow_PORV, h_break_arr, cvcs_mdot, cvcs_makeup_arr, cvcs_letdown_arr, hpsi_mdot_arr,
            lpsi_mdot_arr, si_pumped_mdot,
            T_fuel_arr, T_hot_clad_arr, T_hot_fuel_arr,
            rho_ext_arr, rho_scram_arr, rho_boron_arr, rcs_boron_ppm,
            rho_D_arr, rho_M_arr, rho_net_arr,
            N_fail_DNB, N_fail_clad, N_fail_gap, N_fail_eiv,
            pzr_level_arr, pzr_level_norm_arr, pzr_mdot_surge_arr]:
        try:
            if hasattr(_arr, "__len__") and len(_arr) > _last:
                _prev = _arr[_last - 1]
                _cur  = _arr[_last]
                # Fill if last index is NaN, or if it dropped abruptly to zero
                # while the previous value was nonzero (zero-initialised arrays)
                if np.isnan(_cur) or (_cur == 0.0 and not np.isnan(_prev) and _prev != 0.0):
                    _arr[_last] = _prev
        except (TypeError, ValueError):
            pass

    if _sg_flow_cap_warn_count > 0:
        print(
            f"WARNING: SG primary-flow capacity warning occurred in {_sg_flow_cap_warn_count} timestep(s). "
            "See the diagnostics CSV columns 'SG Flow-Capacity Warning (-)', "
            "'SG Flow-Capacity Limit (MW)', and 'Core Power / SG Flow Capacity (-)'.",
            flush=True,
        )

    # Write a separate diagnostics CSV with SG UA and heat-removal internals.
    try:
        _diag_df = pd.DataFrame({
            "Time (s)": time,
            "SG Enabled (-)": int(sg_flag),
            "SG Table Flag (-)": int(sg_table_flag),
            "SG Dynamic UA Flag (-)": int(sg_dynamic_ua_flag),
            "SG Dynamic UA Active (-)": sg_dynamic_active_arr,
            "SG UA User Supplied (-)": int(_UA_sg_user_supplied),
            "UA_sg_rated Used (W/K)": sg_UA_used,
            "UA_sg_rated Used (MW/K)": sg_UA_used / 1.0e6,
            "UA_sg_dynamic Candidate (W/K)": sg_UA_dynamic,
            "UA_sg_dynamic Candidate (MW/K)": sg_UA_dynamic / 1.0e6,
            "SG Primary-Secondary dT (K)": sg_dT_primary_sec,
            "RCS Temperature (K)": Temperature,
            "SG Secondary Temperature (K)": np.full(number_timesteps, T_sec_K),
            "Total Power Input (MW)": np.full(number_timesteps, float(total_power)),
            "RK Total Power (MW)": rkpower_total / 1.0e6,
            "SG Flow-Capacity Limit (MW)": sg_flow_cap_W / 1.0e6,
            "Core Power / SG Flow Capacity (-)": sg_flow_cap_ratio,
            "SG Flow-Capacity Warning (-)": sg_flow_cap_warn,
            "Pump Rated Speed (rpm)": np.full(number_timesteps, pump_model_config.get("pump_rated_speed_rpm", np.nan)),
            "Pump Initial Speed Input (rpm)": np.full(number_timesteps, float(pump_speed_rpm)),
            "Pump Rated Flow (m3/s)": np.full(number_timesteps, pump_model_config.get("pump_rated_flow_m3s", np.nan)),
            "Pump Rated Head (m)": np.full(number_timesteps, pump_model_config.get("pump_rated_head_m", np.nan)),
            "Pump Rated Torque (N-m)": np.full(number_timesteps, pump_model_config.get("pump_rated_torque_Nm", np.nan)),
            "Pump Rated Density (kg/m3)": np.full(number_timesteps, pump_model_config.get("pump_rated_density_kg_m3", np.nan)),
            "Pump Inertia (kg-m2)": np.full(number_timesteps, pump_model_config.get("pump_inertia_kg_m2", np.nan)),
            "Pump Flow Area (m2)": np.full(number_timesteps, pump_model_config.get("pump_flow_area_m2", np.nan)),
            "Pump Friction Torque Const (N-m)": np.full(number_timesteps, pump_model_config.get("pump_friction_torque_Nm", np.nan)),
            "Pump Friction Torque Speed2 (N-m)": np.full(number_timesteps, pump_model_config.get("pump_friction_speed2_Nm", np.nan)),
            "Pump Orifice Area (-)": np.full(number_timesteps, float(pump_orifice_area)),
            "Pump Mass Flow (kg/s)": pump_mdot,
            "Pump Volumetric Flow Ratio (-)": sg_v_ratio,
            "SG UA_eff Flow Scaled (W/K)": sg_UA_eff,
            "SG Primary Heat Capacity Rate (W/K)": sg_C_primary,
            "SG Explicit Flow Fraction (-)": sg_flow_frac,
            "SG Open Fraction (-)": sg_open_frac,
            "SG Natural Circulation Active (-)": sg_nat_active_arr,
            "SG Q Forced Raw (MW)": sg_Q_forced_raw / 1.0e6,
            "SG Q Forced After Multipliers (MW)": sg_Q_forced / 1.0e6,
            "SG Q Natural Floor (MW)": sg_Q_nat / 1.0e6,
            "SG Q Final Used (MW)": Q_sg / 1.0e6,
        })
        _diag_name = f"{wkstbase}_diag.csv"
        _diag_df.to_csv(_diag_name, index=False)
        print(f"DIAGNOSTICS: wrote {_diag_name}", flush=True)
    except Exception as _diag_err:
        print(f"WARNING: could not write diagnostics CSV: {_diag_err}", flush=True)

    # Pre-compute accumulator level for output dict
    _acc_level_m = (acc_liqvol / acc_area) if (acc_area is not None and float(acc_area) > 0) else np.zeros(number_timesteps)

    # ── Early output CSV write ────────────────────────────────────────────────
    # Write a minimal output CSV directly from pwr_sim, before post_processing
    # is called, so that a crash anywhere in post_processing (Excel writer,
    # matplotlib, source term, dose) still leaves a usable CSV on disk.
    # This mirrors the same guarantee the diag CSV already has.
    # post_processing will overwrite this with the full-fidelity version if it
    # completes successfully.
    _early_csv_name = f"{wkstbase}_out.csv"
    try:
        _early_data = {
            "Time (s)":                    time,
            "RCS Pressure (kPa)":          Pressure / 1e3,
            "RCS Temperature (K)":         Temperature,
            "Break Flow (kg/s)":           massflow_break,
            "Break Enthalpy (kJ/kg)":      h_break_arr * 1e-3,
            "Accumulator Flow (kg/s)":     massflow_in,
            "Accumulator Liquid Volume (m3)": acc_liqvol,
            "Total Mass Scaled":           Total_Mass_scaled,
            "Vessel Level (m)":            ves_ll,
            "Core Power (MW)":             net_heat_total / 1e6,
            "Pump Speed (rpm)":            pump_omega,
            "Pump Mass Flow (kg/s)":       pump_mdot,
            "Core Coolant Velocity (m/s)": pump_velocity,
            "Pump Velocity (m/s)":         pump_velocity,
            "Steam Cooling Velocity (m/s)": steam_vel_arr,
            "SG Heat Removal (MW)":        Q_sg / 1e6,
            "Equilibrium Quality (-)":     x_eq,
            "Void Fraction (-)":           alpha_void,
            "Clad Surface Temp (K)":       TTwall,
            "Clad HTC (W/m2-K)":          alpha,
            "RK Total Power (MW)":         rkpower_total / 1e6,
            "DNBR":                        DNBR,
            "Hot Pin Clad Temp (K)":       T_hot_clad_arr,
            "Hot Pin Fuel Temp (K)":       T_hot_fuel_arr,
            "HPSI Flow (kg/s)":            hpsi_mdot_arr,
            "LPSI Flow (kg/s)":            lpsi_mdot_arr,
            "Pressurizer Level (m)":       pzr_level_arr,
            "Reactivity net (pcm)":        rho_net_arr,
        }
        pd.DataFrame(_early_data).to_csv(_early_csv_name, index=False)
        print(f"Early output CSV written: {_early_csv_name}", flush=True)
    except Exception as _early_err:
        print(f"WARNING: could not write early output CSV: {_early_err}", flush=True)

    print("POST-PROCESSING", flush=True)
    try:
        post_processing(
            time, Pressure, Temperature, massflow_break,
            acc_pres, acc_tgas, massflow_in,
            acc_liqvol, Total_Mass_scaled, ves_ll, net_heat_total,
        R5_p, R5_t, R5_vessel_mass_scaled, R5_mdot,
        R5_accp, R5_acct, R5_massflow_in, R5_power,
        x_eq, wkstbase, title_name_include_line_1="",
        pump_omega=pump_omega, pump_mdot=pump_mdot,
        pump_velocity=pump_velocity, pump_head=pump_head,
        steam_vel_arr=steam_vel_arr,
        Q_sg=Q_sg,
        alpha_void_out=alpha_void,
        TTwall_out=TTwall, alpha_out=alpha,
        rkpower_total_out=rkpower_total,
        core_flag=bool(core_flag),
        DNBR_out=DNBR if core_flag else None,
        q_chf_out=q_chf if core_flag else None,
        q_hot_out=q_hot if core_flag else None,
        F_r=F_r,
        F_z=F_z,
        massflow_PORV_out=massflow_PORV,
        h_break_out=h_break_arr,
        cvcs_mdot_out=cvcs_mdot,
        cvcs_makeup_out=cvcs_makeup_arr,
        cvcs_letdown_out=cvcs_letdown_arr,
        acc_level_out=_acc_level_m,
        hpsi_mdot_out=hpsi_mdot_arr,
        lpsi_mdot_out=lpsi_mdot_arr,
        si_pumped_out=si_pumped_mdot,
        T_fuel_out=T_fuel_arr     if core_flag else None,
        T_hot_clad_out=T_hot_clad_arr if core_flag else None,
        T_hot_fuel_out=T_hot_fuel_arr if core_flag else None,
        rho_ext_out=rho_ext_arr     if core_flag else None,
        rho_scram_out=rho_scram_arr if core_flag else None,
        rho_boron_out=rho_boron_arr if core_flag else None,
        rcs_boron_ppm_out=rcs_boron_ppm if core_flag else None,
        rho_D_out=rho_D_arr         if core_flag else None,
        rho_M_out=rho_M_arr         if core_flag else None,
        rho_net_out=rho_net_arr     if core_flag else None,
        N_fail_DNB_out=N_fail_DNB if core_flag else None,
        N_fail_clad_out=N_fail_clad if core_flag else None,
        N_fail_gap_out=N_fail_gap   if core_flag else None,
        N_fail_eiv_out=N_fail_eiv   if core_flag else None,
        source_term_out=_source_term_df,
        dose_out=_dose_df,
        iodine_spike_dose_out=_iodine_spike_dose_df,
        input_echo_out=_input_echo_log,
        pzr_level_out=pzr_level_arr,
        pzr_level_norm_out=pzr_level_norm_arr,
        pzr_mdot_surge_out=pzr_mdot_surge_arr,
        minimum_output=minimum_output,
        N_pins_out=N_pins,
        d_pin_out=d_pin,
        L_heated_out=L_heated,
        delta_clad_out=delta_clad,
        k_sigma_out=k_sigma,
        Pressure_report_out=np.where(np.isfinite(Pressure_report), Pressure_report, Pressure),
        )
    except Exception as _pp_err:
        print(f"WARNING: post_processing failed: {_pp_err}", flush=True)
        import traceback
        traceback.print_exc()



    # ── Optional FLARECON containment response coupling ──────────────────────
    # If the archived input workbook contains a case-insensitive FLARECON
    # worksheet, run the containment simulator using the just-written FLARE CSV
    # as the mass/energy/H2 source term.  The Risk tool sets
    # FLARE_SKIP_FLARECON=1 because that workflow is dose-consequence focused.
    if os.environ.get("FLARE_SKIP_FLARECON", "").strip() not in ("1", "true", "TRUE", "yes", "YES"):
        try:
            _flare_csv_for_con = Path.cwd() / f"{wkstbase}_out.csv"
            _flare_in_for_con  = _archived_input_xlsx
            if _flare_csv_for_con.exists() and _flare_in_for_con.exists():
                import flarecon_sim as _flarecon_sim
                _con = _flarecon_sim.run_integrated_flarecon(
                    flare_case=wkstbase,
                    flare_input_xlsx=str(_flare_in_for_con),
                    flare_output_csv=str(_flare_csv_for_con),
                    output_dir=Path.cwd(),
                    make_plots=(minimum_output == 0),
                )
                if _con is not None:
                    print("FLARECON: containment response simulation complete", flush=True)
        except Exception as _con_err:
            print(f"WARNING: FLARECON integration failed: {_con_err}", flush=True)


    # Remove the private temporary Excel snapshot. The archived XLSX and CSV
    # snapshots remain in the case output folder for traceability.
    try:
        _tmp_in.unlink()
    except Exception:
        pass


def user_input(prompt):
    while True:
        print(prompt, end="", flush=True)
        val = input().strip()
        if val:
            # Strip trailing _in or _out suffix if user pastes the full sheet name
            for suffix in ("_in.xlsx", "_out.xlsx", "_in", "_out"):
                if val.endswith(suffix):
                    val = val[:-len(suffix)]
                    break
            if val: return val
        print("Input cannot be empty. Please try again.")


print("\n", flush=True)

# (pump_module.py is no longer a separate file)
# ═══════════════════════════════════════════════════════════════════════════════
# NOTBADTRAD  —  Embedded radiological screening module
# Aligned with RG 1.183 Revision 1 (October 2023) Table 3
# 9 chemical groups; Pasquill-Gifford chi/Q; TEDE at EAB, LPZ, Control Room
# Called automatically from pwr_sim() when source term is non-zero.
# ═══════════════════════════════════════════════════════════════════════════════

    # Core inventories from NUREG-1465 Table 2 (3411 MWt PWR at shutdown).
    # Scaled to actual plant power via scale = total_power / 3411.
_NBT_GROUPS = {
    "NG":           dict(label="Noble Gases",        elements="Kr, Xe",
                         t_half_hr=0.78,    inv_ci=3.697e+08, ci_per_fission=1.416509e-18, release=1.00,
                         dcf_inh=6.5e-5,   dcf_cloud=2.8e-2),
    "Halogens":     dict(label="Halogens",           elements="I, Br",
                         t_half_hr=2.41,    inv_ci=1.110e+08, ci_per_fission=1.437732e-18, release=0.40,
                         dcf_inh=1.5e1,    dcf_cloud=3.1e-2),
    "Alkali_metals":dict(label="Alkali Metals",      elements="Cs, Rb",
                         t_half_hr=2080.0,  inv_ci=1.360e7, ci_per_fission=1.470166e-18, release=0.30,
                         dcf_inh=5.0e1,    dcf_cloud=4.7e-3),
    "Te_group":     dict(label="Tellurium Group",    elements="Te, Sb, Se",
                         t_half_hr=5.04,    inv_ci=1.888e+08, ci_per_fission=1.308530e-18, release=0.05,
                         dcf_inh=1.1e1,    dcf_cloud=1.6e-2),
    "Ba_Sr":        dict(label="Barium/Strontium",   elements="Ba, Sr",
                         t_half_hr=3.68,    inv_ci=2.389e+08, ci_per_fission=1.955664e-18, release=0.02,
                         dcf_inh=1.7e1,    dcf_cloud=2.6e-2),
    "Noble_metals": dict(label="Noble Metals",       elements="Ru, Rh, Pd, Mo, Tc",
                         t_half_hr=11.24,   inv_ci=5.241e+08, ci_per_fission=3.185190e-18, release=0.0025,
                         dcf_inh=1.26e1,   dcf_cloud=2.8e-2),
    "Ce_group":     dict(label="Cerium Group",       elements="Ce, Pu, Np",
                         t_half_hr=59.88,   inv_ci=1.544e+09, ci_per_fission=8.014688e-19, release=0.0005,
                         dcf_inh=9.6e0,    dcf_cloud=1.1e-2),
    "Lanthanides":  dict(label="Lanthanides",        elements="La, Zr, Nd, Eu, Nb, Pm, Pr, Sm, Y",
                         t_half_hr=6.67,    inv_ci=7.683e+08, ci_per_fission=5.970780e-18, release=0.0002,
                         dcf_inh=3.0e0,    dcf_cloud=5.0e-2),
    "U_actinides":  dict(label="Uranium/Actinides",  elements="U, Th",
                         t_half_hr=162.0,   inv_ci=8.450e+03, ci_per_fission=0.000000e+00, release=0.0002,
                         dcf_inh=8.0e1,    dcf_cloud=1.0e-5),
}

_NBT_SPRAY_REMOVAL = {
    "NG": 0.0,
    "Halogens": 0.90*1.5e-3 + 0.05*4.17e-5 + 0.05*0.0,
    "Alkali_metals":1.5e-3, "Te_group":1.5e-3, "Ba_Sr":1.5e-3,
    "Noble_metals":1.5e-3,  "Ce_group":1.5e-3, "Lanthanides":1.5e-3,
    "U_actinides":1.5e-3,
}
_NBT_NAT_DEP = {k: (0.0 if k == "NG" else 4.17e-5) for k in _NBT_GROUPS}
_NBT_CR_DF   = {
    "NG": 1.0,
    "Halogens": 0.90*100.0 + 0.05*20.0 + 0.05*20.0,
    **{k: 100.0 for k in ["Alkali_metals","Te_group","Ba_Sr",
                           "Noble_metals","Ce_group","Lanthanides","U_actinides"]},
}



def _nbt_estimate_core_fissions_from_geometry(
        n_fuel_rods=18200.0,
        fuel_radius_m=0.00418,
        fuel_active_length_m=2.408,
        rho_uo2_kg_m3=10400.0,
        u235_enrichment=0.05,
        fissile_burned_fraction=0.05):
    """Reduced-order estimate of accumulated fissions from fuel geometry.

    This is intentionally simple and is used only to normalize source-term
    inventory.  It assumes UO2 fuel, a specified U-235 enrichment, and a
    specified fraction of the initial U-235 inventory fissioned.
    """
    _NA = 6.02214076e23
    _u_mass_frac_in_uo2 = 238.0 / (238.0 + 2.0 * 16.0)
    _fuel_vol_m3 = float(n_fuel_rods) * _math.pi * float(fuel_radius_m)**2 * float(fuel_active_length_m)
    _m_uo2_kg = _fuel_vol_m3 * float(rho_uo2_kg_m3)
    _m_u_kg = _m_uo2_kg * _u_mass_frac_in_uo2
    _m_u235_initial_kg = _m_u_kg * float(u235_enrichment)
    _m_u235_fissioned_kg = _m_u235_initial_kg * float(fissile_burned_fraction)
    return (_m_u235_fissioned_kg / 0.235) * _NA


def _nbt_apply_inventory_model(groups, local_namespace=None):
    """Set NBT group inventories from geometry/fission-fraction defaults.

    Defaults are calibrated to the typPWR Table 12 EOC fission-product inventory
    using Ci/fission coefficients for each RG 1.183/NBT group.  Users may
    override either all default geometry inputs or each group inventory directly.

    Per-group override names:
        nbt_inv_NG_Ci
        nbt_inv_Halogens_Ci
        nbt_inv_Alkali_metals_Ci
        nbt_inv_Te_group_Ci
        nbt_inv_Ba_Sr_Ci
        nbt_inv_Noble_metals_Ci
        nbt_inv_Lanthanides_Ci
        nbt_inv_Ce_group_Ci
        nbt_inv_U_actinides_Ci
    """
    if local_namespace is None:
        local_namespace = {}

    _model = str(get_variable(local_namespace, "nbt_inventory_model",
                              "geometry_typpwr_table12")).strip("'\"").lower()

    if _model in ("fixed", "legacy", "static"):
        return groups, {"model": _model, "n_fissions": None}

    _n_pins_th = float(get_variable(local_namespace, "N_pins", 18200.0))
    _n_rods = float(get_variable(local_namespace, "nbt_n_fuel_rods", _n_pins_th))
    if _n_rods != _n_pins_th:
        print(
            f"WARNING: nbt_n_fuel_rods ({_n_rods:.0f}) differs from N_pins "
            f"({_n_pins_th:.0f}). Source-term inventory will be scaled to "
            f"{_n_rods:.0f} rods while the thermal-hydraulic core model uses "
            f"{_n_pins_th:.0f} rods. Set nbt_n_fuel_rods = N_pins in the input "
            f"file to suppress this warning.",
            flush=True,
        )
    _r_fuel = float(get_variable(local_namespace, "nbt_fuel_radius_m", 0.00418))
    _l_fuel = float(get_variable(local_namespace, "nbt_fuel_active_length_m", 2.408))
    _rho    = float(get_variable(local_namespace, "nbt_rho_uo2_kg_m3", 10400.0))
    _enr    = float(get_variable(local_namespace, "nbt_u235_enrichment", 0.05))
    _burn   = float(get_variable(local_namespace, "nbt_fissile_burned_fraction", 0.05))

    _n_fiss = _nbt_estimate_core_fissions_from_geometry(
        n_fuel_rods=_n_rods,
        fuel_radius_m=_r_fuel,
        fuel_active_length_m=_l_fuel,
        rho_uo2_kg_m3=_rho,
        u235_enrichment=_enr,
        fissile_burned_fraction=_burn)

    for _key, _g in groups.items():
        _default_ci = float(_g.get("ci_per_fission", 0.0)) * _n_fiss
        _override_name = f"nbt_inv_{_key}_Ci"
        _g["inv_ci"] = float(get_variable(local_namespace, _override_name, _default_ci))

    _meta = {
        "model": _model,
        "n_fissions": _n_fiss,
        "n_fuel_rods": _n_rods,
        "fuel_radius_m": _r_fuel,
        "fuel_active_length_m": _l_fuel,
        "rho_uo2_kg_m3": _rho,
        "u235_enrichment": _enr,
        "fissile_burned_fraction": _burn,
    }
    return groups, _meta

def _nbt_solve_group(inv_ci, release_frac, release_duration_hr,
                     containment_volume_ft3, leak_rate_frac_per_day,
                     lam_decay_s, lam_removal_s, end_time_hr, dt_hr=0.05,
                     credit_decay=False, credit_deposition=False):
    """Analytical single-group containment→environment solver.
    credit_decay=False       : conservative — no radioactive decay credited
    credit_deposition=False  : conservative — no spray/natural deposition credited
    """
    _lam_d   = lam_decay_s   if credit_decay       else 0.0
    _lam_rem = lam_removal_s if credit_deposition  else 0.0
    V_m3     = containment_volume_ft3 * 0.0283168
    lam_leak = leak_rate_frac_per_day / 86400.0
    released = inv_ci * release_frac
    rel_s    = release_duration_hr * 3600.0
    src_rate = released / rel_s if rel_s > 0 else 0.0
    lam_c    = _lam_d + _lam_rem + lam_leak
    lam_e    = _lam_d
    end_s    = end_time_hr * 3600.0
    N_c = cum_leaked = 0.0
    t = 0.0
    # Use a timestep no larger than the release duration to avoid
    # integrating the source over a longer interval than it actually lasts.
    dt = min(dt_hr, release_duration_hr / 2.0 if rel_s > 0 else dt_hr) * 3600.0
    dt = max(dt, 1.0)   # floor at 1 s
    while t < end_s:
        dt_a = min(dt, end_s - t)
        S    = src_rate if t < rel_s else 0.0
        x    = lam_c * dt_a
        if x > 1e-6:
            ex = _math.exp(-x)
            if S > 0:
                ss_c    = S / lam_c
                N_c_new = (N_c - ss_c) * ex + ss_c
                avg_Nc  = ss_c + (N_c - ss_c) * (1.0 - ex) / x
            else:
                N_c_new = N_c * ex
                avg_Nc  = N_c * (1.0 - ex) / x
        else:
            # lam_c*dt_a << 1: linear approximation avoids cancellation
            N_c_new = N_c + (S - lam_c * N_c) * dt_a
            avg_Nc  = N_c + (S - lam_c * N_c) * dt_a / 2.0
        # Accumulate Ci leaked to environment.
        # TEDE = cum_leaked [Ci] x chi_q [s/m3] x DCF [rem*m3/(Ci*s)] = rem
        cum_leaked += lam_leak * avg_Nc * dt_a
        N_c, t = N_c_new, t + dt_a
    return cum_leaked


def _nbt_chi_q(distance_m, release_height_m=0.0,
               stability_class="F", wind_speed_m_s=1.0):
    """Pasquill-Gifford atmospheric dispersion factor [s/m³]."""
    pg = {"A":(0.22,0.896,0.20,0.890), "B":(0.16,0.896,0.12,0.890),
          "C":(0.11,0.896,0.08,0.890), "D":(0.08,0.896,0.06,0.890),
          "E":(0.06,0.896,0.03,0.890), "F":(0.04,0.896,0.016,0.890)}
    a_y, b_y, a_z, b_z = pg.get(stability_class, pg["F"])
    x_km   = distance_m / 1000.0
    sig_y  = a_y * (x_km ** b_y) * 1000.0
    sig_z  = min(a_z * (x_km ** b_z) * 1000.0, 5000.0)
    denom  = 2.0 * _math.pi * sig_y * sig_z * wind_speed_m_s
    return (_math.exp(-0.5 * (release_height_m / sig_z) ** 2) / denom
            if denom > 0 else 0.0)


def _notbadtrad(
        power_mwt=3411.0, containment_volume_ft3=2.74e6,
        leak_rate_frac_per_day=0.001, sprays_on=True,
        spray_start_hr=1/60, spray_stop_hr=24.0,
        release_duration_hr=1.8,
        wind_speed_m_s=1.0, stability_class="F", release_height_m=0.0,
        distance_eab_m=914.0, distance_lpz_m=4800.0,
        distance_cr_intake_m=100.0, cr_flow_cfm=100.0,
        cr_volume_ft3=20000.0, cr_filter_on=True,
        end_time_hr=24.0, breathing_rate_m3_s=3.47e-4, dt_hr=0.05,
        eab_integration_hr=None, lpz_integration_hr=None,
        release_overrides=None,
        credit_decay=False,
        credit_deposition=False,
        local_namespace=None):
    """
    NOTBADTRAD screening calculation.
    release_overrides: dict {group_key: total_release_fraction} from FLARE.
    If supplied, overrides the default RG 1.183 release fractions.
    """
    import copy as _copy
    import time as _time
    t0 = _time.perf_counter()

    groups = _copy.deepcopy(_NBT_GROUPS)
    groups, _inventory_meta = _nbt_apply_inventory_model(groups, local_namespace)
    if release_overrides is not None:
        for k in groups:
            groups[k]["release"] = release_overrides.get(k, 0.0)

    scale   = power_mwt / 3411.0
    xq_eab  = _nbt_chi_q(distance_eab_m,  release_height_m, stability_class, wind_speed_m_s)
    xq_lpz  = _nbt_chi_q(distance_lpz_m,  release_height_m, stability_class, wind_speed_m_s)
    xq_cr   = _nbt_chi_q(distance_cr_intake_m, release_height_m, stability_class, wind_speed_m_s)

    # Licensing acceptance intervals are location-specific:
    #   EAB: 0–2 hr by default
    #   LPZ: plume passage / 30 days by default
    #   CR:  user-specified end_time_hr by default
    _eab_time_hr = float(eab_integration_hr if eab_integration_hr is not None else end_time_hr)
    _lpz_time_hr = float(lpz_integration_hr if lpz_integration_hr is not None else end_time_hr)
    _cr_time_hr  = float(end_time_hr)

    tede_eab = tede_lpz = tede_cr = 0.0
    group_doses = {}

    def _cum_chi_for_time(inv_ci, rel_frac, gname, gdata, _time_hr):
        lam_d = _math.log(2) / (gdata["t_half_hr"] * 3600.0)
        if sprays_on and spray_start_hr < _time_hr:
            sp_frac = max(0.0, min(spray_stop_hr, _time_hr) - spray_start_hr) / max(_time_hr, 1e-12)
        else:
            sp_frac = 0.0
        lam_rem = sp_frac * _NBT_SPRAY_REMOVAL[gname] + _NBT_NAT_DEP[gname]
        return _nbt_solve_group(
            inv_ci, rel_frac, release_duration_hr, containment_volume_ft3,
            leak_rate_frac_per_day, lam_d, lam_rem, _time_hr, dt_hr,
            credit_decay=credit_decay, credit_deposition=credit_deposition)

    for gname, gdata in groups.items():
        inv_ci   = gdata["inv_ci"]
        rel_frac = gdata["release"]

        cum_chi_eab = _cum_chi_for_time(inv_ci, rel_frac, gname, gdata, _eab_time_hr)
        cum_chi_lpz = _cum_chi_for_time(inv_ci, rel_frac, gname, gdata, _lpz_time_hr)
        cum_chi_cr  = _cum_chi_for_time(inv_ci, rel_frac, gname, gdata, _cr_time_hr)

        d_eab = cum_chi_eab * xq_eab * (breathing_rate_m3_s * gdata["dcf_inh"] + gdata["dcf_cloud"])
        d_lpz = cum_chi_lpz * xq_lpz * (breathing_rate_m3_s * gdata["dcf_inh"] + gdata["dcf_cloud"])
        df_cr = _NBT_CR_DF[gname] if cr_filter_on else 1.0
        d_cr  = cum_chi_cr * xq_cr / df_cr * breathing_rate_m3_s * gdata["dcf_inh"]

        tede_eab += d_eab; tede_lpz += d_lpz; tede_cr += d_cr
        group_doses[gname] = {"tede_eab": d_eab, "tede_lpz": d_lpz,
                               "tede_cr": d_cr,
                               "cum_chi_eab": cum_chi_eab,
                               "cum_chi_lpz": cum_chi_lpz,
                               "cum_chi_cr":  cum_chi_cr,
                               "released_ci": inv_ci * rel_frac}

    # ── TEDE vs distance table ───────────────────────────────────────────────
    # Provide both EAB-integration and LPZ-integration doses at each distance.
    dist_fit = None
    dist_table = []
    # Representative distance table points.  Include the user-specified
    # EAB/LPZ distances in addition to standard reference points.  The old
    # hard-coded 914 m point is replaced by 1000 m.
    _DIST_POINTS = sorted(set([
        100, 250, 500, 1000, 1600, 3200, 4800, 8000, 16000,
        int(round(float(distance_eab_m))),
        int(round(float(distance_lpz_m))),
    ]))

    if xq_eab > 0:
        _fit_err = 0.0
        dist_fit = {"formula": ("Dose(d) = Dose(ref) * chiQ(d) / chiQ(ref); "
                                "separate columns use EAB and LPZ integration times"),
                    "fit_error_pct": _fit_err}
        for d in _DIST_POINTS:
            _xq_d = _nbt_chi_q(d, release_height_m, stability_class, wind_speed_m_s)
            _eab_d = tede_eab * (_xq_d / xq_eab) if xq_eab > 0 else 0.0
            _lpz_d = tede_lpz * (_xq_d / xq_lpz) if xq_lpz > 0 else 0.0
            dist_table.append({"distance_m": d,
                               "chi_q_s_m3": _xq_d,
                               "eab_integration_tede_rem": _eab_d,
                               "lpz_integration_tede_rem": _lpz_d})

    return {"tede_eab_rem": tede_eab, "tede_lpz_rem": tede_lpz,
            "tede_cr_rem":  tede_cr,  "group_doses":  group_doses,
            "runtime_ms":   (_time.perf_counter() - t0) * 1000.0,
            "chi_q_used":   {"eab": xq_eab, "lpz": xq_lpz, "cr": xq_cr},
            "integration_times_hr": {"eab": _eab_time_hr, "lpz": _lpz_time_hr, "cr": _cr_time_hr},
            "inputs":       {"eab_integration_hr": _eab_time_hr,
                             "lpz_integration_hr": _lpz_time_hr,
                             "cr_integration_hr":  _cr_time_hr,
                             "inventory": _inventory_meta},
            "dist_fit":     dist_fit,
            "dist_table":   dist_table}

# ── end NOTBADTRAD ────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Accept case name as optional command-line argument so the
    # program can be driven non-interactively from batch scripts:
    #   python flare_sim.py CaseSteadyState
    # Without an argument the interactive prompt is shown as before.
    _args = sys.argv[1:]
    _no_figures = "--no-figures" in _args
    _args = [a for a in _args if not a.startswith("-")]
    if _args:
        _arg = _args[0].strip()
        for _sfx in ("_in.xlsx", "_out.xlsx", "_in", "_out"):
            if _arg.endswith(_sfx):
                _arg = _arg[:-len(_sfx)]; break
        wkstbase = _arg
        print(f"Case: {wkstbase}")
    else:
        wkstbase = user_input("Enter name of input file: ")
    try:
        pwr_sim(wkstbase, minimum_output=1 if _no_figures else 0)
    except Exception as _e:
        print(f"\nUnexpected error running case '{wkstbase}':")
        print(f"  {type(_e).__name__}: {_e}")
        print("  Check that the filename and sheet name exactly match the case name.")
