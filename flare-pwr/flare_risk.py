"""
flare_risk.py  —  FLARE Risk Assessment  (F-C Tool)
Run from flare_home.py via ?page=risk routing.
No sidebar — all controls in the main panel.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import subprocess, sys, json, csv, os
import base64
import requests
from pathlib import Path

# WORK_DIR: injected by flare_home exec globals, or derived from __file__
try:
    WORK_DIR = Path(FLARE_WORK_DIR)          # injected by flare_home.py
except NameError:
    try:
        WORK_DIR = Path(__file__).parent
    except NameError:
        WORK_DIR = Path.cwd()

# On Windows, Streamlit may try to write to '.' (cwd) on each rerun.
# Ensure cwd is always the FLARE folder where the user has write access.
import os as _os
_os.chdir(str(WORK_DIR))

# Persist across Streamlit reruns
if "risk_work_dir" not in st.session_state:
    st.session_state.risk_work_dir = WORK_DIR
WORK_DIR = st.session_state.risk_work_dir


def _load_buddy_b64():
    """Return a base64 data URI for the FLARE Buddy icon, or None."""
    for _dir in (WORK_DIR / "icons", WORK_DIR / "Icons"):
        _p = _dir / "FLAREBUDDY.png"
        if _p.exists():
            try:
                return "data:image/png;base64," + base64.b64encode(_p.read_bytes()).decode()
            except Exception:
                return None
    return None

_BUDDY_B64 = _load_buddy_b64()

def _buddy_expander_label(title: str) -> str:
    """Return an expander label with an inline FLARE Buddy thumbnail, or fall back to 🤖."""
    if _BUDDY_B64:
        return (
            f"<img src='{_BUDDY_B64}' style='height:1.3rem;width:1.3rem;"
            f"border-radius:50%;object-fit:cover;vertical-align:middle;"
            f"margin-right:0.4rem;'/>{title}"
        )
    return f"🤖  {title}"

def _runtime_dir(base_dir: Path | None = None) -> Path:
    """Return the FLARE runtime folder, preferring runtime/ then Runtime/."""
    base = Path(base_dir) if base_dir is not None else WORK_DIR
    lower = base / "runtime"
    upper = base / "Runtime"
    if lower.exists():
        rt = lower
    elif upper.exists():
        rt = upper
    else:
        rt = lower
        rt.mkdir(parents=True, exist_ok=True)
    return rt


def _runtime_file(name: str, base_dir: Path | None = None) -> Path:
    return _runtime_dir(base_dir) / name

# ── AI narrative helpers ─────────────────────────────────────────────────────
def _read_config(key):
    """Read KEY=value from runtime/flare_config.txt (or Runtime/flare_config.txt)."""
    cfg = _runtime_file("flare_config.txt")
    if cfg.exists():
        for line in cfg.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k.strip().upper() == key.upper():
                return v.strip().strip('"').strip("'")
    return None

def _load_api_key():
    return (st.session_state.get("user_api_key")
            or _read_config("ANTHROPIC_API_KEY")
            or os.environ.get("ANTHROPIC_API_KEY"))

def _load_model():
    return (_read_config("ANTHROPIC_MODEL")
            or os.environ.get("ANTHROPIC_MODEL")
            or "claude-sonnet-4-5")

def _anthropic_text(system_prompt, user_prompt, max_tokens=4000, timeout_s=90):
    """Call Claude and return (text, stop_reason).

    Raises RuntimeError on API / HTTP errors. The stop_reason lets callers
    detect truncation ('max_tokens') and warn the user.
    """
    api_key = _load_api_key()
    if not api_key:
        raise RuntimeError("No API key found. Add ANTHROPIC_API_KEY to runtime/flare_config.txt (or Runtime/flare_config.txt) or save it on the FLARE Home page.")
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": _load_model(),
            "max_tokens": max_tokens,
            "temperature": 0.2,
            "system": system_prompt,
            "messages": [{"role": "user", "content": user_prompt}],
        },
        timeout=timeout_s,
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"Anthropic API error {resp.status_code}: {resp.text[:600]}")
    data = resp.json()
    text = "\n".join(
        block.get("text", "") for block in data.get("content", [])
        if block.get("type") == "text"
    ).strip()
    stop_reason = data.get("stop_reason", "")
    return text, stop_reason



# ── AI narrative detail / section-generation helpers ─────────────────────────
def _risk_detail_level(detail):
    """Snap a narrative detail slider value to one of 21 detents."""
    try:
        return max(0.0, min(1.0, round(float(detail) / 0.05) * 0.05))
    except Exception:
        return 0.55


def _risk_detail_word_target(detail):
    """Return the intended whole-report word target for the 21 detail levels."""
    detail = _risk_detail_level(detail)
    targets = {
        0.00: 300, 0.05: 400, 0.10: 500, 0.15: 650, 0.20: 800,
        0.25: 1000, 0.30: 1200, 0.35: 1450, 0.40: 1700, 0.45: 2000,
        0.50: 2300, 0.55: 2600, 0.60: 2900, 0.65: 3200, 0.70: 3500,
        0.75: 3800, 0.80: 4100, 0.85: 4400, 0.90: 4700, 0.95: 5000,
        1.00: 5300,
    }
    return targets.get(round(detail, 2), 2600)


def _risk_detail_word_range(detail):
    target = _risk_detail_word_target(detail)
    lo = max(250, int(round(target * 0.85)))
    hi = int(round(target * 1.15))
    return lo, hi


def _risk_detail_descriptor(detail):
    detail = _risk_detail_level(detail)
    if detail <= 0.10:
        return "abstract-level"
    if detail <= 0.30:
        return "executive-summary"
    if detail <= 0.55:
        return "moderate technical"
    if detail <= 0.80:
        return "detailed technical"
    return "full technical report"


def _risk_narrative_max_tokens(word_max):
    """Generous but bounded token budget for a section or full narrative."""
    try:
        return int(min(12000, max(900, round(float(word_max) * 2.8 + 300))))
    except Exception:
        return 2500


def _risk_narrative_timeout_s(word_target):
    try:
        return int(min(420, max(90, 60 + float(word_target) * 0.06)))
    except Exception:
        return 180


def _risk_word_count(text):
    import re as _re
    return len(_re.findall(r"\b[\w./%+\-×⁻]+\b", str(text or "")))


def _risk_clamp_words(text, max_words):
    """Softly clamp excessive section text at a paragraph boundary."""
    import re as _re
    s = str(text or "").strip()
    if not s or _risk_word_count(s) <= max_words:
        return s
    words = _re.findall(r"\S+", s)
    clipped = " ".join(words[:max_words]).rstrip()
    # Prefer ending at the last sentence terminator in the final 25% of the clipped text.
    tail_start = max(0, int(len(clipped) * 0.75))
    last = max(clipped.rfind(". ", tail_start), clipped.rfind("; ", tail_start))
    if last > 0:
        clipped = clipped[:last + 1]
    return clipped.rstrip() + "\n\n*[Narrative shortened to match the selected detail level.]*"


def _risk_sanitize_markdown(text, fallback_title="Risk Narrative Section"):
    """Normalize AI markdown so only actual section headings render as headings."""
    import re as _re
    s = str(text or "").replace("\r\n", "\n").strip()
    if not s:
        return f"### {fallback_title}\n\n"
    # Convert '# Title Body starts here' into a title line plus body paragraph.
    lines = []
    for line in s.split("\n"):
        raw = line.rstrip()
        stripped = raw.strip()
        m = _re.match(r"^(#{1,6})\s+(.+)$", stripped)
        if m:
            content = m.group(2).strip()
            # Split at the first sentence-like boundary after a compact title.
            # Handles model output such as '### Summary The case results...'.
            title_body = _re.match(r"^([A-Z][A-Za-z0-9/\-–—&,() ]{3,80}?)(\s+(?:A|An|The|This|These|In|For|Overall|Across)\b.+)$", content)
            if title_body:
                title = title_body.group(1).strip(" :-–—")
                body = title_body.group(2).strip()
                lines.append(f"### {title}")
                lines.append("")
                lines.append(body)
            else:
                lines.append(f"### {content.lstrip('# ').strip()}")
            continue
        # Demote accidental heading-like bold paragraphs.
        stripped = _re.sub(r"^\*\*(.{40,})\*\*$", r"\1", stripped)
        lines.append(stripped)
    out = "\n".join(lines)
    out = _re.sub(r"\n{3,}", "\n\n", out).strip()
    return out


def _risk_section_plan(detail):
    """Build a deterministic report plan from the 21-level detail setting."""
    detail = _risk_detail_level(detail)
    target = _risk_detail_word_target(detail)
    if target <= 500:
        titles = [
            "Risk Summary and Boundary Margin",
            "Regulatory Interpretation and Conclusions",
        ]
    elif target <= 1000:
        titles = [
            "Risk Summary and Boundary Margin",
            "Dominant Contributors and Event Categories",
            "Regulatory Interpretation and Conclusions",
        ]
    elif target <= 2000:
        titles = [
            "Risk Summary and Boundary Margin",
            "Event Category Review",
            "Dose Contributors and Margins",
            "Regulatory Interpretation and Conclusions",
        ]
    elif target <= 3500:
        titles = [
            "Risk Summary and Boundary Margin",
            "Event Category Review",
            "Dominant Dose Contributors",
            "Cases Near or Beyond Limits",
            "Uncertainty, Frequency Sensitivity, and Conclusions",
        ]
    else:
        titles = [
            "Risk Summary and Boundary Margin",
            "Frequency-Consequence Framework",
            "Event Category Review",
            "Dominant Dose Contributors",
            "Cases Near or Beyond Limits",
            "Uncertainty, Frequency Sensitivity, and Conclusions",
        ]
    base = target // len(titles)
    remainder = target - base * len(titles)
    plan = []
    for i, title in enumerate(titles):
        words = int(base + (remainder if i == len(titles) - 1 else 0))
        plan.append({"title": title, "target_words": words, "max_words": int(round(words * 1.18))})
    return plan


def _risk_build_common_ai_context(sum_df, detail):
    """Prepare compact risk context shared by all section-generation calls."""
    _table_csv = sum_df.to_csv(index=False)
    _n_fail = int((sum_df["Status"].astype(str).str.upper() == "FAIL").sum()) if "Status" in sum_df else 0
    _n_cases = len(sum_df)
    _max_row = None
    try:
        _tmp = sum_df.copy()
        _tmp["_dose_num"] = pd.to_numeric(_tmp["EAB TEDE Total (rem)"].replace("< 1e-6", "0"), errors="coerce")
        _max_row = _tmp.sort_values("_dose_num", ascending=False).head(1).to_dict("records")
    except Exception:
        _max_row = None
    _detail = _risk_detail_level(detail)
    _word_target = _risk_detail_word_target(_detail)
    _lo, _hi = _risk_detail_word_range(_detail)
    return {
        "detail": _detail,
        "detail_label": _risk_detail_descriptor(_detail),
        "word_target": _word_target,
        "word_low": _lo,
        "word_high": _hi,
        "n_cases": _n_cases,
        "n_fail": _n_fail,
        "max_row": _max_row,
        "table_csv": _table_csv,
    }

# ── Constants (must be defined before sidebar) ─────────────────────────────────
_DEFAULT_FREQ  = 1e-3
NEI_FREQ_AOO   = 1e-2
NEI_FREQ_DBE   = 1e-4
NEI_FREQ_SCREEN= 5e-7
NEI_DOSE_AOO   = 1.0
NEI_DOSE_DBE   = 25.0
NEI_DOSE_BDBE  = 1000.0

def classify(freq):
    if freq >= NEI_FREQ_AOO:    return "AOO"
    if freq >= NEI_FREQ_DBE:    return "DBE"
    if freq >= NEI_FREQ_SCREEN: return "BDBE"
    return "Screened"

def dose_limit(freq):
    if freq >= NEI_FREQ_AOO:    return NEI_DOSE_AOO
    if freq >= NEI_FREQ_DBE:    return NEI_DOSE_DBE
    return NEI_DOSE_BDBE

# set_page_config already called by flare_home — skip it entirely

# ── Styling ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;600&display=swap');

:root {
    --bg:       #f5f7fa;
    --surface:  #ffffff;
    --sidebar:  #1a1f2e;
    --border:   #d0d7de;
    --accent:   #0969da;
    --accent2:  #1a7f37;
    --warn:     #9a6700;
    --danger:   #cf222e;
    --text:     #1f2328;
    --muted:    #57606a;
    --sid-text: #e6edf3;
    --sid-muted:#8b949e;
    --dark-surface: #0d1117;
    --dark-border:  #30363d;
}

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
    color: var(--text);
}
.stApp { background: var(--bg); }
header { background: transparent !important; }

section[data-testid="stSidebar"] {
    background: var(--sidebar) !important;
    border-right: 1px solid var(--dark-border);
    color: var(--sid-text);
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label {
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] .stCaption {
    color: var(--sid-muted) !important;
}
/* Keep expander header dark when expanded so label stays visible */
section[data-testid="stSidebar"] details,
section[data-testid="stSidebar"] details[open],
section[data-testid="stSidebar"] [data-testid="stExpander"],
section[data-testid="stSidebar"] [data-testid="stExpander"] > details {
    background: var(--dark-surface) !important;
    border: 1px solid var(--dark-border) !important;
    border-radius: 4px;
}
section[data-testid="stSidebar"] details summary,
section[data-testid="stSidebar"] details[open] summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    background: var(--dark-surface) !important;
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] details summary p,
section[data-testid="stSidebar"] details summary span,
section[data-testid="stSidebar"] details[open] summary p,
section[data-testid="stSidebar"] details[open] summary span {
    color: var(--sid-text) !important;
}
section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] {
    background: var(--dark-surface);
    border: 1px solid var(--dark-border);
    border-radius: 6px;
}
section[data-testid="stSidebar"] .stSelectbox span { color: #ffffff !important; }
div[role="listbox"] { background: var(--dark-surface) !important; }
div[role="option"] { color: #ffffff !important; }
div[role="option"]:hover { background: #21262d !important; }
section[data-testid="stSidebar"] input {
    background: var(--dark-surface) !important;
    color: #ffffff !important;
    border: 1px solid var(--dark-border) !important;
}


/* Sidebar secondary/default buttons: keep Load Previous Run buttons readable on dark theme. */
section[data-testid="stSidebar"] div.stButton > button,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button,
section[data-testid="stSidebar"] button[kind="secondary"] {
    background: #263142 !important;
    color: #ffffff !important;
    border: 1px solid #f97316 !important;
    border-radius: 6px !important;
    opacity: 1 !important;
}
section[data-testid="stSidebar"] div.stButton > button:hover,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button:hover,
section[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background: #f97316 !important;
    color: #ffffff !important;
    border-color: #fb923c !important;
}
section[data-testid="stSidebar"] div.stButton > button:disabled,
section[data-testid="stSidebar"] div[data-testid="stButton"] > button:disabled,
section[data-testid="stSidebar"] button[kind="secondary"]:disabled,
section[data-testid="stSidebar"] button[disabled] {
    background: #374151 !important;
    color: #d1d5db !important;
    border: 1px solid #6b7280 !important;
    opacity: 1 !important;
}

section[data-testid="stSidebar"] button[kind="primary"] {
    background: var(--danger) !important;
    color: white !important;
}
section[data-testid="stSidebar"] button[kind="primary"]:hover {
    background: #a40e26 !important;
}
section[data-testid="stSidebar"] code {
    color: #ffffff !important;
    background: var(--dark-border) !important;
    padding: 2px 6px; border-radius: 4px;
}
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace; }
button[role="tab"] { color: var(--muted); }
button[role="tab"][aria-selected="true"] {
    color: var(--accent);
    border-bottom: 2px solid var(--accent);
}
.console {
    background: #0d1117;
    border: 1px solid #30363d;
    border-radius: 6px;
    padding: 1rem;
    color: #7ee787;
    font-family: 'IBM Plex Mono', monospace;
    white-space: pre-wrap;
    font-size: 0.85rem;
}
.hdiv { border-top: 1px solid var(--border); margin: 1.5rem 0; }
.metric-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:0.75rem; }
.metric-tile {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1rem;
    text-align: center;
}
.metric-tile .val {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.4rem;
    color: var(--accent);
}
.metric-tile.ok  .val { color: var(--accent2); }
.metric-tile.warn .val { color: var(--warn); }
.metric-tile.danger .val { color: var(--danger); }
/* ── Tooltip (?) icon — targets the actual SVG stroke ── */
section[data-testid="stSidebar"] .stTooltipHoverTarget svg.icon {
    stroke: #f97316 !important;
}
section[data-testid="stSidebar"] .stTooltipHoverTarget:hover svg.icon {
    stroke: #ffffff !important;
}
.stTooltipHoverTarget svg.icon {
    stroke: #0969da !important;
}
/* Risk-specific classes */
.risk-title{ font-family:'IBM Plex Mono',monospace; font-size:1.5rem;
             font-weight:600; color:#1f2328; margin-bottom:0.2rem; }
.risk-sub  { font-size:0.85rem; color:#8b949e; margin-bottom:1.2rem; }
.home-btn  { margin-bottom:1rem; }
.cat-aoo   { color:#1a7f37; font-weight:600; font-size:0.8rem; }
.cat-dbe   { color:#0969da; font-weight:600; font-size:0.8rem; }
.cat-bdbe  { color:#9a6700; font-weight:600; font-size:0.8rem; }
.cat-screen{ color:#888;    font-weight:600; font-size:0.8rem; }
.fc-note   { font-size:0.78rem; color:#8b949e; margin-top:0.3rem; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────

# ── NEI 18-04 definitions ──────────────────────────────────────────────────────
#   AOO  : f >= 1e-2 /yr
#   DBE  : 1e-4 <= f < 1e-2 /yr
#   BDBE : 5e-7 <= f < 1e-4 /yr
#   Screened out : f < 5e-7 /yr
#
# F-C Target (EAB TEDE, mean):
#   AOO  dose limit  : 0.05 rem
#   DBE  dose limit  : 2.5  rem
#   BDBE dose limit  : 25.0 rem

def cat_html(cat):
    cls = {"AOO":"cat-aoo","DBE":"cat-dbe","BDBE":"cat-bdbe","Screened":"cat-screen"}[cat]
    return f'<span class="{cls}">{cat}</span>'

# ── Default frequencies keyed on case name stem ────────────────────────────────
_DEFAULTS = {
    "CaseBlowdown":    1e-3,
    "CaseLBLOCA_cold": 1e-4,
    "CaseLBLOCA":      1e-4,
    "CaseFullLOCA":    1e-5,
    "CaseECCS":        1e-3,
    "CaseCoreDNBR":    1e-3,
    "CaseCore":        1e-3,
    "CaseDNBR":        1e-2,
    "CaseCoast":       1e-1,
    "CaseRIA":         1e-4,
    "CaseRodEject":    1e-4,
    "CaseStartup":     1e-1,
    "CaseSS":          1.0,
    "CaseMSLB":        1e-3,
    "CaseLOHS":        1e-2,
    "CaseFeedwater":   1e-1,
    "CaseATWS":        1e-3,
    "CaseSGTR":        1e-3,
    "CasePump":        1e-2,
    "CaseFLAREManager":1e-3,
    "CaseNewLOCA":     1e-4,
}

# ── Risk-run-local persistence files ─────────────────────────────────────────
# Revision note: risk results and PRA frequency tables are intentionally written
# only inside risk_<timestamp> run folders.  The FLARE root folder is kept clean.
# Older root-level files are not created by this UI; if present from a prior
# revision, they are ignored except by manual user action.

def _risk_dirs_newest_first():
    try:
        return sorted(
            [d for d in WORK_DIR.iterdir() if d.is_dir() and d.name.startswith("risk_")],
            key=lambda d: d.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        return []

def load_pra_table():
    """Load frequencies from the most recent risk run folder, if available."""
    for _d in _risk_dirs_newest_first():
        _p = _d / "flare_pra_table.csv"
        if not _p.exists():
            continue
        try:
            df = pd.read_csv(_p)
            return dict(zip(df["CaseName"].astype(str).str.strip(),
                            df["Frequency"].astype(float)))
        except Exception:
            continue
    return {}

def load_results_from_disk():
    """Restore results from the most recent risk run folder, if available."""
    for _d in _risk_dirs_newest_first():
        _p = _d / "flare_risk_results.json"
        if not _p.exists():
            continue
        try:
            data = json.loads(_p.read_text(encoding="utf-8"))
            if any(k in cases for k in data):
                st.session_state.risk_run_dir = str(_d)
                return data
        except Exception:
            continue
    return {}

# ── Discover input cases ───────────────────────────────────────────────────────
# Risk cases are discovered recursively below the FLARE root. Root-level input
# decks are intentionally ignored; generated output/control folders are excluded.
_EXCLUDED_INPUT_DIR_PREFIXES = ("sim_", "risk_", "ua_", ".sim_all_", "__pycache__")

def _is_generated_input_dir(path: Path) -> bool:
    try:
        rel_parts = path.relative_to(WORK_DIR).parts
    except Exception:
        rel_parts = path.parts
    return any(part.startswith(_EXCLUDED_INPUT_DIR_PREFIXES) or part.startswith(".")
               for part in rel_parts)

def discover_input_cases():
    entries = []
    for p in WORK_DIR.rglob("Case*_in.xlsx"):
        if p.parent == WORK_DIR:
            continue
        if p.name.startswith(".") or p.stem.startswith(".~"):
            continue
        if _is_generated_input_dir(p.parent):
            continue
        case = p.stem[:-3] if p.stem.endswith("_in") else p.stem.replace("_in", "")
        entries.append({"case": case, "input_path": str(p), "rel": p.parent.relative_to(WORK_DIR).as_posix()})
    # Preserve one entry per case name; warn later if users create duplicates.
    seen = {}
    for e in sorted(entries, key=lambda x: (x["case"].lower(), x["rel"].lower())):
        seen.setdefault(e["case"], e)
    return list(seen.values())

case_entries = discover_input_cases()
cases = [e["case"] for e in case_entries]

def _active_risk_run_dir():
    _rd = st.session_state.get("risk_run_dir")
    return Path(_rd) if _rd else None

def save_results(results, run_dir=None):
    """Persist results only inside a risk_<timestamp> run folder."""
    try:
        _rd = Path(run_dir) if run_dir is not None else _active_risk_run_dir()
        if _rd is None:
            return
        _rd.mkdir(parents=True, exist_ok=True)
        (_rd / "flare_risk_results.json").write_text(
            json.dumps(results, indent=2), encoding="utf-8")
    except Exception:
        pass

def save_pra_table(freq_dict, run_dir=None):
    """Write frequencies only inside a risk_<timestamp> run folder."""
    try:
        _rd = Path(run_dir) if run_dir is not None else _active_risk_run_dir()
        if _rd is None:
            return False
        _rd.mkdir(parents=True, exist_ok=True)
        rows = [{"CaseName": k, "Frequency": v} for k, v in sorted(freq_dict.items())]
        pd.DataFrame(rows).to_csv(_rd / "flare_pra_table.csv", index=False)
        return True
    except Exception:
        return False


def _extract_eab_from_sheet(_wb, _sheet_name):
    """Extract EAB TEDE from a Dose-style worksheet."""
    if _sheet_name not in _wb.sheetnames:
        return None
    _ws = _wb[_sheet_name]
    for _row in _ws.iter_rows(values_only=True):
        if _row and str(_row[0]).strip() == "EAB":
            _v = _row[1] if len(_row) > 1 else None
            if _v is None:
                return None
            try:
                return float(_v)
            except (TypeError, ValueError):
                return None
    return None


def _extract_risk_eab_dose_components(_xlsx):
    """Return accident, iodine-spike, and total EAB TEDE for risk plotting.

    The F-C plot consequence should be the sum of the accident source-term dose
    from the `Dose` sheet and the pre-existing-coolant iodine-spike dose from
    the `Iodine Spike` sheet.
    """
    import openpyxl as _opxl
    _wb = _opxl.load_workbook(_xlsx, read_only=True, data_only=True)
    try:
        _accident = _extract_eab_from_sheet(_wb, "Dose")
        _iodine   = _extract_eab_from_sheet(_wb, "Iodine Spike")
    finally:
        _wb.close()
    if _accident is None and _iodine is None:
        return None, None, None
    _total = (_accident or 0.0) + (_iodine or 0.0)
    return _accident, _iodine, _total


# ── Durable risk-run worker support ───────────────────────────────────────────
def _risk_status_path(run_dir):
    return Path(run_dir) / "risk_status.json"

def _read_json(path, default=None):
    try:
        p = Path(path)
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {} if default is None else default


def _tail_text_file(path, max_chars=6000):
    """Return the tail of a text file for live console display."""
    try:
        p = Path(path)
        if not p.exists():
            return ""
        # Read as bytes from the end so huge console logs do not slow the UI.
        with open(p, "rb") as f:
            try:
                f.seek(0, 2)
                size = f.tell()
                f.seek(max(size - max_chars, 0))
            except Exception:
                pass
            data = f.read()
        return data.decode("utf-8", errors="replace")
    except Exception as e:
        return f"(Could not read log: {e})"


def _safe_float(v, default=0.0):
    """Return v as float, accepting display strings such as '< 1e-6'."""
    try:
        if v is None:
            return default
        if isinstance(v, str):
            s = v.strip().replace(',', '')
            if s.startswith('<'):
                return default
            return float(s)
        return float(v)
    except Exception:
        return default


def _risk_pdf_path():
    """Choose a persistent PDF path for the current risk results."""
    try:
        csv_path = Path(st.session_state.get("risk_csv_path", "") or "")
        if csv_path and csv_path != Path('.') and csv_path.exists():
            return csv_path.parent / "FLARE_risk_report.pdf"
    except Exception:
        pass
    try:
        run_dir = Path(st.session_state.get("risk_run_dir", "") or "")
        if run_dir and run_dir.exists():
            return run_dir / "FLARE_risk_report.pdf"
    except Exception:
        pass
    return WORK_DIR / "FLARE_risk_report.pdf"


def _export_risk_pdf(results, sum_df, out_pdf):
    """Export the F-C curve and results table to a PDF file.

    This intentionally avoids Plotly static image dependencies.  The F-C chart
    is drawn directly with ReportLab using the same log-axis ranges and boundary
    used in the Streamlit chart.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.lib.utils import ImageReader
    import math
    import tempfile

    out_pdf = Path(out_pdf)
    out_pdf.parent.mkdir(parents=True, exist_ok=True)

    fc_dose = [1.0e-2, 1.0e-1, 1.0e0, 1.0e0, 2.5e1, 1.0e3, 1.0e4]
    fc_freq = [1.0e1,  1.0e0, 1.0e-1, 1.0e-2, 1.0e-4, 5.0e-7, 5.0e-7]
    cat_color = {"AOO": colors.HexColor("#1a7f37"),
                 "DBE": colors.HexColor("#0969da"),
                 "BDBE": colors.HexColor("#9a6700"),
                 "Screened": colors.HexColor("#666666")}

    page_w, page_h = landscape(letter)
    margin = 0.45 * inch
    chart_w = page_w - 2 * margin
    chart_h = page_h - 1.75 * inch
    chart_x = margin
    chart_y = 0.8 * inch
    x_min, x_max = -7.0, 4.0
    y_min, y_max = -8.0, 1.0

    # PDF accessibility / government preflight readability guard.
    # Use >= 14 pt for axis tick numbers and >= 16 pt for axis labels,
    # which is comfortably above a nominal 12 pt minimum plus 15%.
    PDF_AXIS_TICK_FONT_PT  = 14
    PDF_AXIS_LABEL_FONT_PT = 16
    PDF_POINT_LABEL_FONT_PT = 11
    PDF_CHART_TITLE_FONT_PT = 18
    PDF_CHART_NOTE_FONT_PT  = 10

    def xmap(x):
        x = max(float(x), 1e-7)
        return chart_x + (math.log10(x) - x_min) / (x_max - x_min) * chart_w

    def ymap(y):
        y = max(float(y), 1e-8)
        return chart_y + (math.log10(y) - y_min) / (y_max - y_min) * chart_h

    tmp_png = Path(tempfile.gettempdir()) / f"flare_risk_fc_{os.getpid()}.png"
    c = _canvas.Canvas(str(tmp_png.with_suffix('.pdf')), pagesize=landscape(letter))
    c.setTitle("FLARE Risk F-C Chart")
    c.setFont("Helvetica-Bold", PDF_CHART_TITLE_FONT_PT)
    c.drawString(margin, page_h - 0.45 * inch, "FLARE Frequency-Consequence Chart")
    c.setFont("Helvetica", PDF_CHART_NOTE_FONT_PT)
    c.drawString(margin, page_h - 0.65 * inch, "Dose plotted is total EAB TEDE = accident dose + iodine-spike dose, where available.")

    # Plot area.
    c.setStrokeColor(colors.black)
    c.setLineWidth(1)
    c.rect(chart_x, chart_y, chart_w, chart_h)

    # Log grid and labels.
    c.setFont("Helvetica", PDF_AXIS_TICK_FONT_PT)
    c.setStrokeColor(colors.HexColor("#dddddd"))
    for lx in range(int(x_min), int(x_max) + 1):
        xx = chart_x + (lx - x_min) / (x_max - x_min) * chart_w
        c.line(xx, chart_y, xx, chart_y + chart_h)
        c.setFillColor(colors.black)
        c.drawCentredString(xx, chart_y - 18, f"1E{lx}")
    for ly in range(int(y_min), int(y_max) + 1):
        yy = chart_y + (ly - y_min) / (y_max - y_min) * chart_h
        c.setStrokeColor(colors.HexColor("#dddddd"))
        c.line(chart_x, yy, chart_x + chart_w, yy)
        c.setFillColor(colors.black)
        c.drawRightString(chart_x - 8, yy - 5, f"1E{ly}")

    # Boundary.
    c.setStrokeColor(colors.red)
    c.setLineWidth(2)
    for i in range(len(fc_dose) - 1):
        c.line(xmap(fc_dose[i]), ymap(fc_freq[i]), xmap(fc_dose[i+1]), ymap(fc_freq[i+1]))
    c.setFont("Helvetica", 11)
    c.setFillColor(colors.red)
    c.drawString(xmap(2.5e1) + 5, ymap(1e-4) + 6, "Design Objective")

    # Dividers.
    c.setStrokeColor(colors.HexColor("#999999"))
    c.setDash(2, 3)
    for f in [NEI_FREQ_AOO, NEI_FREQ_DBE, NEI_FREQ_SCREEN]:
        yy = ymap(f); c.line(chart_x, yy, chart_x + chart_w, yy)
    for d in [NEI_DOSE_AOO, NEI_DOSE_DBE, NEI_DOSE_BDBE]:
        xx = xmap(d); c.line(xx, chart_y, xx, chart_y + chart_h)
    c.setDash()

    # Points.
    c.setFont("Helvetica", PDF_POINT_LABEL_FONT_PT)
    for cname, r in sorted(results.items()):
        dose = r.get("dose") if r.get("dose") is not None and r.get("dose") > 0 else 1e-6
        freq = r.get("freq", 1e-8)
        cat = r.get("category", "Screened")
        col = cat_color.get(cat, colors.black)
        xx, yy = xmap(dose), ymap(freq)
        c.setFillColor(col)
        c.setStrokeColor(colors.red if str(r.get("status", "")).upper() == "FAIL" else col)
        c.circle(xx, yy, 3.5, fill=1, stroke=1)
        c.setFillColor(col)
        label = cname.replace("Case", "")[:16]
        c.drawString(xx + 4, yy + 4, label)

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", PDF_AXIS_LABEL_FONT_PT)
    c.drawCentredString(chart_x + chart_w / 2, chart_y - 42, "EAB TEDE Total (rem)")
    c.saveState()
    c.translate(chart_x - 60, chart_y + chart_h / 2)
    c.rotate(90)
    c.drawCentredString(0, 0, "Event Frequency (events/yr)")
    c.restoreState()
    c.showPage()
    c.save()

    # Build table document and append the chart PDF as first page via platypus
    # by drawing the chart again in onFirstPage.  This avoids external image/kaleido.
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out_pdf), pagesize=landscape(letter),
                            rightMargin=0.35*inch, leftMargin=0.35*inch,
                            topMargin=0.35*inch, bottomMargin=0.35*inch)

    story = [Paragraph("FLARE Risk Results Table", styles["Title"]),
             Paragraph("Dose plotted in the F-C chart is total EAB TEDE = accident dose + iodine-spike dose, where available.", styles["BodyText"]),
             Spacer(1, 0.15*inch)]

    table_df = sum_df.copy().astype(str)
    cols = list(table_df.columns)
    data = [cols] + table_df.values.tolist()
    # Approximate widths for landscape letter.
    avail_w = page_w - 0.7*inch
    weights = []
    for col in cols:
        if col == "Case": weights.append(1.25)
        elif col == "Error": weights.append(1.25)
        elif "Iodine" in col: weights.append(1.25)
        elif "Accident" in col: weights.append(1.25)
        elif "Total" in col: weights.append(1.15)
        else: weights.append(0.9)
    total_w = sum(weights)
    col_widths = [avail_w * w / total_w for w in weights]
    tbl = Table(data, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
    ]))
    story.append(tbl)

    def _draw_chart_page(canvas, doc_obj):
        # Redraw the chart on the first page, then allow the table story to start on the next page.
        pass

    # Simpler: build a two-page PDF directly with canvas for chart + platypus table is awkward.
    # Instead, create a full canvas PDF manually: chart page above, then table pages via platypus append is not supported.
    # ReportLab can build table-only PDF; for reliability, put chart and table in one platypus PDF by using a custom Flowable.
    from reportlab.platypus import Flowable

    class FCChartFlowable(Flowable):
        def __init__(self):
            Flowable.__init__(self)
            self.width = chart_w
            # Keep the chart flowable inside the landscape-letter frame even
            # after the preflight-readable font increase.  The previous height
            # was 561.6 pt, while the available frame can be about 549.6 pt on
            # some ReportLab page templates.
            self.height = min(chart_h + 0.85*inch, 7.45*inch)
        def draw(self):
            cc = self.canv
            # local origin at current frame lower-left.  The larger left and
            # bottom margins prevent 14-16 pt axis text from clipping.
            x0 = 0.95*inch
            y0 = 0.70*inch
            w = self.width - 1.15*inch
            h = self.height - 1.65*inch
            def xm(x):
                x = max(float(x), 1e-7)
                return x0 + (math.log10(x) - x_min)/(x_max-x_min)*w
            def ym(y):
                y = max(float(y), 1e-8)
                return y0 + (math.log10(y) - y_min)/(y_max-y_min)*h
            cc.setFont("Helvetica-Bold", PDF_CHART_TITLE_FONT_PT)
            cc.drawString(0.1*inch, self.height - 0.32*inch, "FLARE Frequency-Consequence Chart")
            cc.setFont("Helvetica", PDF_CHART_NOTE_FONT_PT)
            cc.drawString(0.1*inch, self.height - 0.55*inch, "Dose plotted is total EAB TEDE = accident dose + iodine-spike dose, where available.")
            cc.setStrokeColor(colors.black); cc.setLineWidth(1); cc.rect(x0, y0, w, h)
            cc.setFont("Helvetica", PDF_AXIS_TICK_FONT_PT)
            for lx in range(int(x_min), int(x_max)+1):
                xx = x0 + (lx-x_min)/(x_max-x_min)*w
                cc.setStrokeColor(colors.HexColor("#dddddd")); cc.line(xx,y0,xx,y0+h)
                cc.setFillColor(colors.black); cc.drawCentredString(xx, y0-20, f"1E{lx}")
            for ly in range(int(y_min), int(y_max)+1):
                yy = y0 + (ly-y_min)/(y_max-y_min)*h
                cc.setStrokeColor(colors.HexColor("#dddddd")); cc.line(x0,yy,x0+w,yy)
                cc.setFillColor(colors.black); cc.drawRightString(x0-10, yy-5, f"1E{ly}")
            cc.setStrokeColor(colors.red); cc.setLineWidth(2)
            for i in range(len(fc_dose)-1): cc.line(xm(fc_dose[i]), ym(fc_freq[i]), xm(fc_dose[i+1]), ym(fc_freq[i+1]))
            cc.setStrokeColor(colors.HexColor("#999999")); cc.setDash(2,3)
            for f in [NEI_FREQ_AOO, NEI_FREQ_DBE, NEI_FREQ_SCREEN]: cc.line(x0,ym(f),x0+w,ym(f))
            for d in [NEI_DOSE_AOO, NEI_DOSE_DBE, NEI_DOSE_BDBE]: cc.line(xm(d),y0,xm(d),y0+h)
            cc.setDash()
            cc.setFont("Helvetica", PDF_POINT_LABEL_FONT_PT)
            for cname, r in sorted(results.items()):
                dose = r.get("dose") if r.get("dose") is not None and r.get("dose") > 0 else 1e-6
                freq = r.get("freq", 1e-8)
                cat = r.get("category", "Screened")
                col = cat_color.get(cat, colors.black)
                xx, yy = xm(dose), ym(freq)
                cc.setFillColor(col); cc.setStrokeColor(colors.red if str(r.get("status","")).upper()=="FAIL" else col)
                cc.circle(xx, yy, 3.5, fill=1, stroke=1)
                cc.setFillColor(col); cc.drawString(xx+4, yy+4, cname.replace("Case", "")[:16])
            cc.setFillColor(colors.black); cc.setFont("Helvetica-Bold", PDF_AXIS_LABEL_FONT_PT)
            cc.drawCentredString(x0+w/2, y0-46, "EAB TEDE Total (rem)")
            cc.saveState(); cc.translate(x0-62, y0+h/2); cc.rotate(90); cc.drawCentredString(0,0,"Event Frequency (events/yr)"); cc.restoreState()

    story = [FCChartFlowable(), PageBreak()] + story
    doc.build(story)
    return out_pdf

def _write_json_atomic(path, payload):
    try:
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(p.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        tmp.replace(p)
    except Exception:
        pass

def _latest_risk_status_dir():
    try:
        dirs = sorted(
            [d for d in WORK_DIR.iterdir() if d.is_dir() and d.name.startswith("risk_")],
            reverse=True,
        )
        for d in dirs:
            if (d / "risk_status.json").exists():
                return d
    except Exception:
        pass
    return None

def _load_worker_results_if_available(run_dir=None):
    """Load partial/final worker-backed risk results into session state."""
    if run_dir is None:
        run_dir = st.session_state.get("risk_run_dir")
    if not run_dir:
        return {}
    run_dir = Path(run_dir)
    status = _read_json(run_dir / "risk_status.json", {})
    # Prefer run-local results only.  Risk outputs are not written to FLARE root.
    results = _read_json(run_dir / "flare_risk_results.json", {})
    if results:
        st.session_state.risk_results = results
        save_results(results, run_dir)
    csv_path = run_dir / "FLARE_risk_results.csv"
    if csv_path.exists():
        st.session_state.risk_csv_path = str(csv_path)
    return status

def _launch_risk_worker(run_dir, fast_mode, source_term_override=None):
    """Launch the durable risk worker as an independent subprocess."""
    worker = WORK_DIR / "flare_risk_worker.py"
    if not worker.exists():
        st.error(
            "`flare_risk_worker.py` was not found in the FLARE folder. "
            "Place the worker file alongside `flare_risk.py`."
        )
        return False

    freqs = {c: float(st.session_state.risk_freqs.get(c, _DEFAULT_FREQ)) for c in cases}
    config = {
        "work_dir": str(WORK_DIR),
        "run_dir": str(run_dir),
        "cases": cases,
        "case_entries": case_entries,
        "freqs": freqs,
        "fast_mode": bool(fast_mode),
        "source_term_override": source_term_override,
        "timeout_s": 600,
    }
    cfg_path = run_dir / "risk_worker_config.json"
    _write_json_atomic(cfg_path, config)

    # Ensure a stale abort-request file cannot kill a newly launched worker.
    try:
        (run_dir / "risk_abort_requested.json").unlink()
    except FileNotFoundError:
        pass
    except Exception:
        pass

    status = {
        "status": "starting",
        "message": "Starting risk worker…",
        "total_runs": len(cases),
        "source_term_override": source_term_override,
        "completed_runs": 0,
        "failed_runs": 0,
        "current_case": None,
        "run_dir": str(run_dir),
    }
    _write_json_atomic(run_dir / "risk_status.json", status)

    stdout_f = open(run_dir / "risk_stdout.log", "ab", buffering=0)
    stderr_f = open(run_dir / "risk_stderr.log", "ab", buffering=0)
    creationflags = 0
    if sys.platform.startswith("win"):
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
    env = {**_os.environ, "PYTHONUTF8": "1", "MPLBACKEND": "Agg"}
    try:
        proc = subprocess.Popen(
            [sys.executable, str(worker), str(cfg_path)],
            cwd=str(WORK_DIR),
            stdout=stdout_f,
            stderr=stderr_f,
            stdin=subprocess.DEVNULL,
            env=env,
            creationflags=creationflags,
            close_fds=(not sys.platform.startswith("win")),
        )
        status.update({
            "status": "running",
            "message": "Risk worker launched.",
            "worker_pid": int(proc.pid),
        })
        _write_json_atomic(run_dir / "risk_status.json", status)
        return True
    except Exception as e:
        st.error(f"Could not launch risk worker: {e}")
        return False


def _terminate_process_tree(pid):
    """Terminate a worker process and its child flare_sim process(es)."""
    try:
        pid = int(pid)
    except Exception:
        return False, "No valid worker PID was available."
    try:
        if sys.platform.startswith("win"):
            # /T kills child processes; /F is important if flare_sim is inside a long solve.
            cp = subprocess.run(
                ["taskkill", "/PID", str(pid), "/T", "/F"],
                capture_output=True, text=True, check=False,
            )
            msg = (cp.stdout or cp.stderr or "").strip()
            return cp.returncode == 0, msg or f"taskkill returned {cp.returncode}"
        else:
            import signal as _signal
            try:
                _os.killpg(pid, _signal.SIGTERM)
            except Exception:
                _os.kill(pid, _signal.SIGTERM)
            return True, "Terminate signal sent."
    except Exception as e:
        return False, str(e)


def _abort_risk_run(run_dir):
    """Request abort and terminate the worker process tree if a PID is known."""
    run_dir = Path(run_dir)
    status_path = run_dir / "risk_status.json"
    status = _read_json(status_path, {})
    abort_file = run_dir / "risk_abort_requested.json"
    _write_json_atomic(abort_file, {
        "abort_requested": True,
        "requested_at": __import__("time").strftime("%Y-%m-%dT%H:%M:%S"),
        "reason": "User clicked Abort Risk Run in the FLARE UI.",
    })

    pid = status.get("worker_pid") or status.get("pid")
    ok, msg = (False, "No worker PID found in risk_status.json.")
    if pid:
        ok, msg = _terminate_process_tree(pid)

    status.update({
        "status": "aborted",
        "message": "Risk run aborted by user." if ok else f"Abort requested; manual process cleanup may be needed: {msg}",
        "abort_requested": True,
        "last_update": __import__("time").strftime("%Y-%m-%dT%H:%M:%S"),
    })
    _write_json_atomic(status_path, status)
    return ok, msg

def _render_worker_progress(status):
    if not status:
        return
    total = max(int(status.get("total_runs", 0) or 0), 1)
    done = int(status.get("completed_runs", 0) or 0)
    state = str(status.get("status", "unknown"))
    msg = status.get("message", "")
    frac = min(max(done / total, 0.0), 1.0)

    def _show_run_folder():
        try:
            st.caption(f"Run folder: `{Path(status.get('run_dir', '')).name}`")
        except Exception:
            pass

    if state in {"running", "starting"}:
        st.progress(frac, text=f"{done}/{total} complete — {msg}")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Completed", f"{done} / {total}")
        c2.metric("Failed", str(status.get("failed_runs", 0)))
        c3.metric("Current case", status.get("current_case") or "—")
        elapsed = status.get("current_case_elapsed_s")
        c4.metric("Case elapsed", f"{elapsed:.0f} s" if isinstance(elapsed, (int, float)) else "—")
        _show_run_folder()

        log_path = status.get("current_case_console_log")
        if log_path:
            with st.expander("Current case console log", expanded=True):
                tail = _tail_text_file(log_path, max_chars=8000)
                st.code(tail or "(No console output yet.)", language="text")

        if st.button("⛔ Abort Risk Run", type="primary", key="abort_risk_run_btn"):
            ok, msg = _abort_risk_run(status.get("run_dir", ""))
            if ok:
                st.warning("Risk run aborted. The worker process tree was terminated.")
            else:
                st.error(f"Abort requested, but process termination could not be confirmed: {msg}")
            st.rerun()
        # Lightweight polling. If the browser disconnects, the worker continues.
        import time as _time
        _time.sleep(1.0)
        st.rerun()
    elif state == "complete":
        st.success(f"Risk run complete — {done}/{total} cases processed.")
        _show_run_folder()
    elif state == "aborted":
        st.warning(f"Risk run aborted — {done}/{total} cases processed.")
        _show_run_folder()
        log_path = status.get("current_case_console_log")
        if log_path:
            with st.expander("Last case console log", expanded=True):
                st.code(_tail_text_file(log_path, max_chars=8000) or "(No console output.)", language="text")
    elif state == "failed":
        st.error(f"Risk worker failed: {msg}")
        _show_run_folder()

with st.sidebar:
    # ── Return to FLARE home ──────────────────────────────────────────────────
    st.markdown(
        f'''<a href="?page=home" target="_self" style="text-decoration:none;display:flex;
            align-items:center;gap:0.55rem;padding:0.35rem 0.6rem;
            border:1px solid #e8530a;border-radius:4px;
            font-family:Share Tech Mono,monospace;font-size:1.6rem;
            font-weight:700;letter-spacing:0.08em;color:#f97316;"
            onmouseover="this.style.background='rgba(232,83,10,0.18)';this.style.boxShadow='0 0 14px rgba(232,83,10,0.45)';this.style.color='#ffffff';"
            onmouseout="this.style.background='transparent';this.style.boxShadow='none';this.style.color='#f97316';">
          {"<img src='" + _BUDDY_B64 + "' style='height:4.5rem;width:auto;object-fit:contain;flex-shrink:0;'/>" if _BUDDY_B64 else "🔥"}
          FLARE Home
        </a>''',
        unsafe_allow_html=True,
    )
    st.divider()
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:1.4rem;'
                'color:#e6edf3;font-weight:600;margin-bottom:2px">⚛ FLARE Risk</div>',
                unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.8rem;color:#8b949e;'
                'margin-bottom:1rem">Risk Assessment</div>',
                unsafe_allow_html=True)

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
    st.markdown("### ▶  Run")
    _fast_mode = st.checkbox(
        "Fast mode (no figures)",
        value=True,
        key="risk_fast_mode",
        help="Skip figure generation for each simulation run. Only CSV and XLSX are written. Recommended for risk batch runs.",
    )

    # Optional run-wide source-term model override. These are intentionally
    # checkboxes, not a radio button, so the user may leave both unchecked and
    # allow each input deck/default to control source_term_model.
    def _risk_best_estimate_changed():
        if st.session_state.get("risk_source_term_best_estimate"):
            st.session_state.risk_source_term_rg1183 = False

    def _risk_rg1183_changed():
        if st.session_state.get("risk_source_term_rg1183"):
            st.session_state.risk_source_term_best_estimate = False

    st.caption("Source-term model override")
    _risk_be = st.checkbox(
        "Best-Estimate",
        value=st.session_state.get("risk_source_term_best_estimate", False),
        key="risk_source_term_best_estimate",
        on_change=_risk_best_estimate_changed,
        help="Override source_term_model to thermal_failure for all risk-run cases.",
    )
    _risk_rg = st.checkbox(
        "RG 1.183",
        value=st.session_state.get("risk_source_term_rg1183", False),
        key="risk_source_term_rg1183",
        on_change=_risk_rg1183_changed,
        help="Override source_term_model to licensing_auto for all risk-run cases.",
    )
    _source_term_override = None
    if st.session_state.get("risk_source_term_best_estimate"):
        _source_term_override = "thermal_failure"
    elif st.session_state.get("risk_source_term_rg1183"):
        _source_term_override = "licensing_auto"

    run_btn = st.button("▶  Run Risk Model", type="primary", width="stretch")
    if st.session_state.get("risk_results"):
        if st.button("🗑  Clear Results", width="stretch"):
            st.session_state.risk_results = {}
            # Results are run-local; clearing the UI does not delete completed run folders.
            st.session_state.pop("risk_run_dir", None)
            st.session_state.pop("risk_csv_path", None)
            st.rerun()

    st.markdown('<div class="hdiv"></div>', unsafe_allow_html=True)
    st.markdown("### 📂  Load Previous Run")

    # Find all previous risk_* subdirectories in WORK_DIR
    _risk_dirs = sorted(
        [d for d in WORK_DIR.iterdir()
         if d.is_dir() and d.name.startswith("risk_")],
        reverse=True   # most recent first
    )

    if _risk_dirs:
        _dir_labels = [d.name for d in _risk_dirs]
        _sel_label  = st.selectbox("Run folder", _dir_labels,
                                   key="load_run_sel",
                                   label_visibility="collapsed")
        _sel_dir    = WORK_DIR / _sel_label

        if st.button("📥  Load", width="stretch", key="load_run_btn"):
            try:
                import openpyxl as _opxl
                _loaded = {}
                st.caption(f"Searching: `{_sel_dir}`")
                _found_files = list(_sel_dir.glob("*_out.xlsx"))
                st.caption(f"Found {len(_found_files)} xlsx file(s)")

                for _xlsx in sorted(set(_found_files)):
                    _cname = _xlsx.stem.replace("_out", "")
                    try:
                        _accident_dose, _iodine_spike_dose, _dose = _extract_risk_eab_dose_components(_xlsx)
                        _status = "pass" if (_dose is not None and _dose > 0.0) else "no release"
                        _freq = st.session_state.risk_freqs.get(_cname, _DEFAULT_FREQ)
                        _cat  = classify(_freq)
                        _lim  = dose_limit(_freq)
                        if _dose is not None and _dose > _lim:
                            _status = "exceeds"
                        _loaded[_cname] = {
                            "freq": _freq, "dose": _dose,
                            "accident_dose": _accident_dose or 0.0,
                            "iodine_spike_dose": _iodine_spike_dose or 0.0,
                            "category": _cat, "limit": _lim,
                            "status": _status, "error": False,
                            "run_error": None,
                        }
                    except Exception as _xe:
                        st.warning(f"Could not read {_xlsx.name}: {_xe}")

                if _loaded:
                    st.session_state.risk_results = _loaded
                    save_results(_loaded, _sel_dir)
                    st.success(f"Loaded {len(_loaded)} case(s) from {_sel_label}")
                    st.rerun()
                else:
                    # Show what files are actually in the folder for diagnosis
                    _all = list(_sel_dir.iterdir()) if _sel_dir.exists() else []
                    _names = [f.name for f in _all[:10]]
                    st.warning(f"No output files found. Folder contains: "
                               f"{', '.join(_names) if _names else '(empty)'}")
            except Exception as _le:
                st.error(f"Load failed: {_le}")
    else:
        st.caption("No previous runs found.")

    st.markdown("---")
    st.caption(f"FLARE · Risk Assessment F-C Tool")

st.markdown('<div class="risk-title">⚛️  FLARE Risk Assessment F-C Tool</div>',
            unsafe_allow_html=True)
st.markdown('<div class="risk-sub">Enter event frequencies, run all cases, view the '
            'Frequency-Consequence chart.</div>', unsafe_allow_html=True)
if not cases:
    st.warning(f"No Case*_in.xlsx files found in FLARE subfolders below: `{WORK_DIR}`. Root-level case inputs are ignored.")
    st.stop()

# Merge: PRA table > _DEFAULTS > _DEFAULT_FREQ
pra_freqs = load_pra_table()

# ── Results persistence ────────────────────────────────────────────────────────
# ── Session state for frequencies ─────────────────────────────────────────────
if "risk_freqs" not in st.session_state:
    st.session_state.risk_freqs = {
        c: pra_freqs.get(c, _DEFAULTS.get(c, _DEFAULT_FREQ))
        for c in cases
    }
if "risk_results" not in st.session_state:
    # Try to restore from disk first (survives page refresh)
    st.session_state.risk_results = load_results_from_disk()

# ── Input table ────────────────────────────────────────────────────────────────
st.markdown("### Case Frequency Input")
st.markdown(
    '<div class="fc-note">NEI 18-04 classification: '
    '<span class="cat-aoo">AOO</span> f ≥ 10⁻² /yr &nbsp;|&nbsp; '
    '<span class="cat-dbe">DBE</span> 10⁻⁴ ≤ f < 10⁻² /yr &nbsp;|&nbsp; '
    '<span class="cat-bdbe">BDBE</span> 5×10⁻⁷ ≤ f < 10⁻⁴ /yr &nbsp;|&nbsp; '
    '<span class="cat-screen">Screened</span> f < 5×10⁻⁷ /yr</div>',
    unsafe_allow_html=True,
)
st.markdown("")

# Column headers
hcol1, hcol2, hcol3, hcol4 = st.columns([4, 3, 1, 2])
hcol1.markdown("**Case Name**")
hcol2.markdown("**Frequency (events/yr)**")
hcol3.markdown("**Category**")
hcol4.markdown("**Last Result**")

st.divider()

for case in cases:
    col1, col2, col3, col4 = st.columns([4, 3, 1, 2])
    freq_val = st.session_state.risk_freqs.get(case, _DEFAULT_FREQ)
    result   = st.session_state.risk_results.get(case, {})

    with col1:
        st.markdown(f"`{case}`")

    with col2:
        b_dn, inp, b_up = st.columns([1, 4, 1])
        with b_dn:
            if st.button("−", key=f"dn_{case}", help="÷10"):
                st.session_state.risk_freqs[case] = max(freq_val / 10, 1e-10)
                st.rerun()
        with inp:
            txt = st.text_input(
                label=f"freq_{case}",
                label_visibility="collapsed",
                value=f"{freq_val:.4e}",
            )
            try:
                parsed = float(txt)
                if 1e-10 <= parsed <= 10.0 and parsed != freq_val:
                    st.session_state.risk_freqs[case] = parsed
            except ValueError:
                pass
        with b_up:
            if st.button("+", key=f"up_{case}", help="×10"):
                st.session_state.risk_freqs[case] = min(freq_val * 10, 10.0)
                st.rerun()

    with col3:
        st.markdown(cat_html(classify(st.session_state.risk_freqs[case])),
                    unsafe_allow_html=True)
    with col4:
        if result:
            dose_str = f"{result['dose']:.3e} rem" if result['dose'] is not None else "0 (no release)"
            status   = result.get("status", "")
            colour   = "green" if status == "PASS" else ("red" if status == "FAIL" else "grey")
            st.markdown(f'<span style="color:{colour};font-size:0.82rem;">'
                        f'{dose_str}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span style="color:#aaa;font-size:0.82rem;">—</span>',
                        unsafe_allow_html=True)

st.divider()

# ── Save frequencies button ────────────────────────────────────────────────────
if st.button("💾  Save Frequencies", width="content"):
    _rd = _active_risk_run_dir()
    if _rd is not None and save_pra_table(st.session_state.risk_freqs, _rd):
        st.success(f"Frequencies saved to `{_rd.name}/flare_pra_table.csv`.", icon="💾")
    else:
        st.info("Frequencies are held in the current session and will be saved inside the next `risk_<timestamp>` run folder when you run the Risk Model.")

# ── Run all cases using durable worker subprocess ─────────────────────────────
if run_btn:
    import time as _time
    _tag     = _time.strftime("%Y%m%d_%H%M%S")
    _run_dir = WORK_DIR / f"risk_{_tag}"
    _run_dir.mkdir(exist_ok=True)
    st.session_state.risk_run_dir = str(_run_dir)
    st.session_state.risk_results = {}
    save_results({}, _run_dir)
    save_pra_table(st.session_state.risk_freqs, _run_dir)

    if _launch_risk_worker(_run_dir, _fast_mode, _source_term_override):
        st.success(f"Risk worker launched in `{_run_dir.name}`.")
        st.rerun()

# Monitor any active worker-backed risk run.
_active_run_dir = st.session_state.get("risk_run_dir")
if not _active_run_dir:
    _latest_status_dir = _latest_risk_status_dir()
    if _latest_status_dir is not None:
        _latest_status = _read_json(_latest_status_dir / "risk_status.json", {})
        if _latest_status.get("status") in {"starting", "running"}:
            st.session_state.risk_run_dir = str(_latest_status_dir)
            _active_run_dir = str(_latest_status_dir)

if _active_run_dir:
    _status = _load_worker_results_if_available(_active_run_dir)
    if _status and _status.get("status") in {"starting", "running", "complete", "failed", "aborted"}:
        st.markdown("### Risk Run Status")
        _render_worker_progress(_status)
        if _status.get("status") in {"complete", "failed", "aborted"}:
            # Keep the completed run selectable/loadable, but stop polling.
            pass

# ── F-C Chart ──────────────────────────────────────────────────────────────────
if st.session_state.risk_results:
    st.markdown("### Frequency-Consequence Chart")

    results = st.session_state.risk_results

    # ── Design Objective boundary ──────────────────────────────────────────────
    fc_dose = [1.0e-2, 1.0e-1, 1.0e0,  1.0e0,  2.5e1,  1.0e3,  1.0e4 ]
    fc_freq = [1.0e1,  1.0e0,  1.0e-1, 1.0e-2, 1.0e-4, 5.0e-7, 5.0e-7]

    # For shading: extend the boundary to the left and bottom edges of the plot
    # so the entire acceptable region (below/left of the curve) is filled.
    # x-axis log range [-7, 4] → 1e-7 to 1e4; y-axis extends below 5e-7.
    _shade_dose = [1e-7] + fc_dose + [1e4,   1e-7]
    _shade_freq = [fc_freq[0]] + fc_freq + [5e-8, 5e-8]

    fig = go.Figure()

    # Shaded acceptance region — full polygon covering acceptable area
    fig.add_trace(go.Scatter(
        x=_shade_dose, y=_shade_freq,
        fill="toself",
        fillcolor="rgba(0,128,0,0.07)",
        line=dict(color="rgba(0,0,0,0)"),
        showlegend=False, hoverinfo="skip",
    ))

    # F-C Target boundary line
    fig.add_trace(go.Scatter(
        x=fc_dose, y=fc_freq,
        mode="lines",
        line=dict(color="#cc0000", width=2.5, dash="dash"),
        name="Design Objective",
    ))

    # Vertical frequency dividers (AOO/DBE/BDBE boundaries)
    for freq_div, label in [
        (NEI_FREQ_AOO,    "AOO / DBE"),
        (NEI_FREQ_DBE,    "DBE / BDBE"),
        (NEI_FREQ_SCREEN, "BDBE / Screened"),
    ]:
        fig.add_hline(y=freq_div, line=dict(color="#aaa", width=1, dash="dot"),
                      annotation_text=label, annotation_position="left",
                      annotation=dict(font_size=10, font_color="#888",
                                      bgcolor="rgba(255,255,255,0.7)"))

    # Vertical consequence dividers
    for dose_div, label in [
        (NEI_DOSE_AOO,  "AOO limit"),
        (NEI_DOSE_DBE,  "DBE limit"),
        (NEI_DOSE_BDBE, "BDBE limit"),
    ]:
        fig.add_vline(x=dose_div, line=dict(color="#aaa", width=1, dash="dot"),
                      annotation_text=label, annotation_position="top left",
                      annotation=dict(font_size=10, font_color="#888",
                                      bgcolor="rgba(255,255,255,0.7)"))

    # Category colours
    cat_color = {"AOO": "#1a7f37", "DBE": "#0969da",
                 "BDBE": "#9a6700", "Screened": "#888"}
    cat_symbol= {"AOO": "circle",   "DBE": "square",
                 "BDBE": "diamond",  "Screened": "x"}

    # Plot each case — group by category for legend
    for cat in ["AOO", "DBE", "BDBE", "Screened"]:
        cat_cases = [(c, r) for c, r in results.items()
                     if r["category"] == cat]
        if not cat_cases:
            continue

        x_vals, y_vals, texts, customdata = [], [], [], []
        for cname, r in cat_cases:
            dose = r["dose"] if r["dose"] is not None and r["dose"] > 0 else 1e-6
            x_vals.append(dose)
            y_vals.append(r["freq"])
            texts.append(cname.replace("Case", ""))
            lim    = r["limit"]
            status = r["status"]
            dose_str = f"{r['dose']:.3e} rem" if r['dose'] is not None else "—"
            customdata.append(
                f"{cname}<br>"
                f"Freq: {r['freq']:.2e} /yr<br>"
                f"Dose: {dose_str}<br>"
                f"Limit: {lim} rem<br>"
                f"Status: {status.upper()}"
            )

        # Marker outline: red if FAIL, else category colour
        marker_line_colors = []
        for cname, r in cat_cases:
            if r["status"] == "FAIL":
                marker_line_colors.append("#cc0000")
            else:
                marker_line_colors.append(cat_color[cat])

        fig.add_trace(go.Scatter(
            x=x_vals, y=y_vals,
            mode="markers+text",
            text=texts,
            textposition="top center",
            textfont=dict(size=10, color=cat_color[cat]),
            marker=dict(
                size=12,
                color=cat_color[cat],
                symbol=cat_symbol[cat],
                opacity=0.85,
                line=dict(width=2, color=marker_line_colors),
            ),
            name=cat,
            hovertemplate="%{customdata}<extra></extra>",
            customdata=customdata,
        ))

    fig.update_layout(
        title=dict(
            text="Frequency-Consequence Chart  —  FLARE Cases",
            font=dict(family="IBM Plex Mono", size=15, color="#1F3864"),
        ),
        xaxis=dict(
            title="EAB TEDE Total (rem)",
            type="log",
            range=[-7, 4],
            showgrid=True, gridcolor="#eee", gridwidth=1,
            minor=dict(showgrid=True, gridcolor="#f5f5f5"),
            tickformat=".0e",
        ),
        yaxis=dict(
            title="Event Frequency (events/yr)",
            type="log",
            range=[-8, 1],
            showgrid=True, gridcolor="#eee", gridwidth=1,
            minor=dict(showgrid=True, gridcolor="#f5f5f5"),
            tickformat=".0e",
        ),
        legend=dict(
            title="Category",
            bordercolor="#ddd", borderwidth=1,
            bgcolor="rgba(255,255,255,0.9)",
            font=dict(size=11),
        ),
        plot_bgcolor="white",
        paper_bgcolor="#f5f7fa",
        height=600,
        margin=dict(l=70, r=40, t=60, b=60),
        font=dict(family="IBM Plex Sans"),
    )

    # Annotations for region labels
    for label, x, y in [
        ("ACCEPTABLE", 1e-5, 5e-2),
        ("UNACCEPTABLE", 10.0, 5e-3),
    ]:
        colour = "#006600" if label == "ACCEPTABLE" else "#cc0000"
        fig.add_annotation(
            x=np.log10(x) if label == "ACCEPTABLE" else 1.5,
            y=np.log10(y),
            xref="x", yref="y",
            text=f"<b>{label}</b>",
            showarrow=False,
            font=dict(size=12, color=colour),
            opacity=0.5,
        )

    st.plotly_chart(fig, width="stretch")

    # ── Summary table ──────────────────────────────────────────────────────────
    st.markdown("### Results Summary")
    sum_rows = []
    for cname, r in sorted(results.items()):
        dose_disp = (f"{r['dose']:.3e}" if r["dose"] is not None and r["dose"] > 0
                     else "< 1e-6")
        err_msg = r.get("run_error") or ""
        acc_disp = (f"{r.get('accident_dose', 0.0):.3e}"
                    if r.get('accident_dose', 0.0) else "0")
        isp_disp = (f"{r.get('iodine_spike_dose', 0.0):.3e}"
                    if r.get('iodine_spike_dose', 0.0) else "0")
        sum_rows.append({
            "Case":           cname,
            "Freq (/yr)":     f"{r['freq']:.2e}",
            "Category":       r["category"],
            "EAB TEDE Total (rem)": dose_disp,
            "Accident EAB TEDE (rem)": acc_disp,
            "Iodine Spike EAB TEDE (rem)": isp_disp,
            "Limit (rem)":    r["limit"],
            "Status":         r["status"].upper(),
            "Error":          err_msg,
        })
    sum_df = pd.DataFrame(sum_rows)

    def _colour_status(v):
        if v == "PASS":       return "color: green; font-weight: bold"
        if v == "FAIL":       return "color: red;   font-weight: bold"
        if v == "NO RELEASE": return "color: grey"
        return "color: orange"

    st.dataframe(
        sum_df.style.map(_colour_status, subset=["Status"]),
        width="stretch", hide_index=True,
    )

    # ── Automatic PDF export: F-C curve + results table ───────────────────────
    try:
        _risk_pdf = _risk_pdf_path()
        _export_risk_pdf(results, sum_df, _risk_pdf)
        st.session_state.risk_pdf_path = str(_risk_pdf)
        st.caption(f"Risk PDF report written automatically: `{_risk_pdf.name}`")
        with open(_risk_pdf, "rb") as _pdf_f:
            st.download_button(
                "📥  Download F-C curve + results table PDF",
                data=_pdf_f.read(),
                file_name=_risk_pdf.name,
                mime="application/pdf",
                key="risk_pdf_download",
            )
    except Exception as _pdf_e:
        st.warning(f"Risk PDF report could not be written: {_pdf_e}")

    # ── AI Risk Narrative ──────────────────────────────────────────────────────
    with st.expander("🤖  AI Risk Narrative", expanded=True):
        if _BUDDY_B64:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:0.6rem;margin-bottom:0.5rem;'>"
                f"<img src='{_BUDDY_B64}' style='height:2.2rem;width:2.2rem;border-radius:50%;object-fit:cover;flex-shrink:0;'/>"
                f"<span style='font-weight:700;font-size:1rem;'>FLARE Buddy — AI Risk Narrative</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
        st.caption(
            "Generate a technical narrative of the F-C chart, dominant contributors, "
            "dose-limit margins, and any cases that challenge the NEI 18-04 target. "
            "The selected detail level uses the same 21-level report-length scale as the "
            "PWR Simulator and UA tools."
        )

        with st.form("risk_ai_generate_form", clear_on_submit=False):
            _risk_detail = st.slider(
                "Narrative detail", 0.0, 1.0, 0.55, 0.05,
                key="risk_ai_detail",
                help=(
                    "Twenty-one detents from 0.00 to 1.00. 0.00 is an abstract-level "
                    "summary of roughly 300 words; 1.00 is a full technical discussion "
                    "of roughly 5,300 words."
                ),
            )
            _risk_detail = _risk_detail_level(_risk_detail)
            _risk_word_target = _risk_detail_word_target(_risk_detail)
            _risk_word_lo, _risk_word_hi = _risk_detail_word_range(_risk_detail)
            st.caption(
                f"Detail {_risk_detail:.2f}: {_risk_detail_descriptor(_risk_detail)} narrative; "
                f"target ~{_risk_word_target:,} words, expected range {_risk_word_lo:,}–{_risk_word_hi:,} words."
            )
            _risk_cols = st.columns([1, 1, 3])
            with _risk_cols[0]:
                _gen_risk_ai = st.form_submit_button("Generate narrative")
            with _risk_cols[1]:
                _clear_risk_ai = st.form_submit_button("Clear")

        if _clear_risk_ai:
            st.session_state.pop("risk_ai_narrative", None)
            st.session_state.pop("risk_ai_narrative_sections", None)
            st.rerun()

        if _gen_risk_ai:
            try:
                _risk_detail = float(st.session_state.get("risk_ai_detail", _risk_detail))
                _risk_detail = _risk_detail_level(_risk_detail)
                _ctx = _risk_build_common_ai_context(sum_df, _risk_detail)
                _plan = _risk_section_plan(_risk_detail)
                _total_sections = len(_plan)
                _total_target = sum(int(s["target_words"]) for s in _plan)
                _progress = st.progress(0.0, text="Planning AI risk narrative sections…")
                _status = st.empty()
                _live = st.empty()
                _status.info(
                    "Section plan: " + "; ".join(
                        f"{i+1}. {s['title']} (~{s['target_words']} words)"
                        for i, s in enumerate(_plan)
                    )
                )

                _system = (
                    "You are FLARE's risk-assessment narrative writer for nuclear safety analysis. "
                    "Write continuous, flowing prose in the style of a safety analysis report. "
                    "Use the supplied data only; do not invent values. "
                    "All numeric measures must be stated explicitly with value and abbreviated unit, "
                    "for example '24.7 rem', '0.31 rem', or '1.2×10⁻⁵/yr'. "
                    "Use Markdown only as follows: start with exactly one level-three heading line "
                    "in the form '### Section Title', then a blank line, then ordinary body paragraphs. "
                    "Do not use bullet lists, numbered lists, tables, block quotes, or bold whole paragraphs. "
                    "Do not place body text on the same line as the heading."
                )
                _prior_sections = []
                _completed_texts = []
                _stop_reasons = []
                _context_header = f"""
Overall report detail setting: {_ctx['detail']:.2f} ({_ctx['detail_label']}).
Whole-report target: about {_ctx['word_target']} words; acceptable whole-report range {_ctx['word_low']} to {_ctx['word_high']} words.
Number of cases: {_ctx['n_cases']}.
Number of FAIL cases: {_ctx['n_fail']}.
Dose plotted in the F-C chart is EAB TEDE Total = accident dose + iodine-spike dose.
NEI 18-04 categories: AOO, DBE, BDBE, Screened.
The plotted target is the FLARE/NEI 18-04 design-objective boundary.
Highest-dose row, if available:
{_ctx['max_row']}

Results table CSV:
{_ctx['table_csv']}
""".strip()

                for _idx, _section in enumerate(_plan, start=1):
                    _title = _section["title"]
                    _target_words = int(_section["target_words"])
                    _max_words = int(_section["max_words"])
                    _progress.progress((_idx - 1) / _total_sections, text=f"Generating section {_idx}/{_total_sections}: {_title}")
                    _status.info(f"Submitting section {_idx}/{_total_sections}: **{_title}** (~{_target_words} words).")
                    _already = "\n\n".join(_prior_sections[-2:])
                    _section_prompt = f"""
Prepare only this section of the FLARE Risk narrative.

Section title: {_title}
Section position: {_idx} of {_total_sections}
Section target length: about {_target_words} words.
Hard maximum section length: {_max_words} words.

Formatting requirements:
- First line must be exactly: ### {_title}
- Second line must be blank.
- Then write normal body paragraph text only.
- Do not use '#', '##', bullets, numbered lists, or bold paragraph formatting.
- Do not repeat sections already written.
- Keep this section self-contained but consistent with prior sections.

Prior completed section excerpts, for continuity:
{_already if _already else '(none yet)'}

Shared risk context:
{_context_header}
""".strip()
                    _tokens = _risk_narrative_max_tokens(_max_words)
                    _timeout = _risk_narrative_timeout_s(_target_words)
                    _section_text, _stop_reason = _anthropic_text(
                        _system, _section_prompt, max_tokens=_tokens, timeout_s=_timeout
                    )
                    _stop_reasons.append(_stop_reason)
                    _section_text = _risk_sanitize_markdown(_section_text, fallback_title=_title)
                    _section_text = _risk_clamp_words(_section_text, _max_words)
                    _completed_texts.append(_section_text)
                    _prior_sections.append(_section_text)
                    _combined_live = "\n\n".join(_completed_texts).strip()
                    _live.markdown(_combined_live)
                    _progress.progress(_idx / _total_sections, text=f"Completed section {_idx}/{_total_sections}: {_title}")

                _final_text = "\n\n".join(_completed_texts).strip()
                _final_text = _risk_sanitize_markdown(_final_text, fallback_title="FLARE Risk Narrative")
                _final_text = _risk_clamp_words(_final_text, int(round(_ctx['word_high'] * 1.05)))
                st.session_state.risk_ai_narrative = _final_text
                st.session_state.risk_ai_narrative_sections = [s["title"] for s in _plan]
                _live.empty()
                _progress.progress(1.0, text="AI risk narrative complete.")
                _status.success(
                    f"AI risk narrative complete: {_total_sections} section(s), "
                    f"{_risk_word_count(_final_text):,} words."
                )
                if any(sr == "max_tokens" for sr in _stop_reasons):
                    st.warning(
                        "One or more narrative sections reached its token budget. The displayed report was retained, "
                        "but consider reducing the detail level if the prose appears incomplete."
                    )
                st.rerun()
            except Exception as _ai_e:
                st.error(f"AI narrative failed: {_ai_e}")

        if st.session_state.get("risk_ai_narrative"):
            _sections = st.session_state.get("risk_ai_narrative_sections") or []
            if _sections:
                st.caption("Report sections: " + "; ".join(_sections))
            st.markdown(st.session_state.risk_ai_narrative)
            st.download_button(
                "⬇  Download narrative (Markdown)",
                data=st.session_state.risk_ai_narrative.encode("utf-8"),
                file_name="FLARE_risk_narrative.md",
                mime="text/markdown",
                key="risk_ai_download",
            )

    # ── Export ─────────────────────────────────────────────────────────────────
    _csv_path = Path(st.session_state.get("risk_csv_path", "") or ".")
    if _csv_path != Path(".") and _csv_path.exists():
        with open(_csv_path, "rb") as _f:
            st.download_button("📥  Download results CSV", data=_f.read(),
                               file_name=_csv_path.name, mime="text/csv")
    else:
        # Fallback: generate from current results
        csv_out = sum_df.to_csv(index=False)
        st.download_button("📥  Download results CSV", data=csv_out,
                           file_name="FLARE_risk_results.csv", mime="text/csv")
